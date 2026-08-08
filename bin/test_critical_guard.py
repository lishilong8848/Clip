# -*- coding: utf-8 -*-
from __future__ import annotations

import io
import datetime as dt
import hashlib
import sys
import tempfile
import threading
import unittest
from pathlib import Path
from unittest.mock import patch

from PIL import Image


BIN_DIR = Path(__file__).resolve().parent
if str(BIN_DIR) not in sys.path:
    sys.path.insert(0, str(BIN_DIR))

from lan_bitable_template_portal.critical_guard import (
    CRITICAL_GUARD_SHEET_NAMES,
    CRITICAL_GUARD_SOURCE_PREVIEW_RENDER_VERSION,
    CriticalGuardError,
    _compose_critical_guard_signatures,
    _excel_column_width_pixels,
    _excel_row_height_pixels,
    _render_workbook_sheet_to_png,
    build_critical_guard_workbook,
    critical_guard_catalog,
    critical_guard_template_path,
    default_response_cells,
    memory_cells_for_new_task,
    render_critical_guard_image,
    render_critical_guard_source_file_artifacts,
    render_critical_guard_source_file_preview,
    render_critical_guard_template_artifacts,
    validate_critical_guard_source_workbook,
    validate_response_for_generation,
)
from lan_bitable_template_portal.state_store import LanPortalStateStore
from lan_bitable_template_portal import portal_service as portal_service_module
from lan_bitable_template_portal.portal_service import (
    MaintenancePortalService,
    PortalConflictError,
)


class CriticalGuardHelperTests(unittest.TestCase):
    def test_catalog_only_exposes_business_sheets(self) -> None:
        catalog = critical_guard_catalog()
        self.assertEqual(
            [item["name"] for item in catalog["sheets"]],
            list(CRITICAL_GUARD_SHEET_NAMES),
        )
        self.assertNotIn("封面", [item["name"] for item in catalog["sheets"]])
        self.assertNotIn("说明", [item["name"] for item in catalog["sheets"]])

    def test_source_file_preview_url_is_content_and_renderer_versioned(self) -> None:
        source = MaintenancePortalService._critical_guard_public_scope_file(
            {
                "file_id": "guard_file_test",
                "scope": "A",
                "sheet_type": "物资检查清单",
                "original_file_name": "物资检查清单.xlsx",
                "sha256": "0123456789abcdef" * 4,
                "size": 123,
            }
        )
        self.assertIn("v=0123456789abcdef", source["preview_url"])
        self.assertIn(
            f"render={CRITICAL_GUARD_SOURCE_PREVIEW_RENDER_VERSION}",
            source["preview_url"],
        )

    def test_abnormal_item_requires_note_and_signature(self) -> None:
        cells = default_response_cells("设备安全", "A", today="2026-08-03")
        first_key = next(iter(cells["checks"]))
        cells["checks"][first_key] = {"status": "abnormal", "note": ""}
        with self.assertRaisesRegex(CriticalGuardError, "异常项必须填写备注"):
            validate_response_for_generation(
                "设备安全",
                cells,
                signature_count=1,
            )
        cells["checks"][first_key]["note"] = "设备状态异常"
        with self.assertRaisesRegex(CriticalGuardError, "检查人签名"):
            validate_response_for_generation(
                "设备安全",
                cells,
                signature_count=0,
            )

    def test_memory_reuses_content_but_resets_scope_and_date(self) -> None:
        memory = default_response_cells("客户重保", "A", today="2025-01-01")
        first_key = next(iter(memory["checks"]))
        memory["checks"][first_key] = {
            "status": "abnormal",
            "note": "保留上次异常说明",
        }
        reused = memory_cells_for_new_task("客户重保", "B", memory)
        self.assertEqual(reused["machine_room"], "南通机房B楼")
        self.assertNotEqual(reused["check_date"], "2025-01-01")
        self.assertEqual(reused["checks"][first_key]["note"], "保留上次异常说明")

    def test_pending_response_date_tracks_the_day_it_is_opened(self) -> None:
        service = MaintenancePortalService.__new__(MaintenancePortalService)
        public = service._critical_guard_public_response(
            {
                "response_id": "response-id",
                "status": "pending",
                "check_date": "2025-01-01",
                "cells": {"check_date": "2025-01-01"},
            }
        )
        self.assertEqual(public["check_date"], dt.date.today().isoformat())
        self.assertEqual(public["cells"]["check_date"], dt.date.today().isoformat())

    def test_signature_confirmation_context_is_shared_and_preview_is_hidden(self) -> None:
        first = {
            "task_id": "task-one",
            "scope": "A",
            "sheet_type": "设备安全",
        }
        second = {**first, "sheet_type": "客户重保"}
        self.assertEqual(
            MaintenancePortalService._critical_guard_signature_context(first),
            MaintenancePortalService._critical_guard_signature_context(second),
        )
        service = MaintenancePortalService.__new__(MaintenancePortalService)
        service._resolve_critical_guard_signatures = lambda **_kwargs: [
            {
                "source": "staff",
                "record_id": "staff-record",
                "name": "检查人",
                "has_signature": True,
                "ready": True,
                "signature_preview_url": "/api/signatures/image?record_id=staff-record",
            }
        ]
        public = service._critical_guard_public_response(
            {
                **first,
                "response_id": "response-one",
                "status": "draft",
                "cells": {},
                "signatures": [
                    {
                        "source": "staff",
                        "record_id": "staff-record",
                        "name": "检查人",
                    }
                ],
            },
            operator_open_id="operator-open-id",
        )
        self.assertNotIn("signature_preview_url", public["selected_signers"][0])

    def test_all_sheets_render_and_long_tables_are_not_clipped(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            output_dir = Path(temp_dir)
            dimensions: dict[str, tuple[int, int]] = {}
            for sheet_name in CRITICAL_GUARD_SHEET_NAMES:
                cells = default_response_cells(sheet_name, "A", today="2026-08-03")
                if sheet_name == "物资检查清单":
                    cells["key_spares"][0]["name"] = "长文本" * 160
                if sheet_name == "重保联络清单":
                    cells["groups"]["onsite"][0]["scope"] = "负责范围" * 160
                output_path = output_dir / f"{sheet_name}.png"
                metadata = render_critical_guard_image(
                    sheet_name=sheet_name,
                    scope="A",
                    task_name="测试重保任务",
                    cells=cells,
                    signatures=None,
                    output_path=output_path,
                )
                self.assertTrue(output_path.is_file())
                self.assertGreater(metadata["size"], 0)
                dimensions[sheet_name] = (metadata["width"], metadata["height"])
            self.assertGreater(dimensions["物资检查清单"][1], 900)
            self.assertGreater(dimensions["重保联络清单"][1], 2200)

    def test_uploaded_scope_file_is_rendered_without_rewriting_workbook(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            source_path = critical_guard_template_path()
            output_path = Path(temp_dir) / "物资检查清单.png"
            validated = validate_critical_guard_source_workbook(
                source_path,
                sheet_name="物资检查清单",
            )
            rendered = render_critical_guard_source_file_artifacts(
                source_path=source_path,
                sheet_name="物资检查清单",
                output_path=output_path,
            )

            generated_workbook = Path(rendered["workbook_path"])
            self.assertTrue(output_path.is_file())
            self.assertTrue(generated_workbook.is_file())
            self.assertEqual(generated_workbook.read_bytes(), source_path.read_bytes())
            self.assertEqual(rendered["source_file_sha256"], validated["sha256"])
            self.assertEqual(rendered["workbook_sha256"], validated["sha256"])
            self.assertGreater(rendered["width"], 600)
            self.assertGreater(rendered["height"], 600)

            preview_path = Path(temp_dir) / "重保联络清单-preview.png"
            preview = render_critical_guard_source_file_preview(
                source_path=source_path,
                sheet_name="重保联络清单",
                output_path=preview_path,
            )
            self.assertTrue(preview_path.is_file())
            self.assertGreater(preview["width"], 600)
            self.assertGreater(preview["height"], 600)

    def test_uploaded_scope_file_accepts_sheet_name_with_outer_spaces(self) -> None:
        from openpyxl import load_workbook

        with tempfile.TemporaryDirectory() as temp_dir:
            source_path = Path(temp_dir) / "material-check.xlsx"
            output_path = Path(temp_dir) / "material-check.png"
            workbook = load_workbook(critical_guard_template_path(), keep_links=False)
            workbook["物资检查清单"].title = "物资检查清单 "
            workbook.save(source_path)
            workbook.close()

            metadata = validate_critical_guard_source_workbook(
                source_path,
                sheet_name="物资检查清单",
            )
            preview = render_critical_guard_source_file_preview(
                source_path=source_path,
                sheet_name="物资检查清单",
                output_path=output_path,
            )

            self.assertEqual(metadata["sheet_type"], "物资检查清单")
            self.assertTrue(output_path.is_file())
            self.assertGreater(preview["width"], 600)
            self.assertGreater(preview["height"], 600)

    def test_uploaded_long_scope_file_reduces_render_scale(self) -> None:
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill

        with tempfile.TemporaryDirectory() as temp_dir:
            source_path = Path(temp_dir) / "long-material-check.xlsx"
            output_path = Path(temp_dir) / "long-material-check.png"
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "物资检查清单"
            for column in ("A", "B", "C"):
                worksheet.column_dimensions[column].width = 16
            for row in range(1, 211):
                worksheet.row_dimensions[row].height = 50
                fill = PatternFill(
                    fill_type="solid",
                    fgColor="DDEBFA" if row % 2 else "FFFFFF",
                )
                for column in range(1, 4):
                    cell = worksheet.cell(
                        row=row,
                        column=column,
                        value=f"物资检查项目 {row}-{column}",
                    )
                    cell.fill = fill
            workbook.save(source_path)
            workbook.close()

            preview = render_critical_guard_source_file_preview(
                source_path=source_path,
                sheet_name="物资检查清单",
                output_path=output_path,
            )

            self.assertTrue(output_path.is_file())
            self.assertLessEqual(preview["height"], 20_000)
            self.assertGreater(preview["height"], 10_000)

    def test_uploaded_scope_file_uses_actual_range_and_rejects_oversized_sheet(self) -> None:
        from openpyxl import load_workbook

        with tempfile.TemporaryDirectory() as temp_dir:
            expanded_path = Path(temp_dir) / "expanded.xlsx"
            workbook = load_workbook(critical_guard_template_path(), keep_links=False)
            workbook["物资检查清单"]["K60"] = "新增检查内容"
            workbook.save(expanded_path)
            workbook.close()
            metadata = validate_critical_guard_source_workbook(
                expanded_path,
                sheet_name="物资检查清单",
            )
            self.assertEqual(metadata["sheet_range"], "A1:K60")

            oversized_path = Path(temp_dir) / "oversized.xlsx"
            workbook = load_workbook(critical_guard_template_path(), keep_links=False)
            workbook["物资检查清单"]["A501"] = "超出范围"
            workbook.save(oversized_path)
            workbook.close()
            with self.assertRaisesRegex(CriticalGuardError, "工作表范围过大"):
                validate_critical_guard_source_workbook(
                    oversized_path,
                    sheet_name="物资检查清单",
                )

    def test_multiple_signatures_render_in_one_check_sheet(self) -> None:
        signature = Image.new("RGBA", (180, 64), (0, 0, 0, 0))
        signature.putpixel((30, 30), (0, 0, 0, 255))
        buffer = io.BytesIO()
        signature.save(buffer, format="PNG")
        with tempfile.TemporaryDirectory() as temp_dir:
            output_path = Path(temp_dir) / "multi-signature.png"
            metadata = render_critical_guard_image(
                sheet_name="设备安全",
                scope="A",
                task_name="多人签名测试",
                cells=default_response_cells("设备安全", "A", today="2026-08-03"),
                signatures=[
                    {"name": "检查人一", "image_bytes": buffer.getvalue()},
                    {"name": "检查人二", "image_bytes": buffer.getvalue()},
                ],
                output_path=output_path,
            )
            self.assertTrue(output_path.is_file())
            self.assertEqual(metadata["signature_count"], 2)

    def test_composed_signatures_only_contain_handwriting_without_name_row(self) -> None:
        signature = Image.new("RGBA", (180, 64), (0, 0, 0, 0))
        signature.putpixel((30, 30), (0, 0, 0, 255))
        buffer = io.BytesIO()
        signature.save(buffer, format="PNG")
        output, width, height, rows = _compose_critical_guard_signatures(
            [
                {"name": "检查人一", "image_bytes": buffer.getvalue()},
                {"name": "检查人二", "image_bytes": buffer.getvalue()},
            ]
        )
        self.assertTrue(output)
        self.assertEqual(width, 360)
        self.assertEqual(height, 64)
        self.assertEqual(rows, 1)

    def test_second_and_third_check_sheets_keep_template_style_and_signatures(self) -> None:
        from openpyxl import load_workbook

        signature = Image.new("RGBA", (180, 64), (0, 0, 0, 0))
        signature.putpixel((30, 30), (0, 0, 0, 255))
        buffer = io.BytesIO()
        signature.save(buffer, format="PNG")
        source_book = load_workbook(critical_guard_template_path(), data_only=False)
        try:
            with tempfile.TemporaryDirectory() as temp_dir:
                for sheet_name, signature_row in (("环境安全", 100), ("客户重保", 36)):
                    cells = default_response_cells(
                        sheet_name,
                        "A",
                        today="2026-08-03",
                    )
                    definition = next(
                        item
                        for item in critical_guard_catalog()["sheets"]
                        if item["name"] == sheet_name
                    )
                    first_item = definition["items"][0]
                    cells["checks"][first_item["key"]] = {
                        "status": "abnormal",
                        "note": "测试异常备注",
                    }
                    image_path = Path(temp_dir) / f"{sheet_name}.png"
                    metadata = render_critical_guard_template_artifacts(
                        sheet_name=sheet_name,
                        scope="A",
                        task_name="第二、第三张检查表测试",
                        cells=cells,
                        signatures=[
                            {"name": "检查人一", "image_bytes": buffer.getvalue()},
                            {"name": "检查人二", "image_bytes": buffer.getvalue()},
                        ],
                        output_path=image_path,
                    )
                    output_path = Path(metadata["workbook_path"])
                    self.assertTrue(image_path.is_file())
                    self.assertTrue(output_path.is_file())
                    self.assertGreater(metadata["size"], 0)
                    self.assertGreater(metadata["width"], 700)
                    self.assertGreater(metadata["height"], 700)
                    self.assertEqual(metadata["signature_count"], 2)
                    output_book = load_workbook(output_path, data_only=False)
                    try:
                        source_sheet = source_book[sheet_name]
                        output_sheet = output_book[sheet_name]
                        self.assertEqual(output_sheet["C3"].value, "南通机房A楼")
                        self.assertEqual(output_sheet["D3"].value, "检查日期：")
                        self.assertEqual(output_sheet["E3"].value, "2026-08-03")
                        self.assertEqual(
                            output_sheet.cell(row=int(first_item["row"]), column=4).value,
                            "异常",
                        )
                        self.assertEqual(
                            output_sheet.cell(row=int(first_item["row"]), column=5).value,
                            "测试异常备注",
                        )
                        self.assertEqual(output_sheet["B4"].style_id, source_sheet["B4"].style_id)
                        self.assertNotIn("D3:E3", {str(item) for item in output_sheet.merged_cells.ranges})
                        self.assertEqual(
                            {str(item) for item in output_sheet.merged_cells.ranges},
                            {
                                str(item)
                                for item in source_sheet.merged_cells.ranges
                                if str(item) != "D3:E3"
                            },
                        )
                        self.assertGreater(
                            len(output_sheet._images),
                            len(source_sheet._images),
                        )
                        self.assertGreater(float(output_sheet.row_dimensions[signature_row].height or 0), 15)
                    finally:
                        output_book.close()
        finally:
            source_book.close()

    def test_disaster_sheet_writes_date_in_cell_right_of_label(self) -> None:
        from openpyxl import load_workbook

        with tempfile.TemporaryDirectory() as temp_dir:
            output_path = Path(temp_dir) / "disaster.xlsx"
            build_critical_guard_workbook(
                sheet_name="灾害专项",
                scope="B",
                cells=default_response_cells("灾害专项", "B", today="2026-08-03"),
                signatures=None,
                output_path=output_path,
            )
            workbook = load_workbook(output_path, data_only=False)
            try:
                sheet = workbook["灾害专项"]
                self.assertEqual(sheet["E6"].value, "检查日期：")
                self.assertEqual(sheet["F6"].value, "2026-08-03")
                self.assertNotIn("E6:F6", {str(item) for item in sheet.merged_cells.ranges})
            finally:
                workbook.close()

    def test_internal_renderer_clips_text_and_resolves_shared_border_once(self) -> None:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Side

        with tempfile.TemporaryDirectory() as temp_dir:
            workbook_path = Path(temp_dir) / "renderer-source.xlsx"
            image_path = Path(temp_dir) / "renderer-output.png"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "设备安全"
            sheet.column_dimensions["A"].width = 12
            sheet.column_dimensions["B"].width = 12
            sheet.row_dimensions[1].height = 15
            sheet.row_dimensions[2].height = 15
            sheet["A1"] = "不会越过单元格边框的超长换行文字" * 8
            sheet["A1"].alignment = Alignment(wrap_text=True, vertical="top")
            sheet["A1"].border = Border(
                right=Side(style="thick", color="FF0000"),
                bottom=Side(style="thin", color="000000"),
            )
            sheet["B1"].border = Border(left=Side(style="thin", color="0000FF"))
            sheet["A2"].border = Border(top=Side(style="thin", color="000000"))
            workbook.save(workbook_path)
            workbook.close()

            _render_workbook_sheet_to_png(
                workbook_path=workbook_path,
                sheet_name="设备安全",
                range_address="A1:B2",
                output_path=image_path,
            )

            margin = 4
            boundary_x = margin + int(round(_excel_column_width_pixels(12) * 2))
            boundary_y = margin + int(round(_excel_row_height_pixels(15) * 2))
            second_row_bottom = boundary_y + int(round(_excel_row_height_pixels(15) * 2))
            with Image.open(image_path).convert("RGB") as image:
                # The thick right side of A1 wins over B1's thin blue left side.
                red, green, blue = image.getpixel((boundary_x, margin + 12))
                self.assertGreater(red, 180)
                self.assertLess(green, 80)
                self.assertLess(blue, 80)
                # A1's overflowing wrapped text is clipped before A2's interior.
                dark_pixels = 0
                for y in range(boundary_y + 6, second_row_bottom - 6):
                    for x in range(margin + 8, boundary_x - 8):
                        pixel = image.getpixel((x, y))
                        if max(pixel) < 120:
                            dark_pixels += 1
                self.assertEqual(dark_pixels, 0)

    def test_internal_renderer_repairs_unreadable_black_table_cells(self) -> None:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill

        with tempfile.TemporaryDirectory() as temp_dir:
            workbook_path = Path(temp_dir) / "black-cell-source.xlsx"
            image_path = Path(temp_dir) / "black-cell-output.png"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "物资检查清单"
            sheet.column_dimensions["A"].width = 12
            sheet.row_dimensions[1].height = 18
            sheet.row_dimensions[2].height = 18
            black_fill = PatternFill(fill_type="solid", fgColor="FF000000")
            sheet["A1"] = "黑底白字"
            sheet["A1"].fill = black_fill
            sheet["A1"].font = Font(color="FFFFFFFF")
            sheet["A2"] = "异常黑底黑字"
            sheet["A2"].fill = black_fill
            sheet["A2"].font = Font(color="FF000000")
            workbook.save(workbook_path)
            workbook.close()

            _render_workbook_sheet_to_png(
                workbook_path=workbook_path,
                sheet_name="物资检查清单",
                range_address="A1:A2",
                output_path=image_path,
            )

            margin = 4
            row_height = int(round(_excel_row_height_pixels(18) * 2))
            sample_x = margin + 6
            with Image.open(image_path).convert("RGB") as image:
                header_fill = image.getpixel((sample_x, margin + row_height - 6))
                repaired_fill = image.getpixel((sample_x, margin + row_height * 2 - 6))
                self.assertLess(max(header_fill), 48)
                self.assertGreater(min(repaired_fill), 240)


class CriticalGuardStateStoreTests(unittest.TestCase):
    def setUp(self) -> None:
        self.tempdir = tempfile.TemporaryDirectory()
        self.store = LanPortalStateStore(
            Path(self.tempdir.name) / "critical_guard.sqlite3"
        )
        self.task_id = "guard_test_task"
        self.response_id = "guard_test_response"
        self.memory_key = "台风橙色预警"
        self.template_version = str(critical_guard_catalog()["template_version"])
        self.initial_cells = default_response_cells(
            "设备安全",
            "A",
            today="2026-08-03",
        )
        self.store.create_critical_guard_task(
            task_id=self.task_id,
            operation_id="guard-operation-1",
            task_name="台风橙色预警",
            memory_key=self.memory_key,
            sheet_types=["设备安全"],
            target_scopes=["A"],
            template_version=self.template_version,
            created_by_open_id="operator-open-id",
            created_by_name="管理员",
            responses=[
                {
                    "response_id": self.response_id,
                    "scope": "A",
                    "sheet_type": "设备安全",
                    "cells": self.initial_cells,
                }
            ],
        )

    def tearDown(self) -> None:
        self.tempdir.cleanup()

    def test_temporary_signature_keeps_company_person_origin(self) -> None:
        service = MaintenancePortalService.__new__(MaintenancePortalService)
        service._state_store = self.store
        notice_key = "critical_guard:guard_test_task:A"

        created = service.create_temporary_signature_session(
            scope="A",
            notice_key=notice_key,
            role="inspector",
            notice_title="台风橙色预警",
            display_name="第三方检查人",
            context_type="critical_guard",
            origin_staff_record_id="staff-record-1",
            origin_staff_open_id="ou_staff_1",
            created_by="operator-open-id",
        )

        self.assertEqual(created["origin_staff_record_id"], "staff-record-1")
        self.assertEqual(created["origin_staff_open_id"], "ou_staff_1")
        listed = service.list_temporary_signatures(
            scope="A",
            notice_key=notice_key,
            created_by="operator-open-id",
        )
        self.assertEqual(len(listed["items"]), 1)
        self.assertEqual(listed["items"][0]["origin_staff_record_id"], "staff-record-1")
        self.assertEqual(listed["items"][0]["origin_staff_open_id"], "ou_staff_1")

    def test_critical_guard_rejects_nonportable_staff_signature(self) -> None:
        service = MaintenancePortalService.__new__(MaintenancePortalService)
        service._state_store = self.store
        service._load_signature_people = lambda force=False: [
            {
                "record_id": "staff-record-legacy",
                "name": "旧签名人员",
                "open_id": "ou_legacy",
                "has_signature": True,
                "signature_version": "legacy-attachment",
                "signature_crypto_version": 1,
                "portable_signature": False,
            }
        ]
        service._load_external_signature_people = lambda force=False: []
        resolved = service._resolve_critical_guard_signatures(
            response={"task_id": self.task_id, "scope": "A"},
            signatures=[
                {
                    "source": "staff",
                    "record_id": "staff-record-legacy",
                    "name": "旧签名人员",
                }
            ],
            operator_open_id="ou_legacy",
        )
        self.assertFalse(resolved[0]["has_signature"])
        self.assertFalse(resolved[0]["ready"])
        self.assertTrue(resolved[0]["signature_requires_resign"])

    def test_scope_files_are_isolated_and_new_tasks_reuse_matching_file(self) -> None:
        def put_file(file_id: str, scope: str, sheet_type: str, content: bytes):
            path = Path(self.tempdir.name) / f"{file_id}.xlsx"
            path.write_bytes(content)
            return self.store.put_critical_guard_scope_file(
                file_id=file_id,
                scope=scope,
                sheet_type=sheet_type,
                original_file_name=f"{scope}-{sheet_type}.xlsx",
                local_file_path=str(path),
                sha256=hashlib.sha256(content).hexdigest(),
                size=len(content),
                uploaded_by_open_id="operator-open-id",
                uploaded_by_name="管理员",
            )

        a_material = put_file("file-a-material", "A", "物资检查清单", b"A-material")
        a_contacts = put_file("file-a-contacts", "A", "重保联络清单", b"A-contacts")
        b_material = put_file("file-b-material", "B", "物资检查清单", b"B-material")

        self.assertEqual(
            self.store.get_latest_critical_guard_scope_file(
                scope="A", sheet_type="物资检查清单"
            )["file_id"],
            a_material["file_id"],
        )
        self.assertEqual(
            self.store.get_latest_critical_guard_scope_file(
                scope="A", sheet_type="重保联络清单"
            )["file_id"],
            a_contacts["file_id"],
        )
        self.assertEqual(
            self.store.get_latest_critical_guard_scope_file(
                scope="B", sheet_type="物资检查清单"
            )["file_id"],
            b_material["file_id"],
        )

        service = MaintenancePortalService.__new__(MaintenancePortalService)
        service._state_store = self.store
        task = service.create_critical_guard_task(
            name="楼栋文件复用测试",
            sheet_types=["物资检查清单", "重保联络清单"],
            target_scopes=["A", "B"],
            operation_id="guard-scope-file-operation",
            operator_open_id="operator-open-id",
            operator_name="管理员",
        )
        responses = {
            (item["scope"], item["sheet_type"]): item
            for item in task["responses"]
        }
        self.assertEqual(
            responses[("A", "物资检查清单")]["source_file"]["file_id"],
            a_material["file_id"],
        )
        self.assertEqual(
            responses[("A", "重保联络清单")]["source_file"]["file_id"],
            a_contacts["file_id"],
        )
        self.assertEqual(
            responses[("B", "物资检查清单")]["source_file"]["file_id"],
            b_material["file_id"],
        )
        self.assertEqual(responses[("B", "重保联络清单")]["source_file"], {})

        newer_a_material = put_file(
            "file-z-a-material",
            "A",
            "物资检查清单",
            b"A-material-newer",
        )
        detail = service.get_critical_guard_task(
            task["task_id"],
            scope="A",
        )
        material_response = next(
            item
            for item in detail["responses"]
            if item["sheet_type"] == "物资检查清单"
        )
        self.assertEqual(
            material_response["source_file"]["file_id"],
            a_material["file_id"],
        )
        self.assertEqual(
            material_response["reusable_source_file"]["file_id"],
            newer_a_material["file_id"],
        )

    def test_scope_file_upload_and_generation_use_uploaded_workbook(self) -> None:
        service = MaintenancePortalService.__new__(MaintenancePortalService)
        service._state_store = self.store
        service._critical_guard_response_locks_guard = threading.RLock()
        service._critical_guard_response_locks = {}
        service._critical_guard_render_semaphore = threading.BoundedSemaphore(2)
        task = service.create_critical_guard_task(
            name="楼栋上传生成测试",
            sheet_types=["物资检查清单"],
            target_scopes=["A"],
            operation_id="guard-scope-upload-operation",
            operator_open_id="operator-open-id",
            operator_name="管理员",
        )
        response = task["responses"][0]
        source_bytes = critical_guard_template_path().read_bytes()

        with patch.object(
            portal_service_module,
            "get_data_file_path",
            side_effect=lambda name: str(Path(self.tempdir.name) / name),
        ):
            uploaded = service.upload_critical_guard_scope_file(
                response["response_id"],
                scope="A",
                file_name="A楼物资检查清单.xlsx",
                content=source_bytes,
                expected_version=1,
                operator_open_id="operator-open-id",
                operator_name="管理员",
            )
            self.assertEqual(uploaded["version"], 2)
            self.assertEqual(uploaded["source_file"]["scope"], "A")
            self.assertEqual(uploaded["source_file"]["sheet_type"], "物资检查清单")
            self.assertNotIn("local_file_path", uploaded["source_file"])
            self.assertIn("/preview", uploaded["source_file"]["preview_url"])
            preview_bytes, preview_name = (
                service.get_critical_guard_scope_file_preview_bytes(
                    uploaded["source_file"]["file_id"],
                    scope="A",
                )
            )
            self.assertTrue(preview_bytes.startswith(b"\x89PNG\r\n\x1a\n"))
            self.assertTrue(preview_name.endswith(".png"))

            generated = service.save_critical_guard_response(
                response["response_id"],
                scope="A",
                cells=uploaded["cells"],
                signature_record_id="",
                signatures=[],
                generate_image=True,
                expected_version=2,
                operator_open_id="operator-open-id",
                operator_name="管理员",
            )
            self.assertEqual(generated["status"], "submitted")
            self.assertTrue(generated["has_image"])
            self.assertTrue(generated["has_workbook"])
            workbook_bytes, workbook_name = service.get_critical_guard_workbook_bytes(
                response["response_id"],
                scope="A",
            )
            self.assertEqual(workbook_bytes, source_bytes)
            self.assertTrue(workbook_name.endswith(".xlsx"))

    def test_task_delete_removes_task_data_but_keeps_memory_and_scope_files(self) -> None:
        data_root = Path(self.tempdir.name) / "critical_guard"
        task_root = data_root / "generated" / self.task_id / "设备安全"
        task_root.mkdir(parents=True, exist_ok=True)
        image_path = task_root / "result.png"
        workbook_path = task_root / "result.xlsx"
        image_path.write_bytes(b"generated-image")
        workbook_path.write_bytes(b"generated-workbook")
        self.store.update_critical_guard_response(
            self.response_id,
            cells=self.initial_cells,
            signatures=[
                {
                    "source": "staff",
                    "record_id": "staff-record-1",
                    "name": "检查人",
                    "role": "inspector",
                }
            ],
            signature_source="staff",
            signature_record_id="staff-record-1",
            signature_name="检查人",
            generated=True,
            generated_image={
                "path": str(image_path),
                "sha256": "image-sha",
                "size": image_path.stat().st_size,
                "width": 320,
                "height": 200,
                "workbook_path": str(workbook_path),
                "workbook_sha256": "workbook-sha",
                "workbook_size": workbook_path.stat().st_size,
            },
            share_signatures=True,
            expected_version=1,
            actor_open_id="operator-open-id",
            actor_name="管理员",
        )
        reusable_file_path = Path(self.tempdir.name) / "A楼物资检查清单.xlsx"
        reusable_file_path.write_bytes(b"reusable-scope-file")
        reusable_file = self.store.put_critical_guard_scope_file(
            file_id="guard-scope-file-kept",
            scope="A",
            sheet_type="物资检查清单",
            original_file_name=reusable_file_path.name,
            local_file_path=str(reusable_file_path),
            sha256=hashlib.sha256(reusable_file_path.read_bytes()).hexdigest(),
            size=reusable_file_path.stat().st_size,
            uploaded_by_open_id="operator-open-id",
            uploaded_by_name="管理员",
        )
        notice_key = f"critical_guard:{self.task_id}:A"
        temporary = self.store.create_mop_temporary_signature_session(
            scope="A",
            notice_key=notice_key,
            role="inspector",
            display_name="临时检查人",
            recipient_open_ids=["operator-open-id"],
            created_by="operator-open-id",
        )
        confirmation = self.store.create_mop_signature_usage_confirmation(
            scope="A",
            notice_key=notice_key,
            role="inspector",
            signer_record_id="staff-record-1",
            signer_open_id="staff-open-id-1",
            requested_by_openid="operator-open-id",
            requested_by_name="管理员",
        )

        service = MaintenancePortalService.__new__(MaintenancePortalService)
        service._state_store = self.store
        with patch.object(
            portal_service_module,
            "get_data_file_path",
            return_value=str(data_root),
        ):
            deleted = service.delete_critical_guard_task(self.task_id)
            repeated = service.delete_critical_guard_task(self.task_id)

        self.assertTrue(deleted["deleted"])
        self.assertEqual(deleted["response_count"], 1)
        self.assertEqual(deleted["removed_artifact_count"], 2)
        self.assertFalse(repeated["deleted"])
        self.assertTrue(repeated["already_deleted"])
        self.assertIsNone(
            self.store.get_critical_guard_task(
                self.task_id,
                include_all_responses=True,
            )
        )
        self.assertIsNone(self.store.get_critical_guard_response(self.response_id))
        self.assertFalse((data_root / "generated" / self.task_id).exists())
        self.assertIsNone(
            self.store.get_mop_temporary_signature_session(
                temp_id=temporary["temp_id"]
            )
        )
        with self.assertRaisesRegex(ValueError, "无效或已过期"):
            self.store.get_mop_signature_usage_confirmation(
                token=confirmation["token"]
            )
        self.assertIsNotNone(
            self.store.get_critical_guard_memory(
                memory_key=self.memory_key,
                scope="A",
                sheet_type="设备安全",
            )
        )
        self.assertEqual(
            self.store.get_critical_guard_scope_file(reusable_file["file_id"])["file_id"],
            reusable_file["file_id"],
        )

    def test_task_idempotency_scope_summary_and_memory(self) -> None:
        duplicate = self.store.create_critical_guard_task(
            task_id="unused-second-id",
            operation_id="guard-operation-1",
            task_name="不会重复创建",
            memory_key="unused",
            sheet_types=["设备安全"],
            target_scopes=["A"],
            template_version="test-version",
            created_by_open_id="operator-open-id",
            created_by_name="管理员",
            responses=[],
        )
        self.assertFalse(duplicate["created"])
        self.assertEqual(duplicate["task_id"], self.task_id)

        tasks = self.store.list_critical_guard_tasks(scope="A", status="active")
        self.assertEqual(len(tasks), 1)
        self.assertEqual(tasks[0]["pending_count"], 1)

        first_key = next(iter(self.initial_cells["checks"]))
        abnormal_cells = default_response_cells("设备安全", "A", today="2026-08-03")
        abnormal_cells["checks"][first_key] = {
            "status": "abnormal",
            "note": "首次异常说明",
        }
        updated = self.store.update_critical_guard_response(
            self.response_id,
            cells=abnormal_cells,
            signatures=[],
            signature_source="",
            signature_record_id="",
            signature_name="",
            generated=False,
            generated_image=None,
            expected_version=1,
            actor_open_id="operator-open-id",
            actor_name="填写人",
        )
        self.assertEqual(updated["version"], 2)

        normal_cells = default_response_cells("设备安全", "A", today="2026-08-03")
        updated = self.store.update_critical_guard_response(
            self.response_id,
            cells=normal_cells,
            signatures=[],
            signature_source="",
            signature_record_id="",
            signature_name="",
            generated=False,
            generated_image=None,
            expected_version=2,
            actor_open_id="operator-open-id",
            actor_name="填写人",
        )
        self.assertEqual(updated["version"], 3)
        memory = self.store.get_critical_guard_memory(
            memory_key=self.memory_key,
            scope="A",
            sheet_type="设备安全",
        )
        self.assertIsNotNone(memory)
        self.assertEqual(memory["template_version"], self.template_version)
        self.assertEqual(memory["cells"]["checks"][first_key]["status"], "normal")
        self.assertEqual(memory["cells"]["checks"][first_key]["note"], "")
        self.assertIsNone(
            self.store.get_critical_guard_memory(
                memory_key=self.memory_key,
                scope="A",
                sheet_type="设备安全",
                template_version="different-template-version",
            )
        )

        with self.assertRaisesRegex(ValueError, "其他用户修改"):
            self.store.update_critical_guard_response(
                self.response_id,
                cells=normal_cells,
                signatures=[],
                signature_source="",
                signature_record_id="",
                signature_name="",
                generated=False,
                generated_image=None,
                expected_version=2,
                actor_open_id="operator-open-id",
                actor_name="填写人",
            )

    def test_multiple_signature_references_are_persisted(self) -> None:
        signatures = [
            {
                "source": "staff",
                "record_id": "staff-record-1",
                "name": "检查人一",
                "role": "inspector",
            },
            {
                "source": "temporary",
                "temp_id": "temporary-record-1",
                "name": "临时检查人一",
                "role": "inspector",
            },
        ]
        updated = self.store.update_critical_guard_response(
            self.response_id,
            cells=self.initial_cells,
            signatures=signatures,
            signature_source="staff",
            signature_record_id="staff-record-1",
            signature_name="检查人一、临时检查人一",
            generated=False,
            generated_image=None,
            expected_version=1,
            actor_open_id="operator-open-id",
            actor_name="填写人",
        )
        self.assertEqual(updated["signatures"], signatures)
        self.assertEqual(updated["signature_record_id"], "staff-record-1")

    def test_signature_set_is_shared_by_all_check_sheets_in_task_scope(self) -> None:
        task_id = "guard_shared_signature_task"
        responses = []
        for index, sheet_name in enumerate(("设备安全", "环境安全", "客户重保"), start=1):
            responses.append(
                {
                    "response_id": f"guard_shared_response_{index}",
                    "scope": "A",
                    "sheet_type": sheet_name,
                    "cells": default_response_cells(sheet_name, "A", today="2026-08-03"),
                }
            )
        self.store.create_critical_guard_task(
            task_id=task_id,
            operation_id="guard-shared-operation",
            task_name="共享签名任务",
            memory_key="共享签名任务",
            sheet_types=["设备安全", "环境安全", "客户重保"],
            target_scopes=["A"],
            template_version="test-version",
            created_by_open_id="operator-open-id",
            created_by_name="管理员",
            responses=responses,
        )
        signatures = [
            {
                "source": "staff",
                "record_id": "staff-record-1",
                "name": "检查人一",
                "role": "inspector",
            },
            {
                "source": "staff",
                "record_id": "staff-record-2",
                "name": "检查人二",
                "role": "inspector",
            },
        ]
        updated = self.store.update_critical_guard_response(
            "guard_shared_response_1",
            cells=responses[0]["cells"],
            signatures=signatures,
            signature_source="staff",
            signature_record_id="staff-record-1",
            signature_name="检查人一、检查人二",
            generated=False,
            generated_image=None,
            share_signatures=True,
            expected_version=1,
            actor_open_id="operator-open-id",
            actor_name="填写人",
        )
        self.assertTrue(updated["signature_set_changed"])
        self.assertEqual(
            set(updated["invalidated_response_ids"]),
            {"guard_shared_response_2", "guard_shared_response_3"},
        )
        self.assertEqual(
            updated["invalidated_response_versions"],
            {"guard_shared_response_2": 2, "guard_shared_response_3": 2},
        )
        task = self.store.get_critical_guard_task(task_id, scope="A")
        self.assertIsNotNone(task)
        self.assertEqual(len(task["responses"]), 3)
        for response in task["responses"]:
            self.assertEqual(response["signatures"], signatures)
            self.assertEqual(response["signature_set_version"], 1)
            self.assertEqual(response["version"], 2)

    def test_stale_generation_cannot_overwrite_current_image(self) -> None:
        service = MaintenancePortalService.__new__(MaintenancePortalService)
        service._state_store = self.store
        service._critical_guard_response_locks_guard = threading.RLock()
        service._critical_guard_response_locks = {}
        service._critical_guard_render_semaphore = threading.BoundedSemaphore(2)
        service._resolve_critical_guard_signatures = lambda **_kwargs: [
            {
                "source": "staff",
                "record_id": "signature-record",
                "name": "检查人",
                "has_signature": True,
                "ready": True,
            }
        ]
        signature = Image.new("RGBA", (160, 60), (0, 0, 0, 0))
        buffer = io.BytesIO()
        signature.save(buffer, format="PNG")
        service.signature_image_bytes = lambda **_kwargs: (buffer.getvalue(), "image/png")

        def fake_template_renderer(*, output_path: Path, signatures, **_kwargs):
            output_path.parent.mkdir(parents=True, exist_ok=True)
            image = Image.new("RGB", (320, 200), "white")
            image.save(output_path, format="PNG")
            workbook_path = output_path.with_suffix(".xlsx")
            workbook_path.write_bytes(b"test-workbook")
            return {
                "path": str(output_path),
                "sha256": "image-sha",
                "size": output_path.stat().st_size,
                "width": 320,
                "height": 200,
                "signature_count": len(signatures or []),
                "workbook_path": str(workbook_path),
                "workbook_sha256": "workbook-sha",
                "workbook_size": workbook_path.stat().st_size,
            }

        with patch.object(
            portal_service_module,
            "get_data_file_path",
            side_effect=lambda name: str(Path(self.tempdir.name) / name),
        ), patch.object(
            portal_service_module,
            "render_critical_guard_template_artifacts",
            side_effect=fake_template_renderer,
        ):
            generated = service.save_critical_guard_response(
                self.response_id,
                scope="A",
                cells=self.initial_cells,
                signature_record_id="signature-record",
                signatures=[
                    {
                        "source": "staff",
                        "record_id": "signature-record",
                        "name": "检查人",
                    }
                ],
                generate_image=True,
                expected_version=1,
                operator_open_id="operator-open-id",
                operator_name="检查人",
            )
            self.assertEqual(generated["version"], 2)
            image_files = list(Path(self.tempdir.name).rglob("*.png"))
            self.assertEqual(len(image_files), 1)
            original_bytes = image_files[0].read_bytes()

            with self.assertRaisesRegex(PortalConflictError, "其他用户修改"):
                service.save_critical_guard_response(
                    self.response_id,
                    scope="A",
                    cells=self.initial_cells,
                    signature_record_id="signature-record",
                    signatures=[
                        {
                            "source": "staff",
                            "record_id": "signature-record",
                            "name": "检查人",
                        }
                    ],
                    generate_image=True,
                    expected_version=1,
                    operator_open_id="operator-open-id",
                    operator_name="检查人",
                )
            image_files_after_conflict = list(Path(self.tempdir.name).rglob("*.png"))
            self.assertEqual(image_files_after_conflict, image_files)
            self.assertEqual(image_files_after_conflict[0].read_bytes(), original_bytes)

    def test_outdated_template_blocks_response_update(self) -> None:
        import sqlite3

        conn = sqlite3.connect(self.store.db_path)
        try:
            conn.execute(
                "UPDATE critical_guard_tasks SET template_version=? WHERE task_id=?",
                ("outdated-template", self.task_id),
            )
            conn.commit()
        finally:
            conn.close()
        service = MaintenancePortalService.__new__(MaintenancePortalService)
        service._state_store = self.store
        service._critical_guard_response_locks_guard = threading.RLock()
        service._critical_guard_response_locks = {}
        with self.assertRaisesRegex(PortalConflictError, "检查模板已更新"):
            service.save_critical_guard_response(
                self.response_id,
                scope="A",
                cells=self.initial_cells,
                signature_record_id="",
                signatures=[],
                generate_image=False,
                expected_version=1,
                operator_open_id="operator-open-id",
                operator_name="检查人",
            )


if __name__ == "__main__":
    unittest.main()
