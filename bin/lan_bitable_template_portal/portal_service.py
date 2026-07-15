# -*- coding: utf-8 -*-
from __future__ import annotations

import ast
import datetime as dt
import base64
import csv
import difflib
import hashlib
import hmac
import io
import json
import os
import re
import secrets
import tempfile
import threading
import time
import uuid
import copy
import zipfile
from contextlib import suppress
from dataclasses import dataclass, field as dataclass_field
from pathlib import Path
from typing import Any, Callable
from urllib.parse import quote, urlparse
import xml.etree.ElementTree as ET

import lark_oapi as lark
from lark_oapi.api.drive.v1 import UploadAllMediaRequest, UploadAllMediaRequestBody

from upload_event_module.config import EVENT_NOTICE_FIELDS, config, get_field_config
from upload_event_module.services.feishu_service import (
    ensure_feishu_token,
    refresh_feishu_token,
)
from upload_event_module.services.feishu_token_manager import TOKEN_ERROR_CODES
from upload_event_module.services.http_client import FeishuHTTPError, FeishuHttpClient
from upload_event_module.services.robot_webhook import send_text_to_open_ids
from upload_event_module.utils import get_data_file_path

from .state_store import LanPortalStateStore
from .identity_utils import (
    canonical_source_record_id,
    canonical_target_record_id,
    is_local_record_id,
    normalize_notice_identity_payload,
)
from .signature_crypto import (
    SIGNATURE_ENCRYPTED_MAGIC,
    SignatureCryptoError,
    SignatureCryptoManager,
    encrypted_signature_file_name,
)


DEFAULT_APP_TOKEN = "HU38bc1vnamMK9sCeOgclUvXnFc"
DEFAULT_TABLE_ID = "tblzk7WrXxNWQy6V"
CHANGE_SOURCE_APP_TOKEN = "JhiVwgfoIimAqEk8YwEc09sknGd"
CHANGE_SOURCE_TABLE_ID = "tblBvg6wCYSX3hcg"
ZHIHANG_CHANGE_APP_TOKEN = "IrIibPkUOa6udGsMhu2cbOqhnWg"
ZHIHANG_CHANGE_TABLE_ID = "tblqMJvYW5dxFFfU"
REPAIR_SOURCE_APP_TOKEN = "AnEBwJlvGiJfDdkOB32cUPuknzg"
REPAIR_SOURCE_TABLE_ID = "tblcws1gwaFQnU6H"
REPAIR_SOURCE_VIEW_ID = ""
REPAIR_MANAGEMENT_TABLE_ID = REPAIR_SOURCE_TABLE_ID
REPAIR_MANAGEMENT_COMPLETED_WORKFLOW = "维修完成"
REPAIR_MANAGEMENT_REQUIRED_FIELD_GROUPS: tuple[tuple[str, ...], ...] = (
    ("维修名称",),
    ("故障发生时间",),
    ("故障维修原因",),
    ("所属专业", "专业（推送消息用）"),
    (
        "所属数据中心/楼栋-使用",
        "所属数据中心/楼栋（关联CMDB唯一ID关联,DE不选）",
    ),
)
REPAIR_MANAGEMENT_FORM_FIELD_NAMES = {
    "维修名称",
    "故障发生时间",
    "故障维修原因",
    "故障发生现象描述",
    "所属专业",
    "所属数据中心/楼栋-使用",
}
REPAIR_MANAGEMENT_EVENT_LINK_FIELD_NAMES = (
    "关联事件单",
)
REPAIR_MANAGEMENT_REPAIR_LINK_FIELD_NAMES = (
    "设备检修关联",
)
REPAIR_MANAGEMENT_CUSTOM_MULTI_SELECT_FIELDS = {
    "对应来源",
    "维修来源",
}
REPAIR_MANAGEMENT_SOURCE_OPTION_ALIASES = {
    "bms": "BMS系统",
    "bms系统": "BMS系统",
    "方舟": "方舟系统",
    "方舟系统": "方舟系统",
    "巡检": "巡检发现",
    "巡检发现": "巡检发现",
}
REPAIR_MANAGEMENT_LEGACY_OPTION_ALIASES = {
    ("所属专业", "optassnaw3"): "电气",
    ("维修来源", "optt3fgrvz"): "BMS系统",
    ("维修方", "optt2azmkp"): "我方",
}
REPAIR_MANAGEMENT_PROTECTED_FIELD_NAMES = {
    "维修单号",
    "流程",
    "关联事件单",
    "设备检修关联",
    "维修跟进记录",
    "CMDB唯一id",
    "是否有唯一id",
    "智航设备名称",
    "当前日期",
    "开始时间",
    "维修结束时间（2026）",
    "维修结束时间",
    "维修周期",
    "对应来源",
    "对应事件等级",
    "当前维修进度",
    "随工人员（或我方维修人员）",
    "维修审批人",
    "楼栋负责人",
    "机房负责人",
    "值班账号",
    "消息推送人",
    "创建日期",
    "记录创建人",
    "修改人",
    "最后修改时间",
    "辅助列-本周打标",
    "辅助列-2个月内",
    "公式",
    "测试ing（勿动）",
}
REPAIR_MANAGEMENT_EVENT_AUTO_FIELD_NAMES = (
    "对应来源",
    "对应事件等级",
    "值班账号",
    "消息推送人",
)
REPAIR_MANAGEMENT_REPAIR_AUTO_FIELD_NAMES = (
    "维修开始时间",
    "开始时间",
    "维修结束时间（2026）",
    "维修结束时间",
    "维修周期",
    "当前维修进度",
    "检修通告名称",
    "随工人员（或我方维修人员）",
)
REPAIR_MANAGEMENT_FOLLOWUP_AUTO_FIELD_NAMES = (
    "维修跟进记录",
    "维修开始时间",
    "开始时间",
    "维修结束时间（2026）",
    "维修结束时间",
    "维修周期",
    "设备名称",
    "设备编号",
    "设备品牌",
    "设备型号",
    "维修方",
    "供应商名称",
    "供应商维修人员",
    "设备生产日期",
    "设备使用年限",
    "设备容量KW/AH",
    "是否质保期内",
    "随工人员（或我方维修人员）",
    "更换备件名称",
    "更换备件数量",
    "故障维修总费用（跟进完成的维修项）",
    "维修审批人",
    "推送群组",
    "维修方案附件",
    "子项链接",
    "跟进项",
    "后续整改措施",
    "最新维修跟进时间",
    "CMDB唯一id",
    "是否有唯一id",
    "智航设备名称",
    "当前维修进度",
    "维修进展描述",
)
REPAIR_MANAGEMENT_PROTECTED_FIELD_NAMES.update(
    REPAIR_MANAGEMENT_FOLLOWUP_AUTO_FIELD_NAMES
)
REPAIR_SYNC_TABLE_ID = "tblSA9euoote8aCA"
REPAIR_MANAGEMENT_REPAIR_TABLE_ID = "tblpaHktT0mn0hwg"
REPAIR_FOLLOWUP_TABLE_ID = "tblG4UpMdgYHykUY"
REPAIR_EQUIPMENT_CATALOG_TABLE_ID = "tblkKnYwajfRmquQ"
REPAIR_CMDB_TABLE_ID = "tblJTRguSUij2RUM"
REPAIR_ROUTING_TABLE_ID = "tblWfnhqLlrtkIIg"
REPAIR_FOLLOWUP_PARENT_ID_FIELD_NAME = "维修汇总记录ID"
REPAIR_FOLLOWUP_CMDB_FIELD_NAME = "CMDB设备唯一ID"
REPAIR_FOLLOWUP_DEVICE_NAME_FIELD_NAME = "设备名称"
REPAIR_FOLLOWUP_DEVICE_NUMBER_FIELD_NAME = "设备编号"
REPAIR_FOLLOWUP_BRAND_FIELD_NAME = "设备品牌"
REPAIR_FOLLOWUP_MODEL_FIELD_NAME = "设备型号"
REPAIR_FOLLOWUP_REPAIR_PARTY_FIELD_NAME = "维修方"
REPAIR_FOLLOWUP_SUPPLIER_FIELD_NAME = "供应商名称"
REPAIR_FOLLOWUP_SUPPLIER_PERSON_FIELD_NAME = "供应商维修人员"
REPAIR_EQUIPMENT_CATALOG_BRAND_FIELD_NAME = "设备品牌"
REPAIR_EQUIPMENT_CATALOG_MODEL_FIELD_NAME = "设备型号"
REPAIR_FOLLOWUP_CATALOG_RUNTIME_KEY = "repair_equipment_brand_model_catalog"
REPAIR_FOLLOWUP_CATALOG_CACHE_TTL_SECONDS = 24 * 60 * 60
REPAIR_FOLLOWUP_CATALOG_MAX_RECORDS = 5000
REPAIR_SNAPSHOT_SOURCE_PROJECTS = "repair_projects"
REPAIR_SNAPSHOT_SOURCE_FOLLOWUPS = "repair_followups"
REPAIR_SNAPSHOT_SOURCE_EVENTS = "repair_events"
REPAIR_SNAPSHOT_SOURCE_NOTICES = "repair_notices"
REPAIR_SNAPSHOT_SOURCE_CMDB = "repair_cmdb"
REPAIR_SNAPSHOT_SOURCE_ROUTING = "repair_routing"
REPAIR_SNAPSHOT_TTL_SECONDS = {
    REPAIR_SNAPSHOT_SOURCE_PROJECTS: 2 * 60,
    REPAIR_SNAPSHOT_SOURCE_FOLLOWUPS: 2 * 60,
    REPAIR_SNAPSHOT_SOURCE_EVENTS: 5 * 60,
    REPAIR_SNAPSHOT_SOURCE_NOTICES: 5 * 60,
    REPAIR_SNAPSHOT_SOURCE_CMDB: 24 * 60 * 60,
    REPAIR_SNAPSHOT_SOURCE_ROUTING: 30 * 60,
}
REPAIR_SNAPSHOT_FAILURE_BACKOFF_SECONDS = 60
REPAIR_FOLLOWUP_SCHEMA_MIGRATION_RUNTIME_KEY = (
    "repair_followup_canonical_fields_v1"
)
REPAIR_MANAGEMENT_PROGRESS_SCHEMA_RUNTIME_KEY = (
    "repair_management_progress_matches_followup_v1"
)
REPAIR_FOLLOWUP_LEGACY_FIELD_MAPPINGS: tuple[tuple[str, str], ...] = (
    ("设备名称-1", REPAIR_FOLLOWUP_DEVICE_NAME_FIELD_NAME),
    ("设备编号-1", REPAIR_FOLLOWUP_DEVICE_NUMBER_FIELD_NAME),
    ("设备品牌 -1", REPAIR_FOLLOWUP_BRAND_FIELD_NAME),
    ("设备型号 -1", REPAIR_FOLLOWUP_MODEL_FIELD_NAME),
    ("维修方 -1", REPAIR_FOLLOWUP_REPAIR_PARTY_FIELD_NAME),
    ("供应商名称 -1", REPAIR_FOLLOWUP_SUPPLIER_FIELD_NAME),
    ("供应商维修人员-1", REPAIR_FOLLOWUP_SUPPLIER_PERSON_FIELD_NAME),
)
REPAIR_FOLLOWUP_CANONICAL_SELECT_SOURCES = {
    REPAIR_FOLLOWUP_DEVICE_NAME_FIELD_NAME: "设备名称-1",
    REPAIR_FOLLOWUP_BRAND_FIELD_NAME: "设备品牌 -1",
    REPAIR_FOLLOWUP_MODEL_FIELD_NAME: "设备型号 -1",
    REPAIR_FOLLOWUP_REPAIR_PARTY_FIELD_NAME: "维修方 -1",
}
REPAIR_FOLLOWUP_CUSTOM_SINGLE_SELECT_FIELDS = {
    REPAIR_FOLLOWUP_DEVICE_NAME_FIELD_NAME,
    REPAIR_FOLLOWUP_BRAND_FIELD_NAME,
    REPAIR_FOLLOWUP_MODEL_FIELD_NAME,
    REPAIR_FOLLOWUP_REPAIR_PARTY_FIELD_NAME,
    "是否质保期内",
    "所属专业",
}
REPAIR_FOLLOWUP_BACKFILL_RUNTIME_KEY = "repair_followup_parent_backfill_v1"
REPAIR_FOLLOWUP_BACKFILL_INTERVAL_SECONDS = 24 * 60 * 60
REPAIR_FOLLOWUP_WRITABLE_FIELD_NAMES = {
    "是否本维修单第一次提交跟进记录",
    REPAIR_FOLLOWUP_DEVICE_NAME_FIELD_NAME,
    REPAIR_FOLLOWUP_DEVICE_NUMBER_FIELD_NAME,
    REPAIR_FOLLOWUP_BRAND_FIELD_NAME,
    REPAIR_FOLLOWUP_MODEL_FIELD_NAME,
    REPAIR_FOLLOWUP_REPAIR_PARTY_FIELD_NAME,
    REPAIR_FOLLOWUP_SUPPLIER_FIELD_NAME,
    REPAIR_FOLLOWUP_SUPPLIER_PERSON_FIELD_NAME,
    "设备生产日期",
    "设备使用年限",
    "设备容量KW/AH",
    "是否质保期内",
    "随工人员（我方维修人员）",
    "更换备件名称",
    "更换备件数量",
    "维修进展描述",
    "维修进度",
    "故障维修总费用",
    "超链接",
    "跟进项（如有）",
    "后续整改措施（如有）",
}
REPAIR_FOLLOWUP_FORM_FIELD_NAMES = {
    REPAIR_FOLLOWUP_DEVICE_NAME_FIELD_NAME,
    REPAIR_FOLLOWUP_DEVICE_NUMBER_FIELD_NAME,
    REPAIR_FOLLOWUP_BRAND_FIELD_NAME,
    REPAIR_FOLLOWUP_MODEL_FIELD_NAME,
    REPAIR_FOLLOWUP_REPAIR_PARTY_FIELD_NAME,
    REPAIR_FOLLOWUP_SUPPLIER_FIELD_NAME,
    REPAIR_FOLLOWUP_SUPPLIER_PERSON_FIELD_NAME,
    "设备生产日期",
    "设备使用年限",
    "设备容量KW/AH",
    "是否质保期内",
    "随工人员（我方维修人员）",
    "更换备件名称",
    "更换备件数量",
    "维修进展描述",
    "维修进度",
    "故障维修总费用",
    "跟进项（如有）",
    "后续整改措施（如有）",
}
REPAIR_FOLLOWUP_READ_FIELD_NAMES = (
    "维修简述",
    "维修名称",
    REPAIR_FOLLOWUP_PARENT_ID_FIELD_NAME,
    REPAIR_FOLLOWUP_CMDB_FIELD_NAME,
    "区域",
    "所属数据中心",
    "维修来源",
    "维修方",
    "维修审批人",
    "推送群组",
    "是否本维修单第一次提交跟进记录",
    "维修开始时间",
    "维修结束时间",
    "维修进展描述",
    "维修进度",
    "所属数据中心",
    "所属专业",
    "楼栋",
    REPAIR_FOLLOWUP_DEVICE_NAME_FIELD_NAME,
    REPAIR_FOLLOWUP_DEVICE_NUMBER_FIELD_NAME,
    REPAIR_FOLLOWUP_BRAND_FIELD_NAME,
    REPAIR_FOLLOWUP_MODEL_FIELD_NAME,
    REPAIR_FOLLOWUP_REPAIR_PARTY_FIELD_NAME,
    REPAIR_FOLLOWUP_SUPPLIER_FIELD_NAME,
    REPAIR_FOLLOWUP_SUPPLIER_PERSON_FIELD_NAME,
    "设备生产日期",
    "设备使用年限",
    "设备容量KW/AH",
    "是否质保期内",
    "随工人员（我方维修人员）",
    "故障维修总费用",
    "更换备件名称",
    "更换备件数量",
    "跟进项（如有）",
    "后续整改措施（如有）",
    "超链接",
    "创建时间",
)
REPAIR_FOLLOWUP_SUMMARY_COPY_MAPPINGS: tuple[tuple[str, tuple[str, ...]], ...] = (
    ("维修简述", ("维修名称",)),
    ("维修名称", ("维修名称",)),
    ("区域", ("区域",)),
    (
        "所属数据中心",
        (
            "所属数据中心",
            "所属数据中心/楼栋-使用",
            "所属数据中心/楼栋（关联CMDB唯一ID关联,DE不选）",
        ),
    ),
    ("所属专业", ("所属专业", "专业（推送消息用）")),
    ("设备名称", ("设备名称",)),
    ("设备品牌", ("设备品牌",)),
    ("设备型号", ("设备型号",)),
    ("设备编号", ("设备编号",)),
    ("设备生产日期", ("设备生产日期",)),
    ("设备使用年限", ("设备使用年限",)),
    ("设备容量KW/AH", ("设备容量KW/AH",)),
    ("是否质保期内", ("是否质保期内",)),
    ("维修方", ("维修方",)),
    ("供应商名称", ("供应商名称",)),
    ("供应商维修人员", ("供应商维修人员",)),
    ("维修开始时间", ("维修开始时间",)),
    ("维修结束时间", ("维修结束时间",)),
    ("结束时间", ("维修结束时间",)),
    ("维修单号", ("维修单号",)),
    ("维修来源", ("对应来源",)),
    ("维修审批人", ("维修审批人",)),
    ("推送群组", ("推送群组",)),
    (
        "随工人员（我方维修人员）",
        ("随工人员（或我方维修人员）", "随工人员（我方维修人员）"),
    ),
    (
        "楼栋",
        (
            "所属数据中心/楼栋-使用",
            "所属数据中心/楼栋（关联CMDB唯一ID关联,DE不选）",
        ),
    ),
)
REPAIR_LINK_FIELD_NAME = "设备检修关联"
REPAIR_LINK_DELAY_SECONDS = 70 * 60
REPAIR_LINK_RETRY_SECONDS = 10 * 60
REPAIR_LINK_RETRY_SLOW_SECONDS = 20 * 60
REPAIR_LINK_FAST_RETRY_ATTEMPTS = 6
REPAIR_LINK_MAX_ATTEMPTS = 18
BITABLE_DATA_NOT_READY_CODE = 1254607
BITABLE_TRANSIENT_RETRY_DELAYS = (1.0, 2.5, 5.0)
MOP_CANDIDATE_CACHE_TTL_SECONDS = 10 * 60
MOP_LOCAL_UPLOAD_MAX_BYTES = 20 * 1024 * 1024
MOP_LOCAL_UPLOAD_EXTENSIONS = {".xlsx", ".xlsm", ".xls"}
MOP_SOURCE_APP_TOKEN = "MliKbC3fXa8PXrsndKscmxjdn1g"
MOP_SOURCE_TABLE_ID = "tblpqwu1kQ0bmi0i"
MOP_SOURCE_VIEW_ID = "vewrHJHl3v"
MOP_TITLE_FIELD_NAME = "文件名"
MOP_ATTACHMENT_FIELD_NAME = "文件"
SIGNATURE_APP_TOKEN = "HU38bc1vnamMK9sCeOgclUvXnFc"
SIGNATURE_TABLE_ID = "tbluozblhRAjbljX"
SIGNATURE_VIEW_ID = ""
SIGNATURE_NAME_FIELD = "姓名"
SIGNATURE_USER_FIELD = "员工姓名"
SIGNATURE_ATTACHMENT_FIELD = "手写签名"
SIGNATURE_KEY_FIELD = "密钥"
SIGNATURE_INACTIVE_FIELD = "离职/异动情况"
SIGNATURE_PEOPLE_CACHE_TTL_SECONDS = 5 * 60
SIGNATURE_LINK_TOKEN_TTL_SECONDS = 60 * 60
TEMP_SIGNATURE_TABLE_ID = "tblC77nllNrprHBY"
TEMP_SIGNATURE_NAME_FIELD = "员工姓名"
TEMP_SIGNATURE_EMPLOYEE_NO_FIELD = "员工工号"
TEMP_SIGNATURE_CERT_FIELD = "持证"
TEMP_SIGNATURE_ATTACHMENT_FIELD = "手写签名"
TEMP_SIGNATURE_KEY_FIELD = "密钥"
TEMP_SIGNATURE_BUILDING_FIELD = "楼栋"
TEMP_SIGNATURE_SPECIALTY_FIELD = "专业"
MOP_SIGNED_ATTACHMENT_FIELD = "维护保养单"
MOP_ENGINEER_CONFIRM_FIELD = "工程师确认"
MOP_SUPERVISOR_CONFIRM_FIELD = "主管确认"
ATTACHMENT_LATEST_FIELD_NAMES = (
    "最新发布时间",
    "最新发布",
    "发布时间",
    "更新时间",
    "最后更新时间",
    "最新更新时间",
)
ATTACHMENT_CACHE_DAILY_MARKER = "attachment_cache_daily_scan.json"
REPAIR_SOURCE_PAGE_SIZE = 200
DEFAULT_IMPACT_TEXT = "对IT业务无影响，不会触发BA和BMS系统相关告警"
DEFAULT_PROGRESS_TEXT = "准备工作已完成，人员已就位，可否开始操作？"
DEFAULT_MAINTENANCE_STATUS = "未开始"
MAINTENANCE_STATUS_ONGOING = "进行中"
WORKBENCH_SOURCE_STATUSES = {DEFAULT_MAINTENANCE_STATUS, MAINTENANCE_STATUS_ONGOING}
MAINTENANCE_COMPLETED_STATUSES = {
    "正常结束",
    "延期结束",
    "延迟结束",
    "已结束",
    "已完成",
    "维修完成",
    "闭环",
}
WORK_TYPE_MAINTENANCE = "maintenance"
WORK_TYPE_CHANGE = "change"
WORK_TYPE_REPAIR = "repair"
WORK_TYPE_POWER = "power"
WORK_TYPE_POLLING = "polling"
WORK_TYPE_ADJUST = "adjust"
NOTICE_TYPE_MAINTENANCE = "维保通告"
NOTICE_TYPE_CHANGE = "变更通告"
NOTICE_HEADING_CHANGE = NOTICE_TYPE_CHANGE
NOTICE_TYPE_REPAIR = "设备检修"
NOTICE_TYPE_POWER = "上下电通告"
NOTICE_TYPE_POWER_UP = "上电通告"
NOTICE_TYPE_POWER_DOWN = "下电通告"
NOTICE_TYPE_POLLING = "设备轮巡"
NOTICE_TYPE_ADJUST = "设备调整"
WORK_TYPE_EVENT = "event"
NOTICE_TYPE_EVENT = "事件通告"
NOTICE_TIME_SEPARATOR = "~"
NOTICE_TEXT_TEMPLATES = {
    WORK_TYPE_MAINTENANCE: {
        "heading": NOTICE_TYPE_MAINTENANCE,
        "fields": (
            ("名称", "title"),
            ("时间", "time_range"),
            ("位置", "location"),
            ("内容", "content"),
            ("原因", "reason"),
            ("影响", "impact"),
            ("进度", "progress"),
        ),
    },
    WORK_TYPE_CHANGE: {
        "heading": NOTICE_HEADING_CHANGE,
        "fields": (
            ("名称", "title"),
            ("等级", "level"),
            ("时间", "time_range"),
            ("位置", "location"),
            ("内容", "content"),
            ("原因", "reason"),
            ("影响", "impact"),
            ("进度", "progress"),
        ),
    },
    WORK_TYPE_REPAIR: {
        "heading": NOTICE_TYPE_REPAIR,
        "fields": (
            ("标题", "title"),
            ("地点", "location"),
            ("紧急程度", "level"),
            ("专业", "specialty"),
            ("发现故障时间", "fault_time"),
            ("期望完成时间", "expected_time"),
            ("维修设备", "repair_device"),
            ("维修故障", "repair_fault"),
            ("故障类型", "fault_type"),
            ("维修方式", "repair_mode"),
            ("影响范围", "impact"),
            ("故障发现方式", "discovery"),
            ("故障现象", "symptom"),
            ("故障原因", "reason"),
            ("解决方案", "solution"),
            ("备件更换情况", "spare_parts"),
            ("完成情况", "progress"),
        ),
    },
    WORK_TYPE_POWER: {
        "heading": "上电通告",
        "fields": (
            ("名称", "title"),
            ("时间", "time_range"),
            ("柜号", "cabinet"),
            ("数量", "quantity"),
            ("进度", "progress"),
        ),
    },
    WORK_TYPE_POLLING: {
        "heading": NOTICE_TYPE_POLLING,
        "fields": (
            ("标题", "title"),
            ("时间", "time_range"),
            ("设备", "device"),
            ("内容", "content"),
            ("影响", "impact"),
            ("进度", "progress"),
        ),
    },
    WORK_TYPE_ADJUST: {
        "heading": NOTICE_TYPE_ADJUST,
        "fields": (
            ("名称", "title"),
            ("时间", "time_range"),
            ("位置", "location"),
            ("内容", "content"),
            ("原因", "reason"),
            ("影响", "impact"),
            ("进度", "progress"),
        ),
    },
}
NOTICE_TYPE_KEYWORD_RULES = {
    WORK_TYPE_MAINTENANCE: (r"维保", r"维护"),
    WORK_TYPE_CHANGE: (r"变更",),
    WORK_TYPE_REPAIR: (r"检修", r"维修"),
    WORK_TYPE_POWER: (r"上下电", r"上电", r"下电"),
    WORK_TYPE_POLLING: (r"轮巡",),
    WORK_TYPE_ADJUST: (r"调整",),
}
NOTICE_WORK_TYPE_LABELS = {
    WORK_TYPE_MAINTENANCE: "维保",
    WORK_TYPE_CHANGE: "变更",
    WORK_TYPE_REPAIR: "检修",
    WORK_TYPE_POWER: "上电",
    WORK_TYPE_POLLING: "轮巡",
    WORK_TYPE_ADJUST: "调整",
    WORK_TYPE_EVENT: "事件",
}
NOTICE_TYPE_BY_WORK_TYPE = {
    WORK_TYPE_MAINTENANCE: NOTICE_TYPE_MAINTENANCE,
    WORK_TYPE_CHANGE: NOTICE_TYPE_CHANGE,
    WORK_TYPE_REPAIR: NOTICE_TYPE_REPAIR,
    WORK_TYPE_POWER: NOTICE_TYPE_POWER_UP,
    WORK_TYPE_POLLING: NOTICE_TYPE_POLLING,
    WORK_TYPE_ADJUST: NOTICE_TYPE_ADJUST,
    WORK_TYPE_EVENT: NOTICE_TYPE_EVENT,
}
WORK_TYPE_BY_NOTICE_TYPE = {
    NOTICE_TYPE_MAINTENANCE: WORK_TYPE_MAINTENANCE,
    "维护通告": WORK_TYPE_MAINTENANCE,
    NOTICE_TYPE_CHANGE: WORK_TYPE_CHANGE,
    "变更通告": WORK_TYPE_CHANGE,
    NOTICE_TYPE_REPAIR: WORK_TYPE_REPAIR,
    "检修通告": WORK_TYPE_REPAIR,
    NOTICE_TYPE_POWER: WORK_TYPE_POWER,
    NOTICE_TYPE_POWER_UP: WORK_TYPE_POWER,
    NOTICE_TYPE_POWER_DOWN: WORK_TYPE_POWER,
    NOTICE_TYPE_POLLING: WORK_TYPE_POLLING,
    "轮巡通告": WORK_TYPE_POLLING,
    NOTICE_TYPE_ADJUST: WORK_TYPE_ADJUST,
    "调整通告": WORK_TYPE_ADJUST,
    NOTICE_TYPE_EVENT: WORK_TYPE_EVENT,
}
END_SITE_PHOTO_REQUIRED_WORK_TYPES = {
    WORK_TYPE_MAINTENANCE,
    WORK_TYPE_CHANGE,
    WORK_TYPE_REPAIR,
}
END_SITE_PHOTO_REQUIRED_NOTICE_TYPES = {
    NOTICE_TYPE_MAINTENANCE,
    NOTICE_TYPE_CHANGE,
    NOTICE_HEADING_CHANGE,
    NOTICE_TYPE_REPAIR,
}
CHANGE_PROGRESS_NOT_STARTED = "未开始"
CHANGE_PROGRESS_ONGOING = "进行中"
CHANGE_PROGRESS_ENDED = "已结束"
CHANGE_WORKBENCH_PROGRESS_VALUES = {CHANGE_PROGRESS_NOT_STARTED, CHANGE_PROGRESS_ONGOING}
CHANGE_DEFAULT_LEVEL = "I3"
CHANGE_PROGRESS_OPTION_FALLBACK = {
    "optXQcI0z3": "未开始",
    "optqHR2ClY": "进行中",
    "optZtdlcZy": "已结束",
    "optqGq0YcR": "退回",
}
ZHIHANG_PROGRESS_ENDED = "已结束"
REPAIR_BUILDING_OPTION_FALLBACK = {
    "optjAjpfmQ": "A楼",
    "optHCxNMDk": "B楼",
    "opt2iUueLp": "C楼",
    "optTUZGlCA": "D楼",
    "opthJp8Wnt": "E楼",
}
REPAIR_SPECIALTY_OPTION_FALLBACK = {
    "optAssNaw3": "电气",
    "opt3jdJJb7": "暖通",
    "opt509Mxgr": "弱电",
    "optyLRPdQS": "消防",
}
FIELD_OPTION_FALLBACK = {
    "所属数据中心/楼栋-使用": REPAIR_BUILDING_OPTION_FALLBACK,
    "专业（推送消息用）": REPAIR_SPECIALTY_OPTION_FALLBACK,
}
REPAIR_DATETIME_FIELDS = {
    "故障发生时间",
    "发现故障时间",
    "维修开始时间",
    "维修结束时间",
    "维修结束时间（2026）",
    "期望完成时间",
}
PLACEHOLDER_TEXT_VALUES = {"", "-", "--", "—", "——", "/", "无", "暂无"}
SOURCE_CACHE_TTL_SECONDS = 30 * 60
RECENT_MONTH_FILTER_LABEL = "本月+上月"
STATE_NS_MEMORY = "notice_memory"
STATE_NS_DAILY_SUMMARY = "notice_daily_summary"
STATE_NS_WORK_STATUS = "notice_work_status"
STATE_NS_HIDDEN_ONGOING = "notice_hidden_ongoing"
STATE_NS_ACTION_JOB = "notice_action_job"
ACTION_JOB_MAX_RETAINED = 120
ACTION_JOB_SUCCESS_RETENTION_SECONDS = 30 * 60
ACTION_JOB_FAILED_RETENTION_SECONDS = 14 * 24 * 60 * 60


def external_mock_enabled() -> bool:
    return os.environ.get("CLIPFLOW_BACKEND_MOCK_EXTERNAL") == "1"


def _payload_list(payload: dict[str, Any], key: str) -> list[Any]:
    value = payload.get(key) if isinstance(payload, dict) else None
    return value if isinstance(value, list) else []


def engineer_mop_fill_kwargs_from_payload(
    payload: dict[str, Any], *, scope: str
) -> dict[str, Any]:
    payload = payload if isinstance(payload, dict) else {}
    return {
        "scope": str(scope or payload.get("scope") or "ALL"),
        "local_file_path": str(payload.get("local_file_path") or ""),
        "mop_record_id": str(payload.get("mop_record_id") or ""),
        "mop_title": str(payload.get("mop_title") or ""),
        "mop_file_name": str(payload.get("mop_file_name") or ""),
        "notice_key": str(payload.get("notice_key") or ""),
        "sheet_name": str(payload.get("sheet_name") or ""),
        "fields": _payload_list(payload, "fields"),
        "checkboxes": _payload_list(payload, "checkboxes"),
        "cell_edits": _payload_list(payload, "cell_edits"),
        "signatures": _payload_list(payload, "signatures"),
    }


def engineer_mop_upload_signed_kwargs_from_payload(
    payload: dict[str, Any],
    *,
    scope: str,
    operator_open_id: str = "",
    operator_name: str = "",
) -> dict[str, Any]:
    kwargs = engineer_mop_fill_kwargs_from_payload(payload, scope=scope)
    kwargs.update(
        {
            "source_record_id": str(payload.get("source_record_id") or ""),
            "notice_title": str(payload.get("notice_title") or ""),
            "notice_key": str(payload.get("notice_key") or ""),
            "operator_open_id": str(operator_open_id or ""),
            "operator_name": str(operator_name or ""),
        }
    )
    return kwargs


def external_real_write_guard() -> dict[str, Any]:
    require_confirm = os.environ.get("CLIPFLOW_REQUIRE_REAL_EXTERNAL_CONFIRM", "1") != "0"
    confirmed = os.environ.get("CLIPFLOW_REAL_EXTERNAL_CONFIRMED") == "1"
    mock = external_mock_enabled()
    allowed = (not mock) and (not require_confirm or confirmed)
    reason = ""
    if mock:
        reason = "当前为外部接口 mock 模式，已拦截真实飞书/多维写入。"
    elif require_confirm and not confirmed:
        reason = (
            "真实飞书/多维写入保护已启用，需设置 "
            "CLIPFLOW_REAL_EXTERNAL_CONFIRMED=1 后才允许执行。"
        )
    return {
        "mock_external": mock,
        "require_confirm": require_confirm,
        "confirmed": confirmed,
        "real_write_allowed": allowed,
        "reason": reason,
    }
LI_SHILONG_OPEN_ID = "ou_902e364a6c2c6c20893c02abe505a7b2"
MA_JINYU_OPEN_ID = "ou_a6644e62a43b916c6bc26148cf74f208"
HANDOVER_PASSWORD_RESET_TTL_SECONDS = 10 * 60
HANDOVER_PASSWORD_RESET_MAX_ATTEMPTS = 5
BUILDING_OPEN_ID_MAP = {
    "110": "ou_913ff495ae8f5a0a55ae5670fd148bbd",
    "A": "ou_3e0a9399a037276f8449cc778f18bb8a",
    "B": "ou_160e991b0fd811cda7260de056d8a363",
    "C": "ou_99391b00c1406917930483d8500af400",
    "D": "ou_b9c26b4da3721c2de11ee175c707e756",
    "E": "ou_f0fe7137818e8b422b0594e51eab13f3",
    "H": "ou_55647926b7fbfe46507ef7e8afa76315",
}
SCOPE_OPTIONS = [
    {"value": "110", "label": "110站"},
    {"value": "A", "label": "A楼"},
    {"value": "B", "label": "B楼"},
    {"value": "C", "label": "C楼"},
    {"value": "D", "label": "D楼"},
    {"value": "E", "label": "E楼"},
    {"value": "H", "label": "H楼"},
    {"value": "CAMPUS", "label": "园区"},
    {"value": "ALL", "label": "全部"},
]
BUILDING_SCOPE_CODES = ("110", "A", "B", "C", "D", "E", "H")


class PortalError(RuntimeError):
    pass


@dataclass
class FieldMeta:
    field_id: str
    field_name: str
    ui_type: str
    field_type: int
    is_primary: bool
    options_map: dict[str, str]
    option_names: list[str]
    has_formula: bool
    property: dict[str, Any] = dataclass_field(default_factory=dict)


class MaintenancePortalService:
    _memory_lock = threading.RLock()
    _summary_lock = threading.RLock()
    _handover_lock = threading.RLock()
    _handover_reset_lock = threading.RLock()
    _hidden_ongoing_lock = threading.RLock()

    def __init__(
        self,
        *,
        app_token: str = DEFAULT_APP_TOKEN,
        table_id: str = DEFAULT_TABLE_ID,
        enable_repair_snapshots: bool = False,
    ) -> None:
        self.app_token = str(app_token or DEFAULT_APP_TOKEN).strip()
        self.table_id = str(table_id or DEFAULT_TABLE_ID).strip()
        self._http_client = FeishuHttpClient()
        self._field_meta_list: list[FieldMeta] = []
        self._field_meta_by_name: dict[str, FieldMeta] = {}
        self._records: list[dict[str, Any]] = []
        self._change_field_meta_list: list[FieldMeta] = []
        self._change_field_meta_by_name: dict[str, FieldMeta] = {}
        self._change_records: list[dict[str, Any]] = []
        self._zhihang_change_field_meta_list: list[FieldMeta] = []
        self._zhihang_change_field_meta_by_name: dict[str, FieldMeta] = {}
        self._zhihang_change_records: list[dict[str, Any]] = []
        self._repair_field_meta_list: list[FieldMeta] = []
        self._repair_field_meta_by_name: dict[str, FieldMeta] = {}
        self._repair_records: list[dict[str, Any]] = []
        self._repair_management_target_cache_lock = threading.RLock()
        self._repair_management_target_cache: dict[str, Any] | None = None
        self._repair_management_event_cache_lock = threading.RLock()
        self._repair_management_event_cache: dict[str, Any] | None = None
        self._repair_management_cmdb_cache_lock = threading.RLock()
        self._repair_management_cmdb_cache: dict[str, Any] | None = None
        self._repair_management_routing_cache_lock = threading.RLock()
        self._repair_management_routing_cache: dict[str, Any] | None = None
        self._repair_management_status_cache_lock = threading.RLock()
        self._repair_management_status_load_lock = threading.Lock()
        self._repair_management_status_cache: dict[str, Any] | None = None
        self._repair_snapshots_enabled = bool(enable_repair_snapshots)
        self._repair_snapshot_locks_guard = threading.RLock()
        self._repair_snapshot_locks: dict[str, threading.Lock] = {}
        self._repair_snapshot_refreshing: set[str] = set()
        self._repair_followup_schema_lock = threading.RLock()
        self._repair_followup_schema_ready = False
        self._repair_management_progress_schema_ready = False
        self._repair_followup_sync_locks_guard = threading.RLock()
        self._repair_followup_sync_locks: dict[str, threading.RLock] = {}
        self._repair_followup_backfill_lock = threading.RLock()
        self._repair_followup_catalog_cache_lock = threading.RLock()
        self._repair_followup_catalog_cache: dict[str, Any] | None = None
        self._maintenance_loaded_once = False
        self._change_loaded_once = False
        self._zhihang_change_loaded_once = False
        self._repair_loaded_once = False
        self._load_warnings: list[str] = []
        self._last_loaded_at = ""
        self._last_loaded_ts = 0.0
        self._refresh_lock = threading.RLock()
        # Legacy JSON paths are read only for one-time migration. Do not create
        # their directories on fresh SQLite-only installs.
        self._memory_dir = Path(get_data_file_path("lan_template_memory"))
        self._summary_dir = Path(get_data_file_path("lan_template_daily_summary"))
        self._work_status_dir = Path(get_data_file_path("lan_template_work_status"))
        self._handover_links_path = Path(get_data_file_path("lan_handover_links.json"))
        self._hidden_ongoing_path = Path(get_data_file_path("lan_template_hidden_ongoing.json"))
        self._state_store = LanPortalStateStore()
        self._state_version_lock = threading.RLock()
        self._state_version = 0
        self._legacy_summary_migrated = False
        self._legacy_work_status_migrated = False
        self._work_status_backfilled = False
        self._work_status_cache_signature: tuple[tuple[str, int], ...] | None = None
        self._work_status_cache_items: list[dict[str, Any]] | None = None
        self._target_record_cache: dict[tuple[str, str], dict[str, Any]] = {}
        self._engineer_mop_cache_lock = threading.RLock()
        self._engineer_mop_cache: dict[str, Any] | None = None
        self._signature_people_cache_lock = threading.RLock()
        self._signature_people_cache: dict[str, Any] | None = None
        self._external_signature_people_cache_lock = threading.RLock()
        self._external_signature_people_cache: dict[str, Any] | None = None
        self._signature_crypto = SignatureCryptoManager()
        self._signature_crypto_migration_lock = threading.RLock()
        self._signature_crypto_migration_running = False
        self._signature_crypto_plain_migrations: set[tuple[str, str]] = set()
        self._attachment_cache_lock = threading.RLock()
        self._attachment_cache_refresh_lock = threading.RLock()
        self._attachment_cache_refresh_running = False
        self._jobs_lock = threading.RLock()
        self._jobs: dict[str, dict[str, Any]] = self._load_action_jobs_from_state()
        self._handover_password_reset: dict[str, Any] | None = None
        self._ensure_signature_crypto_ready()

    def _ensure_signature_crypto_ready(self) -> None:
        try:
            fingerprint = self._signature_crypto.master_key_fingerprint()
            with suppress(Exception):
                self._state_store.put_settings(
                    {"signature_master_key_fingerprint": fingerprint}
                )
            with suppress(Exception):
                self._state_store.put_backend_runtime(
                    "signature_crypto",
                    {
                        "status": "ready",
                        "master_key_exists": True,
                        "fingerprint": fingerprint,
                        "master_key_path": str(self._signature_crypto.master_key_path),
                        "updated_at": dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    },
                )

        except Exception as exc:
            with suppress(Exception):
                self._state_store.put_backend_runtime(
                    "signature_crypto",
                    {
                        "status": "failed",
                        "master_key_exists": self._signature_crypto.master_key_exists(),
                        "error": str(exc),
                        "updated_at": dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    },
                )

    def _invalidate_repair_management_target_cache(self) -> None:
        with self._repair_management_target_cache_lock:
            self._repair_management_target_cache = None

    def _invalidate_repair_management_event_cache(self) -> None:
        with self._repair_management_event_cache_lock:
            self._repair_management_event_cache = None

    def _invalidate_repair_followup_catalog_cache(self) -> None:
        with self._repair_followup_catalog_cache_lock:
            self._repair_followup_catalog_cache = None

    def _invalidate_repair_management_status_cache(self) -> None:
        with self._repair_management_status_cache_lock:
            self._repair_management_status_cache = None

    @staticmethod
    def _repair_snapshot_field_payload(meta: FieldMeta) -> dict[str, Any]:
        return {
            "field_id": str(meta.field_id or ""),
            "field_name": str(meta.field_name or ""),
            "ui_type": str(meta.ui_type or ""),
            "field_type": int(meta.field_type or 0),
            "is_primary": bool(meta.is_primary),
            "options_map": dict(meta.options_map or {}),
            "option_names": list(meta.option_names or []),
            "has_formula": bool(meta.has_formula),
            "property": dict(meta.property or {}),
        }

    @staticmethod
    def _repair_snapshot_field_meta(payload: dict[str, Any]) -> FieldMeta:
        return FieldMeta(
            field_id=str(payload.get("field_id") or ""),
            field_name=str(payload.get("field_name") or ""),
            ui_type=str(payload.get("ui_type") or ""),
            field_type=int(payload.get("field_type") or 0),
            is_primary=bool(payload.get("is_primary")),
            options_map={
                str(key): str(value)
                for key, value in dict(payload.get("options_map") or {}).items()
            },
            option_names=[
                str(value)
                for value in list(payload.get("option_names") or [])
                if str(value or "").strip()
            ],
            has_formula=bool(payload.get("has_formula")),
            property=dict(payload.get("property") or {}),
        )

    def _repair_snapshot_lock(self, source_key: str) -> threading.Lock:
        with self._repair_snapshot_locks_guard:
            lock = self._repair_snapshot_locks.get(source_key)
            if lock is None:
                lock = threading.Lock()
                self._repair_snapshot_locks[source_key] = lock
            return lock

    @classmethod
    def _repair_snapshot_record_payload(
        cls,
        record: dict[str, Any],
        *,
        parent_field_name: str = "",
    ) -> dict[str, Any]:
        display_fields = (
            record.get("display_fields")
            if isinstance(record.get("display_fields"), dict)
            else {}
        )
        record_id = str(
            record.get("record_id") or record.get("source_record_id") or ""
        ).strip()
        parent_record_id = ""
        if parent_field_name:
            parent_record_id = cls._repair_management_plain_text(
                display_fields.get(parent_field_name)
            ).strip()
        title = cls._repair_management_plain_text(
            display_fields.get("维修名称")
            or display_fields.get("名称（标题）")
            or display_fields.get("检修概述")
            or display_fields.get("事件简述")
            or display_fields.get("告警描述")
            or display_fields.get("跟进项")
            or display_fields.get("维修进展描述")
            or display_fields.get("设备名称")
            or ""
        ).strip()
        status = cls._repair_management_plain_text(
            display_fields.get("流程")
            or display_fields.get("维修进度")
            or display_fields.get("当前维修进度")
            or display_fields.get("检修状态")
            or display_fields.get("事件状态")
            or display_fields.get("最终状态")
            or ""
        ).strip()
        building_value = cls._repair_management_plain_text(
            display_fields.get("所属数据中心/楼栋-使用")
            or display_fields.get("楼栋")
            or display_fields.get("机楼")
            or display_fields.get("南通楼栋")
            or title
        )
        scope_codes = cls._repair_building_codes_from_value(building_value)
        sort_value = (
            display_fields.get("最新维修跟进时间")
            or display_fields.get("最后修改时间")
            or display_fields.get("进展更新时间")
            or display_fields.get("事件发生时间")
            or display_fields.get("故障发生时间")
            or display_fields.get("创建时间")
            or ""
        )
        sort_dt = cls._parse_notice_datetime(
            cls._repair_management_plain_text(sort_value)
        )
        search_text = " ".join(
            cls._repair_management_plain_text(value)
            for value in display_fields.values()
        ).strip()
        return {
            "record_id": record_id,
            "parent_record_id": parent_record_id,
            "scope_codes": scope_codes,
            "title": title,
            "status": status,
            "search_text": search_text,
            "sort_time": sort_dt.timestamp() if sort_dt else 0.0,
            "payload": dict(record),
        }

    def _repair_snapshot_from_local(
        self, snapshot: dict[str, Any]
    ) -> tuple[list[FieldMeta], dict[str, FieldMeta], list[dict[str, Any]]]:
        metas = [
            self._repair_snapshot_field_meta(item)
            for item in (snapshot.get("fields") or [])
            if isinstance(item, dict)
        ]
        records = [
            dict(item)
            for item in (snapshot.get("records") or [])
            if isinstance(item, dict)
        ]
        return metas, {meta.field_name: meta for meta in metas}, records

    def _refresh_repair_snapshot_source(
        self,
        *,
        source_key: str,
        app_token: str,
        table_id: str,
        loader: Callable[[], tuple[list[FieldMeta], list[dict[str, Any]]]],
        parent_field_name: str = "",
        force_refresh: bool = False,
    ) -> tuple[list[FieldMeta], dict[str, FieldMeta], list[dict[str, Any]]]:
        ttl = float(REPAIR_SNAPSHOT_TTL_SECONDS.get(source_key) or 120)
        lock = self._repair_snapshot_lock(source_key)
        with lock:
            current = self._state_store.get_repair_snapshot(source_key)
            refreshed_at = float(current.get("refreshed_at") or 0)
            if (
                not force_refresh
                and refreshed_at > 0
                and time.time() - refreshed_at <= ttl
            ):
                return self._repair_snapshot_from_local(current)
            try:
                metas, records = loader()
                self._state_store.replace_repair_snapshot(
                    source_key,
                    records=[
                        self._repair_snapshot_record_payload(
                            record,
                            parent_field_name=parent_field_name,
                        )
                        for record in records
                        if isinstance(record, dict)
                    ],
                    app_token=app_token,
                    table_id=table_id,
                    fields=[
                        self._repair_snapshot_field_payload(meta) for meta in metas
                    ],
                    meta={"source_key": source_key},
                )
                return metas, {meta.field_name: meta for meta in metas}, records
            except Exception as exc:
                self._state_store.mark_repair_snapshot_failed(source_key, str(exc))
                stale = self._state_store.get_repair_snapshot(source_key)
                if stale.get("records"):
                    return self._repair_snapshot_from_local(stale)
                raise

    def _schedule_repair_snapshot_refresh(
        self,
        *,
        source_key: str,
        app_token: str,
        table_id: str,
        loader: Callable[[], tuple[list[FieldMeta], list[dict[str, Any]]]],
        parent_field_name: str = "",
    ) -> None:
        with self._repair_snapshot_locks_guard:
            if source_key in self._repair_snapshot_refreshing:
                return
            self._repair_snapshot_refreshing.add(source_key)

        def run() -> None:
            try:
                self._refresh_repair_snapshot_source(
                    source_key=source_key,
                    app_token=app_token,
                    table_id=table_id,
                    loader=loader,
                    parent_field_name=parent_field_name,
                )
                self._invalidate_repair_management_status_cache()
            except Exception:
                pass
            finally:
                with self._repair_snapshot_locks_guard:
                    self._repair_snapshot_refreshing.discard(source_key)

        threading.Thread(
            target=run,
            name=f"RepairSnapshot-{source_key}",
            daemon=True,
        ).start()

    def _load_repair_snapshot_source(
        self,
        *,
        source_key: str,
        app_token: str,
        table_id: str,
        loader: Callable[[], tuple[list[FieldMeta], list[dict[str, Any]]]],
        parent_field_name: str = "",
        force_refresh: bool = False,
    ) -> tuple[list[FieldMeta], dict[str, FieldMeta], list[dict[str, Any]]]:
        if not self._repair_snapshots_enabled:
            metas, records = loader()
            return metas, {meta.field_name: meta for meta in metas}, records
        snapshot = self._state_store.get_repair_snapshot(source_key)
        now = time.time()
        refreshed_at = float(snapshot.get("refreshed_at") or 0)
        ttl = float(REPAIR_SNAPSHOT_TTL_SECONDS.get(source_key) or 120)
        if (
            not force_refresh
            and refreshed_at > 0
            and now - refreshed_at <= ttl
        ):
            return self._repair_snapshot_from_local(snapshot)
        if not force_refresh and refreshed_at > 0 and snapshot.get("records"):
            failure_backoff = (
                str(snapshot.get("status") or "") == "failed"
                and now - float(snapshot.get("updated_at") or 0)
                < REPAIR_SNAPSHOT_FAILURE_BACKOFF_SECONDS
            )
            if not failure_backoff:
                self._schedule_repair_snapshot_refresh(
                    source_key=source_key,
                    app_token=app_token,
                    table_id=table_id,
                    loader=loader,
                    parent_field_name=parent_field_name,
                )
            return self._repair_snapshot_from_local(snapshot)
        return self._refresh_repair_snapshot_source(
            source_key=source_key,
            app_token=app_token,
            table_id=table_id,
            loader=loader,
            parent_field_name=parent_field_name,
            force_refresh=force_refresh,
        )

    def _upsert_repair_snapshot_fields(
        self,
        *,
        source_key: str,
        record_id: str,
        fields: dict[str, Any],
        parent_record_id: str = "",
    ) -> None:
        if not self._repair_snapshots_enabled:
            return
        normalized_record_id = str(record_id or "").strip()
        if not normalized_record_id:
            return
        snapshot = self._state_store.get_repair_snapshot(
            source_key,
            record_ids=[normalized_record_id],
        )
        existing = next(
            (
                dict(item)
                for item in (snapshot.get("records") or [])
                if isinstance(item, dict)
            ),
            {},
        )
        existing["record_id"] = normalized_record_id
        display_fields = (
            dict(existing.get("display_fields") or {})
            if isinstance(existing.get("display_fields"), dict)
            else {}
        )
        raw_fields = (
            dict(existing.get("raw_fields") or {})
            if isinstance(existing.get("raw_fields"), dict)
            else {}
        )
        display_fields.update(dict(fields or {}))
        raw_fields.update(dict(fields or {}))
        existing["display_fields"] = display_fields
        existing["raw_fields"] = raw_fields
        existing.setdefault("created_time", str(int(time.time() * 1000)))
        existing["last_modified_time"] = str(int(time.time() * 1000))
        envelope = self._repair_snapshot_record_payload(
            existing,
            parent_field_name=(
                REPAIR_FOLLOWUP_PARENT_ID_FIELD_NAME
                if source_key == REPAIR_SNAPSHOT_SOURCE_FOLLOWUPS
                else ""
            ),
        )
        effective_parent = str(parent_record_id or "").strip()
        if effective_parent:
            envelope["parent_record_id"] = effective_parent
        self._state_store.upsert_repair_snapshot_record(
            source_key,
            normalized_record_id,
            existing,
            parent_record_id=str(envelope.get("parent_record_id") or ""),
            scope_codes=list(envelope.get("scope_codes") or []),
            title=str(envelope.get("title") or ""),
            status=str(envelope.get("status") or ""),
            search_text=str(envelope.get("search_text") or ""),
            sort_time=float(envelope.get("sort_time") or 0),
        )

    def _delete_repair_snapshot_item(self, source_key: str, record_id: str) -> None:
        if not self._repair_snapshots_enabled:
            return
        self._state_store.delete_repair_snapshot_record(source_key, record_id)

    def start_repair_snapshot_warmup_async(self, *, delay_seconds: float = 2.0) -> None:
        if not self._repair_snapshots_enabled:
            return
        warmup_key = "__warmup__"
        with self._repair_snapshot_locks_guard:
            if warmup_key in self._repair_snapshot_refreshing:
                return
            self._repair_snapshot_refreshing.add(warmup_key)

        def run() -> None:
            try:
                if delay_seconds > 0:
                    time.sleep(float(delay_seconds))
                loaders: tuple[Callable[[], Any], ...] = (
                    self._load_repair_management_project_records,
                    self._load_repair_followup_snapshot,
                    self._load_repair_management_event_records,
                    self._load_repair_management_target_records,
                    self._load_repair_management_cmdb_records,
                    self._load_repair_management_routing_records,
                )
                for loader in loaders:
                    with suppress(Exception):
                        loader()
            finally:
                with self._repair_snapshot_locks_guard:
                    self._repair_snapshot_refreshing.discard(warmup_key)

        threading.Thread(
            target=run,
            name="RepairSnapshotWarmup",
            daemon=True,
        ).start()

    def state_cache_version(self) -> int:
        with self._state_version_lock:
            return int(self._state_version)

    def signature_crypto_status(self) -> dict[str, Any]:
        runtime = self._state_store.get_backend_runtime("signature_crypto") or {}
        summary = self._state_store.signature_crypto_migration_summary()
        try:
            fingerprint = self._signature_crypto.master_key_fingerprint()
            key_status = "ready"
            error = ""
        except Exception as exc:
            fingerprint = str(runtime.get("fingerprint") or "")
            key_status = "failed"
            error = str(exc)
        return {
            "status": key_status if key_status == "failed" else str(runtime.get("status") or "ready"),
            "master_key_exists": self._signature_crypto.master_key_exists(),
            "fingerprint": fingerprint,
            "master_key_path": str(self._signature_crypto.master_key_path),
            "migration": summary,
            "error": error or str(runtime.get("error") or ""),
            "updated_at": str(runtime.get("updated_at") or ""),
        }

    def _touch_state_cache_version(self) -> None:
        with self._state_version_lock:
            self._state_version += 1

    def _auth_headers(self) -> dict[str, str]:
        token = str(ensure_feishu_token() or config.user_token or "").strip()
        if not token:
            raise PortalError("未配置有效的飞书 user_token。")
        return {"Authorization": f"Bearer {token}"}

    def _request_payload(
        self,
        method: str,
        url: str,
        *,
        context: str,
        headers: dict[str, str],
        params: dict[str, Any] | None = None,
        json_payload: Any = None,
    ) -> dict[str, Any]:
        try:
            return self._http_client.request_json(
                method,
                url,
                headers=headers,
                params=params or {},
                json_payload=json_payload,
            )
        except FeishuHTTPError as exc:
            raise PortalError(f"{context} HTTP失败: {exc}") from exc

    def _request_json(
        self,
        path: str,
        *,
        params: dict[str, Any] | None = None,
        app_token: str | None = None,
        table_id: str | None = None,
    ) -> dict[str, Any]:
        app_token = str(app_token or self.app_token or DEFAULT_APP_TOKEN).strip()
        table_id = str(table_id or self.table_id or DEFAULT_TABLE_ID).strip()
        if not app_token or not table_id:
            raise PortalError("飞书多维请求缺少 app_token/table_id，已阻止空源表请求。")
        url = (
            f"https://open.feishu.cn/open-apis/bitable/v1/apps/"
            f"{app_token}/tables/{table_id}/{path}"
        )

        def do_get() -> dict[str, Any]:
            return self._request_payload(
                "GET",
                url,
                context="飞书接口",
                headers=self._auth_headers(),
                params=params or {},
            )

        payload: dict[str, Any] = {}
        code = 0
        for attempt in range(len(BITABLE_TRANSIENT_RETRY_DELAYS) + 1):
            payload = do_get()
            code = int(payload.get("code") or 0)
            if code in TOKEN_ERROR_CODES:
                refresh_feishu_token()
                payload = do_get()
                code = int(payload.get("code") or 0)
            if (
                code == BITABLE_DATA_NOT_READY_CODE
                and attempt < len(BITABLE_TRANSIENT_RETRY_DELAYS)
            ):
                time.sleep(BITABLE_TRANSIENT_RETRY_DELAYS[attempt])
                continue
            break
        if code != 0:
            if code == BITABLE_DATA_NOT_READY_CODE:
                raise PortalError(
                    "飞书多维表正在计算，数据暂时未准备好，请稍后重试。"
                )
            raise PortalError(
                f"飞书接口失败: code={code}, msg={payload.get('msg') or 'unknown'}"
            )
        return payload

    def _patch_record_fields(
        self,
        *,
        app_token: str,
        table_id: str,
        record_id: str,
        fields: dict[str, Any],
    ) -> dict[str, Any]:
        app_token = str(app_token or "").strip()
        table_id = str(table_id or "").strip()
        record_id = str(record_id or "").strip()
        if not app_token or not table_id or not record_id:
            raise PortalError("更新多维记录缺少 app_token/table_id/record_id。")
        if not isinstance(fields, dict) or not fields:
            raise PortalError("更新多维记录字段不能为空。")
        url = (
            f"https://open.feishu.cn/open-apis/bitable/v1/apps/"
            f"{app_token}/tables/{table_id}/records/{record_id}"
        )

        def do_update() -> dict[str, Any]:
            return self._request_payload(
                "PUT",
                url,
                context="飞书记录更新",
                headers={**self._auth_headers(), "Content-Type": "application/json"},
                json_payload={"fields": fields},
            )

        payload = do_update()
        if int(payload.get("code") or 0) in TOKEN_ERROR_CODES:
            refresh_feishu_token()
            payload = do_update()
        code = payload.get("code", 0)
        if code != 0:
            raise PortalError(
                f"飞书记录更新失败: code={code}, msg={payload.get('msg') or 'unknown'}"
            )
        if table_id == REPAIR_MANAGEMENT_REPAIR_TABLE_ID:
            self._invalidate_repair_management_target_cache()
        if table_id == REPAIR_FOLLOWUP_TABLE_ID:
            self._invalidate_repair_followup_catalog_cache()
        return payload

    @staticmethod
    def _repair_followup_record_value(
        record: dict[str, Any],
        field_name: str,
    ) -> Any:
        display = (
            record.get("display_fields")
            if isinstance(record.get("display_fields"), dict)
            else {}
        )
        raw = (
            record.get("raw_fields")
            if isinstance(record.get("raw_fields"), dict)
            else {}
        )
        value = display.get(field_name)
        if value in (None, "", [], {}):
            value = raw.get(field_name)
        return value

    @staticmethod
    def _repair_followup_option_specs(
        raw_field: dict[str, Any] | None,
        *,
        keep_ids: bool,
    ) -> list[dict[str, Any]]:
        if not isinstance(raw_field, dict):
            return []
        options = (raw_field.get("property") or {}).get("options") or []
        result: list[dict[str, Any]] = []
        seen: set[str] = set()
        for option in options:
            if not isinstance(option, dict):
                continue
            name = str(option.get("name") or "").strip()
            key = name.casefold()
            if not name or key in seen:
                continue
            item: dict[str, Any] = {"name": name}
            if keep_ids and str(option.get("id") or "").strip():
                item["id"] = str(option.get("id") or "").strip()
            color = option.get("color")
            if isinstance(color, int):
                item["color"] = color
            result.append(item)
            seen.add(key)
        return result

    def _update_bitable_field_definition(
        self,
        raw_field: dict[str, Any],
        *,
        app_token: str,
        table_id: str,
        context: str,
        field_type: int,
        ui_type: str | None = None,
        property_payload: dict[str, Any] | None = None,
        options: list[dict[str, Any]] | None = None,
    ) -> None:
        field_name = str(raw_field.get("field_name") or "").strip()
        field_id = str(raw_field.get("field_id") or "").strip()
        if not field_name or not field_id:
            raise PortalError(f"{context}缺少字段名称或 field_id。")
        request_body: dict[str, Any] = {
            "field_name": field_name,
            "type": int(field_type),
        }
        if int(field_type) == 3:
            request_body["ui_type"] = str(ui_type or "SingleSelect")
            request_body["property"] = {"options": list(options or [])}
        else:
            effective_ui_type = str(
                ui_type if ui_type is not None else raw_field.get("ui_type") or ""
            ).strip()
            if effective_ui_type:
                request_body["ui_type"] = effective_ui_type
            effective_property = copy.deepcopy(
                property_payload
                if property_payload is not None
                else raw_field.get("property") or {}
            )
            if effective_property:
                request_body["property"] = effective_property
        description = raw_field.get("description")
        if isinstance(description, str):
            request_body["description"] = description
        url = (
            "https://open.feishu.cn/open-apis/bitable/v1/apps/"
            f"{app_token}/tables/{table_id}/"
            f"fields/{field_id}"
        )

        def do_update() -> dict[str, Any]:
            return self._request_payload(
                "PUT",
                url,
                context=f"{context}({field_name})",
                headers={**self._auth_headers(), "Content-Type": "application/json"},
                json_payload=request_body,
            )

        payload = do_update()
        if int(payload.get("code") or 0) in TOKEN_ERROR_CODES:
            refresh_feishu_token()
            payload = do_update()
        code = int(payload.get("code") or 0)
        if code != 0:
            raise PortalError(
                f"{context}{field_name}失败: "
                f"code={code}, msg={payload.get('msg') or 'unknown'}"
            )

    def _update_repair_followup_field_definition(
        self,
        raw_field: dict[str, Any],
        *,
        field_type: int,
        options: list[dict[str, Any]] | None = None,
    ) -> None:
        self._update_bitable_field_definition(
            raw_field,
            app_token=REPAIR_SOURCE_APP_TOKEN,
            table_id=REPAIR_FOLLOWUP_TABLE_ID,
            context="迁移维修跟进字段",
            field_type=field_type,
            options=options,
        )

    def _ensure_repair_management_progress_field_definition(
        self,
        summary_meta_by_name: dict[str, FieldMeta],
        followup_meta_by_name: dict[str, FieldMeta],
    ) -> bool:
        if self._repair_management_progress_schema_ready:
            return False
        summary_meta = summary_meta_by_name.get("当前维修进度")
        followup_meta = followup_meta_by_name.get("维修进度")
        if summary_meta is None or followup_meta is None:
            return False
        if int(followup_meta.field_type or 0) != 2:
            raise PortalError("维修跟进表的维修进度必须是进度或数字字段。")

        desired_ui_type = str(followup_meta.ui_type or "Progress").strip() or "Progress"
        desired_property = copy.deepcopy(followup_meta.property or {})
        current_property = copy.deepcopy(summary_meta.property or {})
        needs_update = (
            int(summary_meta.field_type or 0) != int(followup_meta.field_type or 0)
            or str(summary_meta.ui_type or "").strip() != desired_ui_type
            or current_property != desired_property
        )
        if needs_update:
            self._update_bitable_field_definition(
                {
                    "field_id": summary_meta.field_id,
                    "field_name": summary_meta.field_name,
                    "ui_type": summary_meta.ui_type,
                    "property": current_property,
                },
                app_token=REPAIR_SOURCE_APP_TOKEN,
                table_id=REPAIR_MANAGEMENT_TABLE_ID,
                context="同步维修项目进度字段格式",
                field_type=int(followup_meta.field_type or 2),
                ui_type=desired_ui_type,
                property_payload=desired_property,
            )
        self._repair_management_progress_schema_ready = True
        with suppress(Exception):
            self._state_store.put_backend_runtime(
                REPAIR_MANAGEMENT_PROGRESS_SCHEMA_RUNTIME_KEY,
                {
                    "status": "complete",
                    "updated": needs_update,
                    "field_type": int(followup_meta.field_type or 2),
                    "ui_type": desired_ui_type,
                    "property": desired_property,
                    "finished_at": dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                },
            )
        return needs_update

    def _delete_repair_followup_field_definition(
        self,
        raw_field: dict[str, Any],
    ) -> None:
        field_name = str(raw_field.get("field_name") or "").strip()
        field_id = str(raw_field.get("field_id") or "").strip()
        if not field_name or not field_id:
            raise PortalError("删除维修跟进重复字段时缺少字段名称或 field_id。")
        url = (
            "https://open.feishu.cn/open-apis/bitable/v1/apps/"
            f"{REPAIR_SOURCE_APP_TOKEN}/tables/{REPAIR_FOLLOWUP_TABLE_ID}/"
            f"fields/{field_id}"
        )

        def do_delete() -> dict[str, Any]:
            return self._request_payload(
                "DELETE",
                url,
                context=f"维修跟进重复字段删除({field_name})",
                headers=self._auth_headers(),
            )

        payload = do_delete()
        if int(payload.get("code") or 0) in TOKEN_ERROR_CODES:
            refresh_feishu_token()
            payload = do_delete()
        code = int(payload.get("code") or 0)
        if code != 0:
            raise PortalError(
                f"删除维修跟进重复字段{field_name}失败: "
                f"code={code}, msg={payload.get('msg') or 'unknown'}"
            )

    @classmethod
    def _repair_followup_migration_text(
        cls,
        value: Any,
        legacy_meta: FieldMeta | None,
    ) -> str:
        text = cls._repair_management_plain_text(value).strip()
        if not text:
            return ""
        if re.fullmatch(r"opt[A-Za-z0-9]+", text):
            mapped = str((legacy_meta.options_map or {}).get(text) or "").strip() \
                if legacy_meta is not None else ""
            if not mapped:
                raise PortalError(f"旧选项 ID {text} 无法还原为文本，已停止字段迁移。")
            return mapped
        return text

    @staticmethod
    def _repair_followup_merge_option_specs(
        *groups: list[dict[str, Any]],
    ) -> list[dict[str, Any]]:
        result: list[dict[str, Any]] = []
        seen: set[str] = set()
        for group in groups:
            for option in group:
                name = str(option.get("name") or "").strip()
                key = name.casefold()
                if not name or key in seen:
                    continue
                result.append(copy.deepcopy(option))
                seen.add(key)
        return result

    def _migrate_repair_followup_duplicate_fields_unlocked(
        self,
        metas: list[FieldMeta],
        meta_by_name: dict[str, FieldMeta],
    ) -> tuple[list[FieldMeta], dict[str, FieldMeta]]:
        raw_payload = self._request_json(
            "fields",
            params={"page_size": 500},
            app_token=REPAIR_SOURCE_APP_TOKEN,
            table_id=REPAIR_FOLLOWUP_TABLE_ID,
        )
        raw_fields = [
            item
            for item in (raw_payload.get("data", {}).get("items") or [])
            if isinstance(item, dict)
        ]
        raw_by_name = {
            str(item.get("field_name") or "").strip(): item
            for item in raw_fields
        }
        legacy_names = [
            legacy_name
            for legacy_name, _canonical_name in REPAIR_FOLLOWUP_LEGACY_FIELD_MAPPINGS
            if legacy_name in raw_by_name
        ]
        expected_types = {
            REPAIR_FOLLOWUP_DEVICE_NAME_FIELD_NAME: 3,
            REPAIR_FOLLOWUP_DEVICE_NUMBER_FIELD_NAME: 1,
            REPAIR_FOLLOWUP_BRAND_FIELD_NAME: 3,
            REPAIR_FOLLOWUP_MODEL_FIELD_NAME: 3,
            REPAIR_FOLLOWUP_REPAIR_PARTY_FIELD_NAME: 3,
            REPAIR_FOLLOWUP_SUPPLIER_FIELD_NAME: 1,
            REPAIR_FOLLOWUP_SUPPLIER_PERSON_FIELD_NAME: 1,
            "是否质保期内": 3,
        }
        missing = [name for name in expected_types if name not in raw_by_name]
        if missing:
            raise PortalError(
                "维修跟进表缺少保留字段：" + "、".join(missing)
            )
        if not legacy_names:
            invalid = [
                f"{name}(当前类型{int(meta_by_name[name].field_type or 0)})"
                for name, field_type in expected_types.items()
                if int(meta_by_name[name].field_type or 0) != field_type
            ]
            if invalid:
                raise PortalError(
                    "维修跟进字段迁移不完整：" + "、".join(invalid)
                )
            current = self._state_store.get_backend_runtime(
                REPAIR_FOLLOWUP_SCHEMA_MIGRATION_RUNTIME_KEY
            ) or {}
            if current.get("status") != "complete":
                self._state_store.put_backend_runtime(
                    REPAIR_FOLLOWUP_SCHEMA_MIGRATION_RUNTIME_KEY,
                    {
                        "status": "complete",
                        "migrated_records": int(current.get("migrated_records") or 0),
                        "deleted_fields": list(current.get("deleted_fields") or []),
                        "finished_at": dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    },
                )
            return metas, meta_by_name

        records = self._load_table_records(
            app_token=REPAIR_SOURCE_APP_TOKEN,
            table_id=REPAIR_FOLLOWUP_TABLE_ID,
            meta_by_name=meta_by_name,
            work_type=WORK_TYPE_REPAIR,
            notice_type=NOTICE_TYPE_REPAIR,
        )
        desired_by_record: dict[str, dict[str, str]] = {}
        for record in records:
            record_id = str(record.get("record_id") or "").strip()
            if not record_id:
                continue
            desired: dict[str, str] = {}
            for legacy_name, canonical_name in REPAIR_FOLLOWUP_LEGACY_FIELD_MAPPINGS:
                legacy_value = self._repair_followup_record_value(record, legacy_name)
                canonical_value = self._repair_followup_record_value(record, canonical_name)
                source_value = (
                    legacy_value
                    if legacy_value not in (None, "", [], {})
                    else canonical_value
                )
                text = self._repair_followup_migration_text(
                    source_value,
                    meta_by_name.get(legacy_name),
                )
                if text:
                    desired[canonical_name] = text
            desired_by_record[record_id] = desired

        desired_options: dict[str, list[dict[str, Any]]] = {}
        for canonical_name, legacy_name in REPAIR_FOLLOWUP_CANONICAL_SELECT_SOURCES.items():
            legacy_specs = self._repair_followup_option_specs(
                raw_by_name.get(legacy_name),
                keep_ids=False,
            )
            current_specs = [
                option
                for option in self._repair_followup_option_specs(
                    raw_by_name.get(canonical_name),
                    keep_ids=True,
                )
                if not re.fullmatch(
                    r"opt[A-Za-z0-9]+",
                    str(option.get("name") or "").strip(),
                )
            ]
            extra_specs = [
                {"name": values[canonical_name]}
                for values in desired_by_record.values()
                if str(values.get(canonical_name) or "").strip()
            ]
            desired_options[canonical_name] = self._repair_followup_merge_option_specs(
                legacy_specs,
                current_specs,
                extra_specs,
            )
            if not desired_options[canonical_name]:
                raise PortalError(f"{canonical_name}没有可迁移的下拉选项。")

        # Persist text targets first. This removes legacy option IDs before any
        # in-place Text -> SingleSelect conversion is attempted.
        for record in records:
            record_id = str(record.get("record_id") or "").strip()
            desired = desired_by_record.get(record_id) or {}
            patch: dict[str, Any] = {}
            for canonical_name, value in desired.items():
                meta = meta_by_name.get(canonical_name)
                if meta is None or int(meta.field_type or 0) != 1:
                    continue
                current = self._repair_followup_migration_text(
                    self._repair_followup_record_value(record, canonical_name),
                    meta_by_name.get(
                        REPAIR_FOLLOWUP_CANONICAL_SELECT_SOURCES.get(
                            canonical_name,
                            "",
                        )
                    ),
                )
                if current != value:
                    patch[canonical_name] = value
            if patch:
                self._patch_record_fields(
                    app_token=REPAIR_SOURCE_APP_TOKEN,
                    table_id=REPAIR_FOLLOWUP_TABLE_ID,
                    record_id=record_id,
                    fields=patch,
                )

        # Existing single-select targets first receive a merged option set so
        # records can be normalized without invalid-option failures.
        for canonical_name, options in desired_options.items():
            raw_field = raw_by_name[canonical_name]
            if int(raw_field.get("type") or 0) != 3:
                continue
            current_specs = self._repair_followup_option_specs(raw_field, keep_ids=True)
            merged_specs = self._repair_followup_merge_option_specs(
                current_specs,
                options,
            )
            current_names = [str(item.get("name") or "") for item in current_specs]
            merged_names = [str(item.get("name") or "") for item in merged_specs]
            if current_names != merged_names:
                self._update_repair_followup_field_definition(
                    raw_field,
                    field_type=3,
                    options=merged_specs,
                )

        metas, meta_by_name = self._load_table_fields(
            app_token=REPAIR_SOURCE_APP_TOKEN,
            table_id=REPAIR_FOLLOWUP_TABLE_ID,
        )
        records = self._load_table_records(
            app_token=REPAIR_SOURCE_APP_TOKEN,
            table_id=REPAIR_FOLLOWUP_TABLE_ID,
            meta_by_name=meta_by_name,
            work_type=WORK_TYPE_REPAIR,
            notice_type=NOTICE_TYPE_REPAIR,
        )
        for record in records:
            record_id = str(record.get("record_id") or "").strip()
            desired = desired_by_record.get(record_id) or {}
            patch = {
                name: value
                for name, value in desired.items()
                if self._repair_management_plain_text(
                    self._repair_followup_record_value(record, name)
                ).strip() != value
            }
            if patch:
                self._patch_record_fields(
                    app_token=REPAIR_SOURCE_APP_TOKEN,
                    table_id=REPAIR_FOLLOWUP_TABLE_ID,
                    record_id=record_id,
                    fields=patch,
                )

        raw_payload = self._request_json(
            "fields",
            params={"page_size": 500},
            app_token=REPAIR_SOURCE_APP_TOKEN,
            table_id=REPAIR_FOLLOWUP_TABLE_ID,
        )
        raw_fields = [
            item
            for item in (raw_payload.get("data", {}).get("items") or [])
            if isinstance(item, dict)
        ]
        raw_by_name = {
            str(item.get("field_name") or "").strip(): item
            for item in raw_fields
        }
        for canonical_name, options in desired_options.items():
            raw_field = raw_by_name[canonical_name]
            current_specs = self._repair_followup_option_specs(raw_field, keep_ids=True)
            current_by_name = {
                str(item.get("name") or "").strip().casefold(): item
                for item in current_specs
            }
            final_specs = [
                copy.deepcopy(
                    current_by_name.get(
                        str(option.get("name") or "").strip().casefold()
                    )
                    or option
                )
                for option in options
            ]
            final_names = [str(item.get("name") or "") for item in final_specs]
            current_names = [str(item.get("name") or "") for item in current_specs]
            if int(raw_field.get("type") or 0) != 3 or current_names != final_names:
                self._update_repair_followup_field_definition(
                    raw_field,
                    field_type=3,
                    options=final_specs,
                )

        metas, meta_by_name = self._load_table_fields(
            app_token=REPAIR_SOURCE_APP_TOKEN,
            table_id=REPAIR_FOLLOWUP_TABLE_ID,
        )
        records = self._load_table_records(
            app_token=REPAIR_SOURCE_APP_TOKEN,
            table_id=REPAIR_FOLLOWUP_TABLE_ID,
            meta_by_name=meta_by_name,
            work_type=WORK_TYPE_REPAIR,
            notice_type=NOTICE_TYPE_REPAIR,
        )
        record_by_id = {
            str(record.get("record_id") or "").strip(): record
            for record in records
        }
        for record_id, desired in desired_by_record.items():
            record = record_by_id.get(record_id)
            if record is None:
                raise PortalError(f"维修跟进记录{record_id}在迁移校验时不存在。")
            select_patch: dict[str, Any] = {}
            for name, value in desired.items():
                current = self._repair_management_plain_text(
                    self._repair_followup_record_value(record, name)
                ).strip()
                if current != value:
                    select_patch[name] = value
            if select_patch:
                self._patch_record_fields(
                    app_token=REPAIR_SOURCE_APP_TOKEN,
                    table_id=REPAIR_FOLLOWUP_TABLE_ID,
                    record_id=record_id,
                    fields=select_patch,
                )

        metas, meta_by_name = self._load_table_fields(
            app_token=REPAIR_SOURCE_APP_TOKEN,
            table_id=REPAIR_FOLLOWUP_TABLE_ID,
        )
        records = self._load_table_records(
            app_token=REPAIR_SOURCE_APP_TOKEN,
            table_id=REPAIR_FOLLOWUP_TABLE_ID,
            meta_by_name=meta_by_name,
            work_type=WORK_TYPE_REPAIR,
            notice_type=NOTICE_TYPE_REPAIR,
        )
        record_by_id = {
            str(record.get("record_id") or "").strip(): record
            for record in records
        }
        verification_errors: list[str] = []
        for name, expected_type in expected_types.items():
            meta = meta_by_name.get(name)
            if meta is None or int(meta.field_type or 0) != expected_type:
                verification_errors.append(f"{name}类型校验失败")
        for canonical_name, options in desired_options.items():
            actual = {
                str(item or "").strip().casefold()
                for item in (meta_by_name[canonical_name].option_names or [])
            }
            missing_options = [
                str(item.get("name") or "").strip()
                for item in options
                if str(item.get("name") or "").strip().casefold() not in actual
            ]
            if missing_options:
                verification_errors.append(
                    f"{canonical_name}缺少{len(missing_options)}个选项"
                )
        for record_id, desired in desired_by_record.items():
            record = record_by_id.get(record_id)
            if record is None:
                verification_errors.append(f"{record_id}记录不存在")
                continue
            for name, value in desired.items():
                current = self._repair_management_plain_text(
                    self._repair_followup_record_value(record, name)
                ).strip()
                if current != value:
                    verification_errors.append(
                        f"{record_id}.{name}={current!r}，期望{value!r}"
                    )
        if verification_errors:
            result = {
                "status": "failed",
                "migrated_records": len(desired_by_record),
                "errors": verification_errors[:50],
                "finished_at": dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            }
            self._state_store.put_backend_runtime(
                REPAIR_FOLLOWUP_SCHEMA_MIGRATION_RUNTIME_KEY,
                result,
            )
            raise PortalError(
                "维修跟进字段迁移校验失败：" + "；".join(verification_errors[:5])
            )

        deletion_errors: list[str] = []
        deleted_fields: list[str] = []
        for legacy_name, _canonical_name in REPAIR_FOLLOWUP_LEGACY_FIELD_MAPPINGS:
            raw_field = raw_by_name.get(legacy_name)
            if not isinstance(raw_field, dict):
                continue
            try:
                self._delete_repair_followup_field_definition(raw_field)
                deleted_fields.append(legacy_name)
            except Exception as exc:
                deletion_errors.append(str(exc))

        metas, meta_by_name = self._load_table_fields(
            app_token=REPAIR_SOURCE_APP_TOKEN,
            table_id=REPAIR_FOLLOWUP_TABLE_ID,
        )
        remaining = [
            name
            for name, _canonical_name in REPAIR_FOLLOWUP_LEGACY_FIELD_MAPPINGS
            if name in meta_by_name
        ]
        result = {
            "status": "complete" if not remaining else "partial",
            "migrated_records": len(desired_by_record),
            "deleted_fields": deleted_fields,
            "remaining_fields": remaining,
            "errors": deletion_errors,
            "finished_at": dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        }
        self._state_store.put_backend_runtime(
            REPAIR_FOLLOWUP_SCHEMA_MIGRATION_RUNTIME_KEY,
            result,
        )
        self._invalidate_repair_followup_catalog_cache()
        return metas, meta_by_name

    def migrate_repair_followup_duplicate_fields(self) -> dict[str, Any]:
        with self._repair_followup_schema_lock:
            metas, meta_by_name = self._load_table_fields(
                app_token=REPAIR_SOURCE_APP_TOKEN,
                table_id=REPAIR_FOLLOWUP_TABLE_ID,
            )
            self._migrate_repair_followup_duplicate_fields_unlocked(
                metas,
                meta_by_name,
            )
            self._repair_followup_schema_ready = True
            return self._state_store.get_backend_runtime(
                REPAIR_FOLLOWUP_SCHEMA_MIGRATION_RUNTIME_KEY
            ) or {"status": "complete"}

    def _ensure_repair_followup_parent_id_field(
        self,
    ) -> tuple[list[FieldMeta], dict[str, FieldMeta]]:
        with self._repair_followup_schema_lock:
            if self._repair_followup_schema_ready and self._repair_snapshots_enabled:
                snapshot = self._state_store.get_repair_snapshot(
                    REPAIR_SNAPSHOT_SOURCE_FOLLOWUPS
                )
                cached_metas, cached_by_name, _cached_records = (
                    self._repair_snapshot_from_local(snapshot)
                )
                cached_parent = cached_by_name.get(
                    REPAIR_FOLLOWUP_PARENT_ID_FIELD_NAME
                )
                if cached_parent is not None and int(cached_parent.field_type or 0) == 1:
                    return cached_metas, cached_by_name
            metas, meta_by_name = self._load_table_fields(
                app_token=REPAIR_SOURCE_APP_TOKEN,
                table_id=REPAIR_FOLLOWUP_TABLE_ID,
            )
            existing = meta_by_name.get(REPAIR_FOLLOWUP_PARENT_ID_FIELD_NAME)
            if existing is None:
                url = (
                    "https://open.feishu.cn/open-apis/bitable/v1/apps/"
                    f"{REPAIR_SOURCE_APP_TOKEN}/tables/{REPAIR_FOLLOWUP_TABLE_ID}/fields"
                )

                def do_create() -> dict[str, Any]:
                    return self._request_payload(
                        "POST",
                        url,
                        context="维修跟进汇总ID字段创建",
                        headers={
                            **self._auth_headers(),
                            "Content-Type": "application/json",
                        },
                        json_payload={
                            "field_name": REPAIR_FOLLOWUP_PARENT_ID_FIELD_NAME,
                            "type": 1,
                        },
                    )

                payload = do_create()
                if int(payload.get("code") or 0) in TOKEN_ERROR_CODES:
                    refresh_feishu_token()
                    payload = do_create()
                code = int(payload.get("code") or 0)
                if code not in {0, 1254014}:
                    raise PortalError(
                        "创建维修跟进汇总记录ID字段失败: "
                        f"code={code}, msg={payload.get('msg') or 'unknown'}"
                    )
                metas, meta_by_name = self._load_table_fields(
                    app_token=REPAIR_SOURCE_APP_TOKEN,
                    table_id=REPAIR_FOLLOWUP_TABLE_ID,
                )
                existing = meta_by_name.get(REPAIR_FOLLOWUP_PARENT_ID_FIELD_NAME)
            if existing is None:
                raise PortalError("维修跟进汇总记录ID字段创建后仍无法读取。")
            if int(existing.field_type or 0) != 1:
                raise PortalError(
                    f"{REPAIR_FOLLOWUP_PARENT_ID_FIELD_NAME} 必须是文本字段。"
                )
            if not self._repair_management_progress_schema_ready:
                _summary_metas, summary_meta_by_name = self._load_table_fields(
                    app_token=REPAIR_SOURCE_APP_TOKEN,
                    table_id=REPAIR_MANAGEMENT_TABLE_ID,
                )
                self._ensure_repair_management_progress_field_definition(
                    summary_meta_by_name,
                    meta_by_name,
                )
            if self._repair_followup_schema_ready:
                return metas, meta_by_name
            migrated = self._migrate_repair_followup_duplicate_fields_unlocked(
                metas,
                meta_by_name,
            )
            self._repair_followup_schema_ready = True
            return migrated

    def _ensure_repair_followup_select_options(
        self,
        fields: dict[str, Any],
        meta_by_name: dict[str, FieldMeta],
    ) -> tuple[list[FieldMeta], dict[str, FieldMeta]]:
        desired_by_name: dict[str, list[str]] = {}
        for field_name, value in (fields or {}).items():
            meta = meta_by_name.get(str(field_name or "").strip())
            if meta is None or int(meta.field_type or 0) not in {3, 4}:
                continue
            values = (
                self._repair_management_select_items(value)
                if int(meta.field_type or 0) == 4
                else [value]
            )
            existing = {str(item or "").strip().casefold() for item in meta.option_names}
            desired: list[str] = []
            for item in values:
                text = self._repair_management_canonical_select_text(
                    meta.field_name,
                    item,
                ).strip()
                if (
                    not text
                    or re.fullmatch(r"opt[A-Za-z0-9]+", text)
                    or text.casefold() in existing
                ):
                    continue
                if text not in desired:
                    desired.append(text)
            if desired:
                desired_by_name[meta.field_name] = desired
        if not desired_by_name:
            return list(meta_by_name.values()), meta_by_name

        with self._repair_followup_schema_lock:
            payload = self._request_json(
                "fields",
                params={"page_size": 500},
                app_token=REPAIR_SOURCE_APP_TOKEN,
                table_id=REPAIR_FOLLOWUP_TABLE_ID,
            )
            raw_fields = payload.get("data", {}).get("items") or []
            raw_by_name = {
                str(item.get("field_name") or "").strip(): item
                for item in raw_fields
                if isinstance(item, dict)
            }
            for field_name, desired in desired_by_name.items():
                raw_field = raw_by_name.get(field_name)
                if not isinstance(raw_field, dict):
                    raise PortalError(f"维修跟进表缺少字段：{field_name}")
                field_id = str(raw_field.get("field_id") or "").strip()
                if not field_id:
                    raise PortalError(f"维修跟进表字段缺少 field_id：{field_name}")
                property_payload = copy.deepcopy(raw_field.get("property") or {})
                options = property_payload.get("options")
                if not isinstance(options, list):
                    options = []
                option_names = {
                    str(option.get("name") or "").strip().casefold()
                    for option in options
                    if isinstance(option, dict)
                }
                changed = False
                for option_name in desired:
                    if option_name.casefold() in option_names:
                        continue
                    options.append({"name": option_name})
                    option_names.add(option_name.casefold())
                    changed = True
                if not changed:
                    continue
                property_payload["options"] = options
                url = (
                    "https://open.feishu.cn/open-apis/bitable/v1/apps/"
                    f"{REPAIR_SOURCE_APP_TOKEN}/tables/{REPAIR_FOLLOWUP_TABLE_ID}/"
                    f"fields/{field_id}"
                )
                request_body: dict[str, Any] = {
                    "field_name": field_name,
                    "type": int(raw_field.get("type") or 0),
                    "property": property_payload,
                }
                ui_type = str(raw_field.get("ui_type") or "").strip()
                if ui_type:
                    request_body["ui_type"] = ui_type
                description = raw_field.get("description")
                if isinstance(description, str):
                    request_body["description"] = description

                def do_update() -> dict[str, Any]:
                    return self._request_payload(
                        "PUT",
                        url,
                        context=f"维修跟进字段选项更新({field_name})",
                        headers={
                            **self._auth_headers(),
                            "Content-Type": "application/json",
                        },
                        json_payload=request_body,
                    )

                update_payload = do_update()
                if int(update_payload.get("code") or 0) in TOKEN_ERROR_CODES:
                    refresh_feishu_token()
                    update_payload = do_update()
                code = int(update_payload.get("code") or 0)
                if code != 0:
                    raise PortalError(
                        f"补齐{field_name}选项失败: code={code}, "
                        f"msg={update_payload.get('msg') or 'unknown'}"
                    )
            return self._load_table_fields(
                app_token=REPAIR_SOURCE_APP_TOKEN,
                table_id=REPAIR_FOLLOWUP_TABLE_ID,
            )

    def _create_record_fields(
        self,
        *,
        app_token: str,
        table_id: str,
        fields: dict[str, Any],
    ) -> dict[str, Any]:
        app_token = str(app_token or "").strip()
        table_id = str(table_id or "").strip()
        if not app_token or not table_id:
            raise PortalError("创建多维记录缺少 app_token/table_id。")
        if not isinstance(fields, dict) or not fields:
            raise PortalError("创建多维记录字段不能为空。")
        url = (
            f"https://open.feishu.cn/open-apis/bitable/v1/apps/"
            f"{app_token}/tables/{table_id}/records"
        )

        def do_create() -> dict[str, Any]:
            return self._request_payload(
                "POST",
                url,
                context="飞书记录创建",
                headers={**self._auth_headers(), "Content-Type": "application/json"},
                json_payload={"fields": fields},
            )

        payload = do_create()
        if int(payload.get("code") or 0) in TOKEN_ERROR_CODES:
            refresh_feishu_token()
            payload = do_create()
        code = payload.get("code", 0)
        if code != 0:
            raise PortalError(
                f"飞书记录创建失败: code={code}, msg={payload.get('msg') or 'unknown'}"
            )
        if table_id == REPAIR_MANAGEMENT_REPAIR_TABLE_ID:
            self._invalidate_repair_management_target_cache()
        if table_id == REPAIR_FOLLOWUP_TABLE_ID:
            self._invalidate_repair_followup_catalog_cache()
        return payload

    def _delete_record_fields(
        self,
        *,
        app_token: str,
        table_id: str,
        record_id: str,
    ) -> dict[str, Any]:
        app_token = str(app_token or "").strip()
        table_id = str(table_id or "").strip()
        record_id = str(record_id or "").strip()
        if not app_token or not table_id or not record_id:
            raise PortalError("删除多维记录缺少 app_token/table_id/record_id。")
        url = (
            f"https://open.feishu.cn/open-apis/bitable/v1/apps/"
            f"{app_token}/tables/{table_id}/records/{record_id}"
        )

        def do_delete() -> dict[str, Any]:
            return self._request_payload(
                "DELETE",
                url,
                context="飞书记录删除",
                headers=self._auth_headers(),
            )

        payload = do_delete()
        if int(payload.get("code") or 0) in TOKEN_ERROR_CODES:
            refresh_feishu_token()
            payload = do_delete()
        code = int(payload.get("code") or 0)
        if code != 0:
            raise PortalError(
                f"飞书记录删除失败: code={code}, msg={payload.get('msg') or 'unknown'}"
            )
        if table_id == REPAIR_MANAGEMENT_REPAIR_TABLE_ID:
            self._invalidate_repair_management_target_cache()
        if table_id == REPAIR_FOLLOWUP_TABLE_ID:
            self._invalidate_repair_followup_catalog_cache()
        return payload

    @staticmethod
    def _extract_option_map(field: dict[str, Any]) -> tuple[dict[str, str], list[str]]:
        prop = field.get("property") or {}
        options = prop.get("options") or []
        if not options:
            nested = (((prop.get("type") or {}).get("ui_property") or {}).get("options")) or []
            options = nested
        option_map: dict[str, str] = {}
        option_names: list[str] = []
        for option in options:
            option_id = str(option.get("id") or "").strip()
            option_name = str(option.get("name") or "").strip()
            if not option_name:
                continue
            if option_id:
                option_map[option_id] = option_name
            option_names.append(option_name)
        deduped = list(dict.fromkeys(option_names))
        return option_map, deduped

    def _load_fields(self) -> list[FieldMeta]:
        payload = self._request_json("fields", params={"page_size": 500})
        metas = self._parse_field_meta(payload)
        self._field_meta_list = metas
        self._field_meta_by_name = {meta.field_name: meta for meta in metas}
        return metas

    def _load_change_fields(self) -> list[FieldMeta]:
        payload = self._request_json(
            "fields",
            params={"page_size": 500},
            app_token=CHANGE_SOURCE_APP_TOKEN,
            table_id=CHANGE_SOURCE_TABLE_ID,
        )
        metas = self._parse_field_meta(payload)
        self._change_field_meta_list = metas
        self._change_field_meta_by_name = {meta.field_name: meta for meta in metas}
        return metas

    def _load_zhihang_change_fields(self) -> list[FieldMeta]:
        payload = self._request_json(
            "fields",
            params={"page_size": 500},
            app_token=ZHIHANG_CHANGE_APP_TOKEN,
            table_id=ZHIHANG_CHANGE_TABLE_ID,
        )
        metas = self._parse_field_meta(payload)
        self._zhihang_change_field_meta_list = metas
        self._zhihang_change_field_meta_by_name = {
            meta.field_name: meta for meta in metas
        }
        return metas

    def _load_repair_fields(self) -> list[FieldMeta]:
        payload = self._request_json(
            "fields",
            params={"page_size": 500},
            app_token=REPAIR_SOURCE_APP_TOKEN,
            table_id=REPAIR_SOURCE_TABLE_ID,
        )
        metas = self._parse_field_meta(payload)
        self._repair_field_meta_list = metas
        self._repair_field_meta_by_name = {meta.field_name: meta for meta in metas}
        return metas

    def _parse_field_meta(self, payload: dict[str, Any]) -> list[FieldMeta]:
        items = payload.get("data", {}).get("items") or []
        return self._parse_field_metas(items)

    def _parse_field_metas(self, items: list[dict[str, Any]]) -> list[FieldMeta]:
        metas: list[FieldMeta] = []
        for item in items:
            if not isinstance(item, dict):
                continue
            option_map, option_names = self._extract_option_map(item)
            prop = item.get("property") or {}
            metas.append(
                FieldMeta(
                    field_id=str(item.get("field_id") or ""),
                    field_name=str(item.get("field_name") or ""),
                    ui_type=str(item.get("ui_type") or ""),
                    field_type=int(item.get("type") or 0),
                    is_primary=bool(item.get("is_primary")),
                    options_map=option_map,
                    option_names=option_names,
                    has_formula=bool(
                        prop.get("formula_expression")
                        or prop.get("formula")
                    ),
                    property=copy.deepcopy(prop),
                )
            )
        return metas

    def _load_records(self) -> list[dict[str, Any]]:
        records: list[dict[str, Any]] = []
        page_token = ""
        while True:
            params = {"page_size": 500}
            if page_token:
                params["page_token"] = page_token
            payload = self._request_json("records", params=params)
            data = payload.get("data", {})
            for item in data.get("items") or []:
                normalized = self._normalize_record(
                    item,
                    meta_by_name=self._field_meta_by_name,
                    work_type=WORK_TYPE_MAINTENANCE,
                    notice_type=NOTICE_TYPE_MAINTENANCE,
                    source_app_token=self.app_token or DEFAULT_APP_TOKEN,
                    source_table_id=self.table_id or DEFAULT_TABLE_ID,
                )
                if self._source_record_matches_month_window(normalized):
                    records.append(normalized)
            if not data.get("has_more"):
                break
            page_token = str(data.get("page_token") or "").strip()
            if not page_token:
                break
        self._records = records
        self._maintenance_loaded_once = True
        return records

    def _load_change_records(self) -> list[dict[str, Any]]:
        records: list[dict[str, Any]] = []
        page_token = ""
        while True:
            params = {"page_size": 500}
            if page_token:
                params["page_token"] = page_token
            payload = self._request_json(
                "records",
                params=params,
                app_token=CHANGE_SOURCE_APP_TOKEN,
                table_id=CHANGE_SOURCE_TABLE_ID,
            )
            data = payload.get("data", {})
            for item in data.get("items") or []:
                normalized = self._normalize_record(
                    item,
                    meta_by_name=self._change_field_meta_by_name,
                    work_type=WORK_TYPE_CHANGE,
                    notice_type=NOTICE_TYPE_CHANGE,
                    source_app_token=CHANGE_SOURCE_APP_TOKEN,
                    source_table_id=CHANGE_SOURCE_TABLE_ID,
                )
                if self._source_record_matches_month_window(normalized):
                    records.append(normalized)
            if not data.get("has_more"):
                break
            page_token = str(data.get("page_token") or "").strip()
            if not page_token:
                break
        self._change_records = records
        self._change_loaded_once = True
        return records

    def _load_zhihang_change_records(self) -> list[dict[str, Any]]:
        records: list[dict[str, Any]] = []
        page_token = ""
        while True:
            params = {"page_size": 500}
            if page_token:
                params["page_token"] = page_token
            payload = self._request_json(
                "records",
                params=params,
                app_token=ZHIHANG_CHANGE_APP_TOKEN,
                table_id=ZHIHANG_CHANGE_TABLE_ID,
            )
            data = payload.get("data", {})
            for item in data.get("items") or []:
                normalized = self._normalize_record(
                    item,
                    meta_by_name=self._zhihang_change_field_meta_by_name,
                    work_type=WORK_TYPE_CHANGE,
                    notice_type=NOTICE_TYPE_CHANGE,
                    source_app_token=ZHIHANG_CHANGE_APP_TOKEN,
                    source_table_id=ZHIHANG_CHANGE_TABLE_ID,
                )
                if self._source_record_matches_month_window(normalized):
                    records.append(normalized)
            if not data.get("has_more"):
                break
            page_token = str(data.get("page_token") or "").strip()
            if not page_token:
                break
        self._zhihang_change_records = records
        self._zhihang_change_loaded_once = True
        return records

    def _load_repair_records(self) -> list[dict[str, Any]]:
        records: list[dict[str, Any]] = []
        page_token = ""
        while True:
            params = {"page_size": REPAIR_SOURCE_PAGE_SIZE}
            if REPAIR_SOURCE_VIEW_ID:
                params["view_id"] = REPAIR_SOURCE_VIEW_ID
            if page_token:
                params["page_token"] = page_token
            payload = self._request_json(
                "records",
                params=params,
                app_token=REPAIR_SOURCE_APP_TOKEN,
                table_id=REPAIR_SOURCE_TABLE_ID,
            )
            data = payload.get("data", {})
            for item in data.get("items") or []:
                normalized = self._normalize_record(
                    item,
                    meta_by_name=self._repair_field_meta_by_name,
                    work_type=WORK_TYPE_REPAIR,
                    notice_type=NOTICE_TYPE_REPAIR,
                    source_app_token=REPAIR_SOURCE_APP_TOKEN,
                    source_table_id=REPAIR_SOURCE_TABLE_ID,
                )
                if self._source_record_matches_month_window(normalized):
                    records.append(normalized)
            if not data.get("has_more"):
                break
            page_token = str(data.get("page_token") or "").strip()
            if not page_token:
                break
        self._repair_records = records
        self._repair_loaded_once = True
        return records

    def _load_table_fields(
        self, *, app_token: str, table_id: str
    ) -> tuple[list[FieldMeta], dict[str, FieldMeta]]:
        payload = self._request_json(
            "fields",
            params={"page_size": 500},
            app_token=app_token,
            table_id=table_id,
        )
        metas = self._parse_field_metas(payload.get("data", {}).get("items") or [])
        return metas, {meta.field_name: meta for meta in metas}

    def _load_table_records(
        self,
        *,
        app_token: str,
        table_id: str,
        meta_by_name: dict[str, FieldMeta],
        work_type: str,
        notice_type: str,
        view_id: str = "",
    ) -> list[dict[str, Any]]:
        records: list[dict[str, Any]] = []
        page_token = ""
        while True:
            params = {"page_size": 500}
            if view_id:
                params["view_id"] = view_id
            if page_token:
                params["page_token"] = page_token
            payload = self._request_json(
                "records",
                params=params,
                app_token=app_token,
                table_id=table_id,
            )
            data = payload.get("data", {})
            for item in data.get("items") or []:
                records.append(
                    self._normalize_record(
                        item,
                        meta_by_name=meta_by_name,
                        work_type=work_type,
                        notice_type=notice_type,
                        source_app_token=app_token,
                        source_table_id=table_id,
                    )
                )
            if not data.get("has_more"):
                break
            page_token = str(data.get("page_token") or "").strip()
            if not page_token:
                break
        return records

    def _search_table_records(
        self,
        *,
        app_token: str,
        table_id: str,
        meta_by_name: dict[str, FieldMeta],
        work_type: str,
        notice_type: str,
        field_names: list[str] | tuple[str, ...],
        sort_field: str = "",
        limit: int = 200,
        filter_payload: dict[str, Any] | None = None,
    ) -> list[dict[str, Any]]:
        max_records = max(1, int(limit or 200))
        url = (
            f"https://open.feishu.cn/open-apis/bitable/v1/apps/"
            f"{app_token}/tables/{table_id}/records/search"
        )
        body: dict[str, Any] = {
            "automatic_fields": False,
            "field_names": [
                name for name in field_names if str(name or "").strip() in meta_by_name
            ],
        }
        if sort_field and sort_field in meta_by_name:
            body["sort"] = [{"field_name": sort_field, "desc": True}]
        if isinstance(filter_payload, dict) and filter_payload:
            body["filter"] = filter_payload

        def do_search(*, page_size: int, page_token: str = "") -> dict[str, Any]:
            params: dict[str, Any] = {
                "page_size": page_size,
                "user_id_type": "open_id",
            }
            if page_token:
                params["page_token"] = page_token
            return self._request_payload(
                "POST",
                url,
                context="飞书记录搜索",
                headers={**self._auth_headers(), "Content-Type": "application/json"},
                params=params,
                json_payload=body,
            )

        records: list[dict[str, Any]] = []
        page_token = ""
        while len(records) < max_records:
            page_size = min(500, max_records - len(records))
            payload: dict[str, Any] = {}
            for attempt, retry_delay in enumerate(
                (*BITABLE_TRANSIENT_RETRY_DELAYS, 0.0)
            ):
                payload = do_search(page_size=page_size, page_token=page_token)
                if int(payload.get("code") or 0) in TOKEN_ERROR_CODES:
                    refresh_feishu_token()
                    payload = do_search(page_size=page_size, page_token=page_token)
                if int(payload.get("code") or 0) != BITABLE_DATA_NOT_READY_CODE:
                    break
                if attempt < len(BITABLE_TRANSIENT_RETRY_DELAYS):
                    time.sleep(retry_delay)
            code = int(payload.get("code") or 0)
            if code != 0:
                raise PortalError(
                    f"飞书记录搜索失败: code={code}, "
                    f"msg={payload.get('msg') or 'unknown'}"
                )
            data = payload.get("data") or {}
            for item in data.get("items") or []:
                if not isinstance(item, dict):
                    continue
                records.append(
                    self._normalize_record(
                        item,
                        meta_by_name=meta_by_name,
                        work_type=work_type,
                        notice_type=notice_type,
                        source_app_token=app_token,
                        source_table_id=table_id,
                    )
                )
                if len(records) >= max_records:
                    break
            if not data.get("has_more"):
                break
            page_token = str(data.get("page_token") or "").strip()
            if not page_token:
                break
        return records

    def _load_table_records_by_ids(
        self,
        *,
        app_token: str,
        table_id: str,
        meta_by_name: dict[str, FieldMeta],
        work_type: str,
        notice_type: str,
        record_ids: list[str] | tuple[str, ...],
    ) -> list[dict[str, Any]]:
        records: list[dict[str, Any]] = []
        for record_id in dict.fromkeys(
            str(item or "").strip() for item in record_ids if str(item or "").strip()
        ):
            payload = self._request_json(
                f"records/{record_id}",
                params={"user_id_type": "open_id"},
                app_token=app_token,
                table_id=table_id,
            )
            raw_record = (payload.get("data") or {}).get("record")
            if not isinstance(raw_record, dict):
                raise PortalError(f"未找到多维记录：{record_id}")
            records.append(
                self._normalize_record(
                    raw_record,
                    meta_by_name=meta_by_name,
                    work_type=work_type,
                    notice_type=notice_type,
                    source_app_token=app_token,
                    source_table_id=table_id,
                )
            )
        return records

    @staticmethod
    def _field_meta_is_readonly(meta: FieldMeta) -> bool:
        ui_type = str(meta.ui_type or "").strip().lower()
        field_name = str(meta.field_name or "").strip()
        field_type = int(meta.field_type or 0)
        if meta.has_formula or field_type in {19, 20, 1001, 1002, 1003, 1004, 1005}:
            return True
        readonly_name_tokens = (
            "创建日期",
            "当前日期",
            "当前维修进度",
            "对应来源",
            "辅助列",
            "公式",
            "记录创建人",
            "最后修改",
            "修改人",
            "是否有唯一id",
        )
        if any(token in field_name for token in readonly_name_tokens):
            return True
        readonly_tokens = (
            "formula",
            "lookup",
            "created",
            "modified",
            "auto",
            "duplex",
            "stage",
        )
        return any(token in ui_type for token in readonly_tokens)

    @classmethod
    def _field_meta_payload(cls, meta: FieldMeta) -> dict[str, Any]:
        return {
            "field_id": meta.field_id,
            "field_name": meta.field_name,
            "ui_type": meta.ui_type,
            "field_type": meta.field_type,
            "is_primary": meta.is_primary,
            "options": meta.option_names,
            "editable": not cls._field_meta_is_readonly(meta),
        }

    @classmethod
    def _repair_management_field_is_readonly(cls, meta: FieldMeta) -> bool:
        field_name = str(meta.field_name or "").strip()
        if field_name not in REPAIR_MANAGEMENT_FORM_FIELD_NAMES:
            return True
        return cls._field_meta_is_readonly(meta)

    @classmethod
    def _repair_management_field_payload(cls, meta: FieldMeta) -> dict[str, Any]:
        payload = cls._field_meta_payload(meta)
        payload["editable"] = not cls._repair_management_field_is_readonly(meta)
        payload["auto_filled"] = not payload["editable"]
        return payload

    @classmethod
    def _repair_management_title(cls, record: dict[str, Any]) -> str:
        fields = record.get("display_fields") if isinstance(record.get("display_fields"), dict) else {}
        for key in ("维修名称", "检修通告名称"):
            value = cls._repair_management_plain_text(fields.get(key))
            if value:
                return value
        for value in fields.values():
            text = cls._repair_management_plain_text(value)
            if text:
                return text[:80]
        return str(record.get("record_id") or "未命名检修记录")

    @classmethod
    def _repair_management_workflow_text(cls, record: dict[str, Any]) -> str:
        for field_container_name in ("display_fields", "raw_fields"):
            fields = record.get(field_container_name)
            if not isinstance(fields, dict):
                continue
            value = cls._repair_management_plain_text(fields.get("流程"))
            if value:
                return value
        return ""

    @classmethod
    def _repair_management_is_completed(cls, record: dict[str, Any]) -> bool:
        workflow = re.sub(r"\s+", "", cls._repair_management_workflow_text(record))
        return workflow == REPAIR_MANAGEMENT_COMPLETED_WORKFLOW

    @classmethod
    def _clean_repair_management_fields(
        cls,
        fields: dict[str, Any],
        meta_by_name: dict[str, FieldMeta],
        *,
        allow_empty: bool = False,
    ) -> dict[str, Any]:
        if not isinstance(fields, dict):
            raise PortalError("检修记录字段必须是对象。")
        cleaned: dict[str, Any] = {}
        for name, value in fields.items():
            field_name = str(name or "").strip()
            if not field_name:
                continue
            meta = meta_by_name.get(field_name)
            if meta is None:
                continue
            if cls._repair_management_field_is_readonly(meta):
                continue
            cleaned[field_name] = value
        if not cleaned and not allow_empty:
            raise PortalError("没有可写入的检修字段。")
        return cleaned

    @classmethod
    def _missing_repair_management_required_fields(
        cls,
        fields: dict[str, Any],
        meta_by_name: dict[str, FieldMeta],
    ) -> list[str]:
        missing: list[str] = []
        for group in REPAIR_MANAGEMENT_REQUIRED_FIELD_GROUPS:
            writable_names = [
                name
                for name in group
                if name in meta_by_name
                and not cls._repair_management_field_is_readonly(meta_by_name[name])
            ]
            if not writable_names:
                continue
            if not any(str(fields.get(name) or "").strip() for name in writable_names):
                missing.append(writable_names[0])
        return missing

    @classmethod
    def _repair_management_schema_warnings(
        cls,
        meta_by_name: dict[str, FieldMeta],
    ) -> list[str]:
        warnings: list[str] = []
        for group in REPAIR_MANAGEMENT_REQUIRED_FIELD_GROUPS:
            available = [name for name in group if name in meta_by_name]
            if not available:
                warnings.append(f"维修项目表缺少字段：{' / '.join(group)}")
            elif all(
                cls._repair_management_field_is_readonly(meta_by_name[name])
                for name in available
            ):
                warnings.append(f"维修项目必填字段不可写：{' / '.join(available)}")
        for name in ("关联事件单", "设备检修关联", "维修跟进记录"):
            if name not in meta_by_name:
                warnings.append(f"维修项目表缺少来源追踪字段：{name}")
        return warnings

    @classmethod
    def _apply_repair_management_event_link(
        cls,
        fields: dict[str, Any],
        meta_by_name: dict[str, FieldMeta],
        source_event_id: str = "",
    ) -> dict[str, Any]:
        source_event_id = str(source_event_id or "").strip()
        if not source_event_id:
            return fields
        linked = dict(fields)
        field_name = REPAIR_MANAGEMENT_EVENT_LINK_FIELD_NAMES[0]
        if field_name in meta_by_name:
            linked[field_name] = source_event_id
        return linked

    @staticmethod
    def _repair_management_record_ids(value: Any) -> list[str]:
        values: list[Any]
        if isinstance(value, list):
            values = value
        elif isinstance(value, dict):
            nested_ids = value.get("record_ids")
            if not isinstance(nested_ids, list):
                nested_ids = value.get("link_record_ids")
            values = nested_ids if isinstance(nested_ids, list) else [value]
        else:
            text = str(value or "").strip()
            if not text:
                return []
            with suppress(Exception):
                decoded = json.loads(text)
                if isinstance(decoded, list):
                    values = decoded
                else:
                    values = [decoded]
                return list(
                    dict.fromkeys(
                        str(item or "").strip()
                        for item in values
                        if str(item or "").strip().startswith("rec")
                    )
                )
            values = re.split(r"[\s,，、;；]+", text)
        result: list[str] = []
        for item in values:
            if isinstance(item, dict):
                nested = item.get("record_ids")
                if not isinstance(nested, list):
                    nested = item.get("link_record_ids")
                if isinstance(nested, list):
                    result.extend(str(record_id or "").strip() for record_id in nested)
                    continue
                item = (
                    item.get("record_id")
                    or item.get("id")
                    or item.get("text")
                    or ""
                )
            record_id = str(item or "").strip()
            if record_id.startswith("rec"):
                result.append(record_id)
        return list(dict.fromkeys(result))

    @staticmethod
    def _repair_management_users(value: Any) -> list[dict[str, Any]]:
        pending: list[Any] = []
        if isinstance(value, list):
            pending.extend(value)
        elif isinstance(value, dict):
            users = value.get("users")
            if not isinstance(users, list):
                users = value.get("value")
            if isinstance(users, list):
                pending.extend(users)
            else:
                pending.append(value)
        result: list[dict[str, Any]] = []
        seen: set[str] = set()
        for item in pending:
            if not isinstance(item, dict):
                continue
            user_id = str(
                item.get("id")
                or item.get("open_id")
                or item.get("user_id")
                or item.get("userId")
                or ""
            ).strip()
            if not user_id or user_id in seen:
                continue
            seen.add(user_id)
            result.append({"id": user_id})
        return result

    @classmethod
    def _apply_repair_management_repair_links(
        cls,
        fields: dict[str, Any],
        meta_by_name: dict[str, FieldMeta],
        repair_record_ids: list[str] | tuple[str, ...] | None,
    ) -> dict[str, Any]:
        ids = list(
            dict.fromkeys(
                str(record_id or "").strip()
                for record_id in (repair_record_ids or [])
                if str(record_id or "").strip().startswith("rec")
            )
        )
        if not ids:
            return fields
        linked = dict(fields)
        field_name = REPAIR_MANAGEMENT_REPAIR_LINK_FIELD_NAMES[0]
        if field_name in meta_by_name:
            linked[field_name] = ids[0]
        return linked

    @classmethod
    def _repair_management_datetime_ms(cls, value: Any) -> int | None:
        if isinstance(value, (int, float)):
            raw = int(value)
            return raw if raw > 10_000_000_000 else raw * 1000
        numeric_text = str(value or "").strip()
        if re.fullmatch(r"\d{10,13}", numeric_text):
            raw = int(numeric_text)
            return raw if raw > 10_000_000_000 else raw * 1000
        parsed = cls._parse_notice_datetime(value)
        return int(parsed.timestamp() * 1000) if parsed else None

    @classmethod
    def _repair_management_plain_text(cls, value: Any) -> str:
        if value in (None, "", [], {}):
            return ""
        if isinstance(value, (list, tuple, set)):
            return "、".join(
                dict.fromkeys(
                    text
                    for item in value
                    if (text := cls._repair_management_plain_text(item))
                )
            )
        if isinstance(value, dict):
            for key in ("text", "name", "value"):
                nested = value.get(key)
                if nested not in (None, "", [], {}):
                    return cls._repair_management_plain_text(nested)
            for key in ("text_arr", "groups", "users"):
                nested = value.get(key)
                if isinstance(nested, list):
                    return cls._repair_management_plain_text(nested)
            return ""
        text = str(value or "").strip()
        if not text:
            return ""
        if text[:1] in {"[", "{"}:
            for decoder in (json.loads, ast.literal_eval):
                with suppress(Exception):
                    decoded = decoder(text)
                    normalized = cls._repair_management_plain_text(decoded)
                    if normalized:
                        return normalized
            matches = re.findall(
                r"['\"](?:text|name)['\"]\s*:\s*['\"]([^'\"]+)['\"]",
                text,
            )
            if matches:
                return "、".join(dict.fromkeys(matches))
        return text

    @classmethod
    def _repair_management_select_items(cls, value: Any) -> list[Any]:
        if value in (None, "", [], {}):
            return []
        if isinstance(value, (list, tuple, set)):
            flattened: list[Any] = []
            for item in value:
                flattened.extend(cls._repair_management_select_items(item))
            return flattened
        if isinstance(value, dict):
            return [value]
        text = str(value or "").strip()
        if not text:
            return []
        if text[:1] in {"[", "{"}:
            for decoder in (json.loads, ast.literal_eval):
                with suppress(Exception):
                    decoded = decoder(text)
                    if decoded != value:
                        return cls._repair_management_select_items(decoded)
        return [
            item.strip()
            for item in re.split(r"[,，、;；]+", text)
            if item.strip()
        ]

    @classmethod
    def _repair_management_canonical_select_text(
        cls,
        field_name: str,
        value: Any,
    ) -> str:
        text = cls._repair_management_plain_text(value)
        legacy = REPAIR_MANAGEMENT_LEGACY_OPTION_ALIASES.get(
            (str(field_name or "").strip(), text.casefold())
        )
        if legacy:
            return legacy
        if field_name not in REPAIR_MANAGEMENT_CUSTOM_MULTI_SELECT_FIELDS:
            return text
        return REPAIR_MANAGEMENT_SOURCE_OPTION_ALIASES.get(text.casefold(), text)

    @classmethod
    def _repair_management_unique_text(cls, values: list[Any]) -> str:
        result: list[str] = []
        for value in values:
            text = cls._repair_management_plain_text(value)
            if text and text not in result:
                result.append(text)
        return "、".join(result)

    @staticmethod
    def _repair_management_match_text(value: Any) -> str:
        text = str(value or "").lower()
        text = re.sub(
            r"ea118(?:[_-]?c01)?机房|巡检发现|bms报|ba报|盯屏发现|"
            r"(?<![a-z0-9])(?:110站|[abcdeh]楼)(?![a-z0-9])|"
            r"检修通告|检修|告警|事件|状态",
            "",
            text,
        )
        return re.sub(r"[^0-9a-z\u4e00-\u9fff]+", "", text)

    @classmethod
    def _repair_management_candidate_score(
        cls,
        event: dict[str, Any],
        repair: dict[str, Any],
    ) -> tuple[int, list[str]]:
        fields = repair.get("display_fields") if isinstance(repair.get("display_fields"), dict) else {}
        event_text = cls._repair_management_match_text(
            cls._repair_management_unique_text(
                [event.get("title"), event.get("alarm_desc")]
            )
        )
        repair_text = cls._repair_management_match_text(
            "".join(
                str(fields.get(name) or "")
                for name in ("名称（标题）", "检修概述", "维修故障", "故障原因", "故障现象")
            )
        )
        reasons: list[str] = []
        score = 0
        if event_text and repair_text:
            ratio = difflib.SequenceMatcher(None, event_text, repair_text).ratio()
            score += int(ratio * 45)
            if event_text in repair_text or repair_text in event_text:
                score += 20
                reasons.append("标题/故障内容高度一致")
            elif ratio >= 0.55:
                reasons.append("标题/故障内容相似")

        event_codes = set(cls._clean_building_codes(event.get("building_codes")))
        repair_codes = set(
            cls._repair_building_codes_from_value(
                fields.get("楼栋") or fields.get("名称（标题）") or fields.get("检修概述")
            )
        )
        if event_codes and repair_codes:
            if event_codes & repair_codes:
                score += 20
                reasons.append("楼栋一致")
            else:
                score -= 35

        event_specialty = str(event.get("specialty") or "").strip()
        repair_specialty = str(fields.get("专业") or "").strip()
        if event_specialty and repair_specialty:
            if event_specialty == repair_specialty:
                score += 12
                reasons.append("专业一致")
            else:
                score -= 10

        event_time = cls._parse_notice_datetime(event.get("occurrence_time"))
        repair_time = cls._parse_notice_datetime(fields.get("发生故障时间"))
        if event_time and repair_time:
            hours = abs((repair_time - event_time).total_seconds()) / 3600
            if hours <= 24:
                score += 15
                reasons.append("故障时间接近")
            elif hours <= 168:
                score += 7
        return max(0, min(score, 100)), reasons

    def _repair_management_repair_source_config(self) -> tuple[str, str]:
        app_token = str(
            getattr(config, "app_token", "")
            or self.app_token
            or DEFAULT_APP_TOKEN
        ).strip()
        table_id = str(
            getattr(config, "table_id_overhaul", "")
            or REPAIR_MANAGEMENT_REPAIR_TABLE_ID
        ).strip()
        if not app_token or not table_id:
            raise PortalError("未配置设备检修目标表 app_token/table_id。")
        return app_token, table_id

    def _load_repair_management_target_records_remote(
        self,
        *,
        force_refresh: bool = False,
    ) -> tuple[list[FieldMeta], dict[str, FieldMeta], list[dict[str, Any]]]:
        app_token, table_id = self._repair_management_repair_source_config()
        now = time.monotonic()
        with self._repair_management_target_cache_lock:
            cached = self._repair_management_target_cache
            if (
                not force_refresh
                and isinstance(cached, dict)
                and str(cached.get("app_token") or "") == app_token
                and str(cached.get("table_id") or "") == table_id
                and now - float(cached.get("loaded_at") or 0) <= 45.0
            ):
                return (
                    list(cached.get("metas") or []),
                    dict(cached.get("meta_by_name") or {}),
                    list(cached.get("records") or []),
                )
            metas, meta_by_name = self._load_table_fields(
                app_token=app_token,
                table_id=table_id,
            )
            records = self._search_table_records(
                app_token=app_token,
                table_id=table_id,
                meta_by_name=meta_by_name,
                work_type=WORK_TYPE_REPAIR,
                notice_type=NOTICE_TYPE_REPAIR,
                field_names=(
                    "检修概述",
                    "专业",
                    "楼栋",
                    "检修状态",
                    "位置",
                    "名称（标题）",
                    "紧急程度",
                    "维修设备",
                    "维修故障",
                    "故障现象",
                    "故障原因",
                    "维修方式",
                    "发生故障时间",
                    "实际开始时间",
                    "实际结束时间",
                    "创建时间",
                ),
                sort_field="创建时间",
                limit=250,
            )
            self._repair_management_target_cache = {
                "loaded_at": time.monotonic(),
                "app_token": app_token,
                "table_id": table_id,
                "metas": list(metas),
                "meta_by_name": dict(meta_by_name),
                "records": list(records),
            }
            return metas, meta_by_name, records

    def _load_repair_management_target_records(
        self,
        *,
        force_refresh: bool = False,
    ) -> tuple[list[FieldMeta], dict[str, FieldMeta], list[dict[str, Any]]]:
        app_token, table_id = self._repair_management_repair_source_config()

        def load_remote() -> tuple[list[FieldMeta], list[dict[str, Any]]]:
            metas, _meta_by_name, records = (
                self._load_repair_management_target_records_remote(
                    force_refresh=force_refresh
                )
            )
            return metas, records

        return self._load_repair_snapshot_source(
            source_key=REPAIR_SNAPSHOT_SOURCE_NOTICES,
            app_token=app_token,
            table_id=table_id,
            loader=load_remote,
            force_refresh=force_refresh,
        )

    def _load_repair_management_event_records_remote(
        self,
        *,
        force_refresh: bool = False,
    ) -> tuple[list[FieldMeta], dict[str, FieldMeta], list[dict[str, Any]]]:
        app_token, table_id, _source_key = self._event_source_config()
        now = time.monotonic()
        with self._repair_management_event_cache_lock:
            cached = self._repair_management_event_cache
            if (
                not force_refresh
                and isinstance(cached, dict)
                and str(cached.get("app_token") or "") == app_token
                and str(cached.get("table_id") or "") == table_id
                and now - float(cached.get("loaded_at") or 0) <= 60.0
            ):
                return (
                    list(cached.get("metas") or []),
                    dict(cached.get("meta_by_name") or {}),
                    list(cached.get("records") or []),
                )
            metas, meta_by_name = self._load_table_fields(
                app_token=app_token,
                table_id=table_id,
            )
            records = self._search_table_records(
                app_token=app_token,
                table_id=table_id,
                meta_by_name=meta_by_name,
                work_type=WORK_TYPE_EVENT,
                notice_type=NOTICE_TYPE_EVENT,
                field_names=(
                    "事件简述",
                    "告警描述",
                    "故障现象",
                    "事件等级",
                    "事件发现来源",
                    "事件发现来源（统一）",
                    "事件发生时间",
                    "事件进展响应时间",
                    "事件状态",
                    "最终状态",
                    "事件目前进展",
                    "事件发生原因",
                    "事件应急措施",
                    "事件解决措施",
                    "备注",
                    "最后更新时间",
                    "进展更新时间",
                    "是否转检修",
                    "机楼",
                    "南通楼栋",
                    "专业",
                    "值班账号",
                    "工程师（消息推送）",
                    "创建时间",
                ),
                sort_field="事件发生时间",
                limit=500,
            )
            self._repair_management_event_cache = {
                "loaded_at": time.monotonic(),
                "app_token": app_token,
                "table_id": table_id,
                "metas": list(metas),
                "meta_by_name": dict(meta_by_name),
                "records": list(records),
            }
            return metas, meta_by_name, records

    def _load_repair_management_event_records(
        self,
        *,
        force_refresh: bool = False,
    ) -> tuple[list[FieldMeta], dict[str, FieldMeta], list[dict[str, Any]]]:
        app_token, table_id, _source_key = self._event_source_config()

        def load_remote() -> tuple[list[FieldMeta], list[dict[str, Any]]]:
            metas, _meta_by_name, records = (
                self._load_repair_management_event_records_remote(
                    force_refresh=force_refresh
                )
            )
            return metas, records

        return self._load_repair_snapshot_source(
            source_key=REPAIR_SNAPSHOT_SOURCE_EVENTS,
            app_token=app_token,
            table_id=table_id,
            loader=load_remote,
            force_refresh=force_refresh,
        )

    @classmethod
    def _repair_management_event_item(
        cls,
        record: dict[str, Any],
    ) -> dict[str, Any]:
        fields = (
            record.get("display_fields")
            if isinstance(record.get("display_fields"), dict)
            else {}
        )
        raw_fields = (
            record.get("raw_fields")
            if isinstance(record.get("raw_fields"), dict)
            else {}
        )
        title = cls._repair_management_plain_text(
            fields.get("事件简述") or fields.get("告警描述")
        )
        alarm_desc = cls._repair_management_plain_text(
            fields.get("告警描述") or title
        )
        building_codes = cls._repair_building_codes_from_value(
            cls._repair_management_plain_text(
                fields.get("机楼") or fields.get("南通楼栋") or title
            )
        )
        source_record_id = str(record.get("record_id") or "").strip()
        return {
            "source_record_id": source_record_id,
            "record_id": source_record_id,
            "title": title,
            "alarm_desc": alarm_desc,
            "building": cls._building_label_from_codes(building_codes),
            "building_codes": building_codes,
            "specialty": cls._repair_management_plain_text(fields.get("专业")),
            "level": cls._repair_management_plain_text(fields.get("事件等级")),
            "source": cls._repair_management_plain_text(
                fields.get("事件发现来源（统一）")
                or fields.get("事件发现来源")
                or ""
            ),
            "status": cls._repair_management_plain_text(
                fields.get("最终状态")
                or fields.get("事件状态")
                or fields.get("事件目前进展")
                or ""
            ),
            "occurrence_time": cls._repair_management_plain_text(
                fields.get("事件发生时间")
            ),
            "sent_time": cls._repair_management_plain_text(
                fields.get("事件进展响应时间")
                or fields.get("创建时间")
                or fields.get("事件发生时间")
            ),
            "progress_update": cls._repair_management_plain_text(
                fields.get("最后更新时间") or fields.get("进展更新时间") or ""
            ),
            "fault_reason": cls._repair_management_plain_text(
                fields.get("事件发生原因") or fields.get("告警描述") or title
            ),
            "fault_phenomenon": cls._repair_management_plain_text(
                fields.get("告警描述") or fields.get("事件简述") or title
            ),
            "transfer_to_overhaul": fields.get("是否转检修"),
            "display_fields": fields,
            "raw_fields": raw_fields,
        }

    def _load_repair_management_cmdb_records_remote(
        self,
        *,
        force_refresh: bool = False,
    ) -> tuple[list[FieldMeta], dict[str, FieldMeta], list[dict[str, Any]]]:
        now = time.monotonic()
        with self._repair_management_cmdb_cache_lock:
            cached = self._repair_management_cmdb_cache
            if (
                not force_refresh
                and isinstance(cached, dict)
                and now - float(cached.get("loaded_at") or 0) <= 120.0
            ):
                return (
                    list(cached.get("metas") or []),
                    dict(cached.get("meta_by_name") or {}),
                    list(cached.get("records") or []),
                )
            metas, meta_by_name = self._load_table_fields(
                app_token=REPAIR_SOURCE_APP_TOKEN,
                table_id=REPAIR_CMDB_TABLE_ID,
            )
            records = self._search_table_records(
                app_token=REPAIR_SOURCE_APP_TOKEN,
                table_id=REPAIR_CMDB_TABLE_ID,
                meta_by_name=meta_by_name,
                work_type=WORK_TYPE_REPAIR,
                notice_type=NOTICE_TYPE_REPAIR,
                field_names=("智航唯一ID", "设备名称", "分类名称", "位置", "楼栋"),
                limit=500,
            )
            self._repair_management_cmdb_cache = {
                "loaded_at": time.monotonic(),
                "metas": list(metas),
                "meta_by_name": dict(meta_by_name),
                "records": list(records),
            }
            return metas, meta_by_name, records

    def _load_repair_management_cmdb_records(
        self,
        *,
        force_refresh: bool = False,
    ) -> tuple[list[FieldMeta], dict[str, FieldMeta], list[dict[str, Any]]]:

        def load_remote() -> tuple[list[FieldMeta], list[dict[str, Any]]]:
            metas, _meta_by_name, records = (
                self._load_repair_management_cmdb_records_remote(
                    force_refresh=force_refresh
                )
            )
            return metas, records

        return self._load_repair_snapshot_source(
            source_key=REPAIR_SNAPSHOT_SOURCE_CMDB,
            app_token=REPAIR_SOURCE_APP_TOKEN,
            table_id=REPAIR_CMDB_TABLE_ID,
            loader=load_remote,
            force_refresh=force_refresh,
        )

    def _load_repair_management_routing_records(
        self,
    ) -> list[dict[str, Any]]:
        def load_remote() -> tuple[list[FieldMeta], list[dict[str, Any]]]:
            now = time.monotonic()
            with self._repair_management_routing_cache_lock:
                cached = self._repair_management_routing_cache
                if (
                    isinstance(cached, dict)
                    and now - float(cached.get("loaded_at") or 0) <= 300.0
                ):
                    return (
                        list(cached.get("metas") or []),
                        list(cached.get("records") or []),
                    )
            metas, meta_by_name = self._load_table_fields(
                app_token=REPAIR_SOURCE_APP_TOKEN,
                table_id=REPAIR_ROUTING_TABLE_ID,
            )
            records = self._search_table_records(
                app_token=REPAIR_SOURCE_APP_TOKEN,
                table_id=REPAIR_ROUTING_TABLE_ID,
                meta_by_name=meta_by_name,
                work_type=WORK_TYPE_REPAIR,
                notice_type=NOTICE_TYPE_REPAIR,
                field_names=("区域", "楼栋", "专业", "通报群组", "维修审核人"),
                limit=500,
            )
            with self._repair_management_routing_cache_lock:
                self._repair_management_routing_cache = {
                    "loaded_at": time.monotonic(),
                    "metas": list(metas),
                    "records": list(records),
                }
            return metas, records

        _metas, _meta_by_name, records = self._load_repair_snapshot_source(
            source_key=REPAIR_SNAPSHOT_SOURCE_ROUTING,
            app_token=REPAIR_SOURCE_APP_TOKEN,
            table_id=REPAIR_ROUTING_TABLE_ID,
            loader=load_remote,
        )
        return records

    def _repair_management_target_record_in_scope(
        self,
        record: dict[str, Any],
        scope: str,
    ) -> bool:
        normalized_scope = self._normalize_scope(scope)
        if normalized_scope == "ALL":
            return True
        fields = record.get("display_fields") if isinstance(record.get("display_fields"), dict) else {}
        codes = self._repair_building_codes_from_value(
            fields.get("楼栋") or fields.get("名称（标题）") or fields.get("检修概述")
        )
        return bool(codes) and self._scope_matches_buildings(normalized_scope, codes)

    def list_repair_management_repair_candidates(
        self,
        *,
        scope: str = "ALL",
        event_record_id: str = "",
        month: str | None = None,
        query: str = "",
        limit: int = 80,
    ) -> dict[str, Any]:
        event: dict[str, Any] = {}
        if str(event_record_id or "").strip():
            event = self._event_snapshot_record_for_repair(
                scope=scope,
                record_id=event_record_id,
                month=month,
            )
        _metas, _meta_by_name, records = self._load_repair_management_target_records()
        query_text = str(query or "").strip().lower()
        candidates: list[dict[str, Any]] = []
        for record in records:
            if not self._repair_management_target_record_in_scope(record, scope):
                continue
            fields = record.get("display_fields") if isinstance(record.get("display_fields"), dict) else {}
            haystack = "\n".join(str(value or "") for value in fields.values()).lower()
            if query_text and query_text not in haystack:
                continue
            score, reasons = self._repair_management_candidate_score(event, record) if event else (0, [])
            building_codes = self._repair_building_codes_from_value(
                fields.get("楼栋") or fields.get("名称（标题）") or fields.get("检修概述")
            )
            candidates.append(
                {
                    "record_id": str(record.get("record_id") or ""),
                    "title": str(fields.get("名称（标题）") or fields.get("检修概述") or "未命名检修"),
                    "building": self._building_label_from_codes(building_codes),
                    "building_codes": building_codes,
                    "specialty": str(fields.get("专业") or ""),
                    "status": str(fields.get("检修状态") or ""),
                    "location": str(fields.get("位置") or ""),
                    "urgency": str(fields.get("紧急程度") or ""),
                    "fault_time": str(fields.get("发生故障时间") or ""),
                    "actual_start_time": str(fields.get("实际开始时间") or ""),
                    "actual_end_time": str(fields.get("实际结束时间") or ""),
                    "repair_device": str(fields.get("维修设备") or ""),
                    "repair_fault": str(fields.get("维修故障") or ""),
                    "score": score,
                    "match_reasons": reasons,
                    "recommended": bool(event and score >= 72),
                }
            )
        candidates.sort(
            key=lambda item: (int(item.get("score") or 0), str(item.get("actual_start_time") or item.get("fault_time") or "")),
            reverse=True,
        )
        max_limit = max(1, min(int(limit or 80), 200))
        visible = candidates[:max_limit]
        auto_selected_ids: list[str] = []
        if visible and int(visible[0].get("score") or 0) >= 70:
            second_score = int(visible[1].get("score") or 0) if len(visible) > 1 else 0
            if int(visible[0].get("score") or 0) - second_score >= 12:
                auto_selected_ids = [str(visible[0].get("record_id") or "")]
        return {
            "scope": self._normalize_scope(scope),
            "event_record_id": str(event_record_id or "").strip(),
            "records": visible,
            "total": len(candidates),
            "returned": len(visible),
            "auto_selected_ids": auto_selected_ids,
        }

    @classmethod
    def _repair_followup_field_payload(cls, meta: FieldMeta) -> dict[str, Any]:
        field_name = str(meta.field_name or "").strip()
        editable = (
            field_name in REPAIR_FOLLOWUP_FORM_FIELD_NAMES
            and not cls._field_meta_is_readonly(meta)
        )
        return {
            "field_id": meta.field_id,
            "field_name": field_name,
            "field_type": int(meta.field_type or 0),
            "ui_type": str(meta.ui_type or ""),
            "options": list(meta.option_names or []),
            "editable": editable,
        }

    @classmethod
    def _repair_followup_brand_model_options(
        cls,
        *,
        records: list[dict[str, Any]],
        summary_record: dict[str, Any] | None,
        meta_by_name: dict[str, FieldMeta],
    ) -> dict[str, list[str]]:
        brand_meta = meta_by_name.get(REPAIR_FOLLOWUP_BRAND_FIELD_NAME)
        model_meta = meta_by_name.get(REPAIR_FOLLOWUP_MODEL_FIELD_NAME)
        if brand_meta is None or model_meta is None:
            return {}

        mapped: dict[str, list[str]] = {}

        def add_pair(brand_value: Any, model_value: Any) -> None:
            brand = cls._repair_management_option_name(brand_meta, brand_value)
            model = cls._repair_management_option_name(model_meta, model_value)
            if not brand or not model:
                return
            models = mapped.setdefault(brand, [])
            if model not in models:
                models.append(model)

        for record in records:
            display = record.get("display_fields") if isinstance(record.get("display_fields"), dict) else {}
            raw = record.get("raw_fields") if isinstance(record.get("raw_fields"), dict) else {}
            add_pair(
                display.get(REPAIR_FOLLOWUP_BRAND_FIELD_NAME)
                or raw.get(REPAIR_FOLLOWUP_BRAND_FIELD_NAME),
                display.get(REPAIR_FOLLOWUP_MODEL_FIELD_NAME)
                or raw.get(REPAIR_FOLLOWUP_MODEL_FIELD_NAME),
            )

        if isinstance(summary_record, dict):
            display = summary_record.get("display_fields") if isinstance(summary_record.get("display_fields"), dict) else {}
            raw = summary_record.get("raw_fields") if isinstance(summary_record.get("raw_fields"), dict) else {}
            add_pair(
                display.get("设备品牌") or raw.get("设备品牌"),
                display.get("设备型号") or raw.get("设备型号"),
            )

        brand_order = {
            name: index for index, name in enumerate(brand_meta.option_names)
        }
        model_order = {
            name: index for index, name in enumerate(model_meta.option_names)
        }
        return {
            brand: sorted(models, key=lambda item: (model_order.get(item, 10**9), item))
            for brand, models in sorted(
                mapped.items(),
                key=lambda item: (brand_order.get(item[0], 10**9), item[0]),
            )
        }

    @classmethod
    def _repair_equipment_catalog_brand_model_options(
        cls,
        records: list[dict[str, Any]],
    ) -> dict[str, list[str]]:
        mapped: dict[str, list[str]] = {}
        for record in records:
            display = (
                record.get("display_fields")
                if isinstance(record.get("display_fields"), dict)
                else {}
            )
            raw = (
                record.get("raw_fields")
                if isinstance(record.get("raw_fields"), dict)
                else {}
            )
            brand = cls._repair_management_plain_text(
                display.get(REPAIR_EQUIPMENT_CATALOG_BRAND_FIELD_NAME)
                or raw.get(REPAIR_EQUIPMENT_CATALOG_BRAND_FIELD_NAME)
            )
            model = cls._repair_management_plain_text(
                display.get(REPAIR_EQUIPMENT_CATALOG_MODEL_FIELD_NAME)
                or raw.get(REPAIR_EQUIPMENT_CATALOG_MODEL_FIELD_NAME)
            )
            if not brand or not model:
                continue
            models = mapped.setdefault(brand, [])
            if model not in models:
                models.append(model)
        return mapped

    @staticmethod
    def _normalize_repair_catalog_mapping(value: Any) -> dict[str, list[str]]:
        if not isinstance(value, dict):
            return {}
        normalized: dict[str, list[str]] = {}
        for brand_value, model_values in value.items():
            brand = str(brand_value or "").strip()
            if not brand or not isinstance(model_values, list):
                continue
            models = list(
                dict.fromkeys(
                    str(model or "").strip()
                    for model in model_values
                    if str(model or "").strip()
                )
            )
            if models:
                normalized[brand] = models
        return normalized

    def _load_repair_followup_brand_model_catalog(
        self,
        meta_by_name: dict[str, FieldMeta],
        *,
        force_refresh: bool = False,
    ) -> dict[str, list[str]]:
        del meta_by_name
        now = time.time()
        with self._repair_followup_catalog_cache_lock:
            cached = self._repair_followup_catalog_cache
            if (
                not force_refresh
                and isinstance(cached, dict)
                and now - float(cached.get("refreshed_at") or 0)
                <= REPAIR_FOLLOWUP_CATALOG_CACHE_TTL_SECONDS
            ):
                return self._normalize_repair_catalog_mapping(cached.get("mapping"))

            persisted = self._state_store.get_backend_runtime(
                REPAIR_FOLLOWUP_CATALOG_RUNTIME_KEY
            ) or {}
            persisted_mapping = self._normalize_repair_catalog_mapping(
                persisted.get("mapping")
            )
            persisted_at = float(
                persisted.get("refreshed_at") or persisted.get("updated_at") or 0
            )
            if (
                not force_refresh
                and persisted_mapping
                and now - persisted_at <= REPAIR_FOLLOWUP_CATALOG_CACHE_TTL_SECONDS
            ):
                self._repair_followup_catalog_cache = {
                    "refreshed_at": persisted_at,
                    "mapping": persisted_mapping,
                }
                return persisted_mapping

            try:
                _metas, catalog_meta_by_name = self._load_table_fields(
                    app_token=REPAIR_SOURCE_APP_TOKEN,
                    table_id=REPAIR_EQUIPMENT_CATALOG_TABLE_ID,
                )
                missing_fields = [
                    field_name
                    for field_name in (
                        REPAIR_EQUIPMENT_CATALOG_BRAND_FIELD_NAME,
                        REPAIR_EQUIPMENT_CATALOG_MODEL_FIELD_NAME,
                    )
                    if field_name not in catalog_meta_by_name
                ]
                if missing_fields:
                    raise PortalError(
                        "设备目录缺少字段：" + "、".join(missing_fields)
                    )
                records = self._search_table_records(
                    app_token=REPAIR_SOURCE_APP_TOKEN,
                    table_id=REPAIR_EQUIPMENT_CATALOG_TABLE_ID,
                    meta_by_name=catalog_meta_by_name,
                    work_type=WORK_TYPE_REPAIR,
                    notice_type=NOTICE_TYPE_REPAIR,
                    field_names=(
                        REPAIR_EQUIPMENT_CATALOG_BRAND_FIELD_NAME,
                        REPAIR_EQUIPMENT_CATALOG_MODEL_FIELD_NAME,
                    ),
                    limit=REPAIR_FOLLOWUP_CATALOG_MAX_RECORDS,
                )
                mapping = self._repair_equipment_catalog_brand_model_options(records)
                if not mapping:
                    raise PortalError("设备目录没有可用的品牌型号数据。")
            except Exception:
                if persisted_mapping:
                    self._repair_followup_catalog_cache = {
                        "refreshed_at": persisted_at,
                        "mapping": persisted_mapping,
                    }
                    return persisted_mapping
                raise

            self._repair_followup_catalog_cache = {
                "refreshed_at": now,
                "mapping": mapping,
            }
            self._state_store.put_backend_runtime(
                REPAIR_FOLLOWUP_CATALOG_RUNTIME_KEY,
                {
                    "source_app_token": REPAIR_SOURCE_APP_TOKEN,
                    "source_table_id": REPAIR_EQUIPMENT_CATALOG_TABLE_ID,
                    "refreshed_at": now,
                    "record_count": len(records),
                    "brand_count": len(mapping),
                    "mapping": mapping,
                },
            )
            return mapping

    @classmethod
    def _repair_followup_parent_ids(cls, record: dict[str, Any]) -> list[str]:
        raw_fields = (
            record.get("raw_fields")
            if isinstance(record.get("raw_fields"), dict)
            else {}
        )
        display_fields = (
            record.get("display_fields")
            if isinstance(record.get("display_fields"), dict)
            else {}
        )
        return cls._repair_management_record_ids(
            raw_fields.get(REPAIR_FOLLOWUP_PARENT_ID_FIELD_NAME)
            or display_fields.get(REPAIR_FOLLOWUP_PARENT_ID_FIELD_NAME)
        )

    @classmethod
    def _repair_followup_order_key(
        cls,
        record: dict[str, Any],
    ) -> tuple[int, int, str]:
        display_fields = (
            record.get("display_fields")
            if isinstance(record.get("display_fields"), dict)
            else {}
        )
        raw_fields = (
            record.get("raw_fields")
            if isinstance(record.get("raw_fields"), dict)
            else {}
        )
        created_ms = cls._repair_management_datetime_ms(
            display_fields.get("创建时间")
            or raw_fields.get("创建时间")
            or record.get("created_time")
        )
        modified_ms = cls._repair_management_datetime_ms(
            record.get("last_modified_time")
            or display_fields.get("最后修改时间")
            or raw_fields.get("最后修改时间")
        )
        effective_created = created_ms or modified_ms or 0
        return (
            effective_created,
            modified_ms or effective_created,
            str(record.get("record_id") or ""),
        )

    def _load_repair_followups_for_summary(
        self,
        summary_record_id: str,
        *,
        limit: int = 200,
        force_refresh: bool = False,
    ) -> tuple[list[FieldMeta], dict[str, FieldMeta], list[dict[str, Any]]]:
        summary_id = str(summary_record_id or "").strip()
        if not summary_id:
            raise PortalError("读取维修跟进记录缺少维修项目记录。")
        if self._repair_snapshots_enabled:
            metas, meta_by_name, records = self._load_repair_followup_snapshot(
                force_refresh=force_refresh
            )
            return (
                metas,
                meta_by_name,
                [
                    record
                    for record in records
                    if summary_id in self._repair_followup_parent_ids(record)
                ][: max(1, min(int(limit or 200), 500))],
            )
        metas, meta_by_name = self._ensure_repair_followup_parent_id_field()
        records = self._search_table_records(
            app_token=REPAIR_SOURCE_APP_TOKEN,
            table_id=REPAIR_FOLLOWUP_TABLE_ID,
            meta_by_name=meta_by_name,
            work_type=WORK_TYPE_REPAIR,
            notice_type=NOTICE_TYPE_REPAIR,
            field_names=REPAIR_FOLLOWUP_READ_FIELD_NAMES,
            sort_field="创建时间",
            limit=max(1, min(int(limit or 200), 500)),
            filter_payload={
                "conjunction": "and",
                "conditions": [
                    {
                        "field_name": REPAIR_FOLLOWUP_PARENT_ID_FIELD_NAME,
                        "operator": "is",
                        "value": [summary_id],
                    }
                ],
            },
        )
        return metas, meta_by_name, records

    def _load_repair_followup_snapshot(
        self,
        *,
        force_refresh: bool = False,
    ) -> tuple[list[FieldMeta], dict[str, FieldMeta], list[dict[str, Any]]]:
        def load_remote() -> tuple[list[FieldMeta], list[dict[str, Any]]]:
            metas, meta_by_name = self._ensure_repair_followup_parent_id_field()
            records = self._load_table_records(
                app_token=REPAIR_SOURCE_APP_TOKEN,
                table_id=REPAIR_FOLLOWUP_TABLE_ID,
                meta_by_name=meta_by_name,
                work_type=WORK_TYPE_REPAIR,
                notice_type=NOTICE_TYPE_REPAIR,
            )
            return metas, records

        return self._load_repair_snapshot_source(
            source_key=REPAIR_SNAPSHOT_SOURCE_FOLLOWUPS,
            app_token=REPAIR_SOURCE_APP_TOKEN,
            table_id=REPAIR_FOLLOWUP_TABLE_ID,
            loader=load_remote,
            parent_field_name=REPAIR_FOLLOWUP_PARENT_ID_FIELD_NAME,
            force_refresh=force_refresh,
        )

    def list_repair_management_cmdb_candidates(
        self,
        *,
        scope: str = "ALL",
        query: str = "",
        limit: int = 160,
    ) -> dict[str, Any]:
        _metas, _meta_by_name, records = self._load_repair_management_cmdb_records()
        normalized_scope = self._normalize_scope(scope)
        query_text = str(query or "").strip().lower()
        candidates: list[dict[str, Any]] = []
        for record in records:
            fields = record.get("display_fields") or {}
            building = str(fields.get("楼栋") or "").strip()
            if (
                normalized_scope != "ALL"
                and building
                and not self._scope_matches_building(normalized_scope, building)
            ):
                continue
            haystack = "\n".join(
                str(fields.get(name) or "")
                for name in ("智航唯一ID", "设备名称", "分类名称", "位置", "楼栋")
            ).lower()
            if query_text and query_text not in haystack:
                continue
            candidates.append(
                {
                    "record_id": str(record.get("record_id") or ""),
                    "title": str(fields.get("设备名称") or "未命名设备"),
                    "unique_id": str(fields.get("智航唯一ID") or ""),
                    "category": str(fields.get("分类名称") or ""),
                    "location": str(fields.get("位置") or ""),
                    "building": building,
                }
            )
        max_limit = max(1, min(int(limit or 160), 500))
        return {
            "scope": normalized_scope,
            "records": candidates[:max_limit],
            "total": len(candidates),
            "returned": min(len(candidates), max_limit),
        }

    def list_repair_followup_people(
        self,
        *,
        scope: str = "ALL",
        query: str = "",
        limit: int = 80,
        refresh: bool = False,
    ) -> dict[str, Any]:
        normalized_scope = self._normalize_scope(scope)
        query_text = re.sub(r"\s+", "", str(query or "")).casefold()
        people = self._load_signature_people(force=bool(refresh))
        matched: list[dict[str, Any]] = []
        for person in people:
            name = str(person.get("name") or "").strip()
            user_id = str(person.get("open_id") or "").strip()
            if not name:
                continue
            if query_text and query_text not in re.sub(r"\s+", "", name).casefold():
                continue
            matched.append(
                {
                    "person_record_id": str(person.get("record_id") or "").strip(),
                    "user_id": user_id,
                    "name": name,
                    "employee_no": str(person.get("employee_no") or "").strip(),
                    "building": str(person.get("building") or "").strip(),
                    "position": str(person.get("position") or "").strip(),
                    "selectable": bool(user_id),
                }
            )
        matched.sort(
            key=lambda item: (
                0 if item.get("selectable") else 1,
                str(item.get("name") or ""),
                str(item.get("employee_no") or ""),
            )
        )
        max_limit = max(1, min(int(limit or 80), 200))
        return {
            "scope": normalized_scope,
            "people": matched[:max_limit],
            "total": len(matched),
            "returned": min(len(matched), max_limit),
        }

    def get_repair_followup_records(
        self,
        *,
        summary_record_id: str,
        scope: str = "ALL",
        query: str = "",
        limit: int = 100,
        offset: int = 0,
        force_refresh: bool = False,
    ) -> dict[str, Any]:
        summary_id = str(summary_record_id or "").strip()
        if not summary_id:
            raise PortalError("读取维修跟进记录缺少维修项目记录。")
        summary_record = self._ensure_repair_management_record_in_scope(summary_id, scope)
        metas, meta_by_name, records = self._load_repair_followups_for_summary(
            summary_id,
            limit=500,
            force_refresh=force_refresh,
        )
        query_text = str(query or "").strip().lower()
        selected: list[dict[str, Any]] = []
        for record in records:
            if summary_id not in self._repair_followup_parent_ids(record):
                continue
            fields = record.get("display_fields") or {}
            if query_text and query_text not in "\n".join(
                str(value or "") for value in fields.values()
            ).lower():
                continue
            raw_fields = record.get("raw_fields") or {}
            cmdb_ids = self._repair_management_record_ids(
                raw_fields.get(REPAIR_FOLLOWUP_CMDB_FIELD_NAME)
            )
            selected.append(
                {
                    "record_id": str(record.get("record_id") or ""),
                    "title": str(
                        self._repair_management_plain_text(fields.get("维修简述"))
                        or self._repair_management_plain_text(fields.get("维修进展描述"))
                        or self._repair_management_plain_text(
                            fields.get(REPAIR_FOLLOWUP_DEVICE_NAME_FIELD_NAME)
                        )
                        or "未命名跟进记录"
                    ),
                    "created_time": str(fields.get("创建时间") or ""),
                    "progress": str(fields.get("维修进度") or ""),
                    "display_fields": fields,
                    "raw_fields": raw_fields,
                    "cmdb_record_ids": cmdb_ids,
                }
            )
        selected.sort(key=lambda item: item.get("created_time") or "", reverse=True)
        max_limit = max(1, min(int(limit or 100), 200))
        page_offset = max(0, int(offset or 0))
        page_records = selected[page_offset : page_offset + max_limit]
        current_mapping = self._repair_followup_brand_model_options(
            records=records,
            summary_record=summary_record,
            meta_by_name=meta_by_name,
        )
        try:
            brand_model_options = self._load_repair_followup_brand_model_catalog(
                meta_by_name
            )
        except PortalError:
            brand_model_options = {}
        for brand, models in current_mapping.items():
            merged_models = brand_model_options.setdefault(brand, [])
            for model in models:
                if model not in merged_models:
                    merged_models.append(model)
        field_payloads = [self._repair_followup_field_payload(meta) for meta in metas]
        for field_payload in field_payloads:
            if field_payload.get("field_name") != REPAIR_FOLLOWUP_BRAND_FIELD_NAME:
                continue
            field_payload["options"] = list(
                dict.fromkeys(
                    [
                        *list(field_payload.get("options") or []),
                        *brand_model_options.keys(),
                    ]
                )
            )
        return {
            "summary_record_id": summary_id,
            "parent_id_field": REPAIR_FOLLOWUP_PARENT_ID_FIELD_NAME,
            "relation_mode": "record_id",
            "fields": field_payloads,
            "records": page_records,
            "total": len(selected),
            "returned": len(page_records),
            "offset": page_offset,
            "has_more": page_offset + len(page_records) < len(selected),
            "brand_model_options": brand_model_options,
        }

    @classmethod
    def _repair_followup_summary_copy_fields(
        cls,
        summary_record: dict[str, Any] | None,
        meta_by_name: dict[str, FieldMeta],
    ) -> dict[str, Any]:
        if not isinstance(summary_record, dict):
            return {}
        summary_display = (
            summary_record.get("display_fields")
            if isinstance(summary_record.get("display_fields"), dict)
            else {}
        )
        summary_raw = (
            summary_record.get("raw_fields")
            if isinstance(summary_record.get("raw_fields"), dict)
            else {}
        )
        copied: dict[str, Any] = {}
        for target_name, source_names in REPAIR_FOLLOWUP_SUMMARY_COPY_MAPPINGS:
            target_meta = meta_by_name.get(target_name)
            if target_meta is None or cls._field_meta_is_readonly(target_meta):
                continue
            prefer_raw = int(target_meta.field_type or 0) in {4, 5, 11, 18, 21}
            source_value: Any = None
            for source_name in source_names:
                primary = summary_raw if prefer_raw else summary_display
                fallback = summary_display if prefer_raw else summary_raw
                source_value = primary.get(source_name)
                if source_value in (None, "", [], {}):
                    source_value = fallback.get(source_name)
                if source_value not in (None, "", [], {}):
                    break
            if source_value not in (None, "", [], {}):
                copied[target_name] = source_value
        return copied

    def _prepare_repair_followup_fields(
        self,
        *,
        summary_record_id: str,
        fields: dict[str, Any],
        cmdb_record_ids: list[str] | tuple[str, ...] | None = None,
        summary_record: dict[str, Any] | None = None,
        existing_record: dict[str, Any] | None = None,
        meta_by_name: dict[str, FieldMeta] | None = None,
    ) -> tuple[dict[str, Any], list[str]]:
        if meta_by_name is None:
            _metas, meta_by_name = self._ensure_repair_followup_parent_id_field()
        cleaned = {
            str(name or "").strip(): value
            for name, value in (fields or {}).items()
            if str(name or "").strip() in REPAIR_FOLLOWUP_WRITABLE_FIELD_NAMES
            and str(name or "").strip() in meta_by_name
        }
        submitted_names = set(cleaned)
        cleaned[REPAIR_FOLLOWUP_PARENT_ID_FIELD_NAME] = summary_record_id

        for field_name, value in self._repair_followup_summary_copy_fields(
            summary_record,
            meta_by_name,
        ).items():
            if field_name not in submitted_names:
                cleaned[field_name] = value

        if existing_record is None and isinstance(summary_record, dict):
            summary_display = (
                summary_record.get("display_fields")
                if isinstance(summary_record.get("display_fields"), dict)
                else {}
            )
            summary_raw = (
                summary_record.get("raw_fields")
                if isinstance(summary_record.get("raw_fields"), dict)
                else {}
            )
            summary_mappings = (
                (REPAIR_FOLLOWUP_DEVICE_NAME_FIELD_NAME, "设备名称", False),
                (REPAIR_FOLLOWUP_DEVICE_NUMBER_FIELD_NAME, "设备编号", False),
                (REPAIR_FOLLOWUP_BRAND_FIELD_NAME, "设备品牌", False),
                (REPAIR_FOLLOWUP_MODEL_FIELD_NAME, "设备型号", False),
                (REPAIR_FOLLOWUP_REPAIR_PARTY_FIELD_NAME, "维修方", False),
                (REPAIR_FOLLOWUP_SUPPLIER_FIELD_NAME, "供应商名称", False),
                (
                    REPAIR_FOLLOWUP_SUPPLIER_PERSON_FIELD_NAME,
                    "供应商维修人员",
                    False,
                ),
                ("设备生产日期", "设备生产日期", True),
                ("设备使用年限", "设备使用年限", True),
                ("设备容量KW/AH", "设备容量KW/AH", False),
                ("是否质保期内", "是否质保期内", False),
                (
                    "随工人员（我方维修人员）",
                    "随工人员（或我方维修人员）",
                    True,
                ),
            )
            for target_name, source_name, prefer_raw in summary_mappings:
                if target_name in submitted_names or target_name not in meta_by_name:
                    continue
                source_value = (
                    summary_raw.get(source_name)
                    if prefer_raw
                    else summary_display.get(source_name)
                )
                if source_value in (None, "", [], {}):
                    source_value = (
                        summary_display.get(source_name)
                        if prefer_raw
                        else summary_raw.get(source_name)
                    )
                if source_value not in (None, "", [], {}):
                    cleaned[target_name] = source_value

        cmdb_ids = self._repair_management_record_ids(list(cmdb_record_ids or []))
        if not cmdb_ids and existing_record:
            existing_raw = existing_record.get("raw_fields") or {}
            cmdb_ids = self._repair_management_record_ids(
                existing_raw.get(REPAIR_FOLLOWUP_CMDB_FIELD_NAME)
            )
        cmdb_warnings: list[str] = []
        if cmdb_ids:
            _cmdb_metas, cmdb_meta_by_name, _cmdb_records = (
                self._load_repair_management_cmdb_records()
            )
            cmdb_records = self._load_table_records_by_ids(
                app_token=REPAIR_SOURCE_APP_TOKEN,
                table_id=REPAIR_CMDB_TABLE_ID,
                meta_by_name=cmdb_meta_by_name,
                work_type=WORK_TYPE_REPAIR,
                notice_type=NOTICE_TYPE_REPAIR,
                record_ids=cmdb_ids,
            )
            cmdb_fields_list = [
                record.get("display_fields") or {}
                for record in cmdb_records
            ]
            categories = list(
                dict.fromkeys(
                    str(item.get("分类名称") or "").strip()
                    for item in cmdb_fields_list
                    if str(item.get("分类名称") or "").strip()
                )
            )
            device_names = list(
                dict.fromkeys(
                    str(item.get("设备名称") or "").strip()
                    for item in cmdb_fields_list
                    if str(item.get("设备名称") or "").strip()
                )
            )
            cleaned[REPAIR_FOLLOWUP_CMDB_FIELD_NAME] = cmdb_ids
            if REPAIR_FOLLOWUP_DEVICE_NAME_FIELD_NAME not in submitted_names:
                if len(categories) == 1:
                    cleaned[REPAIR_FOLLOWUP_DEVICE_NAME_FIELD_NAME] = categories[0]
                elif len(categories) > 1:
                    cmdb_warnings.append(
                        "所选 CMDB 设备属于不同分类，请手动确认设备名称。"
                    )
            if REPAIR_FOLLOWUP_DEVICE_NUMBER_FIELD_NAME not in submitted_names and device_names:
                cleaned[REPAIR_FOLLOWUP_DEVICE_NUMBER_FIELD_NAME] = "、".join(
                    device_names
                )

        existing_display = (
            existing_record.get("display_fields")
            if isinstance(existing_record, dict)
            and isinstance(existing_record.get("display_fields"), dict)
            else {}
        )
        existing_raw = (
            existing_record.get("raw_fields")
            if isinstance(existing_record, dict)
            and isinstance(existing_record.get("raw_fields"), dict)
            else {}
        )
        repair_party_value = cleaned.get(REPAIR_FOLLOWUP_REPAIR_PARTY_FIELD_NAME)
        if repair_party_value in (None, "", [], {}):
            repair_party_value = (
                existing_display.get(REPAIR_FOLLOWUP_REPAIR_PARTY_FIELD_NAME)
                or existing_raw.get(REPAIR_FOLLOWUP_REPAIR_PARTY_FIELD_NAME)
            )
        repair_party_text = self._repair_management_canonical_select_text(
            REPAIR_FOLLOWUP_REPAIR_PARTY_FIELD_NAME,
            self._repair_management_plain_text(repair_party_value),
        )
        if repair_party_text == "我方":
            for supplier_field_name in (
                REPAIR_FOLLOWUP_SUPPLIER_FIELD_NAME,
                REPAIR_FOLLOWUP_SUPPLIER_PERSON_FIELD_NAME,
            ):
                if supplier_field_name in meta_by_name:
                    cleaned[supplier_field_name] = None
        if "维修简述" in meta_by_name:
            start_value = cleaned.get("维修开始时间")
            worker_value = cleaned.get("随工人员（我方维修人员）")
            if worker_value in (None, "", [], {}):
                worker_value = existing_raw.get("随工人员（我方维修人员）")
            worker_name = self._repair_management_plain_text(worker_value)
            start_ms = self._repair_management_datetime_ms(start_value)
            repair_name = self._repair_management_plain_text(cleaned.get("维修名称"))
            if start_ms is not None and worker_name:
                cleaned["维修简述"] = (
                    f"{dt.datetime.fromtimestamp(start_ms / 1000).strftime('%Y/%m/%d')}"
                    f" - {worker_name}"
                )
            elif repair_name:
                cleaned["维修简述"] = repair_name
            elif start_ms is not None:
                cleaned["维修简述"] = dt.datetime.fromtimestamp(
                    start_ms / 1000
                ).strftime("%Y/%m/%d")

        prepared, warnings = self._coerce_repair_management_fields(
            cleaned,
            meta_by_name,
            custom_single_select_fields=REPAIR_FOLLOWUP_CUSTOM_SINGLE_SELECT_FIELDS,
        )
        if not prepared:
            raise PortalError("没有可写入的维修跟进字段。")
        return prepared, list(dict.fromkeys([*cmdb_warnings, *warnings]))

    @staticmethod
    def _created_record_id(payload: dict[str, Any]) -> str:
        return str(
            (((payload.get("data") or {}).get("record") or {}).get("record_id"))
            or ((payload.get("data") or {}).get("record_id"))
            or ""
        ).strip()

    def _create_repair_followup_fields(
        self,
        fields: dict[str, Any],
        meta_by_name: dict[str, FieldMeta],
    ) -> tuple[dict[str, Any], dict[str, Any], list[str]]:
        try:
            payload = self._create_record_fields(
                app_token=REPAIR_SOURCE_APP_TOKEN,
                table_id=REPAIR_FOLLOWUP_TABLE_ID,
                fields=fields,
            )
            return payload, dict(fields), []
        except PortalError as exc:
            if "1254062" not in str(exc):
                raise

        single_select_fields = {
            name: value
            for name, value in fields.items()
            if (meta := meta_by_name.get(name)) is not None
            and (
                int(meta.field_type or 0) == 3
                or "singleselect" in str(meta.ui_type or "").strip().lower()
            )
        }
        base_fields = {
            name: value for name, value in fields.items() if name not in single_select_fields
        }
        if not single_select_fields or not base_fields:
            raise PortalError(
                "新增维修跟进失败：飞书无法转换单选字段，请刷新页面后重新选择。"
            )

        payload = self._create_record_fields(
            app_token=REPAIR_SOURCE_APP_TOKEN,
            table_id=REPAIR_FOLLOWUP_TABLE_ID,
            fields=base_fields,
        )
        record_id = self._created_record_id(payload)
        if not record_id:
            raise PortalError("维修跟进基础记录已提交，但未返回记录 ID。")

        effective_fields = dict(base_fields)
        warnings: list[str] = []
        try:
            for field_name, value in single_select_fields.items():
                try:
                    self._patch_record_fields(
                        app_token=REPAIR_SOURCE_APP_TOKEN,
                        table_id=REPAIR_FOLLOWUP_TABLE_ID,
                        record_id=record_id,
                        fields={field_name: value},
                    )
                    effective_fields[field_name] = value
                except PortalError as exc:
                    if "1254062" not in str(exc):
                        raise
                    warnings.append(
                        f"{field_name}未被飞书接受，已保留跟进记录，请重新选择后更新"
                    )
        except Exception:
            with suppress(Exception):
                self._delete_record_fields(
                    app_token=REPAIR_SOURCE_APP_TOKEN,
                    table_id=REPAIR_FOLLOWUP_TABLE_ID,
                    record_id=record_id,
                )
            raise
        return payload, effective_fields, warnings

    def _sync_repair_management_from_followup(
        self,
        *,
        summary_record_id: str,
        followup_record_id: str = "",
        scope: str = "ALL",
    ) -> list[str]:
        summary_id = str(summary_record_id or "").strip()
        if not summary_id:
            raise PortalError("缺少维修项目记录 ID。")
        with self._repair_followup_sync_locks_guard:
            sync_lock = self._repair_followup_sync_locks.setdefault(
                summary_id,
                threading.RLock(),
            )
        with sync_lock:
            return self._sync_repair_management_from_followup_unlocked(
                summary_record_id=summary_id,
                followup_record_id=followup_record_id,
                scope=scope,
            )

    def _sync_repair_management_from_followup_unlocked(
        self,
        *,
        summary_record_id: str,
        followup_record_id: str = "",
        scope: str = "ALL",
    ) -> list[str]:
        summary = self._ensure_repair_management_record_in_scope(
            summary_record_id,
            scope,
        )
        _metas, meta_by_name, _records = (
            self._load_repair_management_project_records()
        )
        raw_fields = summary.get("raw_fields") or {}
        event_ids = self._repair_management_record_ids(raw_fields.get("关联事件单"))
        repair_ids = self._repair_management_record_ids(raw_fields.get("设备检修关联"))[:1]
        _followup_metas, _followup_meta_by_name, followup_records = (
            self._load_repair_followups_for_summary(summary_record_id, limit=500)
        )
        linked_followups = [
            record
            for record in followup_records
            if summary_record_id in self._repair_followup_parent_ids(record)
        ]
        current_followup_id = str(followup_record_id or "").strip()
        linked_ids = {
            str(record.get("record_id") or "").strip()
            for record in linked_followups
        }
        if current_followup_id and current_followup_id not in linked_ids:
            current_records = self._load_table_records_by_ids(
                app_token=REPAIR_SOURCE_APP_TOKEN,
                table_id=REPAIR_FOLLOWUP_TABLE_ID,
                meta_by_name=_followup_meta_by_name,
                work_type=WORK_TYPE_REPAIR,
                notice_type=NOTICE_TYPE_REPAIR,
                record_ids=[current_followup_id],
            )
            if (
                current_records
                and summary_record_id
                in self._repair_followup_parent_ids(current_records[0])
            ):
                linked_followups.append(current_records[0])
        linked_followups.sort(
            key=self._repair_followup_order_key,
            reverse=True,
        )
        followup_ids = [
            str(record.get("record_id") or "").strip()
            for record in linked_followups
            if str(record.get("record_id") or "").strip()
        ]
        auto = self._build_repair_management_prefill(
            scope=scope,
            event_record_id=event_ids[0] if event_ids else "",
            repair_record_ids=repair_ids,
            followup_record_ids=followup_ids,
            month=None,
            meta_by_name=meta_by_name,
            allow_multiple_followups=True,
        )
        auto_fields = dict(auto.get("fields") or {})
        if "维修跟进记录" in meta_by_name:
            auto_fields["维修跟进记录"] = (
                ",".join(followup_ids) if followup_ids else None
            )
        for field_name in REPAIR_MANAGEMENT_FOLLOWUP_AUTO_FIELD_NAMES:
            if field_name in meta_by_name:
                auto_fields.setdefault(field_name, None)
        prepared, warnings = self._coerce_repair_management_fields(
            auto_fields,
            meta_by_name,
        )
        if prepared:
            self._patch_record_fields(
                app_token=REPAIR_SOURCE_APP_TOKEN,
                table_id=REPAIR_MANAGEMENT_TABLE_ID,
                record_id=summary_record_id,
                fields=prepared,
            )
            self._upsert_repair_snapshot_fields(
                source_key=REPAIR_SNAPSHOT_SOURCE_PROJECTS,
                record_id=summary_record_id,
                fields=prepared,
            )
        self._invalidate_repair_management_status_cache()
        return list(dict.fromkeys([*(auto.get("warnings") or []), *warnings]))

    def create_repair_followup_record(
        self,
        *,
        summary_record_id: str,
        fields: dict[str, Any],
        cmdb_record_ids: list[str] | tuple[str, ...] | None = None,
        operation_id: str = "",
        scope: str = "ALL",
    ) -> dict[str, Any]:
        summary_id = str(summary_record_id or "").strip()
        summary = self._ensure_repair_management_record_in_scope(summary_id, scope)
        existing = self.get_repair_followup_records(
            summary_record_id=summary_id,
            scope=scope,
            limit=1,
        )
        source_fields = dict(fields or {})
        source_fields.setdefault(
            "是否本维修单第一次提交跟进记录",
            "否" if int(existing.get("total") or 0) else "是",
        )
        _metas, meta_by_name = self._ensure_repair_followup_parent_id_field()
        prepared, warnings = self._prepare_repair_followup_fields(
            summary_record_id=summary_id,
            fields=source_fields,
            cmdb_record_ids=cmdb_record_ids,
            summary_record=summary,
            meta_by_name=meta_by_name,
        )
        _metas, meta_by_name = self._ensure_repair_followup_select_options(
            prepared,
            meta_by_name,
        )
        stable_operation_id = str(operation_id or "").strip()
        operation: dict[str, Any] | None = None
        if stable_operation_id:
            operation_payload = {
                "summary_record_id": summary_id,
                "scope": self._normalize_scope(scope),
                "fields": prepared,
                "cmdb_record_ids": list(cmdb_record_ids or []),
            }
            payload_hash = hashlib.sha256(
                json.dumps(
                    operation_payload,
                    ensure_ascii=False,
                    sort_keys=True,
                    separators=(",", ":"),
                    default=str,
                ).encode("utf-8")
            ).hexdigest()
            try:
                operation = self._state_store.begin_repair_management_operation(
                    stable_operation_id,
                    operation_type="followup_create",
                    scope=self._normalize_scope(scope),
                    payload_hash=payload_hash,
                    summary_record_id=summary_id,
                )
            except ValueError as exc:
                raise PortalError(f"维修跟进操作标识冲突：{exc}") from exc

            if not operation.get("created"):
                existing_result = dict(operation.get("result") or {})
                existing_record_id = str(
                    operation.get("record_id")
                    or existing_result.get("record_id")
                    or ""
                ).strip()
                if operation.get("status") == "completed" and existing_record_id:
                    existing_result["idempotent_replay"] = True
                    return existing_result
                if existing_record_id:
                    replay_warnings = [
                        warning
                        for warning in (existing_result.get("warnings") or [])
                        if "维修项目汇总暂未同步" not in str(warning)
                    ]
                    try:
                        replay_warnings.extend(
                            self._sync_repair_management_from_followup(
                                summary_record_id=summary_id,
                                followup_record_id=existing_record_id,
                                scope=scope,
                            )
                        )
                        existing_result.update(
                            {
                                "record_id": existing_record_id,
                                "summary_record_id": summary_id,
                                "warnings": list(dict.fromkeys(replay_warnings)),
                                "summary_sync_pending": False,
                                "idempotent_replay": True,
                            }
                        )
                        self._state_store.update_repair_management_operation(
                            stable_operation_id,
                            status="completed",
                            result=existing_result,
                            error="",
                        )
                    except Exception as exc:
                        replay_warnings.append(
                            f"跟进记录已保存，维修项目汇总暂未同步：{exc}"
                        )
                        existing_result.update(
                            {
                                "record_id": existing_record_id,
                                "summary_record_id": summary_id,
                                "warnings": list(dict.fromkeys(replay_warnings)),
                                "summary_sync_pending": True,
                                "idempotent_replay": True,
                            }
                        )
                        self._state_store.update_repair_management_operation(
                            stable_operation_id,
                            status="sync_pending",
                            result=existing_result,
                            error=str(exc),
                        )
                    return existing_result
                if operation.get("status") == "started":
                    raise PortalError("该维修跟进正在后台创建，请稍后刷新，勿重复提交。")
                raise PortalError(
                    "上一次维修跟进创建结果未确认。为避免重复记录，请刷新跟进列表后再操作。"
                )

        try:
            result, prepared, create_warnings = self._create_repair_followup_fields(
                prepared,
                meta_by_name,
            )
            record_id = self._created_record_id(result)
            if not record_id:
                raise PortalError("维修跟进记录已提交，但未返回记录 ID。")
            self._upsert_repair_snapshot_fields(
                source_key=REPAIR_SNAPSHOT_SOURCE_FOLLOWUPS,
                record_id=record_id,
                fields=prepared,
                parent_record_id=summary_id,
            )
        except Exception as exc:
            if stable_operation_id:
                self._state_store.update_repair_management_operation(
                    stable_operation_id,
                    status="failed",
                    error=str(exc),
                )
            raise

        response = {
            "record_id": record_id,
            "summary_record_id": summary_id,
            "fields": prepared,
            "warnings": list(dict.fromkeys([*warnings, *create_warnings])),
            "summary_sync_pending": False,
        }
        if stable_operation_id:
            self._state_store.update_repair_management_operation(
                stable_operation_id,
                status="remote_written",
                record_id=record_id,
                summary_record_id=summary_id,
                result=response,
                error="",
            )
        try:
            sync_warnings = self._sync_repair_management_from_followup(
                summary_record_id=summary_id,
                followup_record_id=record_id,
                scope=scope,
            )
            response["warnings"] = list(
                dict.fromkeys([*response["warnings"], *sync_warnings])
            )
            if stable_operation_id:
                self._state_store.update_repair_management_operation(
                    stable_operation_id,
                    status="completed",
                    result=response,
                    error="",
                )
        except Exception as exc:
            response["warnings"] = list(
                dict.fromkeys(
                    [
                        *response["warnings"],
                        f"跟进记录已保存，维修项目汇总暂未同步：{exc}",
                    ]
                )
            )
            response["summary_sync_pending"] = True
            if stable_operation_id:
                self._state_store.update_repair_management_operation(
                    stable_operation_id,
                    status="sync_pending",
                    result=response,
                    error=str(exc),
                )
        return response

    def update_repair_followup_record(
        self,
        record_id: str,
        *,
        summary_record_id: str,
        fields: dict[str, Any],
        cmdb_record_ids: list[str] | tuple[str, ...] | None = None,
        scope: str = "ALL",
    ) -> dict[str, Any]:
        summary_id = str(summary_record_id or "").strip()
        _metas, meta_by_name = self._ensure_repair_followup_parent_id_field()
        records = self._load_table_records_by_ids(
            app_token=REPAIR_SOURCE_APP_TOKEN,
            table_id=REPAIR_FOLLOWUP_TABLE_ID,
            meta_by_name=meta_by_name,
            work_type=WORK_TYPE_REPAIR,
            notice_type=NOTICE_TYPE_REPAIR,
            record_ids=[record_id],
        )
        if not records:
            raise PortalError("该维修跟进记录已不存在，请刷新后重试。")
        existing = records[0]
        if summary_id not in self._repair_followup_parent_ids(existing):
            raise PortalError("该维修跟进记录不属于当前检修单。")
        summary = self._ensure_repair_management_record_in_scope(summary_id, scope)
        prepared, warnings = self._prepare_repair_followup_fields(
            summary_record_id=summary_id,
            fields=fields,
            cmdb_record_ids=cmdb_record_ids,
            summary_record=summary,
            existing_record=existing,
            meta_by_name=meta_by_name,
        )
        _metas, meta_by_name = self._ensure_repair_followup_select_options(
            prepared,
            meta_by_name,
        )
        self._patch_record_fields(
            app_token=REPAIR_SOURCE_APP_TOKEN,
            table_id=REPAIR_FOLLOWUP_TABLE_ID,
            record_id=record_id,
            fields=prepared,
        )
        self._upsert_repair_snapshot_fields(
            source_key=REPAIR_SNAPSHOT_SOURCE_FOLLOWUPS,
            record_id=record_id,
            fields=prepared,
            parent_record_id=summary_id,
        )
        summary_sync_pending = False
        try:
            sync_warnings = self._sync_repair_management_from_followup(
                summary_record_id=summary_id,
                followup_record_id=record_id,
                scope=scope,
            )
        except Exception as exc:
            summary_sync_pending = True
            sync_warnings = [f"跟进记录已更新，维修项目汇总暂未同步：{exc}"]
        return {
            "record_id": str(record_id or "").strip(),
            "summary_record_id": summary_id,
            "fields": prepared,
            "warnings": list(dict.fromkeys([*warnings, *sync_warnings])),
            "summary_sync_pending": summary_sync_pending,
        }

    def delete_repair_followup_record(
        self,
        record_id: str,
        *,
        summary_record_id: str,
        scope: str = "ALL",
    ) -> dict[str, Any]:
        summary_id = str(summary_record_id or "").strip()
        _metas, meta_by_name = self._ensure_repair_followup_parent_id_field()
        records = self._load_table_records_by_ids(
            app_token=REPAIR_SOURCE_APP_TOKEN,
            table_id=REPAIR_FOLLOWUP_TABLE_ID,
            meta_by_name=meta_by_name,
            work_type=WORK_TYPE_REPAIR,
            notice_type=NOTICE_TYPE_REPAIR,
            record_ids=[record_id],
        )
        if not records:
            raise PortalError("该维修跟进记录已不存在，请刷新后重试。")
        if summary_id not in self._repair_followup_parent_ids(records[0]):
            raise PortalError("该维修跟进记录不属于当前检修单。")
        self._ensure_repair_management_record_in_scope(summary_id, scope)
        self._delete_record_fields(
            app_token=REPAIR_SOURCE_APP_TOKEN,
            table_id=REPAIR_FOLLOWUP_TABLE_ID,
            record_id=record_id,
        )
        self._delete_repair_snapshot_item(
            REPAIR_SNAPSHOT_SOURCE_FOLLOWUPS,
            record_id,
        )
        summary_sync_pending = False
        try:
            warnings = self._sync_repair_management_from_followup(
                summary_record_id=summary_id,
                scope=scope,
            )
        except Exception as exc:
            summary_sync_pending = True
            warnings = [f"跟进记录已删除，维修项目汇总暂未同步：{exc}"]
        return {
            "record_id": str(record_id or "").strip(),
            "summary_record_id": summary_id,
            "deleted": True,
            "warnings": warnings,
            "summary_sync_pending": summary_sync_pending,
        }

    @classmethod
    def _repair_followup_match_key(cls, value: Any) -> str:
        text = cls._repair_management_plain_text(value).casefold()
        return re.sub(r"[\s\u3000,，。.;；:：\-—_/\\]+", "", text)

    @classmethod
    def _repair_followup_comparable_value(
        cls,
        meta: FieldMeta | None,
        value: Any,
    ) -> Any:
        if value in (None, "", [], {}):
            return None
        field_type = int(meta.field_type or 0) if meta is not None else 0
        ui_type = str(meta.ui_type or "").strip().lower() if meta is not None else ""
        if field_type == 5 or "datetime" in ui_type:
            return cls._repair_management_datetime_ms(value)
        if field_type == 2 or ui_type in {"number", "progress", "currency"}:
            with suppress(TypeError, ValueError):
                return float(value)
        if field_type in {18, 21} or "link" in ui_type:
            return tuple(cls._repair_management_record_ids(value))
        if field_type == 11 or ui_type == "user":
            users = cls._repair_management_users(value)
            return tuple(
                sorted(
                    str(user.get("id") or user.get("name") or "").strip()
                    for user in users
                    if str(user.get("id") or user.get("name") or "").strip()
                )
            )
        if field_type in {3, 4} or "select" in ui_type:
            items = cls._repair_management_select_items(value)
            normalized = [
                cls._repair_management_canonical_select_text(
                    meta.field_name if meta is not None else "",
                    item,
                )
                for item in items
            ]
            return tuple(item for item in normalized if item)
        return cls._repair_management_plain_text(value)

    def backfill_repair_followup_records(
        self,
        *,
        force: bool = False,
    ) -> dict[str, Any]:
        now = time.time()
        runtime = self._state_store.get_backend_runtime(
            REPAIR_FOLLOWUP_BACKFILL_RUNTIME_KEY
        ) or {}
        last_finished = float(runtime.get("finished_ts") or 0)
        if (
            not force
            and last_finished > 0
            and now - last_finished < REPAIR_FOLLOWUP_BACKFILL_INTERVAL_SECONDS
        ):
            return {**runtime, "skipped": True}

        with self._repair_followup_backfill_lock:
            summary_metas, summary_meta_by_name = self._load_table_fields(
                app_token=REPAIR_SOURCE_APP_TOKEN,
                table_id=REPAIR_MANAGEMENT_TABLE_ID,
            )
            summaries = self._load_table_records(
                app_token=REPAIR_SOURCE_APP_TOKEN,
                table_id=REPAIR_MANAGEMENT_TABLE_ID,
                meta_by_name=summary_meta_by_name,
                work_type=WORK_TYPE_REPAIR,
                notice_type=NOTICE_TYPE_REPAIR,
            )
            followup_metas, followup_meta_by_name = (
                self._ensure_repair_followup_parent_id_field()
            )
            followups = self._load_table_records(
                app_token=REPAIR_SOURCE_APP_TOKEN,
                table_id=REPAIR_FOLLOWUP_TABLE_ID,
                meta_by_name=followup_meta_by_name,
                work_type=WORK_TYPE_REPAIR,
                notice_type=NOTICE_TYPE_REPAIR,
            )
            summary_by_id = {
                str(record.get("record_id") or "").strip(): record
                for record in summaries
                if str(record.get("record_id") or "").strip()
            }
            summary_by_order: dict[str, list[dict[str, Any]]] = {}
            summary_by_name: dict[str, list[dict[str, Any]]] = {}
            for summary in summaries:
                display = summary.get("display_fields") or {}
                raw = summary.get("raw_fields") or {}
                order_key = self._repair_followup_match_key(
                    display.get("维修单号") or raw.get("维修单号")
                )
                name_key = self._repair_followup_match_key(
                    display.get("维修名称") or raw.get("维修名称")
                )
                if order_key:
                    summary_by_order.setdefault(order_key, []).append(summary)
                if name_key:
                    summary_by_name.setdefault(name_key, []).append(summary)

            updated = 0
            unchanged = 0
            unresolved = 0
            failed = 0
            warnings: list[str] = []
            linked_by_summary: dict[str, list[str]] = {}
            for followup in followups:
                record_id = str(followup.get("record_id") or "").strip()
                display = followup.get("display_fields") or {}
                raw = followup.get("raw_fields") or {}
                parent_ids = self._repair_followup_parent_ids(followup)
                summary: dict[str, Any] | None = None
                if len(parent_ids) == 1:
                    summary = summary_by_id.get(parent_ids[0])
                if summary is None:
                    order_key = self._repair_followup_match_key(
                        display.get("维修单号") or raw.get("维修单号")
                    )
                    candidates = summary_by_order.get(order_key, []) if order_key else []
                    if len(candidates) != 1:
                        name_key = self._repair_followup_match_key(
                            display.get("维修名称") or raw.get("维修名称")
                        )
                        candidates = summary_by_name.get(name_key, []) if name_key else []
                    if len(candidates) == 1:
                        summary = candidates[0]
                if summary is None:
                    unresolved += 1
                    continue
                summary_id = str(summary.get("record_id") or "").strip()
                if not summary_id or not record_id:
                    unresolved += 1
                    continue
                parent_was_persisted = (
                    len(parent_ids) == 1 and parent_ids[0] == summary_id
                )
                try:
                    prepared, record_warnings = self._prepare_repair_followup_fields(
                        summary_record_id=summary_id,
                        fields={},
                        cmdb_record_ids=self._repair_management_record_ids(
                            raw.get(REPAIR_FOLLOWUP_CMDB_FIELD_NAME)
                        ),
                        summary_record=summary,
                        existing_record=followup,
                        meta_by_name=followup_meta_by_name,
                    )
                    _metas, followup_meta_by_name = (
                        self._ensure_repair_followup_select_options(
                            prepared,
                            followup_meta_by_name,
                        )
                    )
                    patch_fields: dict[str, Any] = {}
                    for field_name, value in prepared.items():
                        meta = followup_meta_by_name.get(field_name)
                        current_value = (
                            raw.get(field_name)
                            if meta is not None
                            and int(meta.field_type or 0) in {2, 3, 4, 5, 11, 15, 17, 18, 21}
                            else display.get(field_name)
                        )
                        if self._repair_followup_comparable_value(
                            meta,
                            current_value,
                        ) != self._repair_followup_comparable_value(meta, value):
                            patch_fields[field_name] = value
                    if patch_fields:
                        self._patch_record_fields(
                            app_token=REPAIR_SOURCE_APP_TOKEN,
                            table_id=REPAIR_FOLLOWUP_TABLE_ID,
                            record_id=record_id,
                            fields=patch_fields,
                        )
                        updated += 1
                    else:
                        unchanged += 1
                    linked_by_summary.setdefault(summary_id, []).append(record_id)
                    warnings.extend(record_warnings)
                except Exception as exc:
                    failed += 1
                    if parent_was_persisted:
                        linked_by_summary.setdefault(summary_id, []).append(record_id)
                    warnings.append(f"{record_id}: {exc}")

            backlink_updates = 0
            backlink_meta = summary_meta_by_name.get("维修跟进记录")
            if backlink_meta is not None and int(backlink_meta.field_type or 0) == 1:
                for summary_id, record_ids in linked_by_summary.items():
                    normalized_ids = list(dict.fromkeys(record_ids))
                    summary = summary_by_id[summary_id]
                    current_ids = self._repair_management_record_ids(
                        (summary.get("raw_fields") or {}).get("维修跟进记录")
                        or (summary.get("display_fields") or {}).get("维修跟进记录")
                    )
                    if current_ids == normalized_ids:
                        continue
                    try:
                        self._patch_record_fields(
                            app_token=REPAIR_SOURCE_APP_TOKEN,
                            table_id=REPAIR_MANAGEMENT_TABLE_ID,
                            record_id=summary_id,
                            fields={
                                "维修跟进记录": (
                                    ",".join(normalized_ids)
                                    if normalized_ids
                                    else None
                                )
                            },
                        )
                        backlink_updates += 1
                    except Exception as exc:
                        failed += 1
                        warnings.append(f"{summary_id}回写维修跟进记录失败: {exc}")

            result = {
                "status": "partial" if failed else "complete",
                "summary_count": len(summaries),
                "followup_count": len(followups),
                "updated": updated,
                "unchanged": unchanged,
                "unresolved": unresolved,
                "failed": failed,
                "backlink_updates": backlink_updates,
                "warnings": list(dict.fromkeys(warnings))[-50:],
                "finished_at": dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "finished_ts": time.time(),
            }
            self._state_store.put_backend_runtime(
                REPAIR_FOLLOWUP_BACKFILL_RUNTIME_KEY,
                result,
            )
            return result

    @classmethod
    def _coerce_repair_management_fields(
        cls,
        fields: dict[str, Any],
        meta_by_name: dict[str, FieldMeta],
        *,
        custom_single_select_fields: set[str] | None = None,
        custom_multi_select_fields: set[str] | None = None,
    ) -> tuple[dict[str, Any], list[str]]:
        result: dict[str, Any] = {}
        warnings: list[str] = []
        custom_single_select_fields = custom_single_select_fields or set()
        custom_multi_select_fields = (
            REPAIR_MANAGEMENT_CUSTOM_MULTI_SELECT_FIELDS
            if custom_multi_select_fields is None
            else custom_multi_select_fields
        )
        for name, value in fields.items():
            meta = meta_by_name.get(str(name or "").strip())
            if meta is None:
                continue
            if value is None or value == "":
                result[meta.field_name] = None
                continue
            ui_type = str(meta.ui_type or "").strip().lower()
            field_type = int(meta.field_type or 0)
            if field_type == 5 or "datetime" in ui_type:
                timestamp = cls._repair_management_datetime_ms(value)
                if timestamp is None:
                    warnings.append(f"{meta.field_name}时间格式无法识别")
                    continue
                result[meta.field_name] = timestamp
                continue
            if field_type == 2 or ui_type == "number":
                try:
                    number = float(value)
                    result[meta.field_name] = int(number) if number.is_integer() else number
                except (TypeError, ValueError):
                    warnings.append(f"{meta.field_name}不是有效数字")
                continue
            if field_type == 4 or "multiselect" in ui_type:
                values = cls._repair_management_select_items(value)
                normalized: list[str] = []
                invalid: list[str] = []
                for item in values:
                    option_name = cls._repair_management_option_name(meta, item)
                    option_name = cls._repair_management_canonical_select_text(
                        meta.field_name,
                        option_name,
                    )
                    if option_name:
                        if option_name not in normalized:
                            normalized.append(option_name)
                        continue
                    text = cls._repair_management_canonical_select_text(
                        meta.field_name,
                        cls._repair_management_plain_text(item),
                    )
                    if text:
                        if meta.field_name in custom_multi_select_fields:
                            if text not in normalized:
                                normalized.append(text)
                        else:
                            invalid.append(text)
                if normalized:
                    result[meta.field_name] = normalized
                if invalid and meta.option_names:
                    warnings.append(
                        f"{meta.field_name}包含无效选项：{'、'.join(dict.fromkeys(invalid))}，已跳过"
                    )
                continue
            if field_type == 3 or "singleselect" in ui_type:
                option_name = cls._repair_management_option_name(meta, value)
                option_name = cls._repair_management_canonical_select_text(
                    meta.field_name,
                    option_name,
                )
                if option_name:
                    result[meta.field_name] = option_name
                else:
                    text = cls._repair_management_plain_text(value)
                    if text and meta.field_name in custom_single_select_fields:
                        result[meta.field_name] = text
                    elif text:
                        warnings.append(
                            f"{meta.field_name}的值“{text}”不在当前单选选项中，已跳过"
                        )
                continue
            if field_type == 11 or ui_type == "user":
                users = cls._repair_management_users(value)
                if users:
                    result[meta.field_name] = users
                elif isinstance(value, (list, dict)):
                    result[meta.field_name] = []
                else:
                    warnings.append(f"{meta.field_name}必须选择飞书人员")
                continue
            if field_type == 15 or ui_type == "url":
                if isinstance(value, dict) and str(value.get("link") or "").strip():
                    result[meta.field_name] = value
                else:
                    link = str(value or "").strip()
                    if link:
                        result[meta.field_name] = {"link": link, "text": link}
                continue
            if field_type in {18, 21} or "link" in ui_type:
                values = value if isinstance(value, list) else cls._repair_management_record_ids(value)
                if values or isinstance(value, list):
                    result[meta.field_name] = values
                continue
            if isinstance(value, (dict, list)):
                result[meta.field_name] = json.dumps(value, ensure_ascii=False, separators=(",", ":"))
            else:
                result[meta.field_name] = value
        return result, list(dict.fromkeys(warnings))

    @classmethod
    def _repair_management_option_name(
        cls,
        meta: FieldMeta,
        value: Any,
    ) -> str:
        candidates: list[str] = []
        if isinstance(value, dict):
            for key in ("name", "text", "value", "id", "option_id"):
                text = cls._repair_management_plain_text(value.get(key))
                if text and text not in candidates:
                    candidates.append(text)
        else:
            text = cls._repair_management_plain_text(value)
            if text:
                candidates.append(text)

        if not meta.option_names:
            return candidates[0] if candidates else ""

        names_by_casefold = {
            str(option or "").strip().casefold(): str(option or "").strip()
            for option in meta.option_names
            if str(option or "").strip()
        }
        for candidate in candidates:
            mapped = str((meta.options_map or {}).get(candidate) or "").strip()
            if mapped:
                return mapped
            exact = names_by_casefold.get(candidate.casefold())
            if exact:
                return exact
        return ""

    @staticmethod
    def _repair_management_prefill_put(
        result: dict[str, Any],
        meta_by_name: dict[str, FieldMeta],
        field_name: str,
        value: Any,
        *,
        overwrite: bool = False,
    ) -> None:
        if field_name not in meta_by_name or value is None or value == "":
            return
        if not overwrite and result.get(field_name) not in (None, "", []):
            return
        result[field_name] = value

    @classmethod
    def _apply_repair_management_form_derivatives(
        cls,
        fields: dict[str, Any],
        meta_by_name: dict[str, FieldMeta],
    ) -> dict[str, Any]:
        derived = dict(fields)
        specialty = cls._repair_management_plain_text(derived.get("所属专业"))
        if specialty and "专业（推送消息用）" in meta_by_name:
            derived["专业（推送消息用）"] = specialty

        building = cls._repair_management_plain_text(
            derived.get("所属数据中心/楼栋-使用")
        )
        if building:
            if "区域" in meta_by_name:
                derived.setdefault("区域", "华东一")
            if "数据中心" in meta_by_name:
                derived.setdefault("数据中心", "南通")
            legacy_name = "所属数据中心/楼栋（关联CMDB唯一ID关联,DE不选）"
            legacy_meta = meta_by_name.get(legacy_name)
            building_codes = cls._repair_building_codes_from_value(building)
            if legacy_meta is not None and len(building_codes) == 1:
                canonical_building = cls._repair_management_building_label(
                    building_codes
                )
                option_name = cls._repair_management_option_name(
                    legacy_meta,
                    canonical_building,
                )
                if option_name:
                    derived[legacy_name] = option_name
        return derived

    @classmethod
    def _repair_management_building_label(cls, codes: list[str]) -> str:
        if not codes:
            return ""
        labels = ["南通110站" if code == "110" else f"南通{code}楼" for code in codes]
        return "、".join(labels)

    def _build_repair_management_prefill(
        self,
        *,
        scope: str,
        event_record_id: str,
        repair_record_ids: list[str] | tuple[str, ...] | None,
        followup_record_ids: list[str] | tuple[str, ...] | None,
        month: str | None,
        meta_by_name: dict[str, FieldMeta],
        allow_multiple_followups: bool = False,
    ) -> dict[str, Any]:
        event_id = str(event_record_id or "").strip()
        event: dict[str, Any] = {}
        if event_id:
            event = self._event_snapshot_record_for_repair(
                scope=scope,
                record_id=event_id,
                month=month,
            )

        requested_repair_ids = list(
            dict.fromkeys(
                str(record_id or "").strip()
                for record_id in (repair_record_ids or [])
                if str(record_id or "").strip().startswith("rec")
            )
        )
        if len(requested_repair_ids) > 1:
            raise PortalError("设备检修关联只能选择一条记录。")
        selected_repairs: list[dict[str, Any]] = []
        if requested_repair_ids:
            _metas, _repair_meta_by_name, repair_records = self._load_repair_management_target_records()
            del repair_records
            repair_app_token, repair_table_id = (
                self._repair_management_repair_source_config()
            )
            selected_repairs = self._load_table_records_by_ids(
                app_token=repair_app_token,
                table_id=repair_table_id,
                meta_by_name=_repair_meta_by_name,
                work_type=WORK_TYPE_REPAIR,
                notice_type=NOTICE_TYPE_REPAIR,
                record_ids=requested_repair_ids,
            )
            for record in selected_repairs:
                if not self._repair_management_target_record_in_scope(record, scope):
                    raise PortalError("当前账号无权关联该楼栋设备检修记录。")

        requested_followup_ids = list(
            dict.fromkeys(
                str(record_id or "").strip()
                for record_id in (followup_record_ids or [])
                if str(record_id or "").strip().startswith("rec")
            )
        )
        if len(requested_followup_ids) > 1 and not allow_multiple_followups:
            raise PortalError("维修跟进记录只能选择一条记录。")
        selected_followups: list[dict[str, Any]] = []
        if requested_followup_ids:
            _metas, followup_meta_by_name = (
                self._ensure_repair_followup_parent_id_field()
            )
            selected_followups = self._load_table_records_by_ids(
                app_token=REPAIR_SOURCE_APP_TOKEN,
                table_id=REPAIR_FOLLOWUP_TABLE_ID,
                meta_by_name=followup_meta_by_name,
                work_type=WORK_TYPE_REPAIR,
                notice_type=NOTICE_TYPE_REPAIR,
                record_ids=requested_followup_ids,
            )
            parent_ids: set[str] = set()
            for record in selected_followups:
                record_parent_ids = self._repair_followup_parent_ids(record)
                if len(record_parent_ids) != 1:
                    raise PortalError("维修跟进记录缺少有效的检修单 ID，请先核对。")
                parent_ids.update(record_parent_ids)
            if len(parent_ids) != 1:
                raise PortalError("维修跟进记录不属于同一个检修单，请先核对。")
            self._ensure_repair_management_record_in_scope(
                next(iter(parent_ids)),
                scope,
            )

        result: dict[str, Any] = {}
        warnings: list[str] = []
        if event:
            event_fields = (
                event.get("display_fields")
                if isinstance(event.get("display_fields"), dict)
                else {}
            )
            event_raw_fields = (
                event.get("raw_fields")
                if isinstance(event.get("raw_fields"), dict)
                else {}
            )
            event_codes = self._clean_building_codes(event.get("building_codes"))
            building_text = (
                self._repair_management_building_label(event_codes)
                if event_codes
                else str(
                    event.get("building")
                    or event_fields.get("机楼")
                    or event_fields.get("南通楼栋")
                    or ""
                ).strip()
            )
            source = self._repair_management_plain_text(
                event_fields.get("事件发现来源（统一）") or event.get("source")
            )
            source = {
                "BMS": "BMS系统",
                "BMS动环系统告警": "BMS系统",
                "方舟": "方舟系统",
            }.get(source, source)
            alarm_desc = self._repair_management_plain_text(
                event.get("alarm_desc") or event.get("title")
            )
            event_summary = self._repair_management_plain_text(
                event_fields.get("事件简述") or event.get("title") or alarm_desc
            )
            fault_reason = self._repair_management_plain_text(
                event.get("fault_reason")
                or event_fields.get("事件发生原因")
                or alarm_desc
            )
            fault_phenomenon = self._repair_management_plain_text(
                event_fields.get("故障现象")
                or event.get("fault_phenomenon")
                or event_fields.get("告警描述")
                or event_summary
            )
            specialty = self._repair_management_plain_text(event.get("specialty"))
            occurrence_time = self._repair_management_plain_text(
                event.get("occurrence_time")
            )
            self._repair_management_prefill_put(result, meta_by_name, "关联事件单", event_id)
            self._repair_management_prefill_put(
                result, meta_by_name, "维修名称", event_summary
            )
            self._repair_management_prefill_put(result, meta_by_name, "对应来源", [source] if source else [])
            self._repair_management_prefill_put(result, meta_by_name, "对应事件等级", event.get("level"))
            self._repair_management_prefill_put(result, meta_by_name, "故障发生时间", occurrence_time)
            self._repair_management_prefill_put(
                result, meta_by_name, "故障维修原因", fault_reason
            )
            self._repair_management_prefill_put(
                result, meta_by_name, "故障发生现象描述", fault_phenomenon
            )
            self._repair_management_prefill_put(result, meta_by_name, "所属专业", specialty)
            self._repair_management_prefill_put(result, meta_by_name, "专业（推送消息用）", specialty)
            self._repair_management_prefill_put(result, meta_by_name, "事件描述", event_summary)
            self._repair_management_prefill_put(result, meta_by_name, "所属数据中心/楼栋-使用", building_text)
            if len(event_codes) == 1:
                self._repair_management_prefill_put(
                    result,
                    meta_by_name,
                    "所属数据中心/楼栋（关联CMDB唯一ID关联,DE不选）",
                    building_text,
                )
            self._repair_management_prefill_put(result, meta_by_name, "区域", "华东一")
            self._repair_management_prefill_put(result, meta_by_name, "数据中心", "南通")
            self._repair_management_prefill_put(
                result,
                meta_by_name,
                "值班账号",
                event_raw_fields.get("值班账号"),
            )
            self._repair_management_prefill_put(
                result,
                meta_by_name,
                "消息推送人",
                event_raw_fields.get("工程师（消息推送）")
                or event_raw_fields.get("值班账号"),
            )

        if selected_repairs:
            repair_fields = [
                item.get("raw_fields") if isinstance(item.get("raw_fields"), dict) else {}
                for item in selected_repairs
            ]
            display_fields = [
                item.get("display_fields") if isinstance(item.get("display_fields"), dict) else {}
                for item in selected_repairs
            ]
            starts = [
                value
                for value in (
                    self._repair_management_datetime_ms(fields.get("实际开始时间"))
                    for fields in repair_fields
                )
                if value is not None
            ]
            ends = [
                value
                for value in (
                    self._repair_management_datetime_ms(fields.get("实际结束时间"))
                    for fields in repair_fields
                )
                if value is not None
            ]
            completed_count = sum(
                1
                for raw, display in zip(repair_fields, display_fields)
                if "结束" in str(display.get("检修状态") or "")
                or self._repair_management_datetime_ms(raw.get("实际结束时间")) is not None
            )
            progress = completed_count / len(selected_repairs)
            start_ms = min(starts) if starts else None
            all_completed = completed_count == len(selected_repairs)
            end_ms = max(ends) if all_completed and len(ends) == len(selected_repairs) else None
            self._repair_management_prefill_put(
                result,
                meta_by_name,
                "设备检修关联",
                requested_repair_ids[0],
                overwrite=True,
            )
            self._repair_management_prefill_put(result, meta_by_name, "维修开始时间", start_ms)
            self._repair_management_prefill_put(result, meta_by_name, "开始时间", start_ms)
            self._repair_management_prefill_put(result, meta_by_name, "维修结束时间（2026）", end_ms)
            self._repair_management_prefill_put(result, meta_by_name, "维修结束时间", end_ms)
            self._repair_management_prefill_put(result, meta_by_name, "当前维修进度", progress)
            if start_ms:
                effective_end = end_ms or int(dt.datetime.now().timestamp() * 1000)
                self._repair_management_prefill_put(
                    result,
                    meta_by_name,
                    "维修周期",
                    max(0, int((effective_end - start_ms) / 86_400_000)),
                )
            self._repair_management_prefill_put(
                result,
                meta_by_name,
                "检修通告名称",
                self._repair_management_unique_text(
                    [fields.get("名称（标题）") or fields.get("检修概述") for fields in display_fields]
                ),
            )
            self._repair_management_prefill_put(
                result,
                meta_by_name,
                "设备名称",
                self._repair_management_unique_text([fields.get("维修设备") for fields in display_fields]),
            )
            self._repair_management_prefill_put(
                result,
                meta_by_name,
                "维修进展描述",
                self._repair_management_unique_text([fields.get("进度（完成情况）") for fields in display_fields]),
            )
            repair_users: list[Any] = []
            for fields in repair_fields:
                users = fields.get("涉及值班账号")
                if isinstance(users, list):
                    repair_users.extend(users)
            if repair_users:
                unique_users: list[Any] = []
                seen_user_ids: set[str] = set()
                for user in repair_users:
                    if isinstance(user, dict):
                        user_key = str(
                            user.get("id")
                            or user.get("open_id")
                            or user.get("user_id")
                            or user
                        )
                    else:
                        user_key = str(user or "")
                    if not user_key or user_key in seen_user_ids:
                        continue
                    seen_user_ids.add(user_key)
                    unique_users.append(user)
                self._repair_management_prefill_put(
                    result,
                    meta_by_name,
                    "随工人员（或我方维修人员）",
                    unique_users,
                )
            self._repair_management_prefill_put(
                result,
                meta_by_name,
                "故障发生现象描述",
                self._repair_management_unique_text([fields.get("故障现象") for fields in display_fields]),
            )
            self._repair_management_prefill_put(
                result,
                meta_by_name,
                "故障维修原因",
                self._repair_management_unique_text(
                    [fields.get("故障原因") or fields.get("维修故障") for fields in display_fields]
                ),
            )
            repair_modes = [
                str(fields.get("维修方式") or "").strip()
                for fields in display_fields
                if str(fields.get("维修方式") or "").strip()
            ]
            if repair_modes and all(mode.startswith("自维") for mode in repair_modes):
                self._repair_management_prefill_put(result, meta_by_name, "维修方", "我方")

        if selected_followups:
            followup_raw_fields = [
                item.get("raw_fields") if isinstance(item.get("raw_fields"), dict) else {}
                for item in selected_followups
            ]
            followup_display_fields = [
                item.get("display_fields")
                if isinstance(item.get("display_fields"), dict)
                else {}
                for item in selected_followups
            ]
            latest_followup = max(
                selected_followups,
                key=self._repair_followup_order_key,
            )
            latest_index = selected_followups.index(latest_followup)
            latest_display = followup_display_fields[latest_index]
            latest_raw = followup_raw_fields[latest_index]

            def latest_value(field_name: str, *, prefer_raw: bool = False) -> Any:
                primary = latest_raw if prefer_raw else latest_display
                fallback = latest_display if prefer_raw else latest_raw
                value = primary.get(field_name)
                if value in (None, "", [], {}):
                    value = fallback.get(field_name)
                return value

            cmdb_record_ids = list(
                dict.fromkeys(
                    record_id
                    for record_id in self._repair_management_record_ids(
                        latest_raw.get(REPAIR_FOLLOWUP_CMDB_FIELD_NAME)
                        or latest_display.get(REPAIR_FOLLOWUP_CMDB_FIELD_NAME)
                    )
                )
            )
            selected_cmdb_records: list[dict[str, Any]] = []
            if cmdb_record_ids:
                _cmdb_metas, cmdb_meta_by_name, _cmdb_records = (
                    self._load_repair_management_cmdb_records()
                )
                selected_cmdb_records = self._load_table_records_by_ids(
                    app_token=REPAIR_SOURCE_APP_TOKEN,
                    table_id=REPAIR_CMDB_TABLE_ID,
                    meta_by_name=cmdb_meta_by_name,
                    work_type=WORK_TYPE_REPAIR,
                    notice_type=NOTICE_TYPE_REPAIR,
                    record_ids=cmdb_record_ids,
                )
                normalized_scope = self._normalize_scope(scope)
                if normalized_scope != "ALL":
                    for cmdb_record in selected_cmdb_records:
                        cmdb_fields = cmdb_record.get("display_fields") or {}
                        building = str(cmdb_fields.get("楼栋") or "").strip()
                        if building and not self._scope_matches_building(
                            normalized_scope, building
                        ):
                            raise PortalError("当前账号无权关联该楼栋 CMDB 设备。")

            self._repair_management_prefill_put(
                result,
                meta_by_name,
                "CMDB唯一id",
                ",".join(cmdb_record_ids),
                overwrite=True,
            )
            self._repair_management_prefill_put(
                result,
                meta_by_name,
                "是否有唯一id",
                "是"
                if cmdb_record_ids
                and any(
                    str((record.get("display_fields") or {}).get("智航唯一ID") or "").strip()
                    not in {"", "00000000"}
                    for record in selected_cmdb_records
                )
                else "否",
                overwrite=True,
            )
            self._repair_management_prefill_put(
                result,
                meta_by_name,
                "智航设备名称",
                self._repair_management_unique_text(
                    [
                        (record.get("display_fields") or {}).get("设备名称")
                        for record in selected_cmdb_records
                    ]
                ),
                overwrite=True,
            )
            self._repair_management_prefill_put(
                result,
                meta_by_name,
                "维修名称",
                latest_value("维修名称"),
                overwrite=True,
            )

            followup_start = self._repair_management_datetime_ms(
                latest_value("维修开始时间", prefer_raw=True)
            )
            followup_end = self._repair_management_datetime_ms(
                latest_value("维修结束时间", prefer_raw=True)
            )
            if followup_start is not None:
                self._repair_management_prefill_put(
                    result,
                    meta_by_name,
                    "维修开始时间",
                    followup_start,
                    overwrite=True,
                )
                self._repair_management_prefill_put(
                    result,
                    meta_by_name,
                    "开始时间",
                    followup_start,
                    overwrite=True,
                )
            if followup_end is not None:
                self._repair_management_prefill_put(
                    result,
                    meta_by_name,
                    "维修结束时间（2026）",
                    followup_end,
                    overwrite=True,
                )
                self._repair_management_prefill_put(
                    result,
                    meta_by_name,
                    "维修结束时间",
                    followup_end,
                    overwrite=True,
                )
            if followup_start is not None:
                effective_end = followup_end or int(dt.datetime.now().timestamp() * 1000)
                self._repair_management_prefill_put(
                    result,
                    meta_by_name,
                    "维修周期",
                    max(0, int((effective_end - followup_start) / 86_400_000)),
                    overwrite=True,
                )

            latest_progress_percent = self._repair_followup_progress_percent(
                latest_value("维修进度", prefer_raw=True)
            )
            if latest_progress_percent is not None:
                self._repair_management_prefill_put(
                    result,
                    meta_by_name,
                    "当前维修进度",
                    latest_progress_percent / 100,
                    overwrite=True,
                )
            self._repair_management_prefill_put(
                result,
                meta_by_name,
                "维修进展描述",
                latest_value("维修进展描述"),
                overwrite=True,
            )
            latest_created_ms = self._repair_followup_order_key(latest_followup)[0]
            if latest_created_ms:
                self._repair_management_prefill_put(
                    result,
                    meta_by_name,
                    "最新维修跟进时间",
                    latest_created_ms,
                    overwrite=True,
                )

            text_mappings = (
                ("设备名称", REPAIR_FOLLOWUP_DEVICE_NAME_FIELD_NAME),
                ("设备编号", REPAIR_FOLLOWUP_DEVICE_NUMBER_FIELD_NAME),
                ("设备品牌", REPAIR_FOLLOWUP_BRAND_FIELD_NAME),
                ("设备型号", REPAIR_FOLLOWUP_MODEL_FIELD_NAME),
                ("供应商名称", REPAIR_FOLLOWUP_SUPPLIER_FIELD_NAME),
                (
                    "供应商维修人员",
                    REPAIR_FOLLOWUP_SUPPLIER_PERSON_FIELD_NAME,
                ),
                ("设备容量KW/AH", "设备容量KW/AH"),
                ("跟进项", "跟进项（如有）"),
                ("后续整改措施", "后续整改措施（如有）"),
            )
            for target_name, source_name in text_mappings:
                self._repair_management_prefill_put(
                    result,
                    meta_by_name,
                    target_name,
                    latest_value(source_name),
                    overwrite=True,
                )

            self._repair_management_prefill_put(
                result,
                meta_by_name,
                "更换备件名称",
                latest_value("更换备件名称"),
                overwrite=True,
            )

            for target_name, source_name in (
                ("维修方", REPAIR_FOLLOWUP_REPAIR_PARTY_FIELD_NAME),
                ("是否质保期内", "是否质保期内"),
            ):
                self._repair_management_prefill_put(
                    result,
                    meta_by_name,
                    target_name,
                    latest_value(source_name),
                    overwrite=True,
                )

            production_date = self._repair_management_datetime_ms(
                latest_value("设备生产日期", prefer_raw=True)
            )
            self._repair_management_prefill_put(
                result,
                meta_by_name,
                "设备生产日期",
                production_date,
                overwrite=True,
            )

            latest_numbers: tuple[tuple[str, str], ...] = (
                ("设备使用年限", "设备使用年限"),
                ("更换备件数量", "更换备件数量"),
            )
            for target_name, source_name in latest_numbers:
                self._repair_management_prefill_put(
                    result,
                    meta_by_name,
                    target_name,
                    latest_value(source_name, prefer_raw=True),
                    overwrite=True,
                )

            completed_costs: list[float] = []
            for raw_fields, display_fields in zip(
                followup_raw_fields,
                followup_display_fields,
            ):
                progress_value = raw_fields.get("维修进度")
                if progress_value in (None, "", [], {}):
                    progress_value = display_fields.get("维修进度")
                progress_percent = self._repair_followup_progress_percent(progress_value)
                if progress_percent is not None and progress_percent >= 100:
                    cost_value = raw_fields.get("故障维修总费用")
                    if cost_value in (None, "", [], {}):
                        cost_value = display_fields.get("故障维修总费用")
                    with suppress(TypeError, ValueError):
                        completed_costs.append(float(cost_value))
            self._repair_management_prefill_put(
                result,
                meta_by_name,
                "故障维修总费用（跟进完成的维修项）",
                sum(completed_costs) if completed_costs else None,
                overwrite=True,
            )

            followup_users = self._repair_management_users(
                latest_value("随工人员（我方维修人员）", prefer_raw=True)
            )
            if followup_users:
                self._repair_management_prefill_put(
                    result,
                    meta_by_name,
                    "随工人员（或我方维修人员）",
                    followup_users,
                    overwrite=True,
                )
            approval_users = self._repair_management_users(
                latest_value("维修审批人", prefer_raw=True)
            )
            if approval_users:
                self._repair_management_prefill_put(
                    result,
                    meta_by_name,
                    "维修审批人",
                    approval_users,
                    overwrite=True,
                )
            groups = latest_value("推送群组", prefer_raw=True)
            if isinstance(groups, dict) and isinstance(groups.get("groups"), list):
                group_names = self._repair_management_unique_text(
                    [item.get("name") for item in groups.get("groups") if isinstance(item, dict)]
                )
                self._repair_management_prefill_put(
                    result,
                    meta_by_name,
                    "推送群组",
                    group_names,
                    overwrite=True,
                )
            self._repair_management_prefill_put(
                result,
                meta_by_name,
                "子项链接",
                latest_value("超链接", prefer_raw=True),
                overwrite=True,
            )

        building_value = str(
            result.get("所属数据中心/楼栋-使用")
            or result.get("所属数据中心/楼栋（关联CMDB唯一ID关联,DE不选）")
            or ""
        ).strip()
        specialty_value = str(
            result.get("所属专业") or result.get("专业（推送消息用）") or ""
        ).strip()
        if building_value and specialty_value:
            routing_matches: list[dict[str, Any]] = []
            for routing_record in self._load_repair_management_routing_records():
                routing_fields = routing_record.get("display_fields") or {}
                if str(routing_fields.get("楼栋") or "").strip() != building_value:
                    continue
                if str(routing_fields.get("专业") or "").strip() != specialty_value:
                    continue
                routing_matches.append(routing_record)
            if routing_matches:
                self._repair_management_prefill_put(
                    result,
                    meta_by_name,
                    "推送群组",
                    self._repair_management_unique_text(
                        [
                            (record.get("display_fields") or {}).get("通报群组")
                            for record in routing_matches
                        ]
                    ),
                    overwrite=False,
                )
                approvers: list[dict[str, Any]] = []
                for record in routing_matches:
                    approvers.extend(
                        self._repair_management_users(
                            (record.get("raw_fields") or {}).get("维修审核人")
                        )
                    )
                self._repair_management_prefill_put(
                    result,
                    meta_by_name,
                    "维修审批人",
                    approvers,
                    overwrite=False,
                )

        self._repair_management_prefill_put(
            result,
            meta_by_name,
            "当前日期",
            dt.datetime.now().strftime("%Y/%m/%d"),
            overwrite=True,
        )
        start_ms = self._repair_management_datetime_ms(result.get("维修开始时间"))
        reason = str(result.get("故障维修原因") or "").strip()
        building = str(result.get("所属数据中心/楼栋-使用") or "").strip()
        if start_ms and reason and building:
            start_text = dt.datetime.fromtimestamp(start_ms / 1000).strftime("%Y-%m-%d %H:%M")
            self._repair_management_prefill_put(
                result,
                meta_by_name,
                "维修名称",
                f"华东一{building}—{start_text}—{reason}",
                overwrite=not bool(selected_followups),
            )
        coerced, coerce_warnings = self._coerce_repair_management_fields(result, meta_by_name)
        warnings.extend(coerce_warnings)
        return {
            "fields": coerced,
            "warnings": list(dict.fromkeys(warnings)),
            "event": event,
            "repair_record_ids": requested_repair_ids,
            "followup_record_ids": requested_followup_ids,
            "repair_records": [
                {
                    "record_id": str(item.get("record_id") or ""),
                    "title": str((item.get("display_fields") or {}).get("名称（标题）") or (item.get("display_fields") or {}).get("检修概述") or "未命名检修"),
                }
                for item in selected_repairs
            ],
            "followup_records": [
                {
                    "record_id": str(item.get("record_id") or ""),
                    "title": str(
                        (item.get("display_fields") or {}).get("维修简述")
                        or (item.get("display_fields") or {}).get("维修名称")
                        or "未命名跟进"
                    ),
                }
                for item in selected_followups
            ],
        }

    def repair_management_combined_prefill(
        self,
        *,
        scope: str = "ALL",
        event_record_id: str = "",
        repair_record_ids: list[str] | tuple[str, ...] | None = None,
        month: str | None = None,
    ) -> dict[str, Any]:
        _metas, meta_by_name, _records = (
            self._load_repair_management_project_records()
        )
        return self._build_repair_management_prefill(
            scope=scope,
            event_record_id=event_record_id,
            repair_record_ids=repair_record_ids,
            followup_record_ids=[],
            month=month,
            meta_by_name=meta_by_name,
        )

    def _repair_management_record_in_scope(
        self,
        record: dict[str, Any],
        scope: str,
    ) -> bool:
        normalized_scope = self._normalize_scope(scope)
        if normalized_scope == "ALL":
            return True
        return self._scope_matches_buildings(
            normalized_scope,
            self._repair_record_building_codes(record),
        )

    def _repair_management_fields_in_scope(
        self,
        fields: dict[str, Any],
        scope: str,
    ) -> bool:
        normalized_scope = self._normalize_scope(scope)
        if normalized_scope == "ALL":
            return True
        codes = self._repair_record_building_codes({"display_fields": fields})
        return not codes or self._scope_matches_buildings(normalized_scope, codes)

    def _load_repair_management_project_records(
        self,
        *,
        force_refresh: bool = False,
    ) -> tuple[list[FieldMeta], dict[str, FieldMeta], list[dict[str, Any]]]:
        def load_remote() -> tuple[list[FieldMeta], list[dict[str, Any]]]:
            metas, meta_by_name = self._load_table_fields(
                app_token=REPAIR_SOURCE_APP_TOKEN,
                table_id=REPAIR_MANAGEMENT_TABLE_ID,
            )
            records = self._load_table_records(
                app_token=REPAIR_SOURCE_APP_TOKEN,
                table_id=REPAIR_MANAGEMENT_TABLE_ID,
                meta_by_name=meta_by_name,
                work_type=WORK_TYPE_REPAIR,
                notice_type=NOTICE_TYPE_REPAIR,
            )
            return metas, records

        return self._load_repair_snapshot_source(
            source_key=REPAIR_SNAPSHOT_SOURCE_PROJECTS,
            app_token=REPAIR_SOURCE_APP_TOKEN,
            table_id=REPAIR_MANAGEMENT_TABLE_ID,
            loader=load_remote,
            force_refresh=force_refresh,
        )

    def _ensure_repair_management_record_in_scope(
        self,
        record_id: str,
        scope: str,
        *,
        meta_by_name: dict[str, FieldMeta] | None = None,
    ) -> dict[str, Any]:
        normalized_scope = self._normalize_scope(scope)
        record_id = str(record_id or "").strip()
        if not record_id:
            raise PortalError("缺少维修项目 ID。")
        _metas, loaded_meta_by_name, records = (
            self._load_repair_management_project_records()
        )
        if meta_by_name is None:
            meta_by_name = loaded_meta_by_name
        for record in records:
            if str(record.get("record_id") or "").strip() != record_id:
                continue
            if normalized_scope != "ALL" and not self._repair_management_record_in_scope(record, normalized_scope):
                raise PortalError("当前账号无权操作该楼栋维修项目。")
            return record
        raise PortalError("未找到维修项目，可能已被删除或当前账号无权访问。")

    def repair_management_notice_prefill(
        self,
        record_id: str,
        *,
        scope: str = "ALL",
    ) -> dict[str, Any]:
        record = self._ensure_repair_management_record_in_scope(record_id, scope)
        fields = (
            record.get("display_fields")
            if isinstance(record.get("display_fields"), dict)
            else {}
        )
        raw_fields = (
            record.get("raw_fields")
            if isinstance(record.get("raw_fields"), dict)
            else {}
        )
        target_record_id = self._repair_target_record_id(record)
        if not target_record_id:
            related_ids = self._repair_management_record_ids(
                raw_fields.get("设备检修关联")
            )[:1]
            target_record_id = related_ids[0] if related_ids else ""

        fault_time = self._format_source_datetime(
            self._repair_first_field(
                fields,
                "故障发生时间",
                "发现故障时间",
                "维修开始时间",
            )
        )
        expected_time = self._format_source_datetime(
            self._repair_first_field(
                fields,
                "期望完成时间",
                "维修结束时间",
                "维修结束时间（2026）",
            )
        )
        spare_parts = self._repair_first_field(
            fields,
            "备件更换情况",
            "备件使用情况",
            "更换备件名称",
        )
        spare_count = self._repair_first_field(fields, "更换备件数量")
        if spare_parts and spare_count and spare_count not in spare_parts:
            spare_parts = f"{spare_parts} × {spare_count}"

        draft = {
            "work_type": WORK_TYPE_REPAIR,
            "notice_type": NOTICE_TYPE_REPAIR,
            "title": self._repair_title(record),
            "location": self._repair_location_text(fields),
            "level": self._repair_first_field(fields, "紧急程度")
            or self._repair_level(fields),
            "specialty": self._repair_specialty(record),
            # The lite workbench keeps repair expected/fault times in
            # start_time/end_time respectively for its existing form contract.
            "start_time": expected_time,
            "end_time": fault_time,
            "content": self._repair_first_field(
                fields, "维修简述", "事件描述", "维修名称"
            ),
            "repair_device": self._repair_device_text(fields),
            "repair_fault": self._repair_first_field(
                fields,
                "维修故障",
                "故障维修原因",
                "故障发生现象描述",
                "事件描述",
            ),
            "fault_type": self._repair_first_field(fields, "故障类型")
            or "设备故障",
            "repair_mode": self._repair_first_field(
                fields, "维修方式", "维修方", "供应商名称"
            ),
            "impact": self._repair_first_field(fields, "影响范围", "影响"),
            "discovery": self._repair_first_field(
                fields, "故障发现方式", "对应来源", "事件发现来源"
            ),
            "symptom": self._repair_first_field(
                fields, "故障发生现象描述", "故障现象", "事件描述"
            ),
            "reason": self._repair_first_field(
                fields, "故障维修原因", "故障原因"
            ),
            "solution": self._repair_first_field(
                fields, "解决方案", "维修方案", "后续整改措施"
            ),
            "spare_parts": spare_parts,
            "progress": self._repair_first_field(
                fields, "维修进展描述", "当前维修进度", "完成情况", "进度"
            ),
        }
        return {
            "repair_management_record_id": str(record_id or "").strip(),
            "source_record_id": str(record_id or "").strip(),
            "target_record_id": target_record_id,
            "action": "update" if target_record_id else "start",
            "draft": {
                key: str(value or "").strip()
                for key, value in draft.items()
                if str(value or "").strip()
            },
        }

    @classmethod
    def _repair_followup_progress_percent(cls, value: Any) -> float | None:
        if value in (None, "", [], {}):
            return None
        text = cls._repair_management_plain_text(value).strip().replace("％", "%")
        if not text:
            return None
        match = re.search(r"-?\d+(?:\.\d+)?", text)
        if not match:
            return None
        try:
            number = float(match.group(0))
        except (TypeError, ValueError):
            return None
        if "%" not in text and abs(number) <= 1:
            number *= 100
        return max(0.0, min(100.0, number))

    def _load_repair_management_status_sources(
        self,
        *,
        force_refresh: bool = False,
    ) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
        request_started_at = time.monotonic()
        now = request_started_at
        with self._repair_management_status_cache_lock:
            cached = self._repair_management_status_cache
            if (
                not force_refresh
                and isinstance(cached, dict)
                and now - float(cached.get("loaded_at") or 0) <= 30.0
            ):
                return (
                    list(cached.get("projects") or []),
                    list(cached.get("followups") or []),
                )

        with self._repair_management_status_load_lock:
            now = time.monotonic()
            with self._repair_management_status_cache_lock:
                cached = self._repair_management_status_cache
                cached_loaded_at = (
                    float(cached.get("loaded_at") or 0)
                    if isinstance(cached, dict)
                    else 0.0
                )
                if (
                    isinstance(cached, dict)
                    and (
                        (not force_refresh and now - cached_loaded_at <= 30.0)
                        or (force_refresh and cached_loaded_at >= request_started_at)
                    )
                ):
                    return (
                        list(cached.get("projects") or []),
                        list(cached.get("followups") or []),
                    )

            _project_metas, _project_meta_by_name, projects = (
                self._load_repair_management_project_records(
                    force_refresh=force_refresh
                )
            )
            _followup_metas, _followup_meta_by_name, followups = (
                self._load_repair_followup_snapshot(
                    force_refresh=force_refresh
                )
            )
            with self._repair_management_status_cache_lock:
                self._repair_management_status_cache = {
                    "loaded_at": time.monotonic(),
                    "projects": list(projects),
                    "followups": list(followups),
                }
            return list(projects), list(followups)

    def _enrich_repair_management_status_events(
        self,
        records: list[dict[str, Any]],
        *,
        force_refresh: bool = False,
    ) -> list[str]:
        requested_ids = {
            str(item.get("source_event_id") or "").strip()
            for item in records
            if str(item.get("source_event_id") or "").strip()
        }
        if not requested_ids:
            return []
        try:
            _metas, meta_by_name, event_records = (
                self._load_repair_management_event_records(
                    force_refresh=force_refresh,
                )
            )
            event_by_id = {
                str(item.get("record_id") or "").strip(): item
                for item in event_records
                if str(item.get("record_id") or "").strip() in requested_ids
            }
            missing_ids = sorted(requested_ids - set(event_by_id))
            if missing_ids:
                app_token, table_id, _source_key = self._event_source_config()
                direct_records = self._load_table_records_by_ids(
                    app_token=app_token,
                    table_id=table_id,
                    meta_by_name=meta_by_name,
                    work_type=WORK_TYPE_EVENT,
                    notice_type=NOTICE_TYPE_EVENT,
                    record_ids=missing_ids,
                )
                for item in direct_records:
                    record_id = str(item.get("record_id") or "").strip()
                    if record_id:
                        event_by_id[record_id] = item

            for item in records:
                event_record = event_by_id.get(
                    str(item.get("source_event_id") or "").strip()
                )
                if not event_record:
                    continue
                event = self._repair_management_event_item(event_record)
                item["title"] = str(
                    event.get("alarm_desc") or event.get("title") or item.get("title") or ""
                ).strip()
                item["event_sent_time"] = str(
                    event.get("sent_time")
                    or event.get("occurrence_time")
                    or item.get("event_sent_time")
                    or ""
                ).strip()
                item["event_title"] = str(event.get("title") or "").strip()
            return []
        except Exception as exc:
            return [f"关联事件信息暂未刷新：{exc}"]

    def get_repair_management_status(
        self,
        *,
        scope: str = "ALL",
        query: str = "",
        state: str = "all",
        period: str = "all",
        limit: int = 100,
        offset: int = 0,
        force_refresh: bool = False,
    ) -> dict[str, Any]:
        projects, followups = self._load_repair_management_status_sources(
            force_refresh=force_refresh,
        )
        followups_by_project: dict[str, list[dict[str, Any]]] = {}
        for followup in followups:
            for project_id in self._repair_followup_parent_ids(followup):
                followups_by_project.setdefault(project_id, []).append(followup)

        normalized_scope = self._normalize_scope(scope)
        local_today = dt.datetime.now().astimezone().date()
        local_week_start = local_today - dt.timedelta(days=local_today.weekday())
        all_status_records: list[dict[str, Any]] = []
        for project in projects:
            workflow_completed = self._repair_management_is_completed(project)
            if not self._repair_management_record_in_scope(project, normalized_scope):
                continue
            project_id = str(project.get("record_id") or "").strip()
            if not project_id:
                continue
            linked = followups_by_project.get(project_id, [])
            latest_followup = (
                max(linked, key=self._repair_followup_order_key)
                if linked
                else {}
            )
            progress_values: list[float | None] = []
            for followup in linked:
                display_fields = (
                    followup.get("display_fields")
                    if isinstance(followup.get("display_fields"), dict)
                    else {}
                )
                raw_fields = (
                    followup.get("raw_fields")
                    if isinstance(followup.get("raw_fields"), dict)
                    else {}
                )
                progress_values.append(
                    self._repair_followup_progress_percent(
                        display_fields.get("维修进度")
                        if display_fields.get("维修进度") not in (None, "", [], {})
                        else raw_fields.get("维修进度")
                    )
                )

            latest_display_fields = (
                latest_followup.get("display_fields")
                if isinstance(latest_followup.get("display_fields"), dict)
                else {}
            )
            latest_raw_fields = (
                latest_followup.get("raw_fields")
                if isinstance(latest_followup.get("raw_fields"), dict)
                else {}
            )
            latest_progress_value = latest_raw_fields.get("维修进度")
            if latest_progress_value in (None, "", [], {}):
                latest_progress_value = latest_display_fields.get("维修进度")
            latest_progress_percent = self._repair_followup_progress_percent(
                latest_progress_value
            )
            progress_completed = bool(linked) and (
                latest_progress_percent is not None
                and latest_progress_percent >= 100
            )
            latest_fields = (
                latest_followup.get("display_fields")
                if isinstance(latest_followup.get("display_fields"), dict)
                else {}
            )
            project_fields = (
                project.get("display_fields")
                if isinstance(project.get("display_fields"), dict)
                else {}
            )
            project_raw_fields = (
                project.get("raw_fields")
                if isinstance(project.get("raw_fields"), dict)
                else {}
            )
            completed_at_ms: int | None = None
            if progress_completed or workflow_completed:
                completion_values = [
                    latest_display_fields.get("维修结束时间"),
                    latest_raw_fields.get("维修结束时间"),
                    project_fields.get("维修结束时间"),
                    project_raw_fields.get("维修结束时间"),
                    project_fields.get("维修结束时间（2026）"),
                    project_raw_fields.get("维修结束时间（2026）"),
                ]
                if progress_completed:
                    completion_values.extend(
                        [
                            latest_display_fields.get("创建时间"),
                            latest_raw_fields.get("创建时间"),
                            latest_followup.get("created_time"),
                        ]
                    )
                for value in completion_values:
                    completed_at_ms = self._repair_management_datetime_ms(value)
                    if completed_at_ms is not None:
                        break
            is_completed = bool(progress_completed or workflow_completed)
            building_codes = self._repair_record_building_codes(project)
            completed_count = sum(
                1 for value in progress_values if value is not None and value >= 100
            )
            progress_percent = round(latest_progress_percent or 0)
            item_state = (
                "completed"
                if is_completed
                else "in_progress" if linked else "without_followup"
            )
            source_event_ids = self._repair_management_record_ids(
                project_raw_fields.get("关联事件单")
            )
            fallback_title = self._repair_management_plain_text(
                project_fields.get("事件描述")
                or project_fields.get("故障发生现象描述")
                or self._repair_management_title(project)
            )
            all_status_records.append(
                {
                    "record_id": project_id,
                    "title": fallback_title,
                    "repair_title": self._repair_management_title(project),
                    "source_event_id": source_event_ids[0] if source_event_ids else "",
                    "event_sent_time": self._repair_management_plain_text(
                        project_fields.get("事件进展响应时间")
                        or project_fields.get("故障发生时间")
                    ),
                    "building": self._building_label_from_codes(building_codes),
                    "building_codes": building_codes,
                    "specialty": self._repair_management_plain_text(
                        project_fields.get("所属专业")
                        or project_fields.get("专业（推送消息用）")
                    ),
                    "fault_time": self._repair_management_plain_text(
                        project_fields.get("故障发生时间")
                    ),
                    "location": self._repair_management_plain_text(
                        project_fields.get("位置")
                        or project_fields.get("地点")
                        or project_fields.get("维修位置")
                    ),
                    "fault_reason": self._repair_management_plain_text(
                        project_fields.get("故障维修原因")
                    ),
                    "state": item_state,
                    "status_label": (
                        "历史已完成"
                        if is_completed
                        else "维修进行中" if linked else "待首次跟进"
                    ),
                    "followup_count": len(linked),
                    "completed_followup_count": completed_count,
                    "progress_percent": progress_percent,
                    "completed_at": completed_at_ms,
                    "latest_followup_time": self._repair_management_plain_text(
                        latest_fields.get("创建时间")
                        or latest_followup.get("last_modified_time")
                        or latest_followup.get("created_time")
                    ),
                    "latest_followup": self._repair_management_plain_text(
                        latest_fields.get("维修进展描述")
                        or latest_fields.get("跟进项（如有）")
                    ),
                }
            )

        active_records = [
            item for item in all_status_records if item.get("state") != "completed"
        ]
        completed_records = [
            item for item in all_status_records if item.get("state") == "completed"
        ]
        normalized_state = str(state or "all").strip().lower()
        normalized_period = str(period or "all").strip().lower()
        if normalized_state == "completed":
            filtered_records = list(completed_records)
        elif normalized_state in {"without_followup", "in_progress"}:
            filtered_records = [
                item for item in active_records if item.get("state") == normalized_state
            ]
        else:
            filtered_records = list(active_records)

        def completion_date(item: dict[str, Any]) -> dt.date | None:
            completed_at = item.get("completed_at")
            if completed_at in (None, ""):
                return None
            try:
                return dt.datetime.fromtimestamp(
                    int(completed_at) / 1000
                ).astimezone().date()
            except (OSError, OverflowError, TypeError, ValueError):
                return None

        def in_completion_period(item: dict[str, Any], period_name: str) -> bool:
            completed_date = completion_date(item)
            if completed_date is None:
                return period_name == "all"
            if period_name == "today":
                return completed_date == local_today
            if period_name == "week":
                return local_week_start <= completed_date <= local_today
            if period_name == "month":
                return (
                    completed_date.year == local_today.year
                    and completed_date.month == local_today.month
                )
            return True

        if normalized_state == "completed" and normalized_period in {
            "today",
            "week",
            "month",
        }:
            filtered_records = [
                item
                for item in filtered_records
                if in_completion_period(item, normalized_period)
            ]

        query_text = str(query or "").strip().casefold()
        enrichment_warnings: list[str] = []
        if query_text:
            enrichment_warnings.extend(
                self._enrich_repair_management_status_events(
                    filtered_records,
                    force_refresh=force_refresh,
                )
            )
            filtered_records = [
                item
                for item in filtered_records
                if query_text
                in "\n".join(
                    str(item.get(name) or "")
                    for name in (
                        "title",
                        "building",
                        "specialty",
                        "fault_reason",
                        "location",
                        "latest_followup",
                    )
                ).casefold()
            ]
        if normalized_state == "completed":
            filtered_records.sort(
                key=lambda item: (
                    int(item.get("completed_at") or 0),
                    str(item.get("latest_followup_time") or ""),
                ),
                reverse=True,
            )
        else:
            filtered_records.sort(
                key=lambda item: (
                    {
                        "without_followup": 0,
                        "in_progress": 1,
                    }.get(str(item.get("state") or ""), 2),
                    int(item.get("progress_percent") or 0),
                    str(item.get("latest_followup_time") or ""),
                )
            )
        max_limit = max(1, min(int(limit or 100), 200))
        page_offset = max(0, int(offset or 0))
        page_records = [
            dict(item)
            for item in filtered_records[page_offset : page_offset + max_limit]
        ]
        enrichment_warnings.extend(
            self._enrich_repair_management_status_events(
                page_records,
                force_refresh=force_refresh,
            )
        )
        warnings = list(dict.fromkeys(enrichment_warnings))
        progress_records = [
            item for item in active_records if item.get("state") == "in_progress"
        ]
        return {
            "scope": normalized_scope,
            "state": normalized_state,
            "period": normalized_period,
            "records": page_records,
            "total": len(filtered_records),
            "returned": len(page_records),
            "offset": page_offset,
            "has_more": page_offset + len(page_records) < len(filtered_records),
            "warnings": warnings,
            "stats": {
                "total": len(active_records),
                "without_followup": sum(
                    1 for item in active_records if item.get("state") == "without_followup"
                ),
                "in_progress": len(progress_records),
                "completed_total": len(completed_records),
                "completed_month": sum(
                    1 for item in completed_records if in_completion_period(item, "month")
                ),
                "completed_week": sum(
                    1 for item in completed_records if in_completion_period(item, "week")
                ),
                "completed_today": sum(
                    1 for item in completed_records if in_completion_period(item, "today")
                ),
                "average_progress": round(
                    sum(int(item.get("progress_percent") or 0) for item in progress_records)
                    / len(progress_records)
                ) if progress_records else 0,
            },
        }

    def _repair_management_record_payload(
        self,
        item: dict[str, Any],
    ) -> dict[str, Any]:
        raw_fields = (
            item.get("raw_fields")
            if isinstance(item.get("raw_fields"), dict)
            else {}
        )
        source_event_ids = self._repair_management_record_ids(
            raw_fields.get("关联事件单")
        )
        source_repair_ids = self._repair_management_record_ids(
            raw_fields.get("设备检修关联")
        )
        return {
            "record_id": str(item.get("record_id") or ""),
            "title": self._repair_management_title(item),
            "created_time": item.get("created_time") or "",
            "last_modified_time": item.get("last_modified_time") or "",
            "building_codes": self._repair_record_building_codes(item),
            "workflow": self._repair_management_workflow_text(item),
            "display_fields": item.get("display_fields") or {},
            "raw_fields": raw_fields,
            "source_event_id": source_event_ids[0] if source_event_ids else "",
            "source_repair_ids": source_repair_ids[:1],
        }

    def get_repair_management_record(
        self,
        record_id: str,
        *,
        scope: str = "ALL",
        force_refresh: bool = False,
    ) -> dict[str, Any]:
        record_id = str(record_id or "").strip()
        if not record_id:
            raise PortalError("缺少维修项目 ID。")
        metas, meta_by_name, records = self._load_repair_management_project_records(
            force_refresh=force_refresh
        )
        record = next(
            (
                item
                for item in records
                if str(item.get("record_id") or "").strip() == record_id
            ),
            None,
        )
        if record is None and self._repair_snapshots_enabled and not force_refresh:
            metas, meta_by_name, records = self._load_repair_management_project_records(
                force_refresh=True
            )
            record = next(
                (
                    item
                    for item in records
                    if str(item.get("record_id") or "").strip() == record_id
                ),
                None,
            )
        if record is None:
            raise PortalError("维修项目已不存在，请刷新列表。")
        if not self._repair_management_record_in_scope(record, scope):
            raise PortalError("当前账号无权查看该楼栋维修项目。")
        return {
            "record": self._repair_management_record_payload(record),
            "fields": [self._repair_management_field_payload(meta) for meta in metas],
            "schema_warnings": self._repair_management_schema_warnings(meta_by_name),
        }

    def get_repair_management_records(
        self,
        *,
        scope: str = "ALL",
        query: str = "",
        limit: int = 200,
        offset: int = 0,
        focus_record_id: str = "",
        force_refresh: bool = False,
    ) -> dict[str, Any]:
        metas, meta_by_name, records = self._load_repair_management_project_records(
            force_refresh=force_refresh
        )
        records = [
            item
            for item in records
            if not self._repair_management_is_completed(item)
            and self._repair_management_record_in_scope(item, scope)
        ]
        query_text = str(query or "").strip().lower()
        if query_text:
            records = [
                item
                for item in records
                if query_text
                in "\n".join(
                    [
                        str(item.get("record_id") or ""),
                        *[
                            self._repair_management_plain_text(value)
                            for value in (
                                (item.get("display_fields") or {}).values()
                                if isinstance(item.get("display_fields"), dict)
                                else []
                            )
                        ],
                    ]
                ).lower()
            ]
        records.sort(
            key=lambda item: str(
                item.get("last_modified_time") or item.get("created_time") or ""
            ),
            reverse=True,
        )
        max_limit = max(1, min(int(limit or 200), 500))
        page_offset = max(0, int(offset or 0))
        focus_id = str(focus_record_id or "").strip()
        if focus_id:
            focus_index = next(
                (
                    index
                    for index, item in enumerate(records)
                    if str(item.get("record_id") or "").strip() == focus_id
                ),
                -1,
            )
            if focus_index >= 0:
                page_offset = (focus_index // max_limit) * max_limit
        payload_records = []
        for item in records[page_offset : page_offset + max_limit]:
            payload_records.append(self._repair_management_record_payload(item))
        return {
            "app_token": REPAIR_SOURCE_APP_TOKEN,
            "table_id": REPAIR_MANAGEMENT_TABLE_ID,
            "fields": [self._repair_management_field_payload(meta) for meta in metas],
            "schema_warnings": self._repair_management_schema_warnings(meta_by_name),
            "records": payload_records,
            "total": len(records),
            "returned": len(payload_records),
            "offset": page_offset,
            "has_more": page_offset + len(payload_records) < len(records),
        }

    def create_repair_management_record(
        self,
        fields: dict[str, Any],
        *,
        operation_id: str = "",
        source_event_id: str = "",
        source_repair_ids: list[str] | tuple[str, ...] | None = None,
        source_month: str | None = None,
        scope: str = "ALL",
    ) -> dict[str, Any]:
        metas, meta_by_name, _records = (
            self._load_repair_management_project_records()
        )
        cleaned = self._clean_repair_management_fields(fields, meta_by_name, allow_empty=True)
        auto = self._build_repair_management_prefill(
            scope=scope,
            event_record_id=source_event_id,
            repair_record_ids=source_repair_ids,
            followup_record_ids=[],
            month=source_month,
            meta_by_name=meta_by_name,
        )
        merged = {**(auto.get("fields") or {}), **cleaned}
        merged = self._apply_repair_management_event_link(
            merged, meta_by_name, source_event_id=source_event_id
        )
        merged = self._apply_repair_management_repair_links(
            merged, meta_by_name, source_repair_ids
        )
        merged = self._apply_repair_management_form_derivatives(
            merged,
            meta_by_name,
        )
        prepared, write_warnings = self._coerce_repair_management_fields(merged, meta_by_name)
        if not prepared:
            raise PortalError("没有可写入的维修汇总字段。")
        if not self._repair_management_fields_in_scope(prepared, scope):
            raise PortalError("当前账号无权创建该楼栋检修记录。")
        missing = self._missing_repair_management_required_fields(prepared, meta_by_name)
        if missing:
            raise PortalError("请先填写检修单必填字段：" + "、".join(missing) + "。")
        stable_operation_id = str(operation_id or "").strip()
        operation: dict[str, Any] | None = None
        if stable_operation_id:
            operation_payload = {
                "scope": self._normalize_scope(scope),
                "source_event_id": str(source_event_id or "").strip(),
                "source_repair_ids": list(source_repair_ids or []),
                "source_month": str(source_month or "").strip(),
                "fields": prepared,
            }
            payload_hash = hashlib.sha256(
                json.dumps(
                    operation_payload,
                    ensure_ascii=False,
                    sort_keys=True,
                    separators=(",", ":"),
                    default=str,
                ).encode("utf-8")
            ).hexdigest()
            try:
                operation = self._state_store.begin_repair_management_operation(
                    stable_operation_id,
                    operation_type="project_create",
                    scope=self._normalize_scope(scope),
                    payload_hash=payload_hash,
                )
            except ValueError as exc:
                raise PortalError(f"维修项目操作标识冲突：{exc}") from exc
            if not operation.get("created"):
                existing_result = dict(operation.get("result") or {})
                existing_record_id = str(
                    operation.get("record_id")
                    or existing_result.get("record_id")
                    or ""
                ).strip()
                if existing_record_id:
                    existing_result.setdefault("record_id", existing_record_id)
                    existing_result["idempotent_replay"] = True
                    existing_result["warnings"] = [
                        warning
                        for warning in (existing_result.get("warnings") or [])
                        if "事件转检修状态暂未同步" not in str(warning)
                    ]
                    event_warning = ""
                    if str(source_event_id or "").strip():
                        try:
                            self.mark_event_transferred_to_repair(
                                record_id=str(source_event_id or "").strip(),
                                month=str(source_month or ""),
                                refresh_snapshot=False,
                            )
                        except Exception as exc:
                            event_warning = f"维修项目已创建，事件转检修状态暂未同步：{exc}"
                    if event_warning:
                        existing_result["warnings"] = list(
                            dict.fromkeys(
                                [*(existing_result.get("warnings") or []), event_warning]
                            )
                        )
                        existing_result["event_transfer_sync_pending"] = True
                        status = "sync_pending"
                    else:
                        existing_result["event_transfer_sync_pending"] = False
                        status = "completed"
                    self._state_store.update_repair_management_operation(
                        stable_operation_id,
                        status=status,
                        result=existing_result,
                        error=event_warning,
                    )
                    return existing_result
                if operation.get("status") == "started":
                    raise PortalError("该维修项目正在后台创建，请稍后刷新，勿重复提交。")
                raise PortalError(
                    "上一次维修项目创建结果未确认。为避免重复记录，请刷新项目列表后再操作。"
                )

        try:
            result = self._create_record_fields(
                app_token=REPAIR_SOURCE_APP_TOKEN,
                table_id=REPAIR_MANAGEMENT_TABLE_ID,
                fields=prepared,
            )
            record_id = str(
                (((result.get("data") or {}).get("record") or {}).get("record_id"))
                or ((result.get("data") or {}).get("record_id"))
                or ""
            ).strip()
            if not record_id:
                raise PortalError("检修记录已提交，但未返回记录 ID。")
            self._upsert_repair_snapshot_fields(
                source_key=REPAIR_SNAPSHOT_SOURCE_PROJECTS,
                record_id=record_id,
                fields=prepared,
            )
        except Exception as exc:
            if stable_operation_id:
                self._state_store.update_repair_management_operation(
                    stable_operation_id,
                    status="failed",
                    error=str(exc),
                )
            raise

        response = {
            "record_id": record_id,
            "fields": prepared,
            "field_count": len(prepared),
            "warnings": list(dict.fromkeys([*(auto.get("warnings") or []), *write_warnings])),
            "editable_fields": [self._repair_management_field_payload(meta) for meta in metas],
            "event_transfer_sync_pending": False,
        }
        if stable_operation_id:
            self._state_store.update_repair_management_operation(
                stable_operation_id,
                status="remote_written",
                record_id=record_id,
                result=response,
                error="",
            )
        if str(source_event_id or "").strip():
            try:
                self.mark_event_transferred_to_repair(
                    record_id=str(source_event_id or "").strip(),
                    month=str(source_month or ""),
                    refresh_snapshot=False,
                )
            except Exception as exc:
                response["event_transfer_sync_pending"] = True
                response["warnings"] = list(
                    dict.fromkeys(
                        [
                            *response["warnings"],
                            f"维修项目已创建，事件转检修状态暂未同步：{exc}",
                        ]
                    )
                )
        self._invalidate_repair_management_status_cache()
        if stable_operation_id:
            self._state_store.update_repair_management_operation(
                stable_operation_id,
                status=(
                    "sync_pending"
                    if response["event_transfer_sync_pending"]
                    else "completed"
                ),
                result=response,
                error=(
                    response["warnings"][-1]
                    if response["event_transfer_sync_pending"] and response["warnings"]
                    else ""
                ),
            )
        return response

    def update_repair_management_record(
        self,
        record_id: str,
        fields: dict[str, Any],
        *,
        source_event_id: str = "",
        source_repair_ids: list[str] | tuple[str, ...] | None = None,
        replace_source_relations: bool = False,
        source_month: str | None = None,
        scope: str = "ALL",
    ) -> dict[str, Any]:
        _metas, meta_by_name, _records = (
            self._load_repair_management_project_records()
        )
        existing = self._ensure_repair_management_record_in_scope(
            record_id,
            scope,
            meta_by_name=meta_by_name,
        )
        cleaned = self._clean_repair_management_fields(fields, meta_by_name, allow_empty=True)
        existing_raw = existing.get("raw_fields") if isinstance((existing or {}).get("raw_fields"), dict) else {}
        summary_id = str(record_id or "").strip()
        effective_event_id = str(source_event_id or "").strip()
        if not replace_source_relations and not effective_event_id:
            existing_event_ids = self._repair_management_record_ids(existing_raw.get("关联事件单"))
            effective_event_id = existing_event_ids[0] if existing_event_ids else ""
        effective_repair_ids = list(source_repair_ids or [])
        if not replace_source_relations and not effective_repair_ids:
            effective_repair_ids = self._repair_management_record_ids(
                existing_raw.get("设备检修关联")
            )[:1]
        _followup_metas, _followup_meta_by_name, linked_followups = (
            self._load_repair_followups_for_summary(summary_id, limit=500)
        )
        effective_followup_ids = [
            str(item.get("record_id") or "").strip()
            for item in linked_followups
            if str(item.get("record_id") or "").strip()
        ]
        auto: dict[str, Any] = {"fields": {}, "warnings": []}
        if effective_event_id or effective_repair_ids or effective_followup_ids:
            auto = self._build_repair_management_prefill(
                scope=scope,
                event_record_id=effective_event_id,
                repair_record_ids=effective_repair_ids,
                followup_record_ids=effective_followup_ids,
                month=source_month,
                meta_by_name=meta_by_name,
                allow_multiple_followups=True,
            )
        auto_fields = auto.get("fields") if isinstance(auto.get("fields"), dict) else {}
        if effective_event_id:
            for field_name in REPAIR_MANAGEMENT_EVENT_AUTO_FIELD_NAMES:
                if field_name in meta_by_name:
                    auto_fields.setdefault(field_name, None)
        if effective_repair_ids:
            for field_name in REPAIR_MANAGEMENT_REPAIR_AUTO_FIELD_NAMES:
                if field_name in meta_by_name:
                    auto_fields.setdefault(field_name, None)
        if effective_followup_ids:
            for field_name in REPAIR_MANAGEMENT_FOLLOWUP_AUTO_FIELD_NAMES:
                if field_name in meta_by_name:
                    auto_fields.setdefault(field_name, None)
        auto["fields"] = auto_fields
        merged = {**(auto.get("fields") or {}), **cleaned}
        merged = self._apply_repair_management_event_link(
            merged, meta_by_name, source_event_id=effective_event_id
        )
        merged = self._apply_repair_management_repair_links(
            merged, meta_by_name, effective_repair_ids
        )
        merged = self._apply_repair_management_form_derivatives(
            merged,
            meta_by_name,
        )
        if replace_source_relations and not effective_event_id:
            field_name = REPAIR_MANAGEMENT_EVENT_LINK_FIELD_NAMES[0]
            meta = meta_by_name.get(field_name)
            if meta is not None:
                merged[field_name] = (
                    []
                    if int(meta.field_type or 0) in {18, 21}
                    or "link" in str(meta.ui_type or "").lower()
                    else None
                )
        if replace_source_relations and not effective_repair_ids:
            field_name = REPAIR_MANAGEMENT_REPAIR_LINK_FIELD_NAMES[0]
            meta = meta_by_name.get(field_name)
            if meta is not None:
                merged[field_name] = (
                    []
                    if int(meta.field_type or 0) in {18, 21}
                    or "link" in str(meta.ui_type or "").lower()
                    else None
                )
        prepared, write_warnings = self._coerce_repair_management_fields(merged, meta_by_name)
        if not prepared:
            raise PortalError("没有可写入的维修汇总字段。")
        effective_fields = {
            **(
                existing.get("display_fields")
                if isinstance(existing.get("display_fields"), dict)
                else {}
            ),
            **existing_raw,
            **prepared,
        }
        missing = self._missing_repair_management_required_fields(
            effective_fields,
            meta_by_name,
        )
        if missing:
            raise PortalError("请先填写检修单必填字段：" + "、".join(missing) + "。")
        if not self._repair_management_fields_in_scope(prepared, scope):
            raise PortalError("当前账号无权把检修记录改到该楼栋。")
        self._patch_record_fields(
            app_token=REPAIR_SOURCE_APP_TOKEN,
            table_id=REPAIR_MANAGEMENT_TABLE_ID,
            record_id=summary_id,
            fields=prepared,
        )
        self._upsert_repair_snapshot_fields(
            source_key=REPAIR_SNAPSHOT_SOURCE_PROJECTS,
            record_id=summary_id,
            fields=prepared,
        )
        self._invalidate_repair_management_status_cache()
        return {
            "record_id": summary_id,
            "fields": prepared,
            "field_count": len(prepared),
            "warnings": list(dict.fromkeys([*(auto.get("warnings") or []), *write_warnings])),
        }

    def delete_repair_management_record(
        self,
        record_id: str,
        *,
        scope: str = "ALL",
    ) -> dict[str, Any]:
        _metas, meta_by_name, _records = (
            self._load_repair_management_project_records()
        )
        self._ensure_repair_management_record_in_scope(
            record_id,
            scope,
            meta_by_name=meta_by_name,
        )
        summary_id = str(record_id or "").strip()
        _followup_metas, _followup_meta_by_name, linked_followups = (
            self._load_repair_followups_for_summary(summary_id, limit=500)
        )
        followup_ids = [
            str(item.get("record_id") or "").strip()
            for item in linked_followups
            if str(item.get("record_id") or "").strip()
        ]
        if followup_ids:
            raise PortalError("该检修单仍有维修跟进记录，请先删除跟进记录。")
        self._delete_record_fields(
            app_token=REPAIR_SOURCE_APP_TOKEN,
            table_id=REPAIR_MANAGEMENT_TABLE_ID,
            record_id=summary_id,
        )
        self._delete_repair_snapshot_item(
            REPAIR_SNAPSHOT_SOURCE_PROJECTS,
            summary_id,
        )
        self._invalidate_repair_management_status_cache()
        return {"record_id": summary_id, "deleted": True}

    def mark_event_transferred_to_repair(
        self,
        *,
        record_id: str,
        month: str = "",
        refresh_snapshot: bool = True,
    ) -> dict[str, Any]:
        record_id = str(record_id or "").strip()
        if not record_id:
            raise PortalError("缺少事件记录 ID。")
        app_token, table_id, _source_key = self._event_source_config()
        field_name = EVENT_NOTICE_FIELDS.get("transfer_to_overhaul", "是否转检修")
        self._patch_record_fields(
            app_token=app_token,
            table_id=table_id,
            record_id=record_id,
            fields={field_name: True},
        )
        self._upsert_repair_snapshot_fields(
            source_key=REPAIR_SNAPSHOT_SOURCE_EVENTS,
            record_id=record_id,
            fields={field_name: True},
        )
        self._invalidate_repair_management_event_cache()
        snapshot: dict[str, Any] = {}
        warning = ""
        if refresh_snapshot:
            try:
                snapshot = self.refresh_event_month_snapshot(month)
            except Exception as exc:
                warning = str(exc)
        return {
            "record_id": record_id,
            "field_name": field_name,
            "transfer_to_overhaul": True,
            "snapshot": snapshot,
            "warning": warning,
        }

    def _source_cache_ttl_seconds(self) -> int:
        try:
            ttl = int(getattr(config, "lan_template_source_cache_ttl_seconds", 0) or 0)
        except (TypeError, ValueError):
            ttl = 0
        if ttl <= 0:
            ttl = SOURCE_CACHE_TTL_SECONDS
        return ttl

    def _source_cache_expired(self) -> bool:
        ttl = self._source_cache_ttl_seconds()
        if not self._last_loaded_ts:
            return True
        return (time.time() - self._last_loaded_ts) >= ttl

    @staticmethod
    def _is_obsolete_empty_source_warning(value: Any) -> bool:
        text = str(value or "")
        normalized = text.replace("\\/", "/")
        return "/apps//tables//" in normalized or "apps//tables//" in normalized

    @staticmethod
    def _is_bitable_data_not_ready_error(value: Any) -> bool:
        text = str(value or "")
        return (
            str(BITABLE_DATA_NOT_READY_CODE) in text
            or "Data not ready" in text
            or "数据暂时未准备好" in text
            or "多维表正在计算" in text
        )

    def _source_sync_warning(self, label: str, exc: Any) -> str:
        if self._is_bitable_data_not_ready_error(exc):
            return (
                f"{label}同步失败: 飞书多维表正在计算，已保留上次成功数据，"
                "请稍后再试。"
            )
        return f"{label}同步失败: {exc}"

    def _clean_load_warnings(self, warnings: list[Any] | None = None) -> list[str]:
        source = self._load_warnings if warnings is None else warnings
        cleaned: list[str] = []
        for item in source or []:
            text = str(item or "").strip()
            if not text or self._is_obsolete_empty_source_warning(text):
                continue
            if text not in cleaned:
                cleaned.append(text)
        return cleaned[-20:]

    def _current_load_warnings(self) -> list[str]:
        cleaned = self._clean_load_warnings()
        if cleaned != list(self._load_warnings or []):
            self._load_warnings = cleaned
        return cleaned

    def _hydrate_source_records_from_sqlite(self) -> bool:
        try:
            snapshot = self._state_store.get_source_scope_snapshot("ALL")
        except Exception as exc:
            if f"SQLite源表快照读取失败: {exc}" not in self._load_warnings:
                self._load_warnings.append(f"SQLite源表快照读取失败: {exc}")
            return False
        if not snapshot.get("exists"):
            return False
        records = [
            dict(item)
            for item in (snapshot.get("records") or [])
            if isinstance(item, dict)
        ]
        zhihang_records = [
            dict(item)
            for item in (snapshot.get("zhihang_records") or [])
            if isinstance(item, dict)
        ]
        self._records = [
            item for item in records if self._record_work_type(item) == WORK_TYPE_MAINTENANCE
        ]
        self._change_records = [
            item for item in records if self._record_work_type(item) == WORK_TYPE_CHANGE
        ]
        self._repair_records = [
            item for item in records if self._record_work_type(item) == WORK_TYPE_REPAIR
        ]
        self._zhihang_change_records = zhihang_records
        self._maintenance_loaded_once = True
        self._change_loaded_once = True
        self._repair_loaded_once = True
        self._zhihang_change_loaded_once = True
        meta = snapshot.get("meta") if isinstance(snapshot.get("meta"), dict) else {}
        self._last_loaded_at = str(meta.get("last_loaded_at") or self._last_loaded_at or "")
        try:
            self._last_loaded_ts = float(snapshot.get("updated_at") or self._last_loaded_ts or 0)
        except Exception:
            pass
        warnings = meta.get("warnings") if isinstance(meta, dict) else []
        if isinstance(warnings, list):
            self._load_warnings = self._clean_load_warnings(warnings)
        return True

    def _snapshot_meta(self) -> dict[str, Any]:
        return {
            "last_loaded_at": self._last_loaded_at,
            "last_loaded_ts": self._last_loaded_ts,
            "source_cache_ttl_seconds": self._source_cache_ttl_seconds(),
            "warnings": self._current_load_warnings(),
        }

    def _save_source_scope_snapshots(self) -> None:
        meta = self._snapshot_meta()
        snapshots: dict[str, dict[str, Any]] = {}
        for option in SCOPE_OPTIONS:
            scope = self._normalize_scope(option.get("value"))
            records = self._workbench_records_from_memory(
                month=RECENT_MONTH_FILTER_LABEL,
                scope=scope,
            )
            zhihang_records = self._filter_zhihang_change_records_from_memory(
                month=RECENT_MONTH_FILTER_LABEL,
                scope=scope,
            )
            snapshots[scope] = {
                "records": records,
                "zhihang_records": zhihang_records,
            }
        self._state_store.replace_all_source_scope_snapshots(snapshots, meta=meta)

    def _source_snapshot_records(self, scope: str) -> list[dict[str, Any]] | None:
        try:
            snapshot = self._state_store.get_source_scope_snapshot(scope)
        except Exception as exc:
            warning = f"SQLite源表快照读取失败: {exc}"
            if warning not in self._load_warnings:
                self._load_warnings.append(warning)
            return None
        if not snapshot.get("exists"):
            return None
        meta = snapshot.get("meta") if isinstance(snapshot.get("meta"), dict) else {}
        if meta.get("last_loaded_at"):
            self._last_loaded_at = str(meta.get("last_loaded_at") or "")
        warnings = meta.get("warnings") if isinstance(meta, dict) else []
        if isinstance(warnings, list):
            self._load_warnings = self._clean_load_warnings(warnings)
        records = [
            dict(item)
            for item in (snapshot.get("records") or [])
            if isinstance(item, dict)
        ]
        return records

    def _source_snapshot_zhihang_records(self, scope: str) -> list[dict[str, Any]] | None:
        try:
            snapshot = self._state_store.get_source_scope_snapshot(scope)
        except Exception as exc:
            warning = f"SQLite智航源表快照读取失败: {exc}"
            if warning not in self._load_warnings:
                self._load_warnings.append(warning)
            return None
        if not snapshot.get("exists"):
            return None
        meta = snapshot.get("meta") if isinstance(snapshot.get("meta"), dict) else {}
        warnings = meta.get("warnings") if isinstance(meta, dict) else []
        if isinstance(warnings, list):
            self._load_warnings = self._clean_load_warnings(warnings)
        return [
            dict(item)
            for item in (snapshot.get("zhihang_records") or [])
            if isinstance(item, dict)
        ]

    def _source_snapshot_exists(self, scope: str = "ALL") -> bool:
        try:
            return bool(self._state_store.get_source_scope_snapshot(scope).get("exists"))
        except Exception:
            return False

    def _payload_version(
        self,
        *,
        scope: str,
        records: list[dict[str, Any]],
        ongoing_items: list[dict[str, Any]],
        daily_summary: dict[str, Any] | None = None,
    ) -> str:
        stats = (daily_summary or {}).get("stats") if isinstance(daily_summary, dict) else {}
        return ":".join(
            [
                str(self.state_cache_version()),
                str(scope or ""),
                str(self._last_loaded_at or ""),
                str(len(records or [])),
                str(len(ongoing_items or [])),
                str((stats or {}).get("ended") or 0),
            ]
        )

    def clear_target_record_cache(self) -> None:
        with self._refresh_lock:
            self._target_record_cache.clear()

    def clear_engineer_mop_cache(self) -> None:
        with self._engineer_mop_cache_lock:
            self._engineer_mop_cache = None

    def refresh(self) -> None:
        with self._refresh_lock:
            warnings: list[str] = []
            self._target_record_cache.clear()
            self._load_fields()
            self._load_records()
            try:
                self._load_change_fields()
                self._load_change_records()
            except Exception as exc:
                self._change_loaded_once = True
                if not self._change_field_meta_list:
                    self._change_field_meta_list = []
                    self._change_field_meta_by_name = {}
                if not self._change_records:
                    self._change_records = []
                warnings.append(self._source_sync_warning("变更源表", exc))
            try:
                self._load_zhihang_change_fields()
                self._load_zhihang_change_records()
            except Exception as exc:
                self._zhihang_change_loaded_once = True
                if not self._zhihang_change_field_meta_list:
                    self._zhihang_change_field_meta_list = []
                    self._zhihang_change_field_meta_by_name = {}
                if not self._zhihang_change_records:
                    self._zhihang_change_records = []
                warnings.append(self._source_sync_warning("智航变更源表", exc))
            try:
                self._load_repair_fields()
                self._load_repair_records()
            except Exception as exc:
                self._repair_loaded_once = True
                if not self._repair_field_meta_list:
                    self._repair_field_meta_list = []
                    self._repair_field_meta_by_name = {}
                if not self._repair_records:
                    self._repair_records = []
                warnings.append(self._source_sync_warning("检修源表", exc))
            now = dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            self._last_loaded_at = now
            self._last_loaded_ts = time.time()
            self._load_warnings = warnings
            if warnings:
                self._state_store.record_failed_source_snapshot(
                    meta=self._snapshot_meta(),
                    error="；".join(warnings),
                )
            else:
                self._save_source_scope_snapshots()
                self._touch_state_cache_version()

    def refresh_repair_source(self) -> dict[str, Any]:
        """Refresh only the repair source table and rewrite SQLite snapshots."""
        with self._refresh_lock:
            if not (
                self._maintenance_loaded_once
                and self._change_loaded_once
                and self._repair_loaded_once
                and self._zhihang_change_loaded_once
            ):
                self._hydrate_source_records_from_sqlite()
            warnings = [
                str(item)
                for item in (self._load_warnings or [])
                if not str(item or "").startswith("检修源表同步失败")
            ]
            try:
                self._load_repair_fields()
                self._load_repair_records()
            except Exception as exc:
                warning = self._source_sync_warning("检修源表", exc)
                if warning not in warnings:
                    warnings.append(warning)
                self._load_warnings = warnings
                raise PortalError(warning) from exc
            now = dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            self._last_loaded_at = now
            self._last_loaded_ts = time.time()
            self._load_warnings = warnings
            self._save_source_scope_snapshots()
            self._touch_state_cache_version()
            return {
                "repair_refreshed_at": now,
                "repair_count": len(self._repair_records),
            }

    def refresh_change_source(self) -> dict[str, Any]:
        """Refresh only change-related source tables and rewrite SQLite snapshots."""
        with self._refresh_lock:
            if not (
                self._maintenance_loaded_once
                and self._change_loaded_once
                and self._repair_loaded_once
                and self._zhihang_change_loaded_once
            ):
                self._hydrate_source_records_from_sqlite()
            warnings = [
                str(item)
                for item in (self._load_warnings or [])
                if not (
                    str(item or "").startswith("变更源表同步失败")
                    or str(item or "").startswith("智航变更源表同步失败")
                )
            ]
            try:
                self._load_change_fields()
                self._load_change_records()
            except Exception as exc:
                warning = self._source_sync_warning("变更源表", exc)
                if warning not in warnings:
                    warnings.append(warning)
                self._load_warnings = warnings
                raise PortalError(warning) from exc
            try:
                self._load_zhihang_change_fields()
                self._load_zhihang_change_records()
            except Exception as exc:
                warning = self._source_sync_warning("智航变更源表", exc)
                if warning not in warnings:
                    warnings.append(warning)
                self._load_warnings = warnings
                raise PortalError(warning) from exc
            now = dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            self._last_loaded_at = now
            self._last_loaded_ts = time.time()
            self._load_warnings = warnings
            self._save_source_scope_snapshots()
            self._touch_state_cache_version()
            return {
                "change_refreshed_at": now,
                "change_count": len(self._change_records),
                "zhihang_change_count": len(self._zhihang_change_records),
            }

    @staticmethod
    def _normalize_event_month(value: Any = None) -> str:
        text = str(value or "").strip()
        match = re.search(r"(\d{4})[-/年](\d{1,2})", text)
        if match:
            return f"{int(match.group(1)):04d}-{int(match.group(2)):02d}"
        return dt.datetime.now().strftime("%Y-%m")

    def _event_source_config(self) -> tuple[str, str, str]:
        app_token = str(getattr(config, "app_token", "") or self.app_token or DEFAULT_APP_TOKEN).strip()
        table_id = str(getattr(config, "table_id_shijian", "") or "").strip()
        if not app_token or not table_id:
            raise PortalError("未配置事件通告表，请先在 Qt 设置中填写事件通告 Table ID。")
        return app_token, table_id, "event_notice"

    def _load_event_table_fields(
        self, *, app_token: str, table_id: str
    ) -> tuple[list[FieldMeta], dict[str, FieldMeta]]:
        payload = self._request_json(
            "fields",
            params={"page_size": 500},
            app_token=app_token,
            table_id=table_id,
        )
        metas = self._parse_field_metas(payload.get("data", {}).get("items") or [])
        return metas, {meta.field_name: meta for meta in metas}

    @classmethod
    def _event_value(cls, fields: dict[str, Any], key: str) -> str:
        field_name = EVENT_NOTICE_FIELDS.get(key, key)
        return str(fields.get(field_name) or "").strip()

    @classmethod
    def _event_month_candidates(cls, item: dict[str, Any]) -> set[str]:
        fields = item.get("display_fields") if isinstance(item.get("display_fields"), dict) else {}
        candidates: set[str] = set()
        for key in (
            "occurrence_time",
            "progress_update",
            "recover_time",
            "end_time",
            "response_time",
        ):
            parsed = cls._parse_notice_datetime(cls._event_value(fields, key))
            if parsed:
                candidates.add(parsed.strftime("%Y-%m"))
        for raw_key in ("created_time", "last_modified_time"):
            raw_value = item.get(raw_key)
            formatted = cls._format_source_datetime(raw_value)
            parsed = cls._parse_notice_datetime(formatted)
            if parsed:
                candidates.add(parsed.strftime("%Y-%m"))
        return candidates

    @classmethod
    def _event_status(cls, fields: dict[str, Any]) -> str:
        if cls._event_value(fields, "end_time"):
            return "已结束"
        if cls._event_value(fields, "recover_time"):
            return "已恢复待闭环"
        return "处理中"

    @staticmethod
    def _event_high_level(level: Any) -> bool:
        text = str(level or "").strip().upper()
        return bool(re.search(r"\bI[123]\b|I1|I2|I3|升级", text))

    @classmethod
    def _event_building_codes_from_text(cls, value: Any) -> list[str]:
        text = str(value or "").strip().upper()
        codes: list[str] = []
        if "110" in text:
            codes.append("110")
        for match in re.finditer(r"(?<![A-Z0-9])([ABCDEH])\s*(?:楼|栋|站|-|\d)", text):
            code = match.group(1)
            if code not in codes:
                codes.append(code)
        return [code for code in BUILDING_SCOPE_CODES if code in codes]

    def _normalize_event_snapshot_record(
        self,
        normalized: dict[str, Any],
        *,
        month: str,
        source_key: str,
        app_token: str,
        table_id: str,
    ) -> dict[str, Any]:
        fields = normalized.get("display_fields") if isinstance(normalized.get("display_fields"), dict) else {}
        raw_fields = normalized.get("raw_fields") if isinstance(normalized.get("raw_fields"), dict) else {}
        alarm_desc = self._event_value(fields, "alarm_desc")
        title = alarm_desc
        if not title:
            title = str(fields.get("标题") or fields.get("名称") or "").strip()
        if not title:
            title = str(normalized.get("record_id") or "未命名事件").strip()
        building = self._event_value(fields, "building")
        if building:
            building_codes = self._building_codes_from_value(building)
        else:
            building_codes = self._event_building_codes_from_text(f"{title}\n{alarm_desc}")
        building_label = self._building_label_from_codes(building_codes) or str(building or "")
        level = self._event_value(fields, "level")
        event_item = {
            "source_key": source_key,
            "source_record_id": str(normalized.get("record_id") or ""),
            "record_id": str(normalized.get("record_id") or ""),
            "source_app_token": app_token,
            "source_table_id": table_id,
            "source_record_url": f"https://vnet.feishu.cn/base/{app_token}?table={table_id}",
            "month": month,
            "title": title,
            "alarm_desc": alarm_desc,
            "level": level,
            "high_level": self._event_high_level(level),
            "building": building_label,
            "building_codes": building_codes,
            "specialty": self._event_value(fields, "specialty"),
            "source": self._event_value(fields, "source"),
            "occurrence_time": self._event_value(fields, "occurrence_time"),
            "response_time": self._event_value(fields, "response_time"),
            "progress_update": self._event_value(fields, "progress_update"),
            "recover_time": self._event_value(fields, "recover_time"),
            "end_time": self._event_value(fields, "end_time"),
            "transfer_to_overhaul": self._event_value(fields, "transfer_to_overhaul"),
            "status": self._event_status(fields),
            "display_fields": fields,
            "raw_fields": raw_fields,
            "created_time": normalized.get("created_time") or "",
            "last_modified_time": normalized.get("last_modified_time") or "",
        }
        return event_item

    def refresh_event_month_snapshot(self, month: str | None = None) -> dict[str, Any]:
        month = self._normalize_event_month(month)
        app_token, table_id, source_key = self._event_source_config()
        meta: dict[str, Any] = {
            "source_key": source_key,
            "source_app_token": app_token,
            "source_table_id": table_id,
            "month": month,
            "warnings": [],
        }
        try:
            _metas, meta_by_name = self._load_event_table_fields(
                app_token=app_token,
                table_id=table_id,
            )
            all_records = self._load_table_records(
                app_token=app_token,
                table_id=table_id,
                meta_by_name=meta_by_name,
                work_type=WORK_TYPE_EVENT,
                notice_type=NOTICE_TYPE_EVENT,
            )
            snapshot_records: list[dict[str, Any]] = []
            for item in all_records:
                candidate_months = self._event_month_candidates(item)
                if candidate_months and month not in candidate_months:
                    continue
                if not candidate_months and month != dt.datetime.now().strftime("%Y-%m"):
                    continue
                snapshot_records.append(
                    self._normalize_event_snapshot_record(
                        item,
                        month=month,
                        source_key=source_key,
                        app_token=app_token,
                        table_id=table_id,
                    )
                )
            snapshot_records.sort(
                key=lambda item: (
                    str(item.get("occurrence_time") or item.get("progress_update") or ""),
                    str(item.get("source_record_id") or ""),
                ),
                reverse=True,
            )
            result = self._state_store.replace_event_month_snapshot(
                month,
                snapshot_records,
                meta=meta,
            )
            self._touch_state_cache_version()
            return {
                **result,
                "event_refreshed_at": dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "event_count": len(snapshot_records),
                "source_key": source_key,
            }
        except Exception as exc:
            warning = f"事件通告表同步失败: {exc}"
            self._state_store.record_failed_event_month_snapshot(
                month,
                meta={**meta, "warnings": [warning]},
                error=warning,
            )
            if isinstance(exc, PortalError):
                raise
            raise PortalError(warning) from exc

    def get_event_monthly_snapshot(
        self,
        *,
        scope: str,
        month: str | None = None,
    ) -> dict[str, Any]:
        month = self._normalize_event_month(month)
        snapshot = self._state_store.get_event_month_snapshot(month)
        records = [
            dict(item)
            for item in snapshot.get("records", [])
            if isinstance(item, dict) and self._scope_matches_item(scope, item)
        ]
        statuses: dict[str, int] = {}
        levels: dict[str, int] = {}
        sources: dict[str, int] = {}
        specialties: dict[str, int] = {}
        high_level = 0
        for item in records:
            status = str(item.get("status") or "未知").strip() or "未知"
            statuses[status] = int(statuses.get(status) or 0) + 1
            level = str(item.get("level") or "未填写").strip() or "未填写"
            levels[level] = int(levels.get(level) or 0) + 1
            source = str(item.get("source") or "未填写").strip() or "未填写"
            sources[source] = int(sources.get(source) or 0) + 1
            specialty = str(item.get("specialty") or "未填写").strip() or "未填写"
            specialties[specialty] = int(specialties.get(specialty) or 0) + 1
            if bool(item.get("high_level")):
                high_level += 1
        stats = {
            "total": len(records),
            "processing": statuses.get("处理中", 0) + statuses.get("已恢复待闭环", 0),
            "ended": statuses.get("已结束", 0),
            "high_level": high_level,
            "statuses": statuses,
            "levels": levels,
            "sources": sources,
            "specialties": specialties,
        }
        config_missing = False
        config_error = ""
        try:
            self._event_source_config()
        except Exception as exc:
            config_missing = True
            config_error = str(exc)
        return {
            "scope": self._normalize_scope(scope),
            "month": month,
            "records": records,
            "stats": stats,
            "snapshot_exists": bool(snapshot.get("exists")),
            "snapshot_id": snapshot.get("snapshot_id") or "",
            "last_refreshed_at": float(snapshot.get("updated_at") or 0),
            "last_failed": snapshot.get("last_failed") or {},
            "source_meta": snapshot.get("meta") or {},
            "config_missing": config_missing,
            "config_error": config_error,
        }

    def _event_snapshot_record_for_repair(
        self,
        *,
        scope: str,
        record_id: str,
        month: str | None = None,
    ) -> dict[str, Any]:
        record_id = str(record_id or "").strip()
        if not record_id:
            raise PortalError("缺少来源事件记录 ID。")
        del month
        event_app_token, event_table_id, _source_key = self._event_source_config()
        _metas, meta_by_name, cached_records = (
            self._load_repair_management_event_records()
        )
        for cached_record in cached_records:
            if str(cached_record.get("record_id") or "").strip() != record_id:
                continue
            event_item = self._repair_management_event_item(cached_record)
            if not self._scope_matches_item(scope, event_item):
                raise PortalError("当前账号无权关联该楼栋事件。")
            return event_item
        payload = self._request_json(
            f"records/{record_id}",
            params={"user_id_type": "open_id"},
            app_token=event_app_token,
            table_id=event_table_id,
        )
        raw_record = (payload.get("data") or {}).get("record")
        if not isinstance(raw_record, dict):
            raise PortalError("未找到来源事件，可能已被删除。")
        normalized = self._normalize_record(
            raw_record,
            meta_by_name=meta_by_name,
            work_type=WORK_TYPE_EVENT,
            notice_type=NOTICE_TYPE_EVENT,
            source_app_token=event_app_token,
            source_table_id=event_table_id,
        )
        event_item = self._repair_management_event_item(normalized)
        if not self._scope_matches_item(scope, event_item):
            raise PortalError("当前账号无权关联该楼栋事件。")
        return event_item

    def list_repair_management_event_candidates(
        self,
        *,
        scope: str = "ALL",
        month: str | None = None,
        query: str = "",
        limit: int = 50,
    ) -> dict[str, Any]:
        month_key = self._normalize_event_month(month)
        _metas, _meta_by_name, source_records = (
            self._load_repair_management_event_records()
        )
        query_text = str(query or "").strip().lower()
        records: list[dict[str, Any]] = []
        for source_record in source_records:
            if not isinstance(source_record, dict):
                continue
            item = self._repair_management_event_item(source_record)
            transfer_text = self._repair_management_plain_text(
                item.get("transfer_to_overhaul")
            )
            if "未转检修" in transfer_text:
                continue
            if not self._scope_matches_item(scope, item):
                continue
            occurrence = self._parse_notice_datetime(item.get("occurrence_time"))
            if occurrence and occurrence.strftime("%Y-%m") != month_key:
                continue
            haystack = "\n".join(
                str(value or "")
                for value in (
                    item.get("title"),
                    item.get("alarm_desc"),
                    item.get("building"),
                    item.get("specialty"),
                    item.get("source"),
                    item.get("level"),
                    item.get("status"),
                    item.get("occurrence_time"),
                )
            ).lower()
            if query_text and query_text not in haystack:
                continue
            records.append(
                {
                    "record_id": str(item.get("source_record_id") or item.get("record_id") or ""),
                    "title": str(item.get("title") or item.get("alarm_desc") or "未命名事件"),
                    "building": str(item.get("building") or ""),
                    "building_codes": item.get("building_codes") or [],
                    "specialty": str(item.get("specialty") or ""),
                    "level": str(item.get("level") or ""),
                    "source": str(item.get("source") or ""),
                    "status": str(item.get("status") or ""),
                    "occurrence_time": str(item.get("occurrence_time") or ""),
                    "progress_update": str(item.get("progress_update") or ""),
                    "transfer_to_overhaul": str(item.get("transfer_to_overhaul") or ""),
                }
            )
        records.sort(
            key=lambda item: str(item.get("occurrence_time") or item.get("progress_update") or ""),
            reverse=True,
        )
        records.sort(key=lambda item: 1 if str(item.get("status") or "") == "已结束" else 0)
        max_limit = max(1, min(int(limit or 50), 200))
        return {
            "scope": self._normalize_scope(scope),
            "month": month_key,
            "records": records[:max_limit],
            "total": len(records),
            "returned": min(len(records), max_limit),
            "snapshot_exists": True,
            "last_refreshed_at": 0,
            "last_failed": {},
        }

    def repair_management_event_prefill(
        self,
        *,
        scope: str = "ALL",
        record_id: str = "",
        month: str | None = None,
    ) -> dict[str, Any]:
        _metas, meta_by_name, _records = (
            self._load_repair_management_project_records()
        )
        payload = self._build_repair_management_prefill(
            scope=scope,
            event_record_id=record_id,
            repair_record_ids=[],
            followup_record_ids=[],
            month=month,
            meta_by_name=meta_by_name,
        )
        event = payload.get("event") if isinstance(payload.get("event"), dict) else {}
        return {
            "event": {
                "record_id": str(event.get("source_record_id") or event.get("record_id") or ""),
                "title": str(event.get("title") or event.get("alarm_desc") or ""),
                "building": event.get("building") or "",
                "building_codes": event.get("building_codes") or [],
                "specialty": event.get("specialty") or "",
                "level": event.get("level") or "",
                "source": event.get("source") or "",
                "occurrence_time": event.get("occurrence_time") or "",
                "status": event.get("status") or "",
            },
            "fields": payload.get("fields") or {},
            "field_count": len(payload.get("fields") or {}),
            "warnings": payload.get("warnings") or [],
            "month": self._normalize_event_month(month),
        }

    @staticmethod
    def _event_stats_for_records(records: list[dict[str, Any]]) -> dict[str, Any]:
        statuses: dict[str, int] = {}
        levels: dict[str, int] = {}
        sources: dict[str, int] = {}
        specialties: dict[str, int] = {}
        high_level = 0
        for item in records:
            status = str(item.get("status") or "未知").strip() or "未知"
            statuses[status] = int(statuses.get(status) or 0) + 1
            level = str(item.get("level") or "未填写").strip() or "未填写"
            levels[level] = int(levels.get(level) or 0) + 1
            source = str(item.get("source") or "未填写").strip() or "未填写"
            sources[source] = int(sources.get(source) or 0) + 1
            specialty = str(item.get("specialty") or "未填写").strip() or "未填写"
            specialties[specialty] = int(specialties.get(specialty) or 0) + 1
            if bool(item.get("high_level")):
                high_level += 1
        return {
            "total": len(records),
            "processing": statuses.get("处理中", 0) + statuses.get("已恢复待闭环", 0),
            "pending": statuses.get("已恢复待闭环", 0),
            "ended": statuses.get("已结束", 0),
            "high_level": high_level,
            "statuses": statuses,
            "levels": levels,
            "sources": sources,
            "specialties": specialties,
        }

    def get_event_monthly_overview(self, *, month: str | None = None) -> dict[str, Any]:
        month = self._normalize_event_month(month)
        snapshot = self._state_store.get_event_month_snapshot(month)
        records = [
            dict(item)
            for item in snapshot.get("records", [])
            if isinstance(item, dict)
        ]
        by_code: dict[str, list[dict[str, Any]]] = {code: [] for code in BUILDING_SCOPE_CODES}
        for item in records:
            codes = item.get("building_codes")
            if not isinstance(codes, list):
                codes = []
            clean_codes = self._clean_building_codes(codes)
            if not clean_codes:
                clean_codes = self._building_codes_from_value(item.get("building"))
            for code in clean_codes:
                if code in by_code:
                    by_code[code].append(item)
        building_stats = []
        for code in BUILDING_SCOPE_CODES:
            code_records = by_code.get(code) or []
            code_stats = self._event_stats_for_records(code_records)
            building_stats.append({
                "code": code,
                "label": self._scope_label(code),
                **code_stats,
            })
        config_missing = False
        config_error = ""
        try:
            self._event_source_config()
        except Exception as exc:
            config_missing = True
            config_error = str(exc)
        return {
            "scope": "ALL",
            "month": month,
            "stats": self._event_stats_for_records(records),
            "building_stats": building_stats,
            "snapshot_exists": bool(snapshot.get("exists")),
            "snapshot_id": snapshot.get("snapshot_id") or "",
            "last_refreshed_at": float(snapshot.get("updated_at") or 0),
            "last_failed": snapshot.get("last_failed") or {},
            "source_meta": snapshot.get("meta") or {},
            "config_missing": config_missing,
            "config_error": config_error,
        }

    def refresh_if_interval_elapsed(self, *, min_interval_seconds: int = 60) -> bool:
        min_interval = max(0, int(min_interval_seconds or 0))
        with self._refresh_lock:
            if (
                min_interval
                and self._last_loaded_ts
                and time.time() - float(self._last_loaded_ts) < min_interval
            ):
                return False
            self.refresh()
            return True

    def ensure_loaded(self, *, refresh_if_expired: bool = False) -> None:
        with self._refresh_lock:
            if (
                not self._maintenance_loaded_once
                or (refresh_if_expired and self._source_cache_expired())
            ):
                if (
                    not refresh_if_expired
                    and not self._maintenance_loaded_once
                    and self._hydrate_source_records_from_sqlite()
                ):
                    return
                self.refresh()
                return
            warnings: list[str] = []
            attempted_optional_load = False
            if not self._change_loaded_once:
                attempted_optional_load = True
                try:
                    self._load_change_fields()
                    self._load_change_records()
                except Exception as exc:
                    self._change_loaded_once = True
                    warnings.append(self._source_sync_warning("变更源表", exc))
            if not self._zhihang_change_loaded_once:
                attempted_optional_load = True
                try:
                    self._load_zhihang_change_fields()
                    self._load_zhihang_change_records()
                except Exception as exc:
                    self._zhihang_change_loaded_once = True
                    warnings.append(self._source_sync_warning("智航变更源表", exc))
            if not self._repair_loaded_once:
                attempted_optional_load = True
                try:
                    self._load_repair_fields()
                    self._load_repair_records()
                except Exception as exc:
                    self._repair_loaded_once = True
                    warnings.append(self._source_sync_warning("检修源表", exc))
            if attempted_optional_load:
                self._load_warnings = warnings

    def ensure_snapshot_loaded(self) -> None:
        if (
            self._maintenance_loaded_once
            and self._change_loaded_once
            and self._repair_loaded_once
            and self._zhihang_change_loaded_once
        ):
            return
        acquired = False
        try:
            acquired = self._refresh_lock.acquire(blocking=False)
            if not acquired:
                self._hydrate_source_records_from_sqlite()
                return
            if not (
                self._maintenance_loaded_once
                and self._change_loaded_once
                and self._repair_loaded_once
                and self._zhihang_change_loaded_once
            ):
                self._hydrate_source_records_from_sqlite()
        finally:
            if acquired:
                self._refresh_lock.release()

    def _normalize_field_value(
        self,
        field_name: str,
        raw_value: Any,
        meta_by_name: dict[str, FieldMeta] | None = None,
    ) -> Any:
        meta_lookup = meta_by_name or self._field_meta_by_name
        meta = meta_lookup.get(field_name)
        option_map = meta.options_map if meta else {}
        if raw_value is None:
            return ""

        fallback_map = FIELD_OPTION_FALLBACK.get(field_name, {})

        def normalize_option_text(value: Any) -> str:
            text = str(value or "").strip()
            if not text:
                return ""
            if field_name == "变更进度" and CHANGE_PROGRESS_OPTION_FALLBACK.get(text):
                return CHANGE_PROGRESS_OPTION_FALLBACK[text]
            if fallback_map.get(text):
                return fallback_map[text]
            return option_map.get(text, text)

        if isinstance(raw_value, str):
            text = raw_value.strip()
            if field_name in REPAIR_DATETIME_FIELDS:
                formatted = self._format_source_datetime(text)
                if formatted:
                    return formatted
            return normalize_option_text(text)
        if isinstance(raw_value, (int, float)):
            if field_name in REPAIR_DATETIME_FIELDS:
                formatted = self._format_source_datetime(raw_value)
                if formatted:
                    return formatted
            if meta and meta.ui_type in {
                "DateTime",
                "CreatedTime",
                "ModifiedTime",
            }:
                try:
                    return dt.datetime.fromtimestamp(raw_value / 1000).strftime(
                        "%Y-%m-%d %H:%M"
                    )
                except Exception:
                    return str(raw_value)
            return str(raw_value)
        if isinstance(raw_value, dict):
            users = raw_value.get("users")
            if isinstance(users, list):
                names = [str(user.get("name") or "").strip() for user in users]
                return "、".join([name for name in names if name])
            for key in ("name", "text", "value"):
                value = str(raw_value.get(key) or "").strip()
                if value:
                    return normalize_option_text(value)
            option_id = str(raw_value.get("id") or "").strip()
            if option_id:
                resolved = normalize_option_text(option_id)
                if resolved:
                    return resolved
            return json.dumps(raw_value, ensure_ascii=False)
        if isinstance(raw_value, list):
            normalized_items: list[str] = []
            for item in raw_value:
                if isinstance(item, dict):
                    if "text" in item:
                        value = normalize_option_text(item.get("text"))
                    elif "name" in item:
                        value = normalize_option_text(item.get("name"))
                    elif "value" in item:
                        value = normalize_option_text(item.get("value"))
                    elif "id" in item:
                        value = normalize_option_text(item.get("id"))
                    else:
                        value = ""
                    if value:
                        normalized_items.append(value)
                    continue
                if isinstance(item, (int, float)) and field_name in REPAIR_DATETIME_FIELDS:
                    item_text = self._format_source_datetime(item)
                    if item_text:
                        normalized_items.append(item_text)
                    continue
                item_text = str(item or "").strip()
                value = normalize_option_text(item_text)
                if value:
                    normalized_items.append(value)
            return "、".join(normalized_items)
        return str(raw_value)

    def _normalize_record(
        self,
        item: dict[str, Any],
        *,
        meta_by_name: dict[str, FieldMeta] | None = None,
        work_type: str = WORK_TYPE_MAINTENANCE,
        notice_type: str = NOTICE_TYPE_MAINTENANCE,
        source_app_token: str = "",
        source_table_id: str = "",
    ) -> dict[str, Any]:
        raw_fields = item.get("fields") or {}
        display_fields = {
            field_name: self._normalize_field_value(
                field_name, raw_value, meta_by_name=meta_by_name
            )
            for field_name, raw_value in raw_fields.items()
        }
        return {
            "record_id": str(item.get("record_id") or item.get("id") or ""),
            "created_time": item.get("created_time") or item.get("created_at") or "",
            "last_modified_time": item.get("last_modified_time")
            or item.get("updated_time")
            or item.get("updated_at")
            or "",
            "raw_fields": raw_fields,
            "display_fields": display_fields,
            "work_type": work_type,
            "notice_type": notice_type,
            "source_app_token": source_app_token,
            "source_table_id": source_table_id,
        }

    @staticmethod
    def _current_month_label() -> str:
        return f"{dt.datetime.now().month}月"

    @staticmethod
    def _month_start(value: dt.datetime | None = None) -> dt.datetime:
        value = value or dt.datetime.now()
        return dt.datetime(value.year, value.month, 1)

    @classmethod
    def _recent_month_starts(cls) -> list[dt.datetime]:
        current = cls._month_start()
        previous_last_day = current - dt.timedelta(days=1)
        previous = dt.datetime(previous_last_day.year, previous_last_day.month, 1)
        return [current, previous]

    @classmethod
    def _recent_month_pairs(cls) -> list[tuple[str, str]]:
        return [
            (f"{month_start.month}月", month_start.strftime("%Y-%m"))
            for month_start in cls._recent_month_starts()
        ]

    @classmethod
    def _recent_month_labels(cls) -> list[str]:
        return [label for label, _key in cls._recent_month_pairs()]

    @classmethod
    def _recent_month_keys(cls) -> set[str]:
        return {key for _label, key in cls._recent_month_pairs()}

    @staticmethod
    def _normalize_month_label(value: Any) -> str:
        text = str(value or "").strip()
        match = re.search(r"(\d{1,2})\s*月", text)
        return f"{int(match.group(1))}月" if match else text

    @classmethod
    def _specific_recent_month_label(cls, value: Any) -> str | None:
        text = cls._normalize_month_label(value)
        if not text or text == RECENT_MONTH_FILTER_LABEL:
            return ""
        return text if text in cls._recent_month_labels() else None

    @classmethod
    def _specific_recent_month_key(cls, value: Any) -> str | None:
        label = cls._specific_recent_month_label(value)
        if label is None:
            return None
        if not label:
            return ""
        for recent_label, recent_key in cls._recent_month_pairs():
            if recent_label == label:
                return recent_key
        return None

    @classmethod
    def _recent_month_filter_options(cls) -> list[str]:
        return [RECENT_MONTH_FILTER_LABEL, *cls._recent_month_labels()]

    @staticmethod
    def _format_input_datetime(value: Any) -> str:
        text = str(value or "").strip()
        return text.replace("T", " ")

    @classmethod
    def _action_response_time(cls, request_payload: dict[str, Any]) -> str:
        explicit = (
            cls._format_input_datetime((request_payload or {}).get("actual_action_time"))
            or cls._format_input_datetime((request_payload or {}).get("response_time"))
        ).strip()
        return explicit or dt.datetime.now().strftime("%Y-%m-%d %H:%M")

    @classmethod
    def _parse_notice_datetime(cls, value: Any) -> dt.datetime | None:
        text = cls._format_input_datetime(value).replace("/", "-").strip()
        if not text:
            return None
        match = re.search(
            r"(\d{4})-(\d{1,2})-(\d{1,2})\s+(\d{1,2}):(\d{1,2})",
            text,
        )
        if not match:
            match = re.search(
                r"(\d{4})[年-](\d{1,2})[月-](\d{1,2})日?\s*"
                r"(\d{1,2})(?:[：:点时.](\d{1,2}))?",
                text,
            )
        if not match:
            return None
        try:
            return dt.datetime(
                int(match.group(1)),
                int(match.group(2)),
                int(match.group(3)),
                int(match.group(4)),
                int(match.group(5) or 0),
            )
        except Exception:
            return None

    @classmethod
    def _validate_minimum_notice_duration(
        cls,
        start_time: Any,
        end_time: Any,
        *,
        start_label: str = "开始时间",
        end_label: str = "结束时间",
    ) -> None:
        start_dt = cls._parse_notice_datetime(start_time)
        end_dt = cls._parse_notice_datetime(end_time)
        if not start_dt or not end_dt:
            return
        if (end_dt - start_dt).total_seconds() < 3600:
            raise PortalError(f"{start_label}和{end_label}之间不能少于1小时。")

    @staticmethod
    def _has_site_photo_payload(payload: dict[str, Any] | None) -> bool:
        payload = payload if isinstance(payload, dict) else {}
        images = payload.get("extra_images")
        if not isinstance(images, list):
            images = payload.get("site_photos")
        if not isinstance(images, list):
            return False
        for item in images:
            if isinstance(item, dict):
                if str(
                    item.get("upload_id")
                    or item.get("file_token")
                    or item.get("token")
                    or ""
                ).strip():
                    return True
            elif str(item or "").strip():
                return True
        return False

    @staticmethod
    def _site_photo_count_value(payload: dict[str, Any] | None) -> int:
        payload = payload if isinstance(payload, dict) else {}
        for key in ("site_photo_count", "site_photos_count", "extra_image_count"):
            try:
                count = int(payload.get(key) or 0)
            except (TypeError, ValueError):
                count = 0
            if count > 0:
                return count
        return 0

    @classmethod
    def _end_site_photo_required(
        cls,
        *,
        notice_type: Any = "",
        work_type: Any = "",
    ) -> bool:
        notice_text = str(notice_type or "").strip()
        work_text = str(work_type or "").strip()
        if notice_text in END_SITE_PHOTO_REQUIRED_NOTICE_TYPES:
            return True
        if work_text in END_SITE_PHOTO_REQUIRED_WORK_TYPES:
            return True
        mapped = WORK_TYPE_BY_NOTICE_TYPE.get(notice_text, "")
        return mapped in END_SITE_PHOTO_REQUIRED_WORK_TYPES

    @classmethod
    def _require_end_site_photo(
        cls,
        request_payload: dict[str, Any],
        action: str,
        *,
        notice_type: Any = "",
        work_type: Any = "",
    ) -> None:
        if str(action or "").strip().lower() != "end":
            return
        if not cls._end_site_photo_required(
            notice_type=notice_type,
            work_type=work_type,
        ):
            return
        if cls._has_site_photo_payload(request_payload):
            return
        raise PortalError("结束通告前必须添加至少一张现场照片。")

    def _has_existing_site_photo_for_notice(
        self,
        request_payload: dict[str, Any],
        *,
        work_type: Any = "",
    ) -> bool:
        request_payload = (
            normalize_notice_identity_payload(request_payload)
            if isinstance(request_payload, dict)
            else {}
        )
        work_type_text = str(work_type or request_payload.get("work_type") or "").strip()
        active_item_id = str(request_payload.get("active_item_id") or "").strip()
        source_record_id = canonical_source_record_id(request_payload)
        target_record_id = canonical_target_record_id(request_payload)
        request_record_id = str(request_payload.get("record_id") or "").strip()
        target_ids = {
            item
            for item in (target_record_id, request_record_id)
            if item and not is_local_record_id(item)
        }

        def matches_notice(payload: dict[str, Any]) -> bool:
            payload = normalize_notice_identity_payload(payload)
            payload_work_type = str(
                payload.get("work_type") or WORK_TYPE_MAINTENANCE
            ).strip()
            if work_type_text and payload_work_type != work_type_text:
                return False
            if active_item_id and str(payload.get("active_item_id") or "").strip() == active_item_id:
                return True
            if source_record_id and canonical_source_record_id(payload) == source_record_id:
                return True
            if target_ids and canonical_target_record_id(payload) in target_ids:
                return True
            return False

        with self._summary_lock:
            status_items = self._load_work_status_items_locked("ALL")
        for item in status_items:
            if (
                isinstance(item, dict)
                and matches_notice(item)
                and (
                    self._has_site_photo_payload(item)
                    or self._site_photo_count_value(item) > 0
                )
            ):
                return True

        try:
            active_items = self._state_store.list_qt_active_items(include_deleted=False)
        except Exception:
            active_items = []
        for row in active_items:
            if not isinstance(row, dict):
                continue
            payload = row.get("payload") if isinstance(row.get("payload"), dict) else {}
            merged_payload = dict(payload)
            merged_payload.setdefault("active_item_id", row.get("active_item_id"))
            merged_payload.setdefault("target_record_id", row.get("record_id"))
            merged_payload.setdefault("record_id", row.get("record_id"))
            if matches_notice(merged_payload) and (
                self._has_site_photo_payload(merged_payload)
                or self._site_photo_count_value(merged_payload) > 0
            ):
                return True
        return False

    def _require_end_site_photo_cumulative(
        self,
        request_payload: dict[str, Any],
        action: str,
        *,
        notice_type: Any = "",
        work_type: Any = "",
    ) -> None:
        if str(action or "").strip().lower() != "end":
            return
        if not self._end_site_photo_required(
            notice_type=notice_type,
            work_type=work_type,
        ):
            return
        if self._has_site_photo_payload(request_payload):
            return
        if self._has_existing_site_photo_for_notice(
            request_payload,
            work_type=work_type,
        ):
            return
        raise PortalError("结束通告前必须至少上传过一张现场照片，可在开始、更新或结束任意一次上传。")

    @staticmethod
    def _site_photo_payload(request_payload: dict[str, Any]) -> list[dict[str, Any]]:
        images = request_payload.get("extra_images")
        if not isinstance(images, list):
            images = request_payload.get("site_photos")
        if not isinstance(images, list):
            return []
        result: list[dict[str, Any]] = []
        for item in images:
            if isinstance(item, dict):
                copied = dict(item)
                if str(
                    copied.get("upload_id")
                    or copied.get("file_token")
                    or copied.get("token")
                    or ""
                ).strip():
                    result.append(copied)
        return result

    @staticmethod
    def _format_source_datetime(value: Any) -> str:
        if value is None:
            return ""
        if isinstance(value, str):
            text = value.strip()
            if not text:
                return ""
            if re.fullmatch(r"\d+(\.\d+)?", text):
                try:
                    value = float(text)
                except Exception:
                    return text.replace("T", " ")
            else:
                return text.replace("T", " ")
        if isinstance(value, (int, float)):
            numeric = float(value)
            try:
                if numeric > 10_000_000_000:
                    return dt.datetime.fromtimestamp(numeric / 1000).strftime(
                        "%Y-%m-%d %H:%M"
                    )
                if numeric > 1_000_000_000:
                    return dt.datetime.fromtimestamp(numeric).strftime(
                        "%Y-%m-%d %H:%M"
                    )
                if 20_000 <= numeric <= 80_000:
                    base = dt.datetime(1899, 12, 30)
                    return (base + dt.timedelta(days=numeric)).strftime(
                        "%Y-%m-%d %H:%M"
                    )
            except Exception:
                return str(value)
        return str(value or "").strip()

    @staticmethod
    def _building_code_from_value(value: Any) -> str:
        text = str(value or "").strip().upper()
        if "110" in text:
            return "110"
        match = re.search(r"[ABCDEH]", text)
        return match.group(0) if match else ""

    @classmethod
    def _building_codes_from_value(cls, value: Any) -> list[str]:
        values = value if isinstance(value, (list, tuple, set)) else [value]
        codes: list[str] = []
        for raw in values:
            raw_text = str(raw or "").strip()
            mapped = REPAIR_BUILDING_OPTION_FALLBACK.get(raw_text)
            text = str(mapped or raw_text).upper()
            if not text:
                continue
            if "110" in text and "110" not in codes:
                codes.append("110")
            for code in re.findall(r"[ABCDEH]", text):
                if code not in codes:
                    codes.append(code)
        return [code for code in BUILDING_SCOPE_CODES if code in codes]

    @classmethod
    def _building_label_from_code(cls, code: str) -> str:
        code = str(code or "").strip().upper()
        if code == "110":
            return "110站"
        if code in {"A", "B", "C", "D", "E", "H"}:
            return f"{code}楼"
        if code == "CAMPUS":
            return "园区"
        return code

    @classmethod
    def _building_label_from_codes(cls, codes: list[str]) -> str:
        return "、".join(cls._building_label_from_code(code) for code in codes if code)

    @classmethod
    def _recipients_for_building(cls, building: str) -> tuple[str, list[str], str]:
        code = cls._building_code_from_value(building)
        if not code:
            return "", [], f"无法从楼栋字段识别 110/A-E/H: {building or '-'}"
        building_open_id = BUILDING_OPEN_ID_MAP.get(code, "")
        if not building_open_id:
            return code, [], f"未配置 {cls._scope_label(code)} openid"
        return code, list(dict.fromkeys([building_open_id, LI_SHILONG_OPEN_ID])), ""

    @classmethod
    def _recipients_for_building_codes(
        cls, building_codes: list[str], *, fallback_building: str = ""
    ) -> tuple[str, list[str], str]:
        codes = [
            str(code or "").strip().upper()
            for code in (building_codes or [])
            if str(code or "").strip()
        ]
        codes = [code for code in BUILDING_SCOPE_CODES if code in dict.fromkeys(codes)]
        if not codes and fallback_building:
            codes = cls._building_codes_from_value(fallback_building)
        if not codes:
            return "", [], f"无法从楼栋字段识别 110/A-E/H: {fallback_building or '-'}"
        recipients: list[str] = []
        missing: list[str] = []
        for code in codes:
            open_id = BUILDING_OPEN_ID_MAP.get(code, "")
            if open_id:
                recipients.append(open_id)
            else:
                missing.append(cls._scope_label(code))
        recipients.append(LI_SHILONG_OPEN_ID)
        recipients = [item for item in dict.fromkeys(str(open_id or "").strip() for open_id in recipients) if item]
        if missing:
            return codes[0] if len(codes) == 1 else "CAMPUS", recipients, f"未配置 {','.join(missing)} openid"
        return codes[0] if len(codes) == 1 else "CAMPUS", recipients, ""

    @staticmethod
    def _normalize_scope(scope: Any) -> str:
        text = str(scope or "ALL").strip().upper()
        if text in {"全部", "ALL", ""}:
            return "ALL"
        if text in {"园区", "CAMPUS", "PARK"}:
            return "CAMPUS"
        if "110" in text:
            return "110"
        match = re.search(r"[ABCDEH]", text)
        return match.group(0) if match else "ALL"

    @staticmethod
    def _scope_label(scope: Any) -> str:
        normalized = MaintenancePortalService._normalize_scope(scope)
        if normalized == "ALL":
            return "全部"
        if normalized == "CAMPUS":
            return "园区"
        if normalized == "110":
            return "110站"
        return f"{normalized}楼"

    @classmethod
    def _handover_scope_options(cls) -> list[dict[str, str]]:
        return [
            {"value": code, "label": cls._scope_label(code)}
            for code in BUILDING_SCOPE_CODES
        ]

    @classmethod
    def _scope_matches_building(cls, scope: Any, building: Any) -> bool:
        return cls._scope_matches_buildings(scope, cls._building_codes_from_value(building))

    @classmethod
    def _scope_matches_buildings(cls, scope: Any, building_codes: list[str] | tuple[str, ...] | set[str]) -> bool:
        normalized = cls._normalize_scope(scope)
        if normalized == "ALL":
            return True
        codes = [str(code or "").strip().upper() for code in building_codes if str(code or "").strip()]
        codes = [code for code in BUILDING_SCOPE_CODES if code in codes]
        if normalized == "CAMPUS":
            return len(codes) >= 2
        return normalized in codes

    @classmethod
    def _scope_matches_item(cls, scope: Any, item: dict[str, Any]) -> bool:
        codes = item.get("building_codes")
        if not isinstance(codes, list):
            codes = []
        if not codes and str(item.get("building_code") or "").strip().upper() == "CAMPUS":
            codes = cls._building_codes_from_value(item.get("building"))
        if not codes:
            codes = cls._building_codes_from_value(item.get("building"))
        return cls._scope_matches_buildings(scope, codes)

    @classmethod
    def _clean_building_codes(cls, building_codes: Any) -> list[str]:
        codes = building_codes if isinstance(building_codes, (list, tuple, set)) else []
        normalized: list[str] = []
        for code in codes:
            value = str(code or "").strip().upper()
            if value in BUILDING_SCOPE_CODES and value not in normalized:
                normalized.append(value)
        return [code for code in BUILDING_SCOPE_CODES if code in normalized]

    def _building_codes_from_request_payload(
        self, request_payload: dict[str, Any]
    ) -> list[str]:
        codes = self._clean_building_codes(request_payload.get("building_codes"))
        if codes:
            return codes
        return self._building_codes_from_value(request_payload.get("building"))

    def _notice_scope_mismatch_message(
        self, *, scope: Any, building: str, building_codes: list[str], work_type_label: str = "通告"
    ) -> str:
        building_label = (
            str(building or "").strip()
            or self._building_label_from_codes(building_codes)
        )
        if not building_codes:
            return f"{work_type_label}未识别到楼栋，请在“楼栋/范围”中选择楼栋后再发送。"
        return (
            f"当前入口是{self._scope_label(scope)}，但{work_type_label}属于"
            f"{building_label or '-'}，请切换到对应楼栋或园区后再发送。"
        )

    def _resolve_notice_submit_building(
        self,
        *,
        scope: Any,
        request_payload: dict[str, Any],
        building: Any = "",
        building_codes: Any = None,
        title: Any = "",
        location: Any = "",
        content: Any = "",
        work_type_label: str = "通告",
        allow_scope_fallback: bool = False,
    ) -> tuple[str, list[str]]:
        codes = self._clean_building_codes(building_codes)
        title_codes = self._building_codes_from_notice_text(
            title,
            request_payload.get("title"),
        )
        if title_codes == ["110"]:
            codes = ["110"]
        payload_codes = self._building_codes_from_request_payload(request_payload)
        if not codes:
            codes = payload_codes

        building_text = (
            str(building or "").strip()
            or str(request_payload.get("building") or "").strip()
        )
        if not codes and building_text:
            codes = self._building_codes_from_value(building_text)
        if not codes:
            codes = self._building_codes_from_notice_text(
                title,
                location,
                request_payload.get("title"),
                request_payload.get("location"),
            )
        normalized_scope = self._normalize_scope(scope)
        if (
            not codes
            and allow_scope_fallback
            and normalized_scope in BUILDING_SCOPE_CODES
        ):
            codes = [normalized_scope]
        building_text = building_text or self._building_label_from_codes(codes)
        if not codes:
            raise PortalError(
                self._notice_scope_mismatch_message(
                    scope=scope,
                    building=building_text,
                    building_codes=codes,
                    work_type_label=work_type_label,
                )
            )
        if not self._scope_matches_buildings(scope, codes):
            raise PortalError(
                self._notice_scope_mismatch_message(
                    scope=scope,
                    building=building_text,
                    building_codes=codes,
                    work_type_label=work_type_label,
                )
            )
        return building_text or self._building_label_from_codes(codes), codes

    def _source_record_in_scope_snapshot(
        self, *, record_id: str, work_type: str, scope: str
    ) -> dict[str, Any] | None:
        record_id = str(record_id or "").strip()
        if not record_id:
            return None
        records = self._source_snapshot_records(scope)
        if records is None:
            return None
        for record in records:
            if (
                self._record_work_type(record) == work_type
                and str(record.get("record_id") or "").strip() == record_id
            ):
                return record
        return None

    def _resolve_scoped_source_building_codes(
        self,
        *,
        scope: str,
        work_type: str,
        record_id: str,
        source_codes: list[str] | tuple[str, ...] | set[str],
        payload_codes: list[str] | tuple[str, ...] | set[str],
    ) -> list[str]:
        source_building_codes = self._clean_building_codes(source_codes)
        payload_building_codes = self._clean_building_codes(payload_codes)
        if self._scope_matches_buildings(scope, source_building_codes):
            return source_building_codes
        if (
            payload_building_codes
            and self._scope_matches_buildings(scope, payload_building_codes)
            and (
                not source_building_codes
                or self._source_record_in_scope_snapshot(
                    record_id=record_id,
                    work_type=work_type,
                    scope=scope,
                )
                is not None
            )
        ):
            return payload_building_codes
        return source_building_codes or payload_building_codes

    def _sorted_unique_option_names(self, field_name: str) -> list[str]:
        if field_name == "计划维护月份":
            return self._recent_month_filter_options()
        records = self._source_snapshot_records("ALL")
        if records is None:
            records = list(self._records)
        values = {
            str(record["display_fields"].get(field_name) or "").strip()
            for record in records
            if self._maintenance_status_allows_workbench(record)
        }
        options = sorted([value for value in values if value])
        if options:
            return options
        meta = self._field_meta_by_name.get(field_name)
        if meta and meta.option_names:
            return [name for name in meta.option_names if name]
        return []

    def _sorted_unique_work_specialties(self) -> list[str]:
        records = self._source_snapshot_records("ALL")
        if records is None:
            maintenance_records = list(self._records)
            change_records = list(self._change_records)
            repair_records = list(self._repair_records)
        else:
            maintenance_records = [
                record for record in records if self._record_work_type(record) == WORK_TYPE_MAINTENANCE
            ]
            change_records = [
                record for record in records if self._record_work_type(record) == WORK_TYPE_CHANGE
            ]
            repair_records = [
                record for record in records if self._record_work_type(record) == WORK_TYPE_REPAIR
            ]
        values = {
            str(record["display_fields"].get("专业类别") or "").strip()
            for record in maintenance_records
            if self._maintenance_status_allows_workbench(record)
            and self._source_record_matches_month_window(record)
        }
        values.update(
            {
                self._change_specialty(record)
                for record in change_records
                if self._change_progress_allows_workbench(record)
                and self._source_record_matches_month_window(record)
            }
        )
        values.update(
            {
                self._repair_specialty(record)
                for record in repair_records
                if self._is_valid_repair_record(record)
                and self._source_record_matches_month_window(record)
            }
        )
        return sorted([value for value in values if value])

    def _filter_records(
        self,
        *,
        month: str = "",
        specialty: str = "",
        building: str = "",
        scope: str = "ALL",
    ) -> list[dict[str, Any]]:
        month = str(month or "").strip()
        specialty = str(specialty or "").strip()
        building = str(building or "").strip()
        scope = self._normalize_scope(scope)
        filtered = []
        for record in self._records:
            display_fields = record["display_fields"]
            if not self._maintenance_status_allows_workbench(record):
                continue
            if not self._maintenance_record_matches_month_window(record, month):
                continue
            if specialty and display_fields.get("专业类别") != specialty:
                continue
            if building and display_fields.get("楼栋") != building:
                continue
            if not self._scope_matches_building(scope, display_fields.get("楼栋")):
                continue
            filtered.append(record)
        return filtered

    @staticmethod
    def _maintenance_status_value(record: dict[str, Any]) -> str:
        fields = record.get("display_fields") or {}
        return str(fields.get("维护实施状态") or "").strip()

    def _maintenance_status_allows_workbench(self, record: dict[str, Any]) -> bool:
        # 维保源表按当前月份全量展示；已完成/结束类状态由前端禁用点击。
        return True

    def _maintenance_status_is_startable(self, record: dict[str, Any]) -> bool:
        status = self._maintenance_status_value(record)
        return not status or "未开始" in status

    def _maintenance_status_is_completed(self, record: dict[str, Any]) -> bool:
        status = self._maintenance_status_value(record)
        if not status:
            return False
        if status in MAINTENANCE_COMPLETED_STATUSES:
            return True
        if "未完成" in status or "未结束" in status:
            return False
        return any(token in status for token in ("已完成", "已结束", "正常结束", "延期结束", "延迟结束", "闭环"))

    def _change_record_building_codes(self, record: dict[str, Any]) -> list[str]:
        fields = record.get("display_fields") or {}
        raw_building = (
            fields.get("变更楼栋")
            or fields.get("楼栋")
            or record.get("building_codes")
            or ""
        )
        return self._building_codes_from_value(raw_building)

    @staticmethod
    def _change_progress_value(record: dict[str, Any]) -> str:
        fields = record.get("display_fields") or {}
        return str(fields.get("变更进度") or "").strip()

    def _change_progress_allows_workbench(self, record: dict[str, Any]) -> bool:
        return self._change_progress_value(record) in CHANGE_WORKBENCH_PROGRESS_VALUES

    def _change_specialty(self, record: dict[str, Any]) -> str:
        fields = record.get("display_fields") or {}
        return self._clean_source_text(fields.get("专业"))

    @classmethod
    def _repair_building_codes_from_value(cls, value: Any) -> list[str]:
        values = value if isinstance(value, (list, tuple, set)) else [value]
        codes: list[str] = []
        for raw in values:
            text = str(raw or "").strip().upper()
            if not text:
                continue
            if re.search(r"110\s*(?:站|楼|机房|数据中心)?", text) and "110" not in codes:
                codes.append("110")
            for code in ("A", "B", "C", "D", "E", "H"):
                patterns = (
                    rf"(?<![A-Z0-9]){code}\s*(?:楼|栋|座|区|机房|数据中心|DC)",
                    rf"(?:楼栋|楼宇|数据中心)\s*{code}(?![A-Z0-9])",
                    rf"^(?:南通)?{code}$",
                )
                if any(re.search(pattern, text) for pattern in patterns) and code not in codes:
                    codes.append(code)
        return [code for code in BUILDING_SCOPE_CODES if code in codes]

    @staticmethod
    def _clean_source_text(value: Any) -> str:
        text = str(value or "").strip()
        if text in PLACEHOLDER_TEXT_VALUES:
            return ""
        if re.fullmatch(r"opt[A-Za-z0-9]{6,}", text):
            return ""
        return text

    def _repair_notice_title(self, fields: dict[str, Any]) -> str:
        return self._clean_source_text(
            fields.get("检修通告名称") or fields.get("维修名称")
        )

    def _repair_title(self, record: dict[str, Any]) -> str:
        fields = record.get("display_fields") or {}
        title = self._repair_notice_title(fields)
        if title:
            return title
        fallback_title = str(record.get("title") or "").strip()
        if fallback_title:
            return fallback_title
        return str(record.get("record_id") or "").strip()

    @staticmethod
    def _repair_level_from_event_level(value: Any) -> str:
        text = str(value or "").strip().upper()
        if re.search(r"(^|[^A-Z0-9])I3([^A-Z0-9]|$)", text) or text == "低":
            return "低"
        if re.search(r"(^|[^A-Z0-9])I2([^A-Z0-9]|$)", text) or text == "中":
            return "中"
        return ""

    def _repair_level(self, fields: dict[str, Any]) -> str:
        return self._repair_level_from_event_level(fields.get("对应事件等级"))

    def _repair_specialty(self, record: dict[str, Any]) -> str:
        fields = record.get("display_fields") or {}
        specialty = self._clean_source_text(fields.get("所属专业"))
        if specialty:
            return specialty
        pushed_raw = str(fields.get("专业（推送消息用）") or "").strip()
        return REPAIR_SPECIALTY_OPTION_FALLBACK.get(
            pushed_raw, self._clean_source_text(pushed_raw)
        )

    def _repair_record_building_codes(self, record: dict[str, Any]) -> list[str]:
        fields = record.get("display_fields") or {}
        for field_name in (
            "所属数据中心/楼栋-使用",
            "所属数据中心/楼栋（关联CMDB唯一ID关联,DE不选）",
        ):
            codes = self._repair_building_codes_from_value(fields.get(field_name))
            if codes:
                return codes
        fallback_codes = self._repair_building_codes_from_value(record.get("building_codes"))
        if fallback_codes:
            return fallback_codes
        return (
            self._repair_building_codes_from_value(self._repair_title(record))
            or self._repair_building_codes_from_value(fields.get("维修名称"))
        )

    def _is_valid_repair_record(self, record: dict[str, Any]) -> bool:
        fields = record.get("display_fields") or {}
        title = self._repair_notice_title(fields)
        if not title:
            return False
        return bool(self._repair_record_building_codes(record))

    def _repair_time_range(self, record: dict[str, Any]) -> tuple[str, str, str]:
        fields = record.get("display_fields") or {}
        start_time = self._clean_source_text(
            fields.get("维修开始时间")
            or fields.get("故障发生时间")
            or fields.get("发现故障时间")
        )
        end_time = self._clean_source_text(
            fields.get("维修结束时间")
            or fields.get("维修结束时间（2026）")
            or fields.get("期望完成时间")
        )
        time_text = f"{start_time}~{end_time}" if start_time or end_time else ""
        return start_time, end_time, time_text

    def _repair_first_field(self, fields: dict[str, Any], *field_names: str) -> str:
        for field_name in field_names:
            value = self._clean_source_text(fields.get(field_name))
            if value:
                return value
        return ""

    def _repair_device_text(self, fields: dict[str, Any]) -> str:
        device_no = self._clean_source_text(fields.get("设备编号"))
        device_name = self._clean_source_text(fields.get("设备名称"))
        if device_no and device_name:
            return f"{device_no}{device_name}"
        if device_no or device_name:
            return device_no or device_name
        return self._repair_first_field(fields, "维修设备", "资产名称", "设备")

    def _repair_location_text(self, fields: dict[str, Any]) -> str:
        return self._repair_first_field(
            fields,
            "地点",
            "位置",
            "区域",
            "数据中心",
            "所属数据中心/楼栋（关联CMDB唯一ID关联,DE不选）",
            "所属数据中心/楼栋-使用",
        )

    def _repair_target_record_id(self, record: dict[str, Any]) -> str:
        raw_value = (record.get("raw_fields") or {}).get("设备检修关联")
        values = raw_value if isinstance(raw_value, list) else [raw_value]
        for item in values:
            if not isinstance(item, dict):
                continue
            record_ids = item.get("record_ids")
            if isinstance(record_ids, list):
                for record_id in record_ids:
                    text = str(record_id or "").strip()
                    if text:
                        return text
            text = str(item.get("record_id") or item.get("id") or "").strip()
            if text:
                return text
        return ""

    @staticmethod
    def _decode_repair_sync_source_id(value: Any) -> str:
        text = str(value or "").strip()
        if not text:
            return ""
        try:
            padded = text + ("=" * (-len(text) % 4))
            return base64.b64decode(padded).decode("utf-8", errors="replace")
        except Exception:
            return ""

    def _repair_sync_record_matches_target(
        self, record: dict[str, Any], target_record_id: str
    ) -> bool:
        target_record_id = str(target_record_id or "").strip()
        if not target_record_id:
            return False
        fields = record.get("display_fields") or record.get("fields") or {}
        source_id = fields.get("SourceID")
        if not isinstance(source_id, str):
            source_id = self._normalize_field_value("SourceID", source_id, {})
        decoded = self._decode_repair_sync_source_id(source_id)
        parts = [part.strip() for part in decoded.split(":")]
        return len(parts) >= 2 and parts[1] == target_record_id

    def _find_repair_sync_record_id(
        self,
        *,
        target_record_id: str,
        sync_app_token: str = REPAIR_SOURCE_APP_TOKEN,
        sync_table_id: str = REPAIR_SYNC_TABLE_ID,
        max_pages: int = 40,
    ) -> str:
        target_record_id = str(target_record_id or "").strip()
        if not target_record_id:
            return ""
        page_token = ""
        pages = 0
        while pages < max_pages:
            pages += 1
            params: dict[str, Any] = {"page_size": 500}
            if page_token:
                params["page_token"] = page_token
            payload = self._request_json(
                "records",
                params=params,
                app_token=sync_app_token,
                table_id=sync_table_id,
            )
            data = payload.get("data") or {}
            for item in data.get("items") or []:
                if not isinstance(item, dict):
                    continue
                if self._repair_sync_record_matches_target(item, target_record_id):
                    return str(item.get("record_id") or item.get("id") or "").strip()
            if not data.get("has_more"):
                break
            page_token = str(data.get("page_token") or "").strip()
            if not page_token:
                break
        return ""

    @staticmethod
    def _linked_record_ids_from_value(value: Any) -> list[str]:
        linked: list[str] = []

        def add(candidate: Any) -> None:
            text = str(candidate or "").strip()
            if text and text not in linked:
                linked.append(text)

        if isinstance(value, dict):
            for key in ("link_record_ids", "record_ids"):
                ids = value.get(key)
                if isinstance(ids, list):
                    for item in ids:
                        add(item)
            add(value.get("record_id") or value.get("id"))
        elif isinstance(value, list):
            for item in value:
                if isinstance(item, dict):
                    for key in ("link_record_ids", "record_ids"):
                        ids = item.get(key)
                        if isinstance(ids, list):
                            for record_id in ids:
                                add(record_id)
                    add(item.get("record_id") or item.get("id"))
                else:
                    add(item)
        else:
            add(value)
        return linked

    def _source_repair_link_record_ids(
        self, *, source_app_token: str, source_table_id: str, source_record_id: str
    ) -> list[str]:
        payload = self._request_json(
            f"records/{source_record_id}",
            app_token=source_app_token,
            table_id=source_table_id,
        )
        record = (payload.get("data") or {}).get("record") or {}
        fields = record.get("fields") or {}
        return self._linked_record_ids_from_value(fields.get(REPAIR_LINK_FIELD_NAME))

    def _write_source_repair_link(
        self,
        *,
        source_app_token: str,
        source_table_id: str,
        source_record_id: str,
        sync_record_id: str,
    ) -> None:
        linked_ids = self._source_repair_link_record_ids(
            source_app_token=source_app_token,
            source_table_id=source_table_id,
            source_record_id=source_record_id,
        )
        if sync_record_id not in linked_ids:
            linked_ids.append(sync_record_id)
        self._patch_record_fields(
            app_token=source_app_token,
            table_id=source_table_id,
            record_id=source_record_id,
            fields={REPAIR_LINK_FIELD_NAME: {"link_record_ids": linked_ids}},
        )

    def _repair_has_started(self, record: dict[str, Any]) -> bool:
        fields = record.get("display_fields") or {}
        return bool(self._clean_source_text(fields.get("维修开始时间")))

    def _repair_has_ended(self, record: dict[str, Any]) -> bool:
        fields = record.get("display_fields") or {}
        return bool(
            self._clean_source_text(
                fields.get("维修结束时间") or fields.get("维修结束时间（2026）")
            )
        )

    def _repair_source_status(self, record: dict[str, Any]) -> str:
        if self._repair_has_ended(record):
            return "已结束"
        return MAINTENANCE_STATUS_ONGOING if self._repair_has_started(record) else DEFAULT_MAINTENANCE_STATUS

    def _filter_change_records(
        self,
        *,
        month: str = "",
        specialty: str = "",
        scope: str = "ALL",
        progress_values: set[str] | None = None,
    ) -> list[dict[str, Any]]:
        specialty = str(specialty or "").strip()
        scope = self._normalize_scope(scope)
        filtered: list[dict[str, Any]] = []
        for record in self._change_records:
            fields = record.get("display_fields") or {}
            progress = self._change_progress_value(record)
            if progress_values is None:
                if not self._change_progress_allows_workbench(record):
                    continue
            elif progress not in progress_values:
                continue
            if not self._source_record_matches_month_window(record, month):
                continue
            if specialty and specialty != self._change_specialty(record):
                continue
            codes = self._change_record_building_codes(record)
            if not self._scope_matches_buildings(scope, codes):
                continue
            filtered.append(record)
        return filtered

    def _zhihang_change_primary_field_name(self) -> str:
        for meta in self._zhihang_change_field_meta_list:
            if meta.is_primary:
                return meta.field_name
        return ""

    def _zhihang_change_title(self, record: dict[str, Any]) -> str:
        fields = record.get("display_fields") or {}
        for name in (
            "标题",
            "名称",
            "变更名称",
            "变更标题",
            "变更简述",
            "工作内容",
            self._zhihang_change_primary_field_name(),
        ):
            value = str(fields.get(name) or "").strip()
            if value:
                return value
        return str(record.get("record_id") or "").strip()

    def _zhihang_change_progress(self, record: dict[str, Any]) -> str:
        fields = record.get("display_fields") or {}
        return str(fields.get("进度") or fields.get("进展") or "").strip()

    def _zhihang_change_building_codes(self, record: dict[str, Any]) -> list[str]:
        title = self._zhihang_change_title(record).upper()
        if not title:
            return []
        codes: list[str] = []
        if "110" in title:
            codes.append("110")
        for match in re.findall(
            r"(?<![A-Z0-9])([ABCDEH]{2,})(?=楼|栋|座|区|园区|[^A-Z0-9]|$)",
            title,
        ):
            for code in match:
                if code not in codes:
                    codes.append(code)
        for code in ("A", "B", "C", "D", "E", "H"):
            if (
                f"{code}楼" in title
                or f"{code}栋" in title
                or f"{code}座" in title
            ) and code not in codes:
                codes.append(code)
        if "园区" in title and len(codes) < 2:
            codes.extend([code for code in ("A", "B", "C") if code not in codes])
        return [code for code in BUILDING_SCOPE_CODES if code in codes]

    def _zhihang_change_scope_all(self, record: dict[str, Any]) -> bool:
        return not self._zhihang_change_building_codes(record)

    def _scope_matches_zhihang_change_record(
        self, scope: Any, record: dict[str, Any]
    ) -> bool:
        scope = self._normalize_scope(scope)
        if scope == "ALL":
            return True
        if self._zhihang_change_scope_all(record):
            return True
        return self._scope_matches_buildings(
            scope, self._zhihang_change_building_codes(record)
        )

    def _linked_zhihang_record_ids(
        self, ongoing_items: list[dict[str, Any]] | None = None
    ) -> set[str]:
        linked: set[str] = set()
        for item in ongoing_items or []:
            if not isinstance(item, dict):
                continue
            for field in ("zhihang_record_id", "lan_zhihang_record_id"):
                value = str(item.get(field) or "").strip()
                if value:
                    linked.add(value)
        with self._summary_lock:
            status_items = self._load_work_status_items_locked("ALL")
        for item in status_items:
            value = str(item.get("zhihang_record_id") or "").strip()
            if value:
                linked.add(value)
        return linked

    def _filter_zhihang_change_records(
        self,
        *,
        month: str = "",
        scope: str = "ALL",
        exclude_record_ids: set[str] | None = None,
    ) -> list[dict[str, Any]]:
        snapshot_records = self._source_snapshot_zhihang_records(scope)
        if snapshot_records is not None:
            return self._filter_zhihang_change_records_list(
                snapshot_records,
                month=month,
                exclude_record_ids=exclude_record_ids,
            )
        return self._filter_zhihang_change_records_from_memory(
            month=month,
            scope=scope,
            exclude_record_ids=exclude_record_ids,
        )

    def _filter_zhihang_change_records_from_memory(
        self,
        *,
        month: str = "",
        scope: str = "ALL",
        exclude_record_ids: set[str] | None = None,
    ) -> list[dict[str, Any]]:
        records = [
            record
            for record in self._zhihang_change_records
            if self._scope_matches_zhihang_change_record(scope, record)
        ]
        return self._filter_zhihang_change_records_list(
            records,
            month=month,
            exclude_record_ids=exclude_record_ids,
        )

    def _filter_zhihang_change_records_list(
        self,
        records: list[dict[str, Any]],
        *,
        month: str = "",
        exclude_record_ids: set[str] | None = None,
    ) -> list[dict[str, Any]]:
        excluded = {str(value or "").strip() for value in (exclude_record_ids or set())}
        filtered: list[dict[str, Any]] = []
        for record in records or []:
            record_id = str(record.get("record_id") or "").strip()
            if record_id and record_id in excluded:
                continue
            progress = self._zhihang_change_progress(record)
            if progress == ZHIHANG_PROGRESS_ENDED:
                continue
            if not self._source_record_matches_month_window(record, month):
                continue
            filtered.append(record)
        return filtered

    def _find_zhihang_change_record_by_id(self, record_id: str) -> dict[str, Any]:
        record_id = str(record_id or "").strip()
        for record in self._zhihang_change_records:
            if str(record.get("record_id") or "").strip() == record_id:
                return record
        snapshot_records = self._source_snapshot_zhihang_records("ALL")
        for record in snapshot_records or []:
            if str(record.get("record_id") or "").strip() == record_id:
                return record
        raise PortalError(f"未找到智航侧变更记录: {record_id}")

    def _filter_repair_records(
        self,
        *,
        month: str = "",
        specialty: str = "",
        scope: str = "ALL",
        include_started: bool | None = None,
    ) -> list[dict[str, Any]]:
        specialty = str(specialty or "").strip()
        scope = self._normalize_scope(scope)
        filtered: list[dict[str, Any]] = []
        for record in self._repair_records:
            if not self._is_valid_repair_record(record):
                continue
            if self._repair_has_ended(record):
                continue
            if include_started is not None and self._repair_has_started(record) != bool(include_started):
                continue
            if self._repair_source_status(record) not in WORKBENCH_SOURCE_STATUSES:
                continue
            if not self._source_record_matches_month_window(record, month):
                continue
            if specialty and specialty != self._repair_specialty(record):
                continue
            if not self._scope_matches_buildings(
                scope, self._repair_record_building_codes(record)
            ):
                continue
            filtered.append(record)
        return filtered

    @staticmethod
    def _normalize_memory_key(value: str) -> str:
        return re.sub(r"\s+", " ", str(value or "").strip()).lower()

    @staticmethod
    def _safe_memory_filename(building: str) -> str:
        raw = str(building or "").strip()
        name = re.sub(r"[^\w\u4e00-\u9fff.-]+", "_", raw)
        name = name.strip("._") or "unknown"
        digest = hashlib.sha1(raw.encode("utf-8")).hexdigest()[:10]
        return f"{name[:70]}_{digest}.json"

    def _building_memory_path(self, building: str) -> Path:
        return self._memory_dir / self._safe_memory_filename(building)

    def _building_memory_key(self, building: str) -> str:
        raw = str(building or "").strip()
        digest = hashlib.sha1(raw.encode("utf-8")).hexdigest()[:10]
        return raw or f"unknown:{digest}"

    def _load_building_memory_locked(self, building: str) -> dict[str, Any]:
        key = self._building_memory_key(building)
        stored = self._state_store.get_document(STATE_NS_MEMORY, key)
        if isinstance(stored, dict):
            items = stored.get("items")
            if not isinstance(items, dict):
                stored["items"] = {}
            stored["building"] = str(stored.get("building") or building or "")
            return stored
        path = self._building_memory_path(building)
        if not path.exists():
            return {"building": building, "items": {}}
        try:
            with path.open("r", encoding="utf-8") as fp:
                payload = json.load(fp)
            if not isinstance(payload, dict):
                return {"building": building, "items": {}}
            items = payload.get("items")
            if not isinstance(items, dict):
                payload["items"] = {}
            self._state_store.put_document(STATE_NS_MEMORY, key, payload)
            return payload
        except Exception:
            return {"building": building, "items": {}}

    def _save_building_memory_locked(self, building: str, payload: dict[str, Any]) -> None:
        self._state_store.put_document(
            STATE_NS_MEMORY, self._building_memory_key(building), payload
        )

    def _summary_path(self, day: str | None = None) -> Path:
        day = str(day or dt.datetime.now().strftime("%Y-%m-%d")).strip()
        day = re.sub(r"[^0-9-]+", "", day) or dt.datetime.now().strftime("%Y-%m-%d")
        return self._summary_dir / f"{day}.json"

    def _load_day_summary_locked(self, day: str | None = None) -> dict[str, Any]:
        path = self._summary_path(day)
        stored = self._state_store.get_document(STATE_NS_DAILY_SUMMARY, path.stem)
        if isinstance(stored, dict):
            if not isinstance(stored.get("items"), list):
                stored["items"] = []
            stored["date"] = str(stored.get("date") or path.stem)
            return stored
        if not path.exists():
            return {"date": path.stem, "items": []}
        try:
            with path.open("r", encoding="utf-8") as fp:
                payload = json.load(fp)
            if not isinstance(payload, dict):
                return {"date": path.stem, "items": []}
            if not isinstance(payload.get("items"), list):
                payload["items"] = []
            payload["date"] = str(payload.get("date") or path.stem)
            self._state_store.put_document(STATE_NS_DAILY_SUMMARY, path.stem, payload)
            return payload
        except Exception:
            return {"date": path.stem, "items": []}

    def _save_day_summary_locked(self, payload: dict[str, Any], day: str | None = None) -> None:
        path = self._summary_path(day or payload.get("date"))
        self._state_store.put_document(STATE_NS_DAILY_SUMMARY, path.stem, payload)
        self._touch_state_cache_version()

    @staticmethod
    def _summary_key(
        *,
        source_record_id: str = "",
        title: str = "",
        building: str = "",
        reason: str = "",
        work_type: str = "",
    ) -> str:
        source_record_id = str(source_record_id or "").strip()
        work_type = str(work_type or "").strip()
        prefix = f"{work_type}:" if work_type else ""
        if source_record_id:
            return f"{prefix}source:{source_record_id}"
        title_key = re.sub(r"\s+", "", str(title or ""))
        building_key = re.sub(r"\s+", "", str(building or ""))
        reason_key = re.sub(r"\s+", "", str(reason or ""))
        suffix = f":reason:{reason_key}" if reason_key else ""
        return f"{prefix}title:{building_key}:{title_key}{suffix}" if title_key else ""

    @staticmethod
    def _work_status_fallback_key(
        *,
        title: str = "",
        building: str = "",
        plan_month: str = "",
        reason: str = "",
    ) -> str:
        title_key = re.sub(r"\s+", "", str(title or ""))
        building_key = re.sub(r"\s+", "", str(building or ""))
        month_key = re.sub(r"\s+", "", str(plan_month or ""))
        reason_key = re.sub(r"\s+", "", str(reason or ""))
        suffix = f":reason:{reason_key}" if reason_key else ""
        return f"{building_key}:{month_key}:{title_key}{suffix}" if title_key else ""

    def _record_fallback_key(self, record: dict[str, Any]) -> str:
        fields = record.get("display_fields") or {}
        building = str(fields.get("楼栋") or "").strip()
        maintenance_total = str(fields.get("维护总项") or "").strip()
        title = f"EA118机房{building}{maintenance_total}" if maintenance_total else ""
        if record.get("work_type") == WORK_TYPE_CHANGE:
            fields = record.get("display_fields") or {}
            return self._work_status_fallback_key(
                title=str(fields.get("变更简述") or record.get("record_id") or ""),
                building=str(fields.get("变更楼栋") or ""),
                plan_month="",
            )
        if record.get("work_type") == WORK_TYPE_REPAIR:
            return self._work_status_fallback_key(
                title=self._repair_title(record),
                building=self._building_label_from_codes(
                    self._repair_record_building_codes(record)
                ),
                plan_month="",
            )
        return self._work_status_fallback_key(
            title=title,
            building=building,
            plan_month=str(fields.get("计划维护月份") or "").strip(),
        )

    def _record_legacy_summary_key(self, record: dict[str, Any]) -> str:
        fields = record.get("display_fields") or {}
        building = str(fields.get("楼栋") or "").strip()
        maintenance_total = str(fields.get("维护总项") or "").strip()
        title = f"EA118机房{building}{maintenance_total}" if maintenance_total else ""
        return self._summary_key(title=title, building=building)

    def _work_status_path(self, building: str = "", building_code: str = "") -> Path:
        code = str(building_code or self._building_code_from_value(building) or "").strip()
        if code:
            name = code
        else:
            raw = str(building or "").strip()
            name = re.sub(r"[^\w\u4e00-\u9fff.-]+", "_", raw).strip("._") or "unknown"
            name = f"{name[:70]}_{hashlib.sha1(raw.encode('utf-8')).hexdigest()[:10]}"
        return self._work_status_dir / f"{name}.json"

    def _load_work_status_locked(self, building: str = "", building_code: str = "") -> dict[str, Any]:
        path = self._work_status_path(building, building_code)
        stored = self._state_store.get_document(STATE_NS_WORK_STATUS, path.stem)
        if isinstance(stored, dict):
            if not isinstance(stored.get("items"), list):
                stored["items"] = []
            stored["version"] = int(stored.get("version") or 1)
            stored["building"] = str(stored.get("building") or building or "")
            stored["building_code"] = str(
                stored.get("building_code")
                or building_code
                or self._building_code_from_value(stored.get("building") or building)
                or ""
            )
            return stored
        if not path.exists():
            code = str(building_code or self._building_code_from_value(building) or "").strip()
            return {"version": 1, "building": building, "building_code": code, "items": []}
        try:
            with path.open("r", encoding="utf-8") as fp:
                payload = json.load(fp)
            if not isinstance(payload, dict):
                payload = {}
        except Exception:
            payload = {}
        if not isinstance(payload.get("items"), list):
            payload["items"] = []
        payload["version"] = int(payload.get("version") or 1)
        payload["building"] = str(payload.get("building") or building or "")
        payload["building_code"] = str(
            payload.get("building_code")
            or building_code
            or self._building_code_from_value(payload.get("building") or building)
            or ""
        )
        self._state_store.put_document(STATE_NS_WORK_STATUS, path.stem, payload)
        return payload

    def _save_work_status_locked(
        self, payload: dict[str, Any], *, building: str = "", building_code: str = ""
    ) -> None:
        path = self._work_status_path(
            building or payload.get("building") or "",
            building_code or payload.get("building_code") or "",
        )
        self._state_store.put_document(STATE_NS_WORK_STATUS, path.stem, payload)
        self._work_status_cache_signature = None
        self._work_status_cache_items = None
        self._touch_state_cache_version()

    def _migrate_legacy_day_summaries_locked(self) -> None:
        if self._legacy_summary_migrated:
            return
        self._legacy_summary_migrated = True
        for path in sorted(self._summary_dir.glob("*.json")):
            key = path.stem
            if self._state_store.get_document(STATE_NS_DAILY_SUMMARY, key):
                continue
            try:
                with path.open("r", encoding="utf-8") as fp:
                    payload = json.load(fp)
            except Exception:
                continue
            if not isinstance(payload, dict):
                continue
            if not isinstance(payload.get("items"), list):
                payload["items"] = []
            payload["date"] = str(payload.get("date") or key)
            self._state_store.put_document(STATE_NS_DAILY_SUMMARY, key, payload)

    def _iter_day_summary_payloads_locked(
        self, *, month: str = ""
    ) -> list[dict[str, Any]]:
        self._migrate_legacy_day_summaries_locked()
        documents = self._state_store.list_documents(
            STATE_NS_DAILY_SUMMARY, key_prefix=month
        )
        payloads: list[dict[str, Any]] = []
        for document in documents:
            payload = document.get("payload")
            if not isinstance(payload, dict):
                continue
            if not isinstance(payload.get("items"), list):
                payload["items"] = []
            payload["date"] = str(payload.get("date") or document.get("key") or "")
            payloads.append(payload)
        return payloads

    def _migrate_legacy_work_status_locked(self) -> None:
        if self._legacy_work_status_migrated:
            return
        self._legacy_work_status_migrated = True
        for path in sorted(self._work_status_dir.glob("*.json")):
            key = path.stem
            if self._state_store.get_document(STATE_NS_WORK_STATUS, key):
                continue
            try:
                with path.open("r", encoding="utf-8") as fp:
                    payload = json.load(fp)
            except Exception:
                continue
            if not isinstance(payload, dict):
                continue
            if not isinstance(payload.get("items"), list):
                payload["items"] = []
            payload["version"] = int(payload.get("version") or 1)
            self._state_store.put_document(STATE_NS_WORK_STATUS, key, payload)

    def _iter_work_status_payloads_locked(self) -> list[dict[str, Any]]:
        self._migrate_legacy_work_status_locked()
        payloads: list[dict[str, Any]] = []
        for document in self._state_store.list_documents(STATE_NS_WORK_STATUS):
            payload = document.get("payload")
            if not isinstance(payload, dict):
                continue
            if not isinstance(payload.get("items"), list):
                payload["items"] = []
            payloads.append(payload)
        return payloads

    def _find_summary_item(
        self,
        items: list[dict[str, Any]],
        *,
        key: str = "",
        active_item_id: str = "",
        title: str = "",
        building: str = "",
        reason: str = "",
        work_type: str = "",
    ) -> dict[str, Any] | None:
        key = str(key or "").strip()
        active_item_id = str(active_item_id or "").strip()
        reason_key = re.sub(r"\s+", "", str(reason or ""))
        if key:
            for item in items:
                if str(item.get("key") or "") == key:
                    return item
        if active_item_id:
            for item in items:
                if str(item.get("active_item_id") or "") == active_item_id:
                    return item
        fallback_key = self._summary_key(
            title=title, building=building, reason=reason, work_type=work_type
        )
        if fallback_key:
            for item in items:
                if str(item.get("fallback_key") or "") == fallback_key:
                    return item
        legacy_fallback_key = self._summary_key(title=title, building=building)
        if legacy_fallback_key and legacy_fallback_key != fallback_key:
            for item in items:
                item_reason_key = re.sub(r"\s+", "", str(item.get("reason") or ""))
                if (
                    str(item.get("fallback_key") or "") == legacy_fallback_key
                    and (not reason_key or item_reason_key == reason_key)
                ):
                    return item
        return None

    def _find_work_status_item(
        self,
        items: list[dict[str, Any]],
        *,
        source_record_id: str = "",
        active_item_id: str = "",
        key: str = "",
        fallback_key: str = "",
        work_fallback_key: str = "",
        work_type: str = "",
    ) -> dict[str, Any] | None:
        source_record_id = str(source_record_id or "").strip()
        active_item_id = str(active_item_id or "").strip()
        key = str(key or "").strip()
        fallback_key = str(fallback_key or "").strip()
        work_fallback_key = str(work_fallback_key or "").strip()
        work_type = str(work_type or WORK_TYPE_MAINTENANCE).strip()

        def _matches_work_type(item: dict[str, Any]) -> bool:
            return self._item_work_type(item) == work_type

        if source_record_id:
            for item in items:
                if (
                    _matches_work_type(item)
                    and str(item.get("source_record_id") or "").strip()
                    == source_record_id
                ):
                    return item
        if active_item_id:
            for item in items:
                if (
                    _matches_work_type(item)
                    and str(item.get("active_item_id") or "").strip()
                    == active_item_id
                ):
                    return item
        if key:
            for item in items:
                if _matches_work_type(item) and str(item.get("key") or "").strip() == key:
                    return item
        if fallback_key:
            for item in items:
                if (
                    _matches_work_type(item)
                    and str(item.get("fallback_key") or "").strip() == fallback_key
                ):
                    return item
        if work_fallback_key:
            for item in items:
                if (
                    _matches_work_type(item)
                    and str(item.get("work_fallback_key") or "").strip()
                    == work_fallback_key
                ):
                    return item
        return None

    @staticmethod
    def _date_part(value: Any) -> str:
        match = re.search(r"\d{4}-\d{1,2}-\d{1,2}", str(value or ""))
        return match.group(0) if match else ""

    def _upsert_work_status_item_locked(
        self,
        incoming: dict[str, Any],
        *,
        action: str = "",
        now: str = "",
    ) -> None:
        if not isinstance(incoming, dict):
            return
        action = str(action or "").strip().lower()
        now = str(now or dt.datetime.now().strftime("%Y-%m-%d %H:%M"))
        building = str(incoming.get("building") or "").strip()
        building_code = str(incoming.get("building_code") or "").strip()
        work_type = self._item_work_type(incoming)
        payload = self._load_work_status_locked(building, building_code)
        if building and not payload.get("building"):
            payload["building"] = building
        if building_code and not payload.get("building_code"):
            payload["building_code"] = building_code
        items = payload.setdefault("items", [])
        source_record_id = str(incoming.get("source_record_id") or "").strip()
        active_item_id = str(incoming.get("active_item_id") or "").strip()
        key = str(incoming.get("key") or "").strip()
        fallback_key = str(incoming.get("fallback_key") or "").strip()
        work_fallback_key = str(incoming.get("work_fallback_key") or "").strip()
        work_fallback_key = work_fallback_key or self._work_status_fallback_key(
            title=str(incoming.get("title") or ""),
            building=building,
            plan_month=str(incoming.get("plan_month") or ""),
            reason=str(incoming.get("reason") or ""),
        )
        item = self._find_work_status_item(
            items,
            source_record_id=source_record_id,
            active_item_id=active_item_id,
            key=key,
            fallback_key=fallback_key,
            work_fallback_key=work_fallback_key,
            work_type=work_type,
        )
        if item is None:
            item = {
                "key": key or fallback_key or uuid.uuid4().hex,
                "fallback_key": fallback_key,
                "work_fallback_key": work_fallback_key,
                "source_record_id": source_record_id,
                "active_item_id": active_item_id,
                "work_type": work_type,
                "actions": [],
            }
            items.append(item)
        item["work_type"] = self._item_work_type(item) if item.get("work_type") else work_type
        if work_fallback_key and not item.get("work_fallback_key"):
            item["work_fallback_key"] = work_fallback_key
        for field in (
            "key",
            "fallback_key",
            "work_fallback_key",
            "work_type",
            "notice_type",
            "source_app_token",
            "source_table_id",
            "source_record_id",
            "target_record_id",
            "active_item_id",
            "title",
            "building",
            "building_code",
            "building_codes",
            "specialty",
            "level",
            "source_progress",
            "zhihang_involved",
            "zhihang_record_id",
            "zhihang_title",
            "zhihang_progress",
            "zhihang_source_app_token",
            "zhihang_source_table_id",
            "plan_month",
            "maintenance_total",
            "maintenance_no",
            "maintenance_project",
            "maintenance_cycle",
            "start_time",
            "end_time",
            "location",
            "content",
            "reason",
            "impact",
            "progress",
            "repair_device",
            "repair_fault",
            "fault_type",
            "repair_mode",
            "discovery",
            "symptom",
            "solution",
            "fault_time",
            "expected_time",
            "text",
        ):
            value = incoming.get(field)
            if value not in (None, ""):
                item[field] = value
        if incoming.get("started_at") and not item.get("started_at"):
            item["started_at"] = incoming["started_at"]
        if incoming.get("started_at"):
            item["started_date"] = self._date_part(incoming.get("started_at"))
        if incoming.get("last_updated_at"):
            item["last_updated_at"] = incoming["last_updated_at"]
        if incoming.get("ended_at"):
            item["ended_at"] = incoming["ended_at"]
            item["completed_date"] = self._date_part(incoming.get("ended_at"))
        if action == "end":
            item["status"] = "已结束"
            item["ended_at"] = str(incoming.get("ended_at") or now)
            item["completed_date"] = self._date_part(item.get("ended_at"))
        elif action in {"start", "update"}:
            if item.get("status") != "已结束":
                item["status"] = "进行中"
        elif incoming.get("status"):
            item["status"] = incoming["status"]
        for incoming_action in incoming.get("actions") or []:
            if not isinstance(incoming_action, dict):
                continue
            job_id = str(incoming_action.get("job_id") or "").strip()
            actions = item.setdefault("actions", [])
            if job_id and any(str(existing.get("job_id") or "").strip() == job_id for existing in actions):
                continue
            actions.append(copy.deepcopy(incoming_action))
        item["updated_at"] = now
        payload["updated_at"] = now
        self._save_work_status_locked(payload, building=building, building_code=building_code)

    def _build_action_frontend_patch(
        self,
        job: dict[str, Any],
        *,
        success: bool,
        message: str = "",
        record_id: str = "",
        active_item_id: str = "",
        target_selection: dict[str, Any] | None = None,
    ) -> dict[str, Any]:
        prepared = job.get("prepared") if isinstance(job.get("prepared"), dict) else {}
        request = job.get("request") if isinstance(job.get("request"), dict) else {}
        action = str(prepared.get("action") or request.get("action") or "").strip().lower()
        merged = {}
        merged.update(request)
        merged.update(prepared)
        if record_id:
            merged["target_record_id"] = str(record_id or "").strip()
            merged["record_id"] = str(record_id or "").strip()
        if active_item_id:
            merged["active_item_id"] = str(active_item_id or "").strip()
        identity = normalize_notice_identity_payload(merged, action=action)
        target_record_id = str(
            record_id
            or identity.get("target_record_id")
            or job.get("target_record_id")
            or ""
        ).strip()
        patch = {
            "kind": "notice_action_result",
            "status": "success" if success else "failed",
            "action": action,
            "work_type": str(
                prepared.get("work_type")
                or request.get("work_type")
                or WORK_TYPE_MAINTENANCE
            ).strip(),
            "notice_type": str(prepared.get("notice_type") or request.get("notice_type") or "").strip(),
            "active_item_id": str(
                active_item_id
                or identity.get("active_item_id")
                or job.get("active_item_id")
                or ""
            ).strip(),
            "source_record_id": canonical_source_record_id(identity),
            "target_record_id": target_record_id,
            "record_id": target_record_id,
            "title": str(prepared.get("title") or request.get("title") or "").strip(),
            "building": str(prepared.get("building") or request.get("building") or "").strip(),
            "specialty": str(prepared.get("specialty") or request.get("specialty") or "").strip(),
            "maintenance_cycle": str(
                prepared.get("maintenance_cycle")
                or request.get("maintenance_cycle")
                or ""
            ).strip(),
            "status_text": str(prepared.get("status") or request.get("status") or "").strip(),
            "site_photo_count": str(
                prepared.get("site_photo_count")
                or request.get("site_photo_count")
                or "0"
            ).strip(),
            "message": str(message or "").strip(),
        }
        if not success and isinstance(target_selection, dict):
            patch["target_record_missing"] = bool(target_selection.get("target_record_missing"))
            patch["needs_target_selection"] = bool(target_selection.get("needs_target_selection"))
            patch["target_selection_message"] = str(target_selection.get("message") or "").strip()
            candidates = target_selection.get("target_candidates")
            if isinstance(candidates, list):
                patch["target_candidates"] = copy.deepcopy(candidates[:20])
        return patch

    def mark_action_upload_result(
        self,
        job_id: str,
        *,
        success: bool,
        message: str = "",
        record_id: str = "",
        active_item_id: str = "",
        target_selection: dict[str, Any] | None = None,
    ) -> None:
        job_id = str(job_id or "").strip()
        if not job_id:
            return
        job = self.get_job(job_id) or {}
        prepared = job.get("prepared") if isinstance(job.get("prepared"), dict) else {}
        notice_text = str(job.get("notice_text") or prepared.get("text") or "").strip()
        message_error = str(job.get("message_error") or "").strip()
        message_warning = str(job.get("message_warning") or "").strip()
        if success and message_error:
            message_warning = (
                f"多维上传成功；个人消息发送失败，可复制通告文本：{message_error}"
            )
            message = message_warning
        phase = "success" if success else "failed"
        frontend_patch = self._build_action_frontend_patch(
            job,
            success=success,
            message=message,
            record_id=record_id,
            active_item_id=active_item_id,
            target_selection=target_selection,
        )
        patch = {
            "phase": phase,
            "error": "" if success else str(message or "上传失败"),
            "upload_message": str(message or ""),
            "record_id": str(record_id or ""),
            "frontend_patch": frontend_patch,
        }
        if success and str(record_id or "").strip():
            patch["target_record_id"] = str(record_id or "").strip()
        if not success and isinstance(target_selection, dict):
            candidates = target_selection.get("target_candidates")
            if isinstance(candidates, list):
                patch["target_candidates"] = candidates[:50]
            patch["target_record_missing"] = bool(target_selection.get("target_record_missing"))
            patch["needs_target_selection"] = bool(target_selection.get("needs_target_selection"))
            patch["target_selection_message"] = str(target_selection.get("message") or "")
        if notice_text:
            patch["notice_text"] = notice_text
            patch["copy_text"] = notice_text
        if message_warning:
            patch["message_warning"] = message_warning
        if message_error:
            patch["message_error"] = message_error
            patch["message_failed"] = True
            patch["message_failed_continue"] = True
        if active_item_id:
            patch["active_item_id"] = str(active_item_id or "")
        if success:
            self.mark_job(job_id, _persist=False, **patch)
            self._record_successful_action(job_id, record_id=record_id, active_item_id=active_item_id)
            with self._jobs_lock:
                self._compact_completed_job_locked(job_id)
        else:
            self.mark_job(job_id, **patch)

    def _record_successful_action(
        self, job_id: str, *, record_id: str = "", active_item_id: str = ""
    ) -> None:
        job = self.get_job(job_id) or {}
        prepared = job.get("prepared") or {}
        if not isinstance(prepared, dict):
            prepared = {}
        action = str(prepared.get("action") or "").strip().lower()
        if action not in {"start", "update", "end"}:
            return
        prepared_identity = normalize_notice_identity_payload(dict(prepared), action=action)
        now = dt.datetime.now().strftime("%Y-%m-%d %H:%M")
        source_record_id = canonical_source_record_id(prepared_identity)
        title = str(prepared.get("title") or "").strip()
        reason = str(prepared.get("reason") or "").strip()
        building = str(prepared.get("building") or "").strip()
        work_type = str(prepared.get("work_type") or WORK_TYPE_MAINTENANCE).strip()
        key = self._summary_key(
            source_record_id=source_record_id,
            title=title,
            building=building,
            reason=reason,
            work_type=work_type,
        )
        fallback_key = self._summary_key(
            title=title, building=building, reason=reason, work_type=work_type
        )
        active_item_id = str(active_item_id or prepared.get("active_item_id") or job.get("active_item_id") or "").strip()
        target_record_id = str(
            record_id
            or job.get("target_record_id")
            or prepared_identity.get("target_record_id")
            or ""
        ).strip()
        with self._summary_lock:
            payload = self._load_day_summary_locked()
            items = payload.setdefault("items", [])
            item = self._find_summary_item(
                items,
                key=key,
                active_item_id=active_item_id,
                title=title,
                building=building,
                reason=reason,
                work_type=work_type,
            )
            if item is None:
                item = {
                    "key": key or fallback_key or uuid.uuid4().hex,
                    "fallback_key": fallback_key,
                    "work_type": work_type,
                    "notice_type": str(prepared.get("notice_type") or ""),
                    "source_app_token": str(prepared.get("source_app_token") or ""),
                    "source_table_id": str(prepared.get("source_table_id") or ""),
                    "source_record_id": source_record_id,
                    "target_record_id": str(
                        target_record_id
                        or prepared_identity.get("target_record_id")
                        or ""
                    ),
                    "active_item_id": active_item_id,
                    "title": title,
                    "building": building,
                    "building_code": str(prepared.get("building_code") or ""),
                    "building_codes": list(prepared.get("building_codes") or []),
                    "specialty": str(prepared.get("specialty") or ""),
                    "status": "进行中",
                    "actions": [],
                }
                items.append(item)
            item["fallback_key"] = item.get("fallback_key") or fallback_key
            item["work_type"] = self._item_work_type(item) if item.get("work_type") else work_type
            for field in (
                "notice_type",
                "source_app_token",
                "source_table_id",
                "building_code",
                "building_codes",
                "target_record_id",
                "level",
                "source_progress",
            ):
                value = prepared.get(field)
                if value not in (None, ""):
                    item[field] = value
            if target_record_id:
                item["target_record_id"] = target_record_id
                item["record_id"] = target_record_id
            if source_record_id and not item.get("source_record_id"):
                item["source_record_id"] = source_record_id
            if active_item_id:
                item["active_item_id"] = active_item_id
            for field in (
                "title",
                "building",
                "building_code",
                "specialty",
                "start_time",
                "end_time",
                "location",
                "content",
                "reason",
                "impact",
                "progress",
                "text",
                "plan_month",
                "maintenance_total",
                "maintenance_no",
                "maintenance_project",
                "maintenance_cycle",
                "level",
                "source_progress",
                "zhihang_involved",
                "zhihang_record_id",
                "zhihang_title",
                "zhihang_progress",
                "zhihang_source_app_token",
                "zhihang_source_table_id",
                "sync_maintenance_target",
                "paired_maintenance_target_record_id",
                "paired_maintenance_original_title",
                "paired_maintenance_actual_start_time",
                "paired_upload_status",
                "paired_upload_warning",
                "repair_device",
                "repair_fault",
                "fault_type",
                "repair_mode",
                "discovery",
                "symptom",
                "solution",
                "fault_time",
                "expected_time",
            ):
                value = prepared.get(field)
                if value not in (None, ""):
                    item[field] = value
            if action == "start":
                item["started_at"] = item.get("started_at") or now
                item["status"] = "进行中"
            elif action == "update":
                item["last_updated_at"] = now
                if item.get("status") != "已结束":
                    item["status"] = "进行中"
            elif action == "end":
                item["ended_at"] = now
                item["status"] = "已结束"
            action_label = {"start": "开始", "update": "更新", "end": "结束"}[action]
            item.setdefault("actions", []).append(
                {
                    "action": action,
                    "label": action_label,
                    "time": now,
                    "job_id": job_id,
                    "record_id": target_record_id,
                    "progress": str(prepared.get("progress") or ""),
                    "text": str(prepared.get("text") or ""),
                }
            )
            payload["updated_at"] = now
            self._save_day_summary_locked(payload)
            self._upsert_work_status_item_locked(item, action=action, now=now)
            if action == "end":
                try:
                    self._state_store.delete_qt_active_item(
                        active_item_id=active_item_id,
                        record_id=target_record_id,
                    )
                except Exception:
                    pass
        try:
            identity_payload = dict(prepared)
            if active_item_id:
                identity_payload["active_item_id"] = active_item_id
            if source_record_id:
                identity_payload["source_record_id"] = source_record_id
            if target_record_id:
                identity_payload["target_record_id"] = target_record_id
                identity_payload["record_id"] = target_record_id
            identity_payload["status"] = "已结束" if action == "end" else "进行中"
            self._state_store.upsert_notice_identity(
                identity_payload,
                origin=str(prepared.get("origin") or "action_success"),
            )
        except Exception:
            pass
        self._schedule_repair_link_task_after_success(
            prepared,
            action=action,
            target_record_id=target_record_id,
            job_id=job_id,
        )

    def _schedule_repair_link_task_after_success(
        self,
        prepared: dict[str, Any],
        *,
        action: str,
        target_record_id: str,
        job_id: str = "",
    ) -> None:
        if action != "start":
            return
        if str((prepared or {}).get("work_type") or "").strip() != WORK_TYPE_REPAIR:
            return
        if self._truthy_flag((prepared or {}).get("manual")):
            return
        source_record_id = canonical_source_record_id(prepared)
        target_record_id = str(target_record_id or "").strip()
        if not source_record_id or not target_record_id:
            return
        source_app_token = str(
            (prepared or {}).get("source_app_token") or REPAIR_SOURCE_APP_TOKEN
        ).strip()
        source_table_id = str(
            (prepared or {}).get("source_table_id") or REPAIR_SOURCE_TABLE_ID
        ).strip()
        task_key = f"repair_link:{source_record_id}:{target_record_id}"
        now = time.time()
        self._state_store.upsert_repair_link_task(
            {
                "task_key": task_key,
                "status": "pending",
                "source_app_token": source_app_token,
                "source_table_id": source_table_id,
                "source_record_id": source_record_id,
                "sync_app_token": REPAIR_SOURCE_APP_TOKEN,
                "sync_table_id": REPAIR_SYNC_TABLE_ID,
                "target_app_token": str(config.app_token or "").strip(),
                "target_table_id": str(config.table_id_overhaul or "").strip(),
                "target_record_id": target_record_id,
                "link_field_name": REPAIR_LINK_FIELD_NAME,
                "due_at": now + REPAIR_LINK_DELAY_SECONDS,
                "attempts": 0,
                "max_attempts": REPAIR_LINK_MAX_ATTEMPTS,
                "last_error": "",
                "job_id": str(job_id or ""),
                "created_at": now,
            }
        )

    @staticmethod
    def _repair_link_retry_delay_seconds(attempts: int) -> int:
        attempts = max(1, int(attempts or 1))
        if attempts <= REPAIR_LINK_FAST_RETRY_ATTEMPTS:
            return REPAIR_LINK_RETRY_SECONDS
        return REPAIR_LINK_RETRY_SLOW_SECONDS

    def process_due_repair_link_tasks(self, *, limit: int = 3) -> dict[str, int]:
        tasks = self._state_store.claim_due_repair_link_tasks(
            limit=limit,
            lease_seconds=5 * 60,
        )
        stats = {"checked": 0, "linked": 0, "rescheduled": 0, "failed": 0}
        for task in tasks:
            stats["checked"] += 1
            task_key = str(task.get("task_key") or "").strip()
            attempts = int(task.get("attempts") or 0) + 1
            max_attempts = int(task.get("max_attempts") or REPAIR_LINK_MAX_ATTEMPTS)
            try:
                source_app_token = str(
                    task.get("source_app_token") or REPAIR_SOURCE_APP_TOKEN
                ).strip()
                source_table_id = str(
                    task.get("source_table_id") or REPAIR_SOURCE_TABLE_ID
                ).strip()
                source_record_id = str(task.get("source_record_id") or "").strip()
                target_record_id = str(task.get("target_record_id") or "").strip()
                sync_app_token = str(
                    task.get("sync_app_token") or REPAIR_SOURCE_APP_TOKEN
                ).strip()
                sync_table_id = str(
                    task.get("sync_table_id") or REPAIR_SYNC_TABLE_ID
                ).strip()
                sync_record_id = str(task.get("sync_record_id") or "").strip()
                if not sync_record_id:
                    sync_record_id = self._find_repair_sync_record_id(
                        target_record_id=target_record_id,
                        sync_app_token=sync_app_token,
                        sync_table_id=sync_table_id,
                    )
                if not sync_record_id:
                    raise PortalError(
                        f"同步表尚未出现目标检修记录 {target_record_id}"
                    )
                linked_ids = self._source_repair_link_record_ids(
                    source_app_token=source_app_token,
                    source_table_id=source_table_id,
                    source_record_id=source_record_id,
                )
                if sync_record_id not in linked_ids:
                    self._write_source_repair_link(
                        source_app_token=source_app_token,
                        source_table_id=source_table_id,
                        source_record_id=source_record_id,
                        sync_record_id=sync_record_id,
                    )
                self._state_store.update_repair_link_task(
                    task_key,
                    {
                        "status": "linked",
                        "attempts": attempts,
                        "sync_record_id": sync_record_id,
                        "last_error": "",
                    },
                )
                stats["linked"] += 1
            except Exception as exc:
                next_status = "failed" if attempts >= max_attempts else "pending"
                retry_delay = self._repair_link_retry_delay_seconds(attempts)
                now = time.time()
                self._state_store.update_repair_link_task(
                    task_key,
                    {
                        "status": next_status,
                        "attempts": attempts,
                        "due_at": now + retry_delay,
                        "last_checked_at": now,
                        "next_retry_seconds": retry_delay if next_status == "pending" else 0,
                        "last_error": str(exc),
                    },
                )
                if next_status == "failed":
                    stats["failed"] += 1
                else:
                    stats["rescheduled"] += 1
        return stats

    def get_daily_summary(
        self, *, scope: str = "ALL", ongoing_items: list[dict[str, Any]] | None = None
    ) -> dict[str, Any]:
        scope = self._normalize_scope(scope)
        with self._summary_lock:
            payload = self._load_day_summary_locked()
            raw_items = [item for item in payload.get("items") or [] if isinstance(item, dict)]
        items = [
            copy.deepcopy(item)
            for item in raw_items
            if self._scope_matches_item(scope, item)
        ]
        ongoing_count = len(ongoing_items or [])
        completed_count = sum(1 for item in items if item.get("status") == "已结束")
        started_count = sum(1 for item in items if item.get("started_at"))
        updated_count = sum(
            1
            for item in items
            if any(str(action.get("action") or "") == "update" for action in item.get("actions") or [])
        )
        items.sort(key=lambda item: str(item.get("ended_at") or item.get("last_updated_at") or item.get("started_at") or ""), reverse=True)
        return {
            "date": payload.get("date") or dt.datetime.now().strftime("%Y-%m-%d"),
            "items": items,
            "stats": {
                "started": started_count,
                "updated": updated_count,
                "ended": completed_count,
                "ongoing": ongoing_count,
            },
        }

    @staticmethod
    def _work_type_counts(items: list[dict[str, Any]]) -> dict[str, int]:
        counts = {
            WORK_TYPE_MAINTENANCE: 0,
            WORK_TYPE_CHANGE: 0,
            WORK_TYPE_REPAIR: 0,
            WORK_TYPE_POWER: 0,
            WORK_TYPE_POLLING: 0,
            WORK_TYPE_ADJUST: 0,
        }
        for item in items or []:
            work_type = MaintenancePortalService._item_work_type(item)
            if work_type not in counts:
                continue
            counts[work_type] += 1
        return counts

    @staticmethod
    def _item_work_type(item: dict[str, Any] | None) -> str:
        item = item if isinstance(item, dict) else {}
        work_type = str(item.get("work_type") or item.get("lan_work_type") or "").strip()
        if work_type in {
            WORK_TYPE_MAINTENANCE,
            WORK_TYPE_CHANGE,
            WORK_TYPE_REPAIR,
            WORK_TYPE_POWER,
            WORK_TYPE_POLLING,
            WORK_TYPE_ADJUST,
        }:
            return work_type
        notice_type = str(item.get("notice_type") or "").strip()
        mapped = WORK_TYPE_BY_NOTICE_TYPE.get(notice_type, "")
        if mapped:
            return mapped
        text = "\n".join(
            str(item.get(key) or "").strip()
            for key in ("text", "content", "title", "name")
            if str(item.get(key) or "").strip()
        )
        if re.search(r"上电通告|下电通告|上下电通告", text):
            return WORK_TYPE_POWER
        if re.search(r"设备轮巡|轮巡通告", text):
            return WORK_TYPE_POLLING
        if re.search(r"设备调整|调整通告", text):
            return WORK_TYPE_ADJUST
        if re.search(r"设备检修|检修通告", text):
            return WORK_TYPE_REPAIR
        if re.search(r"变更通告", text):
            return WORK_TYPE_CHANGE
        return WORK_TYPE_MAINTENANCE

    def get_scope_overview(
        self,
        *,
        ongoing_items: list[dict[str, Any]] | None = None,
        scopes: list[str] | tuple[str, ...] | set[str] | None = None,
        include_prepared: bool = False,
    ) -> dict[str, Any]:
        self.ensure_snapshot_loaded()
        default_month = RECENT_MONTH_FILTER_LABEL
        overview: dict[str, dict[str, Any]] = {}
        prepared_workbenches: dict[str, dict[str, Any]] = {}
        has_scope_filter = scopes is not None
        requested_scopes = {
            self._normalize_scope(scope)
            for scope in (scopes or [])
            if str(scope or "").strip()
        }
        for option in SCOPE_OPTIONS:
            scope = self._normalize_scope(option.get("value"))
            if has_scope_filter and scope not in requested_scopes:
                continue
            records = self._workbench_records(month=default_month, scope=scope)
            merged_ongoing = self._project_ongoing_items(scope, ongoing_items or [])
            daily_summary = self.get_daily_summary(
                scope=scope, ongoing_items=merged_ongoing
            )
            summary_by_record = self._work_status_by_records(
                records,
                scope=scope,
                daily_items=daily_summary.get("items") or [],
            )
            ongoing_sources = {
                (
                    self._item_work_type(item),
                    str(item.get("source_record_id") or "").strip(),
                )
                for item in merged_ongoing
                if str(item.get("source_record_id") or "").strip()
            }
            pending_records = []
            for record in records:
                record_id = str(record.get("record_id") or "").strip()
                work_type = self._record_work_type(record)
                summary = summary_by_record.get(record_id) or {}
                if summary.get("started_at"):
                    continue
                if (work_type, record_id) in ongoing_sources:
                    continue
                pending_records.append(record)
            pending_counts = self._work_type_counts(pending_records)
            ongoing_counts = self._work_type_counts(merged_ongoing)
            ongoing_titles: list[dict[str, str]] = []
            for index, item in enumerate(merged_ongoing[:30]):
                fields = item.get("fields") if isinstance(item.get("fields"), dict) else {}
                title = str(
                    item.get("title")
                    or item.get("name")
                    or fields.get("标题")
                    or fields.get("名称")
                    or item.get("content")
                    or ""
                ).strip()
                if not title:
                    continue
                item_work_type = self._item_work_type(item)
                ongoing_titles.append(
                    {
                        "key": str(item.get("active_item_id") or item.get("record_id") or index),
                        "work_type": item_work_type,
                        "title": title,
                    }
                )
            stats = daily_summary.get("stats") or {}
            overview[scope] = {
                "scope": scope,
                "scope_label": self._scope_label(scope),
                "maintenance_pending": pending_counts[WORK_TYPE_MAINTENANCE],
                "change_pending": pending_counts[WORK_TYPE_CHANGE],
                "repair_pending": pending_counts[WORK_TYPE_REPAIR],
                "power_pending": pending_counts[WORK_TYPE_POWER],
                "polling_pending": pending_counts[WORK_TYPE_POLLING],
                "adjust_pending": pending_counts[WORK_TYPE_ADJUST],
                "maintenance_ongoing": ongoing_counts[WORK_TYPE_MAINTENANCE],
                "change_ongoing": ongoing_counts[WORK_TYPE_CHANGE],
                "repair_ongoing": ongoing_counts[WORK_TYPE_REPAIR],
                "power_ongoing": ongoing_counts[WORK_TYPE_POWER],
                "polling_ongoing": ongoing_counts[WORK_TYPE_POLLING],
                "adjust_ongoing": ongoing_counts[WORK_TYPE_ADJUST],
                "ongoing_title_count": len(merged_ongoing),
                "ongoing_titles": ongoing_titles,
                "closed_today": int(stats.get("ended") or 0),
            }
            if include_prepared:
                prepared_workbenches[scope] = self.get_bootstrap(
                    scope=scope, ongoing_items=ongoing_items or []
                )
        payload = {
            "scope_options": SCOPE_OPTIONS,
            "scopes": overview,
            "last_loaded_at": self._last_loaded_at,
            "payload_version": self._payload_version(
                scope="overview",
                records=[],
                ongoing_items=ongoing_items or [],
            ),
            "source_snapshot_ready": self._source_snapshot_exists("ALL"),
            "source_cache_ttl_seconds": self._source_cache_ttl_seconds(),
            "warnings": self._current_load_warnings(),
        }
        if include_prepared:
            payload["prepared_workbenches"] = prepared_workbenches
        return payload

    def get_handover_links(self) -> dict[str, Any]:
        options = self._handover_scope_options()
        with self._handover_lock:
            payload = self._load_handover_payload_locked()
        links = payload.get("links") if isinstance(payload, dict) else {}
        if not isinstance(links, dict):
            links = {}
        normalized_links = {
            option["value"]: str(links.get(option["value"]) or "").strip()
            for option in options
        }
        return {
            "scope_options": options,
            "links": normalized_links,
            "updated_at": str(payload.get("updated_at") or "") if isinstance(payload, dict) else "",
        }

    def _load_handover_payload_locked(self) -> dict[str, Any]:
        stored = self._state_store.get_handover_payload()
        if isinstance(stored, dict):
            return stored
        try:
            with self._handover_links_path.open("r", encoding="utf-8") as fp:
                payload = json.load(fp)
        except FileNotFoundError:
            return {}
        except Exception:
            return {}
        if not isinstance(payload, dict):
            return {}
        self._state_store.put_handover_payload(payload)
        return payload

    def _save_handover_payload_locked(self, payload: dict[str, Any]) -> None:
        self._state_store.put_handover_payload(payload)

    @staticmethod
    def _hash_handover_password(password: str, salt: str) -> str:
        raw = f"{salt}:{password}".encode("utf-8")
        return hashlib.sha256(raw).hexdigest()

    @staticmethod
    def _new_handover_password_salt() -> str:
        return secrets.token_hex(16)

    def _set_handover_password_locked(self, payload: dict[str, Any], password: str) -> None:
        salt = self._new_handover_password_salt()
        payload["password_salt"] = salt
        payload["password_hash"] = self._hash_handover_password(password, salt)
        payload["password_updated_at"] = dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    def verify_handover_settings_password(self, password: str) -> bool:
        password = str(password or "")
        with self._handover_lock:
            payload = self._load_handover_payload_locked()
        salt = str(payload.get("password_salt") or "")
        password_hash = str(payload.get("password_hash") or "")
        if salt and password_hash:
            digest = self._hash_handover_password(password, salt)
            return hmac.compare_digest(digest, password_hash)
        expected = str(getattr(config, "lan_handover_settings_password", "") or "").strip()
        if not expected:
            return False
        return hmac.compare_digest(password, expected)

    def save_handover_links(self, links: dict[str, Any], *, password: str = "") -> dict[str, Any]:
        if not self.verify_handover_settings_password(password):
            raise PortalError("设置密码错误。")
        if not isinstance(links, dict):
            raise PortalError("交接班链接配置格式错误。")
        options = self._handover_scope_options()
        allowed = {option["value"] for option in options}
        normalized_links: dict[str, str] = {}
        for code in allowed:
            url = str(links.get(code) or "").strip()
            if url and not re.match(r"^https?://", url, flags=re.IGNORECASE):
                raise PortalError(f"{self._scope_label(code)} 链接必须以 http:// 或 https:// 开头。")
            normalized_links[code] = url
        with self._handover_lock:
            payload = self._load_handover_payload_locked()
            payload["updated_at"] = dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            payload["links"] = normalized_links
            self._save_handover_payload_locked(payload)
        return self.get_handover_links()

    def _handover_reset_open_ids(self) -> list[str]:
        configured = getattr(config, "lan_handover_reset_open_ids", None)
        recipients: list[str]
        if isinstance(configured, str):
            recipients = re.split(r"[,，;\s]+", configured)
        elif isinstance(configured, (list, tuple, set)):
            recipients = [str(item or "") for item in configured]
        else:
            recipients = [LI_SHILONG_OPEN_ID, MA_JINYU_OPEN_ID]
        recipients = list(dict.fromkeys([str(item or "").strip() for item in recipients]))
        recipients = [item for item in recipients if item]
        if len(recipients) < 2:
            raise PortalError("未配置李世龙和马进宇两个 openid，无法发送改密验证码。")
        return recipients

    def _clear_expired_handover_reset_locked(self) -> None:
        reset = self._handover_password_reset or {}
        expires_at = float(reset.get("expires_at_ts") or 0)
        if expires_at and time.time() >= expires_at:
            self._handover_password_reset = None

    def request_handover_password_reset(self) -> dict[str, Any]:
        recipients = self._handover_reset_open_ids()
        with self._handover_reset_lock:
            self._clear_expired_handover_reset_locked()
            if self._handover_password_reset:
                raise PortalError("已有密码修改流程进行中，请稍后再试。")
            code = f"{secrets.randbelow(1000000):06d}"
            reset_id = secrets.token_urlsafe(18)
            now = dt.datetime.now()
            expires_at = now + dt.timedelta(seconds=HANDOVER_PASSWORD_RESET_TTL_SECONDS)
            self._handover_password_reset = {
                "reset_id": reset_id,
                "code": code,
                "created_at": now.strftime("%Y-%m-%d %H:%M:%S"),
                "expires_at": expires_at.strftime("%Y-%m-%d %H:%M:%S"),
                "expires_at_ts": time.time() + HANDOVER_PASSWORD_RESET_TTL_SECONDS,
                "attempts": 0,
            }
        text = (
            "南通基地-运维灯塔工作台交接班链接设置密码重置验证码：\n"
            f"{code}\n"
            f"有效期至：{expires_at.strftime('%Y-%m-%d %H:%M:%S')}\n"
            "如果不是本人操作，请忽略。"
        )
        guard = external_real_write_guard()
        if guard["mock_external"]:
            ok, message = True, "mock external send skipped"
        elif not guard["real_write_allowed"]:
            ok, message = False, str(guard["reason"] or "真实外部写入未确认。")
        else:
            ok, message, _ = send_text_to_open_ids(text, recipients)
        if not ok:
            with self._handover_reset_lock:
                reset = self._handover_password_reset or {}
                if reset.get("reset_id") == reset_id:
                    self._handover_password_reset = None
            raise PortalError(f"验证码发送失败: {message}")
        return {
            "reset_id": reset_id,
            "expires_at": expires_at.strftime("%Y-%m-%d %H:%M:%S"),
            "recipients_count": len(recipients),
        }

    def reset_handover_password_with_code(
        self, *, reset_id: str, code: str, new_password: str
    ) -> dict[str, Any]:
        reset_id = str(reset_id or "").strip()
        code = str(code or "").strip()
        new_password = str(new_password or "")
        if len(new_password) < 6:
            raise PortalError("新密码长度不能少于 6 位。")
        with self._handover_reset_lock:
            self._clear_expired_handover_reset_locked()
            reset = self._handover_password_reset
            if not reset:
                raise PortalError("当前没有有效的密码修改流程，请重新获取验证码。")
            if str(reset.get("reset_id") or "") != reset_id:
                raise PortalError("密码修改会话不匹配，请重新获取验证码。")
            attempts = int(reset.get("attempts") or 0) + 1
            reset["attempts"] = attempts
            if attempts > HANDOVER_PASSWORD_RESET_MAX_ATTEMPTS:
                self._handover_password_reset = None
                raise PortalError("验证码错误次数过多，请重新获取验证码。")
            if not hmac.compare_digest(str(reset.get("code") or ""), code):
                raise PortalError("验证码错误。")
            with self._handover_lock:
                payload = self._load_handover_payload_locked()
                self._set_handover_password_locked(payload, new_password)
                self._save_handover_payload_locked(payload)
            self._handover_password_reset = None
        return {"password_updated": True}

    @staticmethod
    def _summary_item_work_type(item: dict[str, Any]) -> str:
        return MaintenancePortalService._item_work_type(item)

    @staticmethod
    def _summary_item_sort_key(item: dict[str, Any]) -> str:
        return str(
            item.get("ended_at")
            or item.get("last_updated_at")
            or item.get("updated_at")
            or item.get("started_at")
            or ""
        )

    def get_history_summary(
        self,
        *,
        scope: str = "ALL",
        month: str = "",
        work_type: str = "all",
    ) -> dict[str, Any]:
        scope = self._normalize_scope(scope)
        month = str(month or dt.datetime.now().strftime("%Y-%m")).strip()
        if not re.fullmatch(r"\d{4}-\d{2}", month):
            raise PortalError("历史月份格式必须是 YYYY-MM。")
        work_type = str(work_type or "all").strip()
        if work_type not in {
            "all",
            WORK_TYPE_MAINTENANCE,
            WORK_TYPE_CHANGE,
            WORK_TYPE_REPAIR,
        }:
            raise PortalError("历史通告类型必须是 all/maintenance/change/repair。")

        days: list[dict[str, Any]] = []
        with self._summary_lock:
            for payload in self._iter_day_summary_payloads_locked(month=f"{month}-"):
                day = str(payload.get("date") or "")
                items = []
                for raw_item in payload.get("items") or []:
                    if not isinstance(raw_item, dict):
                        continue
                    item = copy.deepcopy(raw_item)
                    item_work_type = self._summary_item_work_type(item)
                    if work_type != "all" and item_work_type != work_type:
                        continue
                    item["work_type"] = item_work_type
                    if not self._scope_matches_item(scope, item):
                        continue
                    items.append(item)
                if not items:
                    continue
                items.sort(key=self._summary_item_sort_key, reverse=True)
                days.append(
                    {
                        "date": day,
                        "items": items,
                        "stats": {
                            "started": sum(1 for item in items if item.get("started_at")),
                            "updated": sum(
                                1
                                for item in items
                                if any(
                                    str(action.get("action") or "") == "update"
                                    for action in item.get("actions") or []
                                    if isinstance(action, dict)
                                )
                            ),
                            "ended": sum(
                                1 for item in items if item.get("status") == "已结束"
                            ),
                            "ongoing": sum(
                                1
                                for item in items
                                if item.get("started_at")
                                and item.get("status") != "已结束"
                            ),
                        },
                    }
                )
        days.sort(key=lambda item: str(item.get("date") or ""), reverse=True)
        return {
            "scope": scope,
            "scope_label": self._scope_label(scope),
            "month": month,
            "work_type": work_type,
            "days": days,
            "last_loaded_at": self._last_loaded_at,
            "warnings": self._current_load_warnings(),
        }

    def _summary_by_record_id(self, summary_items: list[dict[str, Any]]) -> dict[str, dict[str, Any]]:
        result: dict[str, dict[str, Any]] = {}
        for item in summary_items:
            record_id = str(item.get("source_record_id") or "").strip()
            if record_id:
                result[record_id] = item
        return result

    def _backfill_work_status_from_daily_summaries_locked(self) -> None:
        if self._work_status_backfilled:
            return
        self._work_status_backfilled = True
        for payload in self._iter_day_summary_payloads_locked():
            day = str(payload.get("date") or "")
            for raw_item in payload.get("items") or []:
                if not isinstance(raw_item, dict):
                    continue
                item = copy.deepcopy(raw_item)
                actions = [action for action in item.get("actions") or [] if isinstance(action, dict)]
                last_action = str(actions[-1].get("action") or "").strip().lower() if actions else ""
                if not item.get("started_at") and last_action == "start":
                    item["started_at"] = f"{day} 00:00"
                if item.get("status") == "已结束" and not item.get("ended_at"):
                    item["ended_at"] = f"{day} 00:00"
                self._upsert_work_status_item_locked(
                    item,
                    action=last_action,
                    now=str(payload.get("updated_at") or item.get("updated_at") or day),
                )

    def _load_work_status_items_locked(self, scope: str = "ALL") -> list[dict[str, Any]]:
        self._backfill_work_status_from_daily_summaries_locked()
        self._migrate_legacy_work_status_locked()
        scope = self._normalize_scope(scope)
        if hasattr(self._state_store, "list_document_meta"):
            documents_meta = self._state_store.list_document_meta(STATE_NS_WORK_STATUS)
        else:
            documents_meta = self._state_store.list_documents(STATE_NS_WORK_STATUS)
        signature = tuple(
            sorted(
                (
                    str(document.get("key") or ""),
                    int(float(document.get("updated_at") or 0) * 1000000),
                )
                for document in documents_meta
            )
        )
        if (
            self._work_status_cache_signature == signature
            and self._work_status_cache_items is not None
        ):
            all_items = self._work_status_cache_items
        else:
            documents = self._state_store.list_documents(STATE_NS_WORK_STATUS)
            all_items: list[dict[str, Any]] = []
            for document in documents:
                payload = document.get("payload")
                if not isinstance(payload, dict):
                    continue
                for item in payload.get("items") or []:
                    if isinstance(item, dict):
                        all_items.append(copy.deepcopy(item))
            self._work_status_cache_signature = signature
            self._work_status_cache_items = all_items
        return [
            copy.deepcopy(item)
            for item in all_items
            if self._scope_matches_item(scope, item)
        ]

    def _work_status_by_records(
        self,
        records: list[dict[str, Any]],
        *,
        scope: str = "ALL",
        daily_items: list[dict[str, Any]] | None = None,
    ) -> dict[str, dict[str, Any]]:
        with self._summary_lock:
            status_items = self._load_work_status_items_locked(scope)
        by_source = {
            (
                self._item_work_type(item),
                str(item.get("source_record_id") or "").strip(),
            ): item
            for item in status_items
            if str(item.get("source_record_id") or "").strip()
        }
        by_fallback: dict[str, dict[str, Any]] = {}
        for item in status_items:
            for key_name in ("work_fallback_key", "fallback_key"):
                fallback_key = str(item.get(key_name) or "").strip()
                if fallback_key and fallback_key not in by_fallback:
                    by_fallback[fallback_key] = item
        for item in daily_items or []:
            if not isinstance(item, dict):
                continue
            source_record_id = str(item.get("source_record_id") or "").strip()
            item_work_type = self._item_work_type(item)
            source_key = (item_work_type, source_record_id)
            if source_record_id and source_key not in by_source:
                by_source[source_key] = copy.deepcopy(item)
            fallback_key = str(item.get("fallback_key") or "").strip()
            if fallback_key and fallback_key not in by_fallback:
                by_fallback[fallback_key] = copy.deepcopy(item)
            work_fallback_key = str(item.get("work_fallback_key") or "").strip()
            if work_fallback_key and work_fallback_key not in by_fallback:
                by_fallback[work_fallback_key] = copy.deepcopy(item)
        result: dict[str, dict[str, Any]] = {}
        for record in records:
            record_id = str(record.get("record_id") or "").strip()
            if not record_id:
                continue
            work_type = self._record_work_type(record)
            item = by_source.get((work_type, record_id))
            if item is None and work_type == WORK_TYPE_MAINTENANCE:
                item = by_source.get(("", record_id))
            if item is None:
                item = by_fallback.get(self._record_fallback_key(record))
            if item is None:
                item = by_fallback.get(self._record_legacy_summary_key(record))
            if item is not None:
                result[record_id] = copy.deepcopy(item)
        return result

    def _work_status_identity_keys(self, item: dict[str, Any]) -> set[str]:
        if not isinstance(item, dict):
            return set()
        item = normalize_notice_identity_payload(item)
        work_type = self._item_work_type(item)
        keys: set[str] = set()
        for kind, field_names in {
            "active": ("active_item_id",),
            "source": ("source_record_id",),
            "target": ("target_record_id",),
            "key": ("key", "fallback_key", "work_fallback_key"),
        }.items():
            for field_name in field_names:
                value = str(item.get(field_name) or "").strip()
                if value:
                    keys.add(f"{work_type}:{kind}:{value}")
        return keys

    @staticmethod
    def _is_completed_work_status_item(item: dict[str, Any]) -> bool:
        return str(item.get("status") or "").strip() == "已结束" or bool(
            item.get("ended_at")
        )

    def _target_record_exists_for_status_item(
        self,
        item: dict[str, Any],
        target_cache: dict[tuple[str, str], set[str] | None],
    ) -> bool | None:
        work_type = self._item_work_type(item)
        notice_type = str(
            item.get("notice_type")
            or NOTICE_TYPE_BY_WORK_TYPE.get(work_type)
            or NOTICE_TYPE_MAINTENANCE
        ).strip()
        target_record_id = canonical_target_record_id(item)
        if not target_record_id:
            return None
        table_id = str(config.get_table_id(notice_type) or "").strip()
        app_token = str(config.app_token or "").strip()
        if not app_token or not table_id:
            return None
        cache_key = (notice_type, work_type)
        if cache_key not in target_cache:
            try:
                records = self._target_records_for_notice_type(
                    notice_type, work_type, force_refresh=False
                )
            except Exception as exc:
                if not self._is_transient_feishu_error(exc):
                    warning = f"{notice_type}目标表孤儿状态校验失败: {exc}"
                    if warning not in self._load_warnings:
                        self._load_warnings.append(warning)
                target_cache[cache_key] = None
            else:
                target_cache[cache_key] = {
                    str(record.get("record_id") or "").strip()
                    for record in records
                    if str(record.get("record_id") or "").strip()
                }
        target_ids = target_cache.get(cache_key)
        if target_ids is None:
            return None
        return target_record_id in target_ids

    @staticmethod
    def _is_transient_feishu_error(exc: Exception) -> bool:
        text = str(exc or "")
        return "code=1254002" in text or "msg=Fail" in text

    def _is_orphan_started_item(
        self,
        item: dict[str, Any],
        *,
        ongoing_keys: set[str],
        target_cache: dict[tuple[str, str], set[str] | None],
    ) -> bool:
        if not isinstance(item, dict):
            return False
        if self._is_completed_work_status_item(item):
            return False
        if not item.get("started_at"):
            return False
        if self._work_status_identity_keys(item) & ongoing_keys:
            return False
        target_exists = self._target_record_exists_for_status_item(item, target_cache)
        return target_exists is False

    def reconcile_orphan_started_items(
        self, *, scope: str = "ALL", ongoing_items: list[dict[str, Any]] | None = None
    ) -> dict[str, Any]:
        scope = self._normalize_scope(scope)
        ongoing_keys: set[str] = set()
        for item in ongoing_items or []:
            if isinstance(item, dict) and self._scope_matches_item(scope, item):
                ongoing_keys.update(self._work_status_identity_keys(item))
        if not ongoing_keys:
            ongoing_keys = set()
        target_cache: dict[tuple[str, str], set[str] | None] = {}
        removed_keys: set[str] = set()
        removed_count = 0
        with self._summary_lock:
            self._backfill_work_status_from_daily_summaries_locked()
            self._migrate_legacy_work_status_locked()
            for document in self._state_store.list_documents(STATE_NS_WORK_STATUS):
                payload = document.get("payload")
                if not isinstance(payload, dict):
                    continue
                items = payload.get("items")
                if not isinstance(items, list):
                    continue
                kept: list[dict[str, Any]] = []
                changed = False
                for item in items:
                    if not isinstance(item, dict):
                        kept.append(item)
                        continue
                    if not self._scope_matches_item(scope, item):
                        kept.append(item)
                        continue
                    if self._is_orphan_started_item(
                        item, ongoing_keys=ongoing_keys, target_cache=target_cache
                    ):
                        removed_keys.update(self._work_status_identity_keys(item))
                        removed_count += 1
                        changed = True
                        continue
                    kept.append(item)
                if changed:
                    payload["items"] = kept
                    payload["updated_at"] = dt.datetime.now().strftime("%Y-%m-%d %H:%M")
                    self._state_store.put_document(
                        STATE_NS_WORK_STATUS, str(document.get("key") or ""), payload
                    )
            if removed_keys:
                for payload in self._iter_day_summary_payloads_locked():
                    if not isinstance(payload, dict):
                        continue
                    items = payload.get("items")
                    if not isinstance(items, list):
                        continue
                    kept = []
                    changed = False
                    for item in items:
                        if (
                            isinstance(item, dict)
                            and not self._is_completed_work_status_item(item)
                            and self._scope_matches_item(scope, item)
                            and (self._work_status_identity_keys(item) & removed_keys)
                        ):
                            changed = True
                            continue
                        kept.append(item)
                    if changed:
                        payload["items"] = kept
                        payload["updated_at"] = dt.datetime.now().strftime("%Y-%m-%d %H:%M")
                        self._save_day_summary_locked(payload)
                self._work_status_cache_signature = None
                self._work_status_cache_items = None
        return {"removed": removed_count}

    def _get_record_memory(self, record: dict[str, Any]) -> dict[str, str]:
        work_type = str(record.get("work_type") or WORK_TYPE_MAINTENANCE)
        fields = record.get("display_fields") or {}
        if work_type == WORK_TYPE_CHANGE:
            building = self._building_label_from_codes(
                self._change_record_building_codes(record)
            )
            memory_name = self._change_title(record)
        elif work_type == WORK_TYPE_REPAIR:
            building = self._building_label_from_codes(
                self._repair_record_building_codes(record)
            )
            memory_name = self._repair_title(record)
        else:
            building = str(fields.get("楼栋") or "").strip()
            memory_name = str(fields.get("维护总项") or "").strip()
            maintenance_cycle = str(
                fields.get("维护周期") or record.get("maintenance_cycle") or ""
            ).strip()
        if not building or not memory_name:
            return {}
        key = self._memory_item_key(
            work_type,
            memory_name,
            maintenance_cycle if work_type == WORK_TYPE_MAINTENANCE else "",
        )
        with self._memory_lock:
            payload = self._load_building_memory_locked(building)
            items = payload.get("items") or {}
            item = items.get(key) or {}
            if not item and work_type == WORK_TYPE_MAINTENANCE:
                item = items.get(self._normalize_memory_key(memory_name)) or {}
        if not isinstance(item, dict):
            return {}
        return {
            "location": str(item.get("location") or ""),
            "content": str(item.get("content") or ""),
            "reason": str(item.get("reason") or ""),
            "impact": str(item.get("impact") or ""),
            "progress": str(item.get("progress") or ""),
            "maintenance_cycle": str(item.get("maintenance_cycle") or ""),
            "specialty": str(item.get("specialty") or ""),
            "level": str(item.get("level") or ""),
            "repair_device": str(item.get("repair_device") or ""),
            "repair_fault": str(item.get("repair_fault") or ""),
            "fault_type": str(item.get("fault_type") or ""),
            "repair_mode": str(item.get("repair_mode") or ""),
            "discovery": str(item.get("discovery") or ""),
            "symptom": str(item.get("symptom") or ""),
            "solution": str(item.get("solution") or ""),
            "zhihang_involved": str(item.get("zhihang_involved") or ""),
            "zhihang_record_id": str(item.get("zhihang_record_id") or ""),
            "zhihang_title": str(item.get("zhihang_title") or ""),
            "zhihang_progress": str(item.get("zhihang_progress") or ""),
            "updated_at": str(item.get("updated_at") or ""),
        }

    def _memory_item_key(
        self, work_type: str, item_name: str, maintenance_cycle: str = ""
    ) -> str:
        normalized = self._normalize_memory_key(item_name)
        work_type = str(work_type or WORK_TYPE_MAINTENANCE).strip() or WORK_TYPE_MAINTENANCE
        if work_type == WORK_TYPE_MAINTENANCE:
            cycle = self._normalize_memory_key(maintenance_cycle)
            if cycle:
                return f"{normalized}|cycle:{cycle}"
            return normalized
        return f"{work_type}:{normalized}"

    def _remember_draft_fields(
        self,
        *,
        building: str,
        maintenance_total: str,
        location: str,
        content: str,
        reason: str,
        impact: str,
        work_type: str = WORK_TYPE_MAINTENANCE,
        item_name: str = "",
        maintenance_cycle: str = "",
        extra_fields: dict[str, Any] | None = None,
    ) -> None:
        work_type = str(work_type or WORK_TYPE_MAINTENANCE).strip() or WORK_TYPE_MAINTENANCE
        building = str(building or "").strip()
        memory_name = str(item_name or maintenance_total or "").strip()
        maintenance_total = str(maintenance_total or memory_name).strip()
        maintenance_cycle = str(maintenance_cycle or "").strip()
        if not building or not memory_name:
            return
        key = self._memory_item_key(work_type, memory_name, maintenance_cycle)
        now = dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        remembered = {
            "work_type": work_type,
            "memory_name": memory_name,
            "maintenance_total": maintenance_total,
            "maintenance_cycle": maintenance_cycle,
            "location": str(location or ""),
            "content": str(content or ""),
            "reason": str(reason or ""),
            "impact": str(impact or ""),
            "updated_at": now,
        }
        for extra_key, extra_value in (extra_fields or {}).items():
            if extra_key in remembered:
                continue
            remembered[str(extra_key)] = str(extra_value or "")
        with self._memory_lock:
            payload = self._load_building_memory_locked(building)
            payload["building"] = building
            payload["updated_at"] = now
            items = payload.setdefault("items", {})
            items[key] = remembered
            self._save_building_memory_locked(building, payload)

    @staticmethod
    def _split_notice_text_blocks(text: str) -> list[str]:
        text = str(text or "").replace("\r\n", "\n").replace("\r", "\n").strip()
        if not text:
            return []
        pattern = re.compile(
            r"(?=【(?:维保通告|变更通告|设备检修|上电通告|下电通告|上下电通告|设备轮巡|设备调整)】\s*状态[:：])"
        )
        parts = [part.strip() for part in pattern.split(text) if part.strip()]
        if len(parts) > 1:
            return parts
        separator_parts = [
            part.strip()
            for part in re.split(r"\n\s*(?:-{4,}|={4,}|#{4,})\s*\n", text)
            if part.strip()
        ]
        return separator_parts or [text]

    @staticmethod
    def _parse_notice_sections(text: str) -> dict[str, str]:
        sections: dict[str, str] = {}
        matches = list(re.finditer(r"【([^】]+)】", text or ""))
        for index, match in enumerate(matches):
            label = str(match.group(1) or "").strip()
            if not label:
                continue
            start = match.end()
            end = matches[index + 1].start() if index + 1 < len(matches) else len(text)
            value = str((text or "")[start:end]).strip()
            value = re.sub(r"^[：:]\s*", "", value).strip()
            sections[label] = value
        return sections

    @staticmethod
    def _notice_section_value(
        sections: dict[str, str], names: list[str], fallback: str = ""
    ) -> str:
        for name in names:
            value = str(sections.get(name) or "").strip()
            if value:
                return value
        return str(fallback or "").strip()

    @classmethod
    def _notice_work_type_from_text(
        cls, text: str, sections: dict[str, str]
    ) -> str:
        raw = str(text or "")
        if "【设备检修】" in raw or cls._notice_section_value(
            sections, ["维修设备", "维修故障", "故障现象", "解决方案"]
        ):
            return WORK_TYPE_REPAIR
        if "【上电通告】" in raw or "【下电通告】" in raw or "【上下电通告】" in raw or cls._notice_section_value(
            sections, ["柜号", "数量"]
        ):
            return WORK_TYPE_POWER
        if "【设备轮巡】" in raw or cls._notice_section_value(
            sections, ["设备"]
        ):
            return WORK_TYPE_POLLING
        if "【设备调整】" in raw:
            return WORK_TYPE_ADJUST
        if "【变更通告】" in raw or cls._notice_section_value(
            sections, ["变更等级", "变更楼栋"]
        ):
            return WORK_TYPE_CHANGE
        return WORK_TYPE_MAINTENANCE

    @classmethod
    def _building_codes_from_notice_text(cls, *values: Any) -> list[str]:
        codes: list[str] = []
        text = "\n".join(str(value or "") for value in values if str(value or "").strip())
        upper = text.upper()
        for label in ("标题", "名称"):
            match = re.search(rf"【{label}】(.*?)(?=【|$)", text, re.S)
            section = str(match.group(1) if match else "").upper()
            if section and re.search(r"110\s*(?:站|楼|机房|数据中心|DC|KV)?", section):
                return ["110"]
        if re.search(r"110\s*(?:站|楼|机房|数据中心|DC|KV)?", upper):
            codes.append("110")
        for code in ("A", "B", "C", "D", "E", "H"):
            patterns = (
                rf"(?<![A-Z0-9]){code}\s*(?:楼|栋|座|区|机房|数据中心|DC)",
                rf"(?:楼栋|楼宇|数据中心)\s*{code}(?![A-Z0-9])",
                rf"(?<![A-Z0-9]){code}[-－]\d",
            )
            if any(re.search(pattern, upper) for pattern in patterns):
                codes.append(code)
        return [code for code in BUILDING_SCOPE_CODES if code in dict.fromkeys(codes)]

    @classmethod
    def _strip_notice_title_suffix(cls, title: str, work_type: str) -> str:
        text = re.sub(r"\s+", " ", str(title or "").strip())
        if not text:
            return ""
        if work_type == WORK_TYPE_MAINTENANCE:
            text = re.sub(r"^EA118机房", "", text, flags=re.IGNORECASE).strip()
            for code in BUILDING_SCOPE_CODES:
                label = cls._building_label_from_code(code)
                text = re.sub(rf"^{re.escape(label)}", "", text).strip()
            text = re.sub(r"(?:维保|维护)?通告$", "", text).strip()
            return text or str(title or "").strip()
        if work_type == WORK_TYPE_CHANGE:
            stripped = re.sub(r"(?:变更)?通告$", "", text).strip()
            return stripped or text
        return text

    def _find_zhihang_record_by_title(self, title: str) -> dict[str, Any] | None:
        target = self._match_text(title)
        if not target:
            return None
        candidates: list[dict[str, Any]] = []
        try:
            candidates.extend(self._source_snapshot_zhihang_records("ALL") or [])
        except Exception:
            pass
        candidates.extend(self._zhihang_change_records)
        seen: set[str] = set()
        for record in candidates:
            if not isinstance(record, dict):
                continue
            record_id = str(record.get("record_id") or "").strip()
            if record_id and record_id in seen:
                continue
            if record_id:
                seen.add(record_id)
            title_text = str(record.get("title") or "").strip() or self._zhihang_change_title(record)
            if self._match_text(title_text) == target:
                return record
        return None

    def _historical_notice_memory_item(
        self, block: str, *, scope: str
    ) -> tuple[dict[str, Any] | None, str]:
        sections = self._parse_notice_sections(block)
        work_type = self._notice_work_type_from_text(block, sections)
        title = self._notice_section_value(sections, ["标题", "名称", "通告名称"])
        location = self._notice_section_value(sections, ["地点", "位置"])
        building_text = self._notice_section_value(
            sections,
            ["楼栋", "变更楼栋", "所属楼栋", "所属数据中心/楼栋-使用"],
        )
        codes = self._building_codes_from_notice_text(building_text, title, location, block)
        if not codes:
            return None, "无法识别楼栋"
        if not self._scope_matches_buildings(scope, codes):
            return None, f"不属于当前入口 {self._scope_label(scope)}"
        building = self._building_label_from_codes(codes)
        specialty = self._notice_section_value(sections, ["专业", "专业类别", "所属专业"])
        level = self._notice_section_value(sections, ["等级", "变更等级", "紧急程度"])
        content = self._notice_section_value(sections, ["内容"], title)
        reason = self._notice_section_value(sections, ["原因", "故障原因"])
        impact = self._notice_section_value(sections, ["影响", "影响范围"])
        maintenance_cycle = self._notice_section_value(sections, ["维保周期", "维护周期"])
        if work_type == WORK_TYPE_MAINTENANCE:
            memory_name = self._notice_section_value(
                sections,
                ["维护总项", "维保总项", "项目", "维护项目"],
                self._strip_notice_title_suffix(title, work_type),
            )
            if not memory_name:
                return None, "维保通告缺少标题或维护总项"
            return {
                "work_type": work_type,
                "building": building,
                "memory_name": memory_name,
                "maintenance_cycle": maintenance_cycle,
                "location": location,
                "content": content,
                "reason": reason,
                "impact": impact,
                "extra_fields": {"specialty": specialty},
            }, ""
        if work_type == WORK_TYPE_CHANGE:
            memory_name = self._notice_section_value(
                sections,
                ["变更简述", "变更标题"],
                self._strip_notice_title_suffix(title, work_type) or title,
            )
            if not memory_name:
                return None, "变更通告缺少标题"
            zhihang_title = self._notice_section_value(
                sections, ["智航变更", "智航侧变更", "互联变更", "互联侧变更"]
            )
            zhihang_record_id = self._notice_section_value(
                sections, ["智航记录ID", "互联记录ID"]
            )
            zhihang_progress = ""
            if zhihang_title and not zhihang_record_id:
                zhihang_record = self._find_zhihang_record_by_title(zhihang_title)
                if zhihang_record:
                    zhihang_record_id = str(zhihang_record.get("record_id") or "").strip()
                    zhihang_title = str(zhihang_record.get("title") or "").strip() or self._zhihang_change_title(zhihang_record)
                    zhihang_progress = str(zhihang_record.get("progress") or "").strip() or self._zhihang_change_progress(zhihang_record)
            return {
                "work_type": work_type,
                "building": building,
                "memory_name": memory_name,
                "location": location,
                "content": content or memory_name,
                "reason": reason,
                "impact": impact,
                "extra_fields": {
                    "specialty": specialty,
                    "level": level,
                    "zhihang_involved": "1" if zhihang_record_id else "",
                    "zhihang_record_id": zhihang_record_id,
                    "zhihang_title": zhihang_title,
                    "zhihang_progress": zhihang_progress,
                },
            }, ""
        if work_type in {WORK_TYPE_POWER, WORK_TYPE_POLLING, WORK_TYPE_ADJUST}:
            memory_name = self._strip_notice_title_suffix(title, work_type) or title
            if not memory_name:
                return None, f"{self._history_work_type_label(work_type)}通告缺少标题"
            return {
                "work_type": work_type,
                "building": building,
                "memory_name": memory_name,
                "location": location,
                "content": content or memory_name,
                "reason": reason,
                "impact": impact,
                "extra_fields": {
                    "specialty": specialty,
                    "device": self._notice_section_value(sections, ["设备"]),
                    "cabinet": self._notice_section_value(sections, ["柜号"]),
                    "quantity": self._notice_section_value(sections, ["数量"]),
                    "progress": self._notice_section_value(sections, ["进度"]),
                },
            }, ""
        memory_name = self._notice_section_value(
            sections,
            ["检修通告名称", "维修名称"],
            title,
        )
        if not memory_name:
            return None, "检修通告缺少标题"
        return {
            "work_type": work_type,
            "building": building,
            "memory_name": memory_name,
            "location": location,
            "content": content or memory_name,
            "reason": reason,
            "impact": impact,
            "extra_fields": {
                "specialty": specialty,
                "level": level,
                "repair_device": self._notice_section_value(sections, ["维修设备"]),
                "repair_fault": self._notice_section_value(sections, ["维修故障"]),
                "fault_type": self._notice_section_value(sections, ["故障类型"]),
                "repair_mode": self._notice_section_value(sections, ["维修方式"]),
                "discovery": self._notice_section_value(sections, ["故障发现方式"]),
                "symptom": self._notice_section_value(sections, ["故障现象"]),
                "solution": self._notice_section_value(sections, ["解决方案"]),
                "spare_parts": self._notice_section_value(
                    sections, ["备件更换情况", "备件使用情况"]
                ),
            },
        }, ""

    def import_historical_notice_memory(
        self,
        *,
        text: str,
        scope: str = "ALL",
        allowed_scopes: list[str] | None = None,
        is_admin: bool = False,
        imported_by: str = "",
    ) -> dict[str, Any]:
        scope = self._normalize_scope(scope)
        blocks = self._split_notice_text_blocks(text)
        if not blocks:
            raise PortalError("请粘贴至少一条历史通告。")
        if len(blocks) > 200:
            raise PortalError("一次最多导入 200 条历史通告，请分批导入。")
        allowed = {self._normalize_scope(item) for item in (allowed_scopes or [])}
        imported: list[dict[str, Any]] = []
        skipped: list[dict[str, Any]] = []
        for index, block in enumerate(blocks, start=1):
            item, reason = self._historical_notice_memory_item(block, scope=scope)
            if not item:
                skipped.append({"index": index, "reason": reason or "无法解析"})
                continue
            item_codes = self._building_codes_from_value(item.get("building"))
            item_scope = "CAMPUS" if len(item_codes) >= 2 else (item_codes[0] if item_codes else "")
            if not is_admin and "ALL" not in allowed and item_scope not in allowed:
                skipped.append(
                    {
                        "index": index,
                        "title": item.get("memory_name") or "",
                        "reason": f"无权导入 {self._scope_label(item_scope)} 历史通告",
                    }
                )
                continue
            extra_fields = item.get("extra_fields") if isinstance(item.get("extra_fields"), dict) else {}
            self._remember_draft_fields(
                work_type=str(item.get("work_type") or WORK_TYPE_MAINTENANCE),
                building=str(item.get("building") or ""),
                maintenance_total=str(item.get("memory_name") or ""),
                item_name=str(item.get("memory_name") or ""),
                maintenance_cycle=str(item.get("maintenance_cycle") or ""),
                location=str(item.get("location") or ""),
                content=str(item.get("content") or ""),
                reason=str(item.get("reason") or ""),
                impact=str(item.get("impact") or ""),
                extra_fields={
                    **extra_fields,
                    "imported_by": imported_by,
                    "imported_from": "historical_notice",
                },
            )
            imported.append(
                {
                    "index": index,
                    "work_type": item.get("work_type") or "",
                    "building": item.get("building") or "",
                    "title": item.get("memory_name") or "",
                    "maintenance_cycle": item.get("maintenance_cycle") or "",
                }
            )
        return {
            "imported_count": len(imported),
            "skipped_count": len(skipped),
            "imported": imported,
            "skipped": skipped,
        }

    @staticmethod
    def _history_work_type_label(work_type: str) -> str:
        return {
            WORK_TYPE_MAINTENANCE: "维保",
            WORK_TYPE_CHANGE: "变更",
            WORK_TYPE_REPAIR: "检修",
            WORK_TYPE_POWER: "上电",
            WORK_TYPE_POLLING: "轮巡",
            WORK_TYPE_ADJUST: "调整",
            WORK_TYPE_EVENT: "事件",
        }.get(str(work_type or ""), "维保")

    @classmethod
    def _coerce_history_work_types(cls, work_types: Any) -> list[str]:
        if isinstance(work_types, str):
            raw_values = [item.strip() for item in work_types.split(",")]
        elif isinstance(work_types, (list, tuple, set)):
            raw_values = [str(item or "").strip() for item in work_types]
        else:
            raw_values = []
        aliases = {
            "maintenance": WORK_TYPE_MAINTENANCE,
            "维保": WORK_TYPE_MAINTENANCE,
            "维保通告": WORK_TYPE_MAINTENANCE,
            "change": WORK_TYPE_CHANGE,
            "变更": WORK_TYPE_CHANGE,
            "变更通告": WORK_TYPE_CHANGE,
            "repair": WORK_TYPE_REPAIR,
            "检修": WORK_TYPE_REPAIR,
            "设备检修": WORK_TYPE_REPAIR,
            "power": WORK_TYPE_POWER,
            "上电": WORK_TYPE_POWER,
            "上电通告": WORK_TYPE_POWER,
            "下电": WORK_TYPE_POWER,
            "下电通告": WORK_TYPE_POWER,
            "上下电通告": WORK_TYPE_POWER,
            "polling": WORK_TYPE_POLLING,
            "轮巡": WORK_TYPE_POLLING,
            "设备轮巡": WORK_TYPE_POLLING,
            "adjust": WORK_TYPE_ADJUST,
            "调整": WORK_TYPE_ADJUST,
            "设备调整": WORK_TYPE_ADJUST,
            "event": WORK_TYPE_EVENT,
            "事件": WORK_TYPE_EVENT,
            "事件通告": WORK_TYPE_EVENT,
        }
        result: list[str] = []
        for value in raw_values:
            normalized = aliases.get(value.lower()) or aliases.get(value)
            if normalized and normalized not in result:
                result.append(normalized)
        return result or [WORK_TYPE_MAINTENANCE, WORK_TYPE_CHANGE, WORK_TYPE_REPAIR]

    @staticmethod
    def _history_month_floor(months: int) -> dt.datetime:
        count = max(1, min(int(months or 3), 12))
        current = dt.datetime(dt.datetime.now().year, dt.datetime.now().month, 1)
        month = current.month - (count - 1)
        year = current.year
        while month <= 0:
            month += 12
            year -= 1
        return dt.datetime(year, month, 1)

    def _history_target_config(self, work_type: str) -> dict[str, str]:
        work_type = str(work_type or WORK_TYPE_MAINTENANCE).strip()
        if work_type == WORK_TYPE_CHANGE:
            return {
                "work_type": WORK_TYPE_CHANGE,
                "notice_type": NOTICE_TYPE_CHANGE,
                "app_token": str(config.app_token or "").strip(),
                "table_id": str(config.table_id_biangeng or "").strip(),
            }
        if work_type == WORK_TYPE_REPAIR:
            return {
                "work_type": WORK_TYPE_REPAIR,
                "notice_type": NOTICE_TYPE_REPAIR,
                "app_token": str(config.app_token or "").strip(),
                "table_id": str(config.table_id_overhaul or "").strip(),
            }
        return {
            "work_type": WORK_TYPE_MAINTENANCE,
            "notice_type": NOTICE_TYPE_MAINTENANCE,
            "app_token": str(config.app_token or "").strip(),
            "table_id": str(config.table_id_weibao or "").strip(),
        }

    def _history_datetime_from_values(self, *values: Any) -> dt.datetime | None:
        for value in values:
            formatted = self._format_source_datetime(value)
            if not formatted:
                continue
            match = re.search(
                r"(\d{4})[-/年](\d{1,2})[-/月](\d{1,2})"
                r"(?:[日\sT]*(\d{1,2})[:点时](\d{1,2})?)?",
                formatted,
            )
            if not match:
                continue
            year, month, day = int(match.group(1)), int(match.group(2)), int(match.group(3))
            hour = int(match.group(4) or 0)
            minute = int(match.group(5) or 0)
            try:
                return dt.datetime(year, month, day, hour, minute)
            except ValueError:
                continue
        return None

    def _target_history_datetime(
        self, record: dict[str, Any], work_type: str
    ) -> tuple[dt.datetime | None, str]:
        fields = record.get("display_fields") or {}
        raw_values: list[Any] = []
        if work_type == WORK_TYPE_CHANGE:
            raw_values.extend(
                [
                    fields.get("变更开始时间"),
                    fields.get("过程更新时间"),
                    fields.get("变更结束时间"),
                    fields.get("计划开始时间"),
                    fields.get("计划开始"),
                ]
            )
        elif work_type == WORK_TYPE_REPAIR:
            raw_values.extend(
                [
                    fields.get("实际开始时间"),
                    fields.get("发生故障时间"),
                    fields.get("期望完成时间"),
                    fields.get("实际结束时间"),
                ]
            )
        else:
            raw_values.extend(
                [
                    fields.get("实际开始时间"),
                    fields.get("计划开始时间"),
                    fields.get("实际结束时间"),
                    fields.get("计划结束时间"),
                ]
            )
        raw_values.extend([record.get("created_time"), record.get("last_modified_time")])
        parsed = self._history_datetime_from_values(*raw_values)
        return parsed, parsed.strftime("%Y-%m-%d %H:%M") if parsed else ""

    def _canonical_history_notice_title(self, value: Any, work_type: str) -> str:
        text = str(value or "").strip()
        text = re.sub(r"【[^】]+】\s*状态[：:]\s*\S+", "", text)
        if work_type == WORK_TYPE_MAINTENANCE:
            text = re.sub(r"(?:通告)+$", "", text).strip()
        else:
            text = self._strip_notice_title_suffix(text, work_type)
        text = re.sub(r"^EA118\s*(?:机房)?", "", text, flags=re.IGNORECASE).strip()
        text = re.sub(r"^(?:南通基地|南通)?", "", text).strip()
        for code in BUILDING_SCOPE_CODES:
            label = self._building_label_from_code(code)
            text = re.sub(rf"^{re.escape(label)}", "", text).strip()
        text = re.sub(r"(?:通告)+$", "", text).strip()
        text = re.sub(r"[\s,，;；:：。.【】（）()《》<>\"'“”‘’\-－_/\\]+", "", text)
        return text.lower()

    def _history_target_memory_fields(
        self, record: dict[str, Any], work_type: str
    ) -> dict[str, str]:
        fields = record.get("display_fields") or {}
        if work_type == WORK_TYPE_REPAIR:
            return {
                "location": str(fields.get("位置") or "").strip(),
                "content": str(fields.get("名称（标题）") or fields.get("名称") or "").strip(),
                "reason": str(fields.get("故障原因") or "").strip(),
                "impact": str(fields.get("影响范围") or fields.get("影响") or "").strip(),
                "progress": str(fields.get("进度（完成情况）") or fields.get("进度") or "").strip(),
                "specialty": self._clean_source_text(fields.get("专业")),
                "level": str(fields.get("紧急程度") or "").strip(),
                "repair_device": str(fields.get("维修设备") or "").strip(),
                "repair_fault": str(fields.get("维修故障") or "").strip(),
                "fault_type": str(fields.get("故障类型") or "").strip(),
                "repair_mode": str(fields.get("维修方式") or "").strip(),
                "discovery": str(fields.get("故障发现方式（来源）") or fields.get("故障发现方式") or "").strip(),
                "symptom": str(fields.get("故障现象") or "").strip(),
                "solution": str(fields.get("解决方案") or "").strip(),
            }
        return {
            "location": str(fields.get("位置") or "").strip(),
            "content": str(fields.get("内容") or "").strip(),
            "reason": str(fields.get("原因") or "").strip(),
            "impact": str(fields.get("影响") or fields.get("影响范围") or "").strip(),
            "progress": str(fields.get("进度") or "").strip(),
            "specialty": self._clean_source_text(fields.get("专业")),
            "level": str(fields.get("阿里-变更等级") or fields.get("智航-变更等级") or "").strip(),
        }

    def _history_candidate_from_target_record(
        self, record: dict[str, Any], *, work_type: str, since: dt.datetime
    ) -> tuple[dict[str, Any] | None, str]:
        fields = record.get("display_fields") or {}
        if work_type == WORK_TYPE_CHANGE:
            title = str(fields.get("名称") or record.get("record_id") or "").strip()
            building_codes = self._building_codes_from_value(fields.get("楼栋") or title)
            maintenance_cycle = ""
        elif work_type == WORK_TYPE_REPAIR:
            title = str(fields.get("名称（标题）") or fields.get("名称") or record.get("record_id") or "").strip()
            building_codes = self._building_codes_from_value(fields.get("楼栋") or title)
            maintenance_cycle = ""
        else:
            title = str(fields.get("名称") or record.get("record_id") or "").strip()
            building_codes = self._building_codes_from_value(fields.get("楼栋") or title)
            maintenance_cycle = str(fields.get("维保周期") or "").strip()
        if not title:
            return None, "历史记录缺少标题"
        if not building_codes:
            return None, f"历史记录无法识别楼栋: {title}"
        business_dt, business_time = self._target_history_datetime(record, work_type)
        if business_dt is None or business_dt < since:
            return None, "不在近3个月范围内"
        if work_type == WORK_TYPE_MAINTENANCE:
            memory_name = re.sub(r"(?:通告)+$", "", title).strip() or title
        else:
            memory_name = self._strip_notice_title_suffix(title, work_type) or title
        canonical = self._canonical_history_notice_title(memory_name, work_type)
        if not canonical:
            return None, "标题标准化后为空"
        memory_fields = self._history_target_memory_fields(record, work_type)
        if work_type == WORK_TYPE_CHANGE:
            memory_fields["zhihang_title"] = str(fields.get("智航变更") or "").strip()
        return {
            "id": f"{work_type}:{record.get('record_id') or uuid.uuid4().hex}",
            "record_id": str(record.get("record_id") or ""),
            "work_type": work_type,
            "work_type_label": self._history_work_type_label(work_type),
            "notice_type": str(record.get("notice_type") or ""),
            "title": title,
            "memory_name": memory_name,
            "canonical_title": canonical,
            "building": self._building_label_from_codes(building_codes),
            "building_codes": building_codes,
            "maintenance_cycle": maintenance_cycle,
            "business_time": business_time,
            "fields": memory_fields,
            "display_fields": fields,
        }, ""

    def _history_source_current_fields(self, record: dict[str, Any]) -> dict[str, str]:
        fields = record.get("display_fields") or {}
        work_type = self._record_work_type(record)
        if work_type == WORK_TYPE_REPAIR:
            return {
                "location": self._repair_location_text(fields),
                "content": self._repair_title(record),
                "reason": self._repair_first_field(fields, "故障原因", "故障维修原因"),
                "impact": self._repair_first_field(fields, "影响范围", "影响"),
                "progress": self._repair_first_field(fields, "完成情况", "进度", "维修进展"),
                "specialty": self._repair_specialty(record),
                "level": self._repair_level(fields),
                "repair_device": self._repair_device_text(fields),
                "repair_fault": self._repair_first_field(fields, "维修故障", "故障维修原因"),
                "fault_type": self._repair_first_field(fields, "故障类型") or "设备故障",
                "repair_mode": self._repair_first_field(fields, "维修方式", "维修方", "供应商名称"),
                "discovery": self._repair_first_field(fields, "对应来源"),
                "symptom": self._repair_first_field(fields, "故障发生现象描述", "故障现象"),
                "solution": self._repair_first_field(fields, "解决方案", "维修方案", "后续整改措施"),
            }
        if work_type == WORK_TYPE_CHANGE:
            return {
                "location": str(fields.get("位置") or fields.get("变更楼栋") or "").strip(),
                "content": self._change_title(record),
                "reason": str(fields.get("变更原因") or fields.get("原因") or "").strip(),
                "impact": str(fields.get("影响") or fields.get("影响范围") or "").strip(),
                "progress": str(fields.get("进度") or fields.get("变更进度") or "").strip(),
                "specialty": self._change_specialty(record),
                "level": str(fields.get("变更等级（阿里）") or fields.get("变更等级") or "").strip(),
                "zhihang_title": "",
                "zhihang_record_id": "",
                "zhihang_progress": "",
            }
        maintenance_total = str(fields.get("维护总项") or "").strip()
        return {
            "location": str(fields.get("位置") or fields.get("楼栋") or "").strip(),
            "content": str(fields.get("内容") or fields.get("维护内容") or maintenance_total).strip(),
            "reason": str(fields.get("原因") or fields.get("维护原因") or "").strip(),
            "impact": str(fields.get("影响") or fields.get("影响范围") or "").strip(),
            "progress": str(fields.get("进度") or fields.get("维护进度") or "").strip(),
            "specialty": self._clean_source_text(fields.get("专业类别"))
            or self._clean_source_text(fields.get("专业")),
            "level": "",
            "maintenance_cycle": str(fields.get("维护周期") or "").strip(),
        }

    def _history_source_item_from_record(self, record: dict[str, Any]) -> dict[str, Any] | None:
        work_type = self._record_work_type(record)
        fields = record.get("display_fields") or {}
        if work_type == WORK_TYPE_CHANGE:
            title = self._change_title(record)
            memory_name = title
            building_codes = self._change_record_building_codes(record)
            maintenance_cycle = ""
            specialty = self._change_specialty(record)
        elif work_type == WORK_TYPE_REPAIR:
            title = self._repair_title(record)
            memory_name = title
            building_codes = self._repair_record_building_codes(record)
            maintenance_cycle = ""
            specialty = self._repair_specialty(record)
        else:
            maintenance_total = str(fields.get("维护总项") or "").strip()
            building = str(fields.get("楼栋") or "").strip()
            maintenance_cycle = str(fields.get("维护周期") or "").strip()
            title = f"{building}{maintenance_total}{('-' + maintenance_cycle) if maintenance_cycle else ''}".strip()
            memory_name = maintenance_total
            building_codes = self._building_codes_from_value(building)
            specialty = str(fields.get("专业类别") or "").strip()
        if not memory_name:
            return None
        canonical = self._canonical_history_notice_title(memory_name, work_type)
        source_id = f"{work_type}:{record.get('record_id') or uuid.uuid4().hex}"
        if work_type == WORK_TYPE_CHANGE:
            source_status = self._change_progress_value(record)
        elif work_type == WORK_TYPE_REPAIR:
            source_status = self._repair_source_status(record)
        else:
            source_status = self._maintenance_status_value(record)
        return {
            "id": source_id,
            "source_record_id": str(record.get("record_id") or ""),
            "work_type": work_type,
            "work_type_label": self._history_work_type_label(work_type),
            "title": title or memory_name,
            "memory_name": memory_name,
            "canonical_title": canonical,
            "building": self._building_label_from_codes(building_codes),
            "building_codes": building_codes,
            "maintenance_cycle": maintenance_cycle,
            "specialty": specialty,
            "source_status": source_status,
            "memory": self._get_record_memory(record),
            "current_fields": self._history_source_current_fields(record),
            "display_fields": fields,
        }

    def _history_source_items(self, work_types: list[str]) -> list[dict[str, Any]]:
        records = self._source_snapshot_records("ALL") or []
        current_month = self._current_month_label()
        items: list[dict[str, Any]] = []
        for record in records:
            work_type = self._record_work_type(record)
            if work_type not in work_types:
                continue
            if not self._source_record_matches_month_window(record, current_month):
                continue
            item = self._history_source_item_from_record(record)
            if item:
                items.append(item)
        return items

    def _history_current_month_source_items(
        self, work_types: list[str]
    ) -> tuple[list[dict[str, Any]], list[str]]:
        """Load current-month source records for the admin memory import page.

        This intentionally does not reuse the workbench active snapshot because
        that snapshot is filtered for day-to-day issuing. The memory import page
        needs all current-month source records, including records that are
        already ended.  The ``work_types`` argument controls which historical
        target tables are scanned as candidates; it must not hide current-month
        source items from the left side of the import page.
        """
        current_month = self._current_month_label()
        source_work_types = [WORK_TYPE_MAINTENANCE, WORK_TYPE_CHANGE, WORK_TYPE_REPAIR]
        records: list[dict[str, Any]] = []
        warnings: list[str] = []
        with self._refresh_lock:
            if WORK_TYPE_MAINTENANCE in source_work_types:
                try:
                    self._load_fields()
                    app_token = str(self.app_token or DEFAULT_APP_TOKEN or "").strip()
                    table_id = str(self.table_id or DEFAULT_TABLE_ID or "").strip()
                    loaded = self._load_table_records(
                        app_token=app_token,
                        table_id=table_id,
                        meta_by_name=self._field_meta_by_name,
                        work_type=WORK_TYPE_MAINTENANCE,
                        notice_type=NOTICE_TYPE_MAINTENANCE,
                    )
                    records.extend(
                        item
                        for item in loaded
                        if self._maintenance_record_matches_month_window(item, current_month)
                    )
                except Exception as exc:
                    warnings.append(f"维保当前月源表全量读取失败: {exc}")
            if WORK_TYPE_CHANGE in source_work_types:
                try:
                    self._load_change_fields()
                    loaded = self._load_table_records(
                        app_token=CHANGE_SOURCE_APP_TOKEN,
                        table_id=CHANGE_SOURCE_TABLE_ID,
                        meta_by_name=self._change_field_meta_by_name,
                        work_type=WORK_TYPE_CHANGE,
                        notice_type=NOTICE_TYPE_CHANGE,
                    )
                    records.extend(
                        item
                        for item in loaded
                        if self._source_record_matches_month_window(item, current_month)
                    )
                except Exception as exc:
                    warnings.append(f"变更当前月源表全量读取失败: {exc}")
        if WORK_TYPE_REPAIR in source_work_types:
            records.extend(
                record
                for record in (self._source_snapshot_records("ALL") or [])
                if self._record_work_type(record) == WORK_TYPE_REPAIR
                and self._source_record_matches_month_window(record, current_month)
            )
        items: list[dict[str, Any]] = []
        seen: set[str] = set()
        for record in records:
            item = self._history_source_item_from_record(record)
            if not item:
                continue
            key = str(item.get("id") or "")
            if key and key in seen:
                continue
            if key:
                seen.add(key)
            items.append(item)
        items.sort(
            key=lambda item: (
                str(item.get("work_type") or ""),
                str(item.get("building") or ""),
                str(item.get("title") or ""),
            )
        )
        return items, warnings

    def _scan_target_history_candidates(
        self, *, work_type: str, months: int
    ) -> tuple[list[dict[str, Any]], list[str]]:
        target = self._history_target_config(work_type)
        app_token = target["app_token"]
        table_id = target["table_id"]
        if not app_token or not table_id:
            return [], [f"{self._history_work_type_label(work_type)}目标表未配置，已跳过。"]
        metas, meta_by_name = self._load_table_fields(app_token=app_token, table_id=table_id)
        records = self._load_table_records(
            app_token=app_token,
            table_id=table_id,
            meta_by_name=meta_by_name,
            work_type=work_type,
            notice_type=target["notice_type"],
        )
        since = self._history_month_floor(months)
        candidates: list[dict[str, Any]] = []
        skipped: dict[str, int] = {}
        for record in records:
            candidate, reason = self._history_candidate_from_target_record(
                record, work_type=work_type, since=since
            )
            if not candidate:
                skipped[reason or "无法解析"] = skipped.get(reason or "无法解析", 0) + 1
                continue
            candidates.append(candidate)
        deduped: dict[tuple, dict[str, Any]] = {}
        for candidate in candidates:
            key = (
                candidate.get("work_type"),
                tuple(candidate.get("building_codes") or []),
                candidate.get("canonical_title"),
                candidate.get("maintenance_cycle") if work_type == WORK_TYPE_MAINTENANCE else "",
            )
            current = deduped.get(key)
            if current is None or str(candidate.get("business_time") or "") >= str(current.get("business_time") or ""):
                deduped[key] = candidate
        warnings = [
            f"{self._history_work_type_label(work_type)}历史记录跳过 {count} 条：{reason}"
            for reason, count in sorted(skipped.items())
            if reason != "不在近3个月范围内"
        ]
        return list(deduped.values()), warnings

    @staticmethod
    def _history_building_score(source: dict[str, Any], candidate: dict[str, Any]) -> int:
        source_codes = set(source.get("building_codes") or [])
        candidate_codes = set(candidate.get("building_codes") or [])
        if not source_codes or not candidate_codes:
            return 5
        if source_codes == candidate_codes:
            return 20
        if source_codes & candidate_codes:
            return 10
        return -100

    def _recommend_history_memory_matches(
        self, source_items: list[dict[str, Any]], candidates: list[dict[str, Any]]
    ) -> list[dict[str, Any]]:
        matches: list[dict[str, Any]] = []
        by_type: dict[str, list[dict[str, Any]]] = {}
        for candidate in candidates:
            by_type.setdefault(str(candidate.get("work_type") or ""), []).append(candidate)
        for source in source_items:
            best: dict[str, Any] | None = None
            best_score = -999
            best_reason = ""
            for candidate in by_type.get(str(source.get("work_type") or ""), []):
                score = self._history_building_score(source, candidate)
                if score < 0:
                    continue
                source_title = str(source.get("canonical_title") or "")
                candidate_title = str(candidate.get("canonical_title") or "")
                if source_title and source_title == candidate_title:
                    score += 70
                    reason = "标题与楼栋匹配"
                elif source_title and candidate_title and (
                    source_title in candidate_title or candidate_title in source_title
                ):
                    score += 45
                    reason = "标题相似"
                else:
                    reason = "楼栋相同"
                if source.get("work_type") == WORK_TYPE_MAINTENANCE:
                    source_cycle = self._normalize_memory_key(str(source.get("maintenance_cycle") or ""))
                    candidate_cycle = self._normalize_memory_key(str(candidate.get("maintenance_cycle") or ""))
                    if source_cycle and candidate_cycle and source_cycle == candidate_cycle:
                        score += 10
                        reason += "，周期一致"
                    elif source_cycle and candidate_cycle and source_cycle != candidate_cycle:
                        score -= 25
                        reason += "，周期不同"
                if score > best_score:
                    best = candidate
                    best_score = score
                    best_reason = reason
            if not best:
                continue
            selected = best_score >= 80
            matches.append(
                {
                    "source_id": source.get("id") or "",
                    "candidate_id": best.get("id") or "",
                    "score": best_score,
                    "confidence": "high" if best_score >= 90 else "medium" if best_score >= 70 else "low",
                    "selected": selected,
                    "reason": best_reason,
                    "fields": copy.deepcopy(best.get("fields") or {}),
                }
            )
        return matches

    def scan_historical_notice_memory_candidates(
        self, *, work_types: Any = None, months: int = 3
    ) -> dict[str, Any]:
        work_type_list = self._coerce_history_work_types(work_types)
        months = max(1, min(int(months or 3), 12))
        candidates: list[dict[str, Any]] = []
        source_items, warnings = self._history_current_month_source_items(work_type_list)
        for work_type in work_type_list:
            try:
                items, item_warnings = self._scan_target_history_candidates(
                    work_type=work_type, months=months
                )
                candidates.extend(items)
                warnings.extend(item_warnings)
            except Exception as exc:
                warnings.append(f"{self._history_work_type_label(work_type)}历史扫描失败: {exc}")
        candidates.sort(
            key=lambda item: (
                str(item.get("work_type") or ""),
                str(item.get("building") or ""),
                str(item.get("business_time") or ""),
            ),
            reverse=True,
        )
        matches = self._recommend_history_memory_matches(source_items, candidates)
        return {
            "months": months,
            "work_types": work_type_list,
            "source_snapshot_ready": bool(self._source_snapshot_records("ALL") is not None),
            "source_items_full_current_month": True,
            "source_items": source_items,
            "candidates": candidates,
            "matches": matches,
            "counts": {
                "source": len(source_items),
                "candidates": len(candidates),
                "recommended": sum(1 for item in matches if item.get("selected")),
            },
            "warnings": warnings,
        }

    def save_historical_notice_memory_matches(
        self, *, matches: list[dict[str, Any]], imported_by: str = ""
    ) -> dict[str, Any]:
        if not isinstance(matches, list):
            raise PortalError("保存参数格式错误。")
        saved: list[dict[str, Any]] = []
        skipped: list[dict[str, Any]] = []
        overwritten_count = 0
        now = dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        for index, match in enumerate(matches, start=1):
            if not isinstance(match, dict) or not bool(match.get("selected", True)):
                skipped.append({"index": index, "reason": "未勾选"})
                continue
            source = match.get("source_item") if isinstance(match.get("source_item"), dict) else {}
            fields = match.get("fields") if isinstance(match.get("fields"), dict) else {}
            work_type = str(source.get("work_type") or match.get("work_type") or WORK_TYPE_MAINTENANCE).strip()
            building = str(source.get("building") or match.get("building") or "").strip()
            memory_name = str(source.get("memory_name") or source.get("title") or match.get("memory_name") or "").strip()
            maintenance_cycle = str(
                source.get("maintenance_cycle")
                or fields.get("maintenance_cycle")
                or match.get("maintenance_cycle")
                or ""
            ).strip()
            if not building or not memory_name:
                skipped.append({"index": index, "reason": "缺少楼栋或记忆名称"})
                continue
            key = self._memory_item_key(work_type, memory_name, maintenance_cycle)
            with self._memory_lock:
                payload = self._load_building_memory_locked(building)
                existed = key in (payload.get("items") or {})
            if existed:
                overwritten_count += 1
            extra_fields = {
                "specialty": str(fields.get("specialty") or ""),
                "level": str(fields.get("level") or ""),
                "progress": str(fields.get("progress") or ""),
                "repair_device": str(fields.get("repair_device") or ""),
                "repair_fault": str(fields.get("repair_fault") or ""),
                "fault_type": str(fields.get("fault_type") or ""),
                "repair_mode": str(fields.get("repair_mode") or ""),
                "discovery": str(fields.get("discovery") or ""),
                "symptom": str(fields.get("symptom") or ""),
                "solution": str(fields.get("solution") or ""),
                "zhihang_involved": str(fields.get("zhihang_involved") or ""),
                "zhihang_record_id": str(fields.get("zhihang_record_id") or ""),
                "zhihang_title": str(fields.get("zhihang_title") or ""),
                "zhihang_progress": str(fields.get("zhihang_progress") or ""),
                "imported_by": imported_by,
                "imported_from": "target_history_scan",
                "history_candidate_id": str(match.get("candidate_id") or ""),
                "history_imported_at": now,
            }
            self._remember_draft_fields(
                work_type=work_type,
                building=building,
                maintenance_total=memory_name,
                item_name=memory_name,
                maintenance_cycle=maintenance_cycle,
                location=str(fields.get("location") or ""),
                content=str(fields.get("content") or ""),
                reason=str(fields.get("reason") or ""),
                impact=str(fields.get("impact") or ""),
                extra_fields=extra_fields,
            )
            saved.append(
                {
                    "index": index,
                    "work_type": work_type,
                    "building": building,
                    "title": memory_name,
                    "maintenance_cycle": maintenance_cycle,
                    "overwritten": existed,
                }
            )
        if saved:
            self._touch_state_cache_version()
        return {
            "saved_count": len(saved),
            "skipped_count": len(skipped),
            "overwritten_count": overwritten_count,
            "saved": saved,
            "skipped": skipped,
        }

    def _serialize_record(
        self, record: dict[str, Any], summary_by_record: dict[str, dict[str, Any]] | None = None
    ) -> dict[str, Any]:
        summary_by_record = summary_by_record or {}
        work_type = str(record.get("work_type") or WORK_TYPE_MAINTENANCE)
        source_work_type = self._record_source_work_type(record)
        if source_work_type == WORK_TYPE_MAINTENANCE:
            source_progress = self._maintenance_status_value(record)
        elif source_work_type == WORK_TYPE_CHANGE:
            source_progress = self._change_progress_value(record)
        elif source_work_type == WORK_TYPE_REPAIR:
            source_progress = self._repair_source_status(record)
        else:
            source_progress = self._maintenance_status_value(record)
        title = (
            self._change_title(record)
            if work_type == WORK_TYPE_CHANGE
            else self._repair_title(record)
            if work_type == WORK_TYPE_REPAIR
            else ""
        )
        return {
            "record_id": record["record_id"],
            "source_record_id": record["record_id"],
            "source_work_type": source_work_type,
            "converted_from_work_type": str(record.get("converted_from_work_type") or ""),
            "converted_to_work_type": str(record.get("converted_to_work_type") or ""),
            "work_type_override_key": str(record.get("work_type_override_key") or ""),
            "work_type_override_title_key": str(record.get("work_type_override_title_key") or ""),
            "work_type": work_type,
            "notice_type": str(record.get("notice_type") or NOTICE_TYPE_MAINTENANCE),
            "source_app_token": str(record.get("source_app_token") or self.app_token),
            "source_table_id": str(record.get("source_table_id") or self.table_id),
            "title": title,
            "display_fields": record["display_fields"],
            "memory": self._get_record_memory(record),
            "work_summary": summary_by_record.get(record["record_id"]) or {},
            "source_progress": source_progress,
            "source_status": source_progress,
            "building_codes": (
                self._change_record_building_codes(record)
                if record.get("work_type") == WORK_TYPE_CHANGE
                else self._repair_record_building_codes(record)
                if record.get("work_type") == WORK_TYPE_REPAIR
                else self._building_codes_from_value(
                    (record.get("display_fields") or {}).get("楼栋")
                )
            ),
        }

    def _serialize_zhihang_change_record(self, record: dict[str, Any]) -> dict[str, Any]:
        fields = record.get("display_fields") or {}
        codes = self._zhihang_change_building_codes(record)
        scope_all = not codes
        return {
            "record_id": str(record.get("record_id") or ""),
            "source_record_id": str(record.get("record_id") or ""),
            "source_app_token": ZHIHANG_CHANGE_APP_TOKEN,
            "source_table_id": ZHIHANG_CHANGE_TABLE_ID,
            "title": self._zhihang_change_title(record),
            "progress": self._zhihang_change_progress(record),
            "building": str(fields.get("楼栋") or ""),
            "building_codes": codes,
            "scope_all": scope_all,
            "scope_label": "全部楼栋" if scope_all else self._building_label_from_codes(codes),
            "level": str(fields.get("变更等级") or ""),
            "change_type": str(fields.get("变更类型") or fields.get("变更类别") or ""),
            "plan_start": str(fields.get("计划开始") or fields.get("计划开始时间") or ""),
            "plan_end": str(fields.get("计划结束") or fields.get("计划结束时间") or ""),
            "display_fields": fields,
        }

    @staticmethod
    def _record_work_type(record: dict[str, Any]) -> str:
        return str(record.get("work_type") or WORK_TYPE_MAINTENANCE).strip()

    @staticmethod
    def _record_source_work_type(record: dict[str, Any]) -> str:
        return str(
            record.get("source_work_type")
            or record.get("original_work_type")
            or record.get("work_type")
            or WORK_TYPE_MAINTENANCE
        ).strip()

    def _maintenance_title(self, record: dict[str, Any]) -> str:
        fields = record.get("display_fields") or {}
        building = str(fields.get("楼栋") or "").strip()
        maintenance_total = str(fields.get("维护总项") or "").strip()
        title = f"EA118机房{building}{maintenance_total}" if maintenance_total else ""
        return self._normalize_110_station_notice_title(
            title or str(record.get("title") or record.get("record_id") or "").strip(),
            building=building,
            building_codes=self._building_codes_from_value(building),
        )

    def _work_type_override_title_key(
        self, title: Any, source_work_type: str = WORK_TYPE_MAINTENANCE
    ) -> str:
        canonical = self._canonical_history_notice_title(title, source_work_type)
        if canonical:
            return canonical
        return re.sub(r"[\s,，;；:：。.【】（）()《》<>\"'“”‘’\-－_/\\]+", "", str(title or "")).lower()

    def _work_type_override_key_for_record(self, record: dict[str, Any]) -> str:
        source_work_type = self._record_source_work_type(record)
        if source_work_type == WORK_TYPE_MAINTENANCE:
            return self._work_type_override_title_key(
                self._maintenance_title(record), source_work_type
            )
        return ""

    def _copy_record_as_change_override(
        self, record: dict[str, Any], override: dict[str, Any]
    ) -> dict[str, Any]:
        copied = copy.deepcopy(record)
        fields = dict(copied.get("display_fields") or {})
        title = self._maintenance_title(record)
        fields.setdefault("变更简述", title)
        fields.setdefault("变更楼栋", fields.get("楼栋") or "")
        fields.setdefault("变更进度", self._maintenance_status_value(record) or DEFAULT_MAINTENANCE_STATUS)
        fields.setdefault("专业", fields.get("专业类别") or "")
        copied["display_fields"] = fields
        copied["source_work_type"] = WORK_TYPE_MAINTENANCE
        copied["original_work_type"] = WORK_TYPE_MAINTENANCE
        copied["work_type"] = WORK_TYPE_CHANGE
        copied["notice_type"] = NOTICE_TYPE_CHANGE
        copied["title"] = title
        copied["converted_from_work_type"] = WORK_TYPE_MAINTENANCE
        copied["converted_to_work_type"] = WORK_TYPE_CHANGE
        copied["work_type_override_key"] = str(override.get("override_key") or "")
        copied["work_type_override_title_key"] = str(override.get("normalized_title") or "")
        return copied

    def _apply_work_type_overrides(
        self, records: list[dict[str, Any]]
    ) -> list[dict[str, Any]]:
        overrides = {
            str(item.get("normalized_title") or ""): item
            for item in self._state_store.list_work_type_overrides(
                source_work_type=WORK_TYPE_MAINTENANCE
            )
            if str(item.get("target_work_type") or "") == WORK_TYPE_CHANGE
        }
        if not overrides:
            return records
        converted: list[dict[str, Any]] = []
        for record in records:
            if self._record_source_work_type(record) != WORK_TYPE_MAINTENANCE:
                converted.append(record)
                continue
            title_key = self._work_type_override_key_for_record(record)
            override = overrides.get(title_key)
            if override:
                converted.append(self._copy_record_as_change_override(record, override))
            else:
                converted.append(record)
        return converted

    def _maintenance_record_is_converted_to_change(
        self, record: dict[str, Any]
    ) -> bool:
        converted = self._apply_work_type_overrides([record])
        return bool(
            converted
            and self._record_work_type(converted[0]) == WORK_TYPE_CHANGE
            and self._record_source_work_type(converted[0]) == WORK_TYPE_MAINTENANCE
        )

    def _change_title(self, record: dict[str, Any]) -> str:
        fields = record.get("display_fields") or {}
        return str(
            fields.get("变更简述")
            or record.get("title")
            or record.get("record_id")
            or ""
        ).strip()

    def _change_time_range(self, record: dict[str, Any]) -> tuple[str, str, str]:
        fields = record.get("display_fields") or {}
        start_time = str(
            fields.get("变更开始日期（阿里）")
            or fields.get("计划开始日期（阿里）")
            or fields.get("计划开始")
            or fields.get("计划开始时间")
            or fields.get("计划延迟开始日期")
            or ""
        ).strip()
        end_time = str(
            fields.get("变更结束日期（阿里）")
            or fields.get("计划结束日期（阿里）")
            or fields.get("计划结束")
            or fields.get("计划结束时间")
            or fields.get("计划延迟结束日期")
            or ""
        ).strip()
        time_text = f"{start_time}至{end_time}" if start_time or end_time else ""
        return start_time, end_time, time_text

    @staticmethod
    def _match_text(value: Any) -> str:
        return re.sub(r"\s+", "", str(value or "")).strip()

    @staticmethod
    def _truthy_flag(value: Any) -> bool:
        if isinstance(value, bool):
            return value
        if isinstance(value, (int, float)):
            return bool(value)
        text = str(value or "").strip().lower()
        return text in {"1", "true", "yes", "y", "on", "是", "涉及"}

    @staticmethod
    def _normalize_110_station_notice_title(
        title: Any,
        *,
        building: Any = "",
        building_codes: list[str] | tuple[str, ...] | None = None,
    ) -> str:
        value = str(title or "").strip()
        if not value:
            return value
        target_prefix = "EA118-110KV阿里中天变"
        if value.startswith(target_prefix):
            return value
        codes = {
            str(code or "").strip().upper()
            for code in (building_codes or [])
            if str(code or "").strip()
        }
        building_text = str(building or "").strip()
        looks_like_110 = bool(
            "110" in codes
            or "110站" in building_text
            or re.match(r"^EA118\s*(?:机房)?\s*[-－]?\s*110\s*(?:站|KV)?", value, re.I)
        )
        if not looks_like_110:
            return value
        normalized = re.sub(
            r"^EA118\s*(?:机房)?\s*[-－]?\s*110\s*(?:站|KV)?\s*",
            target_prefix,
            value,
            count=1,
            flags=re.I,
        ).strip()
        return normalized or value

    def _date_keys_from_values(self, *values: Any) -> set[str]:
        keys: set[str] = set()
        for value in values:
            formatted = self._format_source_datetime(value)
            for year, month, day in re.findall(
                r"(\d{4})[-/年](\d{1,2})[-/月](\d{1,2})", formatted
            ):
                keys.add(f"{int(year):04d}-{int(month):02d}-{int(day):02d}")
        return keys

    def _year_month_keys_from_values(self, *values: Any) -> set[str]:
        keys: set[str] = set()
        for value in values:
            formatted = self._format_source_datetime(value)
            if not formatted:
                continue
            for year, month in re.findall(
                r"(\d{4})[-/年](\d{1,2})(?=[-/月\s日]|$)", formatted
            ):
                keys.add(f"{int(year):04d}-{int(month):02d}")
        return keys

    def _maintenance_record_matches_month_window(
        self, record: dict[str, Any], month: str = ""
    ) -> bool:
        label = self._specific_recent_month_label(month)
        if label is None:
            return False
        fields = record.get("display_fields") or {}
        record_month = self._normalize_month_label(fields.get("计划维护月份"))
        if not record_month:
            return False
        if label:
            return record_month == label
        return record_month in set(self._recent_month_labels())

    def _maintenance_record_matches_mop_month_window(
        self, record: dict[str, Any], month: str = ""
    ) -> bool:
        if self._maintenance_record_matches_month_window(record, month):
            return True
        month_key = self._specific_recent_month_key(month)
        if month_key is None:
            return False
        fields = record.get("display_fields") or {}
        record_keys = self._year_month_keys_from_values(
            fields.get("计划开始时间"),
            fields.get("计划结束时间"),
            fields.get("实际开始时间"),
            fields.get("实际结束时间"),
            fields.get("计划延迟开始日期"),
            fields.get("计划延迟结束日期"),
            fields.get("维护开始时间"),
            fields.get("维护完成时间"),
            fields.get("审核确认时间"),
        )
        if month_key:
            return month_key in record_keys
        return bool(record_keys & self._recent_month_keys())

    def _source_record_month_keys(self, record: dict[str, Any]) -> set[str]:
        fields = record.get("display_fields") or {}
        work_type = self._record_work_type(record)
        if work_type == WORK_TYPE_CHANGE:
            start_time, end_time, _ = self._change_time_range(record)
            return self._year_month_keys_from_values(
                start_time,
                end_time,
                fields.get("变更开始日期（阿里）至变更结束日期（阿里）"),
                fields.get("计划开始日期（阿里）"),
                fields.get("计划结束日期（阿里）"),
                fields.get("计划开始"),
                fields.get("计划开始时间"),
                fields.get("计划结束"),
                fields.get("计划结束时间"),
                fields.get("计划延迟开始日期"),
                fields.get("计划延迟结束日期"),
            )
        if work_type == WORK_TYPE_REPAIR:
            return self._year_month_keys_from_values(
                fields.get("维修开始时间"),
                fields.get("故障发生时间"),
                fields.get("发现故障时间"),
                fields.get("期望完成时间"),
                fields.get("维修结束时间"),
                fields.get("维修结束时间（2026）"),
            )
        return set()

    def _source_record_matches_month_window(
        self, record: dict[str, Any], month: str = ""
    ) -> bool:
        work_type = self._record_work_type(record)
        if work_type == WORK_TYPE_MAINTENANCE:
            return self._maintenance_record_matches_month_window(record, month)
        specific_key = self._specific_recent_month_key(month)
        if specific_key is None:
            return False
        record_keys = self._source_record_month_keys(record)
        if not record_keys:
            return False
        if specific_key:
            return specific_key in record_keys
        return bool(record_keys & self._recent_month_keys())

    @staticmethod
    def _target_status_is_finished(status: Any) -> bool:
        return "结束" in str(status or "").strip()

    def _ongoing_hidden_keys(self, item: dict[str, Any]) -> list[str]:
        if not isinstance(item, dict):
            return []
        work_type = self._item_work_type(item)
        active_item_id = str(item.get("active_item_id") or "").strip()
        if active_item_id:
            return [f"{work_type}:active:{active_item_id}"]
        values = {
            "source": str(item.get("source_record_id") or "").strip(),
            "record": str(
                item.get("target_record_id") or item.get("record_id") or ""
            ).strip(),
        }
        return [
            f"{work_type}:{kind}:{value}"
            for kind, value in values.items()
            if value
        ]

    def _load_hidden_ongoing_locked(self) -> dict[str, Any]:
        stored = self._state_store.get_document(STATE_NS_HIDDEN_ONGOING, "global")
        if isinstance(stored, dict):
            hidden = stored.get("hidden")
            if not isinstance(hidden, dict):
                stored["hidden"] = {}
            return stored
        try:
            with self._hidden_ongoing_path.open("r", encoding="utf-8") as fp:
                payload = json.load(fp)
        except FileNotFoundError:
            payload = {}
        except Exception:
            payload = {}
        if not isinstance(payload, dict):
            payload = {}
        hidden = payload.get("hidden")
        if not isinstance(hidden, dict):
            hidden = {}
        payload["hidden"] = hidden
        self._state_store.put_document(STATE_NS_HIDDEN_ONGOING, "global", payload)
        return payload

    def _save_hidden_ongoing_locked(self, payload: dict[str, Any]) -> None:
        payload["updated_at"] = dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        self._state_store.put_document(STATE_NS_HIDDEN_ONGOING, "global", payload)
        self._touch_state_cache_version()

    def _is_ongoing_hidden(self, item: dict[str, Any]) -> bool:
        keys = self._ongoing_hidden_keys(item)
        if not keys:
            return False
        with self._hidden_ongoing_lock:
            payload = self._load_hidden_ongoing_locked()
            hidden = payload.get("hidden") or {}
            return any(key in hidden for key in keys)

    def hide_ongoing_item(
        self, item: dict[str, Any], *, scope: str = "ALL", deleted_by: str = ""
    ) -> dict[str, Any]:
        if not isinstance(item, dict):
            raise PortalError("删除参数格式错误。")
        scope = self._normalize_scope(scope)
        item = copy.deepcopy(item)
        item["work_type"] = self._item_work_type(item)
        item.setdefault(
            "notice_type",
            NOTICE_TYPE_BY_WORK_TYPE.get(item["work_type"], NOTICE_TYPE_MAINTENANCE),
        )
        keys = self.validate_ongoing_delete_item(item, scope=scope)
        now = dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        with self._hidden_ongoing_lock:
            payload = self._load_hidden_ongoing_locked()
            hidden = payload.setdefault("hidden", {})
            for key in keys:
                hidden[key] = {
                    "deleted_at": now,
                    "deleted_by": str(deleted_by or ""),
                    "scope": scope,
                    "work_type": item["work_type"],
                    "notice_type": str(item.get("notice_type") or ""),
                    "title": str(item.get("title") or ""),
                    "source_record_id": str(item.get("source_record_id") or ""),
                    "active_item_id": str(item.get("active_item_id") or ""),
                    "record_id": str(
                        item.get("target_record_id") or item.get("record_id") or ""
                    ),
                }
            self._save_hidden_ongoing_locked(payload)
        return {"deleted": True, "keys": keys, "deleted_at": now}

    def discard_deleted_ongoing_state(
        self, item: dict[str, Any], *, scope: str = "ALL"
    ) -> dict[str, Any]:
        if not isinstance(item, dict):
            return {"work_status_removed": 0, "daily_summary_removed": 0}
        scope = self._normalize_scope(scope)
        item = copy.deepcopy(item)
        item["work_type"] = self._item_work_type(item)
        item.setdefault(
            "notice_type",
            NOTICE_TYPE_BY_WORK_TYPE.get(item["work_type"], NOTICE_TYPE_MAINTENANCE),
        )
        deleted_keys = self._work_status_identity_keys(item)
        if not deleted_keys:
            return {"work_status_removed": 0, "daily_summary_removed": 0}
        work_status_removed = 0
        daily_summary_removed = 0
        with self._summary_lock:
            self._migrate_legacy_work_status_locked()
            for document in self._state_store.list_documents(STATE_NS_WORK_STATUS):
                payload = document.get("payload")
                if not isinstance(payload, dict) or not isinstance(payload.get("items"), list):
                    continue
                kept: list[dict[str, Any]] = []
                changed = False
                for status_item in payload.get("items") or []:
                    if (
                        isinstance(status_item, dict)
                        and self._scope_matches_item(scope, status_item)
                        and (self._work_status_identity_keys(status_item) & deleted_keys)
                    ):
                        work_status_removed += 1
                        changed = True
                        continue
                    kept.append(status_item)
                if changed:
                    payload["items"] = kept
                    payload["updated_at"] = dt.datetime.now().strftime("%Y-%m-%d %H:%M")
                    self._state_store.put_document(
                        STATE_NS_WORK_STATUS, str(document.get("key") or ""), payload
                    )

            summary_payload = self._load_day_summary_locked()
            summary_items = summary_payload.get("items")
            if isinstance(summary_items, list):
                kept_summary: list[dict[str, Any]] = []
                changed = False
                for summary_item in summary_items:
                    if (
                        isinstance(summary_item, dict)
                        and self._scope_matches_item(scope, summary_item)
                        and (self._work_status_identity_keys(summary_item) & deleted_keys)
                    ):
                        daily_summary_removed += 1
                        changed = True
                        continue
                    kept_summary.append(summary_item)
                if changed:
                    summary_payload["items"] = kept_summary
                    summary_payload["updated_at"] = dt.datetime.now().strftime("%Y-%m-%d %H:%M")
                    self._save_day_summary_locked(summary_payload)
            self._work_status_cache_signature = None
            self._work_status_cache_items = None
        return {
            "work_status_removed": work_status_removed,
            "daily_summary_removed": daily_summary_removed,
        }

    @staticmethod
    def _undo_action_label(action_type: str) -> str:
        return {
            "update": "更新",
            "end": "结束",
            "delete": "删除",
        }.get(str(action_type or "").strip().lower(), "操作")

    def _undo_identity_from_context(
        self, context: dict[str, Any], *, action_type: str = ""
    ) -> dict[str, Any]:
        context = normalize_notice_identity_payload(
            context if isinstance(context, dict) else {},
            action=action_type,
        )
        work_type = self._item_work_type(context)
        notice_type = str(
            context.get("notice_type")
            or NOTICE_TYPE_BY_WORK_TYPE.get(work_type)
            or NOTICE_TYPE_MAINTENANCE
        ).strip()
        active_item_id = str(context.get("active_item_id") or "").strip()
        source_record_id = str(context.get("source_record_id") or "").strip()
        target_record_id = canonical_target_record_id(context)
        if not source_record_id:
            source_record_id = str(context.get("source_id") or "").strip()
        title = str(context.get("title") or context.get("name") or "").strip()
        reason = str(context.get("reason") or "").strip()
        building = str(context.get("building") or "").strip()
        building_codes = context.get("building_codes")
        if not isinstance(building_codes, list):
            building_codes = self._building_codes_from_value(building)
        identity_item = {
            "work_type": work_type,
            "notice_type": notice_type,
            "active_item_id": active_item_id,
            "source_record_id": source_record_id,
            "target_record_id": target_record_id,
            "title": title,
            "reason": reason,
            "building": building,
            "building_code": str(context.get("building_code") or "").strip(),
            "building_codes": building_codes,
        }
        identity_keys = self._work_status_identity_keys(identity_item)
        if target_record_id:
            identity_keys.add(f"{work_type}:record:{target_record_id}")
        fallback_key = self._summary_key(
            title=title,
            building=building,
            reason=reason,
            work_type=work_type,
        )
        if fallback_key:
            identity_item["fallback_key"] = fallback_key
            identity_keys.add(f"{work_type}:key:{fallback_key}")
        if active_item_id:
            identity_key = f"{work_type}:active:{active_item_id}"
        elif target_record_id:
            identity_key = f"{work_type}:target:{target_record_id}"
        elif source_record_id:
            identity_key = f"{work_type}:source:{source_record_id}"
        elif fallback_key:
            identity_key = f"{work_type}:key:{fallback_key}"
        else:
            identity_key = f"{work_type}:title:{hashlib.sha1(title.encode('utf-8', errors='ignore')).hexdigest()}"
        return {
            "identity_key": identity_key,
            "identity_item": identity_item,
            "identity_keys": sorted(identity_keys),
            "work_type": work_type,
            "notice_type": notice_type,
            "active_item_id": active_item_id,
            "source_record_id": source_record_id,
            "target_record_id": target_record_id,
            "title": title,
            "reason": reason,
            "building": building,
            "building_codes": building_codes,
        }

    @staticmethod
    def _items_identity_intersects(keys: set[str], item_keys: set[str]) -> bool:
        return bool(keys and item_keys and keys.intersection(item_keys))

    def _find_qt_active_snapshot(self, identity: dict[str, Any]) -> dict[str, Any] | None:
        active_item_id = str(identity.get("active_item_id") or "").strip()
        target_record_id = str(identity.get("target_record_id") or "").strip()
        source_record_id = str(identity.get("source_record_id") or "").strip()
        title = str(identity.get("title") or "").strip()
        for row in self._state_store.list_qt_active_items(include_deleted=True):
            payload = row.get("payload") if isinstance(row.get("payload"), dict) else {}
            if active_item_id and (
                str(row.get("active_item_id") or "") == active_item_id
                or str(payload.get("active_item_id") or "") == active_item_id
            ):
                return copy.deepcopy(row)
            if target_record_id and (
                str(row.get("record_id") or "") == target_record_id
                or str(payload.get("record_id") or "") == target_record_id
                or str(payload.get("target_record_id") or "") == target_record_id
            ):
                return copy.deepcopy(row)
            if source_record_id and str(payload.get("source_record_id") or "") == source_record_id:
                return copy.deepcopy(row)
            if title and title == str(payload.get("title") or "").strip():
                return copy.deepcopy(row)
        return None

    @staticmethod
    def _first_text_value(*values: Any) -> str:
        for value in values:
            text = str(value or "").strip()
            if text:
                return text
        return ""

    def _source_record_building_codes_for_delete(
        self, work_type: str, source_record_id: str, scope: str
    ) -> list[str]:
        source_record_id = str(source_record_id or "").strip()
        if not source_record_id:
            return []
        record = self._source_record_in_scope_snapshot(
            record_id=source_record_id,
            work_type=work_type,
            scope=scope,
        ) or self._source_record_in_scope_snapshot(
            record_id=source_record_id,
            work_type=work_type,
            scope="ALL",
        )
        if not isinstance(record, dict):
            return []
        if work_type == WORK_TYPE_CHANGE:
            return self._change_record_building_codes(record)
        if work_type == WORK_TYPE_REPAIR:
            return self._repair_record_building_codes(record)
        fields = record.get("display_fields") if isinstance(record.get("display_fields"), dict) else {}
        return self._building_codes_from_value(
            record.get("building_codes")
            or record.get("building")
            or fields.get("楼栋")
            or fields.get("所属楼栋")
        )

    def _enrich_ongoing_identity_item(
        self, item: dict[str, Any], *, scope: str = "ALL"
    ) -> dict[str, Any]:
        enriched = normalize_notice_identity_payload(
            copy.deepcopy(item) if isinstance(item, dict) else {}
        )
        enriched["work_type"] = self._item_work_type(enriched)
        enriched.setdefault(
            "notice_type",
            NOTICE_TYPE_BY_WORK_TYPE.get(enriched["work_type"], NOTICE_TYPE_MAINTENANCE),
        )
        work_type = enriched["work_type"]
        identity = self._undo_identity_from_context(enriched, action_type="delete")
        qt_snapshot = self._find_qt_active_snapshot(identity)
        qt_payload = normalize_notice_identity_payload(
            qt_snapshot.get("payload")
            if isinstance(qt_snapshot, dict) and isinstance(qt_snapshot.get("payload"), dict)
            else {}
        )

        def set_text_if_missing(field: str, *values: Any) -> None:
            if str(enriched.get(field) or "").strip():
                return
            value = self._first_text_value(*values)
            if value:
                enriched[field] = value

        set_text_if_missing("active_item_id", qt_payload.get("active_item_id"), (qt_snapshot or {}).get("active_item_id"))
        set_text_if_missing("notice_type", qt_payload.get("notice_type"), (qt_snapshot or {}).get("notice_type"))
        set_text_if_missing("source_record_id", qt_payload.get("source_record_id"))
        set_text_if_missing(
            "target_record_id",
            canonical_target_record_id(qt_payload),
            (qt_snapshot or {}).get("record_id"),
        )
        qt_sections = self._parse_notice_sections(str(qt_payload.get("text") or ""))
        set_text_if_missing(
            "title",
            qt_payload.get("title"),
            self._notice_section_value(qt_sections, ["名称", "标题", "事件描述"]),
            qt_payload.get("content"),
        )
        set_text_if_missing(
            "reason",
            qt_payload.get("reason"),
            self._notice_section_value(qt_sections, ["原因", "故障原因", "故障维修原因"]),
        )
        set_text_if_missing("building", qt_payload.get("building"))
        set_text_if_missing("building_code", qt_payload.get("building_code"))

        building_codes = self._building_codes_from_request_payload(enriched)
        if not building_codes:
            building_codes = self._clean_building_codes(qt_payload.get("building_codes"))
        if not building_codes:
            building_codes = self._building_codes_from_value(
                qt_payload.get("building")
                or qt_payload.get("title")
                or qt_payload.get("content")
            )
        if not building_codes:
            building_codes = self._source_record_building_codes_for_delete(
                work_type,
                str(enriched.get("source_record_id") or "").strip(),
                scope,
            )
        if building_codes:
            enriched["building_codes"] = building_codes
            if not str(enriched.get("building") or "").strip():
                enriched["building"] = self._building_label_from_codes(building_codes)

        target_record_id = canonical_target_record_id(enriched)
        source_record_id = str(enriched.get("source_record_id") or "").strip()
        current_record_id = str(enriched.get("record_id") or "").strip()
        if not target_record_id:
            target_record_id = self._target_record_id_from_identity_map(
                work_type=work_type,
                active_item_id=str(enriched.get("active_item_id") or "").strip(),
                source_record_id=source_record_id,
                target_record_id="",
            )
        if target_record_id:
            enriched["target_record_id"] = target_record_id
            enriched["record_id"] = target_record_id
        elif source_record_id and current_record_id == source_record_id:
            enriched["_source_only_delete"] = True
        return enriched

    def create_notice_undo_checkpoint(
        self,
        action_type: str,
        context: dict[str, Any],
        *,
        remote_fields: dict[str, Any] | None = None,
        remote_missing: bool = False,
        job_id: str = "",
        created_by: str = "",
        scope: str = "ALL",
    ) -> str:
        action_type = str(action_type or "").strip().lower()
        if action_type not in {"update", "end", "delete"}:
            return ""
        context = copy.deepcopy(context) if isinstance(context, dict) else {}
        scope = self._normalize_scope(scope or context.get("scope") or "ALL")
        context = self._enrich_ongoing_identity_item(context, scope=scope)
        identity = self._undo_identity_from_context(context, action_type=action_type)
        identity_keys = set(identity.get("identity_keys") or [])
        if not identity.get("identity_key") or not identity_keys:
            raise PortalError("创建回退点失败：缺少通告身份。")
        qt_active = self._find_qt_active_snapshot(identity)
        daily_item = None
        daily_document_key = ""
        work_items: list[dict[str, Any]] = []
        with self._summary_lock:
            summary_payload = self._load_day_summary_locked()
            for item in summary_payload.get("items") or []:
                if isinstance(item, dict) and self._items_identity_intersects(
                    identity_keys, self._work_status_identity_keys(item)
                ):
                    daily_item = copy.deepcopy(item)
                    daily_document_key = str(summary_payload.get("date") or "").strip()
                    break
            if daily_item is None:
                for document in self._state_store.list_documents(STATE_NS_DAILY_SUMMARY):
                    payload = document.get("payload")
                    if not isinstance(payload, dict):
                        continue
                    for item in payload.get("items") or []:
                        if isinstance(item, dict) and self._items_identity_intersects(
                            identity_keys, self._work_status_identity_keys(item)
                        ):
                            daily_item = copy.deepcopy(item)
                            daily_document_key = str(document.get("key") or "").strip()
                            break
                    if daily_item is not None:
                        break
            self._migrate_legacy_work_status_locked()
            for document in self._state_store.list_documents(STATE_NS_WORK_STATUS):
                payload = document.get("payload")
                if not isinstance(payload, dict):
                    continue
                for item in payload.get("items") or []:
                    if isinstance(item, dict) and self._items_identity_intersects(
                        identity_keys, self._work_status_identity_keys(item)
                    ):
                        work_items.append(
                            {
                                "document_key": str(document.get("key") or ""),
                                "item": copy.deepcopy(item),
                            }
                        )
        now = time.time()
        title = str(identity.get("title") or (daily_item or {}).get("title") or "").strip()
        payload = {
            "undo_id": uuid.uuid4().hex,
            "identity_key": str(identity.get("identity_key") or ""),
            "identity_keys": sorted(identity_keys),
            "status": "available",
            "action_type": action_type,
            "action_label": self._undo_action_label(action_type),
            "scope": scope,
            "work_type": str(identity.get("work_type") or ""),
            "notice_type": str(identity.get("notice_type") or ""),
            "active_item_id": str(identity.get("active_item_id") or ""),
            "source_record_id": str(identity.get("source_record_id") or ""),
            "target_record_id": str(identity.get("target_record_id") or ""),
            "title": title,
            "reason": str(identity.get("reason") or ""),
            "building": str(identity.get("building") or ""),
            "building_codes": list(identity.get("building_codes") or []),
            "created_by": str(created_by or context.get("_auth_open_id") or ""),
            "job_id": str(job_id or ""),
            "created_at": now,
            "expires_at": now + 7 * 24 * 60 * 60,
            "context": context,
            "remote": {
                "missing": bool(remote_missing),
                "fields": copy.deepcopy(remote_fields or {}),
            },
            "local": {
                "qt_active": qt_active,
                "daily_document_key": daily_document_key,
                "daily_item": daily_item,
                "work_items": work_items,
            },
        }
        undo_id = self._state_store.create_notice_undo_action(payload)
        if not undo_id:
            raise PortalError("创建回退点失败：SQLite 写入失败。")
        self._state_store.append_event_async(
            "notice_undo",
            {
                "event": "checkpoint_created",
                "undo_id": undo_id,
                "action_type": action_type,
                "title": title,
                "target_record_id": str(identity.get("target_record_id") or ""),
            },
        )
        return undo_id

    def _undo_key_candidates(self, item: dict[str, Any]) -> set[str]:
        identity = self._undo_identity_from_context(item, action_type="delete")
        return set(identity.get("identity_keys") or [])

    def _available_undo_map(self, scope: str = "ALL") -> dict[str, dict[str, Any]]:
        scope = self._normalize_scope(scope)
        result: dict[str, dict[str, Any]] = {}
        for undo in self._state_store.list_notice_undo_actions(scope=scope):
            enriched_undo = self._enrich_ongoing_identity_item(undo, scope=scope)
            if not self._scope_matches_item(scope, enriched_undo):
                continue
            keys = set(enriched_undo.get("identity_keys") or undo.get("identity_keys") or [])
            for key in keys:
                if key and key not in result:
                    result[key] = enriched_undo
        return result

    @staticmethod
    def _public_undo_fields(undo: dict[str, Any]) -> dict[str, Any]:
        created_at = float((undo or {}).get("created_at") or 0)
        action_type = str((undo or {}).get("action_type") or "")
        label = MaintenancePortalService._undo_action_label(action_type)
        return {
            "undo_available": True,
            "undo_id": str((undo or {}).get("undo_id") or ""),
            "undo_action_type": action_type,
            "undo_created_at": created_at,
            "undo_label": f"回退上一步{label}",
        }

    def _annotate_undo_items(
        self, items: list[dict[str, Any]], *, scope: str = "ALL"
    ) -> list[dict[str, Any]]:
        undo_map = self._available_undo_map(scope)
        annotated: list[dict[str, Any]] = []
        for item in items or []:
            row = copy.deepcopy(item)
            for key in self._undo_key_candidates(row):
                undo = undo_map.get(key)
                if undo:
                    row.update(self._public_undo_fields(undo))
                    break
            annotated.append(row)
        return annotated

    def list_available_notice_undos(
        self,
        *,
        scope: str = "ALL",
        action_type: str = "",
        since_seconds: float = 0,
    ) -> list[dict[str, Any]]:
        scope = self._normalize_scope(scope)
        action_type = str(action_type or "").strip().lower()
        cutoff = time.time() - float(since_seconds or 0) if float(since_seconds or 0) > 0 else 0
        items: list[dict[str, Any]] = []
        for undo in self._state_store.list_notice_undo_actions(scope=scope):
            enriched_undo = self._enrich_ongoing_identity_item(undo, scope=scope)
            if not self._scope_matches_item(scope, enriched_undo):
                continue
            undo_action_type = str(enriched_undo.get("action_type") or "").strip().lower()
            created_at = float(enriched_undo.get("created_at") or 0)
            if action_type and undo_action_type != action_type:
                continue
            if cutoff and created_at < cutoff:
                continue
            items.append(
                {
                    **self._public_undo_fields(enriched_undo),
                    "title": str(enriched_undo.get("title") or ""),
                    "scope": str(enriched_undo.get("scope") or ""),
                    "work_type": str(enriched_undo.get("work_type") or ""),
                    "notice_type": str(enriched_undo.get("notice_type") or ""),
                    "building": str(enriched_undo.get("building") or ""),
                }
            )
        items.sort(key=lambda row: float(row.get("undo_created_at") or 0), reverse=True)
        return items

    def create_notice_undo_job(
        self,
        undo_id: str,
        *,
        scope: str = "ALL",
        auth_open_id: str = "",
        auth_user_name: str = "",
    ) -> str:
        undo = self._state_store.get_notice_undo_action(undo_id)
        if not undo or str(undo.get("status") or "") != "available":
            raise PortalError("该回退记录不可用或已过期。")
        scope = self._normalize_scope(scope or undo.get("scope") or "ALL")
        undo = self._enrich_ongoing_identity_item(undo, scope=scope)
        if not self._scope_matches_item(scope, undo):
            raise PortalError("当前账号无权回退该通告。")
        request_payload = {
            "action": "undo",
            "undo_id": str(undo.get("undo_id") or ""),
            "scope": scope,
            "work_type": str(undo.get("work_type") or WORK_TYPE_MAINTENANCE),
            "notice_type": str(undo.get("notice_type") or ""),
            "active_item_id": str(undo.get("active_item_id") or ""),
            "source_record_id": str(undo.get("source_record_id") or ""),
            "target_record_id": str(undo.get("target_record_id") or ""),
            "_auth_open_id": str(auth_open_id or ""),
            "_auth_user_name": str(auth_user_name or ""),
        }
        job = self._base_job(request_payload)
        job["phase"] = "undo_queued"
        job["operation_id"] = f"undo:{undo_id}"
        job["target_key"] = f"undo:{str(undo.get('identity_key') or undo_id)}"
        with self._jobs_lock:
            for existing in self._jobs.values():
                if str(existing.get("target_key") or "") != job["target_key"]:
                    continue
                if str(existing.get("phase") or "") in {
                    "accepted",
                    "queued",
                    "sending_message",
                    "message_sent",
                    "upload_queued",
                    "qt_queued",
                    "qt_displaying",
                    "upload_waiting",
                    "uploading",
                    "undo_queued",
                    "undoing_remote",
                    "undoing_local",
                }:
                    raise PortalError("该通告正在处理，请稍后再试。")
            self._jobs[job["job_id"]] = job
            self._persist_action_job_locked(job)
            self._trim_jobs_locked()
        return str(job["job_id"])

    def _remove_hidden_ongoing_keys(self, keys: set[str]) -> None:
        if not keys:
            return
        with self._hidden_ongoing_lock:
            payload = self._load_hidden_ongoing_locked()
            hidden = payload.get("hidden") if isinstance(payload.get("hidden"), dict) else {}
            removed = False
            for key in list(keys):
                if key in hidden:
                    hidden.pop(key, None)
                    removed = True
            if removed:
                payload["hidden"] = hidden
                self._save_hidden_ongoing_locked(payload)

    def restore_notice_undo_local(
        self,
        undo: dict[str, Any],
        *,
        target_record_id: str = "",
        applied_by: str = "",
        job_id: str = "",
    ) -> dict[str, Any]:
        if not isinstance(undo, dict):
            raise PortalError("回退记录格式错误。")
        local = undo.get("local") if isinstance(undo.get("local"), dict) else {}
        identity_keys = set(str(key or "") for key in (undo.get("identity_keys") or []) if str(key or ""))
        now = dt.datetime.now().strftime("%Y-%m-%d %H:%M")
        restored_active = False
        removed_active = False
        with self._summary_lock:
            daily_item = local.get("daily_item") if isinstance(local.get("daily_item"), dict) else None
            daily_document_key = str(
                local.get("daily_document_key") or (daily_item or {}).get("date") or ""
            ).strip()
            summary_payload = self._load_day_summary_locked(daily_document_key or None)
            current_summary = summary_payload.get("items")
            if not isinstance(current_summary, list):
                current_summary = []
            summary_payload["items"] = [
                item
                for item in current_summary
                if not (
                    isinstance(item, dict)
                    and self._items_identity_intersects(
                        identity_keys, self._work_status_identity_keys(item)
                    )
                )
            ]
            if daily_item:
                restored_daily = copy.deepcopy(daily_item)
                if target_record_id:
                    restored_daily["target_record_id"] = target_record_id
                summary_payload["items"].append(restored_daily)
            summary_payload["updated_at"] = now
            self._save_day_summary_locked(summary_payload)

            self._migrate_legacy_work_status_locked()
            for document in self._state_store.list_documents(STATE_NS_WORK_STATUS):
                payload = document.get("payload")
                if not isinstance(payload, dict) or not isinstance(payload.get("items"), list):
                    continue
                kept = [
                    item
                    for item in payload.get("items") or []
                    if not (
                        isinstance(item, dict)
                        and self._items_identity_intersects(
                            identity_keys, self._work_status_identity_keys(item)
                        )
                    )
                ]
                if len(kept) != len(payload.get("items") or []):
                    payload["items"] = kept
                    payload["updated_at"] = now
                    self._state_store.put_document(
                        STATE_NS_WORK_STATUS, str(document.get("key") or ""), payload
                    )
            for saved in local.get("work_items") or []:
                if not isinstance(saved, dict) or not isinstance(saved.get("item"), dict):
                    continue
                document_key = str(saved.get("document_key") or "ALL").strip() or "ALL"
                payload = self._state_store.get_document(STATE_NS_WORK_STATUS, document_key)
                if not isinstance(payload, dict):
                    payload = {"version": 1, "items": []}
                if not isinstance(payload.get("items"), list):
                    payload["items"] = []
                item = copy.deepcopy(saved.get("item") or {})
                if target_record_id:
                    item["target_record_id"] = target_record_id
                payload["items"].append(item)
                payload["updated_at"] = now
                self._state_store.put_document(STATE_NS_WORK_STATUS, document_key, payload)
            self._work_status_cache_signature = None
            self._work_status_cache_items = None
        qt_snapshot = local.get("qt_active") if isinstance(local.get("qt_active"), dict) else None
        if qt_snapshot and isinstance(qt_snapshot.get("payload"), dict):
            qt_payload = copy.deepcopy(qt_snapshot.get("payload") or {})
            if target_record_id:
                qt_payload["record_id"] = target_record_id
                qt_payload["target_record_id"] = target_record_id
            restored_active = self._state_store.upsert_qt_active_item(
                qt_payload,
                section=str(qt_snapshot.get("section") or ""),
                sort_order=int(qt_snapshot.get("sort_order") or 0),
                origin=str(qt_snapshot.get("origin") or ""),
            )
        else:
            removed_active = self._state_store.delete_qt_active_item(
                active_item_id=str(undo.get("active_item_id") or ""),
                record_id=str(target_record_id or undo.get("target_record_id") or ""),
            )
        self._remove_hidden_ongoing_keys(identity_keys)
        self._touch_state_cache_version()
        with suppress(Exception):
            self._state_store.append_event(
                "notice_undo",
                {
                    "event": "undo_applied_local",
                    "undo_id": str(undo.get("undo_id") or ""),
                    "job_id": str(job_id or ""),
                    "applied_by": str(applied_by or ""),
                    "target_record_id": str(target_record_id or undo.get("target_record_id") or ""),
                },
            )
        return {
            "restored_active": restored_active,
            "removed_active": removed_active,
            "active_payload": (qt_snapshot or {}).get("payload") if qt_snapshot else {},
        }

    def validate_ongoing_delete_item(
        self, item: dict[str, Any], *, scope: str = "ALL"
    ) -> list[str]:
        if not isinstance(item, dict):
            raise PortalError("删除参数格式错误。")
        scope = self._normalize_scope(scope)
        enriched = self._enrich_ongoing_identity_item(item, scope=scope)
        item.clear()
        item.update(enriched)
        if not self._scope_matches_item(scope, enriched):
            raise PortalError("当前账号无权删除该楼栋的进行中通告。")
        keys = self._ongoing_hidden_keys(enriched)
        if not keys:
            raise PortalError("该进行中通告缺少可删除身份。")
        return keys

    def _target_record_is_finished(
        self, *, work_type: str, notice_type: str, record_id: str
    ) -> bool:
        record_id = str(record_id or "").strip()
        if not record_id:
            return False
        field_config = get_field_config(notice_type)
        status_field = field_config.get("status", "")
        if not status_field:
            return False
        try:
            target_records = self._target_records_for_notice_type(notice_type, work_type)
        except Exception as exc:
            warning = f"{notice_type}目标表完成状态检查失败: {exc}"
            if warning not in self._load_warnings:
                self._load_warnings.append(warning)
            return False
        for target in target_records:
            if str(target.get("record_id") or "").strip() != record_id:
                continue
            fields = target.get("display_fields") or {}
            return self._target_status_is_finished(fields.get(status_field))
        return False

    def _target_record_building_codes(
        self, fields: dict[str, Any], field_config: dict[str, str]
    ) -> list[str]:
        building_field = (
            field_config.get("building")
            or field_config.get("building_codes")
            or "楼栋"
        )
        return self._building_codes_from_value(fields.get(building_field))

    def _source_target_match_profile(
        self, work_type: str, record: dict[str, Any]
    ) -> tuple[str, set[str], list[str]]:
        fields = record.get("display_fields") or {}
        if work_type == WORK_TYPE_CHANGE:
            start_time, end_time, _ = self._change_time_range(record)
            return (
                self._change_title(record),
                self._date_keys_from_values(start_time, end_time),
                self._change_record_building_codes(record),
            )
        if work_type == WORK_TYPE_REPAIR:
            return (
                self._repair_title(record),
                self._date_keys_from_values(
                    fields.get("维修开始时间"),
                    fields.get("故障发生时间"),
                    fields.get("期望完成时间"),
                    fields.get("维修结束时间"),
                    fields.get("维修结束时间（2026）"),
                ),
                self._repair_record_building_codes(record),
            )
        building = str(fields.get("楼栋") or "").strip()
        title = f"EA118机房{building}{str(fields.get('维护总项') or '').strip()}"
        return (
            title,
            self._date_keys_from_values(
                fields.get("计划开始时间"),
                fields.get("计划结束时间"),
                fields.get("实际开始时间"),
            ),
            self._building_codes_from_value(building),
        )

    def _target_match_date_fields(
        self, work_type: str, field_config: dict[str, str]
    ) -> list[str]:
        if work_type == WORK_TYPE_CHANGE:
            return [
                field_config.get("start_time", ""),
                field_config.get("end_time", ""),
                field_config.get("update_time", ""),
            ]
        if work_type == WORK_TYPE_REPAIR:
            return [
                field_config.get("actual_start", ""),
                field_config.get("fault_time", ""),
                field_config.get("expected_time", ""),
            ]
        if work_type in {WORK_TYPE_POWER, WORK_TYPE_POLLING, WORK_TYPE_ADJUST}:
            return [
                field_config.get("plan_start", ""),
                field_config.get("plan_end", ""),
                field_config.get("actual_start", ""),
                field_config.get("actual_end", ""),
            ]
        if work_type == WORK_TYPE_EVENT:
            return [
                field_config.get("occurrence_time", ""),
                field_config.get("response_time", ""),
                field_config.get("recover_time", ""),
                field_config.get("end_time", ""),
            ]
        return [
            field_config.get("actual_start", ""),
            field_config.get("plan_start", ""),
            field_config.get("plan_end", ""),
        ]

    def _target_records_for_notice_type(
        self, notice_type: str, work_type: str, *, force_refresh: bool = False
    ) -> list[dict[str, Any]]:
        table_id = str(config.get_table_id(notice_type) or "").strip()
        app_token = str(config.app_token or "").strip()
        if not app_token or not table_id:
            return []
        cache_key = (notice_type, table_id)
        with self._refresh_lock:
            cached = self._target_record_cache.get(cache_key) or {}
            if (
                not force_refresh
                and
                cached.get("records") is not None
                and time.time() - float(cached.get("loaded_ts") or 0)
                < self._source_cache_ttl_seconds()
            ):
                return list(cached.get("records") or [])
            _metas, meta_by_name = self._load_table_fields(
                app_token=app_token, table_id=table_id
            )
            records = self._load_table_records(
                app_token=app_token,
                table_id=table_id,
                meta_by_name=meta_by_name,
                work_type=work_type,
                notice_type=notice_type,
            )
            self._target_record_cache[cache_key] = {
                "loaded_ts": time.time(),
                "records": records,
            }
            return list(records)

    @classmethod
    def _normalize_notice_work_type_alias(cls, work_type: Any) -> str:
        text = str(work_type or "").strip()
        aliases = {
            "maintenance": WORK_TYPE_MAINTENANCE,
            "维保": WORK_TYPE_MAINTENANCE,
            "维保通告": WORK_TYPE_MAINTENANCE,
            "change": WORK_TYPE_CHANGE,
            "变更": WORK_TYPE_CHANGE,
            "变更通告": WORK_TYPE_CHANGE,
            "repair": WORK_TYPE_REPAIR,
            "检修": WORK_TYPE_REPAIR,
            "设备检修": WORK_TYPE_REPAIR,
            "power": WORK_TYPE_POWER,
            "上电": WORK_TYPE_POWER,
            "上电通告": WORK_TYPE_POWER,
            "下电": WORK_TYPE_POWER,
            "下电通告": WORK_TYPE_POWER,
            "上下电通告": WORK_TYPE_POWER,
            "polling": WORK_TYPE_POLLING,
            "轮巡": WORK_TYPE_POLLING,
            "设备轮巡": WORK_TYPE_POLLING,
            "adjust": WORK_TYPE_ADJUST,
            "调整": WORK_TYPE_ADJUST,
            "设备调整": WORK_TYPE_ADJUST,
            "event": WORK_TYPE_EVENT,
            "事件": WORK_TYPE_EVENT,
            "事件通告": WORK_TYPE_EVENT,
        }
        return aliases.get(text.lower()) or aliases.get(text) or WORK_TYPE_MAINTENANCE

    def validate_notice_identity_binding(
        self,
        *,
        scope: str,
        work_type: str,
        notice_type: str = "",
        source_record_id: str = "",
        target_record_id: str = "",
        active_item_id: str = "",
    ) -> dict[str, Any]:
        scope = self._normalize_scope(scope)
        work_type = self._normalize_notice_work_type_alias(work_type)
        notice_type = str(notice_type or self._notice_type_for_work_type(work_type)).strip()
        source_record_id = str(source_record_id or "").strip()
        target_record_id = str(target_record_id or "").strip()
        active_item_id = str(active_item_id or "").strip()
        if source_record_id and is_local_record_id(source_record_id):
            source_record_id = ""
        if target_record_id and is_local_record_id(target_record_id):
            target_record_id = ""
        if not (source_record_id or target_record_id or active_item_id):
            raise PortalError("缺少可绑定的源表记录、目标记录或本地进行中条目。")

        source_record: dict[str, Any] | None = None
        if source_record_id:
            source_record = self._source_record_in_scope_snapshot(
                record_id=source_record_id,
                work_type=work_type,
                scope=scope,
            )
            if source_record is None:
                raise PortalError("源表事项不属于当前楼栋或当前通告类型，请重新选择。")

        target_record: dict[str, Any] | None = None
        target_building_codes: list[str] = []
        if target_record_id:
            try:
                target_records = self._target_records_for_notice_type(
                    notice_type,
                    work_type,
                    force_refresh=False,
                )
            except Exception as exc:
                raise PortalError(f"查询{notice_type}目标表失败：{exc}") from exc
            target_record = next(
                (
                    record
                    for record in target_records
                    if str(record.get("record_id") or "").strip() == target_record_id
                ),
                None,
            )
            if target_record is None:
                try:
                    target_records = self._target_records_for_notice_type(
                        notice_type,
                        work_type,
                        force_refresh=True,
                    )
                except Exception as exc:
                    raise PortalError(f"查询{notice_type}目标表失败：{exc}") from exc
                target_record = next(
                    (
                        record
                        for record in target_records
                        if str(record.get("record_id") or "").strip() == target_record_id
                    ),
                    None,
                )
            if target_record is None:
                raise PortalError("目标多维记录不存在，请重新查找并选择正确记录。")
            fields = target_record.get("display_fields") or {}
            field_config = get_field_config(notice_type)
            target_building_codes = self._target_record_building_codes(fields, field_config)
            if not target_building_codes:
                target_building_codes = self._clean_building_codes(
                    target_record.get("building_codes")
                ) or self._building_codes_from_value(target_record.get("building"))
            if target_building_codes and not self._scope_matches_buildings(
                scope, target_building_codes
            ):
                raise PortalError("目标多维记录不属于当前楼栋，请重新选择。")

        return {
            "scope": scope,
            "work_type": work_type,
            "notice_type": notice_type,
            "source_record_id": source_record_id,
            "target_record_id": target_record_id,
            "active_item_id": active_item_id,
            "source_found": bool(source_record),
            "target_found": bool(target_record),
            "target_building_codes": target_building_codes,
            "target_status": str(
                ((target_record or {}).get("display_fields") or {}).get(
                    get_field_config(notice_type).get("status", "")
                )
                or ""
            ),
        }

    def _resolve_target_record_reference(
        self,
        *,
        work_type: str,
        notice_type: str,
        source_record: dict[str, Any],
        include_finished: bool = False,
        only_finished: bool = False,
    ) -> str:
        field_config = get_field_config(notice_type)
        title_field = (
            field_config.get("title")
            or field_config.get("name")
            or "名称"
        )
        status_field = field_config.get("status", "")
        source_title, source_dates, source_codes = self._source_target_match_profile(
            work_type, source_record
        )
        source_title_key = self._match_text(source_title)
        if not source_title_key or not source_dates:
            return ""
        date_fields = [
            field_name
            for field_name in self._target_match_date_fields(work_type, field_config)
            if field_name
        ]
        matches: list[str] = []
        try:
            target_records = self._target_records_for_notice_type(notice_type, work_type)
        except Exception as exc:
            self._load_warnings = list(self._load_warnings or [])
            warning = f"{notice_type}目标表自动引用失败: {exc}"
            if warning not in self._load_warnings:
                self._load_warnings.append(warning)
            return ""
        for target in target_records:
            fields = target.get("display_fields") or {}
            target_finished = bool(
                status_field and self._target_status_is_finished(fields.get(status_field))
            )
            if only_finished and not target_finished:
                continue
            if target_finished and not include_finished and not only_finished:
                continue
            if self._match_text(fields.get(title_field)) != source_title_key:
                continue
            target_dates = self._date_keys_from_values(
                *[fields.get(field_name) for field_name in date_fields]
            )
            if not target_dates or not (source_dates & target_dates):
                continue
            target_codes = self._target_record_building_codes(fields, field_config)
            if source_codes and target_codes and source_codes != target_codes:
                continue
            record_id = str(target.get("record_id") or "").strip()
            if record_id:
                matches.append(record_id)
        return matches[0] if len(set(matches)) == 1 else ""

    def lookup_change_target_candidates(
        self,
        *,
        scope: str,
        title: str,
        start_time: str = "",
        end_time: str = "",
        action: str = "update",
        content: str = "",
        reason: str = "",
        impact: str = "",
        progress: str = "",
        text: str = "",
        limit: int = 30,
    ) -> dict[str, Any]:
        scope = self._normalize_scope(scope)
        title = str(title or "").strip()
        if not title:
            raise PortalError("变更通告缺少【名称】，无法查询目标记录。")
        self.ensure_snapshot_loaded()
        title_key = self._match_text(title)
        query_dates = self._date_keys_from_values(start_time, end_time)
        field_config = get_field_config(NOTICE_TYPE_CHANGE)
        title_field = field_config.get("title") or field_config.get("name") or "名称"
        status_field = field_config.get("status", "")
        date_fields = [
            field_name
            for field_name in self._target_match_date_fields(WORK_TYPE_CHANGE, field_config)
            if field_name
        ]
        lookup_text_values = [content, reason, impact, progress]
        if not any(str(value or "").strip() for value in lookup_text_values):
            lookup_text_values.append(text)
        try:
            target_records = self._target_records_for_notice_type(
                NOTICE_TYPE_CHANGE,
                WORK_TYPE_CHANGE,
                force_refresh=True,
            )
        except Exception as exc:
            raise PortalError(f"查询变更通告目标表失败：{exc}") from exc
        candidates: list[dict[str, Any]] = []
        for target in target_records:
            fields = target.get("display_fields") or {}
            target_title = str(fields.get(title_field) or "").strip()
            title_matched = self._source_candidate_title_matches(title, target_title, WORK_TYPE_CHANGE)
            business_match_count = self._target_business_match_count(
                fields,
                lookup_text_values,
            ) + self._target_named_business_match_count(
                fields,
                {
                    "title": title,
                    "content": content,
                    "reason": reason,
                    "impact": impact,
                    "progress": progress,
                },
                field_config,
            )
            if not title_matched and business_match_count <= 0:
                continue
            target_dates = self._date_keys_from_values(
                *[fields.get(field_name) for field_name in date_fields],
                fields.get("时间"),
                fields.get("计划时间"),
            )
            building_codes = self._target_record_building_codes(fields, field_config)
            if building_codes and not self._scope_matches_buildings(scope, building_codes):
                continue
            record_id = str(target.get("record_id") or "").strip()
            if not record_id:
                continue
            target_start = str(fields.get(field_config.get("start_time", "")) or "").strip()
            target_end = str(fields.get(field_config.get("end_time", "")) or "").strip()
            status = str(fields.get(status_field) or "").strip() if status_field else ""
            detail_fields = self._target_candidate_detail_fields(fields)
            date_matched = bool(query_dates and target_dates and (query_dates & target_dates))
            match_reason = self._target_candidate_match_reason(
                date_matched=date_matched,
                title_matched=title_matched,
                business_match_count=business_match_count,
                building_matched=bool(building_codes),
                status=status,
            )
            candidates.append(
                {
                    "record_id": record_id,
                    "target_record_id": record_id,
                    "title": target_title or title,
                    "building": self._building_label_from_codes(building_codes),
                    "building_codes": building_codes,
                    "status": status,
                    "start_time": target_start,
                    "end_time": target_end,
                    "date_matched": date_matched,
                    "title_matched": title_matched,
                    "business_text_matched": business_match_count > 0,
                    "business_match_count": business_match_count,
                    "match_reason": match_reason,
                    "fields": detail_fields,
                    "field_items": [
                        {"label": key, "value": value}
                        for key, value in detail_fields.items()
                    ],
                }
            )
        candidates.sort(
            key=lambda item: (
                0 if item.get("date_matched") else 1,
                0 if str(item.get("status") or "") != "结束" else 1,
                str(item.get("start_time") or ""),
            )
        )
        candidates, result_meta = self._candidate_result_meta(candidates, limit)
        return {
            "scope": scope,
            "action": str(action or "update").strip().lower(),
            "title": title,
            "start_time": str(start_time or "").strip(),
            "end_time": str(end_time or "").strip(),
            **result_meta,
            "candidates": candidates,
            "source_candidates": self._lookup_change_source_candidates(
                scope=scope,
                title=title,
                start_time=start_time,
                end_time=end_time,
                limit=limit,
            ),
        }

    def lookup_notice_target_candidates(
        self,
        *,
        work_type: str,
        scope: str,
        title: str,
        start_time: str = "",
        end_time: str = "",
        action: str = "update",
        content: str = "",
        reason: str = "",
        impact: str = "",
        progress: str = "",
        text: str = "",
        limit: int = 30,
    ) -> dict[str, Any]:
        work_type = str(work_type or WORK_TYPE_MAINTENANCE).strip()
        aliases = {
            "maintenance": WORK_TYPE_MAINTENANCE,
            "维保": WORK_TYPE_MAINTENANCE,
            "维保通告": WORK_TYPE_MAINTENANCE,
            "change": WORK_TYPE_CHANGE,
            "变更": WORK_TYPE_CHANGE,
            "变更通告": WORK_TYPE_CHANGE,
            "repair": WORK_TYPE_REPAIR,
            "检修": WORK_TYPE_REPAIR,
            "设备检修": WORK_TYPE_REPAIR,
            "power": WORK_TYPE_POWER,
            "上电": WORK_TYPE_POWER,
            "上电通告": WORK_TYPE_POWER,
            "下电": WORK_TYPE_POWER,
            "下电通告": WORK_TYPE_POWER,
            "上下电通告": WORK_TYPE_POWER,
            "polling": WORK_TYPE_POLLING,
            "轮巡": WORK_TYPE_POLLING,
            "设备轮巡": WORK_TYPE_POLLING,
            "adjust": WORK_TYPE_ADJUST,
            "调整": WORK_TYPE_ADJUST,
            "设备调整": WORK_TYPE_ADJUST,
            "event": WORK_TYPE_EVENT,
            "事件": WORK_TYPE_EVENT,
            "事件通告": WORK_TYPE_EVENT,
        }
        work_type = aliases.get(work_type.lower()) or aliases.get(work_type) or WORK_TYPE_MAINTENANCE
        if work_type == WORK_TYPE_CHANGE:
            return self.lookup_change_target_candidates(
                scope=scope,
                title=title,
                start_time=start_time,
                end_time=end_time,
                action=action,
                content=content,
                reason=reason,
                impact=impact,
                progress=progress,
                text=text,
                limit=limit,
            )
        notice_type = (
            NOTICE_TYPE_REPAIR
            if work_type == WORK_TYPE_REPAIR
            else NOTICE_TYPE_POWER
            if work_type == WORK_TYPE_POWER
            else NOTICE_TYPE_POLLING
            if work_type == WORK_TYPE_POLLING
            else NOTICE_TYPE_ADJUST
            if work_type == WORK_TYPE_ADJUST
            else NOTICE_TYPE_EVENT
            if work_type == WORK_TYPE_EVENT
            else NOTICE_TYPE_MAINTENANCE
        )
        scope = self._normalize_scope(scope)
        title = str(title or "").strip()
        if not title:
            raise PortalError(f"{self._history_work_type_label(work_type)}通告缺少标题，无法查询目标记录。")
        title_key = self._match_text(title)
        query_dates = self._date_keys_from_values(start_time, end_time)
        field_config = get_field_config(notice_type)
        title_field = (
            field_config.get("alarm_desc")
            if work_type == WORK_TYPE_EVENT
            else ""
        ) or (
            field_config.get("title")
            or field_config.get("name")
            or "名称"
        )
        status_field = field_config.get("status", "")
        date_fields = [
            field_name
            for field_name in self._target_match_date_fields(work_type, field_config)
            if field_name
        ]
        lookup_text_values = [content, reason, impact, progress]
        if not any(str(value or "").strip() for value in lookup_text_values):
            lookup_text_values.append(text)
        try:
            target_records = self._target_records_for_notice_type(
                notice_type,
                work_type,
                force_refresh=True,
            )
        except Exception as exc:
            raise PortalError(f"查询{notice_type}目标表失败：{exc}") from exc
        candidates: list[dict[str, Any]] = []
        for target in target_records:
            fields = target.get("display_fields") or {}
            target_title = str(
                fields.get(title_field)
                or fields.get("告警描述")
                or fields.get("概述")
                or fields.get("名称")
                or fields.get("标题")
                or fields.get("名称（标题）")
                or fields.get("检修通告名称")
                or fields.get("维修名称")
                or ""
            ).strip()
            title_matched = self._source_candidate_title_matches(title, target_title, work_type)
            business_match_count = self._target_business_match_count(
                fields,
                lookup_text_values,
            ) + self._target_named_business_match_count(
                fields,
                {
                    "title": title,
                    "content": content,
                    "reason": reason,
                    "impact": impact,
                    "progress": progress,
                },
                field_config,
            )
            if not title_matched and business_match_count <= 0:
                continue
            building_codes = self._target_record_building_codes(fields, field_config)
            if building_codes and not self._scope_matches_buildings(scope, building_codes):
                continue
            record_id = str(target.get("record_id") or "").strip()
            if not record_id:
                continue
            target_dates = self._date_keys_from_values(
                *[fields.get(field_name) for field_name in date_fields],
                fields.get("时间"),
                fields.get("计划时间"),
            )
            target_start = str(
                fields.get(field_config.get("occurrence_time", ""))
                or fields.get(field_config.get("plan_start", ""))
                or fields.get(field_config.get("actual_start", ""))
                or fields.get(field_config.get("start_time", ""))
                or fields.get(field_config.get("expected_time", ""))
                or ""
            ).strip()
            target_end = str(
                fields.get(field_config.get("end_time", ""))
                or fields.get(field_config.get("recover_time", ""))
                or fields.get(field_config.get("plan_end", ""))
                or fields.get(field_config.get("actual_end", ""))
                or fields.get(field_config.get("fault_time", ""))
                or ""
            ).strip()
            status = str(fields.get(status_field) or "").strip() if status_field else ""
            if work_type == WORK_TYPE_EVENT:
                status = self._event_status(fields)
            detail_fields = self._target_candidate_detail_fields(fields)
            date_matched = bool(query_dates and target_dates and (query_dates & target_dates))
            match_reason = self._target_candidate_match_reason(
                date_matched=date_matched,
                title_matched=title_matched,
                business_match_count=business_match_count,
                building_matched=bool(building_codes),
                status=status,
            )
            candidates.append(
                {
                    "record_id": record_id,
                    "target_record_id": record_id,
                    "work_type": work_type,
                    "notice_type": notice_type,
                    "title": target_title or title,
                    "building": self._building_label_from_codes(building_codes),
                    "building_codes": building_codes,
                    "status": status,
                    "start_time": target_start,
                    "end_time": target_end,
                    "date_matched": date_matched,
                    "title_matched": title_matched,
                    "business_text_matched": business_match_count > 0,
                    "business_match_count": business_match_count,
                    "match_reason": match_reason,
                    "fields": detail_fields,
                    "field_items": [
                        {"label": key, "value": value}
                        for key, value in detail_fields.items()
                    ],
                }
            )
        candidates.sort(
            key=lambda item: (
                0 if item.get("date_matched") else 1,
                0 if not self._target_status_is_finished(item.get("status")) else 1,
                str(item.get("start_time") or ""),
            )
        )
        candidates, result_meta = self._candidate_result_meta(candidates, limit)
        return {
            "scope": scope,
            "work_type": work_type,
            "notice_type": notice_type,
            "action": str(action or "update").strip().lower(),
            "title": title,
            "start_time": str(start_time or "").strip(),
            "end_time": str(end_time or "").strip(),
            **result_meta,
            "candidates": candidates,
            "source_candidates": self._lookup_notice_source_candidates(
                work_type=work_type,
                scope=scope,
                title=title,
                start_time=start_time,
                end_time=end_time,
                limit=limit,
            ),
        }

    @staticmethod
    def _target_candidate_match_reason(
        *,
        date_matched: bool,
        title_matched: bool,
        business_match_count: int,
        building_matched: bool,
        status: str,
    ) -> str:
        reasons: list[str] = []
        if date_matched:
            reasons.append("时间匹配")
        if title_matched:
            reasons.append("标题匹配")
        if business_match_count > 0:
            reasons.append(f"内容字段匹配 {business_match_count} 项")
        if building_matched:
            reasons.append("楼栋匹配")
        if str(status or "").strip() not in {"结束", "已结束"}:
            reasons.append("未结束优先")
        return " / ".join(reasons) or "相似记录"

    @staticmethod
    def _candidate_result_limit(limit: Any, *, default: int = 30, maximum: int = 80) -> int:
        try:
            value = int(limit if limit not in (None, "") else default)
        except Exception:
            value = default
        return max(1, min(value, maximum))

    @classmethod
    def _candidate_result_meta(cls, candidates: list, limit: Any) -> tuple[list, dict[str, Any]]:
        limit_value = cls._candidate_result_limit(limit)
        total_matched = len(candidates)
        visible = candidates[:limit_value]
        return visible, {
            "count": len(visible),
            "returned_count": len(visible),
            "total_matched": total_matched,
            "limit": limit_value,
            "limited": total_matched > len(visible),
        }

    def _target_business_match_count(
        self,
        fields: dict[str, Any],
        lookup_values: list[Any],
    ) -> int:
        if not isinstance(fields, dict):
            return 0
        blob = self._match_text(
            " ".join(self._clean_source_text(value) for value in fields.values())
        )
        if not blob:
            return 0
        matched = 0
        seen: set[str] = set()
        for value in lookup_values or []:
            key = self._match_text(value)
            if not key or key in seen:
                continue
            seen.add(key)
            if len(key) < 4:
                continue
            if key in blob or (len(key) >= 12 and blob in key):
                matched += 1
        return matched

    def _target_named_business_match_count(
        self,
        fields: dict[str, Any],
        lookup_values: dict[str, Any],
        field_config: dict[str, Any],
    ) -> int:
        if not isinstance(fields, dict) or not isinstance(lookup_values, dict):
            return 0
        aliases = {
            "title": [
                field_config.get("alarm_desc"),
                field_config.get("title"),
                field_config.get("name"),
                "告警描述",
                "概述",
                "名称",
                "名称（标题）",
                "标题",
            ],
            "content": [field_config.get("content"), "内容", "工作内容"],
            "reason": [field_config.get("reason"), "原因", "故障原因"],
            "impact": [field_config.get("impact"), "影响", "影响范围"],
            "progress": [field_config.get("progress"), "进度", "完成情况"],
        }
        matched = 0
        for key, names in aliases.items():
            source_value = str(lookup_values.get(key) or "").strip()
            if not source_value:
                continue
            source_key = self._match_text(source_value)
            if len(source_key) < 4:
                continue
            target_values = []
            for name in names:
                name = str(name or "").strip()
                if name and name in fields:
                    target_values.append(fields.get(name))
            for target_value in target_values:
                target_key = self._match_text(self._clean_source_text(target_value))
                if not target_key:
                    continue
                if source_key in target_key or target_key in source_key:
                    matched += 1
                    break
        return matched

    def _target_candidate_detail_fields(self, fields: dict[str, Any]) -> dict[str, str]:
        if not isinstance(fields, dict):
            return {}
        detail_fields: dict[str, str] = {}
        for key, value in fields.items():
            label = str(key or "").strip()
            if not label:
                continue
            text = self._clean_source_text(value)
            if not text:
                continue
            if len(text) > 500:
                text = text[:500] + "..."
            detail_fields[label] = text
        return detail_fields

    def _lookup_change_source_candidates(
        self,
        *,
        scope: str,
        title: str,
        start_time: str = "",
        end_time: str = "",
        limit: int = 30,
    ) -> list[dict[str, Any]]:
        title_key = self._match_text(title)
        if not title_key:
            return []
        query_dates = self._date_keys_from_values(start_time, end_time)
        candidates: list[dict[str, Any]] = []
        for record in list(self._change_records or []):
            if not isinstance(record, dict):
                continue
            source_title = self._change_title(record)
            if not self._source_candidate_title_matches(title, source_title, WORK_TYPE_CHANGE):
                continue
            building_codes = self._change_record_building_codes(record)
            if not self._scope_matches_buildings(scope, building_codes):
                continue
            source_start, source_end, _ = self._change_time_range(record)
            source_dates = self._date_keys_from_values(source_start, source_end)
            fields = record.get("display_fields") or {}
            detail_fields = self._target_candidate_detail_fields(fields)
            candidates.append(
                {
                    "record_id": str(record.get("record_id") or ""),
                    "source_record_id": str(record.get("record_id") or ""),
                    "source_app_token": CHANGE_SOURCE_APP_TOKEN,
                    "source_table_id": CHANGE_SOURCE_TABLE_ID,
                    "title": source_title or title,
                    "building": self._building_label_from_codes(building_codes),
                    "building_codes": building_codes,
                    "status": self._change_progress_value(record),
                    "start_time": source_start,
                    "end_time": source_end,
                    "date_matched": bool(query_dates and source_dates and (query_dates & source_dates)),
                    "fields": detail_fields,
                    "field_items": [
                        {"label": key, "value": value}
                        for key, value in detail_fields.items()
                    ],
                }
            )
        candidates.sort(
            key=lambda item: (
                0 if item.get("date_matched") else 1,
                0 if str(item.get("status") or "") != CHANGE_PROGRESS_ENDED else 1,
                str(item.get("start_time") or ""),
            )
        )
        return candidates[: max(1, int(limit or 30))]

    def _source_records_for_notice_lookup(self, work_type: str) -> list[dict[str, Any]]:
        if work_type == WORK_TYPE_REPAIR:
            primary = list(self._repair_records or [])
        elif work_type == WORK_TYPE_MAINTENANCE:
            primary = list(self._records or [])
        else:
            return []
        by_id: dict[str, dict[str, Any]] = {}
        for record in primary + list(self._source_snapshot_records("ALL") or []):
            if not isinstance(record, dict):
                continue
            if self._record_work_type(record) != work_type:
                continue
            record_id = str(record.get("record_id") or "").strip()
            if record_id and record_id not in by_id:
                by_id[record_id] = record
        return list(by_id.values())

    def _source_candidate_title(self, record: dict[str, Any], work_type: str) -> str:
        fields = record.get("display_fields") or {}
        if work_type == WORK_TYPE_REPAIR:
            return self._repair_title(record)
        if work_type == WORK_TYPE_MAINTENANCE:
            explicit = str(record.get("title") or fields.get("名称") or fields.get("标题") or "").strip()
            if explicit:
                return explicit
            building = str(fields.get("楼栋") or "").strip()
            total = str(fields.get("维护总项") or "").strip()
            return f"EA118机房{building}{total}".strip()
        return str(record.get("title") or record.get("record_id") or "").strip()

    def _source_candidate_building_codes(self, record: dict[str, Any], work_type: str) -> list[str]:
        if work_type == WORK_TYPE_REPAIR:
            return self._repair_record_building_codes(record)
        if work_type == WORK_TYPE_MAINTENANCE:
            return self._building_codes_from_value((record.get("display_fields") or {}).get("楼栋"))
        return []

    def _source_candidate_status(self, record: dict[str, Any], work_type: str) -> str:
        if work_type == WORK_TYPE_REPAIR:
            return self._repair_source_status(record)
        if work_type == WORK_TYPE_MAINTENANCE:
            return self._maintenance_status_value(record)
        return ""

    def _source_candidate_time_range(self, record: dict[str, Any], work_type: str) -> tuple[str, str]:
        fields = record.get("display_fields") or {}
        if work_type == WORK_TYPE_REPAIR:
            start, end, _ = self._repair_time_range(record)
            return start, end
        if work_type == WORK_TYPE_MAINTENANCE:
            return str(fields.get("计划维护月份") or "").strip(), ""
        return "", ""

    def _source_candidate_title_matches(self, title: str, candidate_title: str, work_type: str) -> bool:
        if self._match_text(candidate_title) == self._match_text(title):
            return True
        return (
            self._canonical_history_notice_title(candidate_title, work_type)
            == self._canonical_history_notice_title(title, work_type)
        )

    def _lookup_notice_source_candidates(
        self,
        *,
        work_type: str,
        scope: str,
        title: str,
        start_time: str = "",
        end_time: str = "",
        limit: int = 30,
    ) -> list[dict[str, Any]]:
        if work_type == WORK_TYPE_CHANGE:
            return self._lookup_change_source_candidates(
                scope=scope,
                title=title,
                start_time=start_time,
                end_time=end_time,
                limit=limit,
            )
        if work_type not in {WORK_TYPE_MAINTENANCE, WORK_TYPE_REPAIR}:
            return []
        title = str(title or "").strip()
        if not title:
            return []
        query_dates = self._date_keys_from_values(start_time, end_time)
        source_app_token = self.app_token if work_type == WORK_TYPE_MAINTENANCE else REPAIR_SOURCE_APP_TOKEN
        source_table_id = self.table_id if work_type == WORK_TYPE_MAINTENANCE else REPAIR_SOURCE_TABLE_ID
        candidates: list[dict[str, Any]] = []
        for record in self._source_records_for_notice_lookup(work_type):
            source_title = self._source_candidate_title(record, work_type)
            if not self._source_candidate_title_matches(title, source_title, work_type):
                continue
            building_codes = self._source_candidate_building_codes(record, work_type)
            if not self._scope_matches_buildings(scope, building_codes):
                continue
            source_start, source_end = self._source_candidate_time_range(record, work_type)
            source_dates = self._date_keys_from_values(source_start, source_end)
            fields = record.get("display_fields") or {}
            detail_fields = self._target_candidate_detail_fields(fields)
            record_id = str(record.get("record_id") or "").strip()
            if not record_id:
                continue
            candidates.append(
                {
                    "record_id": record_id,
                    "source_record_id": record_id,
                    "source_app_token": source_app_token,
                    "source_table_id": source_table_id,
                    "work_type": work_type,
                    "title": source_title or title,
                    "building": self._building_label_from_codes(building_codes),
                    "building_codes": building_codes,
                    "status": self._source_candidate_status(record, work_type),
                    "start_time": source_start,
                    "end_time": source_end,
                    "date_matched": bool(query_dates and source_dates and (query_dates & source_dates)),
                    "fields": detail_fields,
                    "field_items": [
                        {"label": key, "value": value}
                        for key, value in detail_fields.items()
                    ],
                }
            )
        candidates.sort(
            key=lambda item: (
                0 if item.get("date_matched") else 1,
                0 if not self._target_status_is_finished(item.get("status")) else 1,
                str(item.get("start_time") or ""),
            )
        )
        return candidates[: max(1, int(limit or 30))]

    def _target_record_id_from_work_status(
        self, *, work_type: str, source_record_id: str = "", active_item_id: str = ""
    ) -> str:
        work_type = str(work_type or WORK_TYPE_MAINTENANCE).strip()
        source_record_id = str(source_record_id or "").strip()
        active_item_id = str(active_item_id or "").strip()
        if not source_record_id and not active_item_id:
            return ""
        with self._summary_lock:
            status_items = self._load_work_status_items_locked("ALL")
        fallback = ""
        for item in status_items:
            if self._item_work_type(item) != work_type:
                continue
            if source_record_id and str(item.get("source_record_id") or "").strip() != source_record_id:
                continue
            if active_item_id and str(item.get("active_item_id") or "").strip() != active_item_id:
                continue
            target_record_id = canonical_target_record_id(item)
            if not target_record_id:
                continue
            if not fallback:
                fallback = target_record_id
            if not self._is_completed_work_status_item(item):
                return target_record_id
        return fallback

    def _target_record_id_from_identity_map(
        self,
        *,
        work_type: str,
        active_item_id: str = "",
        source_record_id: str = "",
        target_record_id: str = "",
    ) -> str:
        try:
            identity = self._state_store.resolve_notice_identity(
                work_type=work_type,
                active_item_id=active_item_id,
                source_record_id=source_record_id,
                target_record_id=target_record_id,
            )
        except Exception:
            return ""
        if not isinstance(identity, dict):
            return ""
        return str(identity.get("target_record_id") or "").strip()

    def _target_record_id_from_request_payload(
        self,
        request_payload: dict[str, Any],
        *,
        source_record_id: str = "",
        action: str = "",
    ) -> str:
        payload = dict(request_payload or {})
        if source_record_id and not str(payload.get("source_record_id") or "").strip():
            payload["source_record_id"] = source_record_id
        target_record_id = canonical_target_record_id(payload)
        if target_record_id and (
            is_local_record_id(target_record_id)
            or target_record_id == str(source_record_id or "").strip()
        ):
            target_record_id = ""
        if target_record_id:
            return target_record_id
        action = str(action or payload.get("action") or "").strip().lower()
        if action not in {"update", "end", "upload_replace", "delete"}:
            return ""
        raw_record_id = str(payload.get("record_id") or "").strip()
        if (
            raw_record_id
            and not is_local_record_id(raw_record_id)
            and raw_record_id != str(source_record_id or "").strip()
        ):
            return raw_record_id
        return ""

    def _target_record_id_from_unique_candidate(
        self,
        *,
        work_type: str,
        scope: str,
        title: str,
        start_time: str = "",
        end_time: str = "",
        action: str = "update",
    ) -> str:
        title = str(title or "").strip()
        if not title:
            return ""
        try:
            result = self.lookup_notice_target_candidates(
                work_type=work_type,
                scope=scope,
                title=title,
                start_time=start_time,
                end_time=end_time,
                action=action,
                limit=10,
            )
        except Exception:
            return ""
        candidates = [
            item
            for item in (result.get("candidates") or [])
            if isinstance(item, dict)
            and str(item.get("target_record_id") or item.get("record_id") or "").strip()
        ]
        if len(candidates) != 1:
            return ""
        return str(
            candidates[0].get("target_record_id") or candidates[0].get("record_id") or ""
        ).strip()

    def _resolve_target_record_id_for_source_update(
        self,
        *,
        work_type: str,
        notice_type: str,
        source_record: dict[str, Any] | None,
    ) -> str:
        if not isinstance(source_record, dict):
            return ""
        source_record_id = str(source_record.get("record_id") or "").strip()
        target_record_id = self._target_record_id_from_identity_map(
            work_type=work_type,
            source_record_id=source_record_id,
        )
        if target_record_id:
            return target_record_id
        target_record_id = self._target_record_id_from_work_status(
            work_type=work_type, source_record_id=source_record_id
        )
        if target_record_id:
            return target_record_id
        if work_type == WORK_TYPE_REPAIR:
            target_record_id = self._repair_target_record_id(source_record)
            if target_record_id:
                return target_record_id
        return self._resolve_target_record_reference(
            work_type=work_type,
            notice_type=notice_type,
            source_record=source_record,
            include_finished=False,
        )

    def _merge_ongoing_items(
        self,
        scope: str,
        ongoing_items: list[dict[str, Any]] | None,
    ) -> list[dict[str, Any]]:
        scope = self._normalize_scope(scope)
        merged: list[dict[str, Any]] = []
        index_by_key: dict[str, int] = {}
        for item in ongoing_items or []:
            if not isinstance(item, dict):
                continue
            if not self._scope_matches_item(scope, item):
                continue
            copied = normalize_notice_identity_payload(copy.deepcopy(item))
            copied["work_type"] = self._item_work_type(copied)
            copied.setdefault(
                "notice_type",
                NOTICE_TYPE_BY_WORK_TYPE.get(copied["work_type"], NOTICE_TYPE_MAINTENANCE),
            )
            if self._is_ongoing_hidden(copied):
                continue
            identity_keys = self._ongoing_merge_identity_keys(copied)
            duplicate_indexes = [
                index_by_key[key]
                for key in identity_keys
                if key in index_by_key
            ]
            target_index = next(
                (
                    index
                    for index in sorted(set(duplicate_indexes))
                    if not self._ongoing_identity_conflicts(merged[index], copied)
                ),
                None,
            )
            if target_index is not None:
                merged[target_index] = self._merge_duplicate_ongoing_item(
                    merged[target_index],
                    copied,
                )
                for key in identity_keys | self._ongoing_merge_identity_keys(merged[target_index]):
                    index_by_key[key] = target_index
                continue
            merged.append(copied)
            item_index = len(merged) - 1
            for key in identity_keys:
                index_by_key[key] = item_index
        return merged

    def _project_ongoing_items(
        self,
        scope: str,
        ongoing_items: list[dict[str, Any]] | None,
    ) -> list[dict[str, Any]]:
        """Project visible Qt active items to the browser without deduping."""
        scope = self._normalize_scope(scope)
        projected: list[dict[str, Any]] = []
        for item in ongoing_items or []:
            if not isinstance(item, dict):
                continue
            nested_payload = item.get("payload")
            if isinstance(nested_payload, dict):
                copied = copy.deepcopy(nested_payload)
                for key in ("active_item_id", "notice_type", "origin"):
                    value = item.get(key)
                    if value not in (None, "") and not copied.get(key):
                        copied[key] = value
                row_record_id = str(item.get("record_id") or "").strip()
                if row_record_id and not is_local_record_id(row_record_id):
                    copied.setdefault("target_record_id", row_record_id)
                    copied.setdefault("record_id", row_record_id)
            else:
                copied = copy.deepcopy(item)
            copied = normalize_notice_identity_payload(copied)
            if not self._scope_matches_item(scope, copied):
                continue
            copied["work_type"] = self._item_work_type(copied)
            copied.setdefault(
                "notice_type",
                NOTICE_TYPE_BY_WORK_TYPE.get(copied["work_type"], NOTICE_TYPE_MAINTENANCE),
            )
            if self._is_ongoing_hidden(copied):
                continue
            projected.append(copied)
        return projected

    def _ongoing_identity_conflicts(
        self, existing: dict[str, Any], incoming: dict[str, Any]
    ) -> bool:
        existing = normalize_notice_identity_payload(existing or {})
        incoming = normalize_notice_identity_payload(incoming or {})
        existing_work_type = str(
            existing.get("work_type") or existing.get("lan_work_type") or ""
        ).strip()
        incoming_work_type = str(
            incoming.get("work_type") or incoming.get("lan_work_type") or ""
        ).strip()
        if (
            existing_work_type
            and incoming_work_type
            and existing_work_type != incoming_work_type
        ):
            return True
        existing_notice_type = str(existing.get("notice_type") or "").strip()
        incoming_notice_type = str(incoming.get("notice_type") or "").strip()
        if (
            existing_notice_type
            and incoming_notice_type
            and existing_notice_type != incoming_notice_type
        ):
            return True
        existing_active = str(existing.get("active_item_id") or "").strip()
        incoming_active = str(incoming.get("active_item_id") or "").strip()
        if existing_active and incoming_active and existing_active == incoming_active:
            return False
        existing_signature = self._ongoing_exact_duplicate_signature(existing)
        if (
            existing_signature
            and existing_signature == self._ongoing_exact_duplicate_signature(incoming)
        ):
            return False
        existing_target = canonical_target_record_id(existing)
        incoming_target = canonical_target_record_id(incoming)
        if existing_target and incoming_target:
            return existing_target != incoming_target
        existing_source = canonical_source_record_id(existing)
        incoming_source = canonical_source_record_id(incoming)
        if existing_source and incoming_source:
            return existing_source != incoming_source
        if existing_active and incoming_active:
            return existing_active != incoming_active
        return False

    def _ongoing_exact_duplicate_signature(self, item: dict[str, Any]) -> tuple[str, ...]:
        item = normalize_notice_identity_payload(item or {})
        title_key = self._ongoing_business_text_key(
            item.get("title") or item.get("content") or item.get("name") or ""
        )
        if not title_key:
            return ()
        item_fields = item.get("fields") if isinstance(item.get("fields"), dict) else {}
        return (
            self._item_work_type(item),
            str(item.get("notice_type") or "").strip(),
            title_key,
            self._ongoing_business_text_key(item.get("building") or ""),
            self._ongoing_business_text_key(
                item.get("maintenance_cycle") or item_fields.get("维护周期") or ""
            ),
            self._ongoing_business_time_key(item),
            self._ongoing_business_text_key(item.get("location") or ""),
            self._ongoing_business_text_key(item.get("content") or ""),
            self._ongoing_business_text_key(item.get("reason") or ""),
            self._ongoing_business_text_key(item.get("impact") or ""),
        )

    def _ongoing_merge_identity_keys(self, item: dict[str, Any]) -> set[str]:
        item = normalize_notice_identity_payload(item)
        work_type = self._item_work_type(item)
        keys: set[str] = set()
        for kind, value in (
            ("active", item.get("active_item_id")),
            ("source", canonical_source_record_id(item)),
            ("target", canonical_target_record_id(item)),
        ):
            value = str(value or "").strip()
            if value:
                keys.add(f"{work_type}:{kind}:{value}")
        exact_signature = self._ongoing_exact_duplicate_signature(item)
        if exact_signature:
            keys.add(
                f"{work_type}:exact:"
                + hashlib.sha1(
                    json.dumps(
                        exact_signature,
                        ensure_ascii=False,
                        separators=(",", ":"),
                    ).encode("utf-8", errors="ignore")
                ).hexdigest()
            )
        return {key for key in keys if key}

    def _ongoing_business_merge_keys(
        self, item: dict[str, Any], *, work_type: str = ""
    ) -> set[str]:
        """Best-effort merge keys for the same notice coming from different projections.

        ID-based keys remain authoritative. These keys only bridge cases where the
        same notice is present once from a source/work-status projection and once
        from the Qt active projection, but one side has not yet received the other
        side's ID. Reason is kept in the key when present so same-name maintenance
        items with different reasons do not collapse into one row.
        """

        work_type = str(work_type or "").strip() or self._item_work_type(item)
        title_key = self._ongoing_business_text_key(
            item.get("title") or item.get("content") or item.get("name") or ""
        )
        if not title_key:
            return set()
        building_key = self._ongoing_business_text_key(item.get("building") or "")
        reason_key = self._ongoing_business_text_key(item.get("reason") or "")
        cycle_key = ""
        if work_type == WORK_TYPE_MAINTENANCE:
            item_fields = item.get("fields") if isinstance(item.get("fields"), dict) else {}
            cycle_key = self._ongoing_business_text_key(
                item.get("maintenance_cycle") or item_fields.get("维护周期") or ""
            )
        time_key = self._ongoing_business_time_key(item)
        keys: set[str] = set()
        if building_key and reason_key:
            keys.add(f"{work_type}:business:title-building-reason:{building_key}:{title_key}:{cycle_key}:{reason_key}")
        if building_key and time_key and reason_key:
            keys.add(f"{work_type}:business:title-time-reason:{building_key}:{title_key}:{cycle_key}:{time_key}:{reason_key}")
        elif building_key and time_key:
            keys.add(f"{work_type}:business:title-time:{building_key}:{title_key}:{cycle_key}:{time_key}")
        return keys

    @staticmethod
    def _ongoing_business_text_key(value: Any) -> str:
        return re.sub(
            r"[\s,，;；:：。.【】（）()《》<>\"'“”‘’\-－_/\\]+",
            "",
            str(value or ""),
        ).strip().lower()

    @staticmethod
    def _ongoing_business_time_key(item: dict[str, Any]) -> str:
        parts = [
            str(item.get("start_time") or ""),
            str(item.get("time_str") or ""),
            str(item.get("time") or ""),
            str(item.get("end_time") or ""),
        ]
        digits = re.findall(r"\d+", "".join(parts))
        return "".join(chunk.zfill(2) if len(chunk) <= 2 else chunk for chunk in digits)

    @staticmethod
    def _ongoing_item_score(item: dict[str, Any]) -> int:
        item = normalize_notice_identity_payload(item)
        score = 0
        for field in (
            "target_record_id",
            "active_item_id",
            "source_record_id",
            "title",
            "building",
            "specialty",
            "maintenance_cycle",
            "start_time",
            "end_time",
            "location",
            "content",
            "reason",
            "impact",
            "progress",
        ):
            if str(item.get(field) or "").strip():
                score += 1
        if item.get("extra_images"):
            score += 1
        return score

    def _merge_duplicate_ongoing_item(
        self,
        existing: dict[str, Any],
        incoming: dict[str, Any],
    ) -> dict[str, Any]:
        existing = normalize_notice_identity_payload(copy.deepcopy(existing))
        incoming = normalize_notice_identity_payload(copy.deepcopy(incoming))
        if self._ongoing_item_score(incoming) > self._ongoing_item_score(existing):
            base, supplement = incoming, existing
        else:
            base, supplement = existing, incoming
        for key, value in supplement.items():
            if key not in base or base.get(key) in (None, "", [], {}):
                base[key] = copy.deepcopy(value)
        return base

    def _source_record_matches_filters(
        self, record: dict[str, Any], *, month: str = "", specialty: str = ""
    ) -> bool:
        work_type = self._record_work_type(record)
        if work_type == WORK_TYPE_MAINTENANCE:
            if not self._maintenance_status_allows_workbench(record):
                return False
        elif work_type == WORK_TYPE_CHANGE:
            if not self._change_progress_allows_workbench(record):
                return False
        elif work_type == WORK_TYPE_REPAIR:
            if not self._is_valid_repair_record(record):
                return False
            if self._repair_has_ended(record):
                return False
            if self._repair_source_status(record) not in WORKBENCH_SOURCE_STATUSES:
                return False
        if not self._source_record_matches_month_window(record, month):
            return False
        specialty = str(specialty or "").strip()
        if not specialty:
            return True
        fields = record.get("display_fields") or {}
        if work_type == WORK_TYPE_MAINTENANCE:
            return self._clean_source_text(fields.get("专业类别")) == specialty
        if work_type == WORK_TYPE_CHANGE:
            return self._change_specialty(record) == specialty
        if work_type == WORK_TYPE_REPAIR:
            return self._repair_specialty(record) == specialty
        return True

    def _source_record_matches_search(self, record: dict[str, Any], search: str = "") -> bool:
        query = str(search or "").strip().lower()
        if not query:
            return True
        fields = record.get("display_fields") or {}
        terms: list[str] = [
            str(record.get("record_id") or ""),
            str(record.get("source_record_id") or ""),
            str(record.get("target_record_id") or ""),
            str(record.get("work_type") or ""),
            str(record.get("notice_type") or ""),
        ]
        with suppress(Exception):
            terms.append(self._change_title(record))
        with suppress(Exception):
            terms.append(self._repair_title(record))
        for value in fields.values():
            if isinstance(value, (list, tuple, set)):
                terms.extend(str(item or "") for item in value)
            elif isinstance(value, dict):
                terms.extend(str(item or "") for item in value.values())
            else:
                terms.append(str(value or ""))
        return query in " ".join(terms).lower()

    def _source_record_display_priority(self, record: dict[str, Any]) -> int:
        source_work_type = self._record_source_work_type(record)
        if source_work_type == WORK_TYPE_MAINTENANCE:
            progress = self._maintenance_status_value(record)
        elif source_work_type == WORK_TYPE_CHANGE:
            progress = self._change_progress_value(record)
        elif source_work_type == WORK_TYPE_REPAIR:
            progress = self._repair_source_status(record)
        else:
            progress = self._maintenance_status_value(record)
        text = str(progress or "").strip()
        if "延期未开始" in text:
            return 0
        if "未开始" in text:
            return 1
        if "进行中" in text:
            return 2
        return 10

    def _sort_workbench_records(self, records: list[dict[str, Any]]) -> list[dict[str, Any]]:
        return [
            item
            for _, item in sorted(
                enumerate(records or []),
                key=lambda pair: (self._source_record_display_priority(pair[1]), pair[0]),
            )
        ]

    def _workbench_records_from_memory(
        self, *, month: str = "", specialty: str = "", scope: str = "ALL"
    ) -> list[dict[str, Any]]:
        maintenance_records = self._filter_records(
            month=month, specialty=specialty, scope=scope
        )
        change_records = self._filter_change_records(
            month=month,
            specialty=specialty,
            scope=scope,
        )
        repair_records = self._filter_repair_records(
            month=month, specialty=specialty, scope=scope
        )
        return maintenance_records + change_records + repair_records

    def _workbench_records(
        self, *, month: str = "", specialty: str = "", scope: str = "ALL"
    ) -> list[dict[str, Any]]:
        snapshot_records = self._source_snapshot_records(scope)
        if snapshot_records is not None:
            return self._sort_workbench_records(self._apply_work_type_overrides([
                record
                for record in snapshot_records
                if self._source_record_matches_filters(
                    record, month=month, specialty=specialty
                )
            ]))
        return self._sort_workbench_records(self._apply_work_type_overrides(
            self._workbench_records_from_memory(
                month=month, specialty=specialty, scope=scope
            )
        ))

    def _maintenance_options_for_records(self, records: list[dict[str, Any]]) -> list[dict[str, str]]:
        options = []
        for record in records:
            if self._record_work_type(record) != WORK_TYPE_MAINTENANCE:
                continue
            fields = record["display_fields"]
            options.append(
                {
                    "record_id": record["record_id"],
                    "label": fields.get("维护总项") or record["record_id"],
                    "sub_label": " | ".join(
                        filter(
                            None,
                            [
                                fields.get("维护编号", ""),
                                fields.get("维护项目", ""),
                            ],
                        )
                    ),
                }
            )
        return options

    def assert_generated_drafts_allowed(
        self, drafts: list[dict[str, Any]], *, scope: str = "ALL"
    ) -> None:
        self.ensure_snapshot_loaded()
        scope = self._normalize_scope(scope)
        for draft in drafts or []:
            record_id = str((draft or {}).get("record_id") or "").strip()
            if not record_id:
                raise PortalError("存在缺少 record_id 的待生成记录。")
            record = self._find_record_by_id(record_id)
            fields = record.get("display_fields") or {}
            if not self._scope_matches_building(scope, fields.get("楼栋")):
                raise PortalError(
                    f"当前账号无权生成 {record_id} 对应楼栋的通告。"
                )

    def assert_generated_items_allowed(
        self, items: list[dict[str, Any]], *, scope: str = "ALL"
    ) -> None:
        self.ensure_snapshot_loaded()
        scope = self._normalize_scope(scope)
        for item in items or []:
            record_id = str((item or {}).get("record_id") or "").strip()
            if not record_id:
                raise PortalError("存在缺少 record_id 的待发送通告。")
            record = self._find_record_by_id(record_id)
            fields = record.get("display_fields") or {}
            if not self._scope_matches_building(scope, fields.get("楼栋")):
                raise PortalError(
                    f"当前账号无权发送 {record_id} 对应楼栋的通告。"
                )

    def get_bootstrap(
        self,
        *,
        scope: str = "ALL",
        ongoing_items: list[dict[str, Any]] | None = None,
    ) -> dict[str, Any]:
        self.ensure_snapshot_loaded()
        default_month = RECENT_MONTH_FILTER_LABEL
        scope = self._normalize_scope(scope)
        merged_ongoing = self._project_ongoing_items(scope, ongoing_items or [])
        filtered_records = self._workbench_records(month=default_month, scope=scope)
        linked_zhihang_ids = self._linked_zhihang_record_ids(merged_ongoing)
        zhihang_records = self._filter_zhihang_change_records(
            month=default_month, scope=scope, exclude_record_ids=linked_zhihang_ids
        )
        daily_summary = self.get_daily_summary(
            scope=scope, ongoing_items=merged_ongoing
        )
        merged_ongoing = self._annotate_undo_items(merged_ongoing, scope=scope)
        daily_summary["items"] = self._annotate_undo_items(
            daily_summary.get("items") or [], scope=scope
        )
        summary_by_record = self._work_status_by_records(
            filtered_records,
            scope=scope,
            daily_items=daily_summary.get("items") or [],
        )
        record_type_counts = self._work_type_counts(filtered_records)
        ongoing_type_counts = self._work_type_counts(merged_ongoing)
        return {
            "app_token": self.app_token,
            "table_id": self.table_id,
            "scope": scope,
            "scope_label": self._scope_label(scope),
            "scope_options": SCOPE_OPTIONS,
            "default_work_type": (
                WORK_TYPE_CHANGE if scope == "CAMPUS" else WORK_TYPE_MAINTENANCE
            ),
            "last_loaded_at": self._last_loaded_at,
            "payload_version": self._payload_version(
                scope=scope,
                records=filtered_records,
                ongoing_items=merged_ongoing,
                daily_summary=daily_summary,
            ),
            "source_snapshot_ready": self._source_snapshot_exists(scope),
            "source_cache_ttl_seconds": self._source_cache_ttl_seconds(),
            "warnings": self._current_load_warnings(),
            "maintenance_status": DEFAULT_MAINTENANCE_STATUS,
            "field_order": [meta.field_name for meta in self._field_meta_list],
            "fields": [
                {
                    "field_name": meta.field_name,
                    "ui_type": meta.ui_type,
                    "type": meta.field_type,
                    "is_primary": meta.is_primary,
                    "options": meta.option_names,
                    "has_formula": meta.has_formula,
                }
                for meta in self._field_meta_list
            ],
            "filters": {
                "default_month": default_month,
                "months": self._sorted_unique_option_names("计划维护月份"),
                "specialties": self._sorted_unique_work_specialties(),
                "buildings": self._sorted_unique_option_names("楼栋"),
            },
            "records": [
                self._serialize_record(record, summary_by_record)
                for record in filtered_records
            ],
            "zhihang_change_records": [
                self._serialize_zhihang_change_record(record)
                for record in zhihang_records
            ],
            "ongoing": merged_ongoing,
            "daily_summary": daily_summary,
            "record_type_counts": record_type_counts,
            "ongoing_type_counts": ongoing_type_counts,
            "maintenance_options": self._maintenance_options_for_records(filtered_records),
            "defaults": {
                "impact": DEFAULT_IMPACT_TEXT,
                "progress": DEFAULT_PROGRESS_TEXT,
            },
        }

    def query_records(
        self,
        *,
        month: str = "",
        specialty: str = "",
        search: str = "",
        building: str = "",
        scope: str = "ALL",
        ongoing_items: list[dict[str, Any]] | None = None,
        work_type: str = "",
        sections: set[str] | list[str] | tuple[str, ...] | None = None,
        records_page: int | str = 1,
        records_page_size: int | str = 0,
        ongoing_page: int | str = 1,
        ongoing_page_size: int | str = 0,
    ) -> dict[str, Any]:
        self.ensure_snapshot_loaded()
        scope = self._normalize_scope(scope)
        requested_sections = {
            str(item or "").strip().lower()
            for item in (sections or [])
            if str(item or "").strip()
        }
        if not requested_sections:
            requested_sections = {"records", "ongoing", "stats", "zhihang"}
        requested_work_type = str(work_type or "").strip()
        if requested_work_type not in {
            WORK_TYPE_MAINTENANCE,
            WORK_TYPE_CHANGE,
            WORK_TYPE_REPAIR,
            WORK_TYPE_POWER,
            WORK_TYPE_POLLING,
            WORK_TYPE_ADJUST,
        }:
            requested_work_type = ""
        merged_ongoing = self._project_ongoing_items(scope, ongoing_items or [])
        scoped_records = self._workbench_records(
            month=month, specialty=specialty, scope=scope
        )
        linked_zhihang_ids = self._linked_zhihang_record_ids(merged_ongoing)
        zhihang_records = self._filter_zhihang_change_records(
            month=month, scope=scope, exclude_record_ids=linked_zhihang_ids
        )
        if building:
            scoped_records = [
                record
                for record in scoped_records
                if str((record.get("display_fields") or {}).get("楼栋") or "") == building
            ]
        record_type_counts = self._work_type_counts(scoped_records)
        filtered_records = [
            record
            for record in scoped_records
            if not requested_work_type or self._record_work_type(record) == requested_work_type
        ]
        search_text = str(search or "").strip()
        if search_text:
            filtered_records = [
                record
                for record in filtered_records
                if self._source_record_matches_search(record, search_text)
            ]
        daily_summary = self.get_daily_summary(
            scope=scope, ongoing_items=merged_ongoing
        )
        merged_ongoing = self._annotate_undo_items(merged_ongoing, scope=scope)
        if "closed" in requested_sections or "daily" in requested_sections:
            daily_summary["items"] = self._annotate_undo_items(
                daily_summary.get("items") or [], scope=scope
            )
        else:
            daily_summary["items"] = []
        ongoing_type_counts = self._work_type_counts(merged_ongoing)
        filtered_ongoing_for_response = [
            item
            for item in merged_ongoing
            if not requested_work_type
            or self._item_work_type(item) == requested_work_type
        ]
        include_records = "records" in requested_sections
        include_ongoing = "ongoing" in requested_sections
        include_zhihang = (
            "zhihang" in requested_sections
            or requested_work_type == WORK_TYPE_CHANGE
            or not requested_work_type
        )
        def _positive_int(value: Any, default: int = 0) -> int:
            try:
                return max(0, int(float(value)))
            except Exception:
                return default

        def _slice_page(items: list[Any], page: Any, page_size: Any) -> tuple[list[Any], dict[str, int]]:
            total = len(items or [])
            size = _positive_int(page_size, 0)
            if size <= 0:
                return list(items or []), {
                    "page": 1,
                    "page_size": total,
                    "total": total,
                    "total_pages": 1,
                    "offset": 0,
                    "returned": total,
                }
            size = min(size, 200)
            total_pages = max(1, (total + size - 1) // size)
            current_page = min(max(1, _positive_int(page, 1) or 1), total_pages)
            offset = (current_page - 1) * size
            visible = list(items or [])[offset : offset + size]
            return visible, {
                "page": current_page,
                "page_size": size,
                "total": total,
                "total_pages": total_pages,
                "offset": offset,
                "returned": len(visible),
            }

        visible_records, records_pagination = _slice_page(
            filtered_records,
            records_page,
            records_page_size,
        )
        visible_ongoing, ongoing_pagination = _slice_page(
            filtered_ongoing_for_response,
            ongoing_page,
            ongoing_page_size,
        )
        summary_by_record = self._work_status_by_records(
            visible_records,
            scope=scope,
            daily_items=daily_summary.get("items") or [],
        ) if include_records else {}
        ongoing_by_source: dict[tuple[str, str], list[dict[str, Any]]] = {}
        for item in merged_ongoing:
            source_record_id = canonical_source_record_id(item)
            if not source_record_id:
                continue
            key = (self._item_work_type(item), source_record_id)
            ongoing_by_source.setdefault(key, []).append(item)

        serialized_records: list[dict[str, Any]] = []
        if include_records:
            for record in visible_records:
                serialized = self._serialize_record(record, summary_by_record)
                source_record_id = str(record.get("record_id") or "").strip()
                linked_items = ongoing_by_source.get(
                    (self._record_work_type(record), source_record_id),
                    [],
                )
                # A source row may stay visible after its notice has started. Expose
                # the exact active projection only when the source relation is unique,
                # so the browser can use the same target identity and actions as Qt.
                if len(linked_items) == 1:
                    serialized["linked_ongoing"] = copy.deepcopy(linked_items[0])
                serialized_records.append(serialized)
        return {
            "maintenance_status": DEFAULT_MAINTENANCE_STATUS,
            "scope": scope,
            "scope_label": self._scope_label(scope),
            "default_work_type": (
                WORK_TYPE_CHANGE if scope == "CAMPUS" else WORK_TYPE_MAINTENANCE
            ),
            "last_loaded_at": self._last_loaded_at,
            "payload_version": self._payload_version(
                scope=scope,
                records=filtered_records,
                ongoing_items=merged_ongoing,
                daily_summary=daily_summary,
            ),
            "source_snapshot_ready": self._source_snapshot_exists(scope),
            "source_cache_ttl_seconds": self._source_cache_ttl_seconds(),
            "warnings": self._current_load_warnings(),
            "records": serialized_records,
            "zhihang_change_records": [
                self._serialize_zhihang_change_record(record)
                for record in zhihang_records
            ] if include_zhihang else [],
            "ongoing": visible_ongoing if include_ongoing else [],
            "daily_summary": daily_summary,
            "record_type_counts": record_type_counts,
            "ongoing_type_counts": ongoing_type_counts,
            "maintenance_options": self._maintenance_options_for_records(visible_records),
            "records_pagination": records_pagination,
            "ongoing_pagination": ongoing_pagination,
            "filters": {
                "default_month": month or RECENT_MONTH_FILTER_LABEL,
                "specialties": self._sorted_unique_work_specialties(),
            },
            "count": len(filtered_records),
        }

    def get_workbench_closed_items(
        self,
        *,
        scope: str = "ALL",
        work_type: str = "",
        ongoing_items: list[dict[str, Any]] | None = None,
    ) -> dict[str, Any]:
        scope = self._normalize_scope(scope)
        requested_work_type = str(work_type or "").strip()
        daily_summary = self.get_daily_summary(scope=scope, ongoing_items=ongoing_items or [])
        items = self._annotate_undo_items(daily_summary.get("items") or [], scope=scope)
        if requested_work_type:
            items = [
                item
                for item in items
                if self._item_work_type(item) == requested_work_type
            ]
        closed = [
            item
            for item in items
            if str(item.get("status") or "") == "已结束" or bool(item.get("ended_at"))
        ]
        return {
            "scope": scope,
            "items": closed,
            "count": len(closed),
            "stats": daily_summary.get("stats") or {},
        }

    def _engineer_mop_settings(self) -> dict[str, str]:
        settings = self._state_store.get_settings() or {}
        def first_text(*values: Any) -> str:
            for value in values:
                text = str(value or "").strip()
                if text and text.lower() not in {"none", "null", "undefined"}:
                    return text
            return ""

        app_token = first_text(
            settings.get("mop_app_token"),
            getattr(config, "mop_app_token", ""),
            MOP_SOURCE_APP_TOKEN,
        )
        table_id = first_text(
            settings.get("mop_table_id"),
            getattr(config, "mop_table_id", ""),
            MOP_SOURCE_TABLE_ID,
        )
        view_id = first_text(
            settings.get("mop_view_id"),
            getattr(config, "mop_view_id", ""),
            MOP_SOURCE_VIEW_ID,
        )
        return {
            "app_token": app_token,
            "table_id": table_id,
            "view_id": view_id,
            "title_field": first_text(settings.get("mop_title_field"), MOP_TITLE_FIELD_NAME),
            "attachment_field": first_text(settings.get("mop_attachment_field"), MOP_ATTACHMENT_FIELD_NAME),
        }

    def _engineer_notice_key(self, item: dict[str, Any]) -> str:
        work_type = self._item_work_type(item)
        target_record_id = canonical_target_record_id(item)
        source_record_id = str(item.get("source_record_id") or "").strip()
        active_item_id = str(item.get("active_item_id") or "").strip()
        if target_record_id:
            return f"{work_type}:target:{target_record_id}"
        if source_record_id:
            return f"{work_type}:source:{source_record_id}"
        if active_item_id:
            return f"{work_type}:active:{active_item_id}"
        seed = json.dumps(
            [
                work_type,
                str(item.get("notice_type") or ""),
                str(item.get("title") or item.get("content") or ""),
                str(item.get("building") or ""),
                str(item.get("started_at") or item.get("start_time") or ""),
                str(item.get("ended_at") or item.get("end_time") or ""),
                str(item.get("reason") or ""),
            ],
            ensure_ascii=False,
            sort_keys=True,
        )
        return f"{work_type}:generated:{hashlib.sha1(seed.encode('utf-8')).hexdigest()}"

    @staticmethod
    def _normalize_engineer_mop_template_text(value: Any) -> str:
        text = str(value or "").strip()
        if not text:
            return ""
        text = re.sub(r"【[^】]*】", "", text)
        text = re.sub(r"^(EA118|南通基地|机房)+", "", text, flags=re.IGNORECASE)
        text = re.sub(r"(维保通告|维护通告|通告)$", "", text)
        text = re.sub(r"[（）()\[\]【】《》<>，,。；;：:\s_\-~至/]+", "", text)
        return text.lower()

    def _engineer_mop_template_key(
        self,
        item: dict[str, Any],
        *,
        title: str = "",
        building: str = "",
        maintenance_total: str = "",
        maintenance_cycle: str = "",
    ) -> str:
        work_type = self._item_work_type(item)
        if work_type != WORK_TYPE_MAINTENANCE:
            return ""
        fields = item.get("display_fields") if isinstance(item.get("display_fields"), dict) else {}
        building_value = (
            building
            or item.get("building")
            or fields.get("楼栋")
            or fields.get("变更楼栋")
            or ""
        )
        codes = item.get("building_codes") if isinstance(item.get("building_codes"), list) else []
        if not codes:
            codes = self._building_codes_from_value(building_value)
        building_key = codes[0] if len(codes) == 1 else self._normalize_scope(building_value)
        if building_key == "ALL":
            building_key = ""
        total_value = (
            maintenance_total
            or item.get("maintenance_total")
            or item.get("memory_name")
            or fields.get("维护总项")
            or title
            or item.get("title")
            or item.get("content")
            or ""
        )
        cycle_value = (
            maintenance_cycle
            or item.get("maintenance_cycle")
            or fields.get("维护周期")
            or ""
        )
        normalized_total = self._normalize_engineer_mop_template_text(total_value)
        normalized_cycle = self._normalize_engineer_mop_template_text(cycle_value) or "/"
        if not building_key or not normalized_total:
            return ""
        seed = json.dumps(
            ["maintenance", building_key, normalized_total, normalized_cycle],
            ensure_ascii=False,
            sort_keys=True,
        )
        return f"mop-template:{hashlib.sha1(seed.encode('utf-8')).hexdigest()}"

    def _serialize_engineer_notice(self, item: dict[str, Any], *, status: str = "") -> dict[str, Any]:
        work_type = self._item_work_type(item)
        fields = item.get("display_fields") if isinstance(item.get("display_fields"), dict) else {}
        title = str(item.get("title") or item.get("content") or item.get("name") or "").strip()
        start_time = str(item.get("started_at") or item.get("start_time") or item.get("time") or "").strip()
        end_time = str(item.get("ended_at") or item.get("end_time") or "").strip()
        notice_status = status or str(item.get("status") or "").strip() or ("已结束" if end_time else "进行中")
        building = str(item.get("building") or fields.get("楼栋") or "")
        maintenance_total = str(
            item.get("maintenance_total")
            or item.get("memory_name")
            or fields.get("维护总项")
            or ""
        ).strip()
        maintenance_cycle = str(item.get("maintenance_cycle") or fields.get("维护周期") or "").strip()
        mop_template_key = self._engineer_mop_template_key(
            item,
            title=title,
            building=building,
            maintenance_total=maintenance_total,
            maintenance_cycle=maintenance_cycle,
        )
        return {
            "notice_key": self._engineer_notice_key(item),
            "mop_template_key": mop_template_key,
            "work_type": work_type,
            "notice_type": str(item.get("notice_type") or NOTICE_TYPE_MAINTENANCE),
            "title": title,
            "status": notice_status,
            "building": building,
            "building_codes": item.get("building_codes") if isinstance(item.get("building_codes"), list) else [],
            "specialty": str(item.get("specialty") or ""),
            "maintenance_total": maintenance_total,
            "maintenance_cycle": maintenance_cycle,
            "source_record_id": str(item.get("source_record_id") or ""),
            "target_record_id": canonical_target_record_id(item),
            "active_item_id": str(item.get("active_item_id") or ""),
            "start_time": start_time,
            "end_time": end_time,
            "location": str(item.get("location") or ""),
            "content": str(item.get("content") or ""),
            "reason": str(item.get("reason") or ""),
            "progress": str(item.get("progress") or ""),
            "updated_at": str(item.get("ended_at") or item.get("last_updated_at") or item.get("updated_at") or item.get("started_at") or ""),
        }

    def _mop_source_flag(self, value: Any) -> bool:
        if isinstance(value, (list, tuple, set)):
            return any(self._mop_source_flag(item) for item in value)
        if isinstance(value, dict):
            for key in ("checked", "value", "text", "name"):
                if key in value and self._mop_source_flag(value.get(key)):
                    return True
            return bool(value)
        return self._truthy_flag(value)

    def _serialize_engineer_source_maintenance_notice(
        self, record: dict[str, Any]
    ) -> dict[str, Any] | None:
        fields = record.get("display_fields") if isinstance(record.get("display_fields"), dict) else {}
        maintenance_total = str(fields.get("维护总项") or "").strip()
        building = str(fields.get("楼栋") or "").strip()
        if not maintenance_total and not building:
            return None
        maintenance_cycle = str(fields.get("维护周期") or "").strip()
        title = f"{building}{maintenance_total}{('-' + maintenance_cycle) if maintenance_cycle else ''}".strip()
        signed_attachments = self._extract_mop_attachments(fields, MOP_SIGNED_ATTACHMENT_FIELD)
        item = {
            "work_type": WORK_TYPE_MAINTENANCE,
            "notice_type": NOTICE_TYPE_MAINTENANCE,
            "title": title or maintenance_total or str(record.get("record_id") or ""),
            "building": building,
            "building_codes": self._building_codes_from_value(building),
            "specialty": str(fields.get("专业类别") or fields.get("专业") or "").strip(),
            "maintenance_total": maintenance_total,
            "maintenance_cycle": maintenance_cycle,
            "source_record_id": str(record.get("record_id") or ""),
            "display_fields": fields,
            "location": str(fields.get("位置") or building or "").strip(),
            "content": str(fields.get("内容") or fields.get("维护内容") or maintenance_total).strip(),
            "reason": str(fields.get("原因") or fields.get("维护原因") or "").strip(),
            "progress": str(fields.get("进度") or fields.get("维护进度") or "").strip(),
            "updated_at": str(fields.get("更新时间") or fields.get("计划维护月份") or ""),
        }
        notice = self._serialize_engineer_notice(
            item,
            status=self._maintenance_status_value(record) or DEFAULT_MAINTENANCE_STATUS,
        )
        notice.update(
            {
                "mop_uploaded": bool(signed_attachments),
                "mop_attachment_count": len(signed_attachments),
                "mop_engineer_confirmed": self._mop_source_flag(fields.get(MOP_ENGINEER_CONFIRM_FIELD)),
                "mop_supervisor_confirmed": self._mop_source_flag(fields.get(MOP_SUPERVISOR_CONFIRM_FIELD)),
                "mop_source_record": True,
            }
        )
        return notice

    @staticmethod
    def _merge_engineer_mop_notice(
        existing: dict[str, Any], incoming: dict[str, Any]
    ) -> dict[str, Any]:
        merged = dict(existing)
        for key, value in incoming.items():
            if value in (None, "", [], {}):
                continue
            if key in {
                "status",
                "target_record_id",
                "active_item_id",
                "start_time",
                "end_time",
                "location",
                "content",
                "reason",
                "progress",
                "updated_at",
            }:
                merged[key] = value
            elif key not in merged or merged.get(key) in (None, "", [], {}):
                merged[key] = value
        for key in (
            "mop_uploaded",
            "mop_engineer_confirmed",
            "mop_supervisor_confirmed",
            "mop_source_record",
        ):
            merged[key] = bool(existing.get(key) or incoming.get(key))
        merged["mop_attachment_count"] = max(
            int(existing.get("mop_attachment_count") or 0),
            int(incoming.get("mop_attachment_count") or 0),
        )
        return merged

    @staticmethod
    def _engineer_notice_is_ended_status(status: Any) -> bool:
        text = str(status or "").strip()
        if not text or any(token in text for token in ("未结束", "未完成", "未闭环")):
            return False
        return any(
            token in text
            for token in (
                "已结束",
                "正常结束",
                "延迟结束",
                "延期结束",
                "维修完成",
                "已完成",
                "闭环",
            )
        )

    @classmethod
    def _engineer_maintenance_notice_sort_key(cls, item: dict[str, Any]) -> tuple[Any, ...]:
        bound = bool(item.get("mop_binding"))
        uploaded = bool(item.get("mop_uploaded"))
        ended = cls._engineer_notice_is_ended_status(item.get("status"))
        needs_action = not bound or not uploaded
        if not bound and not uploaded:
            priority = 0
        elif not bound:
            priority = 1
        elif not uploaded:
            priority = 2
        else:
            priority = 3
        return (
            0 if ended or needs_action else 1,
            priority,
            0 if ended else 1,
            str(item.get("building") or ""),
            str(item.get("maintenance_cycle") or ""),
            str(item.get("title") or ""),
            str(item.get("updated_at") or ""),
        )

    def _engineer_month_maintenance_notices(
        self, *, scope: str, ongoing_items: list[dict[str, Any]] | None
    ) -> list[dict[str, Any]]:
        merged_ongoing = self._merge_ongoing_items(scope, ongoing_items or [])
        notices_by_key: dict[str, dict[str, Any]] = {}

        def upsert_notice(notice: dict[str, Any]) -> None:
            if not notice:
                return
            match_key = str(notice.get("notice_key") or "").strip()
            source_record_id = str(notice.get("source_record_id") or "").strip()
            template_key = str(notice.get("mop_template_key") or "").strip()
            if source_record_id:
                for key, existing in notices_by_key.items():
                    if str(existing.get("source_record_id") or "").strip() == source_record_id:
                        match_key = key
                        break
            if match_key not in notices_by_key and template_key:
                for key, existing in notices_by_key.items():
                    if str(existing.get("mop_template_key") or "").strip() == template_key:
                        match_key = key
                        break
            if match_key in notices_by_key:
                notices_by_key[match_key] = self._merge_engineer_mop_notice(
                    notices_by_key[match_key],
                    notice,
                )
            else:
                notices_by_key[match_key or uuid.uuid4().hex] = notice

        current_month = self._current_month_label()
        source_records = [
            record
            for record in list(self._records or [])
            if self._record_work_type(record) == WORK_TYPE_MAINTENANCE
            and self._maintenance_record_matches_mop_month_window(record, current_month)
            and self._scope_matches_building(
                scope,
                (record.get("display_fields") or {}).get("楼栋"),
            )
        ]
        if not source_records:
            source_records = [
                record
                for record in (self._source_snapshot_records(scope) or [])
                if self._record_work_type(record) == WORK_TYPE_MAINTENANCE
                and self._maintenance_record_matches_mop_month_window(record, current_month)
            ]
        for record in source_records:
            notice = self._serialize_engineer_source_maintenance_notice(record)
            if notice:
                upsert_notice(notice)

        for item in merged_ongoing:
            if self._item_work_type(item) != WORK_TYPE_MAINTENANCE:
                continue
            notice = self._serialize_engineer_notice(item, status="进行中")
            upsert_notice(notice)
        daily = self.get_daily_summary(scope=scope, ongoing_items=merged_ongoing)
        for item in daily.get("items") or []:
            if not isinstance(item, dict):
                continue
            if self._item_work_type(item) != WORK_TYPE_MAINTENANCE:
                continue
            notice = self._serialize_engineer_notice(item)
            existing = notices_by_key.get(notice["notice_key"])
            if existing and not self._engineer_notice_is_ended_status(existing.get("status")):
                continue
            upsert_notice(notice)
        notices = list(notices_by_key.values())
        notices.sort(key=self._engineer_maintenance_notice_sort_key)
        return notices

    @staticmethod
    def _mop_field_text(fields: dict[str, Any], names: list[str]) -> str:
        for name in names:
            value = fields.get(name)
            if isinstance(value, list):
                text = "、".join(str(item.get("text") or item.get("name") or item) if isinstance(item, dict) else str(item) for item in value)
            else:
                text = str(value or "")
            text = text.strip()
            if text:
                return text
        return ""

    @staticmethod
    def _extract_mop_attachments(fields: dict[str, Any], preferred_field: str = "") -> list[dict[str, Any]]:
        candidates: list[Any] = []
        if preferred_field and preferred_field in fields:
            candidates.append(fields.get(preferred_field))
        else:
            candidates.extend(fields.values())
        attachments: list[dict[str, Any]] = []
        seen: set[str] = set()
        for value in candidates:
            values = value if isinstance(value, list) else [value]
            for item in values:
                if not isinstance(item, dict):
                    continue
                token = str(
                    item.get("file_token")
                    or item.get("token")
                    or item.get("tmp_url")
                    or item.get("url")
                    or item.get("download_url")
                    or ""
                ).strip()
                name = str(item.get("name") or item.get("file_name") or item.get("filename") or "").strip()
                if not token and not name:
                    continue
                dedupe_key = token or name
                if dedupe_key in seen:
                    continue
                seen.add(dedupe_key)
                attachments.append(
                    {
                        "file_token": str(item.get("file_token") or item.get("token") or "").strip(),
                        "name": name or "MOP表格",
                        "size": item.get("size") or item.get("file_size") or "",
                        "mime_type": str(item.get("mime_type") or item.get("type") or "").strip(),
                        "url": str(item.get("download_url") or item.get("url") or item.get("tmp_url") or "").strip(),
                    }
                )
        return attachments

    @classmethod
    def _attachment_latest_publish_time(cls, fields: dict[str, Any]) -> str:
        return cls._mop_field_text(fields, list(ATTACHMENT_LATEST_FIELD_NAMES))

    @staticmethod
    def _attachment_cache_signature(
        *,
        category: str,
        app_token: str,
        table_id: str,
        record_id: str,
        attachment: dict[str, Any],
        latest_publish_time: str = "",
    ) -> dict[str, str]:
        return {
            "category": str(category or ""),
            "app_token": str(app_token or ""),
            "table_id": str(table_id or ""),
            "record_id": str(record_id or ""),
            "file_token": str(attachment.get("file_token") or ""),
            "url": str(attachment.get("url") or ""),
            "name": str(attachment.get("name") or ""),
            "size": str(attachment.get("size") or ""),
            "mime_type": str(attachment.get("mime_type") or ""),
            "latest_publish_time": str(latest_publish_time or attachment.get("latest_publish_time") or ""),
        }

    def _attachment_with_cache_context(
        self,
        attachment: dict[str, Any],
        *,
        category: str,
        app_token: str,
        table_id: str,
        record_id: str,
        latest_publish_time: str = "",
    ) -> dict[str, Any]:
        item = dict(attachment or {})
        item["cache_category"] = category
        item["app_token"] = app_token
        item["table_id"] = table_id
        item["record_id"] = record_id
        item["latest_publish_time"] = latest_publish_time
        return item

    @classmethod
    def _extract_signature_attachments(cls, fields: dict[str, Any]) -> list[dict[str, Any]]:
        attachments = cls._extract_mop_attachments(fields, SIGNATURE_ATTACHMENT_FIELD)
        usable: list[dict[str, Any]] = []
        for item in attachments:
            if not isinstance(item, dict):
                continue
            token = str(item.get("file_token") or "").strip()
            name = str(item.get("name") or "").strip().lower()
            mime_type = str(item.get("mime_type") or "").strip().lower()
            size_text = str(item.get("size") or "").strip()
            if size_text in {"0", "0.0"}:
                continue
            encrypted_like = name.endswith(".sigenc")
            image_like = (
                mime_type.startswith("image/")
                or name.endswith((".png", ".jpg", ".jpeg", ".webp", ".bmp"))
            )
            if token and (image_like or encrypted_like):
                usable.append(item)
        return usable

    @staticmethod
    def _signature_person_inactive(value: Any) -> bool:
        """Return True when the signature person is marked resigned/transferred."""

        if value is None:
            return False
        if isinstance(value, bool):
            return value
        if isinstance(value, (int, float)):
            return bool(value)
        if isinstance(value, (list, tuple, set)):
            return any(MaintenancePortalService._signature_person_inactive(item) for item in value)
        if isinstance(value, dict):
            if any(
                MaintenancePortalService._signature_person_inactive(value.get(key))
                for key in ("checked", "value", "text", "name")
                if key in value
            ):
                return True
            return bool(value)
        text = str(value or "").strip().lower()
        if not text:
            return False
        if text in {"false", "0", "否", "未勾选", "未选", "no", "none", "null"}:
            return False
        return True

    def _load_engineer_mop_candidates(self, *, force: bool = False) -> tuple[list[dict[str, Any]], list[str], dict[str, str]]:
        settings = self._engineer_mop_settings()
        if not settings["app_token"] or not settings["table_id"]:
            return [], ["未配置 MOP 多维表 app_token/table_id，请先在 SQLite settings 中配置 mop_app_token 和 mop_table_id。"], settings
        signature = json.dumps(settings, ensure_ascii=False, sort_keys=True)
        now = time.time()
        with self._engineer_mop_cache_lock:
            cached = self._engineer_mop_cache
            if (
                not force
                and cached
                and cached.get("signature") == signature
                and now - float(cached.get("loaded_ts") or 0.0) < MOP_CANDIDATE_CACHE_TTL_SECONDS
            ):
                return (
                    copy.deepcopy(cached.get("candidates") or []),
                    list(cached.get("warnings") or []),
                    dict(cached.get("settings") or settings),
                )
        warnings: list[str] = []
        metas, meta_by_name = self._load_table_fields(
            app_token=settings["app_token"],
            table_id=settings["table_id"],
        )
        records: list[dict[str, Any]] = []
        page_token = ""
        while True:
            params = {"page_size": 500}
            if settings.get("view_id"):
                params["view_id"] = settings["view_id"]
            if page_token:
                params["page_token"] = page_token
            payload = self._request_json(
                "records",
                params=params,
                app_token=settings["app_token"],
                table_id=settings["table_id"],
            )
            data = payload.get("data", {})
            for item in data.get("items") or []:
                fields = item.get("fields") if isinstance(item.get("fields"), dict) else {}
                records.append(
                    {
                        "record_id": str(item.get("record_id") or ""),
                        "display_fields": fields,
                    }
                )
            if not data.get("has_more"):
                break
            page_token = str(data.get("page_token") or "").strip()
            if not page_token:
                break
        candidates: list[dict[str, Any]] = []
        title_names = [
            settings["title_field"],
            "文件名",
            "MOP名称",
            "MOP表格",
            "名称",
            "标题",
            "维护总项",
        ]
        for record in records:
            fields = record.get("display_fields") if isinstance(record.get("display_fields"), dict) else {}
            record_id = str(record.get("record_id") or "")
            latest_publish_time = self._attachment_latest_publish_time(fields)
            attachments = [
                self._attachment_with_cache_context(
                    attachment,
                    category="mop",
                    app_token=settings["app_token"],
                    table_id=settings["table_id"],
                    record_id=record_id,
                    latest_publish_time=latest_publish_time,
                )
                for attachment in self._extract_mop_attachments(fields, settings["attachment_field"])
            ]
            title = self._mop_field_text(fields, title_names) or str(record.get("record_id") or "MOP表格")
            candidates.append(
                {
                    "record_id": record_id,
                    "title": title,
                    "fields": fields,
                    "attachments": attachments,
                    "attachment_count": len(attachments),
                    "latest_publish_time": latest_publish_time,
                    "app_token": settings["app_token"],
                    "table_id": settings["table_id"],
                    "view_id": settings.get("view_id", ""),
                    "file_no": self._mop_field_text(fields, ["文件编号"]),
                    "specialty": self._mop_field_text(fields, ["专业"]),
                    "maintenance_type": self._mop_field_text(fields, ["维护类型"]),
                    "version": self._mop_field_text(fields, ["版本号"]),
                    "file_status": self._mop_field_text(fields, ["文件状态"]),
                }
            )
        if metas and settings["attachment_field"] not in {meta.field_name for meta in metas}:
            warnings.append(f"MOP附件字段「{settings['attachment_field']}」未在表中找到，已自动扫描所有附件字段。")
        with self._engineer_mop_cache_lock:
            self._engineer_mop_cache = {
                "signature": signature,
                "loaded_ts": now,
                "candidates": copy.deepcopy(candidates),
                "warnings": list(warnings),
                "settings": dict(settings),
            }
        return candidates, warnings, settings

    def engineer_mop_bootstrap(
        self, *, scope: str = "ALL", ongoing_items: list[dict[str, Any]] | None = None
    ) -> dict[str, Any]:
        scope = self._normalize_scope(scope)
        notices = self._engineer_month_maintenance_notices(
            scope=scope,
            ongoing_items=ongoing_items,
        )
        mop_candidates, warnings, settings = self._load_engineer_mop_candidates()
        self._maybe_start_daily_attachment_cache_refresh()
        bindings = self._state_store.list_mop_notice_bindings(
            scope=scope,
            notice_keys=[item["notice_key"] for item in notices],
            template_keys=[str(item.get("mop_template_key") or "") for item in notices],
        )
        bindings_by_notice: dict[str, dict[str, Any]] = {}
        for item in bindings:
            notice_key = str(item.get("notice_key") or "")
            if notice_key and notice_key not in bindings_by_notice:
                bindings_by_notice[notice_key] = item
        bindings_by_template: dict[str, dict[str, Any]] = {}
        for binding in bindings:
            template_key = str(binding.get("template_key") or "").strip()
            if template_key and template_key not in bindings_by_template:
                bindings_by_template[template_key] = binding
        local_candidate_by_id: dict[str, dict[str, Any]] = {}
        for binding in bindings:
            payload = binding.get("payload") if isinstance(binding.get("payload"), dict) else {}
            upload_id = self._upload_id_from_local_mop_ref(
                payload.get("upload_id"),
                allow_plain=True,
            ) or self._upload_id_from_local_mop_ref(
                binding.get("mop_record_id"),
                binding.get("mop_attachment_token"),
            )
            if not upload_id or upload_id in local_candidate_by_id:
                continue
            stored = self._state_store.get_engineer_mop_local_file(upload_id)
            if not stored:
                warnings.append(f"本地上传 MOP 已失效，请重新上传：{binding.get('mop_title') or upload_id}")
                continue
            local_candidate_by_id[upload_id] = self._engineer_mop_local_candidate(stored)
        if local_candidate_by_id:
            existing_ids = {str(item.get("record_id") or "") for item in mop_candidates}
            mop_candidates = [
                *[item for item in local_candidate_by_id.values() if str(item.get("record_id") or "") not in existing_ids],
                *mop_candidates,
            ]
        for notice in notices:
            binding = bindings_by_notice.get(notice["notice_key"])
            template_key = str(notice.get("mop_template_key") or "").strip()
            if binding and template_key and not str(binding.get("template_key") or "").strip() and template_key not in bindings_by_template:
                try:
                    backfill_payload = dict(binding.get("payload") or {})
                    backfill_payload.update(
                        {
                            "scope": binding.get("scope") or scope,
                            "notice_key": notice["notice_key"],
                            "template_key": template_key,
                            "notice_title": notice.get("title") or binding.get("notice_title") or "",
                            "notice_status": notice.get("status") or binding.get("notice_status") or "",
                            "building": notice.get("building") or "",
                            "maintenance_total": notice.get("maintenance_total") or "",
                            "maintenance_cycle": notice.get("maintenance_cycle") or "",
                            "source_record_id": notice.get("source_record_id") or binding.get("source_record_id") or "",
                            "target_record_id": notice.get("target_record_id") or binding.get("target_record_id") or "",
                            "active_item_id": notice.get("active_item_id") or binding.get("active_item_id") or "",
                            "mop_app_token": binding.get("mop_app_token") or "",
                            "mop_table_id": binding.get("mop_table_id") or "",
                            "mop_record_id": binding.get("mop_record_id") or "",
                            "mop_title": binding.get("mop_title") or "",
                            "mop_attachment_token": binding.get("mop_attachment_token") or "",
                            "mop_attachment_name": binding.get("mop_attachment_name") or "",
                            "selected_sheet": binding.get("selected_sheet") or "",
                            "updated_by": binding.get("updated_by") or "template-backfill",
                        }
                    )
                    backfilled = self._state_store.upsert_mop_notice_binding(backfill_payload)
                    binding = backfilled
                    bindings_by_template[template_key] = backfilled
                except Exception as exc:
                    warnings.append(f"MOP绑定模板补齐失败: {exc}")
            if not binding:
                template_binding = bindings_by_template.get(template_key)
                if template_binding:
                    binding = dict(template_binding)
                    binding["inherited"] = True
            notice["mop_binding"] = binding or None
        notices.sort(key=self._engineer_maintenance_notice_sort_key)
        return {
            "scope": scope,
            "scope_label": self._scope_label(scope),
            "date": dt.date.today().isoformat(),
            "notices": notices,
            "mop_candidates": mop_candidates,
            "bindings": bindings,
            "mop_settings": {
                "configured": bool(settings.get("app_token") and settings.get("table_id")),
                "title_field": settings.get("title_field", ""),
                "attachment_field": settings.get("attachment_field", ""),
            },
            "warnings": warnings,
        }

    def bind_engineer_mop_notice(
        self, *, payload: dict[str, Any], updated_by: str = ""
    ) -> dict[str, Any]:
        payload = dict(payload or {})
        if not str(payload.get("template_key") or payload.get("mop_template_key") or "").strip():
            payload["template_key"] = self._engineer_mop_template_key(
                {
                    "work_type": WORK_TYPE_MAINTENANCE,
                    "title": payload.get("notice_title") or payload.get("title") or "",
                    "building": payload.get("building") or "",
                    "maintenance_total": payload.get("maintenance_total") or "",
                    "maintenance_cycle": payload.get("maintenance_cycle") or "",
                },
                title=str(payload.get("notice_title") or payload.get("title") or ""),
                building=str(payload.get("building") or ""),
                maintenance_total=str(payload.get("maintenance_total") or ""),
                maintenance_cycle=str(payload.get("maintenance_cycle") or ""),
            )
        payload["updated_by"] = updated_by
        binding = self._state_store.upsert_mop_notice_binding(payload)
        self._touch_state_cache_version()
        return {"binding": binding}

    @staticmethod
    def _lark_response_is_token_error(response: Any) -> bool:
        code = int(getattr(response, "code", 0) or 0)
        msg = str(getattr(response, "msg", "") or "").lower()
        return code in TOKEN_ERROR_CODES or "token" in msg or "access_token" in msg

    @classmethod
    def _signature_open_id_from_any(cls, value: Any) -> str:
        """Return a Feishu open_id from user fields, text fields, or nested payloads."""

        if isinstance(value, dict):
            candidates = (
                value.get("open_id"),
                value.get("openId"),
                value.get("openid"),
                value.get("user_id"),
                value.get("userId"),
                value.get("id"),
            )
            for candidate in candidates:
                text = str(candidate or "").strip()
                if text.startswith("ou_"):
                    return text
            for nested in value.values():
                found = cls._signature_open_id_from_any(nested)
                if found:
                    return found
            return ""
        if isinstance(value, list):
            for item in value:
                found = cls._signature_open_id_from_any(item)
                if found:
                    return found
            return ""
        match = re.search(r"ou_[A-Za-z0-9_-]+", str(value or ""))
        return match.group(0) if match else ""

    @classmethod
    def _signature_open_id_from_user(cls, value: dict[str, Any]) -> str:
        """Return a real Feishu open_id from a bitable user field payload."""

        return cls._signature_open_id_from_any(value)

    @classmethod
    def _signature_user_info(cls, value: Any) -> dict[str, str]:
        if isinstance(value, list) and value:
            first = value[0]
            if isinstance(first, dict):
                return {
                    "open_id": cls._signature_open_id_from_user(first),
                    "name": str(first.get("name") or first.get("en_name") or "").strip(),
                    "email": str(first.get("email") or "").strip(),
                }
        if isinstance(value, dict):
            return {
                "open_id": cls._signature_open_id_from_user(value),
                "name": str(value.get("name") or value.get("en_name") or "").strip(),
                "email": str(value.get("email") or "").strip(),
            }
        return {"open_id": "", "name": "", "email": ""}

    @staticmethod
    def _decode_signature_png(signature_png: str) -> bytes:
        text = str(signature_png or "").strip()
        if not text:
            raise PortalError("签名图片为空，请先手写签名。")
        if "," in text and text.lower().startswith("data:image/"):
            header, text = text.split(",", 1)
            if "png" not in header.lower():
                raise PortalError("签名图片必须是 PNG 格式。")
        try:
            raw = base64.b64decode(text, validate=True)
        except Exception as exc:
            raise PortalError("签名图片格式无效，请重新签名。") from exc
        if len(raw) < 64:
            raise PortalError("签名图片内容过小，请重新签名。")
        if len(raw) > 2 * 1024 * 1024:
            raise PortalError("签名图片超过 2MB，请清空后重新签名。")
        if not raw.startswith(b"\x89PNG\r\n\x1a\n"):
            raise PortalError("签名图片必须是 PNG 格式。")
        return raw

    @staticmethod
    def _signature_attachment_version(attachment: dict[str, Any]) -> str:
        seed = "|".join(
            str(attachment.get(key) or "")
            for key in ("file_token", "url", "name", "size", "tmp_url")
        )
        return hashlib.sha1(seed.encode("utf-8", errors="ignore")).hexdigest()[:12]

    @staticmethod
    def _signature_preview_url(*, record_id: str, signature_version: str, link_token: str = "") -> str:
        record_id = str(record_id or "").strip()
        signature_version = str(signature_version or "").strip()
        if not record_id or not signature_version:
            return ""
        url = f"/api/signatures/image?record_id={quote(record_id, safe='')}&v={quote(signature_version, safe='')}"
        link_token = str(link_token or "").strip()
        if link_token:
            url += f"&token={quote(link_token, safe='')}"
        return url

    @classmethod
    def _transparent_signature_png(cls, signature_bytes: bytes) -> bytes:
        """Normalize a canvas/image signature to black ink on transparent background."""

        try:
            from PIL import Image, ImageChops, ImageFilter
        except Exception as exc:  # pragma: no cover - dependency bootstrap should provide Pillow.
            raise PortalError("缺少 Pillow 依赖，无法处理透明签名图片。") from exc

        try:
            image = Image.open(io.BytesIO(signature_bytes)).convert("RGBA")
        except Exception as exc:
            raise PortalError("签名图片无法读取，请重新签名。") from exc
        gray = image.convert("L")
        ink_alpha = gray.point(lambda p: 0 if p >= 248 else min(255, max(0, (248 - int(p)) * 7)))
        source_alpha = image.getchannel("A")
        alpha = ImageChops.multiply(ink_alpha, source_alpha)
        alpha = alpha.point(
            lambda p: 0
            if p <= 3
            else min(255, max(160, int(p) * 4))
        )
        alpha = alpha.filter(ImageFilter.MaxFilter(3))
        alpha = alpha.point(lambda p: 0 if p <= 6 else min(255, int(p) * 2))
        transparent = Image.new("RGBA", image.size, (0, 0, 0, 0))
        transparent.putalpha(alpha)
        bbox = alpha.point(lambda p: 255 if p > 8 else 0).getbbox()
        if not bbox:
            raise PortalError("签名图片未识别到有效笔迹，请重新签名。")
        left = max(0, bbox[0] - 8)
        top = max(0, bbox[1] - 8)
        right = min(image.size[0], bbox[2] + 8)
        bottom = min(image.size[1], bbox[3] + 8)
        transparent = transparent.crop((left, top, right, bottom))
        output = io.BytesIO()
        transparent.save(output, format="PNG", optimize=True)
        return output.getvalue()

    def _upload_signature_image(self, *, signature_bytes: bytes, file_name: str) -> str:
        token = str(ensure_feishu_token() or config.user_token or "").strip()
        if not token:
            raise PortalError("未配置有效的飞书 user_token，无法上传签名。")
        temp_file_path = ""
        try:
            with tempfile.NamedTemporaryFile(
                delete=False,
                prefix="clipflow_signature_",
                suffix=".png",
            ) as tmp:
                tmp.write(signature_bytes)
                temp_file_path = tmp.name

            def attempt_upload(upload_token: str) -> Any:
                with open(temp_file_path, "rb") as file_obj:
                    client = (
                        lark.Client.builder()
                        .enable_set_token(True)
                        .log_level(lark.LogLevel.ERROR)
                        .build()
                    )
                    request = (
                        UploadAllMediaRequest.builder()
                        .request_body(
                            UploadAllMediaRequestBody.builder()
                            .file_name(file_name)
                            .parent_type("bitable_image")
                            .parent_node(SIGNATURE_APP_TOKEN)
                            .size(str(len(signature_bytes)))
                            .file(file_obj)
                            .build()
                        )
                        .build()
                    )
                    option = lark.RequestOption.builder().user_access_token(upload_token).build()
                    return client.drive.v1.media.upload_all(request, option)

            response = attempt_upload(token)
            if not response.success() and self._lark_response_is_token_error(response):
                token = str(refresh_feishu_token() or config.user_token or "").strip()
                response = attempt_upload(token)
            if not response.success():
                raise PortalError(
                    f"签名图片上传失败: {response.code} - {response.msg}"
                )
            file_token = str(getattr(response.data, "file_token", "") or "").strip()
            if not file_token:
                raise PortalError("签名图片上传成功但未返回 file_token。")
            return file_token
        finally:
            if temp_file_path:
                with suppress(Exception):
                    os.remove(temp_file_path)

    def _upload_encrypted_signature_file(self, *, encrypted_bytes: bytes, file_name: str) -> str:
        token = str(ensure_feishu_token() or config.user_token or "").strip()
        if not token:
            raise PortalError("未配置有效的飞书 user_token，无法上传签名。")
        if not encrypted_bytes:
            raise PortalError("加密签名文件为空，无法上传。")
        upload_name = str(file_name or "").strip() or encrypted_signature_file_name("signature")
        if not upload_name.lower().endswith(".sigenc"):
            upload_name = f"{upload_name}.sigenc"
        temp_file_path = ""
        try:
            with tempfile.NamedTemporaryFile(
                delete=False,
                prefix="clipflow_signature_",
                suffix=".sigenc",
            ) as tmp:
                tmp.write(encrypted_bytes)
                temp_file_path = tmp.name

            def attempt_upload(upload_token: str) -> Any:
                with open(temp_file_path, "rb") as file_obj:
                    client = (
                        lark.Client.builder()
                        .enable_set_token(True)
                        .log_level(lark.LogLevel.ERROR)
                        .build()
                    )
                    request = (
                        UploadAllMediaRequest.builder()
                        .request_body(
                            UploadAllMediaRequestBody.builder()
                            .file_name(upload_name)
                            .parent_type("bitable_file")
                            .parent_node(SIGNATURE_APP_TOKEN)
                            .size(str(len(encrypted_bytes)))
                            .file(file_obj)
                            .build()
                        )
                        .build()
                    )
                    option = lark.RequestOption.builder().user_access_token(upload_token).build()
                    return client.drive.v1.media.upload_all(request, option)

            response = attempt_upload(token)
            if not response.success() and self._lark_response_is_token_error(response):
                token = str(refresh_feishu_token() or config.user_token or "").strip()
                response = attempt_upload(token)
            if not response.success():
                raise PortalError(
                    f"签名加密文件上传失败: {response.code} - {response.msg}"
                )
            file_token = str(getattr(response.data, "file_token", "") or "").strip()
            if not file_token:
                raise PortalError("签名加密文件上传成功但未返回 file_token。")
            return file_token
        finally:
            if temp_file_path:
                with suppress(Exception):
                    os.remove(temp_file_path)

    def _save_encrypted_signature_record(
        self,
        *,
        table_id: str,
        record_id: str,
        attachment_field: str,
        key_field: str,
        signature_bytes: bytes,
        display_name: str,
        source: str,
        open_id: str = "",
        employee_no: str = "",
    ) -> tuple[str, dict[str, Any]]:
        aad = self._signature_crypto.build_aad(
            app_token=SIGNATURE_APP_TOKEN,
            table_id=table_id,
            record_id=record_id,
            source=source,
            open_id=open_id,
            employee_no=employee_no,
            display_name=display_name,
        )
        encrypted_bytes, metadata = self._signature_crypto.encrypt_signature(
            signature_bytes,
            aad,
        )
        file_token = self._upload_encrypted_signature_file(
            encrypted_bytes=encrypted_bytes,
            file_name=encrypted_signature_file_name(display_name),
        )
        try:
            self._patch_record_fields(
                app_token=SIGNATURE_APP_TOKEN,
                table_id=table_id,
                record_id=record_id,
                fields={
                    attachment_field: [{"file_token": file_token}],
                    key_field: self._signature_crypto.metadata_to_text(metadata),
                },
            )
        except PortalError as exc:
            message = str(exc)
            if key_field in message or "FieldNameNotFound" in message or "字段" in message:
                raise PortalError(
                    f"签名加密字段【{key_field}】不可写或不存在，请在签名多维表中添加文本字段【{key_field}】后重试。"
                ) from exc
            raise
        return file_token, metadata

    def _mark_signature_crypto_migration(
        self,
        *,
        table_id: str,
        record_id: str,
        status: str,
        error: str = "",
        payload: dict[str, Any] | None = None,
    ) -> None:
        with suppress(Exception):
            self._state_store.upsert_signature_crypto_migration(
                table_id=table_id,
                record_id=record_id,
                status=status,
                error=error,
                payload=payload or {},
            )

    def _decode_signature_attachment_bytes(
        self,
        *,
        record_id: str,
        fields: dict[str, Any],
        attachment: dict[str, Any],
        table_id: str,
        source: str,
        person: dict[str, Any] | None = None,
    ) -> bytes:
        key_field = SIGNATURE_KEY_FIELD if table_id == SIGNATURE_TABLE_ID else TEMP_SIGNATURE_KEY_FIELD
        metadata = self._signature_crypto.metadata_from_field(fields.get(key_field))
        signature_sha = str(metadata.get("signature_sha256") or "").strip()
        if self._signature_crypto.is_encrypted_metadata(metadata) and signature_sha:
            cached = self._signature_crypto.read_cache(record_id, signature_sha)
            if cached:
                return cached
        content, _content_type = self._download_mop_attachment(attachment)
        if self._signature_crypto.is_encrypted_bytes(content):
            try:
                plain = self._signature_crypto.decrypt_signature(content, metadata)
            except SignatureCryptoError as exc:
                self._mark_signature_crypto_migration(
                    table_id=table_id,
                    record_id=record_id,
                    status="failed",
                    error=str(exc),
                    payload={"source": source},
                )
                raise PortalError(str(exc)) from exc
            png = self._transparent_signature_png(plain)
            signature_sha = signature_sha or hashlib.sha256(png).hexdigest()
            self._signature_crypto.write_cache(record_id, signature_sha, png)
            self._mark_signature_crypto_migration(
                table_id=table_id,
                record_id=record_id,
                status="encrypted",
                payload={"source": source},
            )
            return png

        png = self._transparent_signature_png(content)
        self._mark_signature_crypto_migration(
            table_id=table_id,
            record_id=record_id,
            status="pending",
            payload={"source": source},
        )
        self._maybe_migrate_plain_signature_async(
            table_id=table_id,
            record_id=record_id,
            fields=fields,
            signature_bytes=png,
            person=person or {},
            source=source,
        )
        return png

    def _maybe_migrate_plain_signature_async(
        self,
        *,
        table_id: str,
        record_id: str,
        fields: dict[str, Any],
        signature_bytes: bytes,
        person: dict[str, Any],
        source: str,
    ) -> None:
        key_field = SIGNATURE_KEY_FIELD if table_id == SIGNATURE_TABLE_ID else TEMP_SIGNATURE_KEY_FIELD
        if self._signature_crypto.is_encrypted_metadata(
            self._signature_crypto.metadata_from_field(fields.get(key_field))
        ):
            return
        migration_key = (str(table_id or ""), str(record_id or ""))
        with self._signature_crypto_migration_lock:
            if migration_key in self._signature_crypto_plain_migrations:
                return
            self._signature_crypto_plain_migrations.add(migration_key)

        def worker() -> None:
            try:
                self._mark_signature_crypto_migration(
                    table_id=table_id,
                    record_id=record_id,
                    status="migrating",
                    payload={"source": source},
                )
                if table_id == SIGNATURE_TABLE_ID:
                    attachment_field = SIGNATURE_ATTACHMENT_FIELD
                    display_name = str(person.get("name") or "signature")
                else:
                    attachment_field = TEMP_SIGNATURE_ATTACHMENT_FIELD
                    display_name = str(person.get("name") or "external_signature")
                file_token, metadata = self._save_encrypted_signature_record(
                    table_id=table_id,
                    record_id=record_id,
                    attachment_field=attachment_field,
                    key_field=key_field,
                    signature_bytes=signature_bytes,
                    display_name=display_name,
                    source=source,
                    open_id=str(person.get("open_id") or ""),
                    employee_no=str(person.get("employee_no") or ""),
                )
                self._signature_crypto.write_cache(
                    record_id,
                    str(metadata.get("signature_sha256") or ""),
                    signature_bytes,
                )
                self._mark_signature_crypto_migration(
                    table_id=table_id,
                    record_id=record_id,
                    status="done",
                    payload={"file_token": file_token, "source": source},
                )
                if table_id == SIGNATURE_TABLE_ID:
                    with self._signature_people_cache_lock:
                        self._signature_people_cache = None
                else:
                    with self._external_signature_people_cache_lock:
                        self._external_signature_people_cache = None
            except Exception as exc:
                self._mark_signature_crypto_migration(
                    table_id=table_id,
                    record_id=record_id,
                    status="failed",
                    error=str(exc),
                    payload={"source": source},
                )
            finally:
                with self._signature_crypto_migration_lock:
                    self._signature_crypto_plain_migrations.discard(migration_key)

        threading.Thread(
            target=worker,
            name=f"signature-crypto-migrate-{record_id}",
            daemon=True,
        ).start()

    def start_signature_crypto_migration_async(self, *, delay_seconds: float = 8.0) -> bool:
        with self._signature_crypto_migration_lock:
            if self._signature_crypto_migration_running:
                return False
            self._signature_crypto_migration_running = True

        def worker() -> None:
            try:
                if delay_seconds > 0:
                    time.sleep(delay_seconds)
                self._run_signature_crypto_migration()
            finally:
                with self._signature_crypto_migration_lock:
                    self._signature_crypto_migration_running = False

        threading.Thread(
            target=worker,
            name="signature-crypto-migration",
            daemon=True,
        ).start()
        return True

    def _run_signature_crypto_migration(self) -> None:
        summary = {"checked": 0, "done": 0, "skipped": 0, "failed": 0}
        tables = (
            (SIGNATURE_TABLE_ID, SIGNATURE_ATTACHMENT_FIELD, SIGNATURE_KEY_FIELD, "staff", self._load_signature_people),
            (TEMP_SIGNATURE_TABLE_ID, TEMP_SIGNATURE_ATTACHMENT_FIELD, TEMP_SIGNATURE_KEY_FIELD, "external", self._load_external_signature_people),
        )
        for table_id, _attachment_field, key_field, source, loader in tables:
            try:
                people = loader(force=True)
            except Exception as exc:
                with suppress(Exception):
                    self._state_store.put_backend_runtime(
                        "signature_crypto",
                        {
                            "status": "migration_failed",
                            "error": str(exc),
                            "summary": summary,
                            "updated_at": dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                        },
                    )
                continue
            for person in people:
                record_id = str(person.get("record_id") or "").strip()
                if not record_id:
                    continue
                fields = person.get("raw_fields") if isinstance(person.get("raw_fields"), dict) else {}
                attachments = [
                    self._attachment_with_cache_context(
                        attachment,
                        category="signature" if table_id == SIGNATURE_TABLE_ID else "temporary_signature",
                        app_token=SIGNATURE_APP_TOKEN,
                        table_id=table_id,
                        record_id=record_id,
                        latest_publish_time=str(person.get("latest_publish_time") or self._attachment_latest_publish_time(fields)),
                    )
                    for attachment in self._extract_signature_attachments(fields)
                ]
                if not attachments:
                    continue
                summary["checked"] += 1
                metadata = self._signature_crypto.metadata_from_field(fields.get(key_field))
                if self._signature_crypto.is_encrypted_metadata(metadata):
                    self._mark_signature_crypto_migration(
                        table_id=table_id,
                        record_id=record_id,
                        status="encrypted",
                        payload={"source": source},
                    )
                    summary["skipped"] += 1
                    continue
                try:
                    content, _content_type = self._download_mop_attachment(attachments[0])
                    if content.startswith(SIGNATURE_ENCRYPTED_MAGIC):
                        raise PortalError("签名附件已加密但密钥字段缺失，无法迁移。")
                    signature_bytes = self._transparent_signature_png(content)
                    self._mark_signature_crypto_migration(
                        table_id=table_id,
                        record_id=record_id,
                        status="migrating",
                        payload={"source": source},
                    )
                    if table_id == SIGNATURE_TABLE_ID:
                        display_name = str(person.get("name") or "signature")
                        attachment_field = SIGNATURE_ATTACHMENT_FIELD
                    else:
                        display_name = str(person.get("name") or "external_signature")
                        attachment_field = TEMP_SIGNATURE_ATTACHMENT_FIELD
                    file_token, saved_metadata = self._save_encrypted_signature_record(
                        table_id=table_id,
                        record_id=record_id,
                        attachment_field=attachment_field,
                        key_field=key_field,
                        signature_bytes=signature_bytes,
                        display_name=display_name,
                        source=source,
                        open_id=str(person.get("open_id") or ""),
                        employee_no=str(person.get("employee_no") or ""),
                    )
                    self._signature_crypto.write_cache(
                        record_id,
                        str(saved_metadata.get("signature_sha256") or ""),
                        signature_bytes,
                    )
                    self._mark_signature_crypto_migration(
                        table_id=table_id,
                        record_id=record_id,
                        status="done",
                        payload={"file_token": file_token, "source": source},
                    )
                    summary["done"] += 1
                except Exception as exc:
                    self._mark_signature_crypto_migration(
                        table_id=table_id,
                        record_id=record_id,
                        status="failed",
                        error=str(exc),
                        payload={"source": source},
                    )
                    summary["failed"] += 1
        with suppress(Exception):
            self._state_store.put_backend_runtime(
                "signature_crypto",
                {
                    "status": "ready",
                    "master_key_exists": self._signature_crypto.master_key_exists(),
                    "fingerprint": self._signature_crypto.master_key_fingerprint(),
                    "summary": summary,
                    "updated_at": dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                },
            )

    def _upload_bitable_file(
        self,
        *,
        file_path: str | Path,
        file_name: str = "",
        app_token: str = DEFAULT_APP_TOKEN,
    ) -> str:
        token = str(ensure_feishu_token() or config.user_token or "").strip()
        if not token:
            raise PortalError("未配置有效的飞书 user_token，无法上传 MOP 文件。")
        path = Path(str(file_path or "")).resolve()
        if not path.is_file():
            raise PortalError("待上传的已签名 MOP 文件不存在。")
        upload_name = str(file_name or path.name).strip() or path.name
        size = path.stat().st_size
        if size <= 0:
            raise PortalError("待上传的已签名 MOP 文件为空。")

        def attempt_upload(upload_token: str) -> Any:
            with open(path, "rb") as file_obj:
                client = (
                    lark.Client.builder()
                    .enable_set_token(True)
                    .log_level(lark.LogLevel.ERROR)
                    .build()
                )
                request = (
                    UploadAllMediaRequest.builder()
                    .request_body(
                        UploadAllMediaRequestBody.builder()
                        .file_name(upload_name)
                        .parent_type("bitable_file")
                        .parent_node(str(app_token or DEFAULT_APP_TOKEN))
                        .size(str(size))
                        .file(file_obj)
                        .build()
                    )
                    .build()
                )
                option = lark.RequestOption.builder().user_access_token(upload_token).build()
                return client.drive.v1.media.upload_all(request, option)

        response = attempt_upload(token)
        if not response.success() and self._lark_response_is_token_error(response):
            token = str(refresh_feishu_token() or config.user_token or "").strip()
            response = attempt_upload(token)
        if not response.success():
            raise PortalError(f"MOP 文件上传失败: {response.code} - {response.msg}")
        file_token = str(getattr(response.data, "file_token", "") or "").strip()
        if not file_token:
            raise PortalError("MOP 文件上传成功但未返回 file_token。")
        return file_token

    def _load_signature_people(self, *, force: bool = False) -> list[dict[str, Any]]:
        now = time.time()
        with self._signature_people_cache_lock:
            cached = self._signature_people_cache
            if (
                not force
                and cached
                and now - float(cached.get("loaded_ts") or 0.0)
                < SIGNATURE_PEOPLE_CACHE_TTL_SECONDS
            ):
                return copy.deepcopy(cached.get("people") or [])

        people: list[dict[str, Any]] = []
        page_token = ""
        while True:
            params = {"page_size": 500}
            if SIGNATURE_VIEW_ID:
                params["view_id"] = SIGNATURE_VIEW_ID
            if page_token:
                params["page_token"] = page_token
            payload = self._request_json(
                "records",
                params=params,
                app_token=SIGNATURE_APP_TOKEN,
                table_id=SIGNATURE_TABLE_ID,
            )
            data = payload.get("data") if isinstance(payload.get("data"), dict) else {}
            for item in data.get("items") or []:
                fields = item.get("fields") if isinstance(item.get("fields"), dict) else {}
                if self._signature_person_inactive(fields.get(SIGNATURE_INACTIVE_FIELD)):
                    continue
                user_info = self._signature_user_info(fields.get(SIGNATURE_USER_FIELD))
                open_id = user_info.get("open_id", "") or self._signature_open_id_from_any(
                    fields.get("openid")
                    or fields.get("OpenID")
                    or fields.get("open_id")
                    or fields.get("飞书OpenID")
                    or fields.get("飞书 openid")
                    or fields.get("飞书openid")
                    or fields.get("人员openid")
                )
                name = (
                    self._mop_field_text(fields, [SIGNATURE_NAME_FIELD])
                    or user_info.get("name")
                    or str(item.get("record_id") or "")
                )
                record_id = str(item.get("record_id") or "")
                latest_publish_time = self._attachment_latest_publish_time(fields)
                attachments = [
                    self._attachment_with_cache_context(
                        attachment,
                        category="signature",
                        app_token=SIGNATURE_APP_TOKEN,
                        table_id=SIGNATURE_TABLE_ID,
                        record_id=record_id,
                        latest_publish_time=latest_publish_time,
                    )
                    for attachment in self._extract_signature_attachments(fields)
                ]
                first_signature = attachments[0] if attachments else None
                signature_version = (
                    self._signature_attachment_version(first_signature)
                    if isinstance(first_signature, dict) and first_signature
                    else ""
                )
                building = self._mop_field_text(fields, ["楼栋", "机楼/专业"])
                has_signature = bool(signature_version)
                person = {
                    "record_id": record_id,
                    "name": name,
                    "open_id": open_id,
                    "email": user_info.get("email", ""),
                    "employee_no": self._mop_field_text(fields, ["员工工号", "工号"]),
                    "building": building,
                    "scope_text": building,
                    "position": self._mop_field_text(fields, ["岗位"]),
                    "team": self._mop_field_text(fields, ["班组"]),
                    "shift": self._mop_field_text(fields, ["班次"]),
                    "has_signature": has_signature,
                    "signature_count": len(attachments) if has_signature else 0,
                    "signature_version": signature_version,
                    "latest_publish_time": latest_publish_time,
                    "signature_preview_url": (
                        self._signature_preview_url(
                            record_id=record_id,
                            signature_version=signature_version,
                        )
                        if record_id and has_signature
                        else ""
                    ),
                    "raw_fields": fields,
                }
                people.append(person)
            if not data.get("has_more"):
                break
            page_token = str(data.get("page_token") or "").strip()
            if not page_token:
                break

        people.sort(
            key=lambda item: (
                0 if item.get("name") else 1,
                str(item.get("building") or ""),
                str(item.get("name") or ""),
            )
        )
        with self._signature_people_cache_lock:
            self._signature_people_cache = {
                "loaded_ts": now,
                "people": copy.deepcopy(people),
            }
        return people

    def signature_people(
        self,
        *,
        scope: str = "",
        query: str = "",
        record_id: str = "",
        link_token: str = "",
        notice_key: str = "",
        operator_open_id: str = "",
        limit: int = 80,
        refresh: bool = False,
    ) -> dict[str, Any]:
        scope = self._normalize_scope(scope or "ALL")
        query_text = re.sub(r"\s+", "", str(query or "")).lower()
        record_id = str(record_id or "").strip()
        notice_key = str(notice_key or "").strip()
        operator_open_id = str(operator_open_id or "").strip()
        people = self._load_signature_people(force=bool(refresh or record_id))
        self._maybe_start_daily_attachment_cache_refresh()

        def matches_query(person: dict[str, Any]) -> bool:
            if record_id and str(person.get("record_id") or "") != record_id:
                return False
            if not query_text:
                return True
            haystack = re.sub(
                r"\s+",
                "",
                "|".join(
                    str(person.get(key) or "")
                    for key in (
                        "record_id",
                        "name",
                        "open_id",
                        "email",
                        "employee_no",
                        "building",
                        "position",
                        "team",
                        "shift",
                    )
                ),
            ).lower()
            return query_text in haystack

        filtered = [
            {
                key: value
                for key, value in person.items()
                if key != "raw_fields"
            }
            for person in people
            if matches_query(person)
        ]
        limited = filtered[: max(1, min(500, int(limit or 80)))]
        link_token = str(link_token or "").strip()
        if link_token and record_id:
            for person in limited:
                if str(person.get("record_id") or "") != record_id:
                    continue
                signature_version = str(person.get("signature_version") or "")
                if signature_version:
                    person["signature_preview_url"] = self._signature_preview_url(
                        record_id=record_id,
                        signature_version=signature_version,
                        link_token=link_token,
                    )
        if notice_key and operator_open_id:
            for person in limited:
                signer_record_id = str(person.get("record_id") or "").strip()
                signer_open_id = str(person.get("open_id") or "").strip()
                confirmed = bool(
                    signer_open_id
                    and signer_open_id == operator_open_id
                )
                usage_status = "confirmed" if confirmed else ""
                if not confirmed and signer_record_id and signer_open_id:
                    usage_status = self._state_store.mop_signature_usage_status(
                        scope=scope,
                        notice_key=notice_key,
                        signer_record_id=signer_record_id,
                        requested_by_openid=operator_open_id,
                    )
                    confirmed = usage_status == "confirmed"
                rejected = usage_status == "rejected"
                pending = usage_status == "pending"
                person["usage_status"] = usage_status
                person["usage_confirmed"] = confirmed
                person["usage_rejected"] = rejected
                person["usage_confirmation_required"] = bool(
                    signer_open_id
                    and signer_open_id != operator_open_id
                    and not confirmed
                    and not rejected
                )
                person["usage_confirmation_pending"] = pending
        return {
            "people": limited,
            "count": len(filtered),
            "returned": len(limited),
            "scope": scope,
            "loaded_at": dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        }

    @staticmethod
    def _external_signature_preview_url(*, record_id: str, signature_version: str) -> str:
        record_id = str(record_id or "").strip()
        signature_version = str(signature_version or "").strip()
        if not record_id or not signature_version:
            return ""
        return (
            f"/api/signatures/temporary/image?record_id={quote(record_id, safe='')}"
            f"&v={quote(signature_version, safe='')}"
        )

    def _load_external_signature_people(self, *, force: bool = False) -> list[dict[str, Any]]:
        now = time.time()
        with self._external_signature_people_cache_lock:
            cached = self._external_signature_people_cache
            if (
                not force
                and cached
                and now - float(cached.get("loaded_ts") or 0.0)
                < SIGNATURE_PEOPLE_CACHE_TTL_SECONDS
            ):
                return copy.deepcopy(cached.get("people") or [])

        people: list[dict[str, Any]] = []
        page_token = ""
        while True:
            params = {"page_size": 500}
            if page_token:
                params["page_token"] = page_token
            payload = self._request_json(
                "records",
                params=params,
                app_token=SIGNATURE_APP_TOKEN,
                table_id=TEMP_SIGNATURE_TABLE_ID,
            )
            data = payload.get("data") if isinstance(payload.get("data"), dict) else {}
            for item in data.get("items") or []:
                fields = item.get("fields") if isinstance(item.get("fields"), dict) else {}
                record_id = str(item.get("record_id") or "").strip()
                if not record_id:
                    continue
                latest_publish_time = self._attachment_latest_publish_time(fields)
                attachments = [
                    self._attachment_with_cache_context(
                        attachment,
                        category="temporary_signature",
                        app_token=SIGNATURE_APP_TOKEN,
                        table_id=TEMP_SIGNATURE_TABLE_ID,
                        record_id=record_id,
                        latest_publish_time=latest_publish_time,
                    )
                    for attachment in self._extract_signature_attachments(fields)
                ]
                if not attachments:
                    continue
                first_signature = attachments[0]
                signature_version = self._signature_attachment_version(first_signature)
                name = (
                    self._mop_field_text(fields, [TEMP_SIGNATURE_NAME_FIELD, "姓名", "名称"])
                    or record_id
                )
                building = self._mop_field_text(fields, [TEMP_SIGNATURE_BUILDING_FIELD, "楼栋"])
                specialty = self._mop_field_text(fields, [TEMP_SIGNATURE_SPECIALTY_FIELD, "专业"])
                people.append(
                    {
                        "source": "external",
                        "record_id": record_id,
                        "name": name,
                        "display_name": name,
                        "building": building,
                        "scope_text": building,
                        "specialty": specialty,
                        "employee_no": self._mop_field_text(fields, [TEMP_SIGNATURE_EMPLOYEE_NO_FIELD, "工号"]),
                        "certificate": self._mop_field_text(fields, [TEMP_SIGNATURE_CERT_FIELD, "持证"]),
                        "has_signature": True,
                        "signature_count": len(attachments),
                        "signature_version": signature_version,
                        "latest_publish_time": latest_publish_time,
                        "signature_preview_url": self._external_signature_preview_url(
                            record_id=record_id,
                            signature_version=signature_version,
                        ),
                        "raw_fields": fields,
                    }
                )
            if not data.get("has_more"):
                break
            page_token = str(data.get("page_token") or "").strip()
            if not page_token:
                break

        people.sort(
            key=lambda item: (
                str(item.get("building") or ""),
                str(item.get("specialty") or ""),
                str(item.get("name") or ""),
            )
        )
        with self._external_signature_people_cache_lock:
            self._external_signature_people_cache = {
                "loaded_ts": now,
                "people": copy.deepcopy(people),
            }
        return people

    def temporary_signature_people(
        self,
        *,
        scope: str = "ALL",
        query: str = "",
        limit: int = 80,
        refresh: bool = False,
    ) -> dict[str, Any]:
        scope = self._normalize_scope(scope or "ALL")
        query_text = re.sub(r"\s+", "", str(query or "")).lower()
        people = self._load_external_signature_people(force=bool(refresh))

        def matches_query(person: dict[str, Any]) -> bool:
            if not query_text:
                return True
            haystack = re.sub(
                r"\s+",
                "",
                "|".join(
                    str(person.get(key) or "")
                    for key in (
                        "record_id",
                        "name",
                        "display_name",
                        "building",
                        "specialty",
                        "employee_no",
                        "certificate",
                    )
                ),
            ).lower()
            return query_text in haystack

        filtered = [
            {key: value for key, value in person.items() if key != "raw_fields"}
            for person in people
            if matches_query(person)
        ]
        limited = filtered[: max(1, min(500, int(limit or 80)))]
        return {
            "people": limited,
            "count": len(filtered),
            "returned": len(limited),
            "scope": scope,
            "loaded_at": dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        }

    def create_signature_link_token(self, *, record_id: str, created_by: str = "") -> dict[str, Any]:
        return self._state_store.create_signature_link_token(
            record_id=record_id,
            created_by=created_by,
            ttl_seconds=SIGNATURE_LINK_TOKEN_TTL_SECONDS,
            payload={"purpose": "signature_link"},
        )

    def validate_signature_link_token(self, *, record_id: str, token: str) -> bool:
        return self._state_store.validate_signature_link_token(
            record_id=record_id,
            token=token,
        )

    def mark_signature_link_token_used(self, *, record_id: str, token: str) -> None:
        self._state_store.mark_signature_link_token_used(
            record_id=record_id,
            token=token,
        )

    def save_signature_for_person(
        self,
        *,
        record_id: str,
        signature_png: str,
        signer_name: str = "",
        link_token: str = "",
        operator_open_id: str = "",
        operator_name: str = "",
        require_operator_match: bool = False,
    ) -> dict[str, Any]:
        record_id = str(record_id or "").strip()
        if not record_id:
            raise PortalError("请选择要保存签名的人员。")
        people = self._load_signature_people(force=True)
        person = next(
            (item for item in people if str(item.get("record_id") or "") == record_id),
            None,
        )
        if not person:
            raise PortalError("签名人员记录不存在或无权访问。")
        operator_open_id = str(operator_open_id or "").strip()
        operator_name = str(operator_name or "").strip()
        person_open_id = str(person.get("open_id") or "").strip()
        if require_operator_match:
            if not operator_open_id:
                raise PortalError("请先登录后再网页手写签名。")
            if not person_open_id or person_open_id != operator_open_id:
                raise PortalError("网页手写只能保存当前登录用户本人的签名，请发送签名链接给对方。")
        signature_bytes = self._transparent_signature_png(
            self._decode_signature_png(signature_png)
        )
        safe_name = self._safe_mop_path_part(
            signer_name or str(person.get("name") or "signature"),
            "signature",
        )
        file_token, metadata = self._save_encrypted_signature_record(
            table_id=SIGNATURE_TABLE_ID,
            record_id=record_id,
            attachment_field=SIGNATURE_ATTACHMENT_FIELD,
            key_field=SIGNATURE_KEY_FIELD,
            signature_bytes=signature_bytes,
            display_name=safe_name,
            source="staff",
            open_id=str(person.get("open_id") or ""),
            employee_no=str(person.get("employee_no") or ""),
        )
        with self._signature_people_cache_lock:
            self._signature_people_cache = None
        signature_version = str(metadata.get("signature_sha256") or hashlib.sha1(file_token.encode("utf-8")).hexdigest())[:12]
        notification_warning = ""
        notification_result: dict[str, Any] = {}
        if operator_open_id or operator_name:
            if person_open_id:
                lines = [
                    "【MOP签名保存提醒】",
                    "",
                    f"签名人员：{person.get('name') or signer_name or record_id}",
                    f"操作人：{operator_name or '未知'}"
                    + (f"（{operator_open_id}）" if operator_open_id else ""),
                    f"保存时间：{dt.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
                    "说明：你的手写签名已由上述登录人通过门户网页保存到签名库。",
                ]
                if external_mock_enabled():
                    ok, message, send_results = True, "mock external send skipped", [
                        {"open_id": person_open_id, "ok": True, "message": "mock external send skipped"}
                    ]
                else:
                    ok, message, send_results = send_text_to_open_ids("\n".join(lines), [person_open_id])
                send_result = (send_results or [{}])[0] if send_results else {}
                notification_result = {
                    "open_id": person_open_id,
                    "ok": bool(ok and send_result.get("ok", ok)),
                    "message": str(send_result.get("message") or message or ""),
                }
                if not notification_result["ok"]:
                    notification_warning = f"签名已保存，但通知签名人员失败：{notification_result['message'] or '发送失败'}"
            else:
                notification_warning = "签名已保存，但该人员缺少 openid，无法发送签名保存提醒。"
        return {
            "record_id": record_id,
            "name": str(person.get("name") or signer_name or ""),
            "file_token": file_token,
            "signature_version": signature_version,
            "signature_preview_url": self._signature_preview_url(
                record_id=record_id,
                signature_version=signature_version,
                link_token=link_token,
            ),
            "has_signature": True,
            "saved_at": dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "notification_result": notification_result,
            "notification_warning": notification_warning,
        }

    def save_external_signature_for_person(
        self,
        *,
        record_id: str,
        signature_png: str,
        signer_name: str = "",
    ) -> dict[str, Any]:
        record_id = str(record_id or "").strip()
        if not record_id:
            raise PortalError("请选择要保存签名的其他人员。")
        signature_bytes = self._transparent_signature_png(
            self._decode_signature_png(signature_png)
        )
        people = self._load_external_signature_people(force=True)
        person = next(
            (item for item in people if str(item.get("record_id") or "") == record_id),
            None,
        )
        if not person:
            raise PortalError("其他人员签名记录不存在。")
        safe_name = self._safe_mop_path_part(
            signer_name or str(person.get("name") or "external_signature"),
            "external_signature",
        )
        file_token, metadata = self._save_encrypted_signature_record(
            table_id=TEMP_SIGNATURE_TABLE_ID,
            record_id=record_id,
            attachment_field=TEMP_SIGNATURE_ATTACHMENT_FIELD,
            key_field=TEMP_SIGNATURE_KEY_FIELD,
            signature_bytes=signature_bytes,
            display_name=safe_name,
            source="external",
            employee_no=str(person.get("employee_no") or ""),
        )
        with self._external_signature_people_cache_lock:
            self._external_signature_people_cache = None
        signature_version = str(metadata.get("signature_sha256") or hashlib.sha1(file_token.encode("utf-8")).hexdigest())[:12]
        name = str(person.get("name") or signer_name or "其他人员")
        return {
            "source": "external",
            "record_id": record_id,
            "name": name,
            "display_name": name,
            "file_token": file_token,
            "signature_version": signature_version,
            "signature_preview_url": self._external_signature_preview_url(
                record_id=record_id,
                signature_version=signature_version,
            ),
            "has_signature": True,
            "signature_count": 1,
            "building": str(person.get("building") or ""),
            "scope_text": str(person.get("scope_text") or person.get("building") or ""),
            "specialty": str(person.get("specialty") or ""),
            "employee_no": str(person.get("employee_no") or ""),
            "certificate": str(person.get("certificate") or ""),
            "saved_at": dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        }

    def signature_image_bytes(self, *, record_id: str) -> tuple[bytes, str]:
        record_id = str(record_id or "").strip()
        if not record_id:
            raise PortalError("缺少签名人员记录。")
        people = self._load_signature_people(force=False)
        person = next(
            (item for item in people if str(item.get("record_id") or "") == record_id),
            None,
        )
        if not person:
            people = self._load_signature_people(force=True)
            person = next(
                (item for item in people if str(item.get("record_id") or "") == record_id),
                None,
            )
        if not person:
            raise PortalError("签名人员记录不存在。")
        fields = person.get("raw_fields") if isinstance(person.get("raw_fields"), dict) else {}
        attachments = [
            self._attachment_with_cache_context(
                attachment,
                category="signature",
                app_token=SIGNATURE_APP_TOKEN,
                table_id=SIGNATURE_TABLE_ID,
                record_id=record_id,
                latest_publish_time=str(person.get("latest_publish_time") or self._attachment_latest_publish_time(fields)),
            )
            for attachment in self._extract_signature_attachments(fields)
        ]
        if not attachments:
            raise PortalError("该人员还没有可用签名。")
        return (
            self._decode_signature_attachment_bytes(
                record_id=record_id,
                fields=fields,
                attachment=attachments[0],
                table_id=SIGNATURE_TABLE_ID,
                source="staff",
                person=person,
            ),
            "image/png",
        )

    def external_signature_image_bytes(self, *, record_id: str) -> tuple[bytes, str]:
        record_id = str(record_id or "").strip()
        if not record_id:
            raise PortalError("缺少其他人员签名记录。")
        people = self._load_external_signature_people(force=False)
        person = next(
            (item for item in people if str(item.get("record_id") or "") == record_id),
            None,
        )
        if not person:
            people = self._load_external_signature_people(force=True)
            person = next(
                (item for item in people if str(item.get("record_id") or "") == record_id),
                None,
            )
        if not person:
            raise PortalError("其他人员签名记录不存在。")
        fields = person.get("raw_fields") if isinstance(person.get("raw_fields"), dict) else {}
        attachments = [
            self._attachment_with_cache_context(
                attachment,
                category="temporary_signature",
                app_token=SIGNATURE_APP_TOKEN,
                table_id=TEMP_SIGNATURE_TABLE_ID,
                record_id=record_id,
                latest_publish_time=str(person.get("latest_publish_time") or self._attachment_latest_publish_time(fields)),
            )
            for attachment in self._extract_signature_attachments(fields)
        ]
        if not attachments:
            raise PortalError("该其他人员还没有可用签名。")
        return (
            self._decode_signature_attachment_bytes(
                record_id=record_id,
                fields=fields,
                attachment=attachments[0],
                table_id=TEMP_SIGNATURE_TABLE_ID,
                source="external",
                person=person,
            ),
            "image/png",
        )

    @staticmethod
    def _url_host_for_display(host: str) -> str:
        host = str(host or "").strip()
        if not host:
            return ""
        if ":" in host and not host.startswith("[") and not host.endswith("]"):
            return f"[{host}]"
        return host

    def _signature_public_base_url(
        self,
        *,
        scope: str = "",
        request_base_url: str = "",
    ) -> str:
        """Build the public portal URL used in signature links.

        The configured handover LAN address is used as the public host,
        because signature links are usually opened from another phone. The
        actual signature endpoint still belongs to this portal service, so the
        portal port is kept instead of reusing a handover/review service port.
        A browser-origin URL is only used as fallback when no public handover
        address has been configured.
        """

        base_text = str(request_base_url or "").strip().rstrip("/")
        if base_text and "://" not in base_text:
            base_text = f"http://{base_text}"
        parsed_base = urlparse(base_text) if base_text else None
        scheme = (
            parsed_base.scheme
            if parsed_base and parsed_base.scheme in {"http", "https"}
            else "http"
        )
        port = (
            parsed_base.port
            if parsed_base and parsed_base.port
            else int(getattr(config, "lan_template_portal_port", 18766) or 18766)
        )
        host = parsed_base.hostname if parsed_base and parsed_base.hostname else ""

        configured_public_host = str(
            getattr(config, "lan_template_public_host", "") or ""
        ).strip().rstrip("/")
        if configured_public_host:
            public_text = configured_public_host
            if "://" not in public_text:
                public_text = f"{scheme}://{public_text}"
            try:
                parsed_public = urlparse(public_text)
                public_host = str(parsed_public.hostname or "").strip()
                if public_host and public_host not in {"0.0.0.0", "::"}:
                    public_scheme = (
                        parsed_public.scheme
                        if parsed_public.scheme in {"http", "https"}
                        else scheme
                    )
                    public_port = parsed_public.port or port
                    return (
                        f"{public_scheme}://"
                        f"{self._url_host_for_display(public_host)}:{public_port}"
                    )
            except Exception:
                pass

        normalized_scope = self._normalize_scope(scope or "ALL")
        handover_links = self.get_handover_links().get("links") or {}
        handover_url = ""
        handover_host_selected = False
        if isinstance(handover_links, dict):
            handover_url = str(handover_links.get(normalized_scope) or "").strip()
            if not handover_url and normalized_scope != "ALL":
                handover_url = str(handover_links.get("ALL") or "").strip()
        try:
            parsed_handover = urlparse(handover_url)
            handover_host = str(parsed_handover.hostname or "").strip()
            if handover_host and handover_host not in {"0.0.0.0", "::"}:
                handover_scheme = (
                    parsed_handover.scheme
                    if parsed_handover.scheme in {"http", "https"}
                    else scheme
                )
                scheme = handover_scheme
                host = handover_host
                handover_host_selected = True
        except Exception:
            pass

        if handover_host_selected:
            return f"{scheme}://{self._url_host_for_display(host)}:{port}"

        if (
            parsed_base
            and parsed_base.scheme in {"http", "https"}
            and parsed_base.netloc
            and str(parsed_base.hostname or "").strip()
            not in {"127.0.0.1", "localhost", "0.0.0.0", "::"}
        ):
            return f"{parsed_base.scheme}://{parsed_base.netloc}"

        configured_host = str(getattr(config, "lan_template_portal_host", "") or "").strip()
        if (
            (not host or host in {"127.0.0.1", "localhost", "0.0.0.0", "::"})
            and configured_host
            and configured_host not in {"0.0.0.0", "::"}
        ):
            host = configured_host
        if not host or host in {"0.0.0.0", "::"}:
            host = "127.0.0.1"
        return f"{scheme}://{self._url_host_for_display(host)}:{port}"

    def _signature_person_for_link(self, *, record_id: str) -> dict[str, Any]:
        record_id = str(record_id or "").strip()
        if not record_id:
            raise PortalError("请选择要发送签名链接的人员。")
        people = self._load_signature_people(force=False)
        person = next(
            (item for item in people if str(item.get("record_id") or "") == record_id),
            None,
        )
        if not person:
            people = self._load_signature_people(force=True)
            person = next(
                (item for item in people if str(item.get("record_id") or "") == record_id),
                None,
            )
        if not person:
            raise PortalError("签名人员记录不存在。")
        open_id = str(person.get("open_id") or "").strip()
        if not open_id:
            raise PortalError("该人员记录缺少员工姓名 openid，无法发送签名链接。")
        return {
            key: value
            for key, value in person.items()
            if key != "raw_fields"
        }

    def build_signature_link_message(
        self,
        *,
        record_id: str,
        signer_name: str = "",
        scope: str = "",
        request_base_url: str = "",
        created_by: str = "",
    ) -> dict[str, Any]:
        person = self._signature_person_for_link(record_id=record_id)
        base_url = self._signature_public_base_url(
            scope=scope,
            request_base_url=request_base_url,
        )
        token_data = self.create_signature_link_token(
            record_id=record_id,
            created_by=created_by,
        )
        link_token = str(token_data.get("token") or "").strip()
        link_url = (
            f"{base_url}/signature?record_id={quote(str(record_id or '').strip(), safe='')}"
            f"&token={quote(link_token, safe='')}"
        )
        name = str(signer_name or person.get("name") or "签名人员").strip()
        text = "\n".join(
            [
                "【线上签名】请完成 MOP 手写签名",
                "",
                f"签名人员：{name}",
                f"签名链接：{link_url}",
                "",
                "请用手机打开链接，在页面手写签名并保存。",
            ]
        )
        return {
            "record_id": str(record_id or "").strip(),
            "person": person,
            "open_id": str(person.get("open_id") or "").strip(),
            "link_url": link_url,
            "expires_at": token_data.get("expires_at"),
            "text": text,
        }

    def build_signature_usage_confirmation_messages(
        self,
        *,
        scope: str,
        notice_key: str,
        notice_title: str,
        signatures: list[dict[str, Any]],
        mop_attachment_name: str = "",
        request_base_url: str = "",
        operator_open_id: str = "",
        operator_name: str = "",
    ) -> dict[str, Any]:
        operator_open_id = str(operator_open_id or "").strip()
        operator_name = str(operator_name or "").strip()
        mop_attachment_name = str(mop_attachment_name or "").strip()
        people = self._load_signature_people(force=False)
        people_by_id = {
            str(item.get("record_id") or "").strip(): item
            for item in people
            if str(item.get("record_id") or "").strip()
        }
        requested: dict[str, set[str]] = {}
        for item in signatures or []:
            if not isinstance(item, dict):
                continue
            if str(item.get("source") or "staff").strip() not in {"", "staff"}:
                continue
            record_id = str(item.get("record_id") or "").strip()
            role = str(item.get("role") or "").strip()
            if not record_id or role not in {"implementer", "auditor"}:
                continue
            requested.setdefault(record_id, set()).add(role)
        missing = [record_id for record_id in requested if record_id not in people_by_id]
        if missing:
            people = self._load_signature_people(force=True)
            people_by_id = {
                str(item.get("record_id") or "").strip(): item
                for item in people
                if str(item.get("record_id") or "").strip()
            }
        base_url = self._signature_public_base_url(
            scope=scope,
            request_base_url=request_base_url,
        )
        messages: list[dict[str, Any]] = []
        skipped: list[dict[str, Any]] = []
        for record_id, roles in requested.items():
            person = people_by_id.get(record_id)
            if not person:
                skipped.append({"record_id": record_id, "reason": "签名人员记录不存在"})
                continue
            person_open_id = str(person.get("open_id") or "").strip()
            if not person_open_id:
                skipped.append({"record_id": record_id, "name": person.get("name") or "", "reason": "缺少 openid"})
                continue
            if operator_open_id and person_open_id == operator_open_id:
                skipped.append({"record_id": record_id, "name": person.get("name") or "", "reason": "本人签名无需确认"})
                continue
            role_text = "、".join(self._mop_role_label(role) for role in ("implementer", "auditor") if role in roles)
            confirmation = self._state_store.create_mop_signature_usage_confirmation(
                scope=scope,
                notice_key=notice_key,
                role="auditor" if "auditor" in roles else "implementer",
                signer_record_id=record_id,
                signer_open_id=person_open_id,
                signer_name=str(person.get("name") or ""),
                requested_by_openid=operator_open_id,
                requested_by_name=operator_name,
                payload={
                    "notice_title": notice_title,
                    "mop_attachment_name": mop_attachment_name,
                    "roles": sorted(roles),
                    "role_text": role_text,
                },
            )
            token = str(confirmation.get("token") or "").strip()
            link_url = f"{base_url}/api/signatures/usage-confirm?token={quote(token, safe='')}"
            text = "\n".join(
                [
                    "【MOP签名使用确认】",
                    "",
                    f"签名人员：{person.get('name') or record_id}",
                    f"签名角色：{role_text or '签名人员'}",
                    f"维护通告：{notice_title or '未命名维保通告'}",
                    f"MOP附件：{mop_attachment_name or '当前选中附件'}",
                    f"操作人：{operator_name or '未知'}"
                    + (f"（{operator_open_id}）" if operator_open_id else ""),
                    "",
                    f"确认链接：{link_url}",
                    "请确认是否允许本次 MOP 使用你的已保存签名。",
                ]
            )
            messages.append(
                {
                    "record_id": record_id,
                    "person": {key: value for key, value in person.items() if key != "raw_fields"},
                    "open_id": person_open_id,
                    "link_url": link_url,
                    "text": text,
                    "confirmation_id": confirmation.get("confirmation_id"),
                }
            )
        return {
            "messages": messages,
            "skipped": skipped,
        }

    def confirm_signature_usage(self, *, token: str) -> dict[str, Any]:
        return self._state_store.confirm_mop_signature_usage(token=token)

    def signature_usage_confirmation(self, *, token: str) -> dict[str, Any]:
        return self._state_store.get_mop_signature_usage_confirmation(token=token)

    def decide_signature_usage(self, *, token: str, decision: str) -> dict[str, Any]:
        return self._state_store.decide_mop_signature_usage(
            token=token,
            decision=decision,
        )

    def reject_signature_usage(self, *, token: str) -> dict[str, Any]:
        return self._state_store.reject_mop_signature_usage(token=token)

    @staticmethod
    def _mop_role_label(role: str) -> str:
        return "维护审核人" if str(role or "") == "auditor" else "维护实施人"

    def _temporary_signature_public_url(
        self,
        *,
        temp_id: str,
        token: str,
        scope: str = "",
        request_base_url: str = "",
    ) -> str:
        base_url = self._signature_public_base_url(
            scope=scope,
            request_base_url=request_base_url,
        )
        return (
            f"{base_url}/signature?temporary_id={quote(str(temp_id or '').strip(), safe='')}"
            f"&token={quote(str(token or '').strip(), safe='')}"
        )

    def _temporary_signature_display_name(
        self,
        *,
        scope: str,
        notice_key: str,
        created_by: str = "",
    ) -> str:
        sessions = self._state_store.list_mop_temporary_signature_sessions(
            scope=scope,
            notice_key=notice_key,
            created_by=created_by,
            include_expired=True,
        )
        return f"临时人员{len(sessions) + 1}"

    def build_temporary_signature_link_message(
        self,
        *,
        scope: str = "ALL",
        notice_key: str = "",
        role: str = "implementer",
        recipient_open_ids: list[str] | None = None,
        notice_title: str = "",
        specialty: str = "",
        display_name: str = "",
        request_base_url: str = "",
        created_by: str = "",
    ) -> dict[str, Any]:
        scope = self._normalize_scope(scope or "ALL")
        notice_key = str(notice_key or "").strip()
        role = str(role or "implementer").strip()
        if role not in {"implementer", "auditor"}:
            raise PortalError("临时签名角色无效。")
        recipients = [
            str(item or "").strip()
            for item in (recipient_open_ids or [])
            if str(item or "").strip()
        ]
        if not recipients:
            raise PortalError("请先选择维护实施人，再发送其他人员签名链接。")
        display_name = str(display_name or "").strip() or self._temporary_signature_display_name(
            scope=scope,
            notice_key=notice_key,
            created_by=created_by,
        )
        session = self._state_store.create_mop_temporary_signature_session(
            scope=scope,
            notice_key=notice_key,
            role=role,
            display_name=display_name,
            recipient_open_ids=recipients,
            created_by=created_by,
            ttl_seconds=SIGNATURE_LINK_TOKEN_TTL_SECONDS,
            payload={
                "notice_title": str(notice_title or ""),
                "specialty": str(specialty or ""),
            },
        )
        link_url = self._temporary_signature_public_url(
            temp_id=str(session.get("temp_id") or ""),
            token=str(session.get("token") or ""),
            scope=scope,
            request_base_url=request_base_url,
        )
        role_label = self._mop_role_label(role)
        text = "\n".join(
            [
                "【线上签名】请现场完成 MOP 其他人员签名",
                "",
                f"签名角色：{role_label}",
                f"临时人员：{display_name}",
                f"维护通告：{notice_title or '未命名维保通告'}",
                f"签名链接：{link_url}",
                "",
                "请用手机打开链接，让现场人员在页面手写签名并保存。",
            ]
        )
        return {
            **{key: value for key, value in session.items() if key != "token"},
            "link_url": link_url,
            "open_ids": recipients,
            "text": text,
            "signature": self._public_temporary_signature_session(
                session,
                link_token=str(session.get("token") or ""),
            ),
        }

    def create_temporary_signature_session(
        self,
        *,
        scope: str,
        notice_key: str,
        role: str,
        notice_title: str = "",
        specialty: str = "",
        display_name: str = "",
        created_by: str = "",
    ) -> dict[str, Any]:
        scope = self._normalize_scope(scope)
        notice_key = str(notice_key or "").strip()
        if not notice_key:
            raise PortalError("当前通告缺少记忆键，无法创建临时签名。")
        role = str(role or "implementer").strip()
        if role not in {"implementer", "auditor"}:
            raise PortalError("临时签名角色无效。")
        display_name = str(display_name or "").strip() or self._temporary_signature_display_name(
            scope=scope,
            notice_key=notice_key,
            created_by=created_by,
        )
        session = self._state_store.create_mop_temporary_signature_session(
            scope=scope,
            notice_key=notice_key,
            role=role,
            display_name=display_name,
            recipient_open_ids=[],
            created_by=created_by,
            ttl_seconds=SIGNATURE_LINK_TOKEN_TTL_SECONDS,
            payload={
                "notice_title": str(notice_title or ""),
                "specialty": str(specialty or ""),
            },
        )
        return self._public_temporary_signature_session(
            session,
            link_token=str(session.get("token") or ""),
        )

    def build_existing_temporary_signature_link_message(
        self,
        *,
        temp_id: str,
        request_base_url: str = "",
    ) -> dict[str, Any]:
        temp_id = str(temp_id or "").strip()
        if not temp_id:
            raise PortalError("请选择要重发链接的临时人员。")
        session = self._state_store.refresh_mop_temporary_signature_session_token(
            temp_id=temp_id,
            ttl_seconds=SIGNATURE_LINK_TOKEN_TTL_SECONDS,
        )
        scope = self._normalize_scope(str(session.get("scope") or "ALL"))
        token = str(session.get("token") or "")
        link_url = self._temporary_signature_public_url(
            temp_id=temp_id,
            token=token,
            scope=scope,
            request_base_url=request_base_url,
        )
        role = str(session.get("role") or "implementer")
        role_label = self._mop_role_label(role)
        payload = session.get("payload") if isinstance(session.get("payload"), dict) else {}
        display_name = str(session.get("display_name") or "临时人员").strip()
        notice_title = str(payload.get("notice_title") or "").strip()
        recipients = [
            str(item or "").strip()
            for item in (session.get("recipient_open_ids") or [])
            if str(item or "").strip()
        ]
        if not recipients:
            raise PortalError("该临时人员缺少接收人 openid，无法重发链接。")
        text = "\n".join(
            [
                "【线上签名】请现场完成 MOP 其他人员签名",
                "",
                f"签名角色：{role_label}",
                f"临时人员：{display_name}",
                f"维护通告：{notice_title or '未命名维保通告'}",
                f"签名链接：{link_url}",
                "",
                "请用手机打开链接，让现场人员在页面手写签名并保存。",
            ]
        )
        return {
            **{key: value for key, value in session.items() if key != "token"},
            "link_url": link_url,
            "open_ids": recipients,
            "text": text,
            "signature": self._public_temporary_signature_session(
                session,
                link_token=token,
            ),
        }

    def _public_temporary_signature_session(
        self,
        session: dict[str, Any],
        *,
        link_token: str = "",
    ) -> dict[str, Any]:
        payload = session.get("payload") if isinstance(session.get("payload"), dict) else {}
        file_token = str(session.get("signature_file_token") or "").strip()
        temp_id = str(session.get("temp_id") or "").strip()
        status = str(session.get("status") or "pending").strip() or "pending"
        preview_url = ""
        if file_token and temp_id:
            preview_url = (
                f"/api/signatures/temporary/image?temporary_id={quote(temp_id, safe='')}"
                f"&v={quote(hashlib.sha1(file_token.encode('utf-8')).hexdigest()[:12], safe='')}"
            )
            if link_token:
                preview_url += f"&token={quote(link_token, safe='')}"
        return {
            "source": "temporary",
            "temp_id": temp_id,
            "record_id": str(session.get("temporary_record_id") or ""),
            "role": str(session.get("role") or ""),
            "display_name": str(session.get("display_name") or ""),
            "name": str(session.get("display_name") or ""),
            "status": status,
            "has_signature": bool(file_token and status == "signed"),
            "signature_file_token": file_token,
            "signature_preview_url": preview_url,
            "expires_at": session.get("expires_at"),
            "notice_title": str(payload.get("notice_title") or ""),
            "specialty": str(payload.get("specialty") or ""),
        }

    def temporary_signature_session(
        self,
        *,
        temp_id: str,
        token: str,
    ) -> dict[str, Any]:
        session = self._state_store.get_mop_temporary_signature_session(
            temp_id=temp_id,
            token=token,
            require_valid_token=True,
        )
        if not session:
            raise PortalError("临时签名链接无效或已过期。")
        return self._public_temporary_signature_session(session, link_token=token)

    def list_temporary_signatures(
        self,
        *,
        scope: str = "ALL",
        notice_key: str = "",
        created_by: str = "",
    ) -> dict[str, Any]:
        scope = self._normalize_scope(scope or "ALL")
        sessions = self._state_store.list_mop_temporary_signature_sessions(
            scope=scope,
            notice_key=str(notice_key or "").strip(),
            created_by=str(created_by or "").strip(),
        )
        return {
            "items": [
                self._public_temporary_signature_session(session)
                for session in sessions
            ],
            "count": len(sessions),
        }

    def _field_option_write_values(
        self,
        meta_by_name: dict[str, FieldMeta],
        field_name: str,
        values: list[str],
    ) -> list[str]:
        meta = meta_by_name.get(field_name)
        if not meta:
            return []
        available = set(meta.option_names or [])
        result: list[str] = []
        for value in values:
            text = str(value or "").strip()
            if text and text in available and text not in result:
                result.append(text)
        return result

    def save_temporary_signature(
        self,
        *,
        temp_id: str,
        token: str,
        signature_png: str,
    ) -> dict[str, Any]:
        require_token = bool(str(token or "").strip())
        session = self._state_store.get_mop_temporary_signature_session(
            temp_id=temp_id,
            token=token,
            require_valid_token=require_token,
        )
        if not session:
            raise PortalError("临时签名链接无效或已过期。")
        signature_bytes = self._transparent_signature_png(
            self._decode_signature_png(signature_png)
        )
        display_name = str(session.get("display_name") or "临时人员").strip()
        _metas, meta_by_name = self._load_table_fields(
            app_token=SIGNATURE_APP_TOKEN,
            table_id=TEMP_SIGNATURE_TABLE_ID,
        )
        payload = session.get("payload") if isinstance(session.get("payload"), dict) else {}
        building_values = self._field_option_write_values(
            meta_by_name,
            TEMP_SIGNATURE_BUILDING_FIELD,
            [self._scope_label(str(session.get("scope") or "")), str(session.get("scope") or "")],
        )
        specialty_values = self._field_option_write_values(
            meta_by_name,
            TEMP_SIGNATURE_SPECIALTY_FIELD,
            [str(payload.get("specialty") or "")],
        )
        fields: dict[str, Any] = {
            TEMP_SIGNATURE_NAME_FIELD: display_name,
        }
        if building_values:
            fields[TEMP_SIGNATURE_BUILDING_FIELD] = building_values
        if specialty_values:
            fields[TEMP_SIGNATURE_SPECIALTY_FIELD] = specialty_values
        existing_record_id = str(session.get("temporary_record_id") or "").strip()
        if existing_record_id:
            self._patch_record_fields(
                app_token=SIGNATURE_APP_TOKEN,
                table_id=TEMP_SIGNATURE_TABLE_ID,
                record_id=existing_record_id,
                fields=fields,
            )
            record_id = existing_record_id
        else:
            created = self._create_record_fields(
                app_token=SIGNATURE_APP_TOKEN,
                table_id=TEMP_SIGNATURE_TABLE_ID,
                fields=fields,
            )
            data = created.get("data") if isinstance(created.get("data"), dict) else {}
            record = data.get("record") if isinstance(data.get("record"), dict) else data
            record_id = str(record.get("record_id") or record.get("id") or "").strip()
            if not record_id:
                raise PortalError("临时签名保存成功但未返回记录 ID。")
        safe_name = self._safe_mop_path_part(display_name, "temporary")
        file_token, metadata = self._save_encrypted_signature_record(
            table_id=TEMP_SIGNATURE_TABLE_ID,
            record_id=record_id,
            attachment_field=TEMP_SIGNATURE_ATTACHMENT_FIELD,
            key_field=TEMP_SIGNATURE_KEY_FIELD,
            signature_bytes=signature_bytes,
            display_name=safe_name,
            source="temporary",
        )
        updated = self._state_store.update_mop_temporary_signature_session(
            temp_id=str(session.get("temp_id") or ""),
            status="signed",
            temporary_record_id=record_id,
            signature_file_token=file_token,
            payload_patch={
                "saved_at": dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "signature_crypto": metadata,
            },
        )
        with self._external_signature_people_cache_lock:
            self._external_signature_people_cache = None
        return self._public_temporary_signature_session(
            updated or {},
            link_token=token,
        )

    def temporary_signature_image_bytes(
        self,
        *,
        temp_id: str,
    ) -> tuple[bytes, str]:
        session = self._state_store.get_mop_temporary_signature_session(temp_id=temp_id)
        if not session:
            raise PortalError("临时签名记录不存在。")
        record_id = str(session.get("temporary_record_id") or "").strip()
        if record_id:
            return self.external_signature_image_bytes(record_id=record_id)
        file_token = str(session.get("signature_file_token") or "").strip()
        if not file_token:
            raise PortalError("临时人员还没有可用签名。")
        content, _content_type = self._download_mop_attachment(
            {"file_token": file_token},
            cache_category="temporary_signature",
            app_token=SIGNATURE_APP_TOKEN,
            table_id=TEMP_SIGNATURE_TABLE_ID,
            record_id=str(session.get("temporary_record_id") or session.get("temp_id") or ""),
            latest_publish_time=str(session.get("updated_at") or ""),
        )
        return self._transparent_signature_png(content), "image/png"

    @staticmethod
    def _column_index(cell_ref: str) -> int:
        letters = "".join(ch for ch in str(cell_ref or "") if ch.isalpha()).upper()
        index = 0
        for ch in letters:
            index = index * 26 + (ord(ch) - 64)
        return max(1, index)

    @classmethod
    def _cell_position(cls, cell_ref: str) -> tuple[int, int] | None:
        match = re.match(r"^\$?([A-Za-z]+)\$?(\d+)$", str(cell_ref or "").strip())
        if not match:
            return None
        col = cls._column_index(match.group(1))
        row = int(match.group(2))
        if row <= 0 or col <= 0:
            return None
        return row, col

    @classmethod
    def _parse_merge_range(cls, ref: str) -> tuple[int, int, int, int] | None:
        parts = str(ref or "").split(":")
        if len(parts) != 2:
            return None
        start = cls._cell_position(parts[0])
        end = cls._cell_position(parts[1])
        if not start or not end:
            return None
        row1, col1 = start
        row2, col2 = end
        return min(row1, row2), min(col1, col2), max(row1, row2), max(col1, col2)

    @staticmethod
    def _column_label(index: int) -> str:
        index = max(1, int(index or 1))
        label = ""
        while index:
            index, remainder = divmod(index - 1, 26)
            label = chr(65 + remainder) + label
        return label

    @staticmethod
    def _excel_points_to_pixels(value: Any, *, fallback: float = 15.0) -> int:
        try:
            points = float(value)
        except Exception:
            points = float(fallback)
        if points <= 0:
            points = float(fallback)
        return max(1, int(round(points * 96 / 72)))

    @classmethod
    def _mop_worksheet_row_height_px(cls, worksheet: Any, *, row: int) -> int:
        row = max(1, int(row or 1))
        default_height = getattr(getattr(worksheet, "sheet_format", None), "defaultRowHeight", None) or 15
        row_height = getattr(getattr(worksheet, "row_dimensions", {}).get(row), "height", None)
        return cls._excel_points_to_pixels(row_height or default_height, fallback=float(default_height or 15))

    @classmethod
    def _mop_signature_max_height_px(cls, worksheet: Any, *, row: int) -> int:
        return max(1, int(round(cls._mop_worksheet_row_height_px(worksheet, row=row) * 1.5)))

    @classmethod
    def _mop_signature_target_column(cls, worksheet: Any, *, row: int, label_col: int) -> int:
        """Return the column immediately to the right of the signature label cell.

        If the label cell is part of a merged range, the signature goes into the
        first cell after that merged range. This keeps the label text intact and
        places signatures in the user's expected fill area.
        """

        row = max(1, int(row or 1))
        label_col = max(1, int(label_col or 1))
        label_ref = f"{cls._column_label(label_col)}{row}"
        for merged_range in getattr(getattr(worksheet, "merged_cells", None), "ranges", []) or []:
            try:
                if label_ref in merged_range:
                    return int(getattr(merged_range, "max_col", label_col)) + 1
            except Exception:
                continue
        return label_col + 1

    def _attachment_cache_paths(
        self,
        *,
        category: str,
        app_token: str,
        table_id: str,
        record_id: str,
        attachment: dict[str, Any],
    ) -> tuple[Path, Path]:
        name = str(attachment.get("name") or "attachment").strip() or "attachment"
        suffix = Path(name).suffix
        if not suffix:
            mime_type = str(attachment.get("mime_type") or "").lower()
            if "png" in mime_type:
                suffix = ".png"
            elif "jpeg" in mime_type or "jpg" in mime_type:
                suffix = ".jpg"
            elif "webp" in mime_type:
                suffix = ".webp"
            elif "csv" in mime_type:
                suffix = ".csv"
            else:
                suffix = ".xlsx"
        stable_seed = json.dumps(
            [
                str(category or ""),
                str(app_token or ""),
                str(table_id or ""),
                str(record_id or ""),
                str(attachment.get("file_token") or ""),
                str(attachment.get("url") or ""),
                name,
            ],
            ensure_ascii=False,
            sort_keys=True,
        )
        digest = hashlib.sha1(stable_seed.encode("utf-8", errors="ignore")).hexdigest()[:16]
        cache_dir = Path(
            get_data_file_path(
                os.path.join(
                    "attachment_cache",
                    self._safe_mop_path_part(category or "attachment", "attachment"),
                    self._safe_mop_path_part(app_token or "app", "app"),
                    self._safe_mop_path_part(table_id or "table", "table"),
                    self._safe_mop_path_part(record_id or "record", "record"),
                )
            )
        )
        cache_name = f"{self._safe_mop_path_part(Path(name).stem or 'attachment', 'attachment')}_{digest}{suffix}"
        return cache_dir / cache_name, cache_dir / f"{cache_name}.meta.json"

    @staticmethod
    def _attachment_meta_matches(expected: dict[str, str], actual: dict[str, Any]) -> bool:
        for key in ("category", "app_token", "table_id", "record_id", "file_token", "url", "name", "size"):
            if str(actual.get(key) or "") != str(expected.get(key) or ""):
                return False
        expected_time = str(expected.get("latest_publish_time") or "")
        actual_time = str(actual.get("latest_publish_time") or "")
        if expected_time and expected_time != actual_time:
            return False
        return True

    def _read_cached_attachment(
        self,
        *,
        category: str,
        app_token: str,
        table_id: str,
        record_id: str,
        attachment: dict[str, Any],
        latest_publish_time: str = "",
    ) -> tuple[bytes, str] | None:
        expected = self._attachment_cache_signature(
            category=category,
            app_token=app_token,
            table_id=table_id,
            record_id=record_id,
            attachment=attachment,
            latest_publish_time=latest_publish_time,
        )
        file_path, meta_path = self._attachment_cache_paths(
            category=category,
            app_token=app_token,
            table_id=table_id,
            record_id=record_id,
            attachment=attachment,
        )
        with self._attachment_cache_lock:
            if not file_path.is_file() or not meta_path.is_file():
                return None
            try:
                meta = json.loads(meta_path.read_text(encoding="utf-8"))
            except Exception:
                return None
            if not isinstance(meta, dict) or not self._attachment_meta_matches(expected, meta):
                return None
            try:
                return file_path.read_bytes(), str(meta.get("content_type") or attachment.get("mime_type") or "")
            except Exception:
                return None

    def _write_cached_attachment(
        self,
        *,
        category: str,
        app_token: str,
        table_id: str,
        record_id: str,
        attachment: dict[str, Any],
        latest_publish_time: str = "",
        content: bytes,
        content_type: str,
    ) -> None:
        if not content:
            return
        expected = self._attachment_cache_signature(
            category=category,
            app_token=app_token,
            table_id=table_id,
            record_id=record_id,
            attachment=attachment,
            latest_publish_time=latest_publish_time,
        )
        file_path, meta_path = self._attachment_cache_paths(
            category=category,
            app_token=app_token,
            table_id=table_id,
            record_id=record_id,
            attachment=attachment,
        )
        meta = {
            **expected,
            "content_type": str(content_type or attachment.get("mime_type") or ""),
            "cached_at": dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "size_bytes": len(content),
            "file_name": file_path.name,
        }
        with self._attachment_cache_lock:
            file_path.parent.mkdir(parents=True, exist_ok=True)
            file_path.write_bytes(content)
            meta_path.write_text(json.dumps(meta, ensure_ascii=False, indent=2), encoding="utf-8")

    @staticmethod
    def _mop_checkbox_value(original: Any, state: str) -> str:
        return MaintenancePortalService._mop_choice_value(original, state)

    @classmethod
    def _mop_choice_options(cls, value: Any) -> list[dict[str, Any]]:
        text = cls._mop_plain_text(value)
        if not text:
            return []
        options: list[dict[str, Any]] = []
        seen: set[str] = set()
        prefix_pattern = re.compile(r"([□☐■☑√✔✓])([^□☐■☑√✔✓\[\]；;，,、/]+)")
        for match in prefix_pattern.finditer(text):
            label = match.group(2).strip()
            if not label or label in seen:
                continue
            seen.add(label)
            checked = match.group(1) in {"☑", "■", "√", "✔", "✓"}
            options.append(
                {
                    "key": f"opt{len(options)}",
                    "label": label,
                    "checked": checked,
                    "style": "prefix_checkbox",
                }
            )
        if len(options) >= 2:
            return options

        options = []
        seen.clear()
        bracket_pattern = re.compile(r"([^\[\]□☐■☑√✔✓；;，,、/]{1,24})\[(.*?)\]")
        for match in bracket_pattern.finditer(text):
            label = match.group(1).strip()
            if "：" in label or ":" in label:
                label = re.split(r"[:：]", label)[-1].strip()
            if not label or label in seen:
                continue
            seen.add(label)
            checked = bool(str(match.group(2) or "").strip())
            options.append(
                {
                    "key": f"opt{len(options)}",
                    "label": label,
                    "checked": checked,
                    "style": "suffix_bracket",
                }
            )
        return options if len(options) >= 2 else []

    @classmethod
    def _mop_choice_label_for_state(cls, value: Any, state: str) -> str:
        state_text = str(state or "").strip()
        state_lower = state_text.lower()
        options = cls._mop_choice_options(value)
        if not options:
            if state_lower == "normal":
                return "正常"
            if state_lower == "abnormal":
                return "异常"
            return state_text
        for option in options:
            label = str(option.get("label") or "")
            key = str(option.get("key") or "")
            if state_text and state_text in {label, key}:
                return label
        if state_lower == "normal":
            for option in options:
                label = str(option.get("label") or "")
                if "正常" in label:
                    return label
        if state_lower == "abnormal":
            for option in options:
                label = str(option.get("label") or "")
                if "异常" in label:
                    return label
        checked = next((str(item.get("label") or "") for item in options if item.get("checked")), "")
        return checked or str(options[0].get("label") or "")

    @classmethod
    def _mop_choice_value(cls, original: Any, state: str) -> str:
        raw_text = str(original or "")
        text = raw_text or "□正常 □异常"
        selected_label = cls._mop_choice_label_for_state(text, state)
        options = cls._mop_choice_options(text)
        if not selected_label and options:
            selected_label = str(options[0].get("label") or "")

        prefix_pattern = re.compile(r"([□☐■☑√✔✓])(\s*)([^□☐■☑√✔✓\[\]；;，,、/\s]+)")
        prefix_changed = False

        def replace_prefix(match: re.Match[str]) -> str:
            nonlocal prefix_changed
            label = match.group(3).strip()
            if not any(str(item.get("label") or "") == label for item in options):
                return match.group(0)
            prefix_changed = True
            mark = "☑" if label == selected_label else "□"
            return f"{mark}{match.group(2)}{match.group(3)}"

        text = prefix_pattern.sub(replace_prefix, text)
        if prefix_changed:
            return text

        bracket_pattern = re.compile(r"([^\[\]□☐■☑√✔✓；;，,、/]{1,24})\[(.*?)\]")
        bracket_changed = False

        def replace_bracket(match: re.Match[str]) -> str:
            nonlocal bracket_changed
            label = match.group(1).strip()
            option_label = re.split(r"[:：]", label)[-1].strip()
            if not any(str(item.get("label") or "") == option_label for item in options):
                return match.group(0)
            bracket_changed = True
            mark = "√" if option_label == selected_label else " "
            return f"{match.group(1)}[{mark}]"

        text = bracket_pattern.sub(replace_bracket, text)
        if bracket_changed:
            return text

        if options:
            return " ".join(
                f"{'☑' if str(item.get('label') or '') == selected_label else '□'}{item.get('label') or ''}"
                for item in options
            )
        normal_mark = "☑" if str(state or "").strip().lower() == "normal" else "□"
        abnormal_mark = "☑" if str(state or "").strip().lower() == "abnormal" else "□"
        return f"{normal_mark}正常 {abnormal_mark}异常"

    @staticmethod
    def _mop_cell_int(value: Any, fallback: int = 0) -> int:
        try:
            return int(value)
        except Exception:
            return fallback

    @staticmethod
    def _safe_mop_path_part(value: str, fallback: str = "MOP") -> str:
        text = str(value or "").strip() or fallback
        text = re.sub(r'[<>:"/\\|?*\x00-\x1f]+', "_", text)
        text = re.sub(r"\s+", " ", text).strip(" ._")
        return (text or fallback)[:96]

    @staticmethod
    def _mop_plain_text(value: Any) -> str:
        return re.sub(r"\s+", "", str(value or "").replace("\u3000", "")).strip()

    @classmethod
    def _is_mop_cover_sheet(cls, sheet_name: str, rows: list[list[str]]) -> bool:
        name = cls._mop_plain_text(sheet_name)
        if any(token in name for token in ("封面", "目录", "说明")):
            return True
        first_cells = "".join(
            cls._mop_plain_text(cell)
            for row in rows[:20]
            for cell in row[:8]
        )
        cover_tokens = ("文件封面", "封面", "版本履历", "修订记录")
        return any(token in first_cells for token in cover_tokens) and not any(
            token in first_cells for token in ("□正常", "□异常", "维护实施人", "维护开始时间", "维护完成时间")
        )

    @classmethod
    def _is_mop_checkbox_cell(cls, value: Any) -> bool:
        return bool(cls._mop_choice_options(value))

    @classmethod
    def _is_mop_datetime_placeholder(cls, value: Any) -> bool:
        text = cls._mop_plain_text(value)
        if not text:
            return False
        return all(token in text for token in ("年", "月", "日", "时")) and (
            "__" in text
            or "年月日时" in text
            or re.search(r"年\s*月\s*日\s*时", str(value or ""))
            or not re.search(r"\d{4}年\d{1,2}月\d{1,2}日\d{1,2}时", text)
        )

    @staticmethod
    def _mop_merged_target_col(
        *,
        row_index: int,
        col_index: int,
        merges: list[dict[str, int]] | None,
    ) -> int:
        for merge in merges or []:
            try:
                merge_row = int(merge.get("row") or 0)
                merge_col = int(merge.get("col") or 0)
                rowspan = max(1, int(merge.get("rowspan") or 1))
                colspan = max(1, int(merge.get("colspan") or 1))
            except Exception:
                continue
            if (
                merge_row <= row_index < merge_row + rowspan
                and merge_col <= col_index < merge_col + colspan
            ):
                return merge_col + colspan
        return col_index + 1

    @classmethod
    def _extract_mop_sheet_targets(
        cls,
        *,
        sheet_name: str,
        rows: list[list[str]],
        merges: list[dict[str, int]] | None = None,
    ) -> dict[str, Any]:
        is_cover = cls._is_mop_cover_sheet(sheet_name, rows)
        checkbox_cells: list[dict[str, Any]] = []
        maintenance_fields: list[dict[str, Any]] = []
        if is_cover:
            return {
                "is_cover": True,
                "checkbox_cells": checkbox_cells,
                "maintenance_fields": maintenance_fields,
                "fillable_count": 0,
            }
        labels = (
            "维护实施人",
            "维护开始时间",
            "维护完成时间",
            "维护完成情况",
            "维护审核人",
            "审核确认时间",
        )
        same_column_time_labels = {"维护完成时间", "审核确认时间"}
        checkbox_columns: list[int] = []
        for row_index, row in enumerate(rows):
            for col_zero, value in enumerate(row):
                if cls._is_mop_checkbox_cell(value):
                    if col_zero not in checkbox_columns:
                        checkbox_columns.append(col_zero)
                    options = cls._mop_choice_options(value)
                    option_labels = {str(item.get("label") or "") for item in options}
                    checkbox_cells.append(
                        {
                            "sheet": sheet_name,
                            "row": row_index,
                            "row_number": row_index + 1,
                            "col": col_zero,
                            "column": cls._column_label(col_zero + 1),
                            "cell_ref": f"{cls._column_label(col_zero + 1)}{row_index + 1}",
                            "value": str(value or ""),
                            "kind": (
                                "normal_abnormal"
                                if {"正常", "异常"}.issubset(option_labels)
                                else "choice_group"
                            ),
                            "options": options,
                        }
                    )
        for row_index, row in enumerate(rows):
            for col_index, value in enumerate(row):
                text = cls._mop_plain_text(value)
                if not text:
                    continue
                if cls._is_mop_checkbox_cell(value):
                    continue
                datetime_placeholder = cls._is_mop_datetime_placeholder(value)
                for label in labels:
                    if label not in text:
                        continue
                    inline_value = ""
                    label_pos = text.find(label)
                    if label_pos >= 0:
                        inline_value = text[label_pos + len(label) :].lstrip(":：")
                    value_col = cls._mop_merged_target_col(
                        row_index=row_index,
                        col_index=col_index,
                        merges=merges,
                    )
                    value_text = ""
                    if label not in {"维护实施人", "维护审核人"}:
                        for candidate_col in range(col_index + 1, min(len(row), col_index + 8)):
                            candidate = str(row[candidate_col] or "").strip()
                            if candidate:
                                value_col = candidate_col
                                value_text = candidate
                                break
                        if inline_value and not value_text:
                            value_col = col_index
                            value_text = inline_value
                    target_value_cols = [value_col]
                    if label in same_column_time_labels and checkbox_columns:
                        target_value_cols = [col_index] if col_index in checkbox_columns else list(checkbox_columns)
                    for target_value_col in target_value_cols:
                        target_value = ""
                        if target_value_col < len(row):
                            target_value = str(row[target_value_col] or "").strip()
                        if target_value_col == value_col:
                            target_value = value_text
                        if inline_value and target_value_col == col_index and not target_value:
                            target_value = inline_value
                        value_col = target_value_col
                        value_text = target_value
                        maintenance_fields.append(
                            {
                                "sheet": sheet_name,
                                "label": label,
                                "row": row_index,
                                "row_number": row_index + 1,
                                "label_col": col_index,
                                "label_cell_ref": f"{cls._column_label(col_index + 1)}{row_index + 1}",
                                "value_col": value_col,
                                "value_column": cls._column_label(value_col + 1),
                                "value_cell_ref": f"{cls._column_label(value_col + 1)}{row_index + 1}",
                                "value": value_text,
                                "kind": "maintenance_meta",
                            }
                        )
                    break
                else:
                    if not datetime_placeholder:
                        continue
                    if any(
                        int(field.get("row") or -1) == row_index
                        and int(field.get("value_col") or -1) == col_index
                        for field in maintenance_fields
                    ):
                        continue
                    label = "日期时间"
                    for candidate_col in range(col_index - 1, max(-1, col_index - 5), -1):
                        candidate = cls._mop_plain_text(row[candidate_col] if candidate_col < len(row) else "")
                        if candidate and not cls._is_mop_checkbox_cell(candidate):
                            label = candidate.rstrip(":：")
                            break
                    maintenance_fields.append(
                        {
                            "sheet": sheet_name,
                            "label": label,
                            "row": row_index,
                            "row_number": row_index + 1,
                            "label_col": col_index,
                            "label_cell_ref": f"{cls._column_label(col_index + 1)}{row_index + 1}",
                            "value_col": col_index,
                            "value_column": cls._column_label(col_index + 1),
                            "value_cell_ref": f"{cls._column_label(col_index + 1)}{row_index + 1}",
                            "value": str(value or ""),
                            "kind": "datetime_placeholder",
                        }
                    )
        return {
            "is_cover": False,
            "checkbox_cells": checkbox_cells,
            "maintenance_fields": maintenance_fields,
            "fillable_count": len(checkbox_cells) + len(maintenance_fields),
        }

    def _save_mop_attachment_copy(
        self,
        content: bytes,
        *,
        scope: str,
        mop_record_id: str,
        file_name: str,
        content_type: str = "",
    ) -> dict[str, Any]:
        scope_part = self._safe_mop_path_part(self._scope_label(self._normalize_scope(scope)) or scope, "ALL")
        date_part = dt.date.today().isoformat()
        original_name = self._safe_mop_path_part(file_name or "MOP表格.xlsx", "MOP表格.xlsx")
        suffix = Path(original_name).suffix
        if not suffix:
            if "csv" in str(content_type or "").lower():
                suffix = ".csv"
            else:
                suffix = ".xlsx"
        stem = self._safe_mop_path_part(Path(original_name).stem or "MOP表格", "MOP表格")
        digest = hashlib.sha1(content).hexdigest()[:10]
        save_dir = Path(get_data_file_path(os.path.join("engineer_mop_files", date_part, scope_part)))
        save_dir.mkdir(parents=True, exist_ok=True)
        save_path = save_dir / f"{stem}_{digest}{suffix}"
        save_path.write_bytes(content)
        data_root = Path(get_data_file_path("")).resolve()
        try:
            relative_path = str(save_path.resolve().relative_to(data_root))
        except Exception:
            relative_path = save_path.name
        return {
            "file_name": save_path.name,
            "path": str(save_path),
            "relative_path": relative_path,
            "size": len(content),
            "saved_at": dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        }

    @staticmethod
    def _local_mop_record_id(upload_id: str) -> str:
        text = str(upload_id or "").strip()
        return f"local:{text}" if text and not text.startswith("local:") else text

    @staticmethod
    def _upload_id_from_local_mop_ref(*values: Any, allow_plain: bool = False) -> str:
        for value in values:
            text = str(value or "").strip()
            if not text:
                continue
            if text.startswith("local:"):
                return text.split(":", 1)[1].strip()
            if allow_plain and re.fullmatch(r"[0-9a-fA-F]{16,64}", text):
                return text
        return ""

    def _engineer_mop_local_file_meta(
        self,
        *,
        path: str,
        file_name: str,
        size: int,
        upload_id: str = "",
    ) -> dict[str, Any]:
        data_root = Path(get_data_file_path("")).resolve()
        source_path = Path(str(path or "")).resolve()
        try:
            relative_path = str(source_path.relative_to(data_root))
        except Exception:
            relative_path = source_path.name
        return {
            "file_name": str(file_name or source_path.name),
            "path": str(source_path),
            "relative_path": relative_path,
            "size": int(size or 0),
            "upload_id": str(upload_id or ""),
            "saved_at": dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        }

    def _engineer_mop_local_candidate(self, item: dict[str, Any]) -> dict[str, Any]:
        upload_id = str(item.get("upload_id") or "").strip()
        file_name = str(item.get("original_file_name") or "本地MOP表格.xlsx").strip()
        record_id = self._local_mop_record_id(upload_id)
        attachment = {
            "name": file_name,
            "file_token": record_id,
            "upload_id": upload_id,
            "size": int(item.get("file_size") or 0),
            "source": "local_upload",
            "local_upload": True,
        }
        return {
            "record_id": record_id,
            "upload_id": upload_id,
            "source": "local_upload",
            "local_upload": True,
            "title": Path(file_name).stem or "本地上传 MOP",
            "app_token": "",
            "table_id": "",
            "file_no": "",
            "specialty": "",
            "maintenance_type": "",
            "version": "",
            "file_status": "本地上传",
            "attachment_count": 1,
            "attachments": [attachment],
            "fields": {
                "来源": "本地上传",
                "文件名": file_name,
            },
            "warnings": item.get("warnings") or [],
            "detected": item.get("detected") or {},
        }

    def _convert_xls_to_xlsx_content(self, content: bytes) -> bytes:
        try:
            import xlrd
            from openpyxl import Workbook
        except Exception as exc:  # pragma: no cover - dependency bootstrap should provide these.
            raise PortalError("缺少 xlrd/openpyxl 依赖，无法识别 xls 文件。") from exc
        try:
            workbook = xlrd.open_workbook(file_contents=content, formatting_info=True)
        except Exception:
            workbook = xlrd.open_workbook(file_contents=content, formatting_info=False)
        output = Workbook()
        default_sheet = output.active
        for sheet_index, sheet in enumerate(workbook.sheets()):
            worksheet = default_sheet if sheet_index == 0 else output.create_sheet()
            safe_title = re.sub(r"[\[\]\:\*\?\/\\]", "_", str(sheet.name or f"Sheet{sheet_index + 1}")).strip() or f"Sheet{sheet_index + 1}"
            worksheet.title = safe_title[:31]
            for row_index in range(sheet.nrows):
                for col_index in range(sheet.ncols):
                    value = sheet.cell_value(row_index, col_index)
                    worksheet.cell(row=row_index + 1, column=col_index + 1).value = value
            for row1, row2, col1, col2 in getattr(sheet, "merged_cells", []) or []:
                if row2 > row1 + 1 or col2 > col1 + 1:
                    with suppress(Exception):
                        worksheet.merge_cells(
                            start_row=row1 + 1,
                            start_column=col1 + 1,
                            end_row=row2,
                            end_column=col2,
                        )
        buffer = io.BytesIO()
        output.save(buffer)
        return buffer.getvalue()

    def _parse_mop_local_file_content(
        self,
        content: bytes,
        *,
        file_name: str,
    ) -> tuple[dict[str, Any], bytes, str]:
        suffix = Path(str(file_name or "")).suffix.lower()
        if suffix == ".xls":
            converted = self._convert_xls_to_xlsx_content(content)
            return self._parse_xlsx_preview(converted), converted, ".xlsx"
        if suffix in {".xlsx", ".xlsm"} or content[:2] == b"PK":
            return self._parse_xlsx_preview(content), content, suffix or ".xlsx"
        raise PortalError("当前仅支持上传 xlsx/xlsm/xls 格式的 MOP 表格。")

    def upload_engineer_mop_local_file(
        self,
        *,
        scope: str = "ALL",
        source_record_id: str = "",
        notice_key: str = "",
        notice_title: str = "",
        file_name: str = "",
        content: bytes = b"",
        created_by_openid: str = "",
    ) -> dict[str, Any]:
        source_record_id = str(source_record_id or "").strip()
        if not source_record_id:
            raise PortalError("请先选择一条维保通告，再上传本地 MOP。")
        if not content:
            raise PortalError("上传文件为空，请重新选择 MOP 文件。")
        if len(content) > MOP_LOCAL_UPLOAD_MAX_BYTES:
            raise PortalError("MOP 文件超过 20MB，请压缩或更换文件后再上传。")
        original_name = str(file_name or "").strip() or "本地MOP表格.xlsx"
        suffix = Path(original_name).suffix.lower()
        if suffix not in MOP_LOCAL_UPLOAD_EXTENSIONS:
            raise PortalError("仅支持上传 xlsx、xlsm、xls 格式的 MOP 表格。")
        with suppress(Exception):
            self._state_store.mark_old_engineer_mop_local_files_deleted(
                older_than_ts=time.time() - 30 * 24 * 60 * 60
            )

        self.ensure_snapshot_loaded()
        record = self._find_record_by_id(source_record_id, WORK_TYPE_MAINTENANCE)
        display_fields = record.get("display_fields") if isinstance(record.get("display_fields"), dict) else {}
        building_codes = self._building_codes_from_value(display_fields.get("楼栋"))
        normalized_scope = self._normalize_scope(scope)
        if not self._scope_matches_buildings(normalized_scope, building_codes):
            raise PortalError("当前账号无权为该楼栋上传 MOP 文件。")

        upload_id = uuid.uuid4().hex
        scope_part = self._safe_mop_path_part(self._scope_label(normalized_scope) or normalized_scope, "ALL")
        source_part = self._safe_mop_path_part(source_record_id, "source")
        save_dir = Path(get_data_file_path(os.path.join("mop_local_uploads", scope_part, source_part, upload_id)))
        save_dir.mkdir(parents=True, exist_ok=True)
        safe_name = self._safe_mop_path_part(original_name, "本地MOP表格.xlsx")
        original_path = save_dir / safe_name
        original_path.write_bytes(content)

        warnings: list[str] = []
        try:
            parsed, working_content, working_suffix = self._parse_mop_local_file_content(
                content,
                file_name=original_name,
            )
        except PortalError:
            raise
        except Exception as exc:
            raise PortalError(f"MOP 文件识别失败：{exc}") from exc

        working_path = original_path
        if suffix == ".xls":
            working_path = save_dir / f"{self._safe_mop_path_part(Path(safe_name).stem, '本地MOP表格')}_converted.xlsx"
            working_path.write_bytes(working_content)
            warnings.append("已将 xls 文件转换为 xlsx 工作副本，最终签名会写入转换后的文件。")
        elif suffix == ".xlsm":
            warnings.append("已识别 xlsm 文件；签名写入时会尽量保留工作簿结构。")

        editable_summary = self._mop_editable_summary(parsed.get("sheets") or [])
        if not editable_summary.get("field_count") and not editable_summary.get("checkbox_count"):
            warnings.append("部分字段未识别，可打开后手动编辑单元格继续。")
        detected = {
            "parser": parsed.get("parser") or ("xls" if suffix == ".xls" else "xlsx"),
            "sheet_count": len(parsed.get("sheets") or []),
            "editable_summary": editable_summary,
        }
        stored = self._state_store.upsert_engineer_mop_local_file(
            {
                "upload_id": upload_id,
                "scope": normalized_scope,
                "source_record_id": source_record_id,
                "notice_key": notice_key,
                "notice_title": notice_title,
                "original_file_name": original_name,
                "local_file_path": str(working_path),
                "file_size": len(content),
                "status": "ready",
                "detected": detected,
                "warnings": warnings,
                "payload": {
                    "original_file_path": str(original_path),
                    "working_file_path": str(working_path),
                    "working_suffix": working_suffix,
                },
                "created_by_openid": created_by_openid,
            }
        )
        candidate = self._engineer_mop_local_candidate(stored)
        return {
            "upload_id": upload_id,
            "file_name": original_name,
            "file_size": len(content),
            "detected_summary": detected,
            "warnings": warnings,
            "local_mop_candidate": candidate,
        }

    @staticmethod
    def _normalize_mop_fill_memory_name(
        *, mop_title: str = "", mop_file_name: str = ""
    ) -> str:
        source = str(mop_file_name or mop_title or "").strip()
        if mop_file_name:
            source = Path(source).stem
        source = re.sub(r"(_[0-9a-fA-F]{8,40})$", "", source)
        source = re.sub(r"_已签名_\d{6}_[0-9a-fA-F]{6,16}$", "", source)
        source = re.sub(r"\.(xlsx|xlsm|xls|csv)$", "", source, flags=re.IGNORECASE)
        source = source.replace("　", " ")
        source = re.sub(r"\s+", "", source)
        source = re.sub(r"[^\w\u4e00-\u9fff]+", "", source, flags=re.UNICODE)
        return source.lower()

    def _mop_fill_memory_key(self, *, mop_title: str = "", mop_file_name: str = "") -> str:
        normalized = self._normalize_mop_fill_memory_name(
            mop_title=mop_title,
            mop_file_name=mop_file_name,
        )
        if not normalized:
            return ""
        return "mop-fill:" + hashlib.sha1(normalized.encode("utf-8")).hexdigest()

    def _build_mop_fill_memory_payload(
        self,
        *,
        scope: str,
        source_record_id: str = "",
        notice_title: str = "",
        notice_key: str = "",
        sheet_name: str = "",
        fields: list[dict[str, Any]] | None = None,
        checkboxes: list[dict[str, Any]] | None = None,
        cell_edits: list[dict[str, Any]] | None = None,
    ) -> dict[str, Any]:
        return {
            "scope": str(scope or "ALL"),
            "source_record_id": str(source_record_id or ""),
            "notice_title": str(notice_title or ""),
            "notice_key": str(notice_key or ""),
            "sheet_name": str(sheet_name or ""),
            "fields": [item for item in (fields or []) if isinstance(item, dict)],
            "checkboxes": [item for item in (checkboxes or []) if isinstance(item, dict)],
            "cell_edits": [item for item in (cell_edits or []) if isinstance(item, dict)],
            "saved_at": dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        }

    @staticmethod
    def _mop_editable_summary(sheets: list[dict[str, Any]]) -> dict[str, Any]:
        checkbox_count = 0
        maintenance_field_count = 0
        editable_sheet_count = 0
        for sheet in sheets:
            if not isinstance(sheet, dict) or sheet.get("is_cover"):
                continue
            sheet_checkbox_count = len(sheet.get("checkbox_cells") or [])
            sheet_field_count = len(sheet.get("maintenance_fields") or [])
            checkbox_count += sheet_checkbox_count
            maintenance_field_count += sheet_field_count
            if sheet_checkbox_count or sheet_field_count:
                editable_sheet_count += 1
        return {
            "editable_sheet_count": editable_sheet_count,
            "checkbox_count": checkbox_count,
            "maintenance_field_count": maintenance_field_count,
            "total": checkbox_count + maintenance_field_count,
        }

    @staticmethod
    def _xml_text(element: ET.Element | None) -> str:
        if element is None:
            return ""
        return "".join(element.itertext()).strip()

    @classmethod
    def _mop_preview_row_heights(
        cls,
        sheet_root: ET.Element,
        ns: dict[str, str],
        *,
        max_rows: int = 0,
    ) -> tuple[int, dict[str, int]]:
        default_height = 15.0
        sheet_format = sheet_root.find("main:sheetFormatPr", ns)
        if sheet_format is not None:
            try:
                default_height = float(sheet_format.attrib.get("defaultRowHeight") or default_height)
            except Exception:
                default_height = 15.0
        default_height_px = cls._excel_points_to_pixels(default_height)
        row_heights: dict[str, int] = {}
        for row in sheet_root.findall("main:sheetData/main:row", ns):
            try:
                row_number = int(row.attrib.get("r") or 0)
            except Exception:
                row_number = 0
            if row_number <= 0 or (max_rows and row_number > max_rows):
                continue
            height = row.attrib.get("ht")
            if not height:
                continue
            row_heights[str(row_number - 1)] = cls._excel_points_to_pixels(height, fallback=default_height)
        return default_height_px, row_heights

    def _parse_xlsx_preview(self, content: bytes, *, max_rows: int = 0, max_cols: int = 0) -> dict[str, Any]:
        ns = {"main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
        rel_ns = {"rel": "http://schemas.openxmlformats.org/package/2006/relationships"}
        with zipfile.ZipFile(io.BytesIO(content)) as archive:
            shared_strings: list[str] = []
            if "xl/sharedStrings.xml" in archive.namelist():
                root = ET.fromstring(archive.read("xl/sharedStrings.xml"))
                for si in root.findall("main:si", ns):
                    shared_strings.append("".join(si.itertext()))
            workbook = ET.fromstring(archive.read("xl/workbook.xml"))
            rels = ET.fromstring(archive.read("xl/_rels/workbook.xml.rels"))
            rel_map = {
                rel.attrib.get("Id"): rel.attrib.get("Target", "")
                for rel in rels.findall("rel:Relationship", rel_ns)
            }
            sheets: list[dict[str, Any]] = []
            for sheet in workbook.findall("main:sheets/main:sheet", ns):
                name = str(sheet.attrib.get("name") or "Sheet")
                rel_id = sheet.attrib.get("{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id")
                target = str(rel_map.get(rel_id) or "").lstrip("/")
                path = target if target.startswith("xl/") else f"xl/{target}"
                if path not in archive.namelist():
                    continue
                sheet_root = ET.fromstring(archive.read(path))
                default_row_height_px, row_heights = self._mop_preview_row_heights(
                    sheet_root,
                    ns,
                    max_rows=max_rows,
                )
                rows_out: list[list[str]] = []
                max_width = 0
                truncated = False
                merges: list[dict[str, int]] = []
                for merge in sheet_root.findall("main:mergeCells/main:mergeCell", ns):
                    parsed = self._parse_merge_range(merge.attrib.get("ref", ""))
                    if not parsed:
                        continue
                    row1, col1, row2, col2 = parsed
                    if (max_rows and row1 > max_rows) or (max_cols and col1 > max_cols):
                        continue
                    row2 = min(row2, max_rows) if max_rows else row2
                    col2 = min(col2, max_cols) if max_cols else col2
                    if row2 <= row1 and col2 <= col1:
                        continue
                    max_width = max(max_width, col2)
                    merges.append(
                        {
                            "row": row1 - 1,
                            "col": col1 - 1,
                            "rowspan": row2 - row1 + 1,
                            "colspan": col2 - col1 + 1,
                        }
                    )
                for row in sheet_root.findall("main:sheetData/main:row", ns):
                    row_values: list[str] = []
                    for cell in row.findall("main:c", ns):
                        col_index = self._column_index(cell.attrib.get("r", "")) or len(row_values) + 1
                        while len(row_values) < col_index - 1:
                            row_values.append("")
                        value = ""
                        cell_type = cell.attrib.get("t")
                        if cell_type == "inlineStr":
                            value = self._xml_text(cell.find("main:is", ns))
                        else:
                            raw = self._xml_text(cell.find("main:v", ns))
                            if cell_type == "s":
                                try:
                                    value = shared_strings[int(raw)]
                                except Exception:
                                    value = raw
                            else:
                                value = raw
                        row_values.append(value)
                    max_width = max(max_width, len(row_values))
                    rows_out.append(row_values[:max_cols] if max_cols else row_values)
                    if max_rows and len(rows_out) >= max_rows:
                        truncated = True
                        break
                targets = self._extract_mop_sheet_targets(sheet_name=name, rows=rows_out, merges=merges)
                target_width = max(
                    [max_width]
                    + [
                        int(field.get("value_col") or 0) + 1
                        for field in targets.get("maintenance_fields", [])
                        if isinstance(field, dict)
                    ]
                    + [
                        int(cell.get("col") or 0) + 1
                        for cell in targets.get("checkbox_cells", [])
                        if isinstance(cell, dict)
                    ]
                )
                sheets.append(
                    {
                        "name": name,
                        "rows": rows_out,
                        "row_count": len(rows_out),
                        "column_count": min(target_width, max_cols) if max_cols else target_width,
                        "columns": [
                            self._column_label(index + 1)
                            for index in range(min(target_width, max_cols) if max_cols else target_width)
                        ],
                        "merges": merges,
                        "default_row_height_px": default_row_height_px,
                        "row_heights": row_heights,
                        "truncated": truncated or bool(max_cols and target_width > max_cols),
                        **targets,
                    }
                )
        return {"sheets": sheets, "parser": "xlsx"}

    def _parse_csv_preview(self, content: bytes, *, file_name: str = "") -> dict[str, Any]:
        text = ""
        for encoding in ("utf-8-sig", "gb18030", "utf-16"):
            try:
                text = content.decode(encoding)
                break
            except Exception:
                continue
        if not text:
            text = content.decode("utf-8", errors="replace")
        rows = [row for row in csv.reader(io.StringIO(text))]
        sheet_name = Path(file_name or "CSV").stem or "CSV"
        targets = self._extract_mop_sheet_targets(sheet_name=sheet_name, rows=rows)
        max_width = max(
            [max([len(row) for row in rows] or [0])]
            + [
                int(field.get("value_col") or 0) + 1
                for field in targets.get("maintenance_fields", [])
                if isinstance(field, dict)
            ]
            + [
                int(cell.get("col") or 0) + 1
                for cell in targets.get("checkbox_cells", [])
                if isinstance(cell, dict)
            ]
        )
        return {
            "sheets": [
                {
                    "name": sheet_name,
                    "rows": rows,
                    "row_count": len(rows),
                    "column_count": max_width,
                    "columns": [self._column_label(index + 1) for index in range(max_width)],
                    "merges": [],
                    "default_row_height_px": self._excel_points_to_pixels(15),
                    "row_heights": {},
                    "truncated": False,
                    **targets,
                }
            ],
            "parser": "csv",
        }

    def _download_mop_attachment(
        self,
        attachment: dict[str, Any],
        *,
        cache_category: str = "",
        app_token: str = "",
        table_id: str = "",
        record_id: str = "",
        latest_publish_time: str = "",
        force_download: bool = False,
    ) -> tuple[bytes, str]:
        cache_category = str(cache_category or attachment.get("cache_category") or "mop").strip()
        app_token = str(app_token or attachment.get("app_token") or "").strip()
        table_id = str(table_id or attachment.get("table_id") or "").strip()
        record_id = str(record_id or attachment.get("record_id") or "").strip()
        latest_publish_time = str(latest_publish_time or attachment.get("latest_publish_time") or "").strip()
        if cache_category and app_token and table_id and record_id and not force_download:
            cached = self._read_cached_attachment(
                category=cache_category,
                app_token=app_token,
                table_id=table_id,
                record_id=record_id,
                attachment=attachment,
                latest_publish_time=latest_publish_time,
            )
            if cached is not None:
                return cached
        url = str(attachment.get("url") or "").strip()
        token = str(attachment.get("file_token") or "").strip()
        headers = self._auth_headers()
        if url:
            content, content_type = self._http_client.request_bytes("GET", url, headers=headers)
        else:
            if not token:
                raise PortalError("MOP附件缺少 file_token 或下载地址。")
            encoded = quote(token, safe="")
            download_url = f"https://open.feishu.cn/open-apis/drive/v1/medias/{encoded}/download"
            content, content_type = self._http_client.request_bytes("GET", download_url, headers=headers)
        if cache_category and app_token and table_id and record_id:
            self._write_cached_attachment(
                category=cache_category,
                app_token=app_token,
                table_id=table_id,
                record_id=record_id,
                attachment=attachment,
                latest_publish_time=latest_publish_time,
                content=content,
                content_type=content_type,
            )
        return content, content_type

    def _attachment_cache_daily_marker_path(self) -> Path:
        return Path(get_data_file_path(os.path.join("attachment_cache", ATTACHMENT_CACHE_DAILY_MARKER)))

    def _maybe_start_daily_attachment_cache_refresh(self) -> None:
        today = dt.date.today().isoformat()
        now = time.time()
        marker_path = self._attachment_cache_daily_marker_path()
        try:
            marker = json.loads(marker_path.read_text(encoding="utf-8")) if marker_path.is_file() else {}
        except Exception:
            marker = {}
        if isinstance(marker, dict) and marker.get("date") == today:
            status = str(marker.get("status") or "").strip()
            if status == "success":
                return
            try:
                next_retry_at = float(marker.get("next_retry_at") or 0)
            except Exception:
                next_retry_at = 0.0
            if next_retry_at > now:
                return
        with self._attachment_cache_refresh_lock:
            if self._attachment_cache_refresh_running:
                return
            self._attachment_cache_refresh_running = True
        thread = threading.Thread(
            target=self._refresh_attachment_cache_daily,
            args=(today,),
            name="clipflow-attachment-cache-refresh",
            daemon=True,
        )
        thread.start()

    def _refresh_attachment_cache_daily(self, today: str) -> None:
        stats = {
            "date": today,
            "started_at": dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "mop_checked": 0,
            "signature_checked": 0,
            "failed": 0,
            "errors": [],
        }
        try:
            try:
                candidates, _warnings, _settings = self._load_engineer_mop_candidates(force=True)
                for candidate in candidates:
                    for attachment in candidate.get("attachments") or []:
                        if not isinstance(attachment, dict):
                            continue
                        stats["mop_checked"] += 1
                        try:
                            self._download_mop_attachment(attachment)
                        except Exception as exc:
                            stats["failed"] += 1
                            if len(stats["errors"]) < 10:
                                stats["errors"].append(f"MOP:{candidate.get('title') or candidate.get('record_id')}: {exc}")
            except Exception as exc:
                stats["failed"] += 1
                stats["errors"].append(f"MOP扫描失败: {exc}")
            try:
                people = self._load_signature_people(force=True)
                for person in people:
                    fields = person.get("raw_fields") if isinstance(person.get("raw_fields"), dict) else {}
                    for attachment in self._extract_signature_attachments(fields):
                        contextual = self._attachment_with_cache_context(
                            attachment,
                            category="signature",
                            app_token=SIGNATURE_APP_TOKEN,
                            table_id=SIGNATURE_TABLE_ID,
                            record_id=str(person.get("record_id") or ""),
                            latest_publish_time=str(person.get("latest_publish_time") or ""),
                        )
                        stats["signature_checked"] += 1
                        try:
                            self._download_mop_attachment(contextual)
                        except Exception as exc:
                            stats["failed"] += 1
                            if len(stats["errors"]) < 10:
                                stats["errors"].append(f"签名:{person.get('name') or person.get('record_id')}: {exc}")
            except Exception as exc:
                stats["failed"] += 1
                stats["errors"].append(f"签名扫描失败: {exc}")
        finally:
            stats["finished_at"] = dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            stats["status"] = "success" if not stats["failed"] else "partial" if (stats["mop_checked"] or stats["signature_checked"]) else "failed"
            if stats["failed"]:
                stats["next_retry_at"] = time.time() + 30 * 60
                stats["next_retry_at_text"] = dt.datetime.fromtimestamp(
                    float(stats["next_retry_at"])
                ).strftime("%Y-%m-%d %H:%M:%S")
            with suppress(Exception):
                marker_path = self._attachment_cache_daily_marker_path()
                marker_path.parent.mkdir(parents=True, exist_ok=True)
                marker_path.write_text(json.dumps(stats, ensure_ascii=False, indent=2), encoding="utf-8")
            with self._attachment_cache_refresh_lock:
                self._attachment_cache_refresh_running = False

    def preview_engineer_mop_attachment(
        self,
        *,
        scope: str = "ALL",
        mop_record_id: str,
        file_token: str = "",
        file_name: str = "",
        upload_id: str = "",
    ) -> dict[str, Any]:
        local_upload_id = self._upload_id_from_local_mop_ref(
            upload_id,
            allow_plain=True,
        ) or self._upload_id_from_local_mop_ref(mop_record_id, file_token)
        if local_upload_id:
            stored = self._state_store.get_engineer_mop_local_file(local_upload_id)
            if not stored:
                raise PortalError("本地上传的 MOP 文件不存在，请重新上传。")
            stored_scope = self._normalize_scope(stored.get("scope") or "ALL")
            request_scope = self._normalize_scope(scope)
            if (
                stored_scope
                and stored_scope != "ALL"
                and request_scope != "ALL"
                and stored_scope != request_scope
            ):
                raise PortalError("当前楼栋与本地 MOP 上传记录不匹配，请重新上传。")
            stored_source_record_id = str(stored.get("source_record_id") or "").strip()
            if stored_source_record_id:
                record = self._find_record_by_id(stored_source_record_id, WORK_TYPE_MAINTENANCE)
                display_fields = record.get("display_fields") if isinstance(record.get("display_fields"), dict) else {}
                if not self._scope_matches_buildings(request_scope, self._building_codes_from_value(display_fields.get("楼栋"))):
                    raise PortalError("当前账号无权预览该本地 MOP 文件。")
            source_path = Path(str(stored.get("local_file_path") or "")).resolve()
            data_root = Path(get_data_file_path("")).resolve()
            try:
                source_path.relative_to(data_root)
            except Exception as exc:
                raise PortalError("本地 MOP 文件路径无效，请重新上传。") from exc
            if not source_path.is_file():
                raise PortalError("本地 MOP 文件已被移动或删除，请重新上传。")
            content = source_path.read_bytes()
            resolved_name = str(file_name or stored.get("original_file_name") or source_path.name).strip()
            suffix = source_path.suffix.lower()
            if suffix in {".xlsx", ".xlsm"} or content[:2] == b"PK":
                parsed = self._parse_xlsx_preview(content)
            else:
                raise PortalError("本地 MOP 工作文件格式无效，请重新上传 xlsx/xlsm 文件。")
            memory_key = self._mop_fill_memory_key(
                mop_title=str(Path(resolved_name).stem or "本地上传 MOP"),
                mop_file_name=resolved_name,
            )
            fill_memory = self._state_store.get_mop_fill_memory(memory_key) if memory_key else None
            parsed.update(
                {
                    "mop_record_id": self._local_mop_record_id(local_upload_id),
                    "mop_title": str(Path(resolved_name).stem or "本地上传 MOP"),
                    "mop_file_name": resolved_name,
                    "mop_source": "local_upload",
                    "upload_id": local_upload_id,
                    "mop_fill_memory_key": memory_key,
                    "fill_memory": fill_memory,
                    "attachment": {
                        "name": resolved_name,
                        "upload_id": local_upload_id,
                        "file_token": self._local_mop_record_id(local_upload_id),
                        "source": "local_upload",
                        "local_upload": True,
                    },
                    "local_file": self._engineer_mop_local_file_meta(
                        path=str(source_path),
                        file_name=source_path.name,
                        size=int(stored.get("file_size") or source_path.stat().st_size),
                        upload_id=local_upload_id,
                    ),
                    "editable_summary": self._mop_editable_summary(parsed.get("sheets") or []),
                    "warnings": stored.get("warnings") or [],
                }
            )
            return parsed

        mop_candidates, warnings, settings = self._load_engineer_mop_candidates()
        if not settings.get("app_token") or not settings.get("table_id"):
            raise PortalError("未配置 MOP 多维表，无法预览。")
        record = next(
            (item for item in mop_candidates if str(item.get("record_id") or "") == str(mop_record_id or "")),
            None,
        )
        if not record:
            raise PortalError("未找到选择的 MOP 表格记录。")
        attachments = record.get("attachments") if isinstance(record.get("attachments"), list) else []
        attachment = None
        for item in attachments:
            if not isinstance(item, dict):
                continue
            if file_token and str(item.get("file_token") or "") == str(file_token):
                attachment = item
                break
            if file_token and str(item.get("url") or "") == str(file_token):
                attachment = item
                break
            if file_token and str(item.get("name") or "") == str(file_token):
                attachment = item
                break
        if attachment is None and attachments:
            attachment = attachments[0]
        if not isinstance(attachment, dict):
            raise PortalError("该 MOP 记录没有可预览的表格附件。")
        resolved_name = str(file_name or attachment.get("name") or "").strip()
        content, content_type = self._download_mop_attachment(attachment)
        local_file = self._save_mop_attachment_copy(
            content,
            scope=scope,
            mop_record_id=str(record.get("record_id") or mop_record_id or ""),
            file_name=resolved_name or str(attachment.get("name") or "MOP表格.xlsx"),
            content_type=content_type,
        )
        lower_name = resolved_name.lower()
        if lower_name.endswith(".csv") or "csv" in content_type:
            parsed = self._parse_csv_preview(content, file_name=resolved_name)
        elif lower_name.endswith(".xlsx") or content[:2] == b"PK":
            parsed = self._parse_xlsx_preview(content)
        else:
            raise PortalError("暂只支持预览 xlsx/csv 格式的 MOP 表格附件。")
        memory_key = self._mop_fill_memory_key(
            mop_title=str(record.get("title") or ""),
            mop_file_name=resolved_name,
        )
        fill_memory = self._state_store.get_mop_fill_memory(memory_key) if memory_key else None
        parsed.update(
            {
                "mop_record_id": record.get("record_id"),
                "mop_title": record.get("title"),
                "mop_file_name": resolved_name,
                "mop_fill_memory_key": memory_key,
                "fill_memory": fill_memory,
                "attachment": attachment,
                "local_file": local_file,
                "editable_summary": self._mop_editable_summary(parsed.get("sheets") or []),
                "warnings": warnings,
            }
        )
        return parsed

    def fill_engineer_mop_file(
        self,
        *,
        scope: str = "ALL",
        notice_key: str = "",
        operator_open_id: str = "",
        local_file_path: str = "",
        mop_record_id: str = "",
        mop_title: str = "",
        mop_file_name: str = "",
        sheet_name: str = "",
        fields: list[dict[str, Any]] | None = None,
        checkboxes: list[dict[str, Any]] | None = None,
        cell_edits: list[dict[str, Any]] | None = None,
        signatures: list[dict[str, Any]] | None = None,
    ) -> dict[str, Any]:
        try:
            from openpyxl import load_workbook
            from openpyxl.drawing.image import Image as ExcelImage
            from PIL import Image
        except Exception as exc:  # pragma: no cover - dependency bootstrap should provide these.
            raise PortalError("缺少 openpyxl/Pillow 依赖，无法生成已签名 MOP。") from exc

        data_root = Path(get_data_file_path("")).resolve()
        source_path = Path(str(local_file_path or "")).resolve()
        try:
            source_path.relative_to(data_root)
        except Exception as exc:
            raise PortalError("MOP 本地文件路径无效。") from exc
        if not source_path.is_file():
            raise PortalError("MOP 本地文件不存在，请重新查看表格后再生成。")
        if source_path.suffix.lower() not in {".xlsx", ".xlsm"}:
            raise PortalError("当前仅支持对 xlsx/xlsm MOP 文件插入签名。")

        fields = [item for item in (fields or []) if isinstance(item, dict)]
        checkboxes = [item for item in (checkboxes or []) if isinstance(item, dict)]
        cell_edits = [item for item in (cell_edits or []) if isinstance(item, dict)]
        signatures = [item for item in (signatures or []) if isinstance(item, dict)]
        if not signatures:
            raise PortalError("请选择至少一个维护实施人或维护审核人签名。")
        if operator_open_id:
            self._ensure_mop_staff_signature_usage_confirmed(
                signatures=signatures,
                scope=scope,
                notice_key=notice_key,
                operator_open_id=operator_open_id,
            )

        role_to_label = {
            "implementer": "维护实施人",
            "auditor": "维护审核人",
        }
        role_fields: dict[str, dict[str, Any]] = {}
        for field in fields:
            label = str(field.get("label") or "")
            for role, role_label in role_to_label.items():
                if role_label in label and role not in role_fields:
                    role_fields[role] = field
        missing_roles = sorted(
            {
                str(item.get("role") or "")
                for item in signatures
                if str(item.get("role") or "") in role_to_label
            }
            - set(role_fields)
        )
        if missing_roles:
            raise PortalError("当前 Sheet 未识别到对应签名位置，请切换到非封面填写页。")

        workbook = load_workbook(source_path)
        worksheet = workbook[sheet_name] if sheet_name and sheet_name in workbook.sheetnames else workbook.active
        protected_cells: set[tuple[int, int]] = set()
        for checkbox in checkboxes:
            row = self._mop_cell_int(checkbox.get("row"), -1) + 1
            col = self._mop_cell_int(checkbox.get("col"), -1) + 1
            if row > 0 and col > 0:
                protected_cells.add((row, col))
        for field in fields:
            row = self._mop_cell_int(field.get("row"), -1) + 1
            label_col = self._mop_cell_int(field.get("label_col"), -1) + 1
            value_col = self._mop_cell_int(
                field.get("value_col"),
                self._mop_cell_int(field.get("label_col"), -1),
            ) + 1
            if row > 0 and label_col > 0:
                protected_cells.add((row, label_col))
            if row > 0 and value_col > 0:
                protected_cells.add((row, value_col))
        inserted = 0
        updated_cells = 0
        temp_paths: list[str] = []
        try:
            for edit in cell_edits:
                row = self._mop_cell_int(edit.get("row"), -1) + 1
                col = self._mop_cell_int(edit.get("col"), -1) + 1
                if row <= 0 or col <= 0 or (row, col) in protected_cells:
                    continue
                worksheet.cell(row=row, column=col).value = str(edit.get("value") or "")
                updated_cells += 1
            for checkbox in checkboxes:
                state = str(
                    checkbox.get("selection")
                    or checkbox.get("selected_label")
                    or checkbox.get("selected_key")
                    or checkbox.get("state")
                    or ""
                ).strip()
                if not state:
                    continue
                row = self._mop_cell_int(checkbox.get("row"), -1) + 1
                col = self._mop_cell_int(checkbox.get("col"), -1) + 1
                if row <= 0 or col <= 0:
                    continue
                cell = worksheet.cell(row=row, column=col)
                cell.value = self._mop_choice_value(cell.value, state)
                updated_cells += 1
            for field in fields:
                fill_value = str(field.get("fill_value") or "").strip()
                if not fill_value:
                    continue
                label = str(field.get("label") or "").strip()
                if label in {"维护实施人", "维护审核人"}:
                    continue
                row = self._mop_cell_int(field.get("row"), -1) + 1
                value_col_zero = self._mop_cell_int(
                    field.get("value_col"),
                    self._mop_cell_int(field.get("label_col"), -1),
                )
                col = value_col_zero + 1
                if row <= 0 or col <= 0:
                    continue
                label_col = self._mop_cell_int(field.get("label_col"), -1) + 1
                cell = worksheet.cell(row=row, column=col)
                value_only = "维护完成时间" in label or "审核确认时间" in label
                cell.value = f"{label}：{fill_value}" if label and label_col == col and not value_only else fill_value
                updated_cells += 1
            for role in ("implementer", "auditor"):
                field = role_fields.get(role)
                if not field:
                    continue
                role_signatures = [
                    item for item in signatures
                    if str(item.get("role") or "") == role
                    and (
                        str(item.get("record_id") or "").strip()
                        or str(item.get("temp_id") or "").strip()
                    )
                ]
                if not role_signatures:
                    continue
                base_row = int(field.get("row") or 0) + 1
                label_col = self._mop_cell_int(field.get("label_col"), -1) + 1
                base_col = self._mop_signature_target_column(
                    worksheet,
                    row=base_row,
                    label_col=label_col,
                )
                max_signature_height = self._mop_signature_max_height_px(worksheet, row=base_row)
                role_images = []
                for signature in role_signatures:
                    source = str(signature.get("source") or "").strip()
                    temp_id = str(signature.get("temp_id") or "").strip()
                    record_id = str(signature.get("record_id") or "").strip()
                    if source == "temporary" or temp_id:
                        image_bytes, _content_type = self.temporary_signature_image_bytes(temp_id=temp_id)
                    elif source == "external":
                        image_bytes, _content_type = self.external_signature_image_bytes(record_id=record_id)
                    else:
                        image_bytes, _content_type = self.signature_image_bytes(record_id=record_id)
                    image = Image.open(io.BytesIO(image_bytes)).convert("RGBA")
                    max_width, max_height = 150, max_signature_height
                    ratio = min(max_width / max(1, image.width), max_height / max(1, image.height), 1.0)
                    if ratio < 1.0:
                        image = image.resize(
                            (max(1, int(image.width * ratio)), max(1, int(image.height * ratio)))
                        )
                    role_images.append(image)
                if not role_images:
                    continue
                gap = 8
                combined_width = max(1, sum(image.width for image in role_images) + gap * (len(role_images) - 1))
                combined_height = max(1, max(image.height for image in role_images))
                combined = Image.new("RGBA", (combined_width, combined_height), (255, 255, 255, 0))
                offset_x = 0
                for image in role_images:
                    combined.alpha_composite(image, (offset_x, max(0, (combined_height - image.height) // 2)))
                    offset_x += image.width + gap
                temp = tempfile.NamedTemporaryFile(delete=False, suffix=".png", prefix="clipflow_mop_signature_")
                temp.close()
                combined.save(temp.name, format="PNG")
                temp_paths.append(temp.name)
                excel_image = ExcelImage(temp.name)
                excel_image.width = combined.width
                excel_image.height = combined.height
                worksheet.add_image(excel_image, f"{self._column_label(base_col)}{base_row}")
                inserted += len(role_images)
            if not inserted:
                raise PortalError("没有可插入的签名。")
            scope_part = self._safe_mop_path_part(self._scope_label(self._normalize_scope(scope)) or scope, "ALL")
            date_part = dt.date.today().isoformat()
            stem = self._safe_mop_path_part(
                Path(source_path.name).stem or mop_title or mop_record_id or "MOP",
                "MOP",
            )
            output_dir = Path(get_data_file_path(os.path.join("engineer_mop_filled", date_part, scope_part)))
            output_dir.mkdir(parents=True, exist_ok=True)
            output_path = output_dir / (
                f"{stem}_已签名_{dt.datetime.now().strftime('%H%M%S')}_{uuid.uuid4().hex[:8]}"
                f"{source_path.suffix}"
            )
            workbook.save(output_path)
        finally:
            for path in temp_paths:
                with suppress(Exception):
                    os.remove(path)

        try:
            relative_path = str(output_path.resolve().relative_to(data_root))
        except Exception:
            relative_path = output_path.name
        return {
            "file_name": output_path.name,
            "path": str(output_path),
            "relative_path": relative_path,
            "inserted": inserted,
            "updated_cells": updated_cells,
            "saved_at": dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        }

    def _mop_signature_people_for_upload(
        self, signatures: list[dict[str, Any]]
    ) -> list[dict[str, Any]]:
        role_to_label = {
            "implementer": "维护实施人",
            "auditor": "维护审核人",
        }
        roles_by_record: dict[str, set[str]] = {}
        temporary_by_id: dict[str, set[str]] = {}
        external_by_record: dict[str, set[str]] = {}
        for item in signatures:
            if not isinstance(item, dict):
                continue
            role = str(item.get("role") or "").strip()
            record_id = str(item.get("record_id") or "").strip()
            temp_id = str(item.get("temp_id") or "").strip()
            source = str(item.get("source") or "").strip()
            if role not in role_to_label:
                continue
            if source == "temporary" or temp_id:
                if temp_id:
                    temporary_by_id.setdefault(temp_id, set()).add(role)
                continue
            if source == "external":
                if record_id:
                    external_by_record.setdefault(record_id, set()).add(role)
                continue
            if record_id:
                roles_by_record.setdefault(record_id, set()).add(role)
        present_roles = {
            role
            for roles in list(roles_by_record.values()) + list(temporary_by_id.values()) + list(external_by_record.values())
            for role in roles
        }
        missing = [label for role, label in role_to_label.items() if role not in present_roles]
        if missing:
            raise PortalError(f"上传已签名 MOP 前请先选择可用签名：{'、'.join(missing)}。")

        people = self._load_signature_people(force=False)
        people_by_id = {
            str(item.get("record_id") or "").strip(): item
            for item in people
            if str(item.get("record_id") or "").strip()
        }
        missing_records = [record_id for record_id in roles_by_record if record_id not in people_by_id]
        if missing_records:
            people = self._load_signature_people(force=True)
            people_by_id = {
                str(item.get("record_id") or "").strip(): item
                for item in people
                if str(item.get("record_id") or "").strip()
            }
        entries: list[dict[str, Any]] = []
        for temp_id, roles in temporary_by_id.items():
            session = self._state_store.get_mop_temporary_signature_session(temp_id=temp_id)
            if not session:
                raise PortalError("临时签名记录不存在，请重新发送签名链接。")
            if not str(session.get("signature_file_token") or "").strip():
                raise PortalError(f"{session.get('display_name') or '临时人员'} 暂无可用签名。")
            # 临时人员没有 openid，只参与 MOP 签名完整性校验和插图，不发送通知。
            continue
        for record_id in external_by_record:
            try:
                self.external_signature_image_bytes(record_id=record_id)
            except Exception as exc:
                raise PortalError("其他人员签名不可用，请重新选择或重新签名。") from exc
            # 其他人员没有可通知 openid，只参与 MOP 签名完整性校验和插图。
            continue
        for record_id, roles in roles_by_record.items():
            person = people_by_id.get(record_id)
            if not person:
                raise PortalError("签名人员记录不存在，请刷新后重试。")
            if not person.get("has_signature"):
                raise PortalError(f"{person.get('name') or record_id} 暂无可用签名。")
            role_labels = [role_to_label[role] for role in ("implementer", "auditor") if role in roles]
            entries.append(
                {
                    "record_id": record_id,
                    "name": str(person.get("name") or record_id),
                    "open_id": str(person.get("open_id") or "").strip(),
                    "roles": [role for role in ("implementer", "auditor") if role in roles],
                    "role_labels": role_labels,
                }
            )
        return entries

    def _ensure_mop_staff_signature_usage_confirmed(
        self,
        *,
        signatures: list[dict[str, Any]],
        scope: str,
        notice_key: str,
        operator_open_id: str,
    ) -> None:
        record_ids = {
            str(item.get("record_id") or "").strip()
            for item in signatures
            if isinstance(item, dict)
            and str(item.get("source") or "staff").strip() in {"", "staff"}
            and str(item.get("record_id") or "").strip()
        }
        if not record_ids:
            return
        operator_open_id = str(operator_open_id or "").strip()
        if not operator_open_id:
            raise PortalError("缺少当前登录人 openid，无法校验他人签名确认。")
        people = self._load_signature_people(force=False)
        people_by_id = {
            str(item.get("record_id") or "").strip(): item
            for item in people
            if str(item.get("record_id") or "").strip()
        }
        if any(record_id not in people_by_id for record_id in record_ids):
            people = self._load_signature_people(force=True)
            people_by_id = {
                str(item.get("record_id") or "").strip(): item
                for item in people
                if str(item.get("record_id") or "").strip()
            }
        missing_confirmations: list[str] = []
        for record_id in sorted(record_ids):
            person = people_by_id.get(record_id)
            if not person:
                raise PortalError("签名人员记录不存在，请刷新后重试。")
            person_open_id = str(person.get("open_id") or "").strip()
            if person_open_id == operator_open_id:
                continue
            if not person_open_id:
                missing_confirmations.append(f"{person.get('name') or record_id}（缺少openid）")
                continue
            if self._state_store.has_confirmed_mop_signature_usage(
                scope=scope,
                notice_key=notice_key,
                signer_record_id=record_id,
                requested_by_openid=operator_open_id,
            ):
                continue
            missing_confirmations.append(str(person.get("name") or record_id))
        if missing_confirmations:
            raise PortalError(
                "使用他人已保存签名前，请先发送确认并等待确认："
                + "、".join(missing_confirmations)
            )

    def _send_mop_signature_upload_notifications(
        self,
        *,
        people: list[dict[str, Any]],
        notice_title: str,
        building: str = "",
        maintenance_cycle: str = "",
        file_name: str = "",
        operator_name: str = "",
        operator_open_id: str = "",
    ) -> tuple[list[dict[str, Any]], str]:
        grouped: dict[str, dict[str, Any]] = {}
        results: list[dict[str, Any]] = []
        for person in people:
            open_id = str(person.get("open_id") or "").strip()
            key = open_id or f"missing:{person.get('record_id') or person.get('name')}"
            target = grouped.setdefault(
                key,
                {
                    "open_id": open_id,
                    "name": str(person.get("name") or ""),
                    "role_labels": [],
                },
            )
            for label in person.get("role_labels") or []:
                if label not in target["role_labels"]:
                    target["role_labels"].append(label)

        for item in grouped.values():
            open_id = str(item.get("open_id") or "").strip()
            role_text = "、".join(item.get("role_labels") or []) or "签名人员"
            if not open_id:
                results.append(
                    {
                        "open_id": "",
                        "name": item.get("name") or "",
                        "roles": role_text,
                        "ok": False,
                        "message": "该人员缺少 openid，无法发送个人消息。",
                    }
                )
                continue
            text_lines = [
                "【MOP签名确认】已上传维护保养单",
                "",
                f"签名角色：{role_text}",
                f"维护通告：{notice_title or '未命名维保通告'}",
            ]
            if building:
                text_lines.append(f"楼栋：{building}")
            if maintenance_cycle:
                text_lines.append(f"维护周期：{maintenance_cycle}")
            if file_name:
                text_lines.append(f"MOP文件：{file_name}")
            text_lines.extend(
                [
                    f"操作人：{operator_name or '未知'}"
                    + (f"（{operator_open_id}）" if operator_open_id else ""),
                    f"时间：{dt.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
                ]
            )
            if external_mock_enabled():
                ok, message, send_results = True, "mock external send skipped", [
                    {"open_id": open_id, "ok": True, "message": "mock external send skipped"}
                ]
            else:
                ok, message, send_results = send_text_to_open_ids("\n".join(text_lines), [open_id])
            send_result = (send_results or [{}])[0] if send_results else {}
            results.append(
                {
                    "open_id": open_id,
                    "name": item.get("name") or "",
                    "roles": role_text,
                    "ok": bool(ok and send_result.get("ok", ok)),
                    "message": str(send_result.get("message") or message or ""),
                }
            )
        failed = [item for item in results if not item.get("ok")]
        warning = ""
        if failed:
            warning = "签名人员通知失败：" + "；".join(
                f"{item.get('name') or item.get('open_id') or '未知人员'}({item.get('roles')})：{item.get('message') or '发送失败'}"
                for item in failed
            )
        return results, warning

    @staticmethod
    def _mop_confirm_field_truthy(value: Any) -> bool:
        if isinstance(value, bool):
            return value
        if isinstance(value, (int, float)):
            return bool(value)
        if isinstance(value, str):
            return value.strip().lower() in {"1", "true", "yes", "y", "是", "已确认", "勾选"}
        if isinstance(value, list):
            return any(MaintenancePortalService._mop_confirm_field_truthy(item) for item in value)
        if isinstance(value, dict):
            for key in ("checked", "value", "text", "name"):
                if key in value and MaintenancePortalService._mop_confirm_field_truthy(value.get(key)):
                    return True
        return False

    @staticmethod
    def _mop_attachment_has_token(value: Any, expected_file_token: str) -> bool:
        expected_file_token = str(expected_file_token or "").strip()
        if not expected_file_token:
            return False
        if isinstance(value, dict):
            token = str(
                value.get("file_token")
                or value.get("token")
                or value.get("id")
                or ""
            ).strip()
            if token == expected_file_token:
                return True
            return any(
                MaintenancePortalService._mop_attachment_has_token(item, expected_file_token)
                for item in value.values()
            )
        if isinstance(value, list):
            return any(
                MaintenancePortalService._mop_attachment_has_token(item, expected_file_token)
                for item in value
            )
        return str(value or "").strip() == expected_file_token

    def _verify_mop_source_record_update(
        self,
        *,
        source_record_id: str,
        file_token: str,
    ) -> dict[str, Any]:
        source_record_id = str(source_record_id or "").strip()
        file_token = str(file_token or "").strip()
        if not source_record_id or not file_token:
            raise PortalError("MOP 上传后校验缺少源记录 ID 或 file_token。")
        guard = external_real_write_guard()
        if guard.get("mock_external"):
            return {
                "ok": True,
                "mock_external": True,
                "message": "mock 外部写入，已跳过源表读回校验。",
                "checked_fields": [
                    MOP_SIGNED_ATTACHMENT_FIELD,
                    MOP_ENGINEER_CONFIRM_FIELD,
                    MOP_SUPERVISOR_CONFIRM_FIELD,
                ],
            }
        payload = self._request_json(
            f"records/{source_record_id}",
            app_token=DEFAULT_APP_TOKEN,
            table_id=DEFAULT_TABLE_ID,
        )
        data = payload.get("data") if isinstance(payload.get("data"), dict) else {}
        record = data.get("record") if isinstance(data.get("record"), dict) else {}
        fields = record.get("fields") if isinstance(record.get("fields"), dict) else {}
        missing: list[str] = []
        if not self._mop_attachment_has_token(
            fields.get(MOP_SIGNED_ATTACHMENT_FIELD),
            file_token,
        ):
            missing.append(MOP_SIGNED_ATTACHMENT_FIELD)
        if not self._mop_confirm_field_truthy(fields.get(MOP_ENGINEER_CONFIRM_FIELD)):
            missing.append(MOP_ENGINEER_CONFIRM_FIELD)
        if not self._mop_confirm_field_truthy(fields.get(MOP_SUPERVISOR_CONFIRM_FIELD)):
            missing.append(MOP_SUPERVISOR_CONFIRM_FIELD)
        if missing:
            raise PortalError(
                "MOP 上传后源表校验失败，以下字段未确认写入："
                + "、".join(missing)
            )
        return {
            "ok": True,
            "mock_external": False,
            "message": "源表附件和确认字段已校验。",
            "source_record_id": source_record_id,
            "file_token": file_token,
            "checked_fields": [
                MOP_SIGNED_ATTACHMENT_FIELD,
                MOP_ENGINEER_CONFIRM_FIELD,
                MOP_SUPERVISOR_CONFIRM_FIELD,
            ],
        }

    def upload_signed_engineer_mop_file(
        self,
        *,
        scope: str = "ALL",
        source_record_id: str = "",
        notice_title: str = "",
        notice_key: str = "",
        operator_open_id: str = "",
        operator_name: str = "",
        local_file_path: str = "",
        mop_record_id: str = "",
        mop_title: str = "",
        mop_file_name: str = "",
        sheet_name: str = "",
        fields: list[dict[str, Any]] | None = None,
        checkboxes: list[dict[str, Any]] | None = None,
        cell_edits: list[dict[str, Any]] | None = None,
        signatures: list[dict[str, Any]] | None = None,
    ) -> dict[str, Any]:
        self.ensure_snapshot_loaded()
        source_record_id = str(source_record_id or "").strip()
        if not source_record_id:
            raise PortalError("缺少维保源表记录 ID，无法上传已签名 MOP。")
        record = self._find_record_by_id(source_record_id, WORK_TYPE_MAINTENANCE)
        display_fields = record.get("display_fields") or {}
        building_codes = self._building_codes_from_value(display_fields.get("楼栋"))
        normalized_scope = self._normalize_scope(scope)
        if not self._scope_matches_buildings(normalized_scope, building_codes):
            raise PortalError("当前账号无权上传该楼栋的 MOP。")

        signature_items = [item for item in (signatures or []) if isinstance(item, dict)]
        self._ensure_mop_staff_signature_usage_confirmed(
            signatures=signature_items,
            scope=scope,
            notice_key=notice_key,
            operator_open_id=operator_open_id,
        )
        signature_people = self._mop_signature_people_for_upload(signature_items)
        filled = self.fill_engineer_mop_file(
            scope=scope,
            local_file_path=local_file_path,
            mop_record_id=mop_record_id,
            mop_title=mop_title,
            mop_file_name=mop_file_name,
            sheet_name=sheet_name,
            fields=fields or [],
            checkboxes=checkboxes or [],
            cell_edits=cell_edits or [],
            signatures=signatures or [],
        )
        file_token = self._upload_bitable_file(
            file_path=str(filled.get("path") or ""),
            file_name=str(filled.get("file_name") or ""),
            app_token=DEFAULT_APP_TOKEN,
        )
        update_fields = {
            MOP_SIGNED_ATTACHMENT_FIELD: [{"file_token": file_token}],
            MOP_ENGINEER_CONFIRM_FIELD: True,
            MOP_SUPERVISOR_CONFIRM_FIELD: True,
        }
        self._patch_record_fields(
            app_token=DEFAULT_APP_TOKEN,
            table_id=DEFAULT_TABLE_ID,
            record_id=source_record_id,
            fields=update_fields,
        )
        verification = self._verify_mop_source_record_update(
            source_record_id=source_record_id,
            file_token=file_token,
        )
        display_fields.update(update_fields)
        for source_record in self._records:
            if str(source_record.get("record_id") or "") == source_record_id:
                source_fields = source_record.get("display_fields")
                if not isinstance(source_fields, dict):
                    source_fields = {}
                    source_record["display_fields"] = source_fields
                source_fields.update(update_fields)
        with suppress(Exception):
            self._state_store.patch_active_source_record_fields(
                source_record_id=source_record_id,
                work_type=WORK_TYPE_MAINTENANCE,
                fields=update_fields,
            )
        memory_warning = ""
        memory_key = self._mop_fill_memory_key(
            mop_title=mop_title,
            mop_file_name=mop_file_name or Path(str(local_file_path or "")).name,
        )
        if memory_key:
            try:
                self._state_store.upsert_mop_fill_memory(
                    {
                        "memory_key": memory_key,
                        "mop_title": mop_title,
                        "mop_file_name": mop_file_name or Path(str(local_file_path or "")).name,
                        "sheet_name": sheet_name,
                        "updated_by": operator_open_id or operator_name,
                        "payload": self._build_mop_fill_memory_payload(
                            scope=scope,
                            source_record_id=source_record_id,
                            notice_title=notice_title,
                            notice_key=notice_key,
                            sheet_name=sheet_name,
                            fields=fields or [],
                            checkboxes=checkboxes or [],
                            cell_edits=cell_edits or [],
                        ),
                    }
                )
            except Exception as exc:
                memory_warning = f"MOP填写记忆保存失败：{exc}"
        self._touch_state_cache_version()
        title = (
            str(notice_title or "").strip()
            or self._maintenance_title(record)
            or str(display_fields.get("维护总项") or "")
        )
        building = self._building_label_from_codes(building_codes)
        maintenance_cycle = str(display_fields.get("维护周期") or "").strip()
        notification_results, notification_warning = self._send_mop_signature_upload_notifications(
            people=signature_people,
            notice_title=title,
            building=building,
            maintenance_cycle=maintenance_cycle,
            file_name=str(filled.get("file_name") or ""),
            operator_name=operator_name,
            operator_open_id=operator_open_id,
        )
        return {
            "source_record_id": source_record_id,
            "notice_key": str(notice_key or ""),
            "notice_title": title,
            "building": building,
            "maintenance_cycle": maintenance_cycle,
            "file_token": file_token,
            "filled_file": filled,
            "updated_fields": list(update_fields.keys()),
            "verification": verification,
            "notification_results": notification_results,
            "notification_warning": notification_warning,
            "memory_key": memory_key,
            "memory_warning": memory_warning,
            "uploaded_at": dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        }

    def reset_engineer_mop_file(
        self,
        *,
        scope: str = "ALL",
        filled_file_path: str = "",
        mop_record_id: str = "",
        file_token: str = "",
        file_name: str = "",
    ) -> dict[str, Any]:
        data_root = Path(get_data_file_path("")).resolve()
        deleted = False
        filled_text = str(filled_file_path or "").strip()
        if filled_text:
            filled_path = Path(filled_text).resolve()
            try:
                relative = filled_path.relative_to(data_root)
                parts = {part.lower() for part in relative.parts}
                if "engineer_mop_filled" in parts and filled_path.is_file():
                    filled_path.unlink()
                    deleted = True
            except Exception as exc:
                raise PortalError("已签名 MOP 文件路径无效，已停止重新签名。") from exc
        preview = self.preview_engineer_mop_attachment(
            scope=scope,
            mop_record_id=mop_record_id,
            file_token=file_token,
            file_name=file_name,
        )
        preview["reset"] = {
            "deleted_old_file": deleted,
            "reset_at": dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        }
        return preview

    def _find_record_by_id(self, record_id: str, work_type: str = WORK_TYPE_MAINTENANCE) -> dict[str, Any]:
        if work_type == WORK_TYPE_CHANGE:
            records = self._change_records
        elif work_type == WORK_TYPE_REPAIR:
            records = self._repair_records
        else:
            records = self._records
        for record in records:
            if record["record_id"] == record_id:
                return record
        snapshot_records = self._source_snapshot_records("ALL")
        for record in snapshot_records or []:
            if (
                self._record_work_type(record) == work_type
                and str(record.get("record_id") or "") == record_id
            ):
                return record
        raise PortalError(f"未找到记录: {record_id}")

    def set_notice_work_type_override(
        self,
        *,
        record_id: str,
        source_work_type: str = WORK_TYPE_MAINTENANCE,
        target_work_type: str = WORK_TYPE_CHANGE,
        scope: str = "ALL",
        updated_by: str = "",
    ) -> dict[str, Any]:
        self.ensure_snapshot_loaded()
        record_id = str(record_id or "").strip()
        source_work_type = str(source_work_type or WORK_TYPE_MAINTENANCE).strip()
        target_work_type = str(target_work_type or WORK_TYPE_CHANGE).strip()
        if source_work_type != WORK_TYPE_MAINTENANCE:
            raise PortalError("当前仅支持将维保源记录转为变更。")
        if target_work_type not in {WORK_TYPE_CHANGE, WORK_TYPE_MAINTENANCE}:
            raise PortalError("转换目标只支持变更或维保。")
        if not record_id:
            raise PortalError("缺少源记录 ID。")
        record = self._find_record_by_id(record_id, WORK_TYPE_MAINTENANCE)
        fields = record.get("display_fields") or {}
        building_codes = self._building_codes_from_value(fields.get("楼栋"))
        normalized_scope = self._normalize_scope(scope)
        if not self._scope_matches_buildings(normalized_scope, building_codes):
            raise PortalError("当前账号无权转换该楼栋通告。")
        title = self._maintenance_title(record)
        normalized_title = self._work_type_override_title_key(
            title, WORK_TYPE_MAINTENANCE
        )
        if not normalized_title:
            raise PortalError("该维保记录缺少可用于长期转换的标题。")
        if target_work_type == WORK_TYPE_MAINTENANCE:
            changed = self._state_store.disable_work_type_override(
                source_work_type=WORK_TYPE_MAINTENANCE,
                normalized_title=normalized_title,
                updated_by=updated_by,
            )
            self._touch_state_cache_version()
            return {
                "record_id": record_id,
                "source_work_type": WORK_TYPE_MAINTENANCE,
                "target_work_type": WORK_TYPE_MAINTENANCE,
                "title": title,
                "normalized_title": normalized_title,
                "changed": changed,
            }
        override = self._state_store.upsert_work_type_override(
            source_work_type=WORK_TYPE_MAINTENANCE,
            normalized_title=normalized_title,
            target_work_type=WORK_TYPE_CHANGE,
            title=title,
            payload={
                "record_id": record_id,
                "building": self._building_label_from_codes(building_codes),
                "building_codes": building_codes,
                "maintenance_total": str(fields.get("维护总项") or ""),
                "maintenance_cycle": str(fields.get("维护周期") or ""),
            },
            updated_by=updated_by,
        )
        self._touch_state_cache_version()
        return {
            "record_id": record_id,
            "source_work_type": WORK_TYPE_MAINTENANCE,
            "target_work_type": WORK_TYPE_CHANGE,
            "title": title,
            "normalized_title": normalized_title,
            "changed": True,
            "override": override,
        }

    def generate_templates(self, drafts: list[dict[str, Any]]) -> list[dict[str, Any]]:
        self.ensure_snapshot_loaded()
        generated: list[dict[str, Any]] = []
        for draft in drafts:
            record_id = str(draft.get("record_id") or "").strip()
            if not record_id:
                raise PortalError("存在缺少 record_id 的待生成记录。")
            record = self._find_record_by_id(record_id)
            fields = record["display_fields"]
            building = str(fields.get("楼栋") or "").strip()
            maintenance_total = str(fields.get("维护总项") or "").strip()
            maintenance_cycle = str(fields.get("维护周期") or "").strip()
            start_time = self._format_input_datetime(draft.get("start_time"))
            end_time = self._format_input_datetime(draft.get("end_time"))
            location = str(draft.get("location") or "").strip()
            content = str(draft.get("content") or "").strip()
            reason = str(draft.get("reason") or "").strip()
            impact = (
                str(draft.get("impact") or "").strip() or DEFAULT_IMPACT_TEXT
            )
            self._remember_draft_fields(
                building=building,
                maintenance_total=maintenance_total,
                location=location,
                content=content,
                reason=reason,
                impact=impact,
                maintenance_cycle=maintenance_cycle,
            )
            text = (
                self.build_notice_text(
                    status="开始",
                    title=f"EA118机房{building}{maintenance_total}",
                    start_time=start_time,
                    end_time=end_time,
                    location=location,
                    content=content,
                    reason=reason,
                    impact=impact,
                    progress=(
                        str(draft.get("progress") or "").strip()
                        or DEFAULT_PROGRESS_TEXT
                    ),
                )
            )
            generated.append(
                {
                    "record_id": record_id,
                    "title": f"EA118机房{building}{maintenance_total}",
                    "building": building,
                    "target_building": (
                        f"{self._building_code_from_value(building)}楼"
                        if self._building_code_from_value(building)
                        else ""
                    ),
                    "text": text,
                }
            )
        return generated

    @staticmethod
    def _format_notice_time_range(start_time: Any, end_time: Any) -> str:
        start_text = MaintenancePortalService._format_input_datetime(start_time)
        end_text = MaintenancePortalService._format_input_datetime(end_time)
        if start_text and end_text:
            return f"{start_text}{NOTICE_TIME_SEPARATOR}{end_text}"
        return start_text or end_text

    @staticmethod
    def _render_notice_text(
        *,
        work_type: str,
        status: str,
        values: dict[str, Any],
        heading_override: str = "",
    ) -> str:
        template = NOTICE_TEXT_TEMPLATES.get(
            str(work_type or "").strip(),
            NOTICE_TEXT_TEMPLATES[WORK_TYPE_MAINTENANCE],
        )
        heading = str(heading_override or template["heading"]).strip()
        lines = [f"【{heading}】状态：{str(status or '').strip()}"]
        for label, key in template["fields"]:
            value = values.get(key, "")
            lines.append(f"【{label}】{str(value or '').strip()}")
        return "\n".join(lines)

    @staticmethod
    def build_notice_text(
        *,
        status: str,
        title: str,
        start_time: str,
        end_time: str,
        location: str,
        content: str,
        reason: str,
        impact: str,
        progress: str,
    ) -> str:
        return MaintenancePortalService._render_notice_text(
            work_type=WORK_TYPE_MAINTENANCE,
            status=status,
            values={
                "title": title,
                "time_range": MaintenancePortalService._format_notice_time_range(
                    start_time, end_time
                ),
                "location": location,
                "content": content,
                "reason": reason,
                "impact": str(impact or "").strip() or DEFAULT_IMPACT_TEXT,
                "progress": str(progress or "").strip() or DEFAULT_PROGRESS_TEXT,
            },
        )

    def _base_job(self, request_payload: dict[str, Any]) -> dict[str, Any]:
        now = dt.datetime.now().strftime("%Y-%m-%d %H:%M")
        now_ts = time.time()
        return {
            "job_id": uuid.uuid4().hex,
            "phase": "accepted",
            "created_at": now,
            "updated_at": now,
            "accepted_at": now_ts,
            "message_started_at": 0.0,
            "message_finished_at": 0.0,
            "upload_queued_at": 0.0,
            "upload_started_at": 0.0,
            "upload_finished_at": 0.0,
            "upload_release_at": 0.0,
            "request": copy.deepcopy(request_payload),
            "operation_id": str(request_payload.get("operation_id") or "").strip(),
            "target_key": self._action_target_key(request_payload),
            "depends_on_job_id": "",
            "depends_on_phase": "",
            "message_sent": False,
            "message_signature": "",
            "message_queue_position": 0,
            "message_queue_size": 0,
            "qt_phase": "",
            "qt_queue_position": 0,
            "qt_queue_size": 0,
            "queue_position": 0,
            "queue_size": 0,
            "upload_queue_position": 0,
            "upload_queue_size": 0,
            "record_id": "",
            "active_item_id": str(request_payload.get("active_item_id") or ""),
            "error": "",
            "error_category": "",
            "error_retryable": False,
            "upload_message": "",
            "prepared": {},
        }

    @staticmethod
    def classify_job_error(message: str) -> dict[str, Any]:
        text = str(message or "").strip()
        lower = text.lower()
        if not text:
            return {"error_category": "", "error_retryable": False}
        if any(key in lower for key in ("timeout", "timed out", "超时", "read timed")):
            return {"error_category": "network_timeout", "error_retryable": True}
        if any(key in lower for key in ("429", "rate limit", "too many requests", "限流", "频率")):
            return {"error_category": "remote_rate_limited", "error_retryable": True}
        if any(key in lower for key in ("connection", "network", "连接", "网络", "10053", "10054")):
            return {"error_category": "network_error", "error_retryable": True}
        if any(key in lower for key in ("重启", "restart")):
            return {"error_category": "process_restart", "error_retryable": False}
        if any(key in lower for key in ("token", "tenant_access_token", "access_token", "授权", "登录态")):
            return {"error_category": "token_or_auth", "error_retryable": False}
        if any(key in lower for key in ("permission", "forbidden", "无权限", "权限不足", "403")):
            return {"error_category": "permission_denied", "error_retryable": False}
        if any(key in lower for key in ("field", "字段", "维保周期", "参数", "invalid")):
            return {"error_category": "field_or_payload", "error_retryable": False}
        if any(key in lower for key in ("record not found", "记录不存在", "远端缺失", "1254002")):
            return {"error_category": "remote_record_missing", "error_retryable": False}
        if any(key in lower for key in ("qt", "主窗口", "主界面", "回调")):
            return {"error_category": "qt_bridge", "error_retryable": True}
        return {"error_category": "unknown", "error_retryable": True}

    @classmethod
    def _manual_payload_notice_work_type(cls, request_payload: dict[str, Any], fallback: str) -> str:
        work_type = str(fallback or WORK_TYPE_MAINTENANCE).strip() or WORK_TYPE_MAINTENANCE
        requested_type = cls._requested_manual_work_type(request_payload)
        if requested_type:
            return requested_type
        return cls._manual_payload_notice_work_type_inferred(request_payload, work_type)

    @classmethod
    def _requested_manual_work_type(cls, request_payload: dict[str, Any]) -> str:
        payload = request_payload if isinstance(request_payload, dict) else {}
        work_type = str(payload.get("work_type") or payload.get("lan_work_type") or "").strip()
        valid_types = {
            WORK_TYPE_MAINTENANCE,
            WORK_TYPE_CHANGE,
            WORK_TYPE_REPAIR,
            WORK_TYPE_POWER,
            WORK_TYPE_POLLING,
            WORK_TYPE_ADJUST,
        }
        return work_type if work_type in valid_types else ""

    @classmethod
    def _manual_payload_notice_work_type_inferred(
        cls, request_payload: dict[str, Any], fallback: str
    ) -> str:
        work_type = str(fallback or WORK_TYPE_MAINTENANCE).strip() or WORK_TYPE_MAINTENANCE
        if work_type not in NOTICE_TEXT_TEMPLATES:
            work_type = WORK_TYPE_MAINTENANCE
        if not cls._truthy_flag((request_payload or {}).get("manual")):
            return work_type
        if work_type != WORK_TYPE_MAINTENANCE:
            return work_type
        head_text = "\n".join(
            str((request_payload or {}).get(name) or "").strip()
            for name in ("notice_type", "title", "content")
            if str((request_payload or {}).get(name) or "").strip()
        )
        if re.search(r"设备检修|检修通告", head_text):
            return WORK_TYPE_REPAIR
        if re.search(r"上电通告|上下电通告|下电通告", head_text):
            return WORK_TYPE_POWER
        if re.search(r"设备轮巡|轮巡通告", head_text):
            return WORK_TYPE_POLLING
        if re.search(r"设备调整|调整通告", head_text):
            return WORK_TYPE_ADJUST
        if re.search(r"变更通告", head_text):
            return WORK_TYPE_CHANGE
        return work_type

    @classmethod
    def _manual_payload_title_text(cls, request_payload: dict[str, Any]) -> str:
        payload = request_payload if isinstance(request_payload, dict) else {}
        direct_title = str(
            payload.get("title")
            or payload.get("name")
            or payload.get("notice_title")
            or ""
        ).strip()
        if direct_title:
            return direct_title
        text = str(payload.get("text") or payload.get("content") or "").strip()
        if text:
            sections = cls._parse_notice_sections(text)
            title = cls._notice_section_value(
                sections,
                ["名称", "标题", "维修名称", "通告名称"],
            )
            if title:
                return title
        return ""

    @staticmethod
    def _semantic_notice_text(value: Any) -> str:
        text = str(value or "").strip()
        text = text.replace("：", ":").replace("；", ";").replace("，", ",")
        text = re.sub(r"\s+", "", text)
        return text.lower()

    @classmethod
    def _semantic_notice_title(cls, request_payload: dict[str, Any]) -> str:
        title = cls._manual_payload_title_text(request_payload)
        title = str(title or "").strip()
        if not title:
            return ""
        title = re.sub(r"^【[^】]+】\s*状态[:：]?\s*(开始|更新|结束)?", "", title).strip()
        title = re.sub(
            r"^EA118(?:机房)?(?:[-－—]?)?(?:110KV阿里中天变|110站|[ABCDEH]楼|[ABCDEH]栋)",
            "",
            title,
            flags=re.IGNORECASE,
        ).strip()
        title = re.sub(r"(?:通告)$", "", title).strip()
        return cls._semantic_notice_text(title)

    @classmethod
    def _semantic_notice_time(cls, value: Any) -> str:
        return cls._semantic_notice_text(cls._format_input_datetime(value) or value)

    @classmethod
    def _semantic_notice_building_key(cls, request_payload: dict[str, Any]) -> str:
        payload = request_payload if isinstance(request_payload, dict) else {}
        codes = [
            str(code or "").strip().upper()
            for code in (payload.get("building_codes") or [])
            if str(code or "").strip()
        ]
        if not codes:
            for value in (
                payload.get("building"),
                payload.get("target_building"),
                payload.get("scope"),
                payload.get("location"),
            ):
                codes.extend(cls._building_codes_from_value(value))
                if codes:
                    break
        if not codes:
            codes.extend(cls._building_codes_from_value(cls._manual_payload_title_text(payload)))
        codes = [code for code in BUILDING_SCOPE_CODES if code in dict.fromkeys(codes)]
        return ",".join(codes)

    @classmethod
    def _manual_start_semantic_key(cls, request_payload: dict[str, Any], work_type: str = "") -> str:
        payload = request_payload if isinstance(request_payload, dict) else {}
        work_type = str(work_type or payload.get("work_type") or payload.get("lan_work_type") or "").strip()
        if not work_type:
            work_type = cls._manual_payload_notice_work_type(
                payload,
                cls._request_work_type_fallback(payload),
            )
        title_key = cls._semantic_notice_title(payload)
        if not title_key:
            return ""
        start_value = (
            payload.get("start_time")
            or payload.get("expected_time")
            or payload.get("fault_time")
            or ""
        )
        end_value = (
            payload.get("end_time")
            or payload.get("fault_time")
            or payload.get("expected_time")
            or ""
        )
        identity_parts = {
            "work_type": work_type,
            "notice_type": str(payload.get("notice_type") or "").strip(),
            "building": cls._semantic_notice_building_key(payload),
            "title": title_key,
            "start_time": cls._semantic_notice_time(start_value),
            "end_time": cls._semantic_notice_time(end_value),
            "location": cls._semantic_notice_text(payload.get("location")),
            "content": cls._semantic_notice_text(payload.get("content")),
            "reason": cls._semantic_notice_text(payload.get("reason")),
            "impact": cls._semantic_notice_text(payload.get("impact")),
            "level": cls._semantic_notice_text(payload.get("level")),
            "specialty": cls._semantic_notice_text(payload.get("specialty")),
            "maintenance_cycle": cls._semantic_notice_text(payload.get("maintenance_cycle")),
            "device": cls._semantic_notice_text(payload.get("device")),
            "cabinet": cls._semantic_notice_text(payload.get("cabinet")),
            "quantity": cls._semantic_notice_text(payload.get("quantity")),
            "repair_device": cls._semantic_notice_text(payload.get("repair_device")),
            "repair_fault": cls._semantic_notice_text(payload.get("repair_fault")),
            "fault_type": cls._semantic_notice_text(payload.get("fault_type")),
            "repair_mode": cls._semantic_notice_text(payload.get("repair_mode")),
            "discovery": cls._semantic_notice_text(payload.get("discovery")),
            "symptom": cls._semantic_notice_text(payload.get("symptom")),
            "solution": cls._semantic_notice_text(payload.get("solution")),
            "spare_parts": cls._semantic_notice_text(payload.get("spare_parts")),
        }
        seed = json.dumps(identity_parts, ensure_ascii=False, sort_keys=True)
        return hashlib.sha256(seed.encode("utf-8", errors="ignore")).hexdigest()

    @classmethod
    def _manual_start_target_key(cls, request_payload: dict[str, Any], work_type: str = "") -> str:
        semantic_key = cls._manual_start_semantic_key(request_payload, work_type)
        if not semantic_key:
            return ""
        work_type = str(work_type or (request_payload or {}).get("work_type") or "").strip()
        if not work_type:
            work_type = cls._manual_payload_notice_work_type(
                request_payload,
                cls._request_work_type_fallback(request_payload),
            )
        return f"{work_type}:manual-start:{semantic_key}"

    @classmethod
    def _manual_notice_type_conflict(
        cls, request_payload: dict[str, Any]
    ) -> dict[str, str] | None:
        payload = request_payload if isinstance(request_payload, dict) else {}
        if not cls._truthy_flag(payload.get("manual")):
            return None
        expected_type = (
            cls._requested_manual_work_type(payload)
            or cls._manual_payload_notice_work_type_inferred(
                payload,
                cls._request_work_type_fallback(payload),
            )
        )
        title = re.sub(r"\s+", "", cls._manual_payload_title_text(payload))
        if not title:
            return None
        for work_type, patterns in NOTICE_TYPE_KEYWORD_RULES.items():
            if work_type == expected_type:
                continue
            for pattern in patterns:
                match = re.search(pattern, title)
                if match:
                    return {
                        "expected_type": expected_type,
                        "actual_type": work_type,
                        "keyword": match.group(0) or NOTICE_WORK_TYPE_LABELS.get(work_type, work_type),
                    }
        return None

    @classmethod
    def validate_manual_notice_type_consistency(
        cls, request_payload: dict[str, Any]
    ) -> None:
        conflict = cls._manual_notice_type_conflict(request_payload)
        if not conflict:
            return
        expected = NOTICE_WORK_TYPE_LABELS.get(
            conflict.get("expected_type", ""),
            conflict.get("expected_type", "") or "当前",
        )
        actual = NOTICE_WORK_TYPE_LABELS.get(
            conflict.get("actual_type", ""),
            conflict.get("actual_type", "") or "其他",
        )
        keyword = conflict.get("keyword", "") or actual
        raise PortalError(
            f"当前选择的是{expected}通告，但标题/名称包含“{keyword}”，"
            f"像是{actual}通告。请改标题或重新选择通告类型。"
        )

    @classmethod
    def _notice_type_for_work_type(cls, work_type: str) -> str:
        return NOTICE_TYPE_BY_WORK_TYPE.get(
            str(work_type or "").strip(),
            NOTICE_TYPE_MAINTENANCE,
        )

    @classmethod
    def _work_type_for_notice_type(cls, notice_type: str) -> str:
        return WORK_TYPE_BY_NOTICE_TYPE.get(str(notice_type or "").strip(), "")

    @classmethod
    def _request_work_type_fallback(cls, request_payload: dict[str, Any]) -> str:
        payload = request_payload if isinstance(request_payload, dict) else {}
        return (
            str(payload.get("work_type") or payload.get("lan_work_type") or "").strip()
            or cls._work_type_for_notice_type(str(payload.get("notice_type") or ""))
            or WORK_TYPE_MAINTENANCE
        )

    @classmethod
    def _normalize_request_notice_type(
        cls, request_payload: dict[str, Any]
    ) -> dict[str, Any]:
        payload = dict(request_payload or {})
        requested_notice_type = str(payload.get("notice_type") or "").strip()
        work_type = cls._manual_payload_notice_work_type(
            payload,
            cls._request_work_type_fallback(payload),
        )
        payload["work_type"] = work_type
        if work_type == WORK_TYPE_POWER and requested_notice_type in {
            NOTICE_TYPE_POWER_UP,
            NOTICE_TYPE_POWER_DOWN,
            NOTICE_TYPE_POWER,
        }:
            if requested_notice_type in {NOTICE_TYPE_POWER_UP, NOTICE_TYPE_POWER_DOWN}:
                payload["power_notice_heading"] = requested_notice_type
                payload["notice_type"] = requested_notice_type
            else:
                payload["notice_type"] = NOTICE_TYPE_POWER_UP
        else:
            payload["notice_type"] = cls._notice_type_for_work_type(work_type)
        return payload

    @classmethod
    def _action_target_key(cls, request_payload: dict[str, Any]) -> str:
        request_payload = normalize_notice_identity_payload(request_payload)
        action = str((request_payload or {}).get("action") or "").strip().lower()
        work_type = cls._manual_payload_notice_work_type(
            request_payload,
            cls._request_work_type_fallback(request_payload),
        )
        if action == "start":
            record_id = str((request_payload or {}).get("record_id") or "").strip()
            manual_id = str((request_payload or {}).get("manual_id") or "").strip()
            if MaintenancePortalService._truthy_flag((request_payload or {}).get("manual")):
                semantic_key = cls._manual_start_target_key(request_payload, work_type)
                if semantic_key:
                    return semantic_key
                if manual_id:
                    return f"{work_type}:manual-start:{manual_id}"
            return f"{work_type}:start:{record_id}" if record_id else ""
        if action in {"update", "end"}:
            target_record_id = canonical_target_record_id(request_payload)
            if target_record_id:
                return f"{work_type}:target:{target_record_id}"
            active_item_id = str(
                (request_payload or {}).get("active_item_id") or ""
            ).strip()
            if active_item_id:
                return f"{work_type}:active:{active_item_id}"
            source_record_id = str(
                (request_payload or {}).get("source_record_id") or ""
            ).strip()
            if action == "update" and source_record_id:
                return f"{work_type}:source-update:{source_record_id}"
        return ""

    def _load_action_jobs_from_state(self) -> dict[str, dict[str, Any]]:
        jobs: dict[str, dict[str, Any]] = {}
        for document in self._state_store.list_documents(STATE_NS_ACTION_JOB):
            job = document.get("payload")
            if not isinstance(job, dict):
                continue
            job_id = str(job.get("job_id") or document.get("key") or "").strip()
            if not job_id:
                continue
            job["job_id"] = job_id
            phase = str(job.get("phase") or "").strip()
            if phase in {"accepted", "queued"} and not float(job.get("message_started_at") or 0):
                job["phase"] = "accepted"
                job["error"] = ""
                job["error_category"] = ""
                job["error_retryable"] = False
                job["restart_recovered"] = True
                job["updated_at"] = dt.datetime.now().strftime("%Y-%m-%d %H:%M")
                self._state_store.put_document(STATE_NS_ACTION_JOB, job_id, job)
            elif phase in {
                "accepted",
                "queued",
                "sending_message",
                "message_sent",
                "upload_queued",
                "qt_queued",
                "qt_displaying",
                "upload_waiting",
                "uploading",
            }:
                job["phase"] = "failed"
                job["error"] = "程序已重启，任务未完成，请核对多维后重试。"
                classified = self.classify_job_error(job["error"])
                job["error_category"] = classified["error_category"]
                job["error_retryable"] = classified["error_retryable"]
                job["restart_recovered"] = False
                job["updated_at"] = dt.datetime.now().strftime("%Y-%m-%d %H:%M")
                self._state_store.put_document(STATE_NS_ACTION_JOB, job_id, job)
            jobs[job_id] = job
        if len(jobs) > ACTION_JOB_MAX_RETAINED:
            terminal_phases = {"success", "failed"}
            terminal_jobs = [
                job
                for job in jobs.values()
                if str(job.get("phase") or "") in terminal_phases
            ]
            terminal_jobs.sort(key=lambda item: str(item.get("created_at") or ""))
            remove_count = min(
                max(0, len(jobs) - ACTION_JOB_MAX_RETAINED),
                len(terminal_jobs),
            )
            for job in terminal_jobs[:remove_count]:
                job_id = str(job.get("job_id") or "")
                jobs.pop(job_id, None)
                self._state_store.delete_document(STATE_NS_ACTION_JOB, job_id)
        return jobs

    def recoverable_action_job_ids(self) -> list[str]:
        with self._jobs_lock:
            return [
                str(job.get("job_id") or "")
                for job in self._jobs.values()
                if isinstance(job, dict)
                and str(job.get("phase") or "") == "accepted"
                and bool(job.get("restart_recovered"))
                and str(job.get("job_id") or "")
            ]

    def _persist_action_job_locked(self, job: dict[str, Any]) -> None:
        job_id = str((job or {}).get("job_id") or "").strip()
        if not job_id:
            return
        self._state_store.put_document(STATE_NS_ACTION_JOB, job_id, job)

    def expand_workbench_action_command(
        self,
        payload: dict[str, Any],
        *,
        scope: str = "ALL",
        ongoing_items: list[dict[str, Any]] | None = None,
    ) -> dict[str, Any]:
        if not isinstance(payload, dict):
            return {}
        if str(payload.get("command_format") or "") != "notice_command":
            return payload
        action = str(payload.get("action") or "").strip().lower()
        patch = payload.get("patch") if isinstance(payload.get("patch"), dict) else {}
        work_type = str(payload.get("work_type") or patch.get("work_type") or WORK_TYPE_MAINTENANCE).strip()
        active_item_id = str(payload.get("active_item_id") or patch.get("active_item_id") or "").strip()
        manual = self._truthy_flag(payload.get("manual")) or self._truthy_flag(patch.get("manual"))
        raw_record_id = str(payload.get("record_id") or patch.get("record_id") or "").strip()
        source_record_id = str(payload.get("source_record_id") or patch.get("source_record_id") or "").strip()
        if manual:
            source_record_id = ""
        if action == "start":
            target_record_id = str(
                payload.get("target_record_id")
                or patch.get("target_record_id")
                or ""
            ).strip()
            if not source_record_id and raw_record_id and not manual and not is_local_record_id(raw_record_id):
                source_record_id = raw_record_id
            if target_record_id and source_record_id and target_record_id == source_record_id:
                target_record_id = ""
        else:
            target_record_id = str(
                payload.get("target_record_id")
                or payload.get("record_id")
                or patch.get("target_record_id")
                or patch.get("record_id")
                or ""
            ).strip()
            if target_record_id and (
                is_local_record_id(target_record_id)
                or (source_record_id and target_record_id == source_record_id)
            ):
                target_record_id = ""
        base: dict[str, Any] = {}
        for item in self._project_ongoing_items(scope, ongoing_items or []):
            item_work_type = self._item_work_type(item)
            if work_type and item_work_type != work_type:
                continue
            if active_item_id and str(item.get("active_item_id") or "").strip() == active_item_id:
                base = item
                break
            if target_record_id and canonical_target_record_id(item) == target_record_id:
                base = item
                break
            if source_record_id and str(item.get("source_record_id") or "").strip() == source_record_id:
                base = item
                break
        if action in {"update", "end"}:
            resolved_active_item_id = active_item_id or str(
                base.get("active_item_id") or ""
            ).strip()
            resolved_source_record_id = source_record_id or canonical_source_record_id(base)
            if not target_record_id:
                target_record_id = canonical_target_record_id(base)
            if target_record_id and (
                is_local_record_id(target_record_id)
                or target_record_id == resolved_source_record_id
            ):
                target_record_id = ""

            identity: dict[str, Any] | None = None
            if resolved_active_item_id or resolved_source_record_id or target_record_id:
                try:
                    identity = self._state_store.resolve_notice_identity(
                        work_type=work_type,
                        active_item_id=resolved_active_item_id,
                        source_record_id=resolved_source_record_id,
                        target_record_id=target_record_id,
                    )
                except Exception:
                    identity = None
            if isinstance(identity, dict):
                identity_payload = (
                    copy.deepcopy(identity.get("payload"))
                    if isinstance(identity.get("payload"), dict)
                    else {}
                )
                for key in (
                    "work_type",
                    "notice_type",
                    "active_item_id",
                    "source_record_id",
                    "target_record_id",
                    "title",
                    "reason",
                    "building_codes",
                    "start_time",
                    "end_time",
                    "status",
                    "origin",
                ):
                    value = identity.get(key)
                    if value not in (None, "", []):
                        identity_payload[key] = copy.deepcopy(value)
                identity_payload = normalize_notice_identity_payload(identity_payload)
                if not base:
                    base = identity_payload
                else:
                    identity_payload.update(base)
                    for key in (
                        "active_item_id",
                        "source_record_id",
                        "target_record_id",
                    ):
                        value = identity.get(key)
                        if value not in (None, ""):
                            identity_payload[key] = value
                    identity_payload = normalize_notice_identity_payload(
                        identity_payload
                    )
                    base = identity_payload
                if not target_record_id:
                    target_record_id = canonical_target_record_id(identity_payload)
                resolved_active_item_id = resolved_active_item_id or str(
                    identity_payload.get("active_item_id") or ""
                ).strip()
                resolved_source_record_id = (
                    resolved_source_record_id
                    or canonical_source_record_id(identity_payload)
                )
                if target_record_id and (
                    is_local_record_id(target_record_id)
                    or target_record_id == resolved_source_record_id
                ):
                    target_record_id = ""

            if not target_record_id:
                try:
                    target_record_id = self._target_record_id_from_work_status(
                        work_type=work_type,
                        source_record_id=resolved_source_record_id,
                        active_item_id=resolved_active_item_id,
                    )
                except Exception:
                    target_record_id = ""
            if target_record_id and (
                is_local_record_id(target_record_id)
                or target_record_id == resolved_source_record_id
            ):
                target_record_id = ""

            active_item_id = active_item_id or resolved_active_item_id
            source_record_id = source_record_id or resolved_source_record_id

        if (
            action in {"update", "end"}
            and not base
            and not target_record_id
            and not active_item_id
            and not source_record_id
        ):
            raise PortalError("更新/结束通告缺少可展开的进行中记录。")
        expanded = copy.deepcopy(base) if base else {}
        expanded.update(copy.deepcopy(patch))
        explicit_site_photos = self._has_site_photo_payload(payload) or self._has_site_photo_payload(patch)
        if not explicit_site_photos:
            for stale_image_key in ("extra_images", "site_photos", "process_site_images"):
                expanded.pop(stale_image_key, None)
        for key in (
            "scope",
            "action",
            "work_type",
            "notice_type",
            "active_item_id",
            "source_record_id",
            "target_record_id",
            "record_id",
            "operation_id",
            "_auth_open_id",
            "_auth_user_name",
        ):
            value = payload.get(key)
            if manual and key == "source_record_id":
                continue
            if value not in (None, ""):
                expanded[key] = value
        if manual:
            expanded.pop("source_record_id", None)
        if action in {"update", "end"} and target_record_id:
            expanded["target_record_id"] = target_record_id
            expanded["record_id"] = target_record_id
        elif action == "start" and target_record_id and target_record_id != source_record_id:
            expanded["target_record_id"] = target_record_id
        if source_record_id:
            expanded["source_record_id"] = source_record_id
        if (
            action == "start"
            and source_record_id
            and str(expanded.get("target_record_id") or "").strip() == source_record_id
        ):
            expanded.pop("target_record_id", None)
        if active_item_id:
            expanded["active_item_id"] = active_item_id
        expanded["scope"] = str(expanded.get("scope") or scope or "ALL")
        expanded["action"] = action
        expanded["work_type"] = str(expanded.get("work_type") or work_type or WORK_TYPE_MAINTENANCE)
        return normalize_notice_identity_payload(expanded, action=action)

    def create_action_job(self, request_payload: dict[str, Any]) -> tuple[str, bool]:
        if not isinstance(request_payload, dict):
            raise PortalError("请求体格式错误。")
        if str(request_payload.get("command_format") or "") == "notice_command":
            patch_payload = request_payload.get("patch")
            patch_payload = patch_payload if isinstance(patch_payload, dict) else {}
            scope = self._normalize_scope(
                request_payload.get("scope") or patch_payload.get("scope") or "ALL"
            )
            try:
                ongoing_items = self._state_store.list_qt_active_items(include_deleted=False)
            except Exception:
                ongoing_items = []
            request_payload = self.expand_workbench_action_command(
                request_payload,
                scope=scope,
                ongoing_items=ongoing_items,
            )
        request_payload = normalize_notice_identity_payload(request_payload)
        self.validate_manual_notice_type_consistency(request_payload)
        request_payload = self._normalize_request_notice_type(request_payload)
        action = str(request_payload.get("action") or "").strip().lower()
        if action not in {"start", "update", "end"}:
            raise PortalError("动作必须是 start/update/end。")
        if action == "start" and not (
            str(request_payload.get("record_id") or "").strip()
            or (
                self._truthy_flag(request_payload.get("manual"))
                and str(request_payload.get("manual_id") or "").strip()
            )
        ):
            raise PortalError("开始通告缺少计划记录ID。")
        if action in {"update", "end"} and not (
            str(request_payload.get("active_item_id") or "").strip()
            or str(request_payload.get("source_record_id") or "").strip()
            or str(request_payload.get("target_record_id") or "").strip()
        ):
            raise PortalError("更新/结束通告缺少主界面条目ID、源记录ID或目标多维record_id。")
        operation_id = str(request_payload.get("operation_id") or "").strip()
        job = self._base_job(request_payload)
        with self._jobs_lock:
            if operation_id:
                for existing in self._jobs.values():
                    if str(existing.get("operation_id") or "") != operation_id:
                        continue
                    phase = str(existing.get("phase") or "")
                    if phase in {
                        "accepted",
                        "queued",
                        "sending_message",
                        "message_sent",
                        "upload_queued",
                        "qt_queued",
                        "qt_displaying",
                        "upload_waiting",
                        "uploading",
                        "success",
                    }:
                        return str(existing.get("job_id") or ""), False
                    if phase == "failed":
                        existing["request"] = copy.deepcopy(request_payload)
                        existing["phase"] = "accepted"
                        existing["error"] = ""
                        existing["error_category"] = ""
                        existing["error_retryable"] = False
                        existing["accepted_at"] = time.time()
                        existing["message_started_at"] = 0.0
                        existing["message_finished_at"] = 0.0
                        existing["upload_queued_at"] = 0.0
                        existing["upload_started_at"] = 0.0
                        existing["upload_finished_at"] = 0.0
                        existing["upload_release_at"] = 0.0
                        existing["qt_phase"] = ""
                        existing["qt_queue_position"] = 0
                        existing["qt_queue_size"] = 0
                        existing["updated_at"] = dt.datetime.now().strftime(
                            "%Y-%m-%d %H:%M"
                        )
                        self._persist_action_job_locked(existing)
                        return str(existing.get("job_id") or ""), True
            target_key = str(job.get("target_key") or "")
            if target_key:
                blocking_phase_order = {
                    "accepted": 1,
                    "queued": 2,
                    "sending_message": 3,
                    "message_sent": 4,
                    "upload_queued": 5,
                    "qt_queued": 6,
                    "qt_displaying": 7,
                    "upload_waiting": 8,
                    "uploading": 9,
                }
                duplicate_start_phases = set(blocking_phase_order) | {"success"}
                blocking_job_id = ""
                blocking_phase = ""
                blocking_rank = -1
                blocking_epoch = 0.0
                for existing in self._jobs.values():
                    if str(existing.get("target_key") or "") != target_key:
                        continue
                    phase = str(existing.get("phase") or "")
                    if action == "start" and phase in duplicate_start_phases:
                        return str(existing.get("job_id") or ""), False
                    if phase not in blocking_phase_order:
                        continue
                    rank = int(blocking_phase_order.get(phase) or 0)
                    epoch = self._job_epoch(existing) or 0.0
                    if rank > blocking_rank or (rank == blocking_rank and epoch >= blocking_epoch):
                        blocking_job_id = str(existing.get("job_id") or "").strip()
                        blocking_phase = phase
                        blocking_rank = rank
                        blocking_epoch = epoch
                if blocking_job_id:
                    job["depends_on_job_id"] = blocking_job_id
                    job["depends_on_phase"] = blocking_phase
            self._jobs[job["job_id"]] = job
            self._persist_action_job_locked(job)
            self._trim_jobs_locked()
        return job["job_id"], True

    def _trim_jobs_locked(self) -> None:
        if len(self._jobs) <= ACTION_JOB_MAX_RETAINED:
            return
        terminal_phases = {"success", "failed"}
        terminal_jobs = [
            job
            for job in self._jobs.values()
            if str(job.get("phase") or "") in terminal_phases
        ]
        terminal_jobs.sort(key=lambda item: self._job_epoch(item) or 0)
        remove_count = max(0, len(self._jobs) - ACTION_JOB_MAX_RETAINED)
        for job in terminal_jobs[:remove_count]:
            job_id = str(job.get("job_id") or "")
            self._jobs.pop(job_id, None)
            self._state_store.delete_document(STATE_NS_ACTION_JOB, job_id)

    def delete_action_job(self, job_id: str) -> bool:
        job_id = str(job_id or "").strip()
        if not job_id:
            return False
        with self._jobs_lock:
            existed = job_id in self._jobs
            self._jobs.pop(job_id, None)
            self._state_store.delete_document(STATE_NS_ACTION_JOB, job_id)
        return existed

    def retry_action_job(self, job_id: str) -> dict[str, Any]:
        job_id = str(job_id or "").strip()
        if not job_id:
            raise PortalError("任务ID不能为空。")
        now_text = dt.datetime.now().strftime("%Y-%m-%d %H:%M")
        now_ts = time.time()
        with self._jobs_lock:
            job = self._jobs.get(job_id)
            if not isinstance(job, dict):
                stored = self._state_store.get_document(STATE_NS_ACTION_JOB, job_id)
                if isinstance(stored, dict):
                    stored["job_id"] = str(stored.get("job_id") or job_id)
                    self._jobs[job_id] = stored
                    job = stored
            if not isinstance(job, dict):
                raise PortalError("任务不存在或已清理。")
            if str(job.get("phase") or "") != "failed":
                raise PortalError("只能重试失败任务。")
            if not bool(job.get("error_retryable")):
                raise PortalError("该失败类型不可自动重试，请人工核对后重新发起。")
            request_payload = (
                job.get("retry_request")
                if isinstance(job.get("retry_request"), dict)
                else job.get("request")
                if isinstance(job.get("request"), dict)
                else {}
            )
            if not request_payload.get("action") or not request_payload.get("work_type"):
                raise PortalError("任务缺少可重试请求内容，请重新发起通告。")
            job["request"] = copy.deepcopy(request_payload)
            job["phase"] = "accepted"
            job["accepted_at"] = now_ts
            job["message_started_at"] = 0.0
            job["message_finished_at"] = 0.0
            job["upload_queued_at"] = 0.0
            job["upload_started_at"] = 0.0
            job["upload_finished_at"] = 0.0
            job["upload_release_at"] = 0.0
            job["message_queue_position"] = 0
            job["message_queue_size"] = 0
            job["qt_phase"] = ""
            job["qt_queue_position"] = 0
            job["qt_queue_size"] = 0
            job["queue_position"] = 0
            job["queue_size"] = 0
            job["upload_queue_position"] = 0
            job["upload_queue_size"] = 0
            job["error"] = ""
            job["error_category"] = ""
            job["error_retryable"] = False
            job["upload_message"] = ""
            job["retry_count"] = int(job.get("retry_count") or 0) + 1
            job["updated_at"] = now_text
            self._persist_action_job_locked(job)
            return copy.deepcopy(job)

    @staticmethod
    def _job_epoch(job: dict[str, Any]) -> float:
        if not isinstance(job, dict):
            return 0.0
        for key in (
            "upload_finished_at",
            "upload_started_at",
            "upload_queued_at",
            "message_finished_at",
            "accepted_at",
        ):
            try:
                value = float(job.get(key) or 0)
            except Exception:
                value = 0.0
            if value > 0:
                return value
        for key in ("updated_at", "created_at"):
            text = str(job.get(key) or "").strip()
            if not text:
                continue
            try:
                return dt.datetime.strptime(text, "%Y-%m-%d %H:%M").timestamp()
            except Exception:
                continue
        return 0.0

    def cleanup_action_jobs(
        self,
        *,
        success_retention_seconds: int = ACTION_JOB_SUCCESS_RETENTION_SECONDS,
        failed_retention_seconds: int = ACTION_JOB_FAILED_RETENTION_SECONDS,
        max_delete: int = 200,
    ) -> dict[str, int]:
        now = time.time()
        success_retention_seconds = max(60, int(success_retention_seconds or 0))
        failed_retention_seconds = max(60, int(failed_retention_seconds or 0))
        max_delete = max(1, min(int(max_delete or 200), 1000))
        removed_success = 0
        removed_failed = 0
        candidates: list[tuple[float, str, str]] = []
        documents = self._state_store.list_documents(STATE_NS_ACTION_JOB)
        for document in documents:
            job = document.get("payload") if isinstance(document, dict) else {}
            if not isinstance(job, dict):
                continue
            job_id = str(job.get("job_id") or document.get("key") or "").strip()
            phase = str(job.get("phase") or "").strip()
            if not job_id or phase not in {"success", "failed"}:
                continue
            epoch = self._job_epoch(job)
            if not epoch:
                continue
            retention = (
                success_retention_seconds
                if phase == "success"
                else failed_retention_seconds
            )
            if now - epoch >= retention:
                candidates.append((epoch, job_id, phase))
        candidates.sort(key=lambda item: item[0])
        with self._jobs_lock:
            for _, job_id, phase in candidates[:max_delete]:
                self._jobs.pop(job_id, None)
                self._state_store.delete_document(STATE_NS_ACTION_JOB, job_id)
                if phase == "success":
                    removed_success += 1
                else:
                    removed_failed += 1
        return {
            "removed_success": removed_success,
            "removed_failed": removed_failed,
            "removed_total": removed_success + removed_failed,
        }

    @staticmethod
    def _compact_job_request(request: dict[str, Any]) -> dict[str, Any]:
        if not isinstance(request, dict):
            return {}
        keep_keys = (
            "_auth_open_id",
            "scope",
            "action",
            "work_type",
            "notice_type",
            "manual",
            "manual_id",
            "record_id",
            "source_record_id",
            "target_record_id",
            "active_item_id",
            "operation_id",
        )
        return {
            key: request.get(key)
            for key in keep_keys
            if request.get(key) not in (None, "")
        }

    def _compact_completed_job_locked(self, job_id: str) -> None:
        job = self._jobs.get(job_id)
        if not isinstance(job, dict) or str(job.get("phase") or "") != "success":
            return
        compacted = {
            "job_id": str(job.get("job_id") or job_id),
            "phase": "success",
            "created_at": str(job.get("created_at") or ""),
            "updated_at": str(job.get("updated_at") or ""),
            "accepted_at": float(job.get("accepted_at") or 0),
            "message_started_at": float(job.get("message_started_at") or 0),
            "message_finished_at": float(job.get("message_finished_at") or 0),
            "upload_queued_at": float(job.get("upload_queued_at") or 0),
            "upload_started_at": float(job.get("upload_started_at") or 0),
            "upload_finished_at": float(job.get("upload_finished_at") or 0),
            "upload_release_at": float(job.get("upload_release_at") or 0),
            "request": self._compact_job_request(job.get("request") or {}),
            "operation_id": str(job.get("operation_id") or ""),
            "target_key": str(job.get("target_key") or ""),
            "depends_on_job_id": str(job.get("depends_on_job_id") or ""),
            "depends_on_phase": "",
            "message_sent": bool(job.get("message_sent")),
            "message_signature": str(job.get("message_signature") or ""),
            "message_failed": bool(job.get("message_failed")),
            "message_failed_continue": bool(job.get("message_failed_continue")),
            "message_error": str(job.get("message_error") or ""),
            "message_warning": str(job.get("message_warning") or ""),
            "message_queue_position": 0,
            "message_queue_size": 0,
            "qt_phase": str(job.get("qt_phase") or ""),
            "qt_queue_position": 0,
            "qt_queue_size": 0,
            "queue_position": 0,
            "queue_size": 0,
            "upload_queue_position": 0,
            "upload_queue_size": 0,
            "record_id": str(job.get("record_id") or ""),
            "active_item_id": str(job.get("active_item_id") or ""),
            "error": "",
            "error_category": "",
            "error_retryable": False,
            "upload_message": str(job.get("upload_message") or ""),
            "paired_upload_status": str(job.get("paired_upload_status") or ""),
            "paired_upload_warning": str(job.get("paired_upload_warning") or ""),
            "paired_maintenance_target_record_id": str(
                job.get("paired_maintenance_target_record_id") or ""
            ),
            "frontend_patch": copy.deepcopy(job.get("frontend_patch") or {}),
            "notice_text": str(
                job.get("notice_text")
                or (job.get("prepared") or {}).get("text")
                or ""
            ),
            "copy_text": str(
                job.get("copy_text")
                or job.get("notice_text")
                or (job.get("prepared") or {}).get("text")
                or ""
            ),
            "prepared": {},
        }
        self._jobs[job_id] = compacted
        self._persist_action_job_locked(compacted)

    def _compact_failed_job_locked(self, job_id: str) -> None:
        job = self._jobs.get(job_id)
        if not isinstance(job, dict) or str(job.get("phase") or "") != "failed":
            return
        retryable = bool(job.get("error_retryable"))
        compacted = {
            "job_id": str(job.get("job_id") or job_id),
            "phase": "failed",
            "created_at": str(job.get("created_at") or ""),
            "updated_at": str(job.get("updated_at") or ""),
            "accepted_at": float(job.get("accepted_at") or 0),
            "message_started_at": float(job.get("message_started_at") or 0),
            "message_finished_at": float(job.get("message_finished_at") or 0),
            "upload_queued_at": float(job.get("upload_queued_at") or 0),
            "upload_started_at": float(job.get("upload_started_at") or 0),
            "upload_finished_at": float(job.get("upload_finished_at") or 0),
            "upload_release_at": float(job.get("upload_release_at") or 0),
            "request": self._compact_job_request(job.get("request") or {}),
            "operation_id": str(job.get("operation_id") or ""),
            "target_key": str(job.get("target_key") or ""),
            "depends_on_job_id": str(job.get("depends_on_job_id") or ""),
            "depends_on_phase": str(job.get("depends_on_phase") or ""),
            "message_sent": bool(job.get("message_sent")),
            "message_signature": str(job.get("message_signature") or ""),
            "message_failed": bool(job.get("message_failed")),
            "message_failed_continue": bool(job.get("message_failed_continue")),
            "message_error": str(job.get("message_error") or ""),
            "message_warning": str(job.get("message_warning") or ""),
            "message_queue_position": 0,
            "message_queue_size": 0,
            "qt_phase": str(job.get("qt_phase") or ""),
            "qt_queue_position": 0,
            "qt_queue_size": 0,
            "queue_position": 0,
            "queue_size": 0,
            "upload_queue_position": 0,
            "upload_queue_size": 0,
            "record_id": str(job.get("record_id") or ""),
            "target_record_id": str(job.get("target_record_id") or ""),
            "active_item_id": str(job.get("active_item_id") or ""),
            "target_record_missing": bool(job.get("target_record_missing")),
            "needs_target_selection": bool(job.get("needs_target_selection")),
            "target_selection_message": str(job.get("target_selection_message") or ""),
            "target_candidates": copy.deepcopy(job.get("target_candidates") or []),
            "error": str(job.get("error") or ""),
            "error_category": str(job.get("error_category") or ""),
            "error_retryable": retryable,
            "upload_message": str(job.get("upload_message") or ""),
            "retry_count": int(job.get("retry_count") or 0),
            "frontend_patch": copy.deepcopy(job.get("frontend_patch") or {}),
            "notice_text": str(
                job.get("notice_text")
                or (job.get("prepared") or {}).get("text")
                or ""
            ),
            "copy_text": str(
                job.get("copy_text")
                or job.get("notice_text")
                or (job.get("prepared") or {}).get("text")
                or ""
            ),
            "prepared": {},
        }
        if retryable:
            compacted["retry_request"] = copy.deepcopy(job.get("request") or {})
        self._jobs[job_id] = compacted
        self._persist_action_job_locked(compacted)

    def get_job(self, job_id: str) -> dict[str, Any] | None:
        job_id = str(job_id or "").strip()
        if not job_id:
            return None
        with self._jobs_lock:
            job = self._jobs.get(job_id)
            if not job:
                stored = self._state_store.get_document(STATE_NS_ACTION_JOB, job_id)
                if isinstance(stored, dict):
                    stored["job_id"] = str(stored.get("job_id") or job_id)
                    self._jobs[job_id] = stored
                    job = stored
            return copy.deepcopy(job) if job else None

    def mark_job(self, job_id: str, **patch: Any) -> None:
        job_id = str(job_id or "").strip()
        if not job_id:
            return
        persist = bool(patch.pop("_persist", True))
        with self._jobs_lock:
            job = self._jobs.get(job_id)
            if not job:
                return
            phase = str(patch.get("phase") or "").strip()
            prior_phase = str(job.get("phase") or "").strip()
            if phase == "accepted":
                patch.setdefault("error_category", "")
                patch.setdefault("error_retryable", False)
            if phase == "failed" or ("error" in patch and str(patch.get("error") or "").strip()):
                classified = self.classify_job_error(str(patch.get("error") or job.get("error") or ""))
                patch.setdefault("error_category", classified["error_category"])
                patch.setdefault("error_retryable", classified["error_retryable"])
            if phase and phase != prior_phase:
                now_ts = time.time()
                if phase == "accepted" and not job.get("accepted_at"):
                    patch["accepted_at"] = now_ts
                elif phase == "sending_message":
                    patch["message_started_at"] = now_ts
                elif phase in {"message_sent", "qt_queued"}:
                    if not job.get("message_finished_at") and job.get("message_started_at"):
                        patch["message_finished_at"] = now_ts
                elif phase == "upload_queued":
                    if not job.get("message_finished_at") and job.get("message_started_at"):
                        patch["message_finished_at"] = now_ts
                    if not job.get("upload_queued_at"):
                        patch["upload_queued_at"] = now_ts
                elif phase == "upload_waiting":
                    if not job.get("upload_queued_at"):
                        patch["upload_queued_at"] = now_ts
                elif phase == "uploading":
                    if not job.get("upload_queued_at"):
                        patch["upload_queued_at"] = now_ts
                    patch["upload_started_at"] = now_ts
                elif phase in {"success", "failed"}:
                    patch["upload_finished_at"] = now_ts
            changed = any(job.get(key) != value for key, value in patch.items())
            if not changed:
                return
            job.update(patch)
            job["updated_at"] = dt.datetime.now().strftime("%Y-%m-%d %H:%M")
            if persist:
                self._persist_action_job_locked(job)
                if str(job.get("phase") or "") == "failed":
                    self._compact_failed_job_locked(job_id)

    @staticmethod
    def _action_message_signature(prepared: dict[str, Any]) -> str:
        payload = {
            "text": str((prepared or {}).get("text") or ""),
            "recipients": sorted(
                {
                    str(open_id or "").strip()
                    for open_id in ((prepared or {}).get("recipients") or [])
                    if str(open_id or "").strip()
                }
            ),
        }
        raw = json.dumps(payload, ensure_ascii=False, sort_keys=True)
        return hashlib.sha256(raw.encode("utf-8")).hexdigest()

    def prepare_action_job(self, job_id: str) -> dict[str, Any]:
        job = self.get_job(job_id)
        if not job:
            raise PortalError("任务不存在。")
        request_payload = job.get("request") or {}
        prepared = self.prepare_workbench_action(request_payload, job_id=job_id)
        message_signature = self._action_message_signature(prepared)
        prepared["message_signature"] = message_signature
        prepared["message_sent"] = bool(job.get("message_sent")) and (
            str(job.get("message_signature") or "") == message_signature
        )
        self.mark_job(
            job_id,
            prepared=prepared,
            record_id=str(prepared.get("record_id") or ""),
            active_item_id=str(prepared.get("active_item_id") or ""),
            message_signature=message_signature,
        )
        return prepared

    def prepare_workbench_action(
        self, request_payload: dict[str, Any], *, job_id: str
    ) -> dict[str, Any]:
        request_payload = normalize_notice_identity_payload(request_payload)
        work_type = self._manual_payload_notice_work_type(
            request_payload,
            self._request_work_type_fallback(request_payload),
        )
        if work_type != str(request_payload.get("work_type") or "").strip():
            request_payload = {**request_payload, "work_type": work_type}
        if work_type == WORK_TYPE_CHANGE:
            return self.prepare_change_action(request_payload, job_id=job_id)
        if work_type == WORK_TYPE_REPAIR:
            return self.prepare_repair_action(request_payload, job_id=job_id)
        if work_type in {WORK_TYPE_POWER, WORK_TYPE_POLLING, WORK_TYPE_ADJUST}:
            return self.prepare_simple_manual_notice_action(request_payload, job_id=job_id)
        return self.prepare_maintenance_action(request_payload, job_id=job_id)

    @staticmethod
    def _simple_notice_profile(work_type: str) -> dict[str, str]:
        work_type = str(work_type or "").strip()
        if work_type == WORK_TYPE_POWER:
            return {
                "notice_type": NOTICE_TYPE_POWER_UP,
                "heading": "上电通告",
                "status_label": "上电状态",
                "title_label": "名称",
            }
        if work_type == WORK_TYPE_POLLING:
            return {
                "notice_type": NOTICE_TYPE_POLLING,
                "heading": NOTICE_TYPE_POLLING,
                "status_label": "轮巡状态",
                "title_label": "标题",
            }
        if work_type == WORK_TYPE_ADJUST:
            return {
                "notice_type": NOTICE_TYPE_ADJUST,
                "heading": NOTICE_TYPE_ADJUST,
                "status_label": "调整状态",
                "title_label": "名称",
            }
        return {
            "notice_type": NOTICE_TYPE_MAINTENANCE,
            "heading": NOTICE_TYPE_MAINTENANCE,
            "status_label": "状态",
            "title_label": "名称",
        }

    @staticmethod
    def build_simple_notice_text(
        *,
        work_type: str,
        status: str,
        title: str,
        start_time: str,
        end_time: str,
        building: str,
        specialty: str,
        location: str = "",
        content: str = "",
        reason: str = "",
        impact: str = "",
        progress: str = "",
        device: str = "",
        cabinet: str = "",
        quantity: str = "",
        notice_type: str = "",
    ) -> str:
        heading = (
            notice_type
            if work_type == WORK_TYPE_POWER
            and str(notice_type or "").strip() in {NOTICE_TYPE_POWER_UP, NOTICE_TYPE_POWER_DOWN}
            else ""
        )
        return MaintenancePortalService._render_notice_text(
            work_type=work_type,
            status=status,
            heading_override=heading,
            values={
                "title": title,
                "time_range": MaintenancePortalService._format_notice_time_range(
                    start_time, end_time
                ),
                "location": location,
                "content": content,
                "reason": reason,
                "impact": impact,
                "progress": progress,
                "device": device,
                "cabinet": cabinet,
                "quantity": quantity,
            },
        )

    def prepare_simple_manual_notice_action(
        self, request_payload: dict[str, Any], *, job_id: str
    ) -> dict[str, Any]:
        request_payload = normalize_notice_identity_payload(request_payload)
        action = str(request_payload.get("action") or "").strip().lower()
        status_map = {"start": "开始", "update": "更新", "end": "结束"}
        status = status_map.get(action)
        if not status:
            raise PortalError("动作必须是 start/update/end。")
        work_type = str(request_payload.get("work_type") or "").strip()
        if work_type not in {WORK_TYPE_POWER, WORK_TYPE_POLLING, WORK_TYPE_ADJUST}:
            raise PortalError("不支持的手填通告类型。")
        profile = self._simple_notice_profile(work_type)
        request_notice_type = str(request_payload.get("notice_type") or "").strip()
        heading_notice_type = (
            request_notice_type
            if work_type == WORK_TYPE_POWER
            and request_notice_type in {NOTICE_TYPE_POWER_UP, NOTICE_TYPE_POWER_DOWN}
            else str(request_payload.get("power_notice_heading") or "").strip()
        )
        notice_type = (
            heading_notice_type
            if work_type == WORK_TYPE_POWER
            and heading_notice_type in {NOTICE_TYPE_POWER_UP, NOTICE_TYPE_POWER_DOWN}
            else profile["notice_type"]
        )
        scope = self._normalize_scope(request_payload.get("scope"))
        manual = self._truthy_flag(request_payload.get("manual"))
        record_id = str(
            request_payload.get("record_id")
            or request_payload.get("manual_id")
            or request_payload.get("active_item_id")
            or ""
        ).strip()
        target_record_id = self._target_record_id_from_request_payload(request_payload)
        if not manual and action == "start":
            raise PortalError(f"{self._history_work_type_label(work_type)}通告目前仅支持前端纯手填或解析发送。")
        if not manual and action != "start" and not target_record_id:
            raise PortalError(f"{self._history_work_type_label(work_type)}通告缺少目标多维 record_id，不能更新/结束。")
        if action != "start" and target_record_id:
            manual = True
        if action == "start" and not record_id:
            raise PortalError("开始通告缺少手填记录ID。")
        if action != "start" and not target_record_id:
            raise PortalError(f"{self._history_work_type_label(work_type)}通告缺少目标多维 record_id，不能更新/结束。")
        title = str(request_payload.get("title") or request_payload.get("content") or "").strip()
        if not title:
            raise PortalError(f"{self._history_work_type_label(work_type)}通告缺少标题。")
        start_time = self._format_input_datetime(request_payload.get("start_time"))
        end_time = self._format_input_datetime(request_payload.get("end_time"))
        if not start_time or not end_time:
            raise PortalError("计划开始时间和计划结束时间不能为空。")
        self._validate_minimum_notice_duration(start_time, end_time)
        self._require_end_site_photo_cumulative(
            request_payload,
            action,
            notice_type=notice_type,
            work_type=work_type,
        )
        specialty = self._clean_source_text(request_payload.get("specialty"))
        location = str(request_payload.get("location") or "").strip()
        content = str(request_payload.get("content") or "").strip()
        reason = str(request_payload.get("reason") or "").strip()
        impact = str(request_payload.get("impact") or "").strip()
        progress = str(request_payload.get("progress") or "").strip()
        device = str(request_payload.get("device") or "").strip()
        cabinet = str(request_payload.get("cabinet") or "").strip()
        quantity = str(request_payload.get("quantity") or "").strip()
        building, building_codes = self._resolve_notice_submit_building(
            scope=scope,
            request_payload=request_payload,
            building=request_payload.get("building"),
            title=title,
            location=location,
            content=content,
            work_type_label=f"{self._history_work_type_label(work_type)}通告",
            allow_scope_fallback=manual,
        )
        title = self._normalize_110_station_notice_title(
            title,
            building=building,
            building_codes=building_codes,
        )
        text = self.build_simple_notice_text(
            work_type=work_type,
            status=status,
            title=title,
            start_time=start_time,
            end_time=end_time,
            building=building,
            specialty=specialty,
            location=location,
            content=content,
            reason=reason,
            impact=impact,
            progress=progress,
            device=device,
            cabinet=cabinet,
            quantity=quantity,
            notice_type=heading_notice_type,
        )
        building_code, recipients, recipient_error = self._recipients_for_building_codes(
            building_codes,
            fallback_building=building,
        )
        if recipient_error:
            raise PortalError(recipient_error)
        response_time = self._action_response_time(request_payload)
        return {
            "job_id": job_id,
            "work_type": work_type,
            "notice_type": notice_type,
            "manual": True,
            "action": action,
            "status": status,
            "scope": scope,
            "scope_label": self._scope_label(scope),
            "record_id": record_id,
            "target_record_id": target_record_id,
            "active_item_id": str(request_payload.get("active_item_id") or "").strip(),
            "title": title,
            "building": building,
            "building_code": building_code,
            "building_codes": building_codes,
            "target_building": self._building_label_from_codes(building_codes),
            "specialty": specialty,
            "start_time": start_time,
            "end_time": end_time,
            "location": location,
            "content": content,
            "reason": reason,
            "impact": impact,
            "progress": progress,
            "device": device,
            "cabinet": cabinet,
            "quantity": quantity,
            "text": text,
            "extra_images": self._site_photo_payload(request_payload),
            "recipients": recipients,
            "skip_personal_message": False,
            "response_time": response_time,
            "operation_id": str(request_payload.get("operation_id") or "").strip() or job_id,
        }

    def prepare_maintenance_action(
        self, request_payload: dict[str, Any], *, job_id: str
    ) -> dict[str, Any]:
        request_payload = normalize_notice_identity_payload(request_payload)
        action = str(request_payload.get("action") or "").strip().lower()
        status_map = {"start": "开始", "update": "更新", "end": "结束"}
        status = status_map.get(action)
        if not status:
            raise PortalError("动作必须是 start/update/end。")
        self.ensure_snapshot_loaded()
        scope = self._normalize_scope(request_payload.get("scope"))
        specialty = self._clean_source_text(request_payload.get("specialty"))
        manual = self._truthy_flag(request_payload.get("manual"))

        record = None
        fields: dict[str, Any] = {}
        record_id = str(
            request_payload.get("record_id")
            or (request_payload.get("manual_id") if manual else "")
            or ""
        ).strip()
        source_record_id = str(request_payload.get("source_record_id") or "").strip()
        target_record_id = self._target_record_id_from_request_payload(
            request_payload,
            source_record_id=source_record_id,
        )
        active_item_id = str(request_payload.get("active_item_id") or "").strip()
        plan_month = ""
        maintenance_total = ""
        maintenance_no = ""
        maintenance_project = ""
        maintenance_cycle = str(request_payload.get("maintenance_cycle") or "").strip()
        if action == "start":
            if not record_id:
                raise PortalError("开始通告缺少计划记录ID。")
            if manual:
                building = str(request_payload.get("building") or "").strip()
                title = str(request_payload.get("title") or "").strip()
                if not title:
                    raise PortalError("纯手填维保通告缺少名称。")
                if not specialty or specialty == "全部":
                    raise PortalError("纯手填维保通告缺少专业。")
                if not maintenance_cycle:
                    raise PortalError("纯手填维保通告必须选择维保周期。")
            else:
                record = self._find_record_by_id(record_id)
                source_record_id = record_id
                fields = record.get("display_fields") or {}
                current_status = str(fields.get("维护实施状态") or "").strip()
                if not self._maintenance_status_is_startable(record):
                    raise PortalError(
                        f"该计划维护项当前状态不是{DEFAULT_MAINTENANCE_STATUS}: {current_status or '-'}"
                    )
                building = str(fields.get("楼栋") or "").strip()
                maintenance_total = str(fields.get("维护总项") or "").strip()
                plan_month = str(fields.get("计划维护月份") or "").strip()
                maintenance_no = str(fields.get("维护编号") or "").strip()
                maintenance_project = str(fields.get("维护项目") or "").strip()
                maintenance_cycle = str(fields.get("维护周期") or maintenance_cycle).strip()
                if not specialty or specialty == "全部":
                    specialty = str(fields.get("专业类别") or "").strip()
                if not specialty or specialty == "全部":
                    raise PortalError("该计划维护项缺少专业类别，无法上传。")
                title = str(request_payload.get("title") or "").strip()
                if not title:
                    title = f"EA118机房{building}{maintenance_total}"
        else:
            if source_record_id:
                try:
                    record = self._find_record_by_id(source_record_id)
                    fields = record.get("display_fields") or {}
                except PortalError:
                    record = None
                    fields = {}
            if record is not None:
                building = str(request_payload.get("building") or fields.get("楼栋") or "").strip()
                maintenance_total = str(fields.get("维护总项") or "").strip()
                plan_month = str(fields.get("计划维护月份") or "").strip()
                maintenance_no = str(fields.get("维护编号") or "").strip()
                maintenance_project = str(fields.get("维护项目") or "").strip()
                maintenance_cycle = str(fields.get("维护周期") or maintenance_cycle).strip()
                if not target_record_id:
                    target_record_id = self._resolve_target_record_id_for_source_update(
                        work_type=WORK_TYPE_MAINTENANCE,
                        notice_type=NOTICE_TYPE_MAINTENANCE,
                        source_record=record,
                    )
            else:
                building = str(request_payload.get("building") or "").strip()
            building = str(request_payload.get("building") or building or "").strip()
            title = str(request_payload.get("title") or "").strip()
            if not title:
                raise PortalError("更新/结束通告缺少名称。")
            if not target_record_id:
                target_record_id = self._target_record_id_from_identity_map(
                    work_type=WORK_TYPE_MAINTENANCE,
                    active_item_id=active_item_id,
                    source_record_id=source_record_id,
                    target_record_id=self._target_record_id_from_request_payload(
                        request_payload,
                        source_record_id=source_record_id,
                    ),
                )
            if not target_record_id and active_item_id:
                target_record_id = self._target_record_id_from_work_status(
                    work_type=WORK_TYPE_MAINTENANCE,
                    active_item_id=active_item_id,
                )
            if not target_record_id and source_record_id:
                target_record_id = self._target_record_id_from_work_status(
                    work_type=WORK_TYPE_MAINTENANCE,
                    source_record_id=source_record_id,
                )
            if not target_record_id:
                target_record_id = self._target_record_id_from_unique_candidate(
                    work_type=WORK_TYPE_MAINTENANCE,
                    scope=scope,
                    title=title,
                    start_time=self._format_input_datetime(request_payload.get("start_time")),
                    end_time=self._format_input_datetime(request_payload.get("end_time")),
                    action=action,
                )
            if not target_record_id:
                target_record_id = self._target_record_id_from_request_payload(
                    request_payload,
                    source_record_id=source_record_id,
                )
            record_id = source_record_id or record_id
            if action != "start" and not target_record_id:
                raise PortalError(
                    "该维保源记录未找到目标维保通告表 record_id，不能更新/结束。"
                )

        start_time = self._format_input_datetime(request_payload.get("start_time"))
        end_time = self._format_input_datetime(request_payload.get("end_time"))
        if not start_time or not end_time:
            raise PortalError("开始时间和结束时间不能为空。")
        self._validate_minimum_notice_duration(start_time, end_time)
        self._require_end_site_photo_cumulative(
            request_payload,
            action,
            notice_type=NOTICE_TYPE_MAINTENANCE,
            work_type=WORK_TYPE_MAINTENANCE,
        )
        location = str(request_payload.get("location") or "").strip()
        content = str(request_payload.get("content") or "").strip()
        reason = str(request_payload.get("reason") or "").strip()
        impact = str(request_payload.get("impact") or "").strip() or DEFAULT_IMPACT_TEXT
        progress = (
            str(request_payload.get("progress") or "").strip()
            or (DEFAULT_PROGRESS_TEXT if action == "start" else "")
        )
        building, building_codes = self._resolve_notice_submit_building(
            scope=scope,
            request_payload=request_payload,
            building=building,
            title=title,
            location=location,
            content=content,
            work_type_label="维保通告",
            allow_scope_fallback=manual,
        )
        title = self._normalize_110_station_notice_title(
            title,
            building=building,
            building_codes=building_codes,
        )
        if action == "start":
            memory_total = (
                str(maintenance_total or fields.get("维护总项") or "").strip()
                or str(request_payload.get("title") or "").strip()
            )
            if not maintenance_total:
                maintenance_total = memory_total
            self._remember_draft_fields(
                building=building,
                maintenance_total=memory_total,
                location=location,
                content=content,
                reason=reason,
                impact=impact,
                maintenance_cycle=maintenance_cycle,
            )

        text = self.build_notice_text(
            status=status,
            title=title,
            start_time=start_time,
            end_time=end_time,
            location=location,
            content=content,
            reason=reason,
            impact=impact,
            progress=progress,
        )
        building_code, recipients, recipient_error = self._recipients_for_building_codes(
            building_codes,
            fallback_building=building,
        )
        if recipient_error:
            raise PortalError(recipient_error)

        response_time = self._action_response_time(request_payload)
        return {
            "job_id": job_id,
            "work_type": WORK_TYPE_MAINTENANCE,
            "notice_type": NOTICE_TYPE_MAINTENANCE,
            "source_app_token": self.app_token if (record_id and not manual) else "",
            "source_table_id": self.table_id if (record_id and not manual) else "",
            "target_app_token": str(config.app_token or "").strip(),
            "target_table_id": str(config.get_table_id(NOTICE_TYPE_MAINTENANCE) or "").strip(),
            "manual": manual,
            "action": action,
            "status": status,
            "scope": scope,
            "scope_label": self._scope_label(scope),
            "record_id": record_id,
            "source_record_id": source_record_id,
            "target_record_id": target_record_id,
            "active_item_id": active_item_id,
            "title": title,
            "building": building,
            "building_code": building_code,
            "building_codes": building_codes,
            "target_building": self._building_label_from_codes(building_codes),
            "specialty": specialty,
            "plan_month": plan_month,
            "maintenance_total": maintenance_total,
            "maintenance_no": maintenance_no,
            "maintenance_project": maintenance_project,
            "maintenance_cycle": maintenance_cycle,
            "source_progress": (
                self._maintenance_status_value(record)
                if record is not None
                else str(request_payload.get("source_progress") or "")
            ),
            "start_time": start_time,
            "end_time": end_time,
            "location": location,
            "content": content,
            "reason": reason,
            "impact": impact,
            "progress": progress,
            "text": text,
            "extra_images": self._site_photo_payload(request_payload),
            "recipients": recipients,
            "response_time": response_time,
            "operation_id": str(request_payload.get("operation_id") or "").strip()
            or job_id,
        }

    @staticmethod
    def build_change_notice_text(
        *,
        status: str,
        title: str,
        level: str,
        start_time: str,
        end_time: str,
        location: str,
        content: str,
        reason: str,
        impact: str,
        progress: str,
    ) -> str:
        return MaintenancePortalService._render_notice_text(
            work_type=WORK_TYPE_CHANGE,
            status=status,
            values={
                "title": title,
                "level": level,
                "time_range": MaintenancePortalService._format_notice_time_range(
                    start_time, end_time
                ),
                "location": location,
                "content": content,
                "reason": reason,
                "impact": str(impact or "").strip() or DEFAULT_IMPACT_TEXT,
                "progress": str(progress or "").strip() or DEFAULT_PROGRESS_TEXT,
            },
        )

    @staticmethod
    def _default_change_level(level: Any) -> str:
        return str(level or "").strip() or CHANGE_DEFAULT_LEVEL

    def _sync_maintenance_target_requested(
        self, request_payload: dict[str, Any], *, source_work_type: str
    ) -> bool:
        if source_work_type != WORK_TYPE_MAINTENANCE:
            return False
        if "sync_maintenance_target" not in request_payload:
            return True
        return self._truthy_flag(request_payload.get("sync_maintenance_target"))

    def _build_paired_maintenance_upload(
        self,
        *,
        request_payload: dict[str, Any],
        job_id: str,
        action: str,
        status: str,
        source_record_id: str,
        target_record_id: str,
        active_item_id: str,
        original_title: str,
        building: str,
        building_codes: list[str],
        specialty: str,
        maintenance_cycle: str,
        start_time: str,
        end_time: str,
        location: str,
        content: str,
        reason: str,
        impact: str,
        progress: str,
        response_time: str,
    ) -> dict[str, Any]:
        paired_actual_start = str(
            request_payload.get("paired_maintenance_actual_start_time") or ""
        ).strip()
        if action == "start" and not paired_actual_start:
            paired_actual_start = response_time
        title = self._normalize_110_station_notice_title(
            original_title or request_payload.get("paired_maintenance_original_title") or "",
            building=building,
            building_codes=building_codes,
        )
        if not title:
            title = self._normalize_110_station_notice_title(
                request_payload.get("title") or "",
                building=building,
                building_codes=building_codes,
            )
        text = self.build_notice_text(
            status=status,
            title=title,
            start_time=start_time,
            end_time=end_time,
            location=location,
            content=content,
            reason=reason,
            impact=impact,
            progress=progress,
        )
        return {
            "job_id": f"{job_id}:paired-maintenance",
            "work_type": WORK_TYPE_MAINTENANCE,
            "notice_type": NOTICE_TYPE_MAINTENANCE,
            "source_work_type": WORK_TYPE_MAINTENANCE,
            "source_app_token": self.app_token if source_record_id else "",
            "source_table_id": self.table_id if source_record_id else "",
            "target_app_token": str(config.app_token or "").strip(),
            "target_table_id": str(config.get_table_id(NOTICE_TYPE_MAINTENANCE) or "").strip(),
            "manual": False,
            "action": action,
            "status": status,
            "scope": self._normalize_scope(request_payload.get("scope")),
            "scope_label": self._scope_label(self._normalize_scope(request_payload.get("scope"))),
            "record_id": source_record_id,
            "source_record_id": source_record_id,
            "target_record_id": target_record_id,
            "active_item_id": active_item_id,
            "title": title,
            "building": building,
            "buildings": [building] if building else [],
            "building_code": (
                "CAMPUS"
                if len(building_codes) >= 2
                else (building_codes[0] if building_codes else "")
            ),
            "building_codes": building_codes,
            "target_building": self._building_label_from_codes(building_codes),
            "specialty": specialty,
            "maintenance_cycle": maintenance_cycle,
            "start_time": start_time,
            "end_time": end_time,
            "location": location,
            "content": content,
            "reason": reason,
            "impact": impact,
            "progress": progress,
            "text": text,
            "extra_images": self._site_photo_payload(request_payload),
            "recipients": [],
            "skip_personal_message": True,
            "response_time": response_time,
            "paired_maintenance_actual_start_time": paired_actual_start,
            "operation_id": (
                str(request_payload.get("operation_id") or "").strip() or job_id
            )
            + ":paired-maintenance",
        }

    def prepare_change_action(
        self, request_payload: dict[str, Any], *, job_id: str
    ) -> dict[str, Any]:
        request_payload = normalize_notice_identity_payload(request_payload)
        action = str(request_payload.get("action") or "").strip().lower()
        status_map = {"start": "开始", "update": "更新", "end": "结束"}
        status = status_map.get(action)
        if not status:
            raise PortalError("动作必须是 start/update/end。")
        self.ensure_snapshot_loaded()
        scope = self._normalize_scope(request_payload.get("scope"))
        manual = self._truthy_flag(request_payload.get("manual"))
        source_work_type = str(
            request_payload.get("source_work_type")
            or request_payload.get("converted_from_work_type")
            or WORK_TYPE_CHANGE
        ).strip()
        if source_work_type not in {WORK_TYPE_MAINTENANCE, WORK_TYPE_CHANGE}:
            source_work_type = WORK_TYPE_CHANGE

        record_id = str(
            request_payload.get("record_id")
            or (request_payload.get("manual_id") if manual else "")
            or ""
        ).strip()
        source_record_id = str(request_payload.get("source_record_id") or "").strip()
        source_identity_record_id = source_record_id or (record_id if action == "start" else "")
        if (
            not manual
            and source_work_type == WORK_TYPE_CHANGE
            and source_identity_record_id
        ):
            try:
                maintenance_source = self._find_record_by_id(
                    source_identity_record_id, WORK_TYPE_MAINTENANCE
                )
            except PortalError:
                maintenance_source = None
            if (
                maintenance_source is not None
                and self._maintenance_record_is_converted_to_change(
                    maintenance_source
                )
            ):
                source_work_type = WORK_TYPE_MAINTENANCE
        target_record_id = ""
        fields: dict[str, Any] = {}
        paired_maintenance_original_title = str(
            request_payload.get("paired_maintenance_original_title") or ""
        ).strip()
        paired_maintenance_target_record_id = str(
            request_payload.get("paired_maintenance_target_record_id") or ""
        ).strip()
        maintenance_cycle = str(request_payload.get("maintenance_cycle") or "").strip()
        if action == "start":
            if not record_id:
                raise PortalError("开始通告缺少变更源记录ID。")
            if manual:
                title = str(request_payload.get("title") or "").strip()
                if not title:
                    raise PortalError("纯手填变更通告缺少名称。")
                building_codes = [
                    str(code or "").strip().upper()
                    for code in (request_payload.get("building_codes") or [])
                    if str(code or "").strip()
                ]
                if not building_codes:
                    building_codes = self._building_codes_from_value(
                        request_payload.get("building")
                    )
                building = (
                    str(request_payload.get("building") or "").strip()
                    or self._building_label_from_codes(building_codes)
                )
                specialty = self._clean_source_text(request_payload.get("specialty"))
                level = str(request_payload.get("level") or "").strip()
                default_start = ""
                default_end = ""
                source_progress = ""
            elif source_work_type == WORK_TYPE_MAINTENANCE:
                record = self._find_record_by_id(record_id, WORK_TYPE_MAINTENANCE)
                source_record_id = record_id
                fields = record.get("display_fields") or {}
                source_progress = self._maintenance_status_value(record)
                if not self._maintenance_status_is_startable(record):
                    existing_target_record_id = (
                        self._target_record_id_from_identity_map(
                            work_type=WORK_TYPE_CHANGE,
                            source_record_id=source_record_id,
                        )
                        or self._target_record_id_from_work_status(
                            work_type=WORK_TYPE_CHANGE,
                            source_record_id=source_record_id,
                        )
                    )
                    if existing_target_record_id:
                        raise PortalError(
                            "该维保转变更已创建目标记录，请从“已开始未结束”发送更新。"
                        )
                    if "进行中" not in source_progress:
                        raise PortalError(
                            f"该维保源记录当前状态不能发起变更通告: {source_progress or '-'}"
                        )
                title = (
                    str(request_payload.get("title") or "").strip()
                    or self._maintenance_title(record)
                )
                paired_maintenance_original_title = (
                    paired_maintenance_original_title or self._maintenance_title(record)
                )
                source_codes = self._building_codes_from_value(fields.get("楼栋"))
                building_codes = self._resolve_scoped_source_building_codes(
                    scope=scope,
                    work_type=WORK_TYPE_MAINTENANCE,
                    record_id=record_id,
                    source_codes=source_codes,
                    payload_codes=self._building_codes_from_request_payload(
                        request_payload
                    ),
                )
                building = (
                    str(request_payload.get("building") or "").strip()
                    or self._building_label_from_codes(building_codes)
                )
                specialty = (
                    self._clean_source_text(request_payload.get("specialty"))
                    or self._clean_source_text(fields.get("专业类别"))
                )
                maintenance_cycle = str(fields.get("维护周期") or maintenance_cycle).strip()
                level = str(request_payload.get("level") or "").strip()
                default_start = ""
                default_end = ""
            else:
                record = self._find_record_by_id(record_id, WORK_TYPE_CHANGE)
                source_record_id = record_id
                fields = record.get("display_fields") or {}
                source_progress = self._change_progress_value(record)
                if source_progress != CHANGE_PROGRESS_NOT_STARTED:
                    raise PortalError(
                        f"该变更当前源进度不是{CHANGE_PROGRESS_NOT_STARTED}: {source_progress or '-'}"
                    )
                title = (
                    str(request_payload.get("title") or "").strip()
                    or self._change_title(record)
                )
                building_codes = self._resolve_scoped_source_building_codes(
                    scope=scope,
                    work_type=WORK_TYPE_CHANGE,
                    record_id=record_id,
                    source_codes=self._change_record_building_codes(record),
                    payload_codes=self._building_codes_from_request_payload(
                        request_payload
                    ),
                )
                building = self._building_label_from_codes(building_codes)
                specialty = self._change_specialty(record)
                level = str(fields.get("变更等级（阿里）") or "").strip()
                default_start, default_end, _ = self._change_time_range(record)
        else:
            active_item_id = str(request_payload.get("active_item_id") or "").strip()
            source_record = None
            if source_record_id:
                try:
                    source_record = self._find_record_by_id(
                        source_record_id, source_work_type
                    )
                    fields = source_record.get("display_fields") or {}
                except PortalError:
                    source_record = None
                    fields = {}
            title = str(request_payload.get("title") or "").strip()
            if not title and source_record is not None:
                title = (
                    self._maintenance_title(source_record)
                    if source_work_type == WORK_TYPE_MAINTENANCE
                    else self._change_title(source_record)
                )
            if source_work_type == WORK_TYPE_MAINTENANCE:
                if source_record is not None:
                    paired_maintenance_original_title = (
                        paired_maintenance_original_title
                        or self._maintenance_title(source_record)
                    )
                    maintenance_cycle = str(
                        fields.get("维护周期") or maintenance_cycle
                    ).strip()
                else:
                    paired_maintenance_original_title = (
                        paired_maintenance_original_title or title
                    )
            if not title:
                raise PortalError("更新/结束通告缺少名称。")
            payload_building_codes = self._building_codes_from_request_payload(
                request_payload
            )
            source_building_codes = (
                (
                    self._building_codes_from_value(fields.get("楼栋"))
                    if source_work_type == WORK_TYPE_MAINTENANCE
                    else self._change_record_building_codes(source_record)
                )
                if source_record is not None
                else []
            )
            building_codes = payload_building_codes or source_building_codes
            if source_record is not None:
                building_codes = self._resolve_scoped_source_building_codes(
                    scope=scope,
                    work_type=source_work_type,
                    record_id=source_record_id,
                    source_codes=source_building_codes,
                    payload_codes=payload_building_codes,
                )
            if not building_codes:
                building_codes = self._building_codes_from_value(request_payload.get("building"))
            building = (
                str(request_payload.get("building") or "").strip()
                or self._building_label_from_codes(building_codes)
            )
            specialty = self._clean_source_text(request_payload.get("specialty"))
            if not specialty and source_record is not None:
                specialty = (
                    self._clean_source_text(fields.get("专业类别"))
                    if source_work_type == WORK_TYPE_MAINTENANCE
                    else self._change_specialty(source_record)
                )
            level = str(request_payload.get("level") or "").strip()
            if not level and source_record is not None:
                level = (
                    ""
                    if source_work_type == WORK_TYPE_MAINTENANCE
                    else str(fields.get("变更等级（阿里）") or "").strip()
                )
            source_progress = str(request_payload.get("source_progress") or "").strip()
            if not source_progress and source_record is not None:
                source_progress = (
                    self._maintenance_status_value(source_record)
                    if source_work_type == WORK_TYPE_MAINTENANCE
                    else self._change_progress_value(source_record)
                )
            default_start = ""
            default_end = ""
            record_id = source_record_id
            target_record_id = self._target_record_id_from_request_payload(
                request_payload,
                source_record_id=source_record_id,
            )
            if not target_record_id:
                target_record_id = self._target_record_id_from_identity_map(
                    work_type=WORK_TYPE_CHANGE,
                    active_item_id=active_item_id,
                    source_record_id=source_record_id,
                    target_record_id=self._target_record_id_from_request_payload(
                        request_payload,
                        source_record_id=source_record_id,
                    ),
                )
            if (
                not target_record_id
                and source_record is not None
                and source_work_type == WORK_TYPE_CHANGE
            ):
                target_record_id = self._resolve_target_record_id_for_source_update(
                    work_type=WORK_TYPE_CHANGE,
                    notice_type=NOTICE_TYPE_CHANGE,
                    source_record=source_record,
                )
            if not target_record_id and active_item_id:
                target_record_id = self._target_record_id_from_work_status(
                    work_type=WORK_TYPE_CHANGE,
                    active_item_id=active_item_id,
                )
            if not target_record_id:
                target_record_id = self._target_record_id_from_unique_candidate(
                    work_type=WORK_TYPE_CHANGE,
                    scope=scope,
                    title=title,
                    start_time=self._format_input_datetime(request_payload.get("start_time")),
                    end_time=self._format_input_datetime(request_payload.get("end_time")),
                    action=action,
                )
            if not target_record_id:
                target_record_id = self._target_record_id_from_request_payload(
                    request_payload,
                    source_record_id=source_record_id,
                )

        building, building_codes = self._resolve_notice_submit_building(
            scope=scope,
            request_payload=request_payload,
            building=building,
            building_codes=building_codes,
            title=title,
            work_type_label="变更通告",
            allow_scope_fallback=manual,
        )
        if action != "start" and source_progress == CHANGE_PROGRESS_ENDED:
            raise PortalError("该变更当前源进度已结束，不能更新/结束。")
        if action != "start" and not target_record_id:
            raise PortalError("该变更缺少目标变更通告表 record_id，不能更新/结束。")
        title = self._normalize_110_station_notice_title(
            title,
            building=building,
            building_codes=building_codes,
        )
        if source_work_type == WORK_TYPE_MAINTENANCE:
            paired_maintenance_original_title = self._normalize_110_station_notice_title(
                paired_maintenance_original_title or title,
                building=building,
                building_codes=building_codes,
            )
        level = self._default_change_level(level)
        zhihang_involved = self._truthy_flag(request_payload.get("zhihang_involved"))
        zhihang_record_id = str(request_payload.get("zhihang_record_id") or "").strip()
        zhihang_title = str(request_payload.get("zhihang_title") or "").strip()
        zhihang_progress = str(request_payload.get("zhihang_progress") or "").strip()
        if action == "start" and zhihang_involved:
            if not zhihang_record_id:
                raise PortalError("该变更已标记涉及智航，请选择右侧智航侧变更记录。")
            zhihang_record = self._find_zhihang_change_record_by_id(zhihang_record_id)
            zhihang_progress = self._zhihang_change_progress(zhihang_record)
            if zhihang_progress == ZHIHANG_PROGRESS_ENDED:
                raise PortalError("所选智航侧变更已结束，不能绑定。")
            if not self._scope_matches_zhihang_change_record(scope, zhihang_record):
                raise PortalError("所选智航侧变更与当前入口不匹配。")
            if zhihang_record_id in self._linked_zhihang_record_ids():
                raise PortalError("所选智航侧变更已绑定到其他阿里侧变更。")
            zhihang_title = self._zhihang_change_title(zhihang_record)

        start_time = (
            self._format_input_datetime(request_payload.get("start_time"))
            or default_start
        )
        end_time = (
            self._format_input_datetime(request_payload.get("end_time"))
            or default_end
        )
        if not start_time or not end_time:
            raise PortalError("开始时间和结束时间不能为空。")
        self._validate_minimum_notice_duration(start_time, end_time)
        self._require_end_site_photo_cumulative(
            request_payload,
            action,
            notice_type=NOTICE_TYPE_CHANGE,
            work_type=WORK_TYPE_CHANGE,
        )
        location = str(request_payload.get("location") or "").strip()
        content = str(request_payload.get("content") or "").strip() or title
        reason = str(request_payload.get("reason") or "").strip()
        impact = str(request_payload.get("impact") or "").strip() or DEFAULT_IMPACT_TEXT
        progress = (
            str(request_payload.get("progress") or "").strip()
            or (DEFAULT_PROGRESS_TEXT if action == "start" else "")
        )
        if action == "start":
            self._remember_draft_fields(
                work_type=WORK_TYPE_CHANGE,
                building=building,
                maintenance_total=title,
                item_name=title,
                location=location,
                content=content,
                reason=reason,
                impact=impact,
                extra_fields={
                    "specialty": specialty,
                    "level": level,
                    "zhihang_involved": "1" if zhihang_involved else "",
                    "zhihang_record_id": zhihang_record_id,
                    "zhihang_title": zhihang_title,
                    "zhihang_progress": zhihang_progress,
                },
            )
        building_code = "CAMPUS" if len(building_codes) >= 2 else (building_codes[0] if building_codes else "")
        text = self.build_change_notice_text(
            status=status,
            title=title,
            level=level,
            start_time=start_time,
            end_time=end_time,
            location=location,
            content=content,
            reason=reason,
            impact=impact,
            progress=progress,
        )
        _, recipients, recipient_error = self._recipients_for_building_codes(
            building_codes,
            fallback_building=building,
        )
        if recipient_error:
            raise PortalError(recipient_error)
        skip_personal_message = False
        response_time = self._action_response_time(request_payload)
        sync_maintenance_target = self._sync_maintenance_target_requested(
            request_payload,
            source_work_type=source_work_type,
        )
        paired_maintenance_upload: dict[str, Any] = {}
        paired_maintenance_actual_start_time = str(
            request_payload.get("paired_maintenance_actual_start_time") or ""
        ).strip()
        if sync_maintenance_target:
            paired_maintenance_upload = self._build_paired_maintenance_upload(
                request_payload=request_payload,
                job_id=job_id,
                action=action,
                status=status,
                source_record_id=source_record_id or record_id,
                target_record_id=paired_maintenance_target_record_id,
                active_item_id=str(request_payload.get("active_item_id") or "").strip(),
                original_title=paired_maintenance_original_title or title,
                building=building,
                building_codes=building_codes,
                specialty=specialty,
                maintenance_cycle=maintenance_cycle,
                start_time=start_time,
                end_time=end_time,
                location=location,
                content=content,
                reason=reason,
                impact=impact,
                progress=progress,
                response_time=response_time,
            )
            paired_maintenance_actual_start_time = str(
                paired_maintenance_upload.get("paired_maintenance_actual_start_time")
                or paired_maintenance_actual_start_time
            ).strip()
        return {
            "job_id": job_id,
            "work_type": WORK_TYPE_CHANGE,
            "notice_type": NOTICE_TYPE_CHANGE,
            "source_work_type": source_work_type,
            "source_app_token": (
                self.app_token
                if (record_id and not manual and source_work_type == WORK_TYPE_MAINTENANCE)
                else CHANGE_SOURCE_APP_TOKEN
                if (record_id and not manual)
                else ""
            ),
            "source_table_id": (
                self.table_id
                if (record_id and not manual and source_work_type == WORK_TYPE_MAINTENANCE)
                else CHANGE_SOURCE_TABLE_ID
                if (record_id and not manual)
                else ""
            ),
            "target_app_token": str(config.app_token or "").strip(),
            "target_table_id": str(config.get_table_id(NOTICE_TYPE_CHANGE) or "").strip(),
            "manual": manual,
            "action": action,
            "status": status,
            "scope": scope,
            "scope_label": self._scope_label(scope),
            "record_id": record_id,
            "source_record_id": source_record_id,
            "target_record_id": target_record_id,
            "active_item_id": str(request_payload.get("active_item_id") or "").strip(),
            "title": title,
            "building": building,
            "building_code": building_code,
            "building_codes": building_codes,
            "target_building": self._building_label_from_code(building_code),
            "specialty": specialty,
            "level": level,
            "source_progress": source_progress,
            "zhihang_involved": zhihang_involved,
            "zhihang_record_id": zhihang_record_id,
            "zhihang_title": zhihang_title,
            "zhihang_progress": zhihang_progress,
            "zhihang_source_app_token": ZHIHANG_CHANGE_APP_TOKEN if zhihang_record_id else "",
            "zhihang_source_table_id": ZHIHANG_CHANGE_TABLE_ID if zhihang_record_id else "",
            "sync_maintenance_target": sync_maintenance_target,
            "paired_maintenance_target_record_id": paired_maintenance_target_record_id,
            "paired_maintenance_original_title": paired_maintenance_original_title,
            "paired_maintenance_actual_start_time": paired_maintenance_actual_start_time,
            "paired_maintenance_upload": paired_maintenance_upload,
            "paired_upload_status": "pending" if paired_maintenance_upload else "skipped",
            "start_time": start_time,
            "end_time": end_time,
            "location": location,
            "content": content,
            "reason": reason,
            "impact": impact,
            "progress": progress,
            "text": text,
            "extra_images": self._site_photo_payload(request_payload),
            "recipients": recipients,
            "skip_personal_message": skip_personal_message,
            "response_time": response_time,
            "operation_id": str(request_payload.get("operation_id") or "").strip()
            or job_id,
        }

    @staticmethod
    def _combine_title_supplement(title: Any, supplement: Any) -> str:
        base = str(title or "").strip()
        extra = str(supplement or "").strip()
        if not base:
            return extra
        if not extra or extra in base:
            return base
        return f"{base}{extra}"

    @staticmethod
    def build_repair_notice_text(
        *,
        status: str,
        title: str,
        specialty: str,
        level: str,
        start_time: str,
        end_time: str,
        location: str,
        content: str,
        reason: str,
        impact: str,
        progress: str,
        repair_device: str,
        repair_fault: str,
        fault_type: str,
        repair_mode: str,
        discovery: str,
        symptom: str,
        solution: str,
        spare_parts: str,
        fault_time: str,
        expected_time: str,
    ) -> str:
        progress_text = str(progress or "").strip()
        reason_text = str(reason or "").strip()
        impact_text = str(impact or "").strip()
        title_text = MaintenancePortalService._combine_title_supplement(title, content)
        return MaintenancePortalService._render_notice_text(
            work_type=WORK_TYPE_REPAIR,
            status=status,
            values={
                "title": title_text,
                "location": location,
                "level": level,
                "specialty": specialty,
                "fault_time": MaintenancePortalService._format_input_datetime(fault_time),
                "expected_time": MaintenancePortalService._format_input_datetime(expected_time),
                "repair_device": repair_device,
                "repair_fault": repair_fault,
                "fault_type": fault_type,
                "repair_mode": repair_mode,
                "impact": impact_text,
                "discovery": discovery,
                "symptom": symptom,
                "reason": reason_text,
                "solution": solution,
                "spare_parts": spare_parts,
                "progress": progress_text,
            },
        )

    def prepare_repair_action(
        self, request_payload: dict[str, Any], *, job_id: str
    ) -> dict[str, Any]:
        request_payload = normalize_notice_identity_payload(request_payload)
        request_payload = self._normalize_repair_time_fields(request_payload)
        action = str(request_payload.get("action") or "").strip().lower()
        status_map = {"start": "开始", "update": "更新", "end": "结束"}
        status = status_map.get(action)
        if not status:
            raise PortalError("动作必须是 start/update/end。")
        self.ensure_snapshot_loaded()
        scope = self._normalize_scope(request_payload.get("scope"))

        def request_text(name: str, default: str = "") -> str:
            if name in request_payload:
                return self._clean_source_text(request_payload.get(name))
            return self._clean_source_text(default)

        def request_text_keep_none_word(name: str, default: str = "") -> str:
            if name in request_payload:
                return str(request_payload.get(name) or "").strip()
            return str(default or "").strip()

        def request_first_text(names: tuple[str, ...], default: str = "") -> str:
            for name in names:
                if name in request_payload:
                    value = str(request_payload.get(name) or "").strip()
                    if value:
                        return value
            return str(default or "").strip()

        record_id = str(request_payload.get("record_id") or "").strip()
        source_record_id = str(request_payload.get("source_record_id") or "").strip()
        manual = self._truthy_flag(request_payload.get("manual"))
        if manual and not record_id:
            record_id = str(request_payload.get("manual_id") or "").strip()
        fields: dict[str, Any] = {}
        record = None
        source_record = None
        if action == "start":
            if not record_id:
                raise PortalError("开始通告缺少检修源记录ID。")
            if manual:
                title = request_first_text(("title", "content"))
                if not title:
                    raise PortalError("纯手填检修通告缺少标题。")
                building_codes = [
                    str(code or "").strip().upper()
                    for code in (request_payload.get("building_codes") or [])
                    if str(code or "").strip()
                ]
                if not building_codes:
                    building_codes = self._repair_building_codes_from_value(
                        request_payload.get("building")
                    )
                building = (
                    str(request_payload.get("building") or "").strip()
                    or self._building_label_from_codes(building_codes)
                )
                specialty = self._clean_source_text(request_payload.get("specialty"))
                level = request_text("level")
                default_start = ""
                default_end = ""
                repair_device = request_text("repair_device")
                repair_fault = request_text("repair_fault")
                fault_type = request_text("fault_type")
                repair_mode = request_text("repair_mode")
                discovery = request_text("discovery")
                symptom = request_text("symptom")
                solution = request_text("solution")
                spare_parts = request_text_keep_none_word("spare_parts")
                fault_time = request_text("fault_time")
                target_record_id = ""
            else:
                try:
                    record = self._find_record_by_id(record_id, WORK_TYPE_REPAIR)
                except PortalError as exc:
                    record = None
                    fields = {}
                    title = request_first_text(("title", "content"))
                    building_codes = [
                        str(code or "").strip().upper()
                        for code in (request_payload.get("building_codes") or [])
                        if str(code or "").strip()
                    ]
                    if not building_codes:
                        building_codes = self._repair_building_codes_from_value(
                            request_payload.get("building")
                        )
                    if not title or not building_codes:
                        raise PortalError(
                            "未找到检修源记录，且前端字段不足，不能发起检修通告。"
                        ) from exc
                    source_progress = request_text(
                        "source_progress", DEFAULT_MAINTENANCE_STATUS
                    ) or DEFAULT_MAINTENANCE_STATUS
                    if source_progress != DEFAULT_MAINTENANCE_STATUS:
                        raise PortalError(
                            f"该检修当前源状态不是{DEFAULT_MAINTENANCE_STATUS}: {source_progress or '-'}"
                        ) from exc
                    building = (
                        str(request_payload.get("building") or "").strip()
                        or self._building_label_from_codes(building_codes)
                    )
                    specialty = self._clean_source_text(request_payload.get("specialty"))
                    level = request_text("level")
                    default_start = ""
                    default_end = ""
                    repair_device = request_text("repair_device")
                    repair_fault = request_text("repair_fault")
                    fault_type = request_text("fault_type") or "设备故障"
                    repair_mode = request_text("repair_mode")
                    discovery = request_text("discovery")
                    symptom = request_text("symptom")
                    solution = request_text("solution")
                    spare_parts = request_text_keep_none_word("spare_parts")
                    fault_time = request_text("fault_time")
                    target_record_id = ""
                else:
                    fields = record.get("display_fields") or {}
                    source_record_id = record_id
                    if not self._is_valid_repair_record(record):
                        raise PortalError("该检修记录缺少有效检修通告名称/维修名称或楼栋，不能发起。")
                    source_progress = self._repair_source_status(record)
                    if source_progress != DEFAULT_MAINTENANCE_STATUS:
                        raise PortalError(
                            f"该检修当前源状态不是{DEFAULT_MAINTENANCE_STATUS}: {source_progress or '-'}"
                        )
                    title = request_first_text(
                        ("title", "content"), default=self._repair_title(record)
                    )
                    building_codes = self._repair_record_building_codes(record)
                    building = self._building_label_from_codes(building_codes)
                    specialty = self._clean_source_text(request_payload.get("specialty"))
                    if not specialty or specialty == "全部":
                        specialty = self._repair_specialty(record)
                    level = request_text("level", self._repair_level(fields))
                    default_start, default_end, _ = self._repair_time_range(record)
                    repair_device = request_text("repair_device", self._repair_device_text(fields))
                    repair_fault = request_text(
                        "repair_fault",
                        self._repair_first_field(fields, "维修故障", "故障维修原因"),
                    )
                    fault_type = request_text(
                        "fault_type", self._repair_first_field(fields, "故障类型") or "设备故障"
                    )
                    repair_mode = request_text(
                        "repair_mode",
                        self._repair_first_field(fields, "维修方式", "维修方", "供应商名称"),
                    )
                    discovery = request_text(
                        "discovery", self._repair_first_field(fields, "对应来源")
                    )
                    symptom = request_text(
                        "symptom", self._repair_first_field(fields, "故障发生现象描述", "故障现象")
                    )
                    solution = request_text(
                        "solution",
                        self._repair_first_field(fields, "解决方案", "维修方案", "后续整改措施"),
                    )
                    spare_parts = request_text(
                        "spare_parts",
                        self._repair_first_field(fields, "备件更换情况", "备件使用情况"),
                    )
                    spare_parts = request_text_keep_none_word(
                        "spare_parts",
                        spare_parts,
                    )
                    source_fault_time = self._repair_first_field(
                        fields, "故障发生时间", "发现故障时间"
                    )
                    fault_time = request_text("fault_time", source_fault_time)
                    target_record_id = self._repair_target_record_id(record)
        else:
            source_record_id = str(request_payload.get("source_record_id") or "").strip()
            active_item_id = str(request_payload.get("active_item_id") or "").strip()
            if source_record_id:
                try:
                    source_record = self._find_record_by_id(source_record_id, WORK_TYPE_REPAIR)
                    fields = source_record.get("display_fields") or {}
                except PortalError:
                    source_record = None
                    fields = {}
            title = str(request_payload.get("title") or "").strip()
            if not title and source_record is not None:
                title = self._repair_title(source_record)
            if not title:
                raise PortalError("更新/结束通告缺少名称。")
            building_codes = [
                str(code or "").strip().upper()
                for code in (request_payload.get("building_codes") or [])
                if str(code or "").strip()
            ]
            if not building_codes and source_record is not None:
                building_codes = self._repair_record_building_codes(source_record)
            if not building_codes:
                building_codes = self._repair_building_codes_from_value(
                    request_payload.get("building")
                )
            building = (
                str(request_payload.get("building") or "").strip()
                or self._building_label_from_codes(building_codes)
            )
            specialty = self._clean_source_text(request_payload.get("specialty"))
            if not specialty and source_record is not None:
                specialty = self._repair_specialty(source_record)
            level = str(request_payload.get("level") or "").strip()
            if not level and source_record is not None:
                level = self._repair_level(fields)
            default_start = ""
            default_end = ""
            record_id = source_record_id
            target_record_id = self._target_record_id_from_request_payload(
                request_payload,
                source_record_id=source_record_id,
            )
            if not target_record_id:
                target_record_id = self._target_record_id_from_identity_map(
                    work_type=WORK_TYPE_REPAIR,
                    active_item_id=active_item_id,
                    source_record_id=source_record_id,
                    target_record_id=self._target_record_id_from_request_payload(
                        request_payload,
                        source_record_id=source_record_id,
                    ),
                )
            if not target_record_id and source_record is not None:
                target_record_id = self._resolve_target_record_id_for_source_update(
                    work_type=WORK_TYPE_REPAIR,
                    notice_type=NOTICE_TYPE_REPAIR,
                    source_record=source_record,
                )
            if not target_record_id and active_item_id:
                target_record_id = self._target_record_id_from_work_status(
                    work_type=WORK_TYPE_REPAIR,
                    active_item_id=active_item_id,
                )
            if not target_record_id:
                target_record_id = self._target_record_id_from_unique_candidate(
                    work_type=WORK_TYPE_REPAIR,
                    scope=scope,
                    title=title,
                    start_time=(
                        self._format_input_datetime(request_payload.get("expected_time"))
                        or self._format_input_datetime(request_payload.get("start_time"))
                    ),
                    end_time=(
                        self._format_input_datetime(request_payload.get("fault_time"))
                        or self._format_input_datetime(request_payload.get("end_time"))
                    ),
                    action=action,
                )
            if not target_record_id:
                target_record_id = self._target_record_id_from_request_payload(
                    request_payload,
                    source_record_id=source_record_id,
                )
            repair_device = request_text("repair_device")
            if not repair_device and source_record is not None:
                repair_device = self._repair_device_text(fields)
            repair_fault = request_text("repair_fault")
            if not repair_fault and source_record is not None:
                repair_fault = self._repair_first_field(fields, "维修故障", "故障维修原因")
            fault_type = (
                request_text("fault_type")
                if "fault_type" in request_payload
                else "设备故障"
            )
            if source_record is not None and fault_type == "设备故障":
                fault_type = self._repair_first_field(fields, "故障类型") or fault_type
            repair_mode = request_text("repair_mode")
            if not repair_mode and source_record is not None:
                repair_mode = self._repair_first_field(fields, "维修方式", "维修方", "供应商名称")
            discovery = request_text("discovery")
            if not discovery and source_record is not None:
                discovery = self._repair_first_field(fields, "对应来源")
            symptom = request_text("symptom")
            if not symptom and source_record is not None:
                symptom = self._repair_first_field(fields, "故障发生现象描述", "故障现象")
            solution = request_text("solution")
            if not solution and source_record is not None:
                solution = self._repair_first_field(fields, "解决方案", "维修方案", "后续整改措施")
            spare_parts = request_text_keep_none_word("spare_parts")
            if not spare_parts and source_record is not None:
                spare_parts = self._repair_first_field(fields, "备件更换情况", "备件使用情况")
            fault_time = request_text("fault_time")
            if not fault_time and source_record is not None:
                fault_time = self._repair_first_field(fields, "故障发生时间", "发现故障时间")

        building, building_codes = self._resolve_notice_submit_building(
            scope=scope,
            request_payload=request_payload,
            building=building,
            building_codes=building_codes,
            title=title,
            location=request_payload.get("location"),
            content=request_payload.get("content"),
            work_type_label="检修通告",
            allow_scope_fallback=manual,
        )
        if action != "start" and source_record is not None and self._repair_has_ended(source_record):
            raise PortalError("该检修当前源状态已结束，不能更新/结束。")
        if action != "start" and not target_record_id:
            raise PortalError("该检修缺少目标设备检修表 record_id，不能更新/结束。")

        requested_expected_time = self._format_input_datetime(
            request_payload.get("expected_time")
        )
        requested_fault_time = self._format_input_datetime(
            request_payload.get("fault_time")
        )
        requested_start_time = self._format_input_datetime(
            request_payload.get("start_time")
        )
        requested_end_time = self._format_input_datetime(request_payload.get("end_time"))
        expected_time = requested_expected_time or requested_start_time or default_end
        fault_time = requested_fault_time or requested_end_time or self._format_input_datetime(fault_time)
        start_time = expected_time or default_start
        end_time = fault_time or default_end
        now_text = dt.datetime.now().strftime("%Y-%m-%d %H:%M")
        if not start_time:
            start_time = now_text
        if not end_time and action == "end":
            end_time = now_text
        self._validate_minimum_notice_duration(
            end_time,
            start_time,
            start_label="发现故障时间",
            end_label="期望完成时间",
        )
        self._require_end_site_photo_cumulative(
            request_payload,
            action,
            notice_type=NOTICE_TYPE_REPAIR,
            work_type=WORK_TYPE_REPAIR,
        )
        location = request_text("location")
        default_content = self._repair_first_field(
            fields, "标题/补充内容", "标题补充内容"
        )
        requested_content = (
            request_text("content") if "content" in request_payload else ""
        )
        if default_content:
            content = requested_content or default_content
        else:
            content = ""
        if content and title.endswith(content):
            title = title[: -len(content)].strip() or title
        title = self._normalize_110_station_notice_title(
            title,
            building=building,
            building_codes=building_codes,
        )
        reason = request_text(
            "reason", self._repair_first_field(fields, "故障原因", "故障维修原因")
        )
        impact = request_text("impact")
        progress = request_text(
            "progress", self._repair_first_field(fields, "维修进展描述", "当前维修进度")
        )
        if action == "start":
            self._remember_draft_fields(
                work_type=WORK_TYPE_REPAIR,
                building=building,
                maintenance_total=title,
                item_name=title,
                location=location,
                content=content,
                reason=reason,
                impact=impact,
                extra_fields={
                    "specialty": specialty,
                    "level": level,
                    "repair_device": repair_device,
                    "repair_fault": repair_fault,
                    "fault_type": fault_type,
                    "repair_mode": repair_mode,
                    "discovery": discovery,
                    "symptom": symptom,
                    "solution": solution,
                    "spare_parts": spare_parts,
                },
            )
        building_code = (
            "CAMPUS"
            if len(building_codes) >= 2
            else (building_codes[0] if building_codes else "")
        )
        text = self.build_repair_notice_text(
            status=status,
            title=title,
            specialty=specialty,
            level=level,
            start_time=start_time,
            end_time=end_time,
            location=location,
            content=content,
            reason=reason,
            impact=impact,
            progress=progress,
            repair_device=repair_device,
            repair_fault=repair_fault,
            fault_type=fault_type,
            repair_mode=repair_mode,
            discovery=discovery,
            symptom=symptom,
            solution=solution,
            spare_parts=spare_parts,
            fault_time=fault_time,
            expected_time=expected_time,
        )
        _, recipients, recipient_error = self._recipients_for_building_codes(
            building_codes,
            fallback_building=building,
        )
        if recipient_error:
            raise PortalError(recipient_error)
        skip_personal_message = False
        response_time = self._action_response_time(request_payload)
        return {
            "job_id": job_id,
            "work_type": WORK_TYPE_REPAIR,
            "notice_type": NOTICE_TYPE_REPAIR,
            "source_app_token": REPAIR_SOURCE_APP_TOKEN if (record_id and not manual) else "",
            "source_table_id": REPAIR_SOURCE_TABLE_ID if (record_id and not manual) else "",
            "target_app_token": str(config.app_token or "").strip(),
            "target_table_id": str(config.get_table_id(NOTICE_TYPE_REPAIR) or "").strip(),
            "manual": manual,
            "action": action,
            "status": status,
            "scope": scope,
            "scope_label": self._scope_label(scope),
            "record_id": record_id,
            "source_record_id": source_record_id,
            "target_record_id": target_record_id,
            "active_item_id": str(request_payload.get("active_item_id") or "").strip(),
            "title": title,
            "building": building,
            "building_code": building_code,
            "building_codes": building_codes,
            "target_building": self._building_label_from_code(building_code),
            "specialty": specialty,
            "level": level,
            "source_progress": str(
                request_payload.get("source_progress")
                or (
                    self._repair_source_status(record)
                    if action == "start" and isinstance(record, dict)
                    else self._repair_source_status(source_record)
                    if isinstance(source_record, dict)
                    else ""
                )
                or self._repair_first_field(fields, "流程", "当前维修进度")
                or ""
            ),
            "start_time": start_time,
            "end_time": end_time,
            "location": location,
            "content": content,
            "reason": reason,
            "impact": impact,
            "progress": progress,
            "repair_device": repair_device,
            "repair_fault": repair_fault,
            "fault_type": fault_type,
            "repair_mode": repair_mode,
            "discovery": discovery,
            "symptom": symptom,
            "solution": solution,
            "spare_parts": spare_parts,
            "fault_time": fault_time,
            "expected_time": expected_time,
            "text": text,
            "extra_images": self._site_photo_payload(request_payload),
            "recipients": recipients,
            "skip_personal_message": skip_personal_message,
            "response_time": response_time,
            "operation_id": str(request_payload.get("operation_id") or "").strip()
            or job_id,
        }

    @staticmethod
    def _normalize_repair_time_fields(payload: dict[str, Any]) -> dict[str, Any]:
        """Keep repair-specific time aliases aligned with the web form fields.

        The lightweight workbench stores repair "发现故障时间" in end_time and
        "期望完成时间" in start_time. The backend also carries explicit
        fault_time/expected_time for source-table and Qt projections. Normalize
        once at the service boundary so generated text and bitable payloads
        always use the latest user-edited values.
        """
        normalized = dict(payload or {})
        start_time = str(normalized.get("start_time") or "").strip()
        end_time = str(normalized.get("end_time") or "").strip()
        fault_time = str(normalized.get("fault_time") or "").strip()
        expected_time = str(normalized.get("expected_time") or "").strip()
        if end_time:
            normalized["fault_time"] = end_time
        elif fault_time:
            normalized["end_time"] = fault_time
        if start_time:
            normalized["expected_time"] = start_time
        elif expected_time:
            normalized["start_time"] = expected_time
        return normalized

    def send_action_personal_message(self, prepared: dict[str, Any]) -> tuple[bool, str]:
        if (prepared or {}).get("skip_personal_message"):
            return True, "无需发送个人消息"
        text = str((prepared or {}).get("text") or "").strip()
        recipients = list(prepared.get("recipients") or [])
        if not text:
            return False, "通告文本为空。"
        if not recipients:
            return False, "缺少飞书接收人。"
        job_id = str((prepared or {}).get("job_id") or "")
        message_signature = str((prepared or {}).get("message_signature") or "")
        job = self.get_job(job_id) or {}
        prior_results = []
        if str(job.get("message_signature") or "") == message_signature:
            prior_results = [
                item
                for item in (job.get("recipient_results") or [])
                if isinstance(item, dict)
            ]
        succeeded = {
            str(item.get("open_id") or "").strip()
            for item in prior_results
            if item.get("ok") and str(item.get("open_id") or "").strip()
        }
        pending_recipients = [
            str(open_id or "").strip()
            for open_id in recipients
            if str(open_id or "").strip() and str(open_id or "").strip() not in succeeded
        ]
        if not pending_recipients:
            self.mark_job(job_id, recipient_results=prior_results)
            return True, "ok"
        guard = external_real_write_guard()
        if guard["mock_external"]:
            new_results = [
                {"open_id": open_id, "ok": True, "message": "mock external send skipped"}
                for open_id in pending_recipients
            ]
            ok = True
            message = "mock external send skipped"
        elif not guard["real_write_allowed"]:
            return False, str(guard["reason"] or "真实外部写入未确认。")
        else:
            ok, message, new_results = send_text_to_open_ids(text, pending_recipients)
        by_open_id = {
            str(item.get("open_id") or "").strip(): dict(item)
            for item in prior_results
            if str(item.get("open_id") or "").strip()
        }
        for item in new_results:
            open_id = str(item.get("open_id") or "").strip()
            if open_id:
                by_open_id[open_id] = dict(item)
        recipient_results = [by_open_id.get(str(open_id).strip()) for open_id in recipients]
        recipient_results = [item for item in recipient_results if item]
        ok = ok and all(bool(item.get("ok")) for item in recipient_results)
        self.mark_job(
            job_id,
            recipient_results=recipient_results,
        )
        return ok, message

    def _validate_generated_item(
        self, item: dict[str, Any]
    ) -> tuple[dict[str, Any] | None, str, str, str, str]:
        record_id = str(item.get("record_id") or "").strip()
        text = str(item.get("text") or "").strip()
        if not record_id:
            return None, "", "", "", "缺少 record_id"
        if not text:
            return None, record_id, "", "", "缺少生成后的通告文本"
        if not text.startswith("【维保通告】状态：开始"):
            return None, record_id, "", "", "通告文本不是维保开始模板"
        record = self._find_record_by_id(record_id)
        fields = record.get("display_fields") or {}
        building = str(fields.get("楼栋") or "").strip()
        maintenance_total = str(fields.get("维护总项") or "").strip()
        expected_name = f"EA118机房{building}{maintenance_total}"
        if expected_name and expected_name not in text:
            return None, record_id, building, expected_name, "通告名称与多维记录不一致"
        return record, record_id, building, expected_name, ""

    def send_generated_templates(
        self,
        items: list[dict[str, Any]],
        *,
        notice_callback: Callable[[dict[str, Any]], Any] | None = None,
    ) -> list[dict[str, Any]]:
        self.ensure_snapshot_loaded()
        if notice_callback is None:
            raise PortalError("主窗口未连接，无法发送后回填主界面。")
        if not isinstance(items, list) or not items:
            raise PortalError("缺少待发送通告。")

        results: list[dict[str, Any]] = []
        for item in items:
            base_result = {
                "record_id": str((item or {}).get("record_id") or "").strip(),
                "title": str((item or {}).get("title") or "").strip(),
                "text": str((item or {}).get("text") or "").strip(),
                "building": "",
                "target_building": "",
                "ok": False,
                "error": "",
                "recipients": [],
            }
            try:
                record, record_id, building, expected_name, error = (
                    self._validate_generated_item(item or {})
                )
                base_result["record_id"] = record_id
                base_result["building"] = building
                base_result["title"] = expected_name or base_result["title"]
                if error:
                    base_result["error"] = error
                    results.append(base_result)
                    continue

                building_code, recipients, recipient_error = self._recipients_for_building(
                    building
                )
                base_result["target_building"] = f"{building_code}楼" if building_code else ""
                base_result["recipients"] = recipients
                if recipient_error:
                    base_result["error"] = recipient_error
                    results.append(base_result)
                    continue

                guard = external_real_write_guard()
                if guard["mock_external"]:
                    ok, message, recipient_results = True, "mock external send skipped", [
                        {
                            "open_id": str(open_id or "").strip(),
                            "ok": True,
                            "message": "mock external send skipped",
                        }
                        for open_id in recipients
                        if str(open_id or "").strip()
                    ]
                elif not guard["real_write_allowed"]:
                    ok, message, recipient_results = False, str(
                        guard["reason"] or "真实外部写入未确认。"
                    ), []
                else:
                    ok, message, recipient_results = send_text_to_open_ids(
                        base_result["text"], recipients
                    )
                base_result["recipient_results"] = recipient_results
                if not ok:
                    base_result["error"] = message
                    results.append(base_result)
                    continue

                notice_callback(
                    {
                        "source": "lan_template_portal",
                        "record_id": record_id,
                        "title": base_result["title"],
                        "building": building,
                        "target_building": base_result["target_building"],
                        "text": base_result["text"],
                    }
                )
                base_result["ok"] = True
                results.append(base_result)
            except Exception as exc:
                base_result["error"] = str(exc)
                results.append(base_result)
        return results
