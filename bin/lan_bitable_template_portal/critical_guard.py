# -*- coding: utf-8 -*-
from __future__ import annotations

import datetime as dt
import hashlib
import io
import math
import re
import shutil
import zipfile
import xml.etree.ElementTree as ET
from copy import copy
from contextlib import suppress
from functools import lru_cache
from pathlib import Path
from typing import Any


CRITICAL_GUARD_TEMPLATE_NAME = "重保戒备检查表.xlsx"
CRITICAL_GUARD_SCOPE_CODES = ("A", "B", "C", "D", "E")
CRITICAL_GUARD_SCOPE_LABELS = {code: f"{code}楼" for code in CRITICAL_GUARD_SCOPE_CODES}
CRITICAL_GUARD_SHEET_NAMES = (
    "设备安全",
    "环境安全",
    "客户重保",
    "灾害专项",
    "物资检查清单",
    "重保联络清单",
)
CRITICAL_GUARD_CHECK_SHEETS = frozenset(
    {"设备安全", "环境安全", "客户重保", "灾害专项"}
)
CRITICAL_GUARD_FILE_SHEETS = frozenset({"物资检查清单", "重保联络清单"})

CRITICAL_GUARD_MAX_XLSX_ENTRIES = 4096
CRITICAL_GUARD_MAX_XLSX_EXPANDED_BYTES = 128 * 1024 * 1024
CRITICAL_GUARD_MAX_SOURCE_ROWS = 500
CRITICAL_GUARD_MAX_SOURCE_COLUMNS = 80
CRITICAL_GUARD_MAX_RENDER_PIXELS = 48_000_000
CRITICAL_GUARD_MAX_RENDER_DIMENSION = 20_000
CRITICAL_GUARD_MIN_RENDER_SCALE = 1.0

_CHECK_SHEET_RULES = {
    "设备安全": {
        "start": 5,
        "end": 83,
        "category_cols": (2,),
        "content_col": 3,
        "result_col": 4,
        "note_col": 5,
        "suggestions_row": 84,
    },
    "环境安全": {
        "start": 5,
        "end": 98,
        "category_cols": (2,),
        "content_col": 3,
        "result_col": 4,
        "note_col": 5,
        "suggestions_row": 99,
    },
    "客户重保": {
        "start": 5,
        "end": 34,
        "category_cols": (2,),
        "content_col": 3,
        "result_col": 4,
        "note_col": 5,
        "suggestions_row": 35,
    },
    "灾害专项": {
        "start": 8,
        "end": 28,
        "category_cols": (2, 3),
        "content_col": 4,
        "result_col": 5,
        "note_col": 6,
        "suggestions_row": 30,
    },
}


class CriticalGuardError(RuntimeError):
    pass


def critical_guard_template_path() -> Path:
    return Path(__file__).resolve().parent / "templates" / CRITICAL_GUARD_TEMPLATE_NAME


def normalize_scope(value: Any) -> str:
    text = str(value or "").strip().upper()
    match = re.search(r"[ABCDE]", text)
    code = match.group(0) if match else ""
    if code not in CRITICAL_GUARD_SCOPE_CODES:
        raise CriticalGuardError("重保管理仅支持 A、B、C、D、E 楼。")
    return code


def normalize_sheet_name(value: Any) -> str:
    text = str(value or "").strip()
    if text == "中报联络清单":
        text = "重保联络清单"
    if text not in CRITICAL_GUARD_SHEET_NAMES:
        raise CriticalGuardError(f"不支持的重保检查表：{text or '未选择'}")
    return text


def _resolve_workbook_sheet_title(workbook: Any, sheet_name: str) -> str:
    """Resolve an uploaded sheet while tolerating accidental outer spaces."""
    normalized_sheet = normalize_sheet_name(sheet_name)
    sheet_names = [str(item) for item in list(workbook.sheetnames or [])]
    if normalized_sheet in sheet_names:
        return normalized_sheet
    matches = [item for item in sheet_names if item.strip() == normalized_sheet]
    if len(matches) == 1:
        return matches[0]
    if len(matches) > 1:
        raise CriticalGuardError(
            f"上传文件中存在多个名称近似“{normalized_sheet}”的工作表，请只保留一个。"
        )
    raise CriticalGuardError(f"上传文件缺少“{normalized_sheet}”工作表。")


def normalize_task_memory_key(value: Any) -> str:
    text = re.sub(r"\s+", "", str(value or "").strip()).lower()
    text = re.sub(r"[，,。；;：:、\-—_（）()【】\[\]]+", "", text)
    return text[:240]


def safe_path_part(value: Any, fallback: str = "item") -> str:
    text = re.sub(r"[\\/:*?\"<>|\r\n]+", "_", str(value or "").strip())
    text = re.sub(r"\s+", " ", text).strip(" ._")
    return (text or fallback)[:120]


def _merged_value(ws: Any, row: int, col: int) -> str:
    cell = ws.cell(row=row, column=col)
    if cell.value not in (None, ""):
        return str(cell.value).strip()
    for merged in ws.merged_cells.ranges:
        if merged.min_row <= row <= merged.max_row and merged.min_col <= col <= merged.max_col:
            value = ws.cell(row=merged.min_row, column=merged.min_col).value
            return str(value or "").strip()
    return ""


def _template_sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


@lru_cache(maxsize=1)
def critical_guard_catalog() -> dict[str, Any]:
    path = critical_guard_template_path()
    if not path.exists():
        raise CriticalGuardError(f"重保检查模板不存在：{path}")
    try:
        from openpyxl import load_workbook
    except Exception as exc:  # pragma: no cover - startup dependency guard
        raise CriticalGuardError("缺少 openpyxl，无法读取重保检查模板。") from exc

    workbook = load_workbook(
        path,
        data_only=False,
        read_only=False,
        keep_links=False,
    )
    sheets: list[dict[str, Any]] = []
    for sheet_name in CRITICAL_GUARD_SHEET_NAMES:
        if sheet_name not in workbook.sheetnames:
            raise CriticalGuardError(f"重保模板缺少 Sheet：{sheet_name}")
        ws = workbook[sheet_name]
        if sheet_name in CRITICAL_GUARD_CHECK_SHEETS:
            rule = _CHECK_SHEET_RULES[sheet_name]
            items: list[dict[str, Any]] = []
            for row in range(int(rule["start"]), int(rule["end"]) + 1):
                content = _merged_value(ws, row, int(rule["content_col"]))
                if not content:
                    continue
                categories = [
                    _merged_value(ws, row, int(col))
                    for col in rule["category_cols"]
                ]
                category = " / ".join(item for item in categories if item)
                items.append(
                    {
                        "key": str(row),
                        "row": row,
                        "category": category,
                        "content": content,
                    }
                )
            sheet = {
                "name": sheet_name,
                "title": str(ws.cell(row=2, column=2).value or sheet_name),
                "kind": "check",
                "items": items,
                "has_signature": True,
                "has_weather": sheet_name == "灾害专项",
            }
            if sheet_name == "灾害专项":
                sheet["weather_fields"] = [
                    {"key": "level1", "label": "一级极端天气"},
                    {"key": "level2", "label": "二级极端天气"},
                    {"key": "current", "label": "当前检查极端天气"},
                ]
            sheets.append(sheet)
            continue

        if sheet_name == "物资检查清单":
            sheets.append(
                {
                    "name": sheet_name,
                    "title": str(ws["A1"].value or sheet_name),
                    "kind": "materials",
                    "input_mode": "file",
                    "has_signature": False,
                    "sections": [
                        {
                            "key": "key_spares",
                            "title": "关键备品备件清单",
                            "row_count": 5,
                            "columns": [
                                {"key": "name", "label": "物品名称"},
                                {"key": "quantity", "label": "物品数量"},
                                {"key": "location", "label": "存放地"},
                            ],
                        },
                        {
                            "key": "emergency_supplies",
                            "title": "应急物资储备清单",
                            "row_count": 5,
                            "columns": [
                                {"key": "name", "label": "物品名称"},
                                {"key": "quantity", "label": "物品数量"},
                                {"key": "location", "label": "存放地"},
                                {"key": "expiry", "label": "保质期"},
                            ],
                        },
                    ],
                }
            )
            continue

        sheets.append(
            {
                "name": sheet_name,
                "title": str(ws["A1"].value or sheet_name),
                "kind": "contacts",
                "input_mode": "file",
                "has_signature": False,
                "duty": {
                    "key": "duty",
                    "title": "7*24小时值班电话",
                    "row_count": 4,
                    "columns": [
                        {"key": "team", "label": "运维团队"},
                        {"key": "contact", "label": "紧急联系人"},
                        {"key": "phone", "label": "联系电话"},
                    ],
                },
                "groups": [
                    {"key": "onsite", "title": "现场保障组"},
                    {"key": "regional", "title": "区域/总部保障组"},
                    {"key": "backup", "title": "关键岗位备份"},
                    {"key": "property", "title": "保安保洁物业团队"},
                    {"key": "supplier", "title": "供应商/外部部门保障组"},
                ],
                "group_row_count": 5,
                "group_columns": [
                    {"key": "name", "label": "姓名"},
                    {"key": "mobile", "label": "手机号"},
                    {"key": "email", "label": "邮箱"},
                    {"key": "scope", "label": "负责范围"},
                ],
            }
        )
    return {
        "template_name": CRITICAL_GUARD_TEMPLATE_NAME,
        "template_version": _template_sha256(path)[:16],
        "sheets": sheets,
    }


def sheet_definition(sheet_name: Any) -> dict[str, Any]:
    normalized = normalize_sheet_name(sheet_name)
    for item in critical_guard_catalog()["sheets"]:
        if item["name"] == normalized:
            return item
    raise CriticalGuardError(f"重保模板缺少 Sheet：{normalized}")


def default_response_cells(sheet_name: Any, scope: Any, *, today: str = "") -> dict[str, Any]:
    definition = sheet_definition(sheet_name)
    scope_code = normalize_scope(scope)
    date_text = str(today or dt.date.today().isoformat()).strip()
    base: dict[str, Any] = {
        "machine_room": f"南通机房{scope_code}楼",
        "check_date": date_text,
    }
    if definition["kind"] == "check":
        base["checks"] = {
            str(item["key"]): {"status": "normal", "note": ""}
            for item in definition.get("items") or []
        }
        base["suggestions"] = ""
        if definition.get("has_weather"):
            base["weather"] = {"level1": "", "level2": "", "current": ""}
        return base
    if definition["kind"] == "materials":
        for section in definition.get("sections") or []:
            base[str(section["key"])] = [
                {str(column["key"]): "" for column in section["columns"]}
                for _ in range(int(section["row_count"]))
            ]
        return base
    base["duty"] = [
        {str(column["key"]): "" for column in definition["duty"]["columns"]}
        for _ in range(int(definition["duty"]["row_count"]))
    ]
    base["groups"] = {
        str(group["key"]): [
            {str(column["key"]): "" for column in definition["group_columns"]}
            for _ in range(int(definition["group_row_count"]))
        ]
        for group in definition.get("groups") or []
    }
    return base


def _clean_text(value: Any, *, limit: int = 2000) -> str:
    return str(value or "").replace("\x00", "").strip()[:limit]


def _normalize_date(value: Any) -> str:
    text = _clean_text(value, limit=20)
    try:
        return dt.date.fromisoformat(text).isoformat()
    except ValueError as exc:
        raise CriticalGuardError("检查日期格式无效。") from exc


def normalize_response_cells(
    sheet_name: Any,
    scope: Any,
    payload: Any,
    *,
    fallback: dict[str, Any] | None = None,
) -> dict[str, Any]:
    definition = sheet_definition(sheet_name)
    source = payload if isinstance(payload, dict) else {}
    result = default_response_cells(sheet_name, scope)
    if isinstance(fallback, dict):
        source = {**fallback, **source}
    result["machine_room"] = f"南通机房{normalize_scope(scope)}楼"
    result["check_date"] = _normalize_date(source.get("check_date") or result["check_date"])
    if definition.get("input_mode") == "file":
        result.update(
            {
                "source_file_id": _clean_text(source.get("source_file_id"), limit=128),
                "source_file_name": _clean_text(source.get("source_file_name"), limit=255),
                "source_file_sha256": _clean_text(
                    source.get("source_file_sha256"), limit=128
                ),
            }
        )

    if definition["kind"] == "check":
        source_checks = source.get("checks") if isinstance(source.get("checks"), dict) else {}
        checks: dict[str, dict[str, str]] = {}
        for item in definition.get("items") or []:
            key = str(item["key"])
            row = source_checks.get(key) if isinstance(source_checks.get(key), dict) else {}
            status = "abnormal" if str(row.get("status") or "").lower() == "abnormal" else "normal"
            checks[key] = {"status": status, "note": _clean_text(row.get("note"))}
        result["checks"] = checks
        result["suggestions"] = _clean_text(source.get("suggestions"), limit=5000)
        if definition.get("has_weather"):
            weather = source.get("weather") if isinstance(source.get("weather"), dict) else {}
            result["weather"] = {
                key: _clean_text(weather.get(key), limit=500)
                for key in ("level1", "level2", "current")
            }
        return result

    if definition["kind"] == "materials":
        for section in definition.get("sections") or []:
            key = str(section["key"])
            source_rows = source.get(key) if isinstance(source.get(key), list) else []
            rows: list[dict[str, str]] = []
            for index in range(int(section["row_count"])):
                row = source_rows[index] if index < len(source_rows) and isinstance(source_rows[index], dict) else {}
                rows.append(
                    {
                        str(column["key"]): _clean_text(row.get(str(column["key"])), limit=500)
                        for column in section["columns"]
                    }
                )
            result[key] = rows
        return result

    source_duty = source.get("duty") if isinstance(source.get("duty"), list) else []
    result["duty"] = []
    for index in range(int(definition["duty"]["row_count"])):
        row = source_duty[index] if index < len(source_duty) and isinstance(source_duty[index], dict) else {}
        result["duty"].append(
            {
                str(column["key"]): _clean_text(row.get(str(column["key"])), limit=500)
                for column in definition["duty"]["columns"]
            }
        )
    source_groups = source.get("groups") if isinstance(source.get("groups"), dict) else {}
    result["groups"] = {}
    for group in definition.get("groups") or []:
        key = str(group["key"])
        source_rows = source_groups.get(key) if isinstance(source_groups.get(key), list) else []
        rows = []
        for index in range(int(definition["group_row_count"])):
            row = source_rows[index] if index < len(source_rows) and isinstance(source_rows[index], dict) else {}
            rows.append(
                {
                    str(column["key"]): _clean_text(row.get(str(column["key"])), limit=500)
                    for column in definition["group_columns"]
                }
            )
        result["groups"][key] = rows
    return result


def validate_response_for_generation(
    sheet_name: Any,
    cells: dict[str, Any],
    *,
    signature_count: int = 0,
) -> None:
    definition = sheet_definition(sheet_name)
    _normalize_date(cells.get("check_date"))
    if definition["kind"] != "check":
        return
    missing_notes: list[int] = []
    checks = cells.get("checks") if isinstance(cells.get("checks"), dict) else {}
    for item in definition.get("items") or []:
        row = checks.get(str(item["key"])) if isinstance(checks.get(str(item["key"])), dict) else {}
        if row.get("status") == "abnormal" and not _clean_text(row.get("note")):
            missing_notes.append(int(item["row"]))
    if missing_notes:
        preview = "、".join(str(row) for row in missing_notes[:8])
        suffix = "等" if len(missing_notes) > 8 else ""
        raise CriticalGuardError(f"异常项必须填写备注：第 {preview} 行{suffix}。")
    if int(signature_count or 0) <= 0:
        raise CriticalGuardError("生成图片前请至少选择一名检查人签名。")


def memory_cells_for_new_task(sheet_name: Any, scope: Any, memory: Any) -> dict[str, Any]:
    source = memory if isinstance(memory, dict) else {}
    result = normalize_response_cells(sheet_name, scope, source)
    result["machine_room"] = f"南通机房{normalize_scope(scope)}楼"
    result["check_date"] = dt.date.today().isoformat()
    return result


def critical_guard_sheet_range(sheet_name: Any) -> str:
    normalized = normalize_sheet_name(sheet_name)
    if normalized == "物资检查清单":
        return "A1:H18"
    if normalized == "重保联络清单":
        return "A1:J53"
    rule = _CHECK_SHEET_RULES[normalized]
    last_column = "F" if normalized == "灾害专项" else "E"
    return f"B2:{last_column}{int(rule['suggestions_row']) + 1}"


def _validate_xlsx_archive(source_path: Path) -> None:
    if not zipfile.is_zipfile(source_path):
        raise CriticalGuardError("上传文件不是有效的 .xlsx 工作簿。")
    try:
        with zipfile.ZipFile(source_path) as archive:
            entries = archive.infolist()
            if len(entries) > CRITICAL_GUARD_MAX_XLSX_ENTRIES:
                raise CriticalGuardError("上传工作簿内部文件过多，请精简后重新上传。")
            expanded_size = 0
            for entry in entries:
                normalized_name = str(entry.filename or "").replace("\\", "/")
                path_parts = [part for part in normalized_name.split("/") if part]
                if normalized_name.startswith("/") or ".." in path_parts:
                    raise CriticalGuardError("上传工作簿包含不安全的内部路径。")
                if int(entry.flag_bits or 0) & 0x1:
                    raise CriticalGuardError("不支持加密的 Excel 工作簿。")
                expanded_size += max(0, int(entry.file_size or 0))
                if expanded_size > CRITICAL_GUARD_MAX_XLSX_EXPANDED_BYTES:
                    raise CriticalGuardError("上传工作簿解压后体积过大，请精简后重新上传。")
    except CriticalGuardError:
        raise
    except (OSError, zipfile.BadZipFile, zipfile.LargeZipFile) as exc:
        raise CriticalGuardError("上传文件损坏或不是有效的 Excel 工作簿。") from exc


def _source_sheet_range(worksheet: Any, sheet_name: str) -> str:
    from openpyxl.utils import get_column_letter, range_boundaries

    ranges = [critical_guard_sheet_range(sheet_name)]
    with suppress(Exception):
        dimension = str(worksheet.calculate_dimension() or "").replace("$", "")
        if ":" in dimension:
            ranges.append(dimension)
    with suppress(Exception):
        print_area = str(worksheet.print_area or "")
        ranges.extend(
            f"{start_col}{start_row}:{end_col}{end_row}"
            for start_col, start_row, end_col, end_row in re.findall(
                r"\$?([A-Z]{1,3})\$?(\d+):\$?([A-Z]{1,3})\$?(\d+)",
                print_area.upper(),
            )
        )

    min_col = min_row = None
    max_col = max_row = 0
    for address in ranges:
        try:
            left, top, right, bottom = range_boundaries(address)
        except (TypeError, ValueError):
            continue
        min_col = left if min_col is None else min(min_col, left)
        min_row = top if min_row is None else min(min_row, top)
        max_col = max(max_col, right)
        max_row = max(max_row, bottom)
    min_col = int(min_col or 1)
    min_row = int(min_row or 1)
    if max_row > CRITICAL_GUARD_MAX_SOURCE_ROWS or max_col > CRITICAL_GUARD_MAX_SOURCE_COLUMNS:
        raise CriticalGuardError(
            "上传工作表范围过大，最多支持 "
            f"{CRITICAL_GUARD_MAX_SOURCE_ROWS} 行、{CRITICAL_GUARD_MAX_SOURCE_COLUMNS} 列。"
        )
    return (
        f"{get_column_letter(min_col)}{min_row}:"
        f"{get_column_letter(max(1, max_col))}{max(1, max_row)}"
    )


def _file_sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _rendered_image_is_usable(path: Path) -> bool:
    if not path.is_file() or path.stat().st_size <= 0:
        return False
    try:
        from PIL import Image, ImageStat

        with Image.open(path) as image:
            if image.width < 100 or image.height < 100:
                return False
            grayscale = image.convert("L")
            extrema = grayscale.getextrema()
            deviation = float(ImageStat.Stat(grayscale).stddev[0] or 0.0)
            return bool(extrema and int(extrema[0]) < 245 and deviation > 2.0)
    except Exception:
        return False


def _compose_critical_guard_signatures(
    signatures: list[dict[str, Any]] | None,
) -> tuple[bytes, int, int, int]:
    from PIL import Image

    items = [
        item
        for item in (signatures or [])
        if isinstance(item, dict)
        and isinstance(item.get("image_bytes"), (bytes, bytearray))
    ]
    if not items:
        return b"", 0, 0, 0
    columns = min(4, len(items))
    rows = math.ceil(len(items) / columns)
    slot_width = 180
    slot_height = 64
    canvas = Image.new(
        "RGBA",
        (slot_width * columns, slot_height * rows),
        (255, 255, 255, 0),
    )
    resampling = getattr(getattr(Image, "Resampling", Image), "LANCZOS")
    for index, item in enumerate(items):
        row_index, column_index = divmod(index, columns)
        x1 = column_index * slot_width
        y1 = row_index * slot_height
        try:
            signature = Image.open(io.BytesIO(bytes(item["image_bytes"]))).convert("RGBA")
            signature.thumbnail((slot_width - 18, 58), resampling)
        except Exception as exc:
            raise CriticalGuardError("检查人签名图片不可用，请重新签名。") from exc
        canvas.alpha_composite(
            signature,
            (
                int(x1 + (slot_width - signature.width) / 2),
                int(y1 + (slot_height - signature.height) / 2),
            ),
        )
    output = io.BytesIO()
    canvas.save(output, format="PNG", optimize=True)
    return output.getvalue(), canvas.width, canvas.height, rows


def _write_value_right_of_label(ws: Any, label_text: str, value: Any) -> str:
    """Keep the template label in place and write its value in the cell to the right."""
    label_cell = None
    normalized_label = str(label_text or "").strip().rstrip("：:")
    for row in ws.iter_rows():
        for cell in row:
            cell_text = str(cell.value or "").strip().rstrip("：:")
            if cell_text == normalized_label:
                label_cell = cell
                break
        if label_cell is not None:
            break
    if label_cell is None:
        raise CriticalGuardError(f"模板缺少“{normalized_label}”单元格。")

    target_column = int(label_cell.column) + 1
    target_border = None
    merged_range = next(
        (
            merged
            for merged in ws.merged_cells.ranges
            if merged.min_row <= label_cell.row <= merged.max_row
            and merged.min_col <= label_cell.column <= merged.max_col
        ),
        None,
    )
    if merged_range is not None:
        if merged_range.min_row != merged_range.max_row:
            raise CriticalGuardError(f"模板“{normalized_label}”区域不支持跨行合并。")
        target_column = max(int(label_cell.column) + 1, int(merged_range.max_col))
        target_border = copy(ws.cell(row=label_cell.row, column=target_column).border)
        ws.unmerge_cells(str(merged_range))

    target_cell = ws.cell(row=label_cell.row, column=target_column)
    if merged_range is not None:
        target_cell._style = copy(label_cell._style)
        if target_border is not None:
            target_cell.border = target_border
        target_cell.alignment = copy(label_cell.alignment)
    label_cell.value = f"{normalized_label}："
    target_cell.value = value
    return target_cell.coordinate


def _write_check_sheet(
    ws: Any,
    *,
    sheet_name: str,
    cells: dict[str, Any],
    signatures: list[dict[str, Any]] | None,
) -> list[Any]:
    from openpyxl.drawing.image import Image as ExcelImage
    from openpyxl.utils import get_column_letter

    rule = _CHECK_SHEET_RULES[sheet_name]
    if sheet_name == "灾害专项":
        weather = cells.get("weather") if isinstance(cells.get("weather"), dict) else {}
        ws["D3"] = _clean_text(weather.get("level1"), limit=500)
        ws["D4"] = _clean_text(weather.get("level2"), limit=500)
        ws["D5"] = _clean_text(weather.get("current"), limit=500)
        ws["D6"] = _clean_text(cells.get("machine_room"), limit=80)
        _write_value_right_of_label(
            ws,
            "检查日期",
            _normalize_date(cells.get("check_date")),
        )
        signature_anchor = "D31"
        signature_row = 31
        signature_start_col = 4
        signature_end_col = 6
    else:
        ws["C3"] = _clean_text(cells.get("machine_room"), limit=80)
        _write_value_right_of_label(
            ws,
            "检查日期",
            _normalize_date(cells.get("check_date")),
        )
        signature_row = int(rule["suggestions_row"]) + 1
        signature_anchor = f"C{signature_row}"
        signature_start_col = 3
        signature_end_col = 5

    checks = cells.get("checks") if isinstance(cells.get("checks"), dict) else {}
    for item in sheet_definition(sheet_name).get("items") or []:
        row_number = int(item["row"])
        row = checks.get(str(item["key"])) if isinstance(checks.get(str(item["key"])), dict) else {}
        abnormal = str(row.get("status") or "").strip().lower() == "abnormal"
        ws.cell(row=row_number, column=int(rule["result_col"])).value = "异常" if abnormal else "正常"
        ws.cell(row=row_number, column=int(rule["note_col"])).value = _clean_text(
            row.get("note"), limit=2000
        )
    ws.cell(
        row=int(rule["suggestions_row"]),
        column=int(rule["content_col"]),
    ).value = _clean_text(cells.get("suggestions"), limit=5000)

    image_handles: list[Any] = []
    signature_png, original_width, original_height, _signature_rows = (
        _compose_critical_guard_signatures(signatures)
    )
    if signature_png:
        available_width_px = sum(
            _excel_column_width_pixels(
                ws.column_dimensions[get_column_letter(column)].width
                or float(getattr(ws.sheet_format, "defaultColWidth", None) or 8.43)
            )
            for column in range(signature_start_col, signature_end_col + 1)
        )
        # Do not enlarge a handwritten signature. Only shrink a crowded signer
        # set to the actual merged-cell width of the source template.
        signature_width_px = max(1, int(min(float(original_width), max(1.0, available_width_px - 8.0))))
        signature_height_px = max(
            1,
            int(round(signature_width_px * original_height / max(1, original_width))),
        )
        row_height_points = max(
            float(ws.row_dimensions[signature_row].height or 15),
            float(signature_height_px * 72.0 / 96.0 + 6.0),
        )
        ws.row_dimensions[signature_row].height = row_height_points
        buffer = io.BytesIO(signature_png)
        excel_image = ExcelImage(buffer)
        excel_image.width = signature_width_px
        excel_image.height = signature_height_px
        ws.add_image(excel_image, signature_anchor)
        image_handles.extend([buffer, excel_image])
    return image_handles


def _write_materials_sheet(ws: Any, cells: dict[str, Any]) -> None:
    mappings = (
        ("key_spares", 5, ("B", "D", "E"), ("name", "quantity", "location")),
        (
            "emergency_supplies",
            14,
            ("B", "D", "E", "G"),
            ("name", "quantity", "location", "expiry"),
        ),
    )
    for key, start_row, columns, fields in mappings:
        rows = cells.get(key) if isinstance(cells.get(key), list) else []
        for index in range(5):
            row = rows[index] if index < len(rows) and isinstance(rows[index], dict) else {}
            for column, field in zip(columns, fields):
                ws[f"{column}{start_row + index}"] = _clean_text(row.get(field), limit=500)


def _write_contacts_sheet(ws: Any, cells: dict[str, Any]) -> None:
    duty_rows = cells.get("duty") if isinstance(cells.get("duty"), list) else []
    for index in range(4):
        row = duty_rows[index] if index < len(duty_rows) and isinstance(duty_rows[index], dict) else {}
        excel_row = 5 + index
        ws[f"A{excel_row}"] = _clean_text(row.get("team"), limit=500)
        ws[f"C{excel_row}"] = _clean_text(row.get("contact"), limit=500)
        ws[f"E{excel_row}"] = _clean_text(row.get("phone"), limit=500)
    groups = cells.get("groups") if isinstance(cells.get("groups"), dict) else {}
    starts = {
        "onsite": 13,
        "regional": 22,
        "backup": 31,
        "property": 40,
        "supplier": 49,
    }
    for key, start_row in starts.items():
        rows = groups.get(key) if isinstance(groups.get(key), list) else []
        for index in range(5):
            row = rows[index] if index < len(rows) and isinstance(rows[index], dict) else {}
            excel_row = start_row + index
            ws[f"B{excel_row}"] = _clean_text(row.get("name"), limit=500)
            ws[f"C{excel_row}"] = _clean_text(row.get("mobile"), limit=500)
            ws[f"E{excel_row}"] = _clean_text(row.get("email"), limit=500)
            ws[f"H{excel_row}"] = _clean_text(row.get("scope"), limit=500)


def build_critical_guard_workbook(
    *,
    sheet_name: str,
    scope: str,
    cells: dict[str, Any],
    signatures: list[dict[str, Any]] | None,
    output_path: Path,
) -> dict[str, Any]:
    try:
        from openpyxl import load_workbook
    except Exception as exc:  # pragma: no cover - startup dependency guard
        raise CriticalGuardError("缺少 openpyxl，无法生成重保检查原表。") from exc

    normalized_sheet = normalize_sheet_name(sheet_name)
    normalized_cells = normalize_response_cells(normalized_sheet, scope, cells)
    template_path = critical_guard_template_path()
    if not template_path.is_file():
        raise CriticalGuardError(f"重保检查模板不存在：{template_path}")
    output_path.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(template_path, output_path)
    workbook = load_workbook(
        output_path,
        data_only=False,
        read_only=False,
        keep_links=False,
    )
    image_handles: list[Any] = []
    try:
        ws = workbook[normalized_sheet]
        if normalized_sheet in CRITICAL_GUARD_CHECK_SHEETS:
            image_handles = _write_check_sheet(
                ws,
                sheet_name=normalized_sheet,
                cells=normalized_cells,
                signatures=signatures,
            )
        elif normalized_sheet == "物资检查清单":
            _write_materials_sheet(ws, normalized_cells)
        else:
            _write_contacts_sheet(ws, normalized_cells)
        ws.sheet_view.showGridLines = False
        ws.print_area = critical_guard_sheet_range(normalized_sheet)
        workbook.active = workbook.sheetnames.index(normalized_sheet)
        workbook.save(output_path)
    except Exception:
        with suppress(OSError):
            output_path.unlink()
        raise
    finally:
        for handle in image_handles:
            with suppress(Exception):
                if hasattr(handle, "close"):
                    handle.close()
        with suppress(Exception):
            workbook.close()
    return {
        "path": str(output_path),
        "sha256": _file_sha256(output_path),
        "size": output_path.stat().st_size,
        "sheet_range": critical_guard_sheet_range(normalized_sheet),
    }


def _excel_column_width_pixels(value: Any) -> float:
    try:
        width = max(0.0, float(value))
    except Exception:
        width = 8.43
    if width <= 0:
        return 0.0
    # Excel stores character widths rather than pixels. This is the same
    # conversion used by the default Calibri/等线 10-11pt layout closely enough
    # for the fixed guard templates.
    return float(math.floor(((256.0 * width + math.floor(128.0 / 7.0)) / 256.0) * 7.0))


def _excel_row_height_pixels(value: Any) -> float:
    try:
        points = max(0.0, float(value))
    except Exception:
        points = 15.0
    return points * 96.0 / 72.0


def _theme_colors(workbook: Any) -> dict[int, str]:
    raw_theme = getattr(workbook, "loaded_theme", None)
    if not raw_theme:
        return {}
    try:
        root = ET.fromstring(raw_theme)
        namespace = {"a": "http://schemas.openxmlformats.org/drawingml/2006/main"}
        scheme = root.find(".//a:clrScheme", namespace)
        if scheme is None:
            return {}
        result: dict[int, str] = {}
        for index, entry in enumerate(list(scheme)):
            color_node = next(iter(entry), None)
            if color_node is None:
                continue
            value = str(
                color_node.attrib.get("lastClr")
                or color_node.attrib.get("val")
                or ""
            ).strip()
            if re.fullmatch(r"[0-9A-Fa-f]{6}", value):
                result[index] = value.upper()
        return result
    except Exception:
        return {}


def _apply_excel_tint(rgb: tuple[int, int, int], tint: Any) -> tuple[int, int, int]:
    try:
        amount = max(-1.0, min(1.0, float(tint or 0.0)))
    except Exception:
        amount = 0.0
    if amount == 0.0:
        return rgb
    values = []
    for channel in rgb:
        if amount < 0:
            adjusted = channel * (1.0 + amount)
        else:
            adjusted = channel * (1.0 - amount) + 255.0 * amount
        values.append(max(0, min(255, int(round(adjusted)))))
    return values[0], values[1], values[2]


def _excel_color_rgb(
    color: Any,
    theme: dict[int, str],
    *,
    default: tuple[int, int, int],
) -> tuple[int, int, int]:
    if color is None:
        return default
    color_type = str(getattr(color, "type", "") or "")
    value = ""
    if color_type == "rgb":
        raw = getattr(color, "rgb", "")
        value = raw if isinstance(raw, str) else ""
    elif color_type == "theme":
        index = getattr(color, "theme", None)
        if isinstance(index, int):
            value = theme.get(index, "")
    elif color_type == "indexed":
        try:
            from openpyxl.styles.colors import COLOR_INDEX

            index = int(getattr(color, "indexed", -1))
            if 0 <= index < len(COLOR_INDEX):
                value = str(COLOR_INDEX[index] or "")
        except Exception:
            value = ""
    if len(value) == 8:
        value = value[-6:]
    if not re.fullmatch(r"[0-9A-Fa-f]{6}", value):
        return default
    rgb = tuple(int(value[index : index + 2], 16) for index in (0, 2, 4))
    return _apply_excel_tint((rgb[0], rgb[1], rgb[2]), getattr(color, "tint", 0.0))


@lru_cache(maxsize=256)
def _excel_pillow_font(
    size_px: int,
    bold: bool,
    italic: bool,
    family: str,
) -> Any:
    from PIL import ImageFont

    normalized = str(family or "").strip().lower()
    if "宋体" in normalized or "simsun" in normalized:
        candidates = ["simsun.ttc", "msyhbd.ttc" if bold else "msyh.ttc"]
    elif "黑体" in normalized or "simhei" in normalized:
        candidates = ["simhei.ttf", "msyhbd.ttc" if bold else "msyh.ttc"]
    elif "等线" in normalized or "deng" in normalized:
        candidates = ["Dengb.ttf" if bold else "Deng.ttf", "msyhbd.ttc" if bold else "msyh.ttc"]
    elif "arial" in normalized:
        candidates = ["arialbi.ttf" if bold and italic else "arialbd.ttf" if bold else "ariali.ttf" if italic else "arial.ttf"]
    elif "calibri" in normalized:
        candidates = ["calibriz.ttf" if bold and italic else "calibrib.ttf" if bold else "calibrii.ttf" if italic else "calibri.ttf"]
    else:
        candidates = ["msyhbd.ttc" if bold else "msyh.ttc", "simsun.ttc"]
    for file_name in candidates:
        path = Path("C:/Windows/Fonts") / file_name
        if path.is_file():
            return ImageFont.truetype(str(path), size=max(7, int(size_px)))
    return ImageFont.load_default()


def _cell_display_text(cell: Any) -> str:
    value = getattr(cell, "value", None)
    if value in (None, ""):
        return ""
    if isinstance(value, dt.datetime):
        return value.strftime("%Y-%m-%d %H:%M")
    if isinstance(value, dt.date):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, dt.time):
        return value.strftime("%H:%M")
    if isinstance(value, bool):
        return "是" if value else "否"
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value)


def _text_width(draw: Any, value: str, font: Any) -> float:
    try:
        return float(draw.textlength(value, font=font))
    except Exception:
        bounds = draw.textbbox((0, 0), value, font=font)
        return float(bounds[2] - bounds[0])


def _wrap_excel_text(
    draw: Any,
    value: str,
    font: Any,
    max_width: float,
    *,
    wrap_text: bool,
) -> list[str]:
    paragraphs = value.replace("\r\n", "\n").replace("\r", "\n").split("\n")
    if not wrap_text:
        return paragraphs or [""]
    lines: list[str] = []
    for paragraph in paragraphs or [""]:
        if not paragraph:
            lines.append("")
            continue
        current = ""
        for character in paragraph:
            candidate = current + character
            if not current or _text_width(draw, candidate, font) <= max_width:
                current = candidate
                continue
            lines.append(current)
            current = character
        lines.append(current)
    return lines or [""]


def _font_line_height(font: Any) -> int:
    try:
        ascent, descent = font.getmetrics()
        return max(1, int(math.ceil(float(ascent) + float(descent))))
    except Exception:
        try:
            bounds = font.getbbox("国Ag", anchor="lt")
            return max(1, int(math.ceil(bounds[3] - bounds[1])))
        except Exception:
            return max(1, int(getattr(font, "size", 12)))


def _draw_excel_text(
    canvas: Any,
    cell: Any,
    box: tuple[int, int, int, int],
    *,
    scale: float,
    theme: dict[int, str],
) -> None:
    from PIL import Image, ImageDraw

    value = _cell_display_text(cell)
    if not value:
        return
    x1, y1, x2, y2 = box
    cell_width = max(1, int(x2 - x1))
    cell_height = max(1, int(y2 - y1))
    # Render each cell on its own transparent layer. This is both the clipping
    # boundary and the safety gap that keeps glyph descenders off cell borders.
    layer = Image.new("RGBA", (cell_width, cell_height), (0, 0, 0, 0))
    draw = ImageDraw.Draw(layer)
    horizontal_padding = max(2, int(round(4.0 * scale)))
    vertical_padding = max(2, int(round(2.5 * scale)))
    indent = max(0, int(getattr(cell.alignment, "indent", 0) or 0))
    indent_width = indent * max(2, int(round(6.0 * scale)))
    available_width = max(1, cell_width - horizontal_padding * 2 - indent_width)
    available_height = max(1, cell_height - vertical_padding * 2)
    point_size = float(getattr(cell.font, "sz", None) or 10.0)
    size_px = max(7, int(round(point_size * 96.0 / 72.0 * scale)))
    minimum_px = max(7, int(round(5.0 * 96.0 / 72.0 * scale)))
    wrap_text = bool(getattr(cell.alignment, "wrap_text", False)) or "\n" in value
    font = _excel_pillow_font(
        size_px,
        bool(getattr(cell.font, "bold", False)),
        bool(getattr(cell.font, "italic", False)),
        str(getattr(cell.font, "name", "") or ""),
    )
    lines = _wrap_excel_text(
        draw,
        value,
        font,
        available_width,
        wrap_text=wrap_text,
    )
    line_gap = max(1, int(round(0.75 * scale)))
    line_height = _font_line_height(font)
    while size_px > minimum_px:
        widest = max((_text_width(draw, line, font) for line in lines), default=0.0)
        total_height = len(lines) * line_height + max(0, len(lines) - 1) * line_gap
        if widest <= available_width + 1 and total_height <= available_height + 1:
            break
        size_px -= 1
        font = _excel_pillow_font(
            size_px,
            bool(getattr(cell.font, "bold", False)),
            bool(getattr(cell.font, "italic", False)),
            str(getattr(cell.font, "name", "") or ""),
        )
        lines = _wrap_excel_text(
            draw,
            value,
            font,
            available_width,
            wrap_text=wrap_text,
        )
        line_height = _font_line_height(font)
    total_height = len(lines) * line_height + max(0, len(lines) - 1) * line_gap
    vertical = str(getattr(cell.alignment, "vertical", "") or "").lower()
    if vertical in {"top", "justify", "distributed"}:
        cursor_y = vertical_padding
    elif vertical == "bottom":
        cursor_y = cell_height - vertical_padding - total_height
    else:
        cursor_y = max(vertical_padding, int((cell_height - total_height) / 2))
    horizontal = str(getattr(cell.alignment, "horizontal", "") or "").lower()
    if not horizontal or horizontal == "general":
        horizontal = "right" if isinstance(getattr(cell, "value", None), (int, float)) else "left"
    color = _excel_color_rgb(cell.font.color, theme, default=(0, 0, 0))
    for line in lines:
        width = _text_width(draw, line, font)
        if horizontal in {"center", "centercontinuous", "distributed"}:
            cursor_x = (cell_width - width) / 2
        elif horizontal == "right":
            cursor_x = cell_width - horizontal_padding - width
        else:
            cursor_x = horizontal_padding + indent_width
        text_position = (int(round(cursor_x)), int(round(cursor_y)))
        try:
            draw.text(
                text_position,
                line,
                font=font,
                fill=(*color, 255),
                anchor="lt",
            )
        except (TypeError, ValueError):
            bounds = draw.textbbox((0, 0), line or "国", font=font)
            draw.text(
                (text_position[0] - bounds[0], text_position[1] - bounds[1]),
                line,
                font=font,
                fill=(*color, 255),
            )
        if bool(getattr(cell.font, "underline", False)) and line:
            underline_y = min(
                cell_height - vertical_padding,
                int(round(cursor_y + line_height)),
            )
            draw.line(
                (int(round(cursor_x)), underline_y, int(round(cursor_x + width)), underline_y),
                fill=(*color, 255),
                width=max(1, int(round(scale))),
            )
        cursor_y += line_height + line_gap
    canvas.alpha_composite(layer, (int(x1), int(y1)))


def _border_width(style: Any, scale: float) -> int:
    normalized = str(style or "").lower()
    if not normalized:
        return 0
    base = {
        "hair": 0.5,
        "dotted": 0.75,
        "dashdotdot": 0.75,
        "dashdot": 0.75,
        "dashed": 0.75,
        "thin": 1.0,
        "mediumdashdotdot": 1.5,
        "mediumdashdot": 1.5,
        "mediumdashed": 1.5,
        "medium": 1.5,
        "thick": 2.25,
        "double": 2.25,
    }.get(normalized, 1.0)
    return max(1, int(round(base * scale)))


def _draw_excel_border_side(
    draw: Any,
    side: Any,
    points: tuple[int, int, int, int],
    *,
    scale: float,
    theme: dict[int, str],
) -> None:
    style = str(getattr(side, "style", "") or "").lower()
    width = _border_width(style, scale)
    if width <= 0:
        return
    color = _excel_color_rgb(getattr(side, "color", None), theme, default=(0, 0, 0))
    draw.line(points, fill=(*color, 255), width=width)
    if style == "double":
        x1, y1, x2, y2 = points
        offset = max(2, width)
        if y1 == y2:
            draw.line((x1, y1 + offset, x2, y2 + offset), fill=(*color, 255), width=max(1, width // 2))
        else:
            draw.line((x1 + offset, y1, x2 + offset, y2), fill=(*color, 255), width=max(1, width // 2))


_BORDER_STYLE_PRIORITY = {
    "hair": 1,
    "dotted": 2,
    "dashdotdot": 3,
    "dashdot": 4,
    "dashed": 5,
    "thin": 6,
    "mediumdashdotdot": 7,
    "mediumdashdot": 8,
    "mediumdashed": 9,
    "medium": 10,
    "thick": 11,
    "double": 12,
}


def _preferred_excel_border_side(*sides: Any, scale: float) -> Any | None:
    candidates = [side for side in sides if str(getattr(side, "style", "") or "").strip()]
    if not candidates:
        return None
    return max(
        candidates,
        key=lambda side: (
            _border_width(getattr(side, "style", ""), scale),
            _BORDER_STYLE_PRIORITY.get(
                str(getattr(side, "style", "") or "").lower(),
                0,
            ),
        ),
    )


def _draw_excel_cell_borders(
    draw: Any,
    worksheet: Any,
    *,
    min_col: int,
    min_row: int,
    max_col: int,
    max_row: int,
    margin: int,
    x_positions: dict[int, int],
    y_positions: dict[int, int],
    merged_cells: dict[tuple[int, int], tuple[int, int]],
    scale: float,
    theme: dict[int, str],
) -> None:
    # Each shared edge is rendered exactly once. Drawing both adjacent cell
    # sides made thin borders look darker or wider than otherwise identical lines.
    for row in range(min_row, max_row + 1):
        y1 = margin + y_positions[row]
        y2 = margin + y_positions[row + 1]
        for boundary in range(min_col, max_col + 2):
            left_anchor = merged_cells.get((row, boundary - 1)) if boundary > min_col else None
            right_anchor = merged_cells.get((row, boundary)) if boundary <= max_col else None
            if left_anchor is not None and left_anchor == right_anchor:
                continue
            left_side = (
                worksheet.cell(row=row, column=boundary - 1).border.right
                if boundary > min_col
                else None
            )
            right_side = (
                worksheet.cell(row=row, column=boundary).border.left
                if boundary <= max_col
                else None
            )
            side = _preferred_excel_border_side(left_side, right_side, scale=scale)
            if side is None:
                continue
            x = margin + x_positions[boundary]
            _draw_excel_border_side(
                draw,
                side,
                (x, y1, x, y2),
                scale=scale,
                theme=theme,
            )

    for boundary in range(min_row, max_row + 2):
        y = margin + y_positions[boundary]
        for column in range(min_col, max_col + 1):
            top_anchor = merged_cells.get((boundary - 1, column)) if boundary > min_row else None
            bottom_anchor = merged_cells.get((boundary, column)) if boundary <= max_row else None
            if top_anchor is not None and top_anchor == bottom_anchor:
                continue
            top_side = (
                worksheet.cell(row=boundary - 1, column=column).border.bottom
                if boundary > min_row
                else None
            )
            bottom_side = (
                worksheet.cell(row=boundary, column=column).border.top
                if boundary <= max_row
                else None
            )
            side = _preferred_excel_border_side(top_side, bottom_side, scale=scale)
            if side is None:
                continue
            x1 = margin + x_positions[column]
            x2 = margin + x_positions[column + 1]
            _draw_excel_border_side(
                draw,
                side,
                (x1, y, x2, y),
                scale=scale,
                theme=theme,
            )


def _render_workbook_sheet_to_png(
    *,
    workbook_path: Path,
    sheet_name: str,
    range_address: str,
    output_path: Path,
    scale: float = 2.0,
) -> None:
    try:
        from openpyxl import load_workbook
        from openpyxl.utils import get_column_letter, range_boundaries
        from PIL import Image, ImageDraw
    except Exception as exc:  # pragma: no cover - startup dependency guard
        raise CriticalGuardError("缺少 openpyxl 或 Pillow，无法生成重保检查图片。") from exc

    workbook = load_workbook(
        workbook_path,
        data_only=False,
        read_only=False,
        keep_links=False,
    )
    try:
        worksheet = workbook[_resolve_workbook_sheet_title(workbook, sheet_name)]
        min_col, min_row, max_col, max_row = range_boundaries(range_address)
        default_column_width = float(
            getattr(worksheet.sheet_format, "defaultColWidth", None) or 8.43
        )
        default_row_height = float(
            getattr(worksheet.sheet_format, "defaultRowHeight", None) or 15.0
        )
        base_column_widths: dict[int, float] = {}
        base_row_heights: dict[int, float] = {}
        for column in range(min_col, max_col + 1):
            dimension = worksheet.column_dimensions[get_column_letter(column)]
            base_column_widths[column] = 0.0 if dimension.hidden else _excel_column_width_pixels(
                dimension.width if dimension.width is not None else default_column_width
            )
        for row in range(min_row, max_row + 1):
            dimension = worksheet.row_dimensions[row]
            base_row_heights[row] = 0.0 if dimension.hidden else _excel_row_height_pixels(
                dimension.height if dimension.height is not None else default_row_height
            )

        requested_scale = max(0.1, float(scale or 2.0))
        minimum_scale = min(requested_scale, CRITICAL_GUARD_MIN_RENDER_SCALE)
        scale = requested_scale
        base_width = sum(base_column_widths.values())
        base_height = sum(base_row_heights.values())
        for _ in range(6):
            candidate_margin = max(2, int(round(2.0 * scale)))
            candidate_width = max(
                1,
                int(round(base_width * scale)) + candidate_margin * 2,
            )
            candidate_height = max(
                1,
                int(round(base_height * scale)) + candidate_margin * 2,
            )
            if (
                candidate_width <= CRITICAL_GUARD_MAX_RENDER_DIMENSION
                and candidate_height <= CRITICAL_GUARD_MAX_RENDER_DIMENSION
                and candidate_width * candidate_height <= CRITICAL_GUARD_MAX_RENDER_PIXELS
            ):
                break
            ratios = [
                CRITICAL_GUARD_MAX_RENDER_DIMENSION / candidate_width,
                CRITICAL_GUARD_MAX_RENDER_DIMENSION / candidate_height,
                (
                    CRITICAL_GUARD_MAX_RENDER_PIXELS
                    / float(candidate_width * candidate_height)
                )
                ** 0.5,
            ]
            next_scale = max(minimum_scale, scale * min(ratios) * 0.98)
            if next_scale >= scale:
                break
            scale = next_scale

        column_widths = {
            column: width * scale for column, width in base_column_widths.items()
        }
        row_heights = {
            row: height * scale for row, height in base_row_heights.items()
        }

        x_positions: dict[int, int] = {min_col: 0}
        for column in range(min_col, max_col + 1):
            x_positions[column + 1] = int(round(x_positions[column] + column_widths[column]))
        y_positions: dict[int, int] = {min_row: 0}
        for row in range(min_row, max_row + 1):
            y_positions[row + 1] = int(round(y_positions[row] + row_heights[row]))

        margin = max(2, int(round(2.0 * scale)))
        image_width = max(1, x_positions[max_col + 1] + margin * 2)
        image_height = max(1, y_positions[max_row + 1] + margin * 2)
        if (
            image_width > CRITICAL_GUARD_MAX_RENDER_DIMENSION
            or image_height > CRITICAL_GUARD_MAX_RENDER_DIMENSION
            or image_width * image_height > CRITICAL_GUARD_MAX_RENDER_PIXELS
        ):
            raise CriticalGuardError(
                "检查表尺寸过大，即使降低清晰度也无法在程序内生成图片。"
            )
        canvas = Image.new("RGBA", (image_width, image_height), (255, 255, 255, 255))
        draw = ImageDraw.Draw(canvas)
        theme = _theme_colors(workbook)

        merged_anchors: dict[tuple[int, int], tuple[int, int, int, int]] = {}
        merged_cells: dict[tuple[int, int], tuple[int, int]] = {}
        for merged in worksheet.merged_cells.ranges:
            if (
                merged.max_col < min_col
                or merged.min_col > max_col
                or merged.max_row < min_row
                or merged.min_row > max_row
            ):
                continue
            clipped = (
                max(min_col, merged.min_col),
                max(min_row, merged.min_row),
                min(max_col, merged.max_col),
                min(max_row, merged.max_row),
            )
            anchor = (merged.min_row, merged.min_col)
            merged_anchors[anchor] = clipped
            for row in range(clipped[1], clipped[3] + 1):
                for column in range(clipped[0], clipped[2] + 1):
                    merged_cells[(row, column)] = anchor

        # Draw fills first so merged-cell backgrounds cover their whole range.
        for row in range(min_row, max_row + 1):
            for column in range(min_col, max_col + 1):
                anchor = merged_cells.get((row, column))
                if anchor and anchor != (row, column):
                    continue
                cell = worksheet.cell(row=row, column=column)
                if anchor:
                    left, top, right, bottom = merged_anchors[anchor]
                else:
                    left = right = column
                    top = bottom = row
                box = (
                    margin + x_positions[left],
                    margin + y_positions[top],
                    margin + x_positions[right + 1],
                    margin + y_positions[bottom + 1],
                )
                if str(getattr(cell.fill, "fill_type", "") or "").lower() == "solid":
                    fill = _excel_color_rgb(cell.fill.fgColor, theme, default=(255, 255, 255))
                    draw.rectangle(box, fill=(*fill, 255))

        # Text is drawn against the complete merged-cell box, matching Excel.
        for row in range(min_row, max_row + 1):
            for column in range(min_col, max_col + 1):
                anchor = merged_cells.get((row, column))
                if anchor and anchor != (row, column):
                    continue
                cell = worksheet.cell(row=row, column=column)
                if anchor:
                    left, top, right, bottom = merged_anchors[anchor]
                else:
                    left = right = column
                    top = bottom = row
                _draw_excel_text(
                    canvas,
                    cell,
                    (
                        margin + x_positions[left],
                        margin + y_positions[top],
                        margin + x_positions[right + 1],
                        margin + y_positions[bottom + 1],
                    ),
                    scale=scale,
                    theme=theme,
                )

        def column_position(zero_based_column: int) -> float:
            one_based = zero_based_column + 1
            if one_based >= min_col:
                total = sum(column_widths.get(column, 0.0) for column in range(min_col, one_based))
            else:
                total = -sum(
                    _excel_column_width_pixels(
                        worksheet.column_dimensions[get_column_letter(column)].width
                        or default_column_width
                    )
                    * scale
                    for column in range(one_based, min_col)
                )
            return float(total)

        def row_position(zero_based_row: int) -> float:
            one_based = zero_based_row + 1
            if one_based >= min_row:
                total = sum(row_heights.get(row, 0.0) for row in range(min_row, one_based))
            else:
                total = -sum(
                    _excel_row_height_pixels(
                        worksheet.row_dimensions[row].height or default_row_height
                    )
                    * scale
                    for row in range(one_based, min_row)
                )
            return float(total)

        emu_per_pixel = 9525.0
        resampling = getattr(getattr(Image, "Resampling", Image), "LANCZOS")
        for excel_image in list(getattr(worksheet, "_images", []) or []):
            try:
                image_bytes = excel_image._data()
                source = Image.open(io.BytesIO(image_bytes)).convert("RGBA")
                anchor = excel_image.anchor
                if isinstance(anchor, str):
                    from openpyxl.utils.cell import coordinate_to_tuple

                    anchor_row, anchor_column = coordinate_to_tuple(anchor)
                    x = column_position(anchor_column - 1)
                    y = row_position(anchor_row - 1)
                    target_width = float(excel_image.width) * scale
                    target_height = float(excel_image.height) * scale
                else:
                    marker = getattr(anchor, "_from", None)
                    if marker is None:
                        continue
                    x = column_position(int(marker.col)) + float(marker.colOff or 0) / emu_per_pixel * scale
                    y = row_position(int(marker.row)) + float(marker.rowOff or 0) / emu_per_pixel * scale
                    target_marker = getattr(anchor, "to", None)
                    extent = getattr(anchor, "ext", None)
                    if target_marker is not None:
                        x2 = column_position(int(target_marker.col)) + float(target_marker.colOff or 0) / emu_per_pixel * scale
                        y2 = row_position(int(target_marker.row)) + float(target_marker.rowOff or 0) / emu_per_pixel * scale
                        target_width = x2 - x
                        target_height = y2 - y
                    elif extent is not None:
                        target_width = float(extent.cx or 0) / emu_per_pixel * scale
                        target_height = float(extent.cy or 0) / emu_per_pixel * scale
                    else:
                        target_width = float(excel_image.width) * scale
                        target_height = float(excel_image.height) * scale
                width = max(1, int(round(target_width)))
                height = max(1, int(round(target_height)))
                source = source.resize((width, height), resampling)
                canvas.alpha_composite(
                    source,
                    (margin + int(round(x)), margin + int(round(y))),
                )
            except Exception as exc:
                raise CriticalGuardError("原表中的图片或签名无法渲染。") from exc

        # Borders stay above both text and floating signature images. Shared
        # edges are resolved once so their color and thickness remain uniform.
        _draw_excel_cell_borders(
            draw,
            worksheet,
            min_col=min_col,
            min_row=min_row,
            max_col=max_col,
            max_row=max_row,
            margin=margin,
            x_positions=x_positions,
            y_positions=y_positions,
            merged_cells=merged_cells,
            scale=scale,
            theme=theme,
        )

        output_path.parent.mkdir(parents=True, exist_ok=True)
        with suppress(OSError):
            output_path.unlink()
        canvas.convert("RGB").save(output_path, format="PNG", optimize=True)
    finally:
        with suppress(Exception):
            workbook.close()


def render_critical_guard_template_artifacts(
    *,
    sheet_name: str,
    scope: str,
    task_name: str,
    cells: dict[str, Any],
    signatures: list[dict[str, Any]] | None,
    output_path: Path,
) -> dict[str, Any]:
    del task_name  # 任务名称由任务目录和页面汇总展示，原模板不额外改标题。
    workbook_path = output_path.with_suffix(".xlsx")
    workbook_metadata = build_critical_guard_workbook(
        sheet_name=sheet_name,
        scope=scope,
        cells=cells,
        signatures=signatures,
        output_path=workbook_path,
    )
    range_address = str(workbook_metadata.get("sheet_range") or "")
    try:
        _render_workbook_sheet_to_png(
            workbook_path=workbook_path,
            sheet_name=normalize_sheet_name(sheet_name),
            range_address=range_address,
            output_path=output_path,
        )
    except Exception as exc:
        with suppress(OSError):
            output_path.unlink()
        with suppress(OSError):
            workbook_path.unlink()
        if isinstance(exc, CriticalGuardError):
            raise
        raise CriticalGuardError("程序内部无法按原表样式生成图片。") from exc
    try:
        from PIL import Image

        with Image.open(output_path) as image:
            width, height = image.size
        if not _rendered_image_is_usable(output_path):
            raise ValueError("rendered image is blank")
    except Exception as exc:
        with suppress(OSError):
            output_path.unlink()
        with suppress(OSError):
            workbook_path.unlink()
        raise CriticalGuardError("生成的重保检查图片不可用。") from exc
    return {
        "path": str(output_path),
        "sha256": _file_sha256(output_path),
        "size": output_path.stat().st_size,
        "width": width,
        "height": height,
        "signature_count": len(signatures or []),
        "workbook_path": str(workbook_path),
        "workbook_sha256": str(workbook_metadata.get("sha256") or ""),
        "workbook_size": int(workbook_metadata.get("size") or 0),
    }


def validate_critical_guard_source_workbook(
    source_path: Path,
    *,
    sheet_name: str,
) -> dict[str, Any]:
    """Validate an immutable per-building workbook before it is stored."""
    try:
        from openpyxl import load_workbook
    except Exception as exc:  # pragma: no cover - startup dependency guard
        raise CriticalGuardError("缺少 openpyxl，无法读取楼栋清单文件。") from exc

    normalized_sheet = normalize_sheet_name(sheet_name)
    if normalized_sheet not in CRITICAL_GUARD_FILE_SHEETS:
        raise CriticalGuardError("只有物资检查清单和重保联络清单可以上传楼栋文件。")
    if not source_path.is_file():
        raise CriticalGuardError("上传的楼栋清单文件不存在。")
    _validate_xlsx_archive(source_path)
    workbook = None
    try:
        workbook = load_workbook(
            source_path,
            data_only=False,
            read_only=True,
            keep_links=False,
        )
        worksheet = workbook[_resolve_workbook_sheet_title(workbook, normalized_sheet)]
        if int(worksheet.max_row or 0) <= 0 or int(worksheet.max_column or 0) <= 0:
            raise CriticalGuardError(f"上传文件中的“{normalized_sheet}”工作表为空。")
        sheet_range = _source_sheet_range(worksheet, normalized_sheet)
    except CriticalGuardError:
        raise
    except Exception as exc:
        raise CriticalGuardError("上传文件损坏或不是有效的 Excel 工作簿。") from exc
    finally:
        if workbook is not None:
            with suppress(Exception):
                workbook.close()
    return {
        "sheet_type": normalized_sheet,
        "sheet_range": sheet_range,
        "sha256": _file_sha256(source_path),
        "size": source_path.stat().st_size,
    }


def render_critical_guard_source_file_artifacts(
    *,
    source_path: Path,
    sheet_name: str,
    output_path: Path,
) -> dict[str, Any]:
    """Render a stored building workbook without modifying its cell contents."""
    metadata = validate_critical_guard_source_workbook(
        source_path,
        sheet_name=sheet_name,
    )
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook_path = output_path.with_suffix(".xlsx")
    with suppress(OSError):
        workbook_path.unlink()
    shutil.copy2(source_path, workbook_path)
    try:
        preview = render_critical_guard_source_file_preview(
            source_path=workbook_path,
            sheet_name=sheet_name,
            output_path=output_path,
        )
    except Exception as exc:
        with suppress(OSError):
            output_path.unlink()
        with suppress(OSError):
            workbook_path.unlink()
        if isinstance(exc, CriticalGuardError):
            raise
        raise CriticalGuardError("程序内部无法生成楼栋清单图片。") from exc
    return {
        "path": str(output_path),
        "sha256": str(preview.get("sha256") or ""),
        "size": int(preview.get("size") or 0),
        "width": int(preview.get("width") or 0),
        "height": int(preview.get("height") or 0),
        "signature_count": 0,
        "workbook_path": str(workbook_path),
        "workbook_sha256": _file_sha256(workbook_path),
        "workbook_size": workbook_path.stat().st_size,
        "source_file_sha256": str(metadata.get("sha256") or ""),
    }


def render_critical_guard_source_file_preview(
    *,
    source_path: Path,
    sheet_name: str,
    output_path: Path,
) -> dict[str, Any]:
    """Render the selected sheet of an immutable building workbook."""
    metadata = validate_critical_guard_source_workbook(
        source_path,
        sheet_name=sheet_name,
    )
    output_path.parent.mkdir(parents=True, exist_ok=True)
    try:
        _render_workbook_sheet_to_png(
            workbook_path=source_path,
            sheet_name=normalize_sheet_name(sheet_name),
            range_address=str(metadata.get("sheet_range") or ""),
            output_path=output_path,
        )
        from PIL import Image

        with Image.open(output_path) as image:
            width, height = image.size
        if not _rendered_image_is_usable(output_path):
            raise ValueError("rendered image is blank")
    except Exception as exc:
        with suppress(OSError):
            output_path.unlink()
        if isinstance(exc, CriticalGuardError):
            raise
        raise CriticalGuardError("程序内部无法生成楼栋清单预览。") from exc
    return {
        "path": str(output_path),
        "sha256": _file_sha256(output_path),
        "size": output_path.stat().st_size,
        "width": width,
        "height": height,
        "source_file_sha256": str(metadata.get("sha256") or ""),
    }


def render_critical_guard_image(
    *,
    sheet_name: str,
    scope: str,
    task_name: str,
    cells: dict[str, Any],
    signatures: list[dict[str, Any]] | None,
    output_path: Path,
) -> dict[str, Any]:
    try:
        from PIL import Image, ImageDraw, ImageFont
    except Exception as exc:  # pragma: no cover - startup dependency guard
        raise CriticalGuardError("缺少 Pillow，无法生成重保检查图片。") from exc

    definition = sheet_definition(sheet_name)
    width = 1800
    margin = 40
    blue = "#2f66b6"
    dark = "#17243a"
    muted = "#53657d"
    line = "#7c91ad"
    header_fill = "#dce9f8"
    group_fill = "#edf4fc"
    normal_fill = "#eef9f4"
    abnormal_fill = "#fff0f0"
    white = "#ffffff"

    font_paths = [
        Path("C:/Windows/Fonts/msyh.ttc"),
        Path("C:/Windows/Fonts/simhei.ttf"),
        Path("C:/Windows/Fonts/simsun.ttc"),
    ]
    bold_paths = [Path("C:/Windows/Fonts/msyhbd.ttc"), *font_paths]

    def font(size: int, *, bold: bool = False):
        for path in (bold_paths if bold else font_paths):
            if path.exists():
                return ImageFont.truetype(str(path), size=size)
        return ImageFont.load_default()

    title_font = font(36, bold=True)
    section_font = font(25, bold=True)
    head_font = font(21, bold=True)
    body_font = font(19)
    body_bold = font(19, bold=True)
    small_font = font(17)
    signature_items = [
        item
        for item in (signatures or [])
        if isinstance(item, dict) and isinstance(item.get("image_bytes"), (bytes, bytearray))
    ]
    signature_columns = 4
    signature_row_height = 112
    signature_rows = max(1, math.ceil(len(signature_items) / signature_columns))
    signature_area_height = max(122, signature_rows * signature_row_height)

    def wrap(draw: Any, text: Any, box_width: int, use_font: Any) -> list[str]:
        value = str(text or "")
        if not value:
            return [""]
        lines: list[str] = []
        for paragraph in value.splitlines() or [""]:
            current = ""
            for char in paragraph:
                candidate = current + char
                if draw.textlength(candidate, font=use_font) <= max(20, box_width - 18):
                    current = candidate
                else:
                    if current:
                        lines.append(current)
                    current = char
            lines.append(current)
        return lines or [""]

    def row_height(draw: Any, values: list[tuple[Any, int, Any]], minimum: int = 46) -> int:
        line_count = max(len(wrap(draw, value, box_width, use_font)) for value, box_width, use_font in values)
        return max(minimum, line_count * 28 + 16)

    def draw_cell(
        draw: Any,
        box: tuple[int, int, int, int],
        value: Any,
        *,
        fill: str = white,
        use_font: Any = body_font,
        color: str = dark,
        align: str = "left",
    ) -> None:
        x1, y1, x2, y2 = box
        draw.rectangle(box, fill=fill, outline=line, width=1)
        lines = wrap(draw, value, x2 - x1, use_font)
        total_height = len(lines) * 27
        y = y1 + max(7, (y2 - y1 - total_height) // 2)
        for line_text in lines:
            text_width = draw.textlength(line_text, font=use_font)
            if align == "center":
                x = x1 + max(8, (x2 - x1 - text_width) / 2)
            elif align == "right":
                x = x2 - text_width - 9
            else:
                x = x1 + 9
            draw.text((x, y), line_text, font=use_font, fill=color)
            y += 27

    # Build once on a tiny canvas for text measurement, then create the final image.
    measure = Image.new("RGB", (width, 100), white)
    measure_draw = ImageDraw.Draw(measure)
    estimated_height = 260
    if definition["kind"] == "check":
        columns = [230, 830, 190, 470]
        for item in definition.get("items") or []:
            check = (cells.get("checks") or {}).get(str(item["key"]), {})
            estimated_height += row_height(
                measure_draw,
                [
                    (item.get("category"), columns[0], body_font),
                    (item.get("content"), columns[1], body_font),
                    ("异常" if check.get("status") == "abnormal" else "正常", columns[2], body_bold),
                    (check.get("note"), columns[3], body_font),
                ],
            )
        estimated_height += 138 + signature_area_height + (150 if definition.get("has_weather") else 0)
    elif definition["kind"] == "materials":
        estimated_height = 260
        for section in definition.get("sections") or []:
            estimated_height += 124
            rows = cells.get(str(section["key"])) if isinstance(cells.get(str(section["key"])), list) else []
            columns = list(section["columns"])
            data_width = int((width - margin * 2 - 90) / max(1, len(columns)))
            for row_index in range(int(section["row_count"])):
                row = rows[row_index] if row_index < len(rows) and isinstance(rows[row_index], dict) else {}
                estimated_height += row_height(
                    measure_draw,
                    [
                        (row.get(str(column["key"])), data_width, body_font)
                        for column in columns
                    ],
                    minimum=56,
                )
        estimated_height += 60
    else:
        estimated_height = 260
        sections = [definition["duty"], *definition.get("groups", [])]
        for section_index, section in enumerate(sections):
            is_duty = section_index == 0
            columns = definition["duty"]["columns"] if is_duty else definition["group_columns"]
            row_count = int(definition["duty"]["row_count"] if is_duty else definition["group_row_count"])
            rows = cells.get("duty") if is_duty else (cells.get("groups") or {}).get(str(section["key"]))
            rows = rows if isinstance(rows, list) else []
            data_width = int((width - margin * 2 - 90) / max(1, len(columns)))
            estimated_height += 118
            for row_index in range(row_count):
                row = rows[row_index] if row_index < len(rows) and isinstance(rows[row_index], dict) else {}
                estimated_height += row_height(
                    measure_draw,
                    [
                        (row.get(str(column["key"])), data_width, body_font)
                        for column in columns
                    ],
                    minimum=54,
                )
        estimated_height += 60

    image = Image.new("RGB", (width, max(600, estimated_height)), white)
    draw = ImageDraw.Draw(image)
    y = 28
    draw.text((margin, y), str(definition.get("title") or sheet_name), font=title_font, fill=dark)
    y += 55
    draw.text((margin, y), f"任务：{task_name}", font=small_font, fill=muted)
    draw.text((width - 520, y), f"检查机房：南通机房{normalize_scope(scope)}楼", font=small_font, fill=muted)
    y += 38
    draw.text((margin, y), f"检查日期：{cells.get('check_date') or ''}", font=small_font, fill=muted)
    y += 44

    if definition["kind"] == "check":
        if definition.get("has_weather"):
            weather = cells.get("weather") if isinstance(cells.get("weather"), dict) else {}
            for key, label in (("level1", "一级极端天气"), ("level2", "二级极端天气"), ("current", "当前检查极端天气")):
                draw_cell(draw, (margin, y, margin + 270, y + 44), label, fill=group_fill, use_font=body_bold)
                draw_cell(draw, (margin + 270, y, width - margin, y + 44), weather.get(key, ""))
                y += 44
            y += 12
        x = [margin, margin + 230, margin + 1060, margin + 1250, width - margin]
        for start, end, label in zip(x, x[1:], ("检查项", "检查内容", "检查结果", "备注（异常现象描述）")):
            draw_cell(draw, (start, y, end, y + 54), label, fill=header_fill, use_font=head_font, color=blue, align="center")
        y += 54
        checks = cells.get("checks") if isinstance(cells.get("checks"), dict) else {}
        items = list(definition.get("items") or [])
        heights: list[int] = []
        for item in items:
            check = checks.get(str(item["key"])) if isinstance(checks.get(str(item["key"])), dict) else {}
            heights.append(
                row_height(
                    draw,
                    [
                        (item.get("content"), x[2] - x[1], body_font),
                        (check.get("note"), x[4] - x[3], body_font),
                    ],
                )
            )
        index = 0
        while index < len(items):
            category = str(items[index].get("category") or "其他")
            group_end = index + 1
            while group_end < len(items) and str(items[group_end].get("category") or "其他") == category:
                group_end += 1
            group_height = sum(heights[index:group_end])
            draw_cell(draw, (x[0], y, x[1], y + group_height), category, fill=group_fill, use_font=body_bold, align="center")
            for item_index in range(index, group_end):
                item = items[item_index]
                height = heights[item_index]
                check = checks.get(str(item["key"])) if isinstance(checks.get(str(item["key"])), dict) else {}
                abnormal = check.get("status") == "abnormal"
                draw_cell(draw, (x[1], y, x[2], y + height), item.get("content"))
                draw_cell(
                    draw,
                    (x[2], y, x[3], y + height),
                    "异常" if abnormal else "正常",
                    fill=abnormal_fill if abnormal else normal_fill,
                    use_font=body_bold,
                    color="#b42318" if abnormal else "#087f5b",
                    align="center",
                )
                draw_cell(draw, (x[3], y, x[4], y + height), check.get("note", ""), fill=abnormal_fill if abnormal else white)
                y += height
            index = group_end
        suggestions_height = row_height(draw, [(cells.get("suggestions"), width - margin * 2 - 250, body_font)], minimum=66)
        draw_cell(draw, (margin, y, margin + 250, y + suggestions_height), "检查意见及建议", fill=group_fill, use_font=body_bold)
        draw_cell(draw, (margin + 250, y, width - margin, y + suggestions_height), cells.get("suggestions", ""))
        y += suggestions_height
        draw_cell(
            draw,
            (margin, y, margin + 250, y + signature_area_height),
            "检查人签字",
            fill=group_fill,
            use_font=body_bold,
        )
        signature_x1 = margin + 250
        signature_x2 = width - margin
        draw_cell(draw, (signature_x1, y, signature_x2, y + signature_area_height), "")
        slot_width = max(1, (signature_x2 - signature_x1) // signature_columns)
        for index, item in enumerate(signature_items):
            row_index, column_index = divmod(index, signature_columns)
            slot_x1 = signature_x1 + column_index * slot_width
            slot_x2 = signature_x2 if column_index == signature_columns - 1 else slot_x1 + slot_width
            slot_y1 = y + row_index * signature_row_height
            slot_y2 = min(y + signature_area_height, slot_y1 + signature_row_height)
            draw.rectangle((slot_x1, slot_y1, slot_x2, slot_y2), outline=line, width=1)
            try:
                signature = Image.open(io.BytesIO(bytes(item["image_bytes"]))).convert("RGBA")
                signature.thumbnail((max(40, slot_width - 24), 72))
                image.paste(
                    signature,
                    (
                        int(slot_x1 + max(8, (slot_x2 - slot_x1 - signature.width) / 2)),
                        int(slot_y1 + 7),
                    ),
                    signature,
                )
            except Exception as exc:
                raise CriticalGuardError("检查人签名图片不可用，请重新签名。") from exc
            signer_name = _clean_text(item.get("name"), limit=40) or f"检查人{index + 1}"
            name_width = draw.textlength(signer_name, font=small_font)
            draw.text(
                (
                    slot_x1 + max(8, (slot_x2 - slot_x1 - name_width) / 2),
                    slot_y2 - 29,
                ),
                signer_name,
                font=small_font,
                fill=muted,
            )
        y += signature_area_height

    elif definition["kind"] == "materials":
        for section in definition.get("sections") or []:
            y += 10
            draw.text((margin, y), str(section["title"]), font=section_font, fill=dark)
            y += 40
            columns = list(section["columns"])
            widths = [90] + [int((width - margin * 2 - 90) / len(columns))] * len(columns)
            x_positions = [margin]
            for cell_width in widths:
                x_positions.append(x_positions[-1] + cell_width)
            x_positions[-1] = width - margin
            labels = ["序号", *[str(column["label"]) for column in columns]]
            for idx, label in enumerate(labels):
                draw_cell(draw, (x_positions[idx], y, x_positions[idx + 1], y + 50), label, fill=header_fill, use_font=head_font, color=blue, align="center")
            y += 50
            rows = cells.get(str(section["key"])) if isinstance(cells.get(str(section["key"])), list) else []
            for row_index in range(int(section["row_count"])):
                row = rows[row_index] if row_index < len(rows) and isinstance(rows[row_index], dict) else {}
                height = row_height(
                    draw,
                    [(row.get(str(column["key"])), x_positions[index + 2] - x_positions[index + 1], body_font) for index, column in enumerate(columns)],
                    minimum=56,
                )
                draw_cell(draw, (x_positions[0], y, x_positions[1], y + height), row_index + 1, fill=group_fill, align="center")
                for index, column in enumerate(columns):
                    draw_cell(draw, (x_positions[index + 1], y, x_positions[index + 2], y + height), row.get(str(column["key"]), ""))
                y += height
            y += 24
    else:
        sections = [definition["duty"], *definition.get("groups", [])]
        for section_index, section in enumerate(sections):
            y += 10
            draw.text((margin, y), str(section["title"]), font=section_font, fill=dark)
            y += 40
            is_duty = section_index == 0
            columns = definition["duty"]["columns"] if is_duty else definition["group_columns"]
            row_count = int(definition["duty"]["row_count"] if is_duty else definition["group_row_count"])
            rows = cells.get("duty") if is_duty else (cells.get("groups") or {}).get(str(section["key"]))
            rows = rows if isinstance(rows, list) else []
            widths = [90] + [int((width - margin * 2 - 90) / len(columns))] * len(columns)
            x_positions = [margin]
            for cell_width in widths:
                x_positions.append(x_positions[-1] + cell_width)
            x_positions[-1] = width - margin
            labels = ["序号", *[str(column["label"]) for column in columns]]
            for idx, label in enumerate(labels):
                draw_cell(draw, (x_positions[idx], y, x_positions[idx + 1], y + 48), label, fill=header_fill, use_font=head_font, color=blue, align="center")
            y += 48
            for row_index in range(row_count):
                row = rows[row_index] if row_index < len(rows) and isinstance(rows[row_index], dict) else {}
                height = row_height(
                    draw,
                    [
                        (
                            row.get(str(column["key"])),
                            x_positions[index + 2] - x_positions[index + 1],
                            body_font,
                        )
                        for index, column in enumerate(columns)
                    ],
                    minimum=54,
                )
                draw_cell(draw, (x_positions[0], y, x_positions[1], y + height), row_index + 1, fill=group_fill, align="center")
                for index, column in enumerate(columns):
                    draw_cell(draw, (x_positions[index + 1], y, x_positions[index + 2], y + height), row.get(str(column["key"]), ""))
                y += height
            y += 20

    final_height = min(image.height, max(240, y + 36))
    image = image.crop((0, 0, width, final_height))
    output_path.parent.mkdir(parents=True, exist_ok=True)
    image.save(output_path, format="PNG", optimize=True)
    content = output_path.read_bytes()
    return {
        "path": str(output_path),
        "sha256": hashlib.sha256(content).hexdigest(),
        "size": len(content),
        "width": image.width,
        "height": image.height,
        "signature_count": len(signature_items),
    }
